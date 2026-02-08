"""
Phase-1 Clean Point Design UI (Streamlit)

Professional, single-command UI for:
- Point Designer: evaluate one operating point and show pass/fail constraint dashboard
- Scan Lab: run parameter scans and explore results
- Results Explorer: filter/sort/export feasible points

Design goals:
- No JS toolchain required (runs on pure Python).
- Physics and models live in src/ (imported as a library).
- All models remain explicit proxies (Phase-1), with conservative pass/fail gates.
"""

from __future__ import annotations

# ---- import path bootstrap (must be before any local imports) ----
import os as _os, sys as _sys
_ROOT = _os.path.abspath(_os.path.join(_os.path.dirname(__file__), ".."))
_SRC = _os.path.join(_ROOT, "src")
if _SRC not in _sys.path:
    _sys.path.insert(0, _SRC)
# ---------------------------------------------------------------

# Expose repo root for downstream helpers
ROOT = _ROOT
SRC = _SRC

import io
import json
import os
import math
import datetime
import time
import concurrent.futures
import queue
import threading
from dataclasses import asdict, fields, replace
from typing import Dict, Any, List, Tuple

import pandas as pd
from decision.kpis import headline_kpis

import streamlit as st

# ---------------------------------------------------------------------------
# Session-state initialization (prevents AttributeError on first run)
# ---------------------------------------------------------------------------
def _init_session_state() -> None:

    # NOTE: phase1_core import happens later in this file. We defensively import
    # PointInputs here to avoid NameError during early session-state init.
    try:
        from phase1_core import PointInputs as _PointInputs
    except Exception:
        _PointInputs = None

    defaults = {
        # If PointInputs is not available for any reason, fall back to None and let
        # downstream code initialize it in a controlled path.
        "last_point_inp": (_PointInputs(R0_m=1.85, a_m=0.57, kappa=1.8, Bt_T=12.2, Ip_MA=8.0, Ti_keV=15.0, fG=0.8, Paux_MW=20.0)
                           if _PointInputs is not None else None),
        "last_point_out": None,
        "last_solver_log": None,
        "explain_mode": True,
        "expert_mode": False,
        "scan_df": None,
        "scan_meta": None,
        "scan_log_lines": [],
        "scan_log_text": "",
        "scan_progress": 0.0,
        "scan_queue": [],
        "scan_running": False,
        "scan_future": None,
        "scan_executor": None,

        # Phase 7+ UI persistence
        "de_best_inputs": None,
        "de_best_out": None,
        "de_history": None,
        "robustness_result": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_session_state()
# _sync_point_designer_from_last_point_inp()


# Stable Streamlit keys for Point Designer widgets (presets rely on these)
PD_KEYS = {
    "R0_m": "pd_R0_m",
    "a_m": "pd_a_m",
    "kappa": "pd_kappa",
    "delta": "pd_delta",
    "Bt_T": "pd_Bt_T",
    "Ti_keV": "pd_Ti_keV",
    "Paux_MW": "pd_Paux_MW",
    "H98_tgt": "pd_target_H98",
    "Q_tgt": "pd_target_Q",
    "Ip_lo": "pd_Ip_lo",
    "Ip_hi": "pd_Ip_hi",
    "fG_lo": "pd_fG_lo",
    "fG_hi": "pd_fG_hi",
    # Additional stable keys (used for preset propagation)
    "Ti_over_Te": "pd_Ti_over_Te",
    "Paux_for_Q": "pd_Paux_for_Q",
}
# One-shot synchronization: when a preset is loaded we set this flag, and on the next
# rerun we push preset values into Point Designer widget keys.
def _sync_point_designer_from_last_point_inp() -> None:
    if not st.session_state.get("pd_needs_sync", False):
        return
    base = st.session_state.get("last_point_inp", None)
    if base is None:
        st.session_state["pd_needs_sync"] = False
        return
    # Overwrite widget state (Point Designer "Inputs") from the loaded preset.
    try:
        st.session_state[PD_KEYS["R0_m"]] = float(getattr(base, "R0_m"))
        st.session_state[PD_KEYS["a_m"]] = float(getattr(base, "a_m"))
        st.session_state[PD_KEYS["kappa"]] = float(getattr(base, "kappa"))
        st.session_state[PD_KEYS["delta"]] = float(getattr(base, "delta", 0.0) or 0.0)
        st.session_state[PD_KEYS["Bt_T"]] = float(getattr(base, "Bt_T"))
        st.session_state[PD_KEYS["Ti_keV"]] = float(getattr(base, "Ti_keV"))
        st.session_state[PD_KEYS["Paux_MW"]] = float(getattr(base, "Paux_MW"))
    except Exception:
        pass
    # Bounds from preset Ip/fG
    try:
        ip = float(getattr(base, "Ip_MA"))
        st.session_state[PD_KEYS["Ip_lo"]] = max(0.1, 0.80 * ip)
        st.session_state[PD_KEYS["Ip_hi"]] = max(0.2, 1.20 * ip)
    except Exception:
        pass
    try:
        fg = float(getattr(base, "fG"))
        st.session_state[PD_KEYS["fG_lo"]] = max(0.0, fg - 0.20)
        st.session_state[PD_KEYS["fG_hi"]] = min(2.0, fg + 0.20)
    except Exception:
        pass
    # Aux/Q denominator and Ti/Te
    try:
        st.session_state[PD_KEYS["Paux_for_Q"]] = float(getattr(base, "Paux_MW"))
    except Exception:
        pass
    try:
        st.session_state[PD_KEYS["Ti_over_Te"]] = float(getattr(base, "Ti_over_Te", 1.0))
    except Exception:
        pass
    st.session_state["pd_needs_sync"] = False




_sync_point_designer_from_last_point_inp()  # deferred until function is defined
# Plotting (matplotlib only; keep dependencies minimal)
try:
    import matplotlib.pyplot as plt  # type: ignore
    _HAVE_MPL = True
except Exception:
    plt = None  # type: ignore
    _HAVE_MPL = False


# ---- Exit / Shutdown ----
try:
    import signal, os
    if st.sidebar.button("Exit UI", help="Stop the Streamlit server and close this UI."):
        st.sidebar.success("Shutting down...")
        # Try graceful termination first
        try:
            os.kill(os.getpid(), signal.SIGTERM)
        except Exception:
            os._exit(0)
except Exception:
    pass




import subprocess
import time

def _verification_report_paths():
    rep = os.path.join(ROOT, "verification", "report.json")
    reqs = os.path.join(ROOT, "requirements", "SHAMS_REQS.yaml")
    runner = os.path.join(ROOT, "verification", "run_verification.py")
    return rep, reqs, runner

def _verification_needs_run():
    rep, reqs, runner = _verification_report_paths()
    if not os.path.exists(rep):
        return True
    try:
        rep_m = os.path.getmtime(rep)
        deps = [p for p in [reqs, runner] if os.path.exists(p)]
        if not deps:
            return False
        dep_m = max(os.path.getmtime(p) for p in deps)
        return rep_m < dep_m
    except Exception:
        return False

def _run_verification_capture():
    """
    Run verification runner using the current Python interpreter.
    Returns: (ok: bool, stdout: str, stderr: str, seconds: float)
    """
    rep, reqs, runner = _verification_report_paths()
    t0 = time.time()
    if not os.path.exists(runner):
        return False, "", f"Missing verification runner: {runner}", 0.0
    try:
        proc = subprocess.run(
            [sys.executable, runner],
            cwd=ROOT,
            capture_output=True,
            text=True,
            check=False,
        )
        dt = time.time() - t0
        ok = (proc.returncode == 0) and os.path.exists(rep)
        return ok, (proc.stdout or ""), (proc.stderr or ""), dt
    except Exception as e:
        dt = time.time() - t0
        return False, "", f"{type(e).__name__}: {e}", dt


from phase1_core import (
    PointInputs,
    hot_ion_point,
    solve_Ip_for_H98_with_Q_match,
    solve_Ip_for_H98_with_Q_match_stream,
    solve_sparc_envelope,
    solve_for_targets,
    solve_for_targets_stream,
    SolveResult,
    optimize_design,
)
from frontier.frontier import find_nearest_feasible
from models.reference_machines import REFERENCE_MACHINES
from phase1_models import BH_COEFFS
from constraints.constraints import evaluate_constraints
from solvers.optimize import scan_feasible_and_pareto, pareto_front
from docs.variable_registry import registry_dataframe
from shams_io.run_artifact import build_run_artifact
from shams_io.plotting import plot_radial_build_from_artifact, plot_summary_pdf
from solvers.sensitivity import finite_difference_sensitivities


# --- Defensive constructor: UI may pass knobs that are absent in older/newer src/PointInputs
# This keeps the UI stable across PointInputs refactors (extra fields are ignored).
_POINTINPUTS_FIELDS = {f.name for f in fields(PointInputs)}

def make_point_inputs(**kwargs) -> PointInputs:
    """Create PointInputs using only supported dataclass fields."""
    filtered = {k: v for k, v in kwargs.items() if k in _POINTINPUTS_FIELDS}
    return PointInputs(**filtered)



# -----------------------------
# UI helpers
# -----------------------------
def _num(label: str, value: float, step: float, fmt: str = None, help: str = None, min_value=None, max_value=None, key: str = None):
    # Streamlit raises if default value is outside [min_value, max_value].
    # Clamp defensively so the UI remains robust even if bounds change.
    v = float(value)
    if min_value is not None:
        try:
            v = max(v, float(min_value))
        except Exception:
            pass
    if max_value is not None:
        try:
            v = min(v, float(max_value))
        except Exception:
            pass
    kwargs = {}
    if fmt: kwargs["format"] = fmt
    if help: kwargs["help"] = help
    if min_value is not None: kwargs["min_value"] = min_value
    if max_value is not None: kwargs["max_value"] = max_value
    return st.number_input(label, value=v, step=float(step), key=key, **kwargs) if key else st.number_input(label, value=v, step=float(step), **kwargs)


def _warn_unrealistic_point_inputs(pi: Any, context: str = "") -> None:
    """Non-blocking, UI-only warnings for obviously unrealistic user inputs.

    This must not change any model/solver behavior; it only surfaces warnings.
    """
    if pi is None:
        return
    # (lo, hi, message)
    checks = [
        ("R0_m", 0.5, 15.0, "Major radius R0 [m] looks unusual"),
        ("a_m", 0.1, 5.0, "Minor radius a [m] looks unusual"),
        ("kappa", 1.0, 3.5, "Elongation κ looks unusual"),
        ("delta", -0.8, 0.8, "Triangularity δ looks unusual"),
        ("Bt_T", 0.5, 25.0, "Toroidal field Bt [T] looks unusual"),
        ("Ip_MA", 0.1, 30.0, "Plasma current Ip [MA] looks unusual"),
        ("Ti_keV", 0.1, 40.0, "Ion temperature Ti [keV] looks unusual"),
        ("fG", 0.05, 1.5, "Greenwald fraction fG looks unusual"),
        ("Paux_MW", 0.0, 300.0, "Auxiliary power Paux [MW] looks unusual"),
        ("t_shield_m", 0.05, 2.0, "Shield thickness t_shield [m] looks unusual"),
    ]
    warns: List[str] = []
    for name, lo, hi, msg in checks:
        if not hasattr(pi, name):
            continue
        try:
            v = float(getattr(pi, name))
        except Exception:
            continue
        if (v < lo) or (v > hi):
            warns.append(f"- {msg}: **{name}={v:g}** (expected roughly {lo:g}–{hi:g})")
    if warns:
        title = "Unrealistic inputs" + (f" ({context})" if context else "")
        st.warning(title + "\n" + "\n".join(warns))


# -----------------------------
# Scan Lab parameter metadata (UI-only)
# -----------------------------

# Human-friendly physics block names (used in tooltips + mapping table)
_PHYS_BLOCKS: Dict[str, str] = {
    "Geometry": "Machine geometry / size assumptions",
    "Magnets & radial build": "TF/HTS coil build, inboard stack closure, peak field mapping, stress",
    "0-D plasma core": "0-D profiles, fusion power, temperatures, density, basic scalings",
    "Confinement": "Energy confinement (IPB98-like) + confinement multipliers",
    "H-mode access": "L-H threshold (Martin-08-like) + margin screening",
    "Stability & limits": "q95, \u03b2N, bootstrap fraction and related operational screens",
    "Power balance & radiation": "Zeff/dilution/radiation and alpha deposition assumptions",
    "Divertor / SOL": "SOL power loading proxy (PSOL/R) and divertor heat-flux screen",
    "Neutronics": "TBR proxy + HTS fluence/lifetime proxy",
    "Electrical balance": "Recirculating power closure and net electric power screen",
    "Numerics": "Solver bounds/tolerance and feasibility filtering",
}

# For Scan Lab UI: which parameters are mandatory vs optional + which physics blocks they affect.
_SCAN_PARAM_META: Dict[str, Dict[str, Any]] = {
    # Machine / plasma assumptions
    "R0": {"req": True, "blocks": ["Geometry", "Magnets & radial build", "0-D plasma core", "Divertor / SOL"]},
    "B0": {"req": True, "blocks": ["Magnets & radial build", "0-D plasma core", "Confinement", "Stability & limits"]},
    "tshield": {"req": True, "blocks": ["Magnets & radial build", "Neutronics"]},
    "Paux": {"req": True, "blocks": ["0-D plasma core", "Power balance & radiation", "H-mode access", "Electrical balance"]},
    "Paux_for_Q": {"req": True, "blocks": ["0-D plasma core"]},
    "Ti_over_Te": {"req": True, "blocks": ["0-D plasma core", "Power balance & radiation"]},

    # Axes
    "Ti": {"req": True, "blocks": ["0-D plasma core", "Confinement", "Power balance & radiation"]},
    "H98": {"req": True, "blocks": ["Confinement"]},
    "a": {"req": True, "blocks": ["Geometry", "0-D plasma core", "Stability & limits", "Divertor / SOL", "Magnets & radial build"]},
    "Q": {"req": True, "blocks": ["0-D plasma core", "Electrical balance"]},
    "g_conf": {"req": True, "blocks": ["Confinement"]},

    # Solver bounds & screens
    "Ip_bounds": {"req": True, "blocks": ["Numerics", "0-D plasma core", "Stability & limits", "Magnets & radial build"]},
    "fG_bounds": {"req": True, "blocks": ["Numerics", "0-D plasma core"]},
    "tol": {"req": True, "blocks": ["Numerics"]},

    # Screening knobs (plasma)
    "Zeff": {"req": True, "blocks": ["Power balance & radiation"]},
    "dilution_fuel": {"req": True, "blocks": ["Power balance & radiation", "0-D plasma core"]},
    "extra_rad_factor": {"req": True, "blocks": ["Power balance & radiation"]},
    "alpha_loss_frac": {"req": True, "blocks": ["Power balance & radiation"]},
    "kappa": {"req": True, "blocks": ["Stability & limits", "0-D plasma core"]},
    "q95_min": {"req": True, "blocks": ["Stability & limits"]},
    "betaN_max": {"req": True, "blocks": ["Stability & limits"]},
    "C_bs": {"req": True, "blocks": ["Stability & limits"]},
    "f_bs_max": {"req": True, "blocks": ["Stability & limits"]},
    "PSOL_over_R_max": {"req": True, "blocks": ["Divertor / SOL"]},

    # Optional toggle
    "require_Hmode": {"req": False, "blocks": ["H-mode access"]},
    "PLH_margin": {"req": False, "blocks": ["H-mode access"]},

    # Clean design knobs (engineering screens)
    "tblanket_m": {"req": False, "blocks": ["Magnets & radial build", "Neutronics"]},
    "t_vv_m": {"req": False, "blocks": ["Magnets & radial build"]},
    "t_gap_m": {"req": False, "blocks": ["Magnets & radial build"]},
    "t_tf_struct_m": {"req": False, "blocks": ["Magnets & radial build"]},
    "t_tf_wind_m": {"req": False, "blocks": ["Magnets & radial build"]},
    "Bpeak_factor": {"req": False, "blocks": ["Magnets & radial build"]},
    "sigma_allow_MPa": {"req": False, "blocks": ["Magnets & radial build"]},
    "Tcoil_K": {"req": False, "blocks": ["Magnets & radial build"]},
    "hts_margin_min": {"req": False, "blocks": ["Magnets & radial build"]},
    "Vmax_kV": {"req": False, "blocks": ["Magnets & radial build"]},
    "q_div_max_MW_m2": {"req": False, "blocks": ["Divertor / SOL"]},
    "q_midplane_max_MW_m2": {"req": False, "blocks": ["Divertor / SOL"]},
    "TBR_min": {"req": False, "blocks": ["Neutronics"]},
    "hts_lifetime_min_yr": {"req": False, "blocks": ["Neutronics"]},
    "P_net_min_MW": {"req": False, "blocks": ["Electrical balance"]},
}


def _scan_badge(param_key: str) -> str:
    meta = _SCAN_PARAM_META.get(param_key)
    # Default to optional if unknown
    is_req = bool(meta.get("req")) if isinstance(meta, dict) else False
    return "🟥 Mandatory" if is_req else "⬜ Optional"


def _scan_blocks(param_key: str) -> List[str]:
    meta = _SCAN_PARAM_META.get(param_key)
    if not isinstance(meta, dict):
        return []
    return [b for b in meta.get("blocks", []) if b in _PHYS_BLOCKS]


def _scan_label(base: str, param_key: str) -> str:
    # number_input labels do not render markdown; keep it simple + consistent.
    return f"{base}  ·  {_scan_badge(param_key)}"


def _scan_help(base_help: str, param_key: str) -> str:
    blocks = _scan_blocks(param_key)
    if not blocks:
        return base_help
    lines = [base_help.strip(), "", "Maps to physics blocks:"]
    for b in blocks:
        lines.append(f"- {b}: {_PHYS_BLOCKS[b]}")
    return "\n".join(lines).strip()

def kpi_row(items: List[Tuple[str, Any]]):
    cols = st.columns(len(items))
    for c, (k, v) in zip(cols, items):
        c.metric(k, v)


def _numeric_cols(df: pd.DataFrame) -> List[str]:
    cols = []
    for c in df.columns:
        try:
            if pd.api.types.is_numeric_dtype(df[c]):
                cols.append(c)
        except Exception:
            pass
    return cols


def plot_scatter(df: pd.DataFrame, x: str, y: str, color: str | None = None, title: str | None = None):
    """Matplotlib scatter with optional numeric color."""
    if df is None or df.empty:
        st.info("No data to plot.")
        return
    if x not in df or y not in df:
        st.warning("Select valid x/y columns.")
        return

    d = df[[x, y] + ([color] if color and color in df else [])].dropna()
    if d.empty:
        st.info("No finite rows for this plot (after dropping NaNs).")
        return
    # Matplotlib is optional: if missing, fall back to Streamlit's built-in charts.
    if not _HAVE_MPL:
        st.warning("Plotting is limited because 'matplotlib' is not installed. Install it (pip install matplotlib) for full plotting.")
        # Streamlit fallback (no colorbar support)
        try:
            st.scatter_chart(d, x=x, y=y)
        except Exception:
            st.line_chart(d[[x, y]].rename(columns={x: "x", y: "y"}))
        return

    fig = plt.figure(figsize=(6.8, 4.6))
    ax = plt.gca()
    if color and color in d and pd.api.types.is_numeric_dtype(d[color]):
        sc = ax.scatter(d[x], d[y], c=d[color], s=22, alpha=0.85)
        cb = plt.colorbar(sc, ax=ax)
        cb.set_label(color)
    else:
        ax.scatter(d[x], d[y], s=22, alpha=0.85)

    ax.set_xlabel(x)
    ax.set_ylabel(y)
    if title:
        ax.set_title(title)
    ax.grid(True, alpha=0.25)
    st.pyplot(fig, clear_figure=True)


def plot_bars(values: Dict[str, float], title: str):
    keys = [k for k in values.keys() if isinstance(values.get(k), (int, float)) and math.isfinite(float(values.get(k))) ]
    if not keys:
        st.caption("No plottable values available.")
        return
    if not _HAVE_MPL:
        st.warning("Bar charts are limited because 'matplotlib' is not installed. Install it (pip install matplotlib) for full plotting.")
        import pandas as _pd
        s = _pd.Series({k: float(values[k]) for k in keys})
        try:
            st.bar_chart(s)
        except Exception:
            st.write(s)
        return

    fig = plt.figure(figsize=(6.8, 4.4))
    ax = plt.gca()
    ax.bar(range(len(keys)), [float(values[k]) for k in keys])
    ax.set_xticks(range(len(keys)), keys, rotation=35, ha="right")
    ax.set_title(title)
    ax.grid(True, axis="y", alpha=0.25)
    st.pyplot(fig, clear_figure=True)

def badge(check):
    """Render PASS/FAIL/WARN/SKIPPED badge.
    Accepts either a check dict with 'status' or a legacy ok flag.
    """
    if isinstance(check, dict):
        stt = check.get('status')
        if stt == 'SKIPPED':
            return '⚪ SKIPPED'
        if stt == 'WARN':
            return '🟡 WARN'
        if stt == 'FAIL':
            return '❌ FAIL'
        if stt == 'PASS':
            return '✅ PASS'
        # fallback
        ok = check.get('ok')
    else:
        ok = check
    if ok is None:
        return '⚪ SKIPPED'
    return '✅ PASS' if ok else '❌ FAIL'


def finite(x):
    return isinstance(x, (int, float)) and math.isfinite(x)

def top_violations(checks: List[Dict[str, Any]], n: int = 3) -> List[Dict[str, Any]]:
    bad = [c for c in checks if c.get('status') == 'FAIL']

    # sort by relative violation if available
    def score(c):
        if c.get("value") is None or c.get("limit") is None:
            return 0.0
        v, lim = c["value"], c["limit"]
        if not (finite(v) and finite(lim)) or lim == 0:
            return 0.0
        if c.get("sense") == "max":
            return (v - lim) / abs(lim)
        if c.get("sense") == "min":
            return (lim - v) / abs(lim)
        return 0.0
    bad.sort(key=score, reverse=True)
    return bad[:n]

def top_warnings(checks: List[Dict[str, Any]], n: int = 3) -> List[Dict[str, Any]]:
    ws = [c for c in checks if c.get('status') == 'WARN']
    def score(c):
        v = c.get('value'); lim = c.get('limit'); sense = c.get('sense')
        if v is None or lim is None or not finite(v) or not finite(lim):
            return 0.0
        if sense == 'max':
            return max(0.0, (v - lim) / abs(lim))
        if sense == 'min':
            return max(0.0, (lim - v) / abs(lim))
        return 0.0
    ws.sort(key=score, reverse=True)
    return ws[:n]

def compute_checks(out: Dict[str, float]) -> List[Dict[str, Any]]:
    """
    Turn flat outputs into a structured constraint list.

    IMPORTANT:
    - Checks are only evaluated when the needed physics is enabled *and* the value is finite.
    - If a model/physics block is disabled (or produced NaN), the check is marked SKIPPED.
    - Checks can be WARNING-level or HARD FAIL depending on how far they are from the limit.
    """
    checks: List[Dict[str, Any]] = []


    # global warning behavior knobs (can be overridden per-check via explicit warn_limit)
    warn_frac_max = float(out.get("_warn_frac_max", 0.90))  # for max constraints: WARN if v > warn_frac*limit
    warn_frac_min = float(out.get("_warn_frac_min", 1.10))  # for min constraints: WARN if v < warn_frac*limit

    def add(name, status, value=None, limit=None, sense=None, notes="", severity="hard", warn_limit=None):
        """status in {'PASS','FAIL','WARN','SKIPPED'}"""
        ok = None
        if status == "PASS":
            ok = True
        elif status == "FAIL":
            ok = False
        elif status == "WARN":
            ok = True  # warn is still 'ok' for legacy consumers
        elif status == "SKIPPED":
            ok = None
        checks.append({
            "name": name,
            "status": status,
            "ok": ok,
            "value": value,
            "limit": limit,
            "warn_limit": warn_limit,
            "sense": sense,
            "notes": notes,
            "severity": severity,
        })

        # ok can be True/False/None (None => skipped)
        checks.append({"name": name, "ok": ok if ok in (True, False, None) else bool(ok),
                       "value": value, "limit": limit, "sense": sense, "notes": notes})

    def fin(x) -> bool:
        return isinstance(x, (int, float)) and math.isfinite(x)
    def eval_max(name, key_value, key_limit, notes="", severity="hard", warn_limit=None):
        v = out.get(key_value)
        lim = out.get(key_limit)
        if (not fin(v)) or (not fin(lim)):
            add(name, "SKIPPED", v, lim, "max", notes, severity=severity, warn_limit=warn_limit)
            return
        wl = warn_limit
        if wl is None:
            wl = warn_frac_max * lim
        if v > lim:
            add(name, "FAIL", v, lim, "max", notes, severity=severity, warn_limit=wl)
        elif v > wl:
            add(name, "WARN", v, lim, "max", notes, severity=severity, warn_limit=wl)
        else:
            add(name, "PASS", v, lim, "max", notes, severity=severity, warn_limit=wl)

    def eval_min(name, key_value, key_limit, notes="", severity="hard", warn_limit=None):
        v = out.get(key_value)
        lim = out.get(key_limit)
        if (not fin(v)) or (not fin(lim)):
            add(name, "SKIPPED", v, lim, "min", notes, severity=severity, warn_limit=warn_limit)
            return
        wl = warn_limit
        if wl is None:
            wl = warn_frac_min * lim
        # for min constraints, warn if v is between lim and wl (wl > lim)
        if v < lim:
            add(name, "FAIL", v, lim, "min", notes, severity=severity, warn_limit=wl)
        elif v < wl:
            add(name, "WARN", v, lim, "min", notes, severity=severity, warn_limit=wl)
        else:
            add(name, "PASS", v, lim, "min", notes, severity=severity, warn_limit=wl)    # --- Build closure ---
    if "radial_build_ok" in out:
        v = out.get("radial_build_ok")
        if not fin(v):
            add("Radial build closure", "SKIPPED", out.get("rb_total_inboard_m"), out.get("rinboard_available_m"), "max",
                "Requires inboard stack thickness ≤ available inboard space (R0 - a).")
        else:
            add("Radial build closure", "PASS" if (v > 0.5) else "FAIL", out.get("rb_total_inboard_m"), out.get("rinboard_available_m"), "max",
                "Requires inboard stack thickness ≤ available inboard space (R0 - a).")

    # --- Magnet stress ---


    if "sigma_hoop_MPa" in out and "sigma_allow_MPa" in out:
        eval_max("TF hoop stress", "sigma_hoop_MPa", "sigma_allow_MPa",
                 "Hoop stress proxy must be below allowable structural stress.")

    # --- HTS margin ---
    if "hts_margin" in out and "hts_margin_min" in out:
        eval_min("HTS margin", "hts_margin", "hts_margin_min",
                 "HTS operating margin proxy vs (B,T) must exceed minimum.")

    # --- Dump voltage ---
    if "V_dump_kV" in out and "Vmax_kV" in out:
        eval_max("Dump voltage", "V_dump_kV", "Vmax_kV",
                 "Fast discharge voltage must not exceed protection limit.")

    # --- Divertor heat flux ---

    if "q_div_MW_m2" in out and "q_div_max_MW_m2" in out:
        eval_max("Divertor heat flux", "q_div_MW_m2", "q_div_max_MW_m2",
                 "Peak divertor heat flux proxy must be below limit.")

    # --- Tritium breeding ratio ---
    if "TBR" in out and "TBR_min" in out:
        eval_min("TBR", "TBR", "TBR_min",
                 "Tritium breeding ratio proxy must exceed minimum.")

    # --- HTS lifetime ---
    if "hts_lifetime_yr" in out and "hts_lifetime_min_yr" in out:
        eval_min("HTS lifetime", "hts_lifetime_yr", "hts_lifetime_min_yr",
                 "Neutron lifetime proxy of HTS must exceed minimum.")

    # --- Net electric power ---
    if "P_net_e_MW" in out and "P_net_min_MW" in out:
        eval_min("Net electric power", "P_net_e_MW", "P_net_min_MW",
                 "Net electric power must exceed minimum (system closure).")

    # --- H-mode access (only if enforced) ---
    # If require_Hmode is False or physics is disabled, this becomes SKIPPED.
    if "require_Hmode" in out and "LH_ok" in out:
        req = out.get("require_Hmode")
        lh_ok = out.get("LH_ok")
        if not (fin(req) and req > 0.5):
            add("H‑mode access", None, lh_ok, 1.0, "min",
                "H‑mode not required (or LH physics disabled).")
        else:
            if not fin(lh_ok):
                add("H‑mode access", None, lh_ok, 1.0, "min",
                    "H‑mode required, but LH physics not available for this point.")
            else:
                add("H‑mode access", lh_ok > 0.5, lh_ok, 1.0, "min",
                    "If H‑mode required, point must be above LH threshold with margin.")

    return checks


# -----------------------------
# Scan runner (UI-native)
# -----------------------------
def frange(start: float, stop: float, step: float) -> List[float]:
    vals: List[float] = []
    if step == 0:
        return [start]
    x = start
    if step > 0:
        while x <= stop + 1e-12:
            vals.append(float(x))
            x += step
    else:
        while x >= stop - 1e-12:
            vals.append(float(x))
            x += step
    return vals

def run_scan(spec: Dict[str, Any]) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    A refactor of the CLI scan loop into a UI-callable function.

    Returns:
      df_feasible: rows of extended-feasible points (same spirit as 'feasible_ext' sheet)
      meta: dict with scan settings + summary stats
    """
    Ti_grid = frange(spec["Ti_start"], spec["Ti_stop"], spec["Ti_step"] if spec["Ti_stop"] >= spec["Ti_start"] else -abs(spec["Ti_step"]))
    H_grid = frange(spec["H98_start"], spec["H98_stop"], abs(spec["H98_step"]))
    a_grid = frange(spec["a_min"], spec["a_max"], abs(spec["a_step"]))
    Q_grid = frange(spec["Q_start"], spec["Q_stop"], abs(spec["Q_step"]))
    g_grid = frange(spec["gconf_start"], spec["gconf_stop"], abs(spec["gconf_step"]))


    # --- Scan Lab: optional UI progress + logging hooks (kept no-op for non-UI use) ---
    progress_cb = spec.get("_progress_cb")  # callable(fraction: float, info: dict) -> None
    log_cb = spec.get("_log_cb")            # callable(line: str) -> None
    log_lines: List[str] = []

    def _log(line: str) -> None:
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        s = f"[{ts}] {line}"
        log_lines.append(s)
        if callable(log_cb):
            try:
                log_cb(s)
            except Exception:
                pass

    def _progress(i: int, n: int, **info: Any) -> None:
        if callable(progress_cb) and n > 0:
            try:
                progress_cb(min(max(i / n, 0.0), 1.0), info)
            except Exception:
                pass

    n_total = max(1, len(g_grid) * len(Ti_grid) * len(H_grid) * len(a_grid) * len(Q_grid))
    _log(f"Scan initialized: {len(g_grid)} g_conf × {len(Ti_grid)} Ti × {len(H_grid)} H98 × {len(a_grid)} a × {len(Q_grid)} Q  => {n_total} evaluations")
    i_eval = 0

    rows: List[Dict[str, Any]] = []
    best_g = None


    for g_conf in g_grid:
        for Ti in Ti_grid:
            for Hreq in H_grid:
                for a in a_grid:
                    for Qtar in Q_grid:
                        # Solve at reduced target (same logic as CLI driver)

                        i_eval += 1
                        _progress(i_eval, n_total,
                                  stage="setup",
                                  g_conf=float(g_conf), Ti_keV=float(Ti), H98_req=float(Hreq), a_m=float(a), Q_target=float(Qtar))
                        _log(f"Eval {i_eval}/{n_total}: g_conf={g_conf:.3g}, Ti={Ti:.3g} keV, H98_req={Hreq:.3g}, a={a:.3g} m, Q={Qtar:.3g}")
                        _log("  - Building point inputs (geometry, fields, density/temperature assumptions)")

                        H_base_target = Hreq / max(g_conf, 1e-9)

                        base = make_point_inputs(
                            R0_m=spec["R0"],
                            a_m=a,
                            kappa=spec["kappa"],
                            Bt_T=spec["B0"],
                            Ip_MA=0.5*(spec["Ip_min"]+spec["Ip_max"]),
                            Ti_keV=Ti,
                            fG=0.8,
                            t_shield_m=spec["tshield"],
                            Paux_MW=spec["Paux"],
                            Ti_over_Te=spec["Ti_over_Te"],
                            zeff=spec["Zeff"],
                            dilution_fuel=spec["dilution_fuel"],
                            fuel_mode="DT",
                            include_secondary_DT=include_secondary_DT,
                            tritium_retention=0.5,
                            tau_T_loss_s=5.0,
                            extra_rad_factor=spec["extra_rad_factor"],
                            alpha_loss_frac=spec["alpha_loss_frac"],
                            C_bs=spec["C_bs"],
                            require_Hmode=spec["require_Hmode"],
                            PLH_margin=spec["PLH_margin"],
                            # --- Clean design knobs (passed through PointInputs defaults if present in your src)
                            **spec.get("clean_knobs", {}),
                        )


                        _log("  - Solving nested system: outer Ip for H98, inner fG for Q (bisection)")
                        _progress(i_eval, n_total, stage="solve")
                        sol_inp, sol_out, ok = solve_Ip_for_H98_with_Q_match(
                            base=base,
                            target_H98=H_base_target,
                            target_Q=Qtar,
                            Ip_min=spec["Ip_min"],
                            Ip_max=spec["Ip_max"],
                            fG_min=spec["fG_min"],
                            fG_max=spec["fG_max"],
                            tol=spec["tol"],
                            Paux_for_Q_MW=spec["Paux_for_Q"],
                        )
                        if not ok:
                            _log("  - Solver failed to bracket/converge for this combo (skipping)")
                            continue

                        # Effective confinement
                        H98_eff = g_conf * sol_out["H98"]
                        sol_out["H98_eff"] = H98_eff


                        _log("  - Evaluating physics proxies (power balance, confinement, operational limits)")
                        _progress(i_eval, n_total, stage="evaluate")
                        # Standard ext checks from CLI
                        ok_ext = True
                        if sol_out["ne20"] > 1.2:
                            ok_ext = False
                        if sol_out.get("q95_proxy", 1e9) < spec["q95_min"]:
                            ok_ext = False
                        if sol_out.get("betaN_proxy", 0.0) > spec["betaN_max"]:
                            ok_ext = False
                        if sol_out.get("f_bs_proxy", 0.0) > spec["f_bs_max"]:
                            ok_ext = False
                        PSOL_over_R = sol_out["Ploss_MW"] / spec["R0"]
                        sol_out["PSOL_over_R"] = PSOL_over_R
                        if PSOL_over_R > spec["PSOL_over_R_max"]:
                            ok_ext = False
                        if spec["require_Hmode"] and sol_out.get("LH_ok", 1.0) < 0.5:
                            ok_ext = False
                        if H98_eff < Hreq:
                            ok_ext = False
                        if sol_out["Q_DT_eqv"] < Qtar:
                            ok_ext = False

                        # Clean design checks (if present)
                        checks = compute_checks(sol_out)
                        if any((not c["ok"]) for c in checks):
                            ok_ext = False

                        if not ok_ext:
                            _log("  - Failed screening checks (skipping)")
                            continue

                        if best_g is None or g_conf < best_g:
                            best_g = g_conf

                        _log("  - Feasible point found ✓ (adding to results)")
                        _progress(i_eval, n_total, stage="record")
                        row = dict(sol_out)
                        row.update({
                            "g_conf": g_conf,
                            "Ti_keV": Ti,
                            "Q_target": Qtar,
                            "H98_required": Hreq,
                            "a_m": a,
                            "Ip_MA": sol_inp.Ip_MA,
                            "f_G": sol_inp.fG,
                            "Paux_MW": sol_inp.Paux_MW,
                            "Paux_for_Q_MW": spec["Paux_for_Q"],
                            "H98_eff": H98_eff,
                        })
                        rows.append(row)

    df = pd.DataFrame(rows)
    meta = dict(spec)
    meta["best_g_conf_found"] = best_g if best_g is not None else "NONE"
    meta["n_feasible"] = int(len(df))

    # Strip non-serializable UI callbacks from meta and attach log text
    meta.pop("_progress_cb", None)
    meta.pop("_log_cb", None)
    meta["scan_log_text"] = "\n".join(log_lines)
    _log(f"Scan complete: feasible={meta['n_feasible']}  best_g_conf_found={meta['best_g_conf_found']}")

    return df, meta

def df_to_excel_bytes(df: pd.DataFrame, meta: Dict[str, Any]) -> bytes:
    """
    Export feasible dataframe + meta into an Excel workbook (in-memory).
    """
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "feasible_ext"

    if df.empty:
        ws.append(["NO_FEASIBLE_POINTS"])
    else:
        ws.append(list(df.columns))
        for c in range(1, len(df.columns)+1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        ws.freeze_panes = "A2"
        for _, r in df.iterrows():
            ws.append([r.get(c) for c in df.columns])

    # meta sheet
    wsM = wb.create_sheet("meta")
    wsM.append(["key", "value"])
    wsM["A1"].font = Font(bold=True)
    for k, v in meta.items():
        # keep meta compact
        if isinstance(v, dict):
            continue
        if isinstance(v, list):
            continue
        wsM.append([k, v])
    wsM.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -----------------------------
# Streamlit UI
# -----------------------------
st.set_page_config(page_title="HTS Compact Tokamak Design Studio", layout="wide")

# ---- UI polish (CSS) ----
# Streamlit theming is usually done via .streamlit/config.toml, but we keep it self-contained.
st.markdown(
    """
<style>
  /* Slightly tighter overall spacing */
  .block-container { padding-top: 1.2rem; padding-bottom: 2.0rem; }

  /* Nicer header spacing */
  h1, h2, h3 { letter-spacing: -0.02em; }

  /* Metric cards: increase contrast and soften edges */
  [data-testid="stMetric"] {
    padding: 0.75rem 0.75rem;
    border-radius: 16px;
    border: 1px solid rgba(49, 51, 63, 0.15);
    background: rgba(255, 255, 255, 0.03);
  }

  /* Buttons */
  div.stButton > button {
    border-radius: 14px;
    padding: 0.55rem 0.9rem;
    font-weight: 650;
  }

  /* Expander */
  details {
    border-radius: 14px;
    border: 1px solid rgba(49, 51, 63, 0.12);
    padding: 0.25rem 0.4rem;
  }

  /* Code blocks */
  pre {
    border-radius: 14px;
  }

  /* Dataframes */
  .stDataFrame { border-radius: 14px; overflow: hidden; }
</style>
    """,
    unsafe_allow_html=True,
)
st.title("HTS Compact Tokamak — Clean 0‑D Point Design Studio")

st.sidebar.header("Mode")
st.session_state.explain_mode = st.sidebar.toggle(
    "Explain mode (show equations & reasons)",
    value=bool(st.session_state.get("explain_mode", True)),
    help="Teaching mode: show model equations, assumptions, and why constraints bind."
)
st.session_state.expert_mode = st.sidebar.toggle(
    "Expert controls",
    value=bool(st.session_state.get("expert_mode", False)),
    help="Expose solver tolerances and optimizer internals."
)




# ---------------------------------------------------------------------------
# Verification (requirements compliance)
# ---------------------------------------------------------------------------
with st.sidebar.expander("Verification (requirements)", expanded=False):
    rep_path, reqs_path, runner_path = _verification_report_paths()

    st.session_state["auto_verify"] = st.checkbox(
        "Auto-run verification when needed",
        value=bool(st.session_state.get("auto_verify", True)),
        help="Runs python verification/run_verification.py if report.json is missing or older than the requirements/runner.",
    )

    needs = _verification_needs_run()
    st.caption(("Report status: " + ("needs update" if needs else "up to date")) if os.path.exists(rep_path) else "Report status: missing")

    colv1, colv2 = st.columns(2)
    force = colv1.button("Run now", use_container_width=True)
    show_logs = colv2.toggle("Show logs", value=bool(st.session_state.get("verify_show_logs", False)))
    st.session_state["verify_show_logs"] = show_logs

    # Auto-run once per session if enabled
    if st.session_state.get("auto_verify", True) and needs and not st.session_state.get("_auto_verify_done", False):
        with st.spinner("Running verification..."):
            ok, out, err, dt = _run_verification_capture()
        st.session_state["_auto_verify_done"] = True
        st.session_state["_last_verify_ok"] = ok
        st.session_state["_last_verify_out"] = out
        st.session_state["_last_verify_err"] = err
        st.session_state["_last_verify_dt"] = dt

    if force:
        with st.spinner("Running verification..."):
            ok, out, err, dt = _run_verification_capture()
        st.session_state["_auto_verify_done"] = True
        st.session_state["_last_verify_ok"] = ok
        st.session_state["_last_verify_out"] = out
        st.session_state["_last_verify_err"] = err
        st.session_state["_last_verify_dt"] = dt
        st.rerun()

    if st.session_state.get("_last_verify_dt") is not None:
        ok = bool(st.session_state.get("_last_verify_ok", False))
        st.success(f"Last run: {'PASS' if ok else 'FAIL'} ({st.session_state.get('_last_verify_dt', 0.0):.2f}s)") if ok else st.error(
            f"Last run: FAIL ({st.session_state.get('_last_verify_dt', 0.0):.2f}s)"
        )

    if show_logs:
        st.text_area("stdout", value=str(st.session_state.get("_last_verify_out", "")), height=160)
        st.text_area("stderr", value=str(st.session_state.get("_last_verify_err", "")), height=160)


# ---------------------------------------------------------------------------
# Fidelity + Calibration (PROCESS-inspired, transparent)
# ---------------------------------------------------------------------------
with st.sidebar.expander("Fidelity (model detail)", expanded=False):
    fid = st.session_state.get("fidelity_config", {})
    plasma = st.selectbox("Plasma", ["0D","1/2D"], index=0 if fid.get("plasma","0D")=="0D" else 1)
    magnets = st.selectbox("Magnets", ["limits","stress"], index=0 if fid.get("magnets","limits")=="limits" else 1)
    exhaust = st.selectbox("Exhaust", ["proxy","enriched"], index=0 if fid.get("exhaust","proxy")=="proxy" else 1)
    neutronics = st.selectbox("Neutronics", ["proxy","enriched"], index=0 if fid.get("neutronics","proxy")=="proxy" else 1)
    profiles = st.selectbox("Profiles", ["off","analytic"], index=0 if fid.get("profiles","off")=="off" else 1)
    economics = st.selectbox("Economics", ["proxy","enriched"], index=0 if fid.get("economics","proxy")=="proxy" else 1)
    st.session_state["fidelity_config"] = {
        "plasma": plasma,
        "magnets": magnets,
        "exhaust": exhaust,
        "neutronics": neutronics,
        "profiles": profiles,
        "economics": economics,
    }

with st.sidebar.expander("Reference calibration (optional)", expanded=False):
    st.caption("Transparent multiplicative factors (default 1.0). Not a black-box fit.")
    st.session_state["calib_confinement"] = st.slider("Confinement factor", 0.5, 1.5, float(st.session_state.get("calib_confinement", 1.0)), 0.01)
    st.session_state["calib_divertor"] = st.slider("Divertor factor", 0.5, 1.5, float(st.session_state.get("calib_divertor", 1.0)), 0.01)
    st.session_state["calib_bootstrap"] = st.slider("Bootstrap factor", 0.5, 1.5, float(st.session_state.get("calib_bootstrap", 1.0)), 0.01)

st.sidebar.markdown("### Reference presets")
_preset_names = list(REFERENCE_MACHINES.keys())

# ---- Performance (opt-in; no effect on results unless enabled) ----
with st.sidebar.expander("Performance", expanded=False):
    cache_enabled = st.checkbox("Enable evaluation cache", value=True, help="Read-only memoization; does not change numerical results.")
    cache_max = int(st.number_input("Cache size (LRU)", value=256, min_value=0, max_value=100000, step=64, help="Max cached evaluations. 0 disables effectively."))
    solver_backend = st.selectbox("Solver backend (Systems Mode)", ["hybrid_newton", "broyden"], index=0, help="Default preserves existing behavior. Broyden is opt-in for speed.")
    show_perf_panel = st.checkbox("Show Performance Panel", value=True, help="Displays cache stats, solver backend, wall time, and iteration counts.")

_preset_sel = st.sidebar.selectbox("Preset", _preset_names, index=0, key="ref_preset_sel")
if st.sidebar.button("Load preset", key="load_ref_preset"):
    _d = dict(REFERENCE_MACHINES.get(_preset_sel, {}))
    # Ensure required core fields exist; fall back to SPARC-class defaults
    _fallback = REFERENCE_MACHINES.get("SPARC-class (compact HTS)", {})
    for _k in ["R0_m","a_m","kappa","Bt_T","Ip_MA","Ti_keV","fG","Paux_MW"]:
        if _k not in _d:
            _d[_k] = _fallback.get(_k)
    try:
        st.session_state.last_point_inp = PointInputs(**_d)
        st.session_state["pd_needs_sync"] = True
        # Update Point Designer widget state so visible inputs change immediately.
        # Streamlit widgets keep their own state once created; setting these keys
        # before rerun ensures the UI reflects the preset.
        _key_map = {
            "R0_m": PD_KEYS["R0_m"],
            "a_m": PD_KEYS["a_m"],
            "kappa": PD_KEYS["kappa"],
            "Bt_T": PD_KEYS["Bt_T"],
            "Ti_keV": PD_KEYS["Ti_keV"],
            "Paux_MW": PD_KEYS["Paux_MW"],
        }
        for _k, _key in _key_map.items():
            try:
                if _k in _d and _d[_k] is not None:
                    st.session_state[_key] = float(_d[_k])
            except Exception:
                pass

        # Presets provide nominal Ip and fG; Point Designer uses bounds for the target solver.
        # Set reasonable bounds around the preset nominal values.
        try:
            ip = float(_d.get("Ip_MA", 0.0))
            if ip > 0:
                st.session_state[PD_KEYS["Ip_lo"]] = max(0.1, 0.80 * ip)
                st.session_state[PD_KEYS["Ip_hi"]] = max(0.2, 1.20 * ip)
        except Exception:
            pass
        try:
            fg = float(_d.get("fG", 0.0))
            if fg > 0:
                st.session_state[PD_KEYS["fG_lo"]] = max(0.0, fg - 0.20)
                st.session_state[PD_KEYS["fG_hi"]] = min(2.0, fg + 0.20)
        except Exception:
            pass

        # Also keep Q denominator aligned with Paux by default.
        # Use an explicit key to avoid label-based widget-key drift.
        try:
            st.session_state[PD_KEYS["Paux_for_Q"]] = float(_d.get("Paux_MW", 0.0))
        except Exception:
            pass

        # Keep Ti/Te at 1.0 unless a preset explicitly provides it.
        try:
            if "Ti_over_Te" in _d and _d["Ti_over_Te"] is not None:
                st.session_state[PD_KEYS["Ti_over_Te"]] = float(_d["Ti_over_Te"])
        except Exception:
            pass
        st.sidebar.success(f"Loaded: {_preset_sel}")
        st.rerun()
    except Exception as _e:
        st.sidebar.error(f"Preset failed to load: {_e}")

tab_point, tab_systems, tab_scan, tab_pareto, tab_compare, tab_more = st.tabs([
    "Point Designer",
    "Systems Mode",
    "Scan Lab",
    "Pareto Lab",
    "Compare",
    "More",
])

with tab_more:
    st.caption("Reference, documentation, validation, artifacts, and advanced tools.")
    tab_studies = st.expander("Studies", expanded=False)
    tab_model = st.expander("0-D Physics Model", expanded=False)
    tab_bench = st.expander("Benchmarks", expanded=False)
    tab_registry = st.expander("Variable Registry", expanded=False)
    tab_validation = st.expander("Validation", expanded=False)
    tab_compliance = st.expander("Compliance", expanded=False)
    tab_docs = st.expander("Docs", expanded=False)
    tab_artifacts = st.expander("Artifacts Explorer", expanded=False)
    tab_deck = st.expander("Case Deck Runner", expanded=False)
    tab_delta = st.expander("Scenario Delta Viewer", expanded=False)
    tab_library = st.expander("Run Library", expanded=False)
    tab_constraints = st.expander("Constraint Cockpit", expanded=False)
    tab_constraint_inspector = st.expander("Constraint Inspector", expanded=False)
    tab_sensitivity = st.expander("Sensitivity Explorer", expanded=False)
    tab_feasmap = st.expander("Feasibility Map", expanded=False)
    tab_decision = st.expander("Decision Builder", expanded=False)
    tab_nonfeas = st.expander("Non-Feasibility Guide", expanded=False)
    tab_cprov = st.expander("Constraint Provenance", expanded=False)
    tab_knobs = st.expander("Knob Trade-Space", expanded=False)
    tab_regress = st.expander("Regression Viewer", expanded=False)
    tab_study_dash = st.expander("Study Dashboard", expanded=False)
    tab_maturity = st.expander("Maturity Heatmap", expanded=False)
    tab_assumptions = st.expander("Assumption Toggles", expanded=False)
    tab_export = st.expander("Export / Share", expanded=False)
    tab_solver = st.expander("Solver Introspection", expanded=False)

# Shared state
if "last_point_out" not in st.session_state:
    st.session_state.last_point_out = None
if "scan_df" not in st.session_state:
    st.session_state.scan_df = pd.DataFrame()
if "scan_meta" not in st.session_state:
    st.session_state.scan_meta = {}
if "studies" not in st.session_state:
    st.session_state.studies = []  # list of study config dicts
if "compare_artifacts" not in st.session_state:
    st.session_state.compare_artifacts = {"A": None, "B": None}

# -----------------------------
# Point Designer
# -----------------------------
with tab_point:
    left, mid, right = st.columns([1.1, 1.6, 1.1])

    # Use the latest loaded preset / last point as the UI default for Point Designer.
    # This makes preset loads robust even if widget state keys change or are newly created.
    _base_pd = st.session_state.get("last_point_inp")
    if _base_pd is None:
        _base_pd = PointInputs(R0_m=1.81, a_m=0.62, kappa=1.8, Bt_T=10.0, Ip_MA=8.0, Ti_keV=10.0, fG=0.8, Paux_MW=50.0)

    with left:
        st.subheader("Inputs")
        with st.expander("Plasma & geometry", expanded=True):
            R0 = _num("Major radius R₀ (m)", float(_base_pd.R0_m), 0.01, help="Distance from tokamak centerline to plasma magnetic axis (major radius).", key=PD_KEYS["R0_m"])
            a = _num("Minor radius a (m)", float(_base_pd.a_m), 0.01, min_value=0.1, help="Plasma minor radius (a). Together with R₀ sets aspect ratio.", key=PD_KEYS["a_m"])
            kappa = _num("Elongation κ (–)", float(_base_pd.kappa), 0.05, min_value=1.0, max_value=3.2, help="Plasma elongation κ. Used in volume/area and stability proxies.", key=PD_KEYS["kappa"])
            delta = _num("Triangularity δ (–)", float(getattr(_base_pd, "delta", 0.0) or 0.0), 0.02, min_value=0.0, max_value=0.8, help="Triangularity δ. Used only in the transparent inboard radial-build clearance proxy (stack closure). Default 0.0 preserves legacy behavior.", key=PD_KEYS["delta"])
            B0 = _num("Toroidal field on axis B₀ (T)", float(_base_pd.Bt_T), 0.1, min_value=0.5, max_value=25.0, help="Toroidal field at plasma axis (B₀). Drives confinement and magnet sizing.", key=PD_KEYS["Bt_T"])
            Ti = _num("Ion temperature Tᵢ (keV)", float(_base_pd.Ti_keV), 0.25, min_value=1.0, max_value=40.0, help="Core ion temperature proxy. Drives fusion reactivity and stored energy.", key=PD_KEYS["Ti_keV"])
            Ti_over_Te = _num("Ion-to-electron temperature ratio Tᵢ/Tₑ (–)", float(getattr(_base_pd, "Ti_over_Te", 1.0)), 0.1, min_value=0.5, help="Assumed ratio Tᵢ/Tₑ; sets electron temperature for radiation estimate.", key=PD_KEYS["Ti_over_Te"])
        with st.expander("Model options (PROCESS-like)", expanded=False):
            confinement_model = st.selectbox(
                "Confinement scaling",
                options=["ipb98y2", "iter89p"],
                index=0,
                help="Select confinement scaling used to compute H-factor. Default preserves current behavior."
            )

            confinement_scaling = st.selectbox(
                "H-factor reference scaling (for H_scaling)",
                options=["IPB98y2", "ITER89P", "KG", "NEOALC", "SHIMOMURA", "MIRNOV"],
                index=0,
                help="Select which empirical scaling is used for the reported H_scaling comparator. H98 remains relative to IPB98(y,2).",
            )
            profile_model = st.selectbox(
                "Analytic profiles (½-D scaffold)",
                options=["none", "parabolic", "pedestal"],
                index=0,
                help="If enabled, SHAMS computes simple analytic profiles and adds profile-integrated fusion diagnostics."
            )
            profile_peaking_ne = _num("nₑ peaking (alpha)", 1.0, 0.1, min_value=0.0, help="Parabolic/pedestal core peaking control for density.")
            profile_peaking_T = _num("T peaking (alpha)", 1.5, 0.1, min_value=0.0, help="Parabolic/pedestal core peaking control for temperature.")
            bootstrap_model = st.selectbox(
                "Bootstrap proxy model",
                options=["proxy", "improved"],
                index=0,
                help="Select bootstrap fraction proxy. 'proxy' preserves current behavior."
            )



        with st.expander("Power & composition", expanded=True):
            Paux = _num("Auxiliary heating power P_aux (MW)", float(_base_pd.Paux_MW), 1.0, min_value=0.0, max_value=500.0, help="Auxiliary heating power delivered to the plasma (MW).", key=PD_KEYS["Paux_MW"])
            Paux_for_Q = _num("Aux power used in Q definition (MW)", float(getattr(_base_pd, "Paux_MW", 0.0)), 1.0, min_value=0.0, help="Denominator power for Q = P_fus,DT(adj)/P_aux_for_Q (MW).", key=PD_KEYS["Paux_for_Q"])

            with st.expander("Physics include/exclude", expanded=True):
                st.caption("Disable a block to SKIP its related physics *and* its checks.")
                include_radiation = st.checkbox("Include core radiation + impurities/dilution model", value=True)
                include_alpha_loss = st.checkbox("Include alpha-loss fraction model", value=True)
                include_hmode_physics = st.checkbox("Include H-mode access physics (P_LH / LH_ok)", value=True)
                use_lambda_q = st.checkbox("Include SOL width (λq) proxy", value=True)

            if include_radiation:
                Zeff = _num("Effective charge Z_eff (–)", 1.5, 0.1, min_value=1.0, help="Effective ion charge Z_eff; used for brems proxy (diagnostic) and radiation screens when enabled.")
                dilution_fuel = _num("Fuel dilution fraction (DT-equivalent) (–)", 0.85, 0.01, min_value=0.0, max_value=1.0, help="Multiplicative penalty on DT-equivalent fusion power due to dilution/impurities.")
                f_rad_core = _num("Core radiation fraction f_rad,core (–)", 0.20, 0.01, min_value=0.0, max_value=0.95, help="If enabled, Prad_core = f_rad_core * Pin (simple screening model).")
            
                radiation_model = st.selectbox(
                    "Radiation model",
                    options=["fractional", "physics"],
                    index=0,
                    help="fractional: Prad_core = f_rad_core * Pin (legacy). physics: brem + (optional) synchrotron + simple impurity line radiation."
                )
                impurity_species = st.selectbox("Impurity species (for line radiation)", options=["C","Ne","Ar","W"], index=0)
                impurity_frac = _num("Impurity fraction (rough)", 0.0, 0.001, min_value=0.0, help="Rough number fraction for line radiation placeholder model.")
                include_synchrotron = st.checkbox("Include synchrotron radiation (rough)", value=True)

                zeff_mode = st.selectbox(
                    "Z_eff handling",
                    options=["fixed", "from_impurity", "from_mix"],
                    index=0,
                    help="fixed: use Z_eff input directly. from_impurity: estimate Z_eff from (species, frac). from_mix: estimate Z_eff from impurity_mix dict.",
                )
                impurity_mix = st.text_input(
                    "Impurity mix (optional JSON dict)",
                    value="",
                    help="Optional multi-impurity number fractions, e.g. {\"C\":0.01, \"Ne\":0.002}. Used by the physics radiation model and (if selected) to estimate Z_eff.",
                )

                st.markdown("**Power-channel bookkeeping (transparent; totals unchanged)**")
                f_alpha_to_ion = st.slider("Alpha deposition to ions f_α→i", min_value=0.0, max_value=1.0, value=0.85, step=0.01)
                f_aux_to_ion = st.slider("Aux deposition to ions f_aux→i", min_value=0.0, max_value=1.0, value=0.50, step=0.01)
                include_P_ie = st.checkbox("Include ion↔electron equilibration P_ie (diagnostic)", value=True)

                st.markdown("**Particle sustainability (optional diagnostic closure)**")
                include_particle_balance = st.checkbox("Enable particle balance closure (diagnostic)", value=False)
                tau_p_over_tauE = _num("τ_p / τ_E,eff (–)", 3.0, 0.2, min_value=0.0, help="Proxy: particle confinement time τ_p = (τ_p/τ_E,eff)·τ_E,eff.")
                S_fuel_max_1e22_per_s = _num("Max fueling source S_max (1e22/s) (optional)", float('nan'), 0.1, min_value=0.0, help="If set, SHAMS enforces S_required ≤ S_max as a feasibility constraint (only when particle closure enabled).")

            else:
                Zeff = 1.0
                dilution_fuel = 1.0
                f_rad_core = 0.0
                zeff_mode = "fixed"
                impurity_species = "C"
                impurity_frac = 0.0
                impurity_mix = ""
                include_synchrotron = False
                f_alpha_to_ion = 0.85
                f_aux_to_ion = 0.50
                include_P_ie = True
                include_particle_balance = False
                tau_p_over_tauE = 3.0
                S_fuel_max_1e22_per_s = float("nan")

            if include_alpha_loss:
                alpha_loss_frac = _num("Alpha heating loss fraction (–)", 0.05, 0.01, min_value=0.0, max_value=1.0, help="If enabled, fraction of alpha heating assumed lost (not deposited in core).")
            else:
                alpha_loss_frac = 0.0

            # Optional fast-particle / ash closures (PROCESS-inspired; defaults preserve legacy behavior)
            with st.expander("Advanced fast-particle / ash closures (optional)", expanded=False):
                st.caption("All options here are **opt-in**; defaults preserve current SHAMS behavior.")
                alpha_loss_model = st.selectbox(
                    "Alpha prompt-loss model",
                    options=["fixed", "rho_star"],
                    index=0,
                    help="fixed: use alpha_loss_frac directly. rho_star: alpha_loss_frac_eff = alpha_loss_frac + k·rho* (transparent proxy).",
                )
                alpha_prompt_loss_k = _num(
                    "Prompt-loss slope k (–)",
                    0.0,
                    0.01,
                    min_value=0.0,
                    max_value=1.0,
                    help="Used only if alpha_loss_model='rho_star'. Effective alpha loss is clipped to [0,0.9].",
                )
                alpha_partition_model = st.selectbox(
                    "Alpha ion/electron partition proxy",
                    options=["fixed", "Te_ratio"],
                    index=0,
                    help="Bookkeeping only: affects Palpha_i/Palpha_e reporting (Pin unchanged).",
                )
                alpha_partition_k = _num(
                    "Partition slope k (–)",
                    0.0,
                    0.01,
                    min_value=0.0,
                    max_value=2.0,
                    help="Used only if alpha_partition_model='Te_ratio'.",
                )

                ash_dilution_mode = st.selectbox(
                    "Helium-ash dilution penalty",
                    options=["off", "fixed_fraction"],
                    index=0,
                    help="off: no additional penalty. fixed_fraction: Pfus_for_Q *= (1-f_He_ash)^2 (transparent proxy).",
                )
                f_He_ash = _num(
                    "Helium-ash fraction f_He_ash (–)",
                    0.0,
                    0.01,
                    min_value=0.0,
                    max_value=0.9,
                    help="Used only if ash_dilution_mode='fixed_fraction'.",
                )
            if include_hmode_physics:
                require_Hmode = st.checkbox("Require H-mode access (enforce P_aux ≥ (1+margin)·P_LH)", value=False)
                PLH_margin = _num("P_LH margin (–)", 0.0, 0.05, min_value=0.0, max_value=5.0, help="If Require H-mode is enabled: require P_aux ≥ (1+margin)·P_LH.")
            else:
                require_Hmode = False
                PLH_margin = 0.0
        with st.expander("Operating targets (solver)", expanded=False):
            fuel_mode_label = st.radio(
                "Fuel / design mode",
                ["DT performance (targets Q & net electric)", "DD feasibility (includes secondary DT from DD-produced T)"],
                index=0,
            )
            fuel_mode = "DT" if fuel_mode_label.startswith("DT") else "DD"
            if fuel_mode == "DD":
                include_secondary_DT = st.checkbox("Include secondary DT from DD-produced tritium", value=True)
                if include_secondary_DT:
                    tritium_retention = _num("Tritium retention fraction f_ret (–)", 0.5, 0.05, min_value=0.0, max_value=1.0,
                                             help="Fraction of DD-produced tritium retained/available to burn in secondary DT.")
                    tau_T_loss_s = _num("Effective tritium loss time τ_T (s)", 5.0, 0.5, min_value=0.1,
                                        help="Effective confinement/retention time for produced tritium before loss/removal.")
                else:
                    tritium_retention = 0.0
                    tau_T_loss_s = 1.0
            else:
                include_secondary_DT = False
                tritium_retention = 0.0
                tau_T_loss_s = 1.0

            # Mode-specific safe defaults (DD mode prioritizes feasibility screens over performance)
            default_Q = 2.0 if fuel_mode == "DT" else 0.05
            default_H98 = 1.15 if fuel_mode == "DT" else 1.0
            Q_target = _num("Target Q (fusion gain proxy) [-]", default_Q, 0.05, min_value=0.0)
            H98_target = _num("Target H98 [-]", default_H98, 0.05, min_value=0.1, help="Required confinement factor H98. Solver adjusts Ip and f_G to meet this target.")
            use_envelope = st.checkbox("Design envelope solve (SPARC-like)", value=False, help="Use PROCESS-like bounded vector solve to hit targets by varying Ip, fG, and optionally Paux.")
            Pfus_target = None
            Pnet_target = None
            if use_envelope:
                Pfus_target = _num("Target fusion power P_fus (MW)", 140.0, 10.0, min_value=0.0)
                Pnet_target = _num("Target net electric power P_net (MW) (optional)", -1.0, 10.0, help="Set to <0 to ignore. If >0, solver will try to meet it by varying Paux as needed.", min_value=-1e6)

            # -------------------------------------------------------------
            # Optimization (PROCESS-like): search within bounds for a better design
            # -------------------------------------------------------------
            st.markdown("**Optimization (experimental)**")
            do_opt = st.checkbox("Run constrained optimization (random search)", value=False,
                                 help="Searches over (Ip, fG, Paux) within bounds to improve an objective while satisfying constraints.")
            opt_objective = st.selectbox("Objective", ["min_R0", "min_Bpeak", "max_Pnet", "min_recirc"], index=1)
            opt_iters = int(_num("Optimization iterations", 200, 10, min_value=20.0))
            opt_seed = int(_num("Optimization seed", 1, 1, min_value=0.0))
            st.divider()

            # Defaults track the currently loaded base point so preset loads immediately feel consistent.
            _ip0 = float(getattr(_base_pd, "Ip_MA", 8.0) or 8.0)
            _fg0 = float(getattr(_base_pd, "fG", 0.8) or 0.8)
            Ip_min = _num("Plasma current lower bound I_p,min (MA)", max(0.1, 0.80 * _ip0), 1.0, min_value=0.1, key=PD_KEYS["Ip_lo"])
            Ip_max = _num("Plasma current upper bound I_p,max (MA)", max(0.2, 1.20 * _ip0), 0.5, min_value=0.1, key=PD_KEYS["Ip_hi"])
            fG_min = _num("Greenwald fraction lower bound f_G,min (–)", max(0.0, _fg0 - 0.20), 0.01, min_value=0.0, max_value=2.0, key=PD_KEYS["fG_lo"])
            fG_max = _num("Greenwald fraction upper bound f_G,max (–)", min(2.0, _fg0 + 0.20), 0.01, min_value=0.0, max_value=2.0, key=PD_KEYS["fG_hi"])
            tol = _num("solver tol [-]", 1e-3, 1e-4, min_value=1e-6, fmt="%.1e")
            show_solver_live = st.checkbox(
                "Show solver physics live (step-by-step)",
                value=True,
                help=(
                    "Visualize how the nested solver converges: outer bisection on Ip to hit the target H98, "
                    "with an inner solve on fG to match the target Q at each Ip evaluation."
                ),
            )
        with st.expander("Engineering & plant feasibility (optional)", expanded=False):
            # These names are passed through via PointInputs **kwargs, so they must exist in your src version.
            # We keep them optional. If missing, they are simply ignored by PointInputs.
            tshield = _num("Neutron shield thickness (m)", 0.8, 0.01, min_value=0.0, help="Effective neutron shield thickness used for neutronics/HTS lifetime proxies.")
            # A small representative set; add more once you confirm exact fields in src/phase1_systems.py
            # We still allow user to run without them.

            # --- Engineering & plant feasibility (optional): per-subsystem toggles + confidence presets ---
            st.markdown("#### Engineering & plant feasibility (optional)")
            confidence = st.radio(
                "Confidence level",
                ["Conservative", "Nominal", "Aggressive"],
                index=1,
                horizontal=True,
                help="Controls default assumptions and warning bands (WARN vs FAIL). Conservative is stricter; aggressive is more permissive."
            )
            warn_fracs = {
                "Conservative": {"max": 0.85, "min": 1.20},
                "Nominal":      {"max": 0.90, "min": 1.10},
                "Aggressive":   {"max": 0.95, "min": 1.05},
            }[confidence]

            c1, c2 = st.columns(2)
            with c1:
                include_build = st.checkbox("Build & radial build", value=True)
                include_magnets = st.checkbox("Magnets & HTS", value=True)
                include_divertor = st.checkbox("Divertor / SOL", value=True)
            with c2:
                include_neutronics = st.checkbox("Neutronics (TBR, lifetime)", value=True)
                include_net_power = st.checkbox("Net power / electrical balance", value=True)

            preset = {
                "Conservative": {
                    "tblanket_m": 0.60, "t_vv_m": 0.08, "t_gap_m": 0.03, "t_tf_struct_m": 0.18, "t_tf_wind_m": 0.12,
                    "Bpeak_factor": 1.30, "sigma_allow_MPa": 800.0, "Tcoil_K": 20.0, "hts_margin_min": 0.20, "Vmax_kV": 18.0,
                    "q_div_max_MW_m2": 7.0, "TBR_min": 1.10, "hts_lifetime_min_yr": 5.0, "P_net_min_MW": 0.0,
                },
                "Nominal": {
                    "tblanket_m": 0.50, "t_vv_m": 0.06, "t_gap_m": 0.02, "t_tf_struct_m": 0.15, "t_tf_wind_m": 0.10,
                    "Bpeak_factor": 1.25, "sigma_allow_MPa": 850.0, "Tcoil_K": 20.0, "hts_margin_min": 0.15, "Vmax_kV": 20.0,
                    "q_div_max_MW_m2": 10.0, "TBR_min": 1.05, "hts_lifetime_min_yr": 3.0, "P_net_min_MW": 0.0,
                },
                "Aggressive": {
                    "tblanket_m": 0.40, "t_vv_m": 0.05, "t_gap_m": 0.015, "t_tf_struct_m": 0.12, "t_tf_wind_m": 0.08,
                    "Bpeak_factor": 1.20, "sigma_allow_MPa": 900.0, "Tcoil_K": 20.0, "hts_margin_min": 0.10, "Vmax_kV": 25.0,
                    "q_div_max_MW_m2": 15.0, "TBR_min": 1.00, "hts_lifetime_min_yr": 1.0, "P_net_min_MW": 0.0,
                },
            }[confidence]

            def _maybe(x: float, enabled: bool) -> float:
                return float(x) if enabled else float("nan")

            clean_knobs = {
                # Build & radial build
                "tblanket_m": _maybe(float(_num("Blanket thickness (inboard) (m)", preset["tblanket_m"], 0.01, min_value=0.0)), include_build),
                "t_vv_m": _maybe(float(_num("Vacuum vessel thickness (inboard) (m)", preset["t_vv_m"], 0.005, min_value=0.0)), include_build),
                "t_gap_m": _maybe(float(_num("Inboard gap / clearance (m)", preset["t_gap_m"], 0.005, min_value=0.0)), include_build),
                "t_tf_struct_m": _maybe(float(_num("TF structure thickness (inboard) (m)", preset["t_tf_struct_m"], 0.01, min_value=0.0)), include_build),
                "t_tf_wind_m": _maybe(float(_num("TF winding pack thickness (inboard) (m)", preset["t_tf_wind_m"], 0.01, min_value=0.0)), include_build),

                # Magnets & HTS
                "Bpeak_factor": _maybe(float(_num("Peak-field mapping factor B_peak/B₀ (–)", preset["Bpeak_factor"], 0.01, min_value=1.0)), include_magnets),
                "sigma_allow_MPa": _maybe(float(_num("Allowable coil hoop stress (MPa)", preset["sigma_allow_MPa"], 10.0, min_value=10.0)), include_magnets),
                "Tcoil_K": _maybe(float(_num("HTS operating temperature (K)", preset["Tcoil_K"], 1.0, min_value=4.0)), include_magnets),
                "hts_margin_min": _maybe(float(_num("Minimum HTS critical-current margin (–)", preset["hts_margin_min"], 0.01, min_value=0.0)), include_magnets),
                "include_hts_critical_surface": bool(st.checkbox("Use HTS critical-surface model (Jc(B,T,ε))", value=False, disabled=not include_magnets, help="Off by default (legacy behavior). When enabled, computes hts_margin_cs using Jc(B,T,ε_tf)/Jop and applies the same hts_margin_min threshold.")),
                "Vmax_kV": _maybe(float(_num("Max dump voltage limit (kV)", preset["Vmax_kV"], 1.0, min_value=1.0)), include_magnets),

                # Divertor / SOL
                "q_div_max_MW_m2": _maybe(float(_num("Max divertor heat flux limit (MW/m²)", preset["q_div_max_MW_m2"], 0.5, min_value=0.1)), include_divertor),

                # Neutronics
                "TBR_min": _maybe(float(_num("Minimum tritium breeding ratio (TBR)", preset["TBR_min"], 0.01, min_value=0.0)), include_neutronics),
                "hts_lifetime_min_yr": _maybe(float(_num("Minimum HTS lifetime (years)", preset["hts_lifetime_min_yr"], 0.5, min_value=0.0)), include_neutronics),

                # Net power / electrical balance
                "P_net_min_MW": _maybe(float(_num("Minimum net electric power (MW)", preset["P_net_min_MW"], 10.0, min_value=-1e6)), include_net_power),

                # propagate UI choices to output for check logic
                "_warn_frac_max": float(warn_fracs["max"]),
                "_warn_frac_min": float(warn_fracs["min"]),
                "_subsystem_enabled": {
                    "build": bool(include_build),
                    "magnets": bool(include_magnets),
                    "divertor": bool(include_divertor),
                    "neutronics": bool(include_neutronics),
                    "net_power": bool(include_net_power),
                },
            }

            # (Button moved outside this expander.)

        # Evaluate button is intentionally *outside* the optional engineering section so
        # users don't have to expand engineering knobs just to run Point Designer.
        run_btn = st.button("Evaluate Point", type="primary", use_container_width=True)


    with mid:
        st.subheader("Results")
        if run_btn:
            # Solve Ip + fG to match (H98_target, Q_target)
            log_lines: List[str] = []

            def _log(line: str) -> None:
                """Append a single line to the expandable solver log."""
                try:
                    log_lines.append(str(line))
                except Exception:
                    pass

            _log("Point Designer solver log")
            _log(
                f"Targets: H98={H98_target:.6g}, Q={Q_target:.6g}; bounds: Ip=[{Ip_min:.6g},{Ip_max:.6g}] MA, fG=[{fG_min:.6g},{fG_max:.6g}]; tol={tol:.3e}"
            )
            _log(
                f"Machine: R0={R0:.6g} m, a={a:.6g} m, kappa={kappa:.6g}, B0={B0:.6g} T; Paux={Paux:.6g} MW (Q denom={Paux_for_Q:.6g} MW); Ti={Ti:.6g} keV; Ti/Te={Ti_over_Te:.6g}; fuel_mode={fuel_mode}"
            )
            base = make_point_inputs(
                R0_m=R0, a_m=a, kappa=kappa, delta=delta, Bt_T=B0,
                Ip_MA=0.5*(Ip_min+Ip_max),
                Ti_keV=Ti, fG=0.8,
                t_shield_m=tshield,
                Paux_MW=Paux,
                Ti_over_Te=Ti_over_Te,
                zeff=Zeff,
                dilution_fuel=dilution_fuel,
                f_rad_core=f_rad_core,
                include_radiation=include_radiation,
                radiation_model=radiation_model,
                impurity_species=impurity_species,
                impurity_frac=impurity_frac,
                include_synchrotron=include_synchrotron,
                confinement_model=confinement_model,
                confinement_scaling=confinement_scaling,
                profile_model=profile_model,
            profile_peaking_ne=profile_peaking_ne,
            profile_peaking_T=profile_peaking_T,
                bootstrap_model=bootstrap_model,
                fuel_mode=fuel_mode,
                include_secondary_DT=include_secondary_DT,
                tritium_retention=tritium_retention,
                tau_T_loss_s=tau_T_loss_s,
                alpha_loss_frac=alpha_loss_frac,
                alpha_loss_model=alpha_loss_model,
                alpha_prompt_loss_k=alpha_prompt_loss_k,
                alpha_partition_model=alpha_partition_model,
                alpha_partition_k=alpha_partition_k,
                ash_dilution_mode=ash_dilution_mode,
                f_He_ash=f_He_ash,
                include_alpha_loss=include_alpha_loss,
                include_hmode_physics=include_hmode_physics,
                require_Hmode=require_Hmode,
                PLH_margin=PLH_margin,
                use_lambda_q=use_lambda_q,
                **clean_knobs,
            )

            # UI-only guardrails: warn on obviously unrealistic knobs (does not block).
            _warn_unrealistic_point_inputs(base, context="Point Designer")
            if do_opt:
                _log(f"Optimization enabled: objective={opt_objective}, iters={opt_iters}, seed={opt_seed}")
                var_bounds = {"Ip_MA": (Ip_min, Ip_max), "fG": (fG_min, fG_max), "Paux_MW": (0.0, max(Paux, 1e-6)*2.0)}
                best_inp, best_out = optimize_design(
                    base,
                    objective=opt_objective,
                    variables=var_bounds,
                    n_iter=opt_iters,
                    seed=opt_seed,
                )
                base = best_inp
                _log(f"Optimization chose: Ip={best_inp.Ip_MA:.4g} MA, fG={best_inp.fG:.4g}, Paux={best_inp.Paux_MW:.4g} MW")
                _log(f"Optimized outputs: Bpeak={best_out.get('B_peak_T', float('nan')):.4g} T, Pnet={best_out.get('P_e_net_MW', float('nan')):.4g} MW")
            # Optional: show solver progress so the user can "see physics happening".
            if show_solver_live:
                with st.expander("Solver convergence (live)", expanded=False):
                    status = st.empty()
                    prog = st.progress(0)
                    # Live convergence diagnostics
                    chart = st.empty()
                    table = st.empty()
                    latest = st.empty()

                trace_rows = []
                sol_inp, out, ok = None, {}, False

                # Select solver iterator (legacy stream solver vs envelope solve)
                if use_envelope:
                    tgt = {'Q_DT_eqv': Q_target, 'H98': H98_target}
                    sol_inp_env, out_env, ok_env, msg_env = solve_sparc_envelope(
                        base, tgt, vary=['Ip_MA','fG'],
                        bounds={'Ip_MA': (Ip_min, Ip_max), 'fG': (fG_min, fG_max)},
                        tol=tol, max_iter=40,
                    )
                    def _env_events():
                        yield {'event':'iter','it':0,'Ip_MA': sol_inp_env.Ip_MA, 'fG': sol_inp_env.fG, 'H98': out_env.get('H98', float('nan')), 'Q_DT_eqv': out_env.get('Q_DT_eqv', float('nan'))}
                        yield {'event':'done','sol': sol_inp_env, 'out': out_env, 'ok': ok_env, 'message': msg_env}
                    event_iter = _env_events()
                else:
                    event_iter = solve_Ip_for_H98_with_Q_match_stream(
                    base=base,
                    target_H98=H98_target,
                    target_Q=Q_target,
                    Ip_min=Ip_min, Ip_max=Ip_max,
                    fG_min=fG_min, fG_max=fG_max,
                    tol=tol,
                    Paux_for_Q_MW=Paux_for_Q,
                )
                for ev in event_iter:
                    if ev.get("event") == "bracket":
                        okb = bool(ev.get("ok"))
                        try:
                            _log(
                                f"BRACKET: H98(Ip_lo={ev.get('Ip_lo'):.6g})={ev.get('H98_lo'):.6g}, H98(Ip_hi={ev.get('Ip_hi'):.6g})={ev.get('H98_hi'):.6g} -> {'OK' if okb else 'NO_BRACKET'}"
                            )
                        except Exception:
                            _log(f"BRACKET: ok={okb}")
                        status.info(
                            f"Bracketing H98 target: H98(Ip_min={ev.get('Ip_lo'):.3g})={ev.get('H98_lo'):.3g}, "
                            f"H98(Ip_max={ev.get('Ip_hi'):.3g})={ev.get('H98_hi'):.3g}  →  "
                            f"{'OK' if okb else 'NO BRACKET'}"
                        )
                    elif ev.get("event") == "iter":
                        try:
                            _log(
                                f"ITER {int(ev.get('iter', 0)):>3d}: Ip={ev.get('Ip_MA'):.8g} MA, fG={ev.get('fG'):.8g}, H98={ev.get('H98'):.8g}, Q={ev.get('Q'):.8g}, residual={ev.get('residual'):.8g}"
                            )
                        except Exception:
                            _log(f"ITER {ev.get('iter')}: {ev}")
                        trace_rows.append({
                            "iter": ev.get("iter"),
                            "Ip_MA": ev.get("Ip_MA"),
                            "fG": ev.get("fG"),
                            "H98": ev.get("H98"),
                            "Q": ev.get("Q"),
                            "residual": ev.get("residual"),
                        })
                        it = int(ev.get("iter", 0))
                        prog.progress(min(1.0, (it + 1) / 80.0))
                        latest.metric("Current guess Ip (MA)", f"{ev.get('Ip_MA', float('nan')):.4g}")
                        if trace_rows:
                            df = pd.DataFrame(trace_rows)
                            # Two quick plots: residual and key state variables
                            chart.line_chart(df.set_index("iter")[["residual"]])
                            table.dataframe(df.tail(10), use_container_width=True)
                    elif ev.get("event") == "done":
                        sol_inp = ev.get("sol")
                        out = ev.get("out", {})
                        ok = True
                        try:
                            _log(
                                f"DONE: Ip={out.get('Ip_MA', float('nan')):.8g} MA, fG={out.get('fG', float('nan')):.8g}, H98={out.get('H98', float('nan')):.8g}, Q_DT_eqv={out.get('Q_DT_eqv', float('nan')):.8g}"
                            )
                        except Exception:
                            _log("DONE")
                        if bool(out.get("_solver_clamped")) or bool(out.get("_solver_clamped_Q")):
                            status.warning("Solver returned a point by clamping to the nearest bound (target not achievable within bounds). See log/details below.")
                        else:
                            status.success("Solver converged.")
                        prog.progress(1.0)
                        break
                    elif ev.get("event") == "fail":
                        reason = ev.get("reason", "solver_failed")
                        _log("FAIL EVENT: " + json.dumps(ev, sort_keys=True))
                        it_fail = ev.get("it", None)
                        mi_fail = ev.get("max_iter", None)
                        extra = ""
                        if it_fail is not None and mi_fail is not None:
                            extra = f" (it={it_fail}/{mi_fail})"
                        status.error(f"Solver failed ({reason}){extra}. Try widening Ip/fG bounds or relaxing targets.")
                        ok = False
                        break
            else:
                sol_inp, out, ok = solve_Ip_for_H98_with_Q_match(
                    base=base,
                    target_H98=H98_target,
                    target_Q=Q_target,
                    Ip_min=Ip_min, Ip_max=Ip_max,
                    fG_min=fG_min, fG_max=fG_max,
                    tol=tol,
                    Paux_for_Q_MW=Paux_for_Q,
                )
                # Minimal log summary when running in non-stream mode.
                if ok:
                    try:
                        _log(
                            f"DONE: Ip={out.get('Ip_MA', float('nan')):.8g} MA, fG={out.get('fG', float('nan')):.8g}, H98={out.get('H98', float('nan')):.8g}, Q_DT_eqv={out.get('Q_DT_eqv', float('nan')):.8g}"
                        )
                    except Exception:
                        _log("DONE")
                else:
                    _log("FAIL: solver_failed")

            # Always show expandable log for this run.
            solver_log_text = "\n".join(log_lines).strip() + "\n"
            st.session_state.last_solver_log = solver_log_text
            with st.expander("Solver log (expand to view)", expanded=False):
                st.download_button(
                    "Download log",
                    data=solver_log_text,
                    file_name="point_designer_solver.log",
                    mime="text/plain",
                    use_container_width=True,
                )
                st.code(solver_log_text)
            if not ok:
                # Provide best-effort diagnostics if available (e.g., H98 at bounds)
                msg = "Solver failed to converge for (Ip, fG) at the requested (H98, Q) targets."
                try:
                    if isinstance(out, dict) and ("H98_at_Ip_min" in out or "H98_at_Ip_max" in out):
                        msg += f"  H98(Ip_min)={out.get('H98_at_Ip_min')}, H98(Ip_max)={out.get('H98_at_Ip_max')}"
                except Exception:
                    pass
                st.error(msg)

                # -----------------------------------------------------------------
                # PROCESS-like feasibility frontier suggestion
                # -----------------------------------------------------------------
                with st.expander("Try to find nearest feasible point (frontier)", expanded=False):
                    st.markdown(
                        "If the solver cannot hit the requested (H98, Q) targets inside the bounds, "
                        "SHAMS can still search for the *nearest feasible* design within your (Ip,fG) bounds. "
                        "This does **not** change your inputs automatically; it only proposes a candidate."
                    )
                    if st.button("Search nearest feasible within bounds", key="pd_frontier_btn", use_container_width=True):
                        try:
                            fr = find_nearest_feasible(
                                base,
                                levers={"Ip_MA": (Ip_min, Ip_max), "fG": (fG_min, fG_max)},
                                targets={"H98": float(H98_target), "Q_DT_eqv": float(Q_target)},
                                n_random=80,
                                seed=0,
                            )
                            st.session_state["pd_frontier_last"] = fr.report
                        except Exception as e:
                            st.session_state["pd_frontier_last"] = {"status": "error", "message": str(e)}

                    rep = st.session_state.get("pd_frontier_last")
                    if isinstance(rep, dict) and rep:
                        if rep.get("status") == "error":
                            st.error(rep.get("message", "frontier error"))
                        else:
                            cols = st.columns(3)
                            cols[0].metric("Best Ip (MA)", f"{rep.get('best_levers', {}).get('Ip_MA', float('nan')):.4g}")
                            cols[1].metric("Best fG", f"{rep.get('best_levers', {}).get('fG', float('nan')):.4g}")
                            cols[2].metric("Feasible?", "YES" if rep.get("best_ok") else "NO")
                            ach = rep.get("best_achieved", {}) or {}
                            st.write("Best achieved targets at proposed point:")
                            st.json(ach)
                            st.caption("Tip: widen bounds or relax targets if the frontier is still infeasible.")
            else:
                # Attach UI-only meta for checks (not used by physics core)
                try:
                    out['_warn_frac_max'] = float(clean_knobs.get('_warn_frac_max', 0.90))
                    out['_warn_frac_min'] = float(clean_knobs.get('_warn_frac_min', 1.10))
                    out['_subsystem_enabled'] = clean_knobs.get('_subsystem_enabled', {})
                except Exception:
                    pass
                st.session_state.last_point_out = out

                # -----------------------------------------------------------------
                # PROCESS-inspired canonical output artifact (SHAMS-native JSON)
                # -----------------------------------------------------------------
                try:
                    inputs_dict = dict(base.__dict__)
                except Exception:
                    inputs_dict = {}
                try:
                    constraints_list = evaluate_constraints(out)
                except Exception:
                    constraints_list = []
                try:
                    solver_meta = None
                    try:
                        solver_meta = dict(out.get("_solver")) if isinstance(out.get("_solver"), dict) else None
                    except Exception:
                        solver_meta = None
                    if solver_meta is not None:
                        # Attach UI log if available
                        try:
                            solver_meta.setdefault("ui_log", st.session_state.get("last_solver_log", ""))
                        except Exception:
                            pass
                    artifact = build_run_artifact(
                        inputs=inputs_dict,
                        outputs=dict(out),
                        constraints=constraints_list,
                        meta=None,
                        baseline_inputs=inputs_dict,
                        subsystems={"fidelity": st.session_state.get("fidelity_config", {}), "calibration": {"confinement": float(st.session_state.get("calib_confinement",1.0)), "divertor": float(st.session_state.get("calib_divertor",1.0)), "bootstrap": float(st.session_state.get("calib_bootstrap",1.0))}},
                        solver=solver_meta,
                    )
                except Exception:
                    artifact = {"inputs": inputs_dict, "outputs": dict(out), "constraints": []}

                # Provide downloadable artifacts and reports (no side effects unless user clicks)
                with st.expander("Exports (PROCESS-style artifacts)", expanded=False):
                    st.download_button(
                        "Download run artifact JSON",
                        data=json.dumps(artifact, indent=2, sort_keys=True),
                        file_name="shams_run_artifact.json",
                        mime="application/json",
                        use_container_width=True,
                    )

                    # Radial build PNG
                    try:
                        import tempfile
                        tmpdir = tempfile.mkdtemp(prefix="shams_export_")
                        radial_path = os.path.join(tmpdir, "radial_build.png")
                        plot_radial_build_from_artifact(artifact, radial_path)
                        with open(radial_path, "rb") as f:
                            st.download_button(
                                "Download radial build PNG",
                                data=f,
                                file_name="shams_radial_build.png",
                                mime="image/png",
                                use_container_width=True,
                            )
                    except Exception as _e:
                        st.caption("Radial-build export unavailable for this point.")

                    # Summary PDF
                    try:
                        import tempfile
                        tmpdir2 = tempfile.mkdtemp(prefix="shams_export_")
                        pdf_path = os.path.join(tmpdir2, "summary.pdf")
                        plot_summary_pdf(artifact, pdf_path)
                        with open(pdf_path, "rb") as f:
                            st.download_button(
                                "Download summary PDF",
                                data=f,
                                file_name="shams_summary.pdf",
                                mime="application/pdf",
                                use_container_width=True,
                            )
                    except Exception:
                        st.caption("PDF summary export unavailable for this point.")

                with st.expander("Point summary", expanded=True):
                    # Standardized KPI set shared with PDF (see decision.kpis.KPI_SET)
                    kpis = headline_kpis(out)
                    for i in range(0, len(kpis), 4):
                        kpi_row(kpis[i:i+4])
                        if i + 4 < len(kpis):
                            st.divider()
                    st.divider()
                    kpi_row([
                        ("HTS margin", f"{out.get('hts_margin', float('nan')):.3f}"),
                        ("Lifetime [yr]", f"{out.get('hts_lifetime_yr', float('nan')):.2f}"),
                        ("Vdump [kV]", f"{out.get('V_dump_kV', float('nan')):.1f}"),
                        ("P_net_e [MW]", f"{out.get('P_net_e_MW', float('nan')):.1f}"),
                    ])


                # PROCESS-style local sensitivities (finite differences)

                # -----------------------------------------------------------------
                # Constraint dashboard (PROCESS-like) with margins + suggestions
                # -----------------------------------------------------------------
                with st.expander("Constraints & margins (why pass/fail)", expanded=True):
                    if not constraints_list:
                        st.info("No constraints evaluated (missing keys).")
                    else:
                        rows_c = []
                        for c in constraints_list:
                            try:
                                margin = float(getattr(c, "margin"))
                            except Exception:
                                margin = float("nan")
                            rows_c.append({
                                "constraint": c.name,
                                "sense": c.sense,
                                "value": c.value,
                                "limit": c.limit,
                                "units": c.units,
                                "passed": bool(c.passed),
                                "margin_frac": margin,
                                "severity": getattr(c, "severity", "hard"),
                                "note": c.note,
                            })
                        dfc = pd.DataFrame(rows_c)
                        # sort: hard fails first, then smallest margin
                        try:
                            dfc = dfc.sort_values(by=["passed", "severity", "margin_frac"], ascending=[True, True, True])
                        except Exception:
                            pass
                        st.dataframe(dfc, use_container_width=True)

                        failed = [r for r in rows_c if not r["passed"] and r.get("severity","hard") == "hard"]
                        if failed:
                            st.error(f"{len(failed)} hard constraint(s) failed. See suggestions below.")
                        soft_failed = [r for r in rows_c if not r["passed"] and r.get("severity") == "soft"]
                        if soft_failed:
                            st.warning(f"{len(soft_failed)} soft constraint(s) failed (screening only).")

                        def _suggest(name: str) -> str:
                            n = name.lower()
                            if "q_div" in n or "p_sol" in n:
                                return "Reduce P_SOL (increase radiation, reduce aux), increase R0, or increase lambda_q (design/multiplier)."
                            if "hts" in n or "b_peak" in n or "sigma" in n:
                                return "Reduce B_peak (increase coil build/R0, reduce Bt), reduce stress (increase thickness, reduce B_peak), or raise HTS margin (lower Top or improve conductor)."
                            if "tbr" in n:
                                return "Increase blanket/shield thickness or improve breeding/coverage assumptions."
                            if "nwl" in n:
                                return "Reduce fusion power density (increase size R0 or reduce performance targets) or improve shielding."
                            if "beta" in n:
                                return "Increase size R0 or reduce Ip/pressure (lower Ti or fG) to bring beta below limit."
                            if "q95" in n:
                                return "Increase q95 (reduce Ip or increase Bt/R0) for stability margin."
                            if "fg" in n:
                                return "Reduce density target (lower fG) or increase Ip to raise Greenwald limit."
                            if "p_net" in n:
                                return "Increase Pfus (within constraints), increase thermal efficiency, or reduce recirculating loads."
                            if "t_flat" in n:
                                return "Increase available flux swing (CS design), reduce loop voltage (improve resistivity/current profile), or allow lower Ip."
                            return "Adjust major radius / field / current / aux power to recover feasibility."

                        if failed or soft_failed:
                            st.markdown("**Actionable suggestions (rule-of-thumb):**")
                            for r in failed + soft_failed:
                                st.write("- **{}**: {}".format(r["constraint"], _suggest(r["constraint"])))
                with st.expander("Local sensitivities (finite difference)", expanded=False):
                    st.caption("Local derivatives around the current point. Useful for design intuition; not a global optimization result.")
                    try:
                        params = ["R0_m","a_m","kappa","B0_T","Ip_MA","fG","H98","eta_CD","n_neu_frac","Zeff"]
                        outs = ["Q_DT_eqv","P_net_e_MW","betaN","q_div_MW_m2","B_peak_T"]
                        def _eval(pi):
                            return hot_ion_point(pi)
                        sens = finite_difference_sensitivities(base, _eval, params=params, outputs=outs, rel_step=1e-3)
                        # Show a compact table: normalized sensitivities (per 1% change), where possible
                        rows = []
                        for o in outs:
                            base_y = float(sens.get("_base", {}).get(o, float("nan")))
                            for p in params:
                                if p not in sens.get(o, {}):
                                    continue
                                dydx = float(sens[o][p])
                                x0 = float(getattr(base, p)) if hasattr(base, p) and getattr(base, p) is not None else float("nan")
                                # normalized: (dY/Y) / (dX/X)  = (dY/dX) * (X/Y)
                                norm = float("nan")
                                if x0 == x0 and base_y == base_y and x0 != 0.0 and base_y != 0.0:
                                    norm = dydx * (x0 / base_y)
                                rows.append({"output": o, "param": p, "dY/dX": dydx, "elasticity": norm})
                        if rows:
                            df_s = pd.DataFrame(rows)
                            st.dataframe(df_s.sort_values(["output","param"]), use_container_width=True)
                        else:
                            st.info("Sensitivities unavailable for this point (missing keys or non-finite outputs).")
                    except Exception as e:
                        st.warning(f"Sensitivity calculation failed: {e}")

                with st.expander("Plots", expanded=True):
                    st.markdown("### Plot dashboard")
                    ptab1, ptab2, ptab3 = st.tabs(["Power balance", "Stability & limits", "Geometry / build"])

                    with ptab1:
                        st.caption(
                            "Quick visual breakdown of where power is going in this 0‑D point (all Phase‑1 proxies)."
                        )
                        power_vals = {
                            "Paux [MW]": out.get("Paux_MW"),
                            "Pfus (DT-eqv) [MW]": out.get("Pfus_DT_adj_MW"),
                            "Pα dep [MW]": out.get("Palpha_dep_MW"),
                            "Prad_core [MW]": out.get("Prad_core_MW"),
                            "P_SOL [MW]": out.get("P_SOL_MW"),
                            "P_net_e [MW]": out.get("P_net_e_MW"),
                        }
                        plot_bars(power_vals, "Power balance (MW)")
                        with st.expander("Physical meaning (with literature)", expanded=False):
                            st.markdown(
                                """
    **Q (fusion gain proxy)** is defined as fusion power divided by auxiliary heating power (here the UI uses *Paux_for_Q* as the denominator).  
    **H98** is a confinement multiplier relative to the empirical **IPB98(y,2)** ELMy H‑mode scaling used as an ITER physics-basis reference. citeturn1view0turn0search16

    **P_LH / H‑mode access** comparisons in this app follow the multi‑machine ITPA threshold scaling (often referred to as “Martin‑2008 / PLH‑08”). citeturn3search18

    If you enable SOL-width physics, the app’s λq proxy is motivated by the multi‑machine H‑mode power‑falloff width scaling (Eich‑2013). citeturn2search3
                                """
                            )

                    with ptab2:
                        st.caption("Screening metrics vs common operational ‘guardrails’ (Phase‑1 proxies).")
                        stab_vals = {
                            "q95": out.get("q95_proxy"),
                            "βN": out.get("betaN_proxy"),
                            "f_bs": out.get("f_bs_proxy"),
                        }
                        plot_bars(stab_vals, "Stability / operational metrics")
                        with st.expander("Physical meaning (with literature)", expanded=False):
                            st.markdown(
                                """
    **q95** (safety factor near 95% flux) is a standard operational metric used as a proxy for MHD margin; lower q tends to reduce kink/tearing stability margin.

    **Normalized beta βN** is a widely used performance/stability figure of merit that scales pressure relative to magnetic field and current (often discussed in terms of the “Troyon” aB/I scaling). citeturn0search19

    **Bootstrap fraction f_bs** indicates how much of the plasma current is self‑driven by pressure gradients (important for steady‑state operation). This UI uses a simple proxy coefficient (C_bs) rather than a full neoclassical calculation.
                                """
                            )

                    with ptab3:
                        st.caption("A few geometry/build proxies that drive magnet and shield feasibility checks.")
                        geom_vals = {
                            "R0 [m]": out.get("R0_m"),
                            "a [m]": out.get("a_m"),
                            "B0 [T]": out.get("Bt_T"),
                            "Bpeak [T]": out.get("Bpeak_T"),
                            "σ_hoop [MPa]": out.get("sigma_hoop_MPa"),
                            "t_shield [m]": out.get("t_shield_m"),
                        }
                        plot_bars(geom_vals, "Key geometry/build scalars")
                        with st.expander("Physical meaning (with literature)", expanded=False):
                            st.markdown(
                                """
    **Greenwald fraction fG** (used internally by the solver) expresses density as a fraction of the empirical tokamak density limit scaling with I_p and minor radius (often called the Greenwald limit). citeturn2search12turn2search8

    The *radial build* and **Bpeak/B0** mapping are engineering proxies; they’re not meant to replace detailed coil/stress finite‑element analysis.
                                """
                            )

                    st.markdown("### Full outputs")
                    st.dataframe(pd.DataFrame([out]).T.rename(columns={0: "value"}), use_container_width=True)

    with right:
        st.subheader("Checks & explain")
        with st.expander("Checks & explain", expanded=False):
            out = st.session_state.last_point_out
            if out is None:
                st.info("Run **Evaluate Point** to see constraint checks.")
            else:
                checks = compute_checks(out)
                for c in checks:
                    with st.expander(f"{c.get('name', 'Check')}", expanded=False):
                        st.write(f"**{c['name']}** — {badge(c)}")
                        v = c.get("value")
                        lim = c.get("limit")
                        wl = c.get("warn_limit")
                        if isinstance(v, (int, float)) and isinstance(lim, (int, float)) and math.isfinite(v) and math.isfinite(lim):
                            if isinstance(wl, (int, float)) and math.isfinite(wl):
                                st.caption(f"value={v:.4g}  warn={wl:.4g}  limit={lim:.4g}  ({c.get('sense','')})")
                            else:
                                st.caption(f"value={v:.4g}  limit={lim:.4g}  ({c.get('sense','')})")
                        if c.get("notes"):
                            st.caption(c["notes"])
                        st.divider()

                with st.expander("Check summary", expanded=True):
                    bad = top_violations(checks, 3)
                    if bad:
                        st.markdown("### Top violations")
                        for c in bad:
                            st.write(f"- **{c['name']}**: value={c.get('value')} vs limit={c.get('limit')}")
                    else:
                        st.success("All enabled checks passed for this point (per Phase‑1 proxy models).")


# -----------------------------
# Scan Lab
# -----------------------------
# -----------------------------
# Systems Mode (PROCESS-like coupled targeting)
# -----------------------------
with tab_systems:
    st.subheader("Systems Mode (constraint-driven solve)")
    st.markdown(
        "Solve for a self-consistent operating point by adjusting **iteration variables** "
        "to hit **targets** (e.g., Q, H98, net electric) while reporting engineering and physics margins. "
        "This is inspired by PROCESS's constraint-driven workflow, but remains SHAMS-native and transparent."
    )

    base0 = st.session_state.last_point_inp
    if base0 is None:
        base0 = PointInputs(R0_m=1.85, a_m=0.57, kappa=1.8, Bt_T=12.2, Ip_MA=8.0, Ti_keV=15.0, fG=0.8, Paux_MW=20.0)

    with st.expander("Base design (starting point)", expanded=True):
        colA, colB, colC = st.columns(3)
        with colA:
            R0_m = _num("R0 [m]", float(base0.R0_m), 0.01, help="Major radius.")
            a_m = _num("a [m]", float(base0.a_m), 0.01, help="Minor radius.")
            kappa = _num("κ [-]", float(base0.kappa), 0.01, help="Elongation.")
            delta = _num("δ [-]", float(getattr(base0, "delta", 0.0) or 0.0), 0.02, min_value=0.0, max_value=0.8, help="Triangularity δ used only in the inboard radial-build clearance proxy.")
        with colB:
            Bt_T = _num("Bt [T]", float(base0.Bt_T), 0.1, help="On-axis toroidal field.")
            Ti_keV = _num("Ti [keV]", float(base0.Ti_keV), 0.5, help="Ion temperature (volume-average input in 0-D mode).")
            Ti_over_Te = _num("Ti/Te [-]", float(getattr(base0, "Ti_over_Te", 2.0)), 0.05, help="Temperature ratio; sets Te.")
        with colC:
            t_shield_m = _num("Shield thickness [m]", float(getattr(base0, "t_shield_m", 0.70)), 0.01, help="Inboard shielding thickness proxy (affects neutronics/HTS lifetime).")
            steady_state = st.checkbox("Steady-state (no CS pulse constraint)", value=bool(getattr(base0, "steady_state", True)))
            P_net_min_MW = _num("Minimum net electric [MW(e)]", float(getattr(base0, "P_net_min_MW", 0.0)), 5.0, help="Optional requirement; 0 disables hard requirement.")
        # model options
        st.markdown("**Model options**")
        m1, m2, m3 = st.columns(3)
        with m1:
            confinement_model = st.selectbox("Confinement scaling", ["ipb98y2","iter89p"], index=0)
        with m2:
            profile_model = st.selectbox("Profiles (½-D)", ["none","parabolic","pedestal"], index=0)
        with m3:
            zeff_mode = st.selectbox("Zeff mode", ["fixed","from_impurity"], index=0)
        profile_peaking_ne = _num("n peaking (alpha)", float(getattr(base0, "profile_peaking_ne", 1.0)), 0.1, help="Profile peaking control (if profiles enabled).")
        profile_peaking_T  = _num("T peaking (alpha)", float(getattr(base0, "profile_peaking_T", 1.5)), 0.1, help="Profile peaking control (if profiles enabled).")

        # Optional: compute TF Jop from winding-pack geometry (screening proxy)
        with st.expander("TF winding-pack Jop (optional)", expanded=False):
            tf_Jop_from_wp_geometry = st.checkbox(
                "Compute TF Jop from required ampere-turns and winding-pack area",
                value=bool(getattr(base0, "tf_Jop_from_wp_geometry", False)),
                help="If enabled, SHAMS derives an engineering current density from Bt,R0 and an explicit winding-pack area proxy (no detailed magnet model).",
            )
            tf_wp_width_m = _num("TF winding-pack width [m]", float(getattr(base0, "tf_wp_width_m", 0.25)), 0.01, min_value=0.05, help="Radial width of the winding pack used for Jop-from-geometry proxy.")
            tf_wp_height_factor = _num("TF winding-pack height factor [-]", float(getattr(base0, "tf_wp_height_factor", 2.4)), 0.05, min_value=0.5, help="Height proxy: H_wp = factor * (a*κ).")
            tf_wp_fill_factor = _num("TF winding-pack fill factor [-]", float(getattr(base0, "tf_wp_fill_factor", 1.0)), 0.05, min_value=0.05, max_value=1.0, help="Fraction of winding-pack area treated as conducting cross-section in the Jop-from-geometry proxy.")

        base = PointInputs(
            R0_m=R0_m, a_m=a_m, kappa=kappa, delta=delta, Bt_T=Bt_T,
            tf_Jop_from_wp_geometry=tf_Jop_from_wp_geometry,
            tf_wp_width_m=tf_wp_width_m,
            tf_wp_height_factor=tf_wp_height_factor,
            tf_wp_fill_factor=tf_wp_fill_factor,
            Ip_MA=float(getattr(base0, "Ip_MA", 8.0)),
            Ti_keV=Ti_keV,
            fG=float(getattr(base0, "fG", 0.8)),
            Paux_MW=float(getattr(base0, "Paux_MW", 20.0)),
            t_shield_m=t_shield_m,
            Ti_over_Te=Ti_over_Te,
            confinement_model=confinement_model,
            profile_model=profile_model,
            profile_peaking_ne=profile_peaking_ne,
            profile_peaking_T=profile_peaking_T,
            zeff_mode=zeff_mode,
            steady_state=steady_state,
            P_net_min_MW=P_net_min_MW,
            calib_confinement=float(st.session_state.get('calib_confinement', 1.0)),
            calib_divertor=float(st.session_state.get('calib_divertor', 1.0)),
            calib_bootstrap=float(st.session_state.get('calib_bootstrap', 1.0)),
        )

    with st.expander("Targets and iteration variables", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            use_Q = st.checkbox("Target Q", value=True)
            Q_t = _num("Q target [-]", 10.0, 0.5, help="Target fusion gain Q.", key=PD_KEYS["Q_tgt"])
            # Default is intentionally conservative to help first-run success:
            # a single target (Q) with a single solved variable (Paux).
            use_H = st.checkbox("Target H98", value=False)
            H_t = _num("H98 target [-]", 1.15, 0.05, help="Target confinement H-factor.", key=PD_KEYS["H98_tgt"])
        with col2:
            use_Pnet = st.checkbox("Target net electric", value=False)
            Pnet_t = _num("P_net target [MW(e)]", 50.0, 5.0, help="Target net electric power.")
            # iteration vars
            st.markdown("**Iteration variables (solved)**")
            solve_Ip = st.checkbox("Solve Ip [MA]", value=False)
            solve_fG = st.checkbox("Solve fG [-]", value=False)
            solve_Paux = st.checkbox("Solve Paux [MW]", value=True)

        targets = {}
        if use_Q:
            targets["Q_DT_eqv"] = float(Q_t)
        if use_H:
            targets["H98"] = float(H_t)
        if use_Pnet:
            targets["P_e_net_MW"] = float(Pnet_t)

        variables = {}
        if solve_Ip:
            variables["Ip_MA"] = (float(base.Ip_MA), 0.5*float(base.Ip_MA), 1.8*float(base.Ip_MA))
        if solve_fG:
            variables["fG"] = (float(base.fG), 0.2, 1.2)
        if solve_Paux:
            variables["Paux_MW"] = (float(base.Paux_MW), 0.0, max(200.0, 3.0*float(base.Paux_MW)))

        tol = _num("Solver tolerance", 1e-3, 1e-3, help="Absolute tolerance on each target residual.", min_value=1e-5, max_value=1e-1)
        damping = _num("Damping", 0.6, 0.05, help="Newton step damping for robustness.", min_value=0.1, max_value=1.0)

        
        max_iter = int(_num("Max iterations", 35.0, 1.0, help="Maximum Newton iterations for Systems solve.", min_value=1.0, max_value=500.0))
        override_trust = st.checkbox(
            "Override trust-region Δ (scaled)",
            value=False,
            help="Optional step-size cap in scaled variable space. Lower Δ for harder/brittle solves; raise Δ for faster convergence when stable.",
        )
        trust_delta = None
        if override_trust:
            trust_delta = _num(
                "Trust-region Δ (scaled)",
                5.0,
                0.5,
                help="Caps max(|dx_scaled|) per iteration. Smaller = safer steps; larger = more aggressive.",
                min_value=0.1,
                max_value=50.0,
            )
        st.caption("Solver trace will show `trust_region` events when steps are clipped or Δ adapts.")
        block_solve = st.checkbox(
            "Block-ordered solve (density → power → confinement → exhaust)",
            value=False,
            help="Runs a staged solve to reduce singular Jacobians. Stages are heuristic and fully traced.",
        )

        do_precheck = st.checkbox(
            "Feasibility-first precheck (explicit)",
            value=True,
            help="Before running Newton iterations, evaluate targets/constraints at variable bounds to detect obviously impossible target combinations. This does not change physics or solver behavior; it only exits early with an explicit reason when infeasibility is detected within the declared bounds.",
        )
        do_continuation = st.checkbox(
            "Continuation ramp to targets (path-following)",
            value=True,
            help="For coupled solves, ramp targets from the starting-point values toward the requested targets in small steps. Each step is solved explicitly and logged as `cont_step` / `cont_result`. This is a UI-side workflow for robustness; physics/models are unchanged.",
        )
        cont_steps = int(
            _num(
                "Continuation steps",
                10.0,
                1.0,
                help="Number of continuation increments (only used when continuation is enabled and the solve is coupled).",
                min_value=2.0,
                max_value=50.0,
            )
        )
        st.caption("Continuation is applied only when there is more than one target or more than one solved variable.")

    run = st.button("Run systems solve", type="primary", use_container_width=True, disabled=(len(targets)==0 or len(variables)==0))
    if run:
        _warn_unrealistic_point_inputs(base, context="Systems")
        st.info("Running coupled solve…")
        log = st.empty()
        last = None

        coupled = (len(targets) > 1) or (len(variables) > 1)
        base_for_solve = base

        # -----------------------------
        # Feasibility-first precheck (explicit; UI-side only)
        # -----------------------------
        if do_precheck and len(variables) > 0:
            import itertools as _it

            def _set_vars(inp0: PointInputs, assign: dict) -> PointInputs:
                d = dict(inp0.__dict__)
                d.update(assign)
                return PointInputs(**d)

            var_items = [(k, float(v[1]), float(v[2])) for k, v in variables.items()]  # (name, lo, hi)

            # Build corner points (bounded to keep work predictable)
            if len(var_items) <= 4:
                combos = list(_it.product(*[(lo, hi) for _, lo, hi in var_items]))
            else:
                # Too many corners: take a small, deterministic subset
                combos = []
                lows = tuple(lo for _, lo, _ in var_items)
                highs = tuple(hi for _, _, hi in var_items)
                combos.append(lows)
                combos.append(highs)
                for i in range(min(6, len(var_items))):
                    mid = list(lows)
                    mid[i] = highs[i]
                    combos.append(tuple(mid))

            corner_points = []
            for vals in combos:
                assign = {var_items[i][0]: float(vals[i]) for i in range(len(var_items))}
                corner_points.append(_set_vars(base_for_solve, assign))

            corner_outs = []
            corner_constraints = []
            for cp in corner_points:
                o = hot_ion_point(cp)
                corner_outs.append(o)
                corner_constraints.append(evaluate_constraints(o))

            # Target reach check (range on corners; explicit and conservative)
            unreachable = []
            for tk, tv in targets.items():
                vs = []
                for o in corner_outs:
                    try:
                        vv = float(o.get(tk, float("nan")))
                    except Exception:
                        vv = float("nan")
                    if vv == vv and abs(vv) != float("inf"):
                        vs.append(vv)
                if not vs:
                    unreachable.append({"target": tk, "reason": "not_computed_at_corners"})
                    continue
                vmin, vmax = min(vs), max(vs)
                if float(tv) < vmin or float(tv) > vmax:
                    unreachable.append({"target": tk, "target_value": float(tv), "corner_min": float(vmin), "corner_max": float(vmax)})

            # Hard constraints impossible within bounds (fail at all corners)
            always_failed = []
            names = {}
            for clist in corner_constraints:
                for c in clist:
                    try:
                        sev = getattr(c, "severity", "hard")
                    except Exception:
                        sev = "hard"
                    if sev != "hard":
                        continue
                    names.setdefault(c.name, []).append(bool(getattr(c, "passed", False)))
            for cname, passed_list in names.items():
                if passed_list and (not any(passed_list)):
                    always_failed.append(cname)

            if unreachable or always_failed:
                fail = {
                    "event": "fail",
                    "reason": "precheck_infeasible",
                    "unreachable_targets": unreachable,
                    "hard_constraints_failed_at_all_corners": always_failed,
                }
                log.code(json.dumps(fail, indent=2, sort_keys=True))
                st.error("Precheck: target combination and/or hard constraints appear infeasible within the declared variable bounds. Adjust bounds/targets and retry.")
                st.stop()

        # -----------------------------
        # Continuation ramp to targets (explicit; UI-side only)
        # -----------------------------
        if coupled and do_continuation and cont_steps >= 2:
            try:
                out0 = hot_ion_point(base_for_solve)
            except Exception:
                out0 = {}

            start_targets = {}
            for k in targets.keys():
                try:
                    v = float(out0.get(k, float("nan")))
                except Exception:
                    v = float("nan")
                if v == v and abs(v) != float("inf"):
                    start_targets[k] = v

            def _make_req(_base: PointInputs, _t: dict) -> SolverRequest:
                opts = {"multistart": True, "restarts": 8}
                if trust_delta is not None:
                    opts["trust_delta"] = float(trust_delta)
                if block_solve:
                    opts["block_solve"] = True
                return SolverRequest(base=_base, targets=_t, variables=variables, max_iter=max_iter, tol=float(tol), damping=float(damping), options=opts)

            base_stage = base_for_solve
            for s in range(1, int(cont_steps)):
                alpha = float(s) / float(cont_steps)
                step_targets = {}
                for k, final in targets.items():
                    if k in start_targets:
                        step_targets[k] = float(start_targets[k] + alpha * (float(final) - float(start_targets[k])))
                    else:
                        step_targets[k] = float(final)

                log.code(json.dumps({"event": "cont_step", "step": float(s), "n_steps": float(cont_steps), "alpha": alpha, "targets": step_targets}, indent=2, sort_keys=True))
                _res = solve_request(_make_req(base_stage, step_targets), backend=DefaultTargetSolverBackend())
                log.code(json.dumps({"event": "cont_result", "step": float(s), "ok": bool(_res.ok), "iters": float(_res.iters), "message": _res.message}, indent=2, sort_keys=True))
                if not _res.ok:
                    fail = {"event": "fail", "reason": "continuation_step_fail", "step": float(s), "alpha": alpha, "message": _res.message}
                    log.code(json.dumps(fail, indent=2, sort_keys=True))
                    st.error("Continuation step failed. Adjust targets/bounds or disable continuation.")
                    st.stop()
                base_stage = _res.inp

            base_for_solve = base_stage
        try:
            for step in solve_for_targets_stream(
                base_for_solve,
                targets=targets,
                variables=variables,
                max_iter=max_iter,
                tol=float(tol),
                damping=float(damping),
                trust_delta=(float(trust_delta) if trust_delta is not None else None),
            ):
                last = step
                log.code(json.dumps(step, indent=2, sort_keys=True))
            req = SolverRequest(base=base_for_solve, targets=targets, variables=variables, max_iter=max_iter, tol=float(tol), damping=float(damping), options={"multistart": True, "restarts": 8, **({"trust_delta": float(trust_delta)} if trust_delta is not None else {}), **({"block_solve": True} if block_solve else {})})
            import time as _time
            t_solve0 = _time.perf_counter()
            res = solve_request(req, backend=DefaultTargetSolverBackend())
            wall_s = float(_time.perf_counter() - t_solve0)
            inp_sol = res.inp
            out_sol = res.out
            st.success(f"Done. Converged={res.ok}, iterations={res.iters}")

            if not res.ok and ("Ip_MA" in variables and "fG" in variables) and ("H98" in targets and "Q_DT_eqv" in targets):
                with st.expander("Target feasibility at (Iₚ, f_G) bound corners", expanded=True):
                    try:
                        from solvers.constraint_solver import evaluate_targets_at_corners
                        lo0, hi0 = float(variables["Ip_MA"][1]), float(variables["Ip_MA"][2])
                        lo1, hi1 = float(variables["fG"][1]), float(variables["fG"][2])
                        rows = evaluate_targets_at_corners(base, {"H98": float(targets["H98"]), "Q_DT_eqv": float(targets["Q_DT_eqv"])}, ("Ip_MA", lo0, hi0), ("fG", lo1, hi1))
                        import pandas as _pd  # type: ignore
                        st.dataframe(_pd.DataFrame(rows), use_container_width=True)
                    except Exception as _e:
                        st.caption(f"Corner table unavailable: {_e}")
            st.session_state.last_point_inp = inp_sol
            st.session_state.last_point_out = out_sol

            constraints_list = evaluate_constraints(out_sol)
            solver_meta = {"message": res.message, "trace": res.trace or []}
            artifact = build_run_artifact(inputs=dict(inp_sol.__dict__), outputs=dict(out_sol), constraints=constraints_list, meta={"mode":"systems"}, solver=solver_meta, baseline_inputs=dict(base.__dict__), subsystems={"fidelity": st.session_state.get("fidelity_config", {}), "calibration": {"confinement": float(st.session_state.get("calib_confinement",1.0)), "divertor": float(st.session_state.get("calib_divertor",1.0)), "bootstrap": float(st.session_state.get("calib_bootstrap",1.0))}})
            st.download_button(
                "Download systems-mode run artifact JSON",
                data=json.dumps(artifact, indent=2, sort_keys=True),
                file_name="shams_run_artifact_systems.json",
                mime="application/json",
                use_container_width=True,
            )

            st.markdown("### Key results")
            kcols = st.columns(4)
            def _k(metric, key, fmt="{:.3g}"):
                v = float(out_sol.get(key, float("nan")))
                with metric:
                    st.metric(key, fmt.format(v) if v==v else "NaN")
            _k(kcols[0], "Q_DT_eqv", "{:.3g}")
            _k(kcols[1], "H98", "{:.3g}")
            _k(kcols[2], "P_e_net_MW", "{:.3g}")
            _k(kcols[3], "q_div_MW_m2", "{:.3g}")

            # constraints dashboard
            with st.expander("Constraints & margins (systems mode)", expanded=True):
                rows_c = []
                for c in constraints_list:
                    try:
                        margin = float(getattr(c, "margin"))
                    except Exception:
                        margin = float("nan")
                    rows_c.append({
                        "constraint": c.name,
                        "sense": c.sense,
                        "value": c.value,
                        "limit": c.limit,
                        "units": c.units,
                        "passed": bool(c.passed),
                        "margin_frac": margin,
                        "severity": getattr(c, "severity", "hard"),
                        "note": c.note,
                    })
                dfc = pd.DataFrame(rows_c)
                st.dataframe(dfc, use_container_width=True)

            # Sankey + radial build
            with st.expander("Plots (radial build + power balance)", expanded=False):
                try:
                    import tempfile, os
                    tmpdir = tempfile.mkdtemp(prefix="shams_systems_")
                    rb = os.path.join(tmpdir, "radial_build.png")
                    plot_radial_build_from_artifact(artifact, rb)
                    st.image(rb, caption="Radial build (proxy)", use_container_width=True)
                except Exception as e:
                    st.warning(f"Radial build plot unavailable: {e}")
                try:
                    import tempfile, os
                    from shams_io.sankey import build_power_balance_sankey
                    import plotly.graph_objects as go
                    sank = build_power_balance_sankey(artifact)
                    fig = go.Figure(data=[go.Sankey(**sank)])
                    st.plotly_chart(fig, use_container_width=True)
                except Exception as e:
                    st.warning(f"Sankey unavailable: {e}")

        except Exception as e:
            st.error(f"Systems solver error: {e}")

with tab_scan:
    st.subheader("Scan Lab")
    st.markdown(
        """
This tab runs **grid scans** across your chosen axes (Ti, H98, a, Q target, and confinement gain *g_conf*).  
Each grid point is solved with the same Phase‑1 proxy models as **Point Designer**, then screened by the
**solver bounds** and optional **clean‑design knobs**.

**Workflow**
1) Set machine / plasma assumptions (R₀, B₀, shield, Paux, Ti/Te).  
2) Define scan axes as **start → stop** with **step** (inclusive grid).  
3) Set solver bounds (Iₚ, fG) and numerical tolerance.  
4) Optional: require H‑mode access with a PLH margin.  
5) Run the scan → inspect feasible points, export CSV, and push any point to *Point Designer*.

**Tips**
- If you get zero feasible points, first widen Iₚ bounds and relax βN/q95/PSOL/R, then refine steps later.
- Use coarser steps to explore, then tighten steps around a promising region.
        """
    )

    with st.expander("Parameter guide (units, meaning, min/max)"):
        st.markdown(
            """
Below are the **recommended** ranges used for input validation in this UI.  
They are intentionally broad to avoid over‑constraining early exploration.

| Parameter | Meaning | Recommended min | Recommended max |
|---|---|---:|---:|
| R₀ (m) | Major radius | 0.5 | 10 |
| B₀ (T) | Toroidal field on axis | 1 | 25 |
| Shield (m) | Neutron shield thickness | 0 | 2 |
| P_aux (MW) | External heating power | 0 | 200 |
| P_aux for Q (MW) | Power used in Q = P_fus / P_aux_for_Q | 0 | 200 |
| Tᵢ/Tₑ (–) | Ion/electron temperature ratio | 0.5 | 5 |
| Ti_start/stop (keV) | Ion temperature scan bounds | 1 | 40 |
| Ti_step (keV) | Ion temperature step | 0.05 | 5 |
| H98_start/stop (–) | H98y2 confinement multiplier bounds | 0.5 | 3 |
| H98_step (–) | H98 step | 0.01 | 0.5 |
| a_min/a_max (m) | Minor radius scan bounds | 0.2 | 5 |
| a_step (m) | Minor radius step | 0.001 | 1 |
| Q_start/stop (–) | Target Q scan bounds (screening target) | 0.1 | 100 |
| Q_step (–) | Q step | 0.05 | 20 |
| g_conf start/stop (–) | Additional confinement gain factor | 0.5 | 5 |
| g_conf step (–) | g_conf step | 0.01 | 1 |
| Iₚ bounds (MA) | Solver search bounds for plasma current | 1 | 50 |
| fG bounds (–) | Greenwald fraction screening bounds | 0.01 | 1.5 |
| tol (–) | Numerical tolerance for the solver | 1e-6 | 1e-2 |
| Zeff (–) | Effective charge | 1.0 | 4.0 |
| dilution_fuel (–) | Fuel dilution factor (≤1) | 0.2 | 1.0 |
| extra_rad_factor (–) | Extra radiation multiplier | 0 | 2 |
| alpha_loss_frac (–) | Fraction of alpha power lost | 0 | 0.5 |
| kappa (–) | Elongation | 1.0 | 3.0 |
| q95_min (–) | Minimum q95 constraint | 1.5 | 10 |
| betaN_max (–) | Maximum normalized beta constraint | 1.0 | 8 |
| C_bs (–) | Bootstrap coefficient proxy | 0 | 1 |
| f_bs_max (–) | Max bootstrap fraction | 0 | 1 |
| PSOL/R max (MW/m) | SOL power per major radius limit | 0 | 200 |
| PLH_margin (–) | Extra margin over PLH if H‑mode required | 0 | 1 |
            """
        )

    with st.expander("Scan Lab → Physics block mapping (what each parameter affects)"):
        st.caption("UI-only helper: shows which Phase‑1 physics/systems blocks each Scan Lab parameter feeds.")
        rows = []
        # Keep the ordering aligned with the form layout.
        ordered = [
            ("R0", "Major radius R₀ (m)"),
            ("B0", "Toroidal field on axis B₀ (T)"),
            ("tshield", "Neutron shield thickness (m)"),
            ("Paux", "Auxiliary heating power P_aux (MW)"),
            ("Paux_for_Q", "Aux power used in Q definition (MW)"),
            ("Ti_over_Te", "Ion-to-electron temperature ratio Tᵢ/Tₑ (–)"),
            ("Ti", "Ti axis (Ti_start/stop/step)"),
            ("H98", "H98 axis (H98_start/stop/step)"),
            ("a", "a axis (a_min/a_max/a_step)"),
            ("Q", "Q axis (Q_start/stop/Q_step)"),
            ("g_conf", "g_conf axis (start/stop/step)"),
            ("Ip_bounds", "Iₚ bounds (I_p,min / I_p,max)"),
            ("fG_bounds", "fG bounds (fG_min / fG_max)"),
            ("tol", "tol"),
            ("Zeff", "Zeff"),
            ("dilution_fuel", "dilution_fuel"),
            ("extra_rad_factor", "extra_rad_factor"),
            ("alpha_loss_frac", "alpha_loss_frac"),
            ("kappa", "kappa"),
            ("q95_min", "q95_min"),
            ("betaN_max", "betaN_max"),
            ("C_bs", "C_bs"),
            ("f_bs_max", "f_bs_max"),
            ("PSOL_over_R_max", "PSOL/R max"),
            ("require_Hmode", "Require H-mode access"),
            ("PLH_margin", "PLH_margin"),
            ("tblanket_m", "Blanket thickness (inboard)"),
            ("t_vv_m", "Vacuum vessel thickness (inboard)"),
            ("t_gap_m", "Inboard gap / clearance"),
            ("t_tf_struct_m", "TF structure thickness (inboard)"),
            ("t_tf_wind_m", "TF winding pack thickness (inboard)"),
            ("Bpeak_factor", "Bpeak_factor"),
            ("sigma_allow_MPa", "Allowable coil hoop stress"),
            ("Tcoil_K", "HTS operating temperature"),
            ("hts_margin_min", "HTS margin min"),
            ("Vmax_kV", "Max dump voltage limit"),
            ("q_div_max_MW_m2", "Max divertor heat flux limit"),
            ("TBR_min", "TBR_min"),
            ("hts_lifetime_min_yr", "Minimum HTS lifetime"),
            ("P_net_min_MW", "Minimum net electric power"),
        ]
        for k, label in ordered:
            blocks = _scan_blocks(k)
            rows.append(
                {
                    "Parameter": label,
                    "Badge": _scan_badge(k),
                    "Physics blocks": ", ".join(blocks) if blocks else "(unmapped)",
                }
            )
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    c1, c2 = st.columns([1.2, 1.0])
    with c1:
        st.markdown("### Scan inputs")
        with st.form("scan_form"):
            with st.expander("Scan inputs", expanded=True):
                with st.expander("Geometry & field", expanded=True):
                    R0 = _num(
                        _scan_label("Major radius R₀ (m)", "R0"),
                        1.81,
                        0.01,
                        min_value=0.5,
                        max_value=10.0,
                        help=_scan_help("Major radius of the tokamak (sets size/scale).", "R0"),
                    )
                    B0 = _num(
                        _scan_label("Toroidal field on axis B₀ (T)", "B0"),
                        10.0,
                        0.1,
                        min_value=1.0,
                        max_value=25.0,
                        help=_scan_help("Toroidal magnetic field at the magnetic axis.", "B0"),
                    )
                    tshield = _num(
                        _scan_label("Neutron shield thickness (m)", "tshield"),
                        0.8,
                        0.01,
                        min_value=0.0,
                        max_value=2.0,
                        help=_scan_help("Inboard neutron shielding thickness used in radial build and constraints.", "tshield"),
                    )

                with st.expander("Heating & assumptions", expanded=True):
                    Paux = _num(
                        _scan_label("Auxiliary heating power P_aux (MW)", "Paux"),
                        48.0,
                        1.0,
                        min_value=0.0,
                        max_value=200.0,
                        help=_scan_help("External heating power applied during the operating point evaluation.", "Paux"),
                    )
                    Paux_for_Q = _num(
                        _scan_label("Aux power used in Q definition (MW)", "Paux_for_Q"),
                        48.0,
                        1.0,
                        min_value=0.0,
                        max_value=200.0,
                        help=_scan_help("Denominator power for Q = P_fus / P_aux_for_Q (can differ from P_aux if desired).", "Paux_for_Q"),
                    )
                    Ti_over_Te = _num(
                        _scan_label("Ion-to-electron temperature ratio Tᵢ/Tₑ (–)", "Ti_over_Te"),
                        2.0,
                        0.1,
                        min_value=0.5,
                        max_value=5.0,
                        help=_scan_help("Assumed ratio of ion to electron temperature (Ti = (Ti/Te)*Te).", "Ti_over_Te"),
                    )

                with st.expander("Axes & targets", expanded=True):
                    st.markdown("**Axes**")
                    Ti_start = _num(_scan_label("Ti_start [keV]", "Ti"), 12.0, 0.25, min_value=1.0, max_value=40.0, help=_scan_help("Start of ion temperature scan (keV).", "Ti"))
                    Ti_stop  = _num(_scan_label("Ti_stop [keV]", "Ti"), 12.0, 0.25, min_value=1.0, max_value=40.0, help=_scan_help("Stop of ion temperature scan (keV).", "Ti"))
                    Ti_step  = _num(_scan_label("Ti_step [keV]", "Ti"), 1.0, 0.05, min_value=0.05, max_value=5.0, help=_scan_help("Ion temperature step size (keV).", "Ti"))

                    H98_start = _num(_scan_label("H98_start", "H98"), 1.15, 0.05, min_value=0.5, max_value=3.0, help=_scan_help("Start of H98 confinement multiplier scan.", "H98"))
                    H98_stop  = _num(_scan_label("H98_stop", "H98"), 1.15, 0.05, min_value=0.5, max_value=3.0, help=_scan_help("Stop of H98 confinement multiplier scan.", "H98"))
                    H98_step  = _num(_scan_label("H98_step", "H98"), 0.1, 0.01, min_value=0.01, max_value=0.5, help=_scan_help("H98 step size.", "H98"))

                    a_min  = _num(_scan_label("a_min [m]", "a"), 0.57, 0.01, min_value=0.2, max_value=5.0, help=_scan_help("Minor radius scan lower bound (m).", "a"))
                    a_max  = _num(_scan_label("a_max [m]", "a"), 0.57, 0.01, min_value=0.2, max_value=5.0, help=_scan_help("Minor radius scan upper bound (m).", "a"))
                    a_step = _num(_scan_label("a_step [m]", "a"), 0.05, 0.005, min_value=0.001, max_value=1.0, help=_scan_help("Minor radius step size (m).", "a"))

                    Q_start = _num(_scan_label("Q_start", "Q"), 10.0, 0.1, min_value=0.1, max_value=100.0, help=_scan_help("Start of target Q scan (used as screening target).", "Q"))
                    Q_stop  = _num(_scan_label("Q_stop", "Q"), 10.0, 0.1, min_value=0.1, max_value=100.0, help=_scan_help("Stop of target Q scan (used as screening target).", "Q"))
                    Q_step  = _num(_scan_label("Q_step", "Q"), 1.0, 0.05, min_value=0.05, max_value=20.0, help=_scan_help("Target Q step size.", "Q"))

                    gconf_start = _num(_scan_label("g_conf start", "g_conf"), 1.0, 0.05, min_value=0.5, max_value=5.0, help=_scan_help("Start of additional confinement gain factor scan.", "g_conf"))
                    gconf_stop  = _num(_scan_label("g_conf stop", "g_conf"), 2.5, 0.05, min_value=0.5, max_value=5.0, help=_scan_help("Stop of additional confinement gain factor scan.", "g_conf"))
                    gconf_step  = _num(_scan_label("g_conf step", "g_conf"), 0.05, 0.01, min_value=0.01, max_value=1.0, help=_scan_help("g_conf step size.", "g_conf"))

                with st.expander("Solver & screening", expanded=True):
                    st.markdown("**Solver bounds & screening**")
                    Ip_min = _num(_scan_label("Plasma current lower bound I_p,min (MA)", "Ip_bounds"), 10.0, 1.0, min_value=1.0, max_value=50.0, help=_scan_help("Lower bound for solver search over plasma current.", "Ip_bounds"))
                    Ip_max = _num(_scan_label("Plasma current upper bound I_p,max (MA)", "Ip_bounds"), 120.0, 1.0, min_value=1.0, max_value=50.0, help=_scan_help("Upper bound for solver search over plasma current.", "Ip_bounds"))
                    fG_min = _num(_scan_label("fG_min", "fG_bounds"), 0.01, 0.01, min_value=0.01, max_value=1.5, help=_scan_help("Minimum Greenwald fraction considered feasible.", "fG_bounds"))
                    fG_max = _num(_scan_label("fG_max", "fG_bounds"), 1.2, 0.01, min_value=0.01, max_value=1.5, help=_scan_help("Maximum Greenwald fraction considered feasible.", "fG_bounds"))
                    tol = _num(_scan_label("tol", "tol"), 1e-3, 1e-4, fmt="%.1e", min_value=1e-6, max_value=1e-2, help=_scan_help("Numerical tolerance used in the root-finding/constraint solver (smaller is stricter but can reduce convergence).", "tol"))

                    Zeff = _num(_scan_label("Zeff", "Zeff"), 1.8, 0.1, min_value=1.0, max_value=4.0, help=_scan_help("Effective plasma charge state used in radiation/transport proxies.", "Zeff"))
                    dilution_fuel = _num(_scan_label("dilution_fuel", "dilution_fuel"), 0.85, 0.01, min_value=0.2, max_value=1.0, help=_scan_help("Fuel dilution factor (fraction of plasma that is D-T fuel).", "dilution_fuel"))
                    extra_rad_factor = _num(_scan_label("extra_rad_factor", "extra_rad_factor"), 0.2, 0.05, min_value=0.0, max_value=2.0, help=_scan_help("Multiplier for additional radiation losses beyond the base model.", "extra_rad_factor"))
                    alpha_loss_frac = _num(_scan_label("alpha_loss_frac", "alpha_loss_frac"), 0.05, 0.01, min_value=0.0, max_value=0.5, help=_scan_help("Fraction of alpha heating power assumed lost (not deposited in plasma).", "alpha_loss_frac"))

                    kappa = _num(_scan_label("kappa", "kappa"), 1.8, 0.05, min_value=1.0, max_value=3.0, help=_scan_help("Plasma elongation κ.", "kappa"))
                    q95_min = _num(_scan_label("q95_min", "q95_min"), 3.0, 0.1, min_value=1.5, max_value=10.0, help=_scan_help("Minimum allowed q95 for MHD/operational margin screening.", "q95_min"))
                    betaN_max = _num(_scan_label("betaN_max", "betaN_max"), 4.0, 0.1, min_value=1.0, max_value=8.0, help=_scan_help("Maximum normalized beta βN allowed for screening.", "betaN_max"))
                    C_bs = _num(_scan_label("C_bs", "C_bs"), 0.15, 0.01, min_value=0.0, max_value=1.0, help=_scan_help("Proxy coefficient for bootstrap current fraction estimation.", "C_bs"))
                    f_bs_max = _num(_scan_label("f_bs_max", "f_bs_max"), 0.60, 0.01, min_value=0.0, max_value=1.0, help=_scan_help("Maximum allowed bootstrap current fraction (screening).", "f_bs_max"))
                    PSOL_over_R_max = _num(_scan_label("PSOL/R max [MW/m]", "PSOL_over_R_max"), 80.0, 1.0, min_value=0.0, max_value=200.0, help=_scan_help("Upper limit on SOL power normalized by major radius (proxy for divertor loading).", "PSOL_over_R_max"))

                    require_Hmode = st.checkbox(_scan_label("Require H-mode access (Paux ≥ (1+margin)·PLH)", "require_Hmode"), value=False, help=_scan_help("Enable additional screening that requires crossing the L-H threshold with a margin.", "require_Hmode"))
                    PLH_margin = _num(_scan_label("PLH_margin", "PLH_margin"), 0.0, 0.05, min_value=0.0, max_value=1.0, help=_scan_help("If H-mode is required, enforce Paux ≥ (1+margin)*PLH. Margin is this value.", "PLH_margin"))

                    st.markdown("**Clean design knobs** (used as additional pass/fail checks)")
                    clean_knobs = {
                        "tblanket_m": float(_num(_scan_label("Blanket thickness (inboard) (m)", "tblanket_m"), 0.5, 0.01, min_value=0.0, max_value=2.0, help=_scan_help("Inboard blanket thickness used in radial build screening.", "tblanket_m"))),
                        "t_vv_m": float(_num(_scan_label("Vacuum vessel thickness (inboard) (m)", "t_vv_m"), 0.06, 0.005, min_value=0.0, max_value=0.5, help=_scan_help("Inboard vacuum vessel thickness for radial build.", "t_vv_m"))),
                        "t_gap_m": float(_num(_scan_label("Inboard gap / clearance (m)", "t_gap_m"), 0.02, 0.005, min_value=0.0, max_value=0.5, help=_scan_help("Assembly/clearance gap between components on inboard side.", "t_gap_m"))),
                        "t_tf_struct_m": float(_num(_scan_label("TF structure thickness (inboard) (m)", "t_tf_struct_m"), 0.15, 0.01, min_value=0.0, max_value=2.0, help=_scan_help("Structural support thickness for TF coil on inboard side.", "t_tf_struct_m"))),
                        "t_tf_wind_m": float(_num(_scan_label("TF winding pack thickness (inboard) (m)", "t_tf_wind_m"), 0.10, 0.01, min_value=0.0, max_value=2.0, help=_scan_help("Winding pack thickness for TF coil on inboard side.", "t_tf_wind_m"))),
                        "Bpeak_factor": float(_num(_scan_label("Bpeak_factor", "Bpeak_factor"), 1.25, 0.01, min_value=1.0, max_value=2.0, help=_scan_help("Ratio of peak field on coil to on-axis field (geometry/coil build proxy).", "Bpeak_factor"))),
                        "sigma_allow_MPa": float(_num(_scan_label("Allowable coil hoop stress (MPa)", "sigma_allow_MPa"), 850.0, 10.0, min_value=10.0, max_value=2000.0, help=_scan_help("Maximum allowable TF coil hoop stress used in stress screening.", "sigma_allow_MPa"))),
                        "Tcoil_K": float(_num(_scan_label("HTS operating temperature (K)", "Tcoil_K"), 20.0, 1.0, min_value=4.0, max_value=80.0, help=_scan_help("Operating temperature for HTS performance proxy.", "Tcoil_K"))),
                        "hts_margin_min": float(_num(_scan_label("HTS margin min", "hts_margin_min"), 0.15, 0.01, min_value=0.0, max_value=0.8, help=_scan_help("Minimum required HTS operating margin (fraction).", "hts_margin_min"))),
                        "Vmax_kV": float(_num(_scan_label("Max dump voltage limit (kV)", "Vmax_kV"), 20.0, 1.0, min_value=1.0, max_value=200.0, help=_scan_help("Upper limit on coil dump voltage during a fast discharge.", "Vmax_kV"))),
                        "q_div_max_MW_m2": float(_num(_scan_label("Max divertor heat flux limit (MW/m²)", "q_div_max_MW_m2"), 10.0, 0.5, min_value=0.1, max_value=50.0, help=_scan_help("Screening limit for divertor heat flux proxy.", "q_div_max_MW_m2"))),
                        "TBR_min": float(_num(_scan_label("TBR_min", "TBR_min"), 1.05, 0.01, min_value=0.0, max_value=2.0, help=_scan_help("Minimum tritium breeding ratio required for screening.", "TBR_min"))),
                        "hts_lifetime_min_yr": float(_num(_scan_label("Minimum HTS lifetime (years)", "hts_lifetime_min_yr"), 3.0, 0.5, min_value=0.0, max_value=50.0, help=_scan_help("Minimum required HTS lifetime (years) for screening.", "hts_lifetime_min_yr"))),
                        "P_net_min_MW": float(_num(_scan_label("Minimum net electric power (MW)", "P_net_min_MW"), 0.0, 10.0, min_value=-1e6, max_value=1e6, help=_scan_help("Minimum net electric power requirement (set to 0 to require net-positive).", "P_net_min_MW"))),
                    }

            run_scan_btn = st.form_submit_button("Run Scan", type="primary", use_container_width=True)

        # Build the scan specification dict from UI inputs.
        # This must exist even before the button is pressed so the async runner can use it.
        spec: Dict[str, Any] = {
            # Machine / fixed assumptions
            "R0": float(R0),
            "B0": float(B0),
            "kappa": float(kappa),
            "tshield": float(tshield),
            "Paux": float(Paux),
            "Paux_for_Q": float(Paux_for_Q),
            "Ti_over_Te": float(Ti_over_Te),
            "Zeff": float(Zeff),
            "dilution_fuel": float(dilution_fuel),
            "extra_rad_factor": float(extra_rad_factor),
            "alpha_loss_frac": float(alpha_loss_frac),
            "C_bs": float(C_bs),
            "require_Hmode": bool(require_Hmode),
            "PLH_margin": float(PLH_margin),

            # Scan axes / targets
            "Ti_start": float(Ti_start),
            "Ti_stop": float(Ti_stop),
            "Ti_step": float(Ti_step),
            "H98_start": float(H98_start),
            "H98_stop": float(H98_stop),
            "H98_step": float(H98_step),
            "a_min": float(a_min),
            "a_max": float(a_max),
            "a_step": float(a_step),
            "Q_start": float(Q_start),
            "Q_stop": float(Q_stop),
            "Q_step": float(Q_step),
            "gconf_start": float(gconf_start),
            "gconf_stop": float(gconf_stop),
            "gconf_step": float(gconf_step),

            # Solver bounds
            "Ip_min": float(Ip_min),
            "Ip_max": float(Ip_max),
            "fG_min": float(fG_min),
            "fG_max": float(fG_max),
            "tol": float(tol),

            # Screening
            "q95_min": float(q95_min),
            "betaN_max": float(betaN_max),
            "f_bs_max": float(f_bs_max),
            "PSOL_over_R_max": float(PSOL_over_R_max),

            # Extra knobs
            "clean_knobs": dict(clean_knobs),
        }

        # --- Async Scan Runner (keeps UI responsive; tabs stay clickable) ---
        def _ensure_scan_executor() -> None:
            if ("scan_executor" not in st.session_state) or (st.session_state.get("scan_executor") is None):
                st.session_state.scan_executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
            if "scan_future" not in st.session_state:
                st.session_state.scan_future = None
            if "scan_queue" not in st.session_state:
                st.session_state.scan_queue = None
            if "scan_running" not in st.session_state:
                st.session_state.scan_running = False
            if "scan_progress" not in st.session_state:
                st.session_state.scan_progress = {"frac": 0.0, "info": {}}
            if "scan_log_lines" not in st.session_state:
                st.session_state.scan_log_lines = []

        _ensure_scan_executor()

        def _start_scan_async(_spec: Dict[str, Any]) -> None:
            # UI-only guardrails: warn on unrealistic scan baselines / ranges.
            try:
                Ip0 = 0.5 * (float(_spec.get("Ip_min", 0.0)) + float(_spec.get("Ip_max", 0.0)))
                fG0 = 0.5 * (float(_spec.get("fG_min", 0.0)) + float(_spec.get("fG_max", 0.0)))
                a0 = 0.5 * (float(_spec.get("a_min", 0.0)) + float(_spec.get("a_max", 0.0)))
                Ti0 = 0.5 * (float(_spec.get("Ti_start", 0.0)) + float(_spec.get("Ti_stop", 0.0)))
                pi0 = make_point_inputs(
                    R0_m=float(_spec.get("R0", 0.0)),
                    a_m=float(a0),
                    kappa=float(_spec.get("kappa", 0.0)),
                    Bt_T=float(_spec.get("B0", 0.0)),
                    Ip_MA=float(Ip0),
                    Ti_keV=float(Ti0),
                    fG=float(fG0),
                    Paux_MW=float(_spec.get("Paux", 0.0)),
                    t_shield_m=float(_spec.get("tshield", 0.0)),
                    Ti_over_Te=float(_spec.get("Ti_over_Te", 2.0)),
                    zeff=float(_spec.get("Zeff", 1.8)),
                    dilution_fuel=float(_spec.get("dilution_fuel", 0.85)),
                )
                _warn_unrealistic_point_inputs(pi0, context="Scan (baseline)")

                def _npts(lo, hi, step):
                    try:
                        lo = float(lo); hi = float(hi); step = float(step)
                        if step <= 0:
                            return 0
                        if hi < lo:
                            return 0
                        return int(round((hi - lo) / step)) + 1
                    except Exception:
                        return 0

                nTi = _npts(_spec.get("Ti_start"), _spec.get("Ti_stop"), _spec.get("Ti_step", 1.0))
                nH = _npts(_spec.get("H98_start"), _spec.get("H98_stop"), _spec.get("H98_step", 0.05))
                na = _npts(_spec.get("a_min"), _spec.get("a_max"), _spec.get("a_step", 0.01))
                nQ = _npts(_spec.get("Q_start"), _spec.get("Q_stop"), _spec.get("Q_step", 1.0))
                n_total = max(0, nTi) * max(0, nH) * max(0, na) * max(0, nQ)
                if n_total > 50000:
                    st.warning(
                        "Large scan requested" + f" (≈{n_total:,} points). "
                        "This may take a long time. Consider narrowing ranges or using fewer steps."
                    )
            except Exception:
                pass
            q: "queue.Queue[tuple]" = queue.Queue()
            st.session_state.scan_queue = q
            st.session_state.scan_log_lines = []
            st.session_state.scan_progress = {"frac": 0.0, "info": {"stage": "starting"}}
            st.session_state.scan_running = True

            def _progress_cb(frac: float, info: Dict[str, Any]) -> None:
                try:
                    q.put(("progress", float(frac), dict(info)), block=False)
                except Exception:
                    pass

            def _log_cb(line: str) -> None:
                try:
                    q.put(("log", str(line)), block=False)
                except Exception:
                    pass

            # Inject callbacks into the scan spec so run_scan() can stream updates.
            _spec["_progress_cb"] = _progress_cb
            _spec["_log_cb"] = _log_cb

            def _job():
                return run_scan(_spec)

            st.session_state.scan_future = st.session_state.scan_executor.submit(_job)

        def _drain_scan_queue(max_items: int = 400) -> None:
            q = st.session_state.scan_queue
            if q is None:
                return
            n = 0
            while n < max_items:
                try:
                    item = q.get_nowait()
                except Exception:
                    break
                n += 1
                kind = item[0]
                if kind == "progress":
                    _, frac, info = item
                    st.session_state.scan_progress = {"frac": frac, "info": info}
                elif kind == "log":
                    _, line = item
                    st.session_state.scan_log_lines.append(line)

        # Start scan when user presses "Run Scan" (button is inside the form; execution here is OUTSIDE the form)
        if run_scan_btn and (not st.session_state.scan_running):
            _start_scan_async(spec)

        # Render live panel + log while scan runs
        if st.session_state.scan_running:
            _drain_scan_queue()

            st.markdown("### Live run: physics & models")
            with st.expander("Physics & model pipeline (live)", expanded=True):
                st.markdown(
                """
This scan iterates over your chosen parameter grid. For **each** combination it runs:

1. **Build inputs** (geometry, fields, assumed profiles, bounds)
2. **Nested solve**:
   - **Outer**: find **Iₚ** such that predicted **H98** hits your requested value
   - **Inner** (for each Iₚ trial): find **fG** such that predicted **Q** hits your target
3. **Evaluate physics proxies** (power balance, confinement, stability/limits, divertor load proxies)
4. **Screening checks** (q95, βN, bootstrap fraction, PSOL/R, etc.)
5. **Record feasible points** into the results table
                """
            )

            frac = float(st.session_state.scan_progress.get("frac", 0.0))
            info = dict(st.session_state.scan_progress.get("info", {}) or {})
            stage = info.get("stage", "…")
            g_conf = info.get("g_conf")
            Ti = info.get("Ti_keV")
            Hreq = info.get("H98_req")
            a = info.get("a_m")
            Q = info.get("Q")

            st.info(f"Stage: **{stage}**")
            cols = st.columns(6)
            cols[0].metric("Progress", f"{100.0*frac:.1f}%")
            cols[1].metric("g_conf", "—" if g_conf is None else f"{g_conf:.3g}")
            cols[2].metric("Ti (keV)", "—" if Ti is None else f"{Ti:.3g}")
            cols[3].metric("H98 target", "—" if Hreq is None else f"{Hreq:.3g}")
            cols[4].metric("a (m)", "—" if a is None else f"{a:.3g}")
            cols[5].metric("Q target", "—" if Q is None else f"{Q:.3g}")
            st.progress(min(max(frac, 0.0), 1.0))

            with st.expander("Scan log (expand)", expanded=False):
                log_txt = "".join(st.session_state.scan_log_lines[-2000:])
                st.text_area("scan_lab.log", value=log_txt, height=260, label_visibility="collapsed")
                st.download_button(
                    "Download scan log",
                    data=log_txt,
                    file_name="scan_lab.log",
                    mime="text/plain",
                    use_container_width=True,
                )

            # Best-effort auto-refresh while background scan runs
            try:
                from streamlit_autorefresh import st_autorefresh
                st_autorefresh(interval=500, key="scan_autorefresh")
            except Exception:
                pass

            fut = st.session_state.scan_future
            if fut is not None and fut.done():
                try:
                    df_done, meta_done = fut.result()
                except Exception as e:
                    st.session_state.scan_running = False
                    st.session_state.scan_future = None
                    st.error(f"Scan failed: {e}")
                else:
                    st.session_state.scan_df = df_done
                    st.session_state.scan_meta = meta_done
                    st.session_state.scan_log_text = meta_done.get("scan_log_text", "") or "".join(st.session_state.scan_log_lines)
                    st.session_state.scan_running = False
                    st.session_state.scan_future = None

        st.divider()
        st.markdown("### Results Explorer")
        df = st.session_state.scan_df
        meta = st.session_state.get("scan_meta", {}) or {}

if df is None or df.empty:
    st.info("No scan results yet. Run a scan in **Scan Lab**.")
else:
    # Downloads + quick plotting (kept inside Scan Lab tab)
    top_row = st.columns([1.2, 1.0, 1.0])
    with top_row[0]:
        st.caption("Explore feasible points, export, and build quick x–y plots.")
    with top_row[1]:
        try:
            excel_bytes = df_to_excel_bytes(df, meta)
            st.download_button(
                "Download Excel",
                data=excel_bytes,
                file_name="phase1_scan_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception:
            pass
    with top_row[2]:
        st.download_button(
            "Download CSV",
            data=df.to_csv(index=False).encode("utf-8"),
            file_name="phase1_scan_results.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with st.expander("Filters", expanded=True):
        cols = st.columns(4)
        # basic, robust filters (only shown if columns exist)
        f = {}
        with cols[0]:
            if "Q_DT_eqv" in df:
                mn, mx = float(df["Q_DT_eqv"].min()), float(df["Q_DT_eqv"].max())
                f["Q_DT_eqv"] = st.slider("Q_DT_eqv", mn, mx, (mn, mx))
        with cols[1]:
            if "H98" in df:
                mn, mx = float(df["H98"].min()), float(df["H98"].max())
                f["H98"] = st.slider("H98", mn, mx, (mn, mx))
        with cols[2]:
            if "Ip_MA" in df:
                mn, mx = float(df["Ip_MA"].min()), float(df["Ip_MA"].max())
                f["Ip_MA"] = st.slider("Iₚ (MA)", mn, mx, (mn, mx))
        with cols[3]:
            if "q_div_MW_m2" in df:
                mn, mx = float(df["q_div_MW_m2"].min()), float(df["q_div_MW_m2"].max())
                f["q_div_MW_m2"] = st.slider("q_div (MW/m²)", mn, mx, (mn, mx))

        dff = df.copy()
        # apply filters
        for k, (lo, hi) in f.items():
            dff = dff[(dff[k] >= lo) & (dff[k] <= hi)]

    with st.expander("Filtered results table", expanded=True):
        st.dataframe(dff, use_container_width=True, height=420)

    with st.expander("Plot Studio", expanded=True):
        st.caption("Choose any two numeric columns to plot. Optional: color by a third column.")
        ncols = _numeric_cols(dff)
        if len(ncols) < 2:
            st.info("Not enough numeric columns to plot.")
        else:
            c1, c2, c3 = st.columns([1.1, 1.1, 1.0])
            with c1:
                xcol = st.selectbox("x-axis", options=ncols, index=0)
            with c2:
                ycol = st.selectbox("y-axis", options=ncols, index=1 if len(ncols) > 1 else 0)
            with c3:
                ccol = st.selectbox("color (optional)", options=["(none)"] + ncols, index=0)

            plot_df = dff[[xcol, ycol] + ([] if ccol == "(none)" else [ccol])].dropna()
            if plot_df.empty:
                st.info("No points after filtering/NaN removal.")
            else:
                if _HAVE_MPL and plt is not None:
                    fig = plt.figure()
                    ax = fig.add_subplot(111)
                    if ccol == "(none)":
                        ax.scatter(plot_df[xcol], plot_df[ycol], s=16)
                    else:
                        sc = ax.scatter(plot_df[xcol], plot_df[ycol], c=plot_df[ccol], s=16)
                        fig.colorbar(sc, ax=ax, label=ccol)
                    ax.set_xlabel(xcol)
                    ax.set_ylabel(ycol)
                    ax.grid(True, alpha=0.25)
                    st.pyplot(fig, use_container_width=True)
                else:
                    st.scatter_chart(plot_df, x=xcol, y=ycol)
# -----------------------------
# Pareto Lab (PROCESS-like design space exploration)
# -----------------------------
with tab_pareto:
    st.subheader("Pareto Lab (feasible design space + Pareto front)")
    st.markdown(
        "Run a lightweight **Latin-hypercube** sampling study inside bounds, filter feasible points "
        "using the constraint system, then compute the **Pareto front** for selected objectives "
        "(maximization/minimization)."
    )

    base0 = st.session_state.last_point_inp
    if base0 is None:
        base0 = PointInputs(R0_m=1.85, a_m=0.57, kappa=1.8, Bt_T=12.2, Ip_MA=8.0, Ti_keV=15.0, fG=0.8, Paux_MW=20.0)

    with st.expander("Bounds (sampling hyper-rectangle)", expanded=True):
        st.caption("Bounds are applied to the chosen variables during sampling.")
        bcols = st.columns(4)
        b_R0 = (float(base0.R0_m*0.8), float(base0.R0_m*1.25))
        b_Bt = (float(base0.Bt_T*0.7), float(base0.Bt_T*1.15))
        b_Ip = (float(base0.Ip_MA*0.6), float(base0.Ip_MA*1.6))
        b_fG = (0.3, 1.1)
        R0_lo = _num("R0 min [m]", b_R0[0], 0.01)
        R0_hi = _num("R0 max [m]", b_R0[1], 0.01)
        Bt_lo = _num("Bt min [T]", b_Bt[0], 0.1)
        Bt_hi = _num("Bt max [T]", b_Bt[1], 0.1)
        Ip_lo = _num("Ip min [MA]", b_Ip[0], 0.1)
        Ip_hi = _num("Ip max [MA]", b_Ip[1], 0.1)
        fG_lo = _num("fG min [-]", b_fG[0], 0.05)
        fG_hi = _num("fG max [-]", b_fG[1], 0.05)

        bounds = {
            "R0_m": (float(R0_lo), float(R0_hi)),
            "Bt_T": (float(Bt_lo), float(Bt_hi)),
            "Ip_MA": (float(Ip_lo), float(Ip_hi)),
            "fG": (float(fG_lo), float(fG_hi)),
        }

    with st.expander("Objectives (Pareto)", expanded=True):
        st.caption("Choose objectives and direction. At least 2 objectives recommended.")
        obj_defaults = {"R0_m":"min", "B_peak_T":"min", "P_e_net_MW":"max"}
        objectives = {}
        c1,c2,c3 = st.columns(3)
        with c1:
            objectives["R0_m"] = st.selectbox("Objective 1", ["min","max"], index=0)
        with c2:
            objectives["B_peak_T"] = st.selectbox("Objective 2", ["min","max"], index=0)
        with c3:
            objectives["P_e_net_MW"] = st.selectbox("Objective 3", ["max","min"], index=0)

        n_samples = int(st.slider("Samples", min_value=50, max_value=2000, value=100, step=50))
        seed = int(st.number_input("Random seed", value=1, step=1))

    if st.button("Run Pareto study", type="primary", use_container_width=True):
        import time
        t0=time.time()
        try:
            from solvers.optimize import pareto_optimize
            res = pareto_optimize(base0, bounds=bounds, objectives=objectives, n_samples=n_samples, seed=seed)
            feasible = res.get("feasible", [])
            front = res.get("pareto", [])
            st.success(f"Done. Feasible points: {len(feasible)} / {n_samples}. Pareto front: {len(front)}. ({time.time()-t0:.1f}s)")
            if feasible:
                dfF = pd.DataFrame(feasible)
                st.dataframe(dfF, use_container_width=True, height=280)
            if front:
                st.markdown("### Pareto front")
                dfP = pd.DataFrame(front)
                st.dataframe(dfP, use_container_width=True, height=280)

                # Simple scatter plot (choose x,y)
                xkey = st.selectbox("x-axis", options=list(dfP.columns), index=list(dfP.columns).index("R0_m") if "R0_m" in dfP.columns else 0)
                ykey = st.selectbox("y-axis", options=list(dfP.columns), index=list(dfP.columns).index("P_e_net_MW") if "P_e_net_MW" in dfP.columns else 1)
                st.scatter_chart(dfP[[xkey,ykey]].dropna(), x=xkey, y=ykey)

                st.download_button(
                    "Download Pareto front (CSV)",
                    data=dfP.to_csv(index=False),
                    file_name="shams_pareto_front.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
        except Exception as e:
            st.error(f"Pareto study error: {e}")

    # -----------------------------
    # Big leap: Global optimization + robustness (Phase 7+ capability)
    # -----------------------------
    with st.expander("Global optimization (Differential Evolution)", expanded=False):
        st.markdown(
            "Search within bounds for a **single best design** under hard constraints using a "
            "Windows-native **Differential Evolution** global optimizer (no SciPy). "
            "Infeasible points are assigned +∞ objective."
        )

        # Objective selection
        obj_key = st.selectbox(
            "Objective metric",
            options=[
                "R0_m (min)",
                "B_peak_T (min)",
                "q_div_MW_m2 (min)",
                "COE_proxy_USD_per_MWh (min)",
                "CAPEX_proxy_MUSD (min)",
                "-P_e_net_MW (min = maximize Pnet)",
                "-Q (min = maximize Q)",
            ],
            index=3,
        )
        c1, c2, c3 = st.columns(3)
        with c1:
            n_pop = int(st.number_input("Population (n_pop)", min_value=8, max_value=200, value=12, step=2))
        with c2:
            n_gen = int(st.number_input("Generations (n_gen)", min_value=5, max_value=400, value=10, step=5))
        with c3:
            de_seed = int(st.number_input("Optimizer seed", value=1, step=1))

        def _objective_from_out(out: Dict[str, Any]) -> float:
            # Keep objective definitions explicit and scientific.
            if obj_key.startswith("R0_m"):
                return float(out.get("R0_m", float("inf")))
            if obj_key.startswith("B_peak_T"):
                return float(out.get("B_peak_T", float("inf")))
            if obj_key.startswith("q_div"):
                return float(out.get("q_div_MW_m2", float("inf")))
            if obj_key.startswith("COE_proxy"):
                return float(out.get("COE_proxy_USD_per_MWh", float("inf")))
            if obj_key.startswith("CAPEX_proxy"):
                return float(out.get("CAPEX_proxy_MUSD", float("inf")))
            if obj_key.startswith("-P_e_net"):
                return -float(out.get("P_e_net_MW", -float("inf")))
            if obj_key.startswith("-Q"):
                return -float(out.get("Q", -float("inf")))
            return float("inf")

        if st.button("Run global optimization", type="primary", use_container_width=True, key="run_de"):
            try:
                from solvers.optimize import differential_evolution_optimize
                _warn_unrealistic_point_inputs(base0, context="Global optimization")
                base_for_opt = base0
                if obj_key.startswith("COE_proxy") or obj_key.startswith("CAPEX_proxy"):
                    base_for_opt = replace(base_for_opt, include_economics=True)
                t0 = time.time()
                res = differential_evolution_optimize(
                    base=base0,
                    bounds=bounds,
                    objective=_objective_from_out,
                    n_pop=n_pop,
                    n_gen=n_gen,
                    seed=de_seed,
                )
                st.session_state.de_best_inputs = res.get("best_inputs")
                st.session_state.de_best_out = res.get("best_out")
                st.session_state.de_history = res.get("history")
                st.success(f"Optimization complete in {time.time()-t0:.1f}s")
            except Exception as e:
                st.error(f"Global optimization error: {e}")

        best_inp = st.session_state.get("de_best_inputs")
        best_out = st.session_state.get("de_best_out")
        hist = st.session_state.get("de_history")
        if best_inp is not None and best_out is not None:
            st.markdown("#### Best design (inputs)")
            st.json(asdict(best_inp), expanded=False)
            st.markdown("#### Best design (key outputs)")
            show_keys = ["Q", "P_e_net_MW", "COE_proxy_USD_per_MWh", "CAPEX_proxy_MUSD", "B_peak_T", "q_div_MW_m2", "min_constraint_margin"]
            st.json({k: best_out.get(k) for k in show_keys if k in best_out}, expanded=False)

            # Provide artifact export
            try:
                art = build_run_artifact(inputs=dict(asdict(best_inp)), outputs=dict(best_out), constraints=evaluate_constraints(best_out), subsystems={"fidelity": st.session_state.get("fidelity_config", {}), "calibration": {"confinement": float(st.session_state.get("calib_confinement",1.0)), "divertor": float(st.session_state.get("calib_divertor",1.0)), "bootstrap": float(st.session_state.get("calib_bootstrap",1.0))}})
                st.download_button(
                    "Download best-point artifact (JSON)",
                    data=json.dumps(art, indent=2, sort_keys=True),
                    file_name="shams_best_point_artifact.json",
                    mime="application/json",
                    use_container_width=True,
                )
            except Exception:
                pass

        if hist:
            dfh = pd.DataFrame(hist)
            st.markdown("#### Objective history")
            st.line_chart(dfh.set_index("gen")["best_obj"])

    with st.expander("Robustness check (Monte Carlo)", expanded=False):
        st.markdown(
            "Estimate **feasibility probability** under input uncertainty using Monte Carlo sampling. "
            "Samples are drawn from Normal(base, σ·base) and clamped to a minimum fraction of base."
        )

        # Choose which point to analyze
        point_choice = st.radio(
            "Base point",
            ["Last evaluated point", "Best from global optimization"],
            index=0,
            horizontal=True,
        )
        base_for_mc = st.session_state.last_point_inp
        if point_choice == "Best from global optimization" and st.session_state.get("de_best_inputs") is not None:
            base_for_mc = st.session_state.get("de_best_inputs")

        if base_for_mc is None:
            st.info("No base point available yet. Evaluate a point or run global optimization first.")
        else:
            st.caption("Select uncertain inputs and set σ (fraction of base) and minimum clamp.")
            mc_cols = st.columns(4)
            perturb = {}
            # Common uncertainty knobs
            for i, k in enumerate(["Bt_T", "Ip_MA", "fG", "Paux_MW"]):
                with mc_cols[i % 4]:
                    use = st.checkbox(f"Uncertain: {k}", value=(k in ("Bt_T", "Ip_MA")), key=f"mc_use_{k}")
                    if use:
                        sigma = float(
                            st.number_input(
                                f"σ/{k}", min_value=0.0, max_value=0.5, value=0.05, step=0.01, key=f"mc_sig_{k}"
                            )
                        )
                        mn = float(
                            st.number_input(
                                f"min/{k}", min_value=0.0, max_value=1.0, value=0.7, step=0.05, key=f"mc_min_{k}"
                            )
                        )
                        perturb[k] = (sigma, mn)

            c1, c2, c3 = st.columns(3)
            with c1:
                mc_n = int(st.number_input("Samples (n)", min_value=50, max_value=5000, value=50, step=50))
            with c2:
                mc_seed = int(st.number_input("MC seed", value=1, step=1))
            with c3:
                st.write("")

            if st.button("Run robustness Monte Carlo", type="primary", use_container_width=True, key="run_mc_robust"):
                try:
                    from solvers.optimize import robust_feasibility_monte_carlo
                    t0 = time.time()
                    res = robust_feasibility_monte_carlo(base=base_for_mc, perturb=perturb, n=mc_n, seed=mc_seed)
                    st.session_state.robustness_result = res
                    st.success(f"Done in {time.time()-t0:.1f}s")
                except Exception as e:
                    st.error(f"Robustness error: {e}")

            res = st.session_state.get("robustness_result")
            if res:
                st.metric("Feasibility probability", f"{100*float(res.get('feasible_prob', 0.0)):.1f}%")
                ms = res.get("margin_stats", {})
                if ms:
                    dfm = pd.DataFrame(ms).T.reset_index().rename(columns={"index": "metric"})
                    st.dataframe(dfm, use_container_width=True, height=240)
                st.download_button(
                    "Download robustness result (JSON)",
                    data=json.dumps(res, indent=2, sort_keys=True),
                    file_name="shams_robustness_mc.json",
                    mime="application/json",
                    use_container_width=True,
                )

with tab_compare:
    st.header("Compare artifacts")
    st.write("Upload two `shams_run_artifact.json` files to compare key outputs and constraints.")
    colA, colB = st.columns(2)
    with colA:
        upA = st.file_uploader("Artifact A (JSON)", type=["json"], key="cmpA")
    with colB:
        upB = st.file_uploader("Artifact B (JSON)", type=["json"], key="cmpB")

    def _load_art(uploaded):
        if uploaded is None:
            return None
        try:
            return json.loads(uploaded.getvalue().decode("utf-8"))
        except Exception:
            try:
                return json.loads(uploaded.getvalue())
            except Exception:
                return None

    artA = _load_art(upA)
    artB = _load_art(upB)

    if artA and artB:
        outA = artA.get("outputs", {}) or {}
        outB = artB.get("outputs", {}) or {}
        keys = ["Q", "Pfus_total_MW", "P_e_net_MW", "betaN", "q95", "Bpeak_TF_T", "q_div_MW_m2", "neutron_wall_load_MW_m2", "COE_proxy_USD_per_MWh"]
        rows = []
        for k in keys:
            a = outA.get(k, float("nan"))
            b = outB.get(k, float("nan"))
            try:
                da = float(a); db = float(b)
                d = db - da
            except Exception:
                d = ""
            rows.append({"metric": k, "A": a, "B": b, "B-A": d})
        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True)

        consA = pd.DataFrame(artA.get("constraints", []) or [])
        consB = pd.DataFrame(artB.get("constraints", []) or [])
        st.markdown("### Constraints (worst margins first)")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Artifact A**")
            if len(consA):
                st.dataframe(consA.sort_values("residual", ascending=False).head(20), use_container_width=True)
        with c2:
            st.markdown("**Artifact B**")
            if len(consB):
                st.dataframe(consB.sort_values("residual", ascending=False).head(20), use_container_width=True)

        # simple markdown diff export
        diff_md = ["# SHAMS Artifact Comparison", "", "## Key metrics", ""]
        diff_md.append(df.to_markdown(index=False))
        diff_md.append("")
        diff_md.append("## Worst constraints (A)")
        diff_md.append("")
        if len(consA):
            diff_md.append(consA.sort_values("residual", ascending=False).head(20).to_markdown(index=False))
        diff_md.append("")
        diff_md.append("## Worst constraints (B)")
        diff_md.append("")
        if len(consB):
            diff_md.append(consB.sort_values("residual", ascending=False).head(20).to_markdown(index=False))
        st.download_button("Download comparison (markdown)", data="\n".join(diff_md), file_name="artifact_comparison.md", mime="text/markdown", use_container_width=True)

with tab_studies:
    st.header("Studies manager")
    st.write("Save, load, and organize study configurations (scan/pareto) as JSON. This keeps studies reproducible across sessions.")
    if "studies" not in st.session_state:
        st.session_state.studies = []

    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("Save current PointInputs as study", use_container_width=True):
            if st.session_state.get("last_point_inp") is not None:
                try:
                    inp_obj = st.session_state.last_point_inp
                    # dataclass -> dict
                    d = {k: getattr(inp_obj, k) for k in inp_obj.__dataclass_fields__.keys()}  # type: ignore
                    st.session_state.studies.append({"type": "point", "created": datetime.datetime.now().isoformat(), "inputs": d})
                    st.success("Saved.")
                except Exception as e:
                    st.error(f"Could not save: {e}")
            else:
                st.warning("Run a point first so `last_point_inp` exists.")

    with c2:
        up = st.file_uploader("Import studies JSON", type=["json"], key="studies_import")
        if up is not None:
            try:
                imported = json.loads(up.getvalue().decode("utf-8"))
                if isinstance(imported, list):
                    st.session_state.studies.extend(imported)
                elif isinstance(imported, dict):
                    st.session_state.studies.append(imported)
                st.success("Imported.")
            except Exception as e:
                st.error(f"Import failed: {e}")

    with c3:
        if st.session_state.studies:
            st.download_button(
                "Download studies JSON",
                data=json.dumps(st.session_state.studies, indent=2, sort_keys=True),
                file_name="shams_studies.json",
                mime="application/json",
                use_container_width=True,
            )

    st.markdown("### Saved studies")
    if st.session_state.studies:
        df = pd.DataFrame([{"i": i, "type": s.get("type","?"), "created": s.get("created",""), "notes": s.get("notes","")} for i,s in enumerate(st.session_state.studies)])
        st.dataframe(df, use_container_width=True)
        idx = st.number_input("Select index to view", min_value=0, max_value=max(0, len(st.session_state.studies)-1), value=0, step=1)
        st.json(st.session_state.studies[int(idx)])
        if st.button("Delete selected", use_container_width=True):
            try:
                st.session_state.studies.pop(int(idx))
                st.experimental_rerun()
            except Exception:
                pass
    else:
        st.info("No studies saved yet.")


st.subheader("Operating envelope check (multi-point)")
st.caption("Evaluates startup / nominal / end-of-life proxy points and reports the worst constraint.")
colA, colB = st.columns([1,3])
with colA:
    run_env = st.button("Run envelope check", use_container_width=True)
if run_env:
    try:
        from envelope.points import default_envelope_points
        from physics.hot_ion import hot_ion_point
        from constraints.system import build_constraints_from_outputs, summarize_constraints
        base_inp = st.session_state.get("last_point_inp", None)
        if base_inp is None:
            st.warning("No current point inputs available.")
        else:
            _warn_unrealistic_point_inputs(base_inp, context="Envelope check")
            pts = default_envelope_points(base_inp)
            env_rows = []
            worst = None
            for i, p in enumerate(pts):
                out = hot_ion_point(p)
                cs = build_constraints_from_outputs(out)
                summ = summarize_constraints(cs)
                dom = summ.get("dominant", {})
                row = {
                    "point": i,
                    "all_ok": bool(summ.get("all_ok", False)),
                    "dominant": dom.get("name", ""),
                    "residual": dom.get("residual", float("nan")),
                    "margin": dom.get("margin", float("nan")),
                }
                env_rows.append(row)
                if worst is None or (row["residual"] == row["residual"] and row["residual"] > worst["residual"]):
                    worst = row
            st.dataframe(env_rows, use_container_width=True)
            if worst:
                st.info(f"Worst point: #{worst['point']} — {worst['dominant']} (residual={worst['residual']:.3g})")
    except Exception as e:
        st.error(f"Envelope check failed: {e}")


with tab_model:
    st.header("0-D Tokamak Physics Model (Phase‑1)")

    with st.expander("0‑D Physical Models — explanations", expanded=True):
        _pm = os.path.join(ROOT, "docs", "PHYSICAL_MODELS_0D.md")
        try:
            with open(_pm, "r", encoding="utf-8") as _f:
                st.markdown(_f.read())
        except Exception as _e:
            st.error(f"Failed to load physical model doc: {_e}")


    st.markdown(r"""
    This tab is written to be **actionable**: each section maps to code in `src/physics/`, `src/phase1_models.py`,
    `src/phase1_systems.py`, and `src/solvers/`.

    SHAMS remains a **0‑D / volume‑averaged / steady‑state** point-design model at its core, intended for *fast feasibility scanning*.
    Over time we have added several **PROCESS‑inspired** upgrades that remain lightweight and Windows‑friendly:

    - **Optional analytic profiles ("½‑D")** for $n_e(\rho)$, $T_i(\rho)$, $T_e(\rho)$ with **normalization to the chosen volume averages**,
      plus derived averages like peaking factors and $\langle n_e^2 \rangle/\langle n_e \rangle^2$.
    - **Radiation options:** legacy fractional radiation (stable for scans) and a physics‑based path (brem + synchrotron + simple impurity line radiation).
    - **Constraint system:** engineering and plasma constraints are represented as reusable objects (PROCESS‑like), usable by scans and vector solvers.
    - **Solvers:** classic nested 1‑D solves are still available, plus a more general bounded "targets → variables" solve primitive.

    It is **not** a full transport / equilibrium / neutronics code, but it is designed to grow in that direction while staying usable.
    """)

    st.caption("Tip: expand only the models you care about — each block is independent.")

    # --- Geometry ---
    with st.expander("Geometry: volume and surface area (implemented)", expanded=True):
        st.markdown(r"""
        Implemented helpers:

        **Plasma volume** (`tokamak_volume`)
        $$
        V \approx 2\pi^2\,R\,a^2\,\kappa
        $$

        **Plasma surface area** (`tokamak_surface_area`)
        $$
        S \approx 4\pi^2\,R\,a\,\kappa
        $$

        Notes:
        - These are **engineering approximations** intended to preserve correct monotonic trends.
        - Units: $R,a$ in m, $V$ in m$^3$, $S$ in m$^2$.
        """)

    # --- Confinement ---
    with st.expander("Energy confinement: IPB98(y,2) (implemented)"):
        st.markdown(r"""
        Implemented model: `tauE_ipb98y2`.

        $$
        \tau_E = 0.0562\, I_p^{0.93} B_t^{0.15} \bar{n}^{0.41} P_{loss}^{-0.69} R^{1.97} \epsilon^{0.58} \kappa^{0.78} M^{0.19}
        $$
        where $\epsilon=a/R$.

        **Units (must match the implementation):**
        - $I_p$ in MA
        - $B_t$ in T
        - $\bar{n}$ in units of $10^{20}\,\mathrm{m^{-3}}$ (i.e. `ne20`)
        - $P_{loss}$ in MW
        - $R,a$ in m
        - $M$ in amu (default 2.5)

        Output: $\tau_E$ in seconds.
        """)

    # --- L-H threshold ---
    with st.expander("H-mode access: Martin-2008 L–H threshold (implemented)"):
        st.markdown(r"""
        Implemented model: `p_LH_martin08`.

        $$
        P_{LH} = 0.0488\, \bar{n}^{0.717} B_t^{0.803} S^{0.941}\,\left(\frac{2}{A_{eff}}\right)
        $$

        **Units:**
        - $\bar{n}$ in $10^{20}\,\mathrm{m^{-3}}$ (line-averaged)
        - $B_t$ in T
        - $S$ in m$^2$ (uses the same proxy as the geometry block)
        - $A_{eff}$ dimensionless (defaults to 2.0)

        Output: $P_{LH}$ in MW.
        """)

    # --- Greenwald ---
    with st.expander("Density limit: Greenwald (implemented)"):
        st.markdown(r"""
        Implemented helper: `greenwald_density_20`.

        $$
        n_{GW}\,[10^{20}\,\mathrm{m^{-3}}] = \frac{I_p\,[\mathrm{MA}]}{\pi a^2\,[\mathrm{m^2}]}
        $$

        In scans, an operating fraction is typically applied:
        $$
        \bar{n} = f_{nG}\,n_{GW},\qquad 0 < f_{nG} \le 1.
        $$
        """)

    # --- Screening proxies ---
    with st.expander("Screening proxies: q95, βN, bootstrap fraction (implemented proxies)"):
        st.markdown(r"""
        These are explicitly labeled **proxies** (trend-correct, not equilibrium/transport solutions).

        **q95 proxy** (`q95_proxy_cyl`)
        $$
        q_{95} \approx \left(\frac{2\pi R B_t}{\mu_0 I_p}\right)\left(\frac{a}{R}\right)\frac{1}{\kappa}
        $$
        with $I_p$ converted to amperes internally.

        **Normalized beta** (`betaN_from_beta`)
        $$
        \beta_N = \beta(\%)\,\frac{a\,B_t}{I_p}
        \qquad\text{with}\qquad \beta(\%)=100\,\beta
        $$
        where $\beta$ is the *fractional* beta.

        **Bootstrap fraction proxy** (`bootstrap_fraction_proxy`)
        $$
        f_{bs} \approx C_{bs}\,\frac{\beta_N}{q_{95}}
        $$
        then clamped to a configured range (default 0 to 0.95).
        """)

    # --- Fusion reactivity ---
    with st.expander("Fusion reactivity: Bosch–Hale ⟨σv⟩ (implemented)"):
        st.markdown(r"""
        Implemented function: `bosch_hale_sigmav(T_i, reaction)`.

        This uses the Bosch–Hale parameterization for Maxwellian-averaged reactivity:
        $$
        \langle\sigma v\rangle(T_i)\;[\mathrm{m^3/s}]
        $$

        Internally, the implementation computes intermediate variables ($\theta$, $\xi$) from a
        reaction-specific coefficient set and returns a strictly non-negative value.

        **Important for UI users:**
        - Input $T_i$ is in **keV**.
        - Output is in **m$^3$/s**.
        """)

        # Bosch–Hale coefficient values used by the implementation (from `BH_COEFFS`)
        _bh_rows = []
        for _rxn in ["DT", "DD_Tp", "DD_He3n"]:
            _c = BH_COEFFS[_rxn]
            _bh_rows.append({"Reaction": _rxn, **asdict(_c)})
        _bh_df = pd.DataFrame(_bh_rows).set_index("Reaction")
        st.caption("Bosch–Hale coefficients used for DT and DD channels (exact values as implemented).")
        st.dataframe(_bh_df, use_container_width=True)

    # --- Fusion power / gain symbols (fixing the screenshot issue) ---
    with st.expander("Fusion power & gain definitions: P_f, P_α, Q (notation)"):
        st.markdown(r"""
        **What these symbols mean (and how they relate):**

        **Fusion power, $P_f$**  
        Total thermal power released by fusion reactions occurring in the plasma:
        $$
        P_f \;=\; \dot{N}_{\text{fus}}\,E_{\text{fus}}
        $$
        where $\dot{N}_{\text{fus}}$ is the fusion reaction rate [1/s] and $E_{\text{fus}}$ is the energy released per reaction.
        For D‑T, $E_{\text{fus}} = 17.6\,\mathrm{MeV}$.

        **Alpha heating power, $P_\alpha$**  
        Part of $P_f$ carried by *charged* alpha particles and deposited back into the plasma (self‑heating):
        $$
        P_\alpha \;=\; f_\alpha\,P_f
        $$
        For D‑T, $f_\alpha = \frac{3.5}{17.6} \approx 0.199$, so $P_\alpha \approx 0.20\,P_f$.
        (The rest is mainly neutron power: $P_n \approx 0.80\,P_f$.)

        **Fusion gain, $Q$**  
        In this UI, $Q$ is the standard *plasma gain*:
        $$
        Q \;=\; \frac{P_f}{P_{\mathrm{aux}}}
        $$
        where $P_{\mathrm{aux}}$ is the **externally applied** auxiliary heating power (e.g., NBI/RF) required to sustain the operating point.
        This is distinct from “wall‑plug” gain, which would include plant efficiencies and non‑plasma power draws.

        **How to interpret in scans**
        - Increasing $P_f$ increases $P_\alpha$ proportionally (more self‑heating).  
        - $Q$ improves only when $P_f$ grows faster than the required $P_{\mathrm{aux}}$.
        """)

    # --- SOL width metric ---
    with st.expander("Optional divertor/SOL risk metric: Eich λq (implemented)"):
        st.markdown(r"""
        Implemented metric: `lambda_q_eich14_mm`.

        $$
        \lambda_q\,[\mathrm{mm}] \approx \text{factor}\times 0.63\,B_{pol}^{-1.19}
        $$

        with $B_{pol}$ approximated by:
        $$
        B_{pol} \approx \frac{\mu_0 I_p}{2\pi a}
        $$

        This is **not** a self‑consistent divertor / edge power‑exhaust model — it’s a compact, order‑of‑magnitude **screening proxy** for quickly comparing design points.
        """)

    st.info(
        "If you want the *full* step-by-step closure shown here (power balance → temperatures → Pf/Q), "
        "tell me which exact function in `src/phase1_core.py` you want treated as the single source of truth, "
        "and I’ll mirror it line-for-line in this tab."
    )

# -----------------------------
# Benchmarks
# -----------------------------
with tab_bench:
    st.subheader("Regression Benchmarks")
    st.write("Run a small suite of SPARC-like cases to ensure recent physics/solver changes haven't broken behavior.")

    import json
    from pathlib import Path

    bench_dir = Path(__file__).resolve().parent.parent / "benchmarks"
    cases_path = bench_dir / "cases.json"
    golden_path = bench_dir / "golden.json"

    diff_path = bench_dir / "last_diff_report.json"
    with st.expander("Latest diff report (from last run)", expanded=False):
        if diff_path.exists():
            try:
                rep = json.loads(diff_path.read_text(encoding="utf-8"))
                st.caption(f"Generated at unix={rep.get('created_unix'):.0f} | failures={rep.get('n_failed',0)}")
                rows = rep.get("rows", [])
                if rows:
                    import pandas as pd

                    df_rep = pd.DataFrame(rows)
                    # show worst first
                    if "ok" in df_rep.columns and "rel_err" in df_rep.columns:
                        df_rep = df_rep.sort_values(by=["ok","rel_err"], ascending=[True, False])
                    st.dataframe(df_rep, use_container_width=True, height=260)
                # Structural diffs (constraints/model cards) vs golden artifacts, if present
                ss = rep.get("structural_summary")
                if ss:
                    st.markdown("**Structural diffs vs golden artifacts**")
                    st.write({k: ss.get(k) for k in ["n_cases","n_with_changes","total_added_constraints","total_removed_constraints","total_changed_constraints","total_modelcard_changes"]})
                structural = rep.get("structural") or {}
                if structural:
                    with st.expander("Show structural diffs by case", expanded=False):
                        for cname, d in structural.items():
                            cadd = d.get("constraints", {}).get("added", [])
                            crem = d.get("constraints", {}).get("removed", [])
                            cchg = d.get("constraints", {}).get("changed_meta", [])
                            mc = d.get("model_cards", {})
                            mcchg = (mc.get("added", []) or []) + (mc.get("removed", []) or []) + (mc.get("changed", []) or [])
                            if not (cadd or crem or cchg or mcchg or (d.get("schema_version", {}).get("new") != d.get("schema_version", {}).get("old"))):
                                continue
                            with st.expander(f"{cname}", expanded=False):
                                if cadd: st.write({"constraints_added": cadd})
                                if crem: st.write({"constraints_removed": crem})
                                if cchg: st.write({"constraint_meta_changes": cchg})
                                if mc.get("added"): st.write({"model_cards_added": mc.get("added")})
                                if mc.get("removed"): st.write({"model_cards_removed": mc.get("removed")})
                                if mc.get("changed"): st.write({"model_cards_changed": mc.get("changed")})

                st.download_button("Download diff report JSON", data=diff_path.read_bytes(), file_name="last_diff_report.json")
            except Exception as e:
                st.warning(f"Could not read diff report: {e}")
        else:
            st.info("No diff report yet. Run benchmarks to generate one.")


    # Release notes (auto-generated)
    with st.expander("Release notes (auto)", expanded=False):
        import subprocess, sys
        from pathlib import Path

        repo_root = Path(__file__).resolve().parent.parent
        out_md = repo_root / "RELEASE_NOTES.md"
        old_default = str((repo_root.parent / "SHAMS_old").resolve()) if (repo_root.parent / "SHAMS_old").exists() else r"..\SHAMS_old"
        old_path = st.text_input("Old SHAMS repo path", value=st.session_state.get("release_notes_old", old_default))
        st.session_state["release_notes_old"] = old_path

        auto = st.checkbox("Auto-generate if missing/out-of-date", value=True, key="release_notes_auto")
        run_now_rn = st.button("Generate release notes now", key="btn_release_notes_now")

        def _needs_rn() -> bool:
            if not out_md.exists():
                return True
            try:
                m_out = out_md.stat().st_mtime
                # regenerate if diff report is newer, or tool changed
                tool_p = repo_root / "tools" / "release_notes.py"
                diff_p = repo_root / "benchmarks" / "last_diff_report.json"
                newest = max([p.stat().st_mtime for p in [tool_p, diff_p] if p.exists()] + [0])
                return newest > m_out
            except Exception:
                return False

        if (auto and _needs_rn() and not st.session_state.get("_rn_ran_this_session", False)) or run_now_rn:
            cmd = [sys.executable, str(repo_root / "tools" / "release_notes.py"), "--old", old_path, "--new", str(repo_root), "--out", str(out_md)]
            st.caption("Running: " + " ".join(cmd))
            try:
                p = subprocess.run(cmd, capture_output=True, text=True, cwd=str(repo_root))
                st.session_state["_rn_last_stdout"] = p.stdout
                st.session_state["_rn_last_stderr"] = p.stderr
                st.session_state["_rn_last_rc"] = p.returncode
                st.session_state["_rn_ran_this_session"] = True
            except Exception as e:
                st.session_state["_rn_last_stderr"] = str(e)
                st.session_state["_rn_last_rc"] = 1

        rc = st.session_state.get("_rn_last_rc")
        if rc is not None:
            if rc == 0:
                st.success("Release notes generated.")
            else:
                st.warning("Release notes generation had issues (see logs).")
            with st.expander("Logs", expanded=False):
                st.code((st.session_state.get("_rn_last_stdout") or "") + "\n" + (st.session_state.get("_rn_last_stderr") or ""))

        if out_md.exists():
            st.markdown(out_md.read_text(encoding="utf-8", errors="ignore"))
            st.download_button("Download RELEASE_NOTES.md", data=out_md.read_bytes(), file_name="RELEASE_NOTES.md", mime="text/markdown")
        else:
            st.info("RELEASE_NOTES.md not found yet.")

    with st.expander("Regression comparisons", expanded=True):
        colA, colB = st.columns([1,1])
        with colA:
            run_now = st.button("Run benchmarks")
        with colB:
            regen = st.button("Regenerate golden (intentional changes)")
    
        def _safe(v):
            try:
                return float(v)
            except Exception:
                return float("nan")
    
        if cases_path.exists():
            _cases_raw = json.loads(cases_path.read_text())
        else:
            _cases_raw = {}
    
        # Normalize benchmark cases into a list[dict] with keys: name, inputs
        # Supports dict-form (name -> inputs), list-form (dicts), or list-form (names).
        cases = []
        if isinstance(_cases_raw, dict):
            for _name, _inp in _cases_raw.items():
                if isinstance(_inp, dict):
                    cases.append({"name": str(_name), "inputs": _inp})
        elif isinstance(_cases_raw, list):
            for i, _c in enumerate(_cases_raw):
                if isinstance(_c, dict):
                    _name = _c.get("name", f"case_{i}")
                    _inp = _c.get("inputs", _c.get("inp", _c.get("input", {})))
                    if isinstance(_inp, dict):
                        cases.append({"name": str(_name), "inputs": _inp})
                else:
                    cases.append({"name": str(_c), "inputs": {}})
    
        if not cases:
            # Always provide at least one default case so the UI doesn't crash
            cases = [{"name": "default", "inputs": {"R0_m": 1.85, "a_m": 0.6, "kappa": 1.75, "Bt_T": 12.0, "Ip_MA": 8.0, "Ti_keV": 10.0, "fG": 0.85, "t_shield_m": 0.25, "Paux_MW": 25.0}}]
    
        base = PointInputs(R0_m=1.85, a_m=0.6, kappa=1.75, Bt_T=12.0, Ip_MA=8.0, Ti_keV=10.0, fG=0.85, t_shield_m=0.25, Paux_MW=25.0)
        if run_now or regen:
            results = {}
            for _case in cases:
                name = _case.get("name","case")
                overrides = _case.get("inputs", {})
                # Defensive: apply only existing fields
                d = base.__dict__.copy()
                for k, v in overrides.items():
                    if k in d:
                        d[k] = v
                inp_case = PointInputs(**d)
                results[name] = hot_ion_point(inp_case)
    
            if regen:
                golden_path.write_text(json.dumps(results, indent=2))
                st.success(f"Wrote golden: {golden_path}")
            else:
                if not golden_path.exists():
                    st.error("golden.json not found. Click 'Regenerate golden' once to create it.")
                else:
                    golden = json.loads(golden_path.read_text())
                    CURATED = ["Q_DT_eqv","H98","P_fus_MW","P_SOL_MW","q_div_MW_m2","B_peak_T","sigma_hoop_MPa","hts_margin_cs","J_eng_A_mm2","t_flat_s","P_net_MW"]
                    rows = []
                    failed = 0
                    for name, cur in results.items():
                        ref = golden.get(name, {})
                        for k in CURATED:
                            a = _safe(cur.get(k))
                            b = _safe(ref.get(k))
                            if not (math.isfinite(a) and math.isfinite(b)):
                                continue
                            atol = 1e-6
                            rtol = 5e-3
                            ok = abs(a-b) <= max(atol, rtol*max(abs(b),1e-9))
                            if not ok:
                                failed += 1
                            rows.append({"case":name,"key":k,"got":a,"golden":b,"rel_err":(abs(a-b)/max(abs(b),1e-9)),"ok":ok})
                    import pandas as pd

                    df = pd.DataFrame(rows)
                    st.dataframe(df, use_container_width=True)

                    # Write a machine-readable diff report (used by CI and the UI)
                    try:
                        import time as _time
                        report = {
                            "created_unix": _time.time(),
                            "rtol": 5e-3,
                            "atol": 1e-6,
                            "n_rows": int(len(rows)),
                            "n_failed": int(failed),
                            "rows": rows,
                        }
                        (bench_dir / "last_diff_report.json").write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")
                    except Exception:
                        pass

                    if failed==0:
                        st.success("All benchmark comparisons passed (within tolerances).")
                    else:
                        st.warning(f"{failed} comparisons exceeded tolerance. See table.")
    
    
    st.divider()
    with st.expander("Sensitivity and uncertainty (Monte Carlo)", expanded=False):
        st.subheader("Sensitivity (Monte Carlo)")
        st.write("Runs a lightweight uncertainty scan around a selected benchmark case (Windows-native).")
    
        from analysis.sensitivity import monte_carlo_feasibility
        from models.inputs import PointInputs
    
        case_names = [c.get("name", f"case_{i}") for i,c in enumerate(cases)]
        case_sel = st.selectbox("Benchmark case for sensitivity", case_names, index=0, key="sens_case_sel")
        n_mc = st.number_input("Samples", min_value=50, max_value=2000, value=50, step=50, key="sens_n")
        if st.button("Run Monte Carlo", key="run_mc_bench"):
            c = cases[case_names.index(case_sel)]
            base_inp = PointInputs(**c["inputs"])
            res = monte_carlo_feasibility(base_inp, n=int(n_mc), seed=42)
            st.metric("Feasible probability", f"{res['p_feasible']*100:.1f}%")
            st.write("Most frequently violated constraints:")
            st.dataframe(res["worst_constraints"], use_container_width=True)
    
    st.divider()
    with st.expander("Pareto search (design studies)", expanded=False):
        st.subheader("Pareto Search (LHS)")
        st.write("Finds a feasible Pareto set for a small set of design knobs around a benchmark case.")
        from solvers.optimize import pareto_optimize
    
        case_sel2 = st.selectbox("Benchmark case for Pareto", case_names, index=0, key="pareto_case_sel")
        n_lhs = st.number_input("LHS samples", min_value=50, max_value=5000, value=100, step=50, key="pareto_n")
        # simple bounds
        colp1, colp2 = st.columns(2)
        with colp1:
            R0_lo = st.number_input("R0 min [m]", value=1.5, step=0.1, key="R0_lo")
            Ip_lo = st.number_input("Ip min [MA]", value=5.0, step=0.5, key="Ip_lo")
        with colp2:
            R0_hi = st.number_input("R0 max [m]", value=2.5, step=0.1, key="R0_hi")
            Ip_hi = st.number_input("Ip max [MA]", value=12.0, step=0.5, key="Ip_hi")
        fG_lo = st.number_input("fG min", value=0.4, step=0.05, key="fG_lo")
        fG_hi = st.number_input("fG max", value=1.2, step=0.05, key="fG_hi")
    
        if st.button("Run Pareto search", key="run_pareto"):
            c = cases[case_names.index(case_sel2)]
            base_inp = PointInputs(**c["inputs"])
            bounds = {"R0_m": (float(R0_lo), float(R0_hi)), "Ip_MA": (float(Ip_lo), float(Ip_hi)), "fG": (float(fG_lo), float(fG_hi))}
            objectives = {"R0_m": "min", "B_peak_T": "min", "P_e_net_MW": "max"}
            res = pareto_optimize(base_inp, bounds=bounds, objectives=objectives, n_samples=int(n_lhs), seed=1)
            st.write(f"Feasible points: {len(res['feasible'])}  |  Pareto points: {len(res['pareto'])}")
            st.dataframe(res["pareto"], use_container_width=True)
    
    # -----------------------------
    # Variable Registry (auditable meanings/units/sources)
    # -----------------------------
with tab_registry:
    st.subheader("Variable Registry")
    st.markdown(
        "A PROCESS-style registry of key SHAMS variables with units, meaning, and model provenance. "
        "Use this to keep the code **auditable** as physics/engineering fidelity increases."
    )
    q = st.text_input("Search variables", value="", placeholder="e.g., H98, q_div, HTS, TBR")
    try:
        df = registry_dataframe(q)
        st.dataframe(df, use_container_width=True, height=520)
        st.download_button(
            "Download registry (CSV)",
            data=df.to_csv(index=False),
            file_name="shams_variable_registry.csv",
            mime="text/csv",
            use_container_width=True,
        )
    except Exception as e:
        st.error(f"Registry unavailable: {e}")

# -----------------------------
# Validation (envelopes)
# -----------------------------
with tab_validation:
    st.subheader("Validation envelopes")
    st.markdown(
        "Decision-grade validation in SHAMS is **envelope-based**: we check whether a solution lies within "
        "a broad reference band for key metrics, rather than trying to match a single reference point. "
        "This is robust to proxy changes and is aligned with PROCESS-style workflows."
    )
    try:
        from validation.envelopes import default_envelopes
        envs = default_envelopes()
        env_name = st.selectbox("Select envelope", list(envs.keys()), index=0, key="validation_env_sel")
        env = envs[env_name]
        st.caption(env.notes)

        out = st.session_state.get("last_point_out")
        if not out:
            st.info("Run a Point Designer solve first. The latest outputs will be checked here.")
        else:
            report = env.check(out)
            import pandas as pd

            rows = []
            n_fail = 0
            for k, r in report.items():
                if not r.get("ok"):
                    n_fail += 1
                rows.append({
                    "metric": k,
                    "value": r.get("value"),
                    "lo": r.get("lo"),
                    "hi": r.get("hi"),
                    "ok": bool(r.get("ok")),
                })
            df = pd.DataFrame(rows)
            st.dataframe(df, use_container_width=True, height=360)
            if n_fail == 0:
                st.success("All selected envelope checks passed.")
            else:
                st.warning(f"{n_fail} envelope checks failed. This indicates the *targets/bounds* are outside the reference band (not a code error).")
    except Exception as e:
        st.error(f"Validation module unavailable: {e}")


# -----------------------------
# Compliance (requirements + model cards)
# -----------------------------
with tab_compliance:
    st.subheader("Verification & Compliance")
    st.caption("Shows the latest verification/compliance matrix from verification/report.json (if present).")

    def _load_verification_report_ui():
        try:
            here = Path(__file__).resolve()
            root = here.parent.parent  # ui/ -> repo root
            rp = root / "verification" / "report.json"
            if rp.exists():
                return json.loads(rp.read_text(encoding="utf-8"))
        except Exception:
            return None
        return None

    report = _load_verification_report_ui()
    if not report:
        st.info("No verification/report.json found. Run: `python verification/run_verification.py` to generate it.")
    else:
        meta = report.get("meta", {})
        st.write({
            "generated_unix": meta.get("generated_unix"),
            "python": meta.get("python"),
            "platform": meta.get("platform"),
            "git_commit": meta.get("git_commit"),
        })

        # Summary
        summary = report.get("summary", {})
        cols = st.columns(4)
        cols[0].metric("Requirements", int(summary.get("n_requirements", 0)))
        cols[1].metric("Passed", int(summary.get("n_pass", 0)))
        cols[2].metric("Failed", int(summary.get("n_fail", 0)))
        cols[3].metric("Overall", "PASS" if summary.get("all_pass") else "FAIL")

        # Detailed table
        rows = report.get("results", [])
        if rows:
            df = pd.DataFrame(rows)
            # Friendly columns ordering
            keep = [c for c in ["req_id","title","status","details","linked_model_cards"] if c in df.columns]
            df = df[keep] if keep else df
            st.dataframe(df, use_container_width=True, height=520)

        # Download JSON
        st.download_button(
            "Download verification report.json",
            data=json.dumps(report, indent=2, sort_keys=True),
            file_name="verification_report.json",
            mime="application/json",
        )


with tab_docs:
    st.header("Docs")
    st.caption("Built-in documentation bundled with this repository (no internet required).")

    doc_options = {
        "Upgrade plan (PROCESS-inspired)": os.path.join(ROOT, "docs", "SHAMS_upgrade_plan_from_PROCESS.md"),
        "Lessons learned from PROCESS": os.path.join(ROOT, "docs", "PROCESS_lessons.md"),
        "0‑D Physical Models (Phase‑1)": os.path.join(ROOT, "docs", "PHYSICAL_MODELS_0D.md"),
        "Engineering closures": os.path.join(ROOT, "docs", "ENGINEERING_CLOSURES.md"),
        "Operating envelope (multi-point)": os.path.join(ROOT, "docs", "ENVELOPE.md"),
        "Studies workflows": os.path.join(ROOT, "docs", "STUDIES.md"),
        "Model cards (auditability)": os.path.join(ROOT, "docs", "MODEL_CARDS.md"),
        "Compliance & verification": os.path.join(ROOT, "docs", "COMPLIANCE.md"),
        "Regression & golden benchmarks": os.path.join(ROOT, "docs", "REGRESSION.md"),
        "Release notes generation": os.path.join(ROOT, "docs", "RELEASE_NOTES.md"),
        "UI quickstart": os.path.join(ROOT, "README_UI.md"),
    }

    doc_sel = st.selectbox("Select a document", list(doc_options.keys()), index=0, key="doc_select")
    doc_path = doc_options.get(doc_sel)

    if doc_path and os.path.exists(doc_path):
        try:
            with open(doc_path, "r", encoding="utf-8") as f:
                st.markdown(f.read())
        except Exception as _e:
            st.error(f"Failed to read doc: {_e}")
    else:
        st.warning("Document file not found in this checkout.")


# -----------------------------
# Artifacts Explorer (new)
# -----------------------------
def _load_json_from_upload(uploaded) -> Dict[str, Any] | None:
    if uploaded is None:
        return None
    try:
        raw = uploaded.getvalue()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8")
        return json.loads(raw)
    except Exception:
        return None


def _safe_df(rows: List[Dict[str, Any]]) -> pd.DataFrame:
    try:
        return pd.DataFrame(rows)
    except Exception:
        return pd.DataFrame()


def _as_float(x) -> float | None:
    try:
        if x is None:
            return None
        if isinstance(x, bool):
            return None
        return float(x)
    except Exception:
        return None


def _numeric_delta_table(base_out: Dict[str, Any], scen_out: Dict[str, Any], limit: int = 40) -> pd.DataFrame:
    keys = sorted(set(base_out.keys()) | set(scen_out.keys()))
    rows = []
    for k in keys:
        a = _as_float(base_out.get(k))
        b = _as_float(scen_out.get(k))
        if a is None or b is None:
            continue
        d = b - a
        # skip near-identical
        if abs(d) < 1e-12:
            continue
        rows.append({"metric": k, "baseline": a, "scenario": b, "delta": d, "delta_frac": (d / a) if abs(a) > 1e-12 else None})
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df = df.reindex(df["delta"].abs().sort_values(ascending=False).index)
    return df.head(limit)


with tab_artifacts:
    st.header("Artifacts Explorer")
    st.caption("Load a SHAMS run artifact and inspect new v50+ artifact sections (constraint ledger, model set, standardized tables).")

    up = st.file_uploader("Upload shams_run_artifact.json", type=["json"], key="ae_upload")
    art = _load_json_from_upload(up)

    col_a, col_b = st.columns([1.2, 1.0])
    with col_a:
        alt_path = st.text_input("...or load from local path", value="", key="ae_path")
    with col_b:
        load_btn = st.button("Load from path", key="ae_load_path")

    if load_btn and alt_path:
        try:
            with open(alt_path, "r", encoding="utf-8") as f:
                art = json.load(f)
        except Exception as e:
            st.error(f"Failed to load JSON: {type(e).__name__}: {e}")
            art = None

    if not art:
        st.info("Upload an artifact JSON (or provide a path) to explore.")
    else:
        meta = art.get("meta", {}) or {}
        prov = art.get("provenance", {}) or {}
        st.subheader("Metadata")
        st.write({
            "schema_version": art.get("schema_version"),
            "label": meta.get("label"),
            "mode": meta.get("mode"),
            "git_commit": prov.get("git_commit"),
            "python": prov.get("python"),
            "platform": prov.get("platform"),
            "repo_version": prov.get("repo_version"),
        })

        # --- Constraint ledger ---
        st.subheader("Constraint Margin Ledger")
        ledger = art.get("constraint_ledger") or {}
        if isinstance(ledger, dict) and ledger.get("entries"):
            st.caption(f"schema={ledger.get('schema_version','(missing)')}  fingerprint={ledger.get('ledger_fingerprint_sha256','(missing)')}")
            top = ledger.get("top_blockers") or []
            if top:
                st.markdown("**Top blockers**")
                st.dataframe(_safe_df(top), use_container_width=True)
            with st.expander("All ledger entries"):
                st.dataframe(_safe_df(ledger.get("entries") or []), use_container_width=True)
        else:
            st.info("No constraint_ledger found in this artifact.")

        # --- Model set / registry ---
        st.subheader("Model Set")
        model_set = art.get("model_set") or {}
        model_registry = art.get("model_registry") or {}
        if model_set:
            st.caption(f"schema={model_set.get('schema_version','(missing)')}")
            st.json(model_set)
        else:
            st.info("No model_set embedded in this artifact.")
        with st.expander("Model Registry"):
            if model_registry:
                st.caption(f"schema={model_registry.get('schema_version','(missing)')}")
                st.json(model_registry)
            else:
                st.info("No model_registry embedded in this artifact.")

        # --- Standard tables ---
        st.subheader("Standard Tables")
        tables = art.get("tables") or {}
        if isinstance(tables, dict) and tables:
            for k in ["plasma", "power_balance", "tritium"]:
                if k in tables:
                    st.markdown(f"**{k}**")
                    t = tables.get(k)
                    if isinstance(t, dict):
                        st.dataframe(pd.DataFrame([t]), use_container_width=True)
                    elif isinstance(t, list):
                        st.dataframe(pd.DataFrame(t), use_container_width=True)
                    else:
                        st.json(t)
        else:
            st.info("No tables.v1 section found in this artifact.")

        with st.expander("Full artifact JSON"):
            st.json(art)


# -----------------------------
# Case Deck Runner (new)
# -----------------------------
with tab_deck:
    st.header("Case Deck Runner")
    st.caption("Run a case_deck.v1 YAML/JSON deck and view the resolved config + artifact outputs.")

    up_deck = st.file_uploader("Upload case_deck.yaml / .json", type=["yaml", "yml", "json"], key="deck_upload")
    out_root = os.path.join(ROOT, "ui_runs")
    os.makedirs(out_root, exist_ok=True)
    out_name = st.text_input("Output folder name (under ui_runs/)", value=f"deck_{int(time.time())}", key="deck_out_name")
    run_btn = st.button("Run Case Deck", key="deck_run")

    if run_btn:
        if up_deck is None:
            st.error("Please upload a case deck file first.")
        else:
            try:
                deck_path = os.path.join(out_root, f"_uploaded_{up_deck.name}")
                with open(deck_path, "wb") as f:
                    f.write(up_deck.getvalue())
                out_dir = os.path.join(out_root, out_name)
                os.makedirs(out_dir, exist_ok=True)
                runner = os.path.join(ROOT, "tools", "run_case_deck.py")
                proc = subprocess.run(
                    [sys.executable, runner, deck_path, "--out", out_dir],
                    cwd=ROOT,
                    capture_output=True,
                    text=True,
                    check=False,
                )
                st.code(proc.stdout or "", language="text")
                if proc.returncode != 0:
                    st.error("Case deck run failed.")
                    st.code(proc.stderr or "", language="text")
                else:
                    art_path = os.path.join(out_dir, "shams_run_artifact.json")
                    cfg_path = os.path.join(out_dir, "run_config_resolved.json")
                    st.success(f"Wrote outputs to: {out_dir}")
                    if os.path.exists(cfg_path):
                        st.subheader("Resolved config")
                        with open(cfg_path, "r", encoding="utf-8") as f:
                            st.json(json.load(f))
                    if os.path.exists(art_path):
                        st.subheader("Run artifact (preview)")
                        with open(art_path, "r", encoding="utf-8") as f:
                            st.json(json.load(f))
            except Exception as e:
                st.error(f"{type(e).__name__}: {e}")


# -----------------------------
# Scenario Delta Viewer (new)
# -----------------------------
with tab_delta:
    st.header("Scenario Delta Viewer")
    st.caption("Compare two run artifacts (baseline vs scenario). Uses embedded scenario_delta when available; otherwise computes a transparent diff.")

    col1, col2 = st.columns(2)
    with col1:
        up_base = st.file_uploader("Baseline shams_run_artifact.json", type=["json"], key="delta_base")
    with col2:
        up_scen = st.file_uploader("Scenario shams_run_artifact.json", type=["json"], key="delta_scen")

    base = _load_json_from_upload(up_base)
    scen = _load_json_from_upload(up_scen)

    if not base or not scen:
        st.info("Upload both baseline and scenario artifacts to view deltas.")
    else:
        st.subheader("Embedded scenario_delta")
        sd = scen.get("scenario_delta")
        if sd:
            st.json(sd)
        else:
            st.info("No embedded scenario_delta found; computing diffs from inputs/outputs.")

        st.subheader("Changed inputs")
        bi = base.get("inputs") or {}
        si = scen.get("inputs") or {}
        changed = []
        for k in sorted(set(bi.keys()) | set(si.keys())):
            if bi.get(k) != si.get(k):
                changed.append({"field": k, "baseline": bi.get(k), "scenario": si.get(k)})
        if changed:
            st.dataframe(pd.DataFrame(changed), use_container_width=True)
        else:
            st.info("No input differences detected.")

        st.subheader("Numeric output deltas")
        bo = base.get("outputs") or {}
        so = scen.get("outputs") or {}
        df = _numeric_delta_table(bo, so)
        if not df.empty:
            st.dataframe(df, use_container_width=True)
        else:
            st.info("No numeric output differences detected.")



        st.subheader("Structural / schema diff (read-only)")
        st.caption("Reports *structure* changes (constraints added/removed/meta changes, model cards) without numeric tolerances.")

        try:
            from shams_io.structural_diff import structural_diff as _structural_diff
            sd = _structural_diff(new_artifact=scen, old_artifact=base)
        except Exception as e:
            sd = None
            st.error(f"Structural diff failed: {e}")

        if isinstance(sd, dict):
            # Constraints changes
            cchg = (sd.get("constraints") or {})
            added = cchg.get("added") or []
            removed = cchg.get("removed") or []
            changed = cchg.get("changed_meta") or []
            cols = st.columns(3)
            cols[0].metric("constraints added", str(len(added)))
            cols[1].metric("constraints removed", str(len(removed)))
            cols[2].metric("constraints meta changed", str(len(changed)))

            if added:
                with st.expander("Added constraints", expanded=False):
                    st.write(added)
            if removed:
                with st.expander("Removed constraints", expanded=False):
                    st.write(removed)
            if changed:
                with st.expander("Changed constraint metadata", expanded=False):
                    st.dataframe(pd.DataFrame(changed), use_container_width=True, hide_index=True)

            # Model cards changes
            mc = (sd.get("model_cards") or {})
            mc_added = mc.get("added") or []
            mc_removed = mc.get("removed") or []
            mc_changed = mc.get("changed") or []
            cols2 = st.columns(3)
            cols2[0].metric("model cards added", str(len(mc_added)))
            cols2[1].metric("model cards removed", str(len(mc_removed)))
            cols2[2].metric("model cards changed", str(len(mc_changed)))
            if mc_added or mc_removed or mc_changed:
                with st.expander("Model card diffs", expanded=False):
                    st.json({"added": mc_added, "removed": mc_removed, "changed": mc_changed}, expanded=False)

            with st.expander("Raw structural diff JSON (audit)", expanded=False):
                st.json(sd, expanded=False)


# -----------------------------
# Run Library (Workspace)
# -----------------------------
with tab_library:
    st.header("Run Library")
    st.caption("Browse a workspace directory of SHAMS run/study artifacts (no physics changes; read-only).")

    def _scan_workspace(root: Path):
        runs = []
        studies = []
        if not root.exists():
            return runs, studies

        # Run artifacts
        for p in root.rglob("*.json"):
            if p.name.lower() in {"shams_run_artifact.json"} or p.name.lower().startswith("case_") or p.name.lower().endswith("_artifact.json"):
                try:
                    art = read_run_artifact(p)
                    k = art.get("kpis", {}) if isinstance(art, dict) else {}
                    prov = art.get("provenance", {}) if isinstance(art, dict) else {}
                    runs.append({
                        "type": "run",
                        "path": str(p),
                        "created_unix": float(art.get("created_unix", prov.get("created_unix", float("nan")))) if isinstance(art, dict) else float("nan"),
                        "hard_ok": bool(k.get("hard_ok", False)),
                        "hard_worst_margin": k.get("hard_worst_margin", None),
                        "Q": k.get("Q_DT_eqv", k.get("Q", None)),
                        "H98": k.get("H98", None),
                        "message": ((art.get("solver") or {}).get("message") if isinstance(art.get("solver"), dict) else ""),
                    })
                except Exception:
                    continue

        # Study indexes
        for p in root.rglob("index.json"):
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                if isinstance(data, dict) and data.get("schema_version") == "study_index.v1":
                    prov = data.get("provenance", {}) if isinstance(data.get("provenance"), dict) else {}
                    studies.append({
                        "type": "study",
                        "path": str(p),
                        "created_unix": float(data.get("created_unix", prov.get("created_unix", float('nan')))),
                        "n_cases": int(data.get("n_cases", 0)),
                        "elapsed_s": float(data.get("elapsed_s", float('nan'))),
                    })
            except Exception:
                continue
        return runs, studies

    default_ws = str((Path.cwd()/ "ui_runs").resolve())
    ws = st.text_input("Workspace folder", value=st.session_state.get("ui_workspace", default_ws))
    st.session_state.ui_workspace = ws
    root = Path(ws)

    colA, colB = st.columns([1, 1])
    with colA:
        do_scan = st.button("Scan workspace", use_container_width=True)
    with colB:
        st.write("")
        st.write("")

    if do_scan:
        runs, studies = _scan_workspace(root)
        st.session_state._ws_runs = runs
        st.session_state._ws_studies = studies

    runs = st.session_state.get("_ws_runs", [])
    studies = st.session_state.get("_ws_studies", [])

    st.subheader("Runs")
    if not runs:
        st.info("No run artifacts found yet. Tip: point runs write artifacts under your chosen output directory; studies write case_XXXX.json under the study out folder.")
    else:
        df = pd.DataFrame(runs)
        # Sort: newest first when available
        if "created_unix" in df.columns:
            df = df.sort_values("created_unix", ascending=False, na_position="last")
        st.dataframe(df, use_container_width=True, hide_index=True)

        sel = st.text_input("Select a run artifact path to open", value=st.session_state.get("selected_artifact_path", ""))
        if st.button("Open selected run", use_container_width=True):
            p = Path(sel)
            if p.exists():
                try:
                    art = read_run_artifact(p)
                    st.session_state.selected_artifact = art
                    st.session_state.selected_artifact_path = str(p)
                    st.success("Loaded run artifact into session.")
                except Exception as e:
                    st.error(f"Failed to read artifact: {e}")
            else:
                st.error("Path does not exist.")

    st.subheader("Studies")
    if studies:
        st.dataframe(pd.DataFrame(studies).sort_values("created_unix", ascending=False, na_position="last"), use_container_width=True, hide_index=True)
        ssel = st.text_input("Select a study index.json path to open", value=st.session_state.get("selected_study_index_path", ""))
        if st.button("Open selected study", use_container_width=True):
            p = Path(ssel)
            if p.exists():
                try:
                    st.session_state.selected_study_index_path = str(p)
                    st.session_state.selected_study_index = json.loads(p.read_text(encoding="utf-8"))
                    st.success("Loaded study index into session.")
                except Exception as e:
                    st.error(f"Failed to read study index: {e}")
            else:
                st.error("Path does not exist.")
    else:
        st.caption("No study indexes found in this workspace.")

# -----------------------------
# Constraint Cockpit
# -----------------------------
with tab_constraints:
    st.header("Constraint Cockpit")
    st.caption("Interactively triage constraints using the embedded constraint ledger (read-only).")

    art = st.session_state.get("selected_artifact")
    if not isinstance(art, dict):
        st.info("Load a run artifact first (Run Library or Artifacts Explorer).")
    else:
        ledger = art.get("constraint_ledger", {})
        entries = ledger.get("entries", []) if isinstance(ledger, dict) else []
        if not entries:
            st.warning("This artifact has no constraint ledger. (It should be present in v39+ artifacts.)")
        else:
            df = pd.DataFrame(entries)
            # Basic filters
            c1, c2, c3 = st.columns([1,1,1])
            with c1:
                sev = st.multiselect("Severity", sorted(df.get("severity", pd.Series(["hard"])).dropna().unique().tolist()), default=["hard","soft"] if "soft" in df.get("severity", pd.Series([])).unique() else ["hard"])
            with c2:
                grp = st.multiselect("Group", sorted(df.get("group", pd.Series(["general"])).dropna().unique().tolist()), default=[])
            with c3:
                show_only_failed = st.checkbox("Only failed constraints", value=True)

            view = df.copy()
            if sev:
                view = view[view["severity"].isin(sev)]
            if grp:
                view = view[view["group"].isin(grp)]
            if show_only_failed and "passed" in view.columns:
                view = view[view["passed"] == False]

            # Sort: worst first by margin_frac or margin
            if "margin_frac" in view.columns:
                view = view.sort_values("margin_frac", ascending=True, na_position="last")
            elif "margin" in view.columns:
                view = view.sort_values("margin", ascending=True, na_position="last")

            st.subheader("Ledger")
            st.dataframe(view, use_container_width=True, hide_index=True)

            st.subheader("Top blockers")
            top = ledger.get("top_blockers", []) if isinstance(ledger, dict) else []
            if top:
                st.dataframe(pd.DataFrame(top), use_container_width=True, hide_index=True)
            fp = ledger.get("ledger_fingerprint_sha256")
            if fp:
                st.caption(f"Ledger fingerprint: `{fp}`")


# -----------------------------
# Constraint Inspector (read-only)
# -----------------------------
with tab_constraint_inspector:
    st.header("Constraint Inspector")
    st.caption("Read-only, equation-first inspection of a single constraint: raw inequality, margin, meaning, knobs, and provenance (when available).")

    art = st.session_state.get("selected_artifact")
    if not isinstance(art, dict):
        st.info("Load a run artifact first (Run Library or Artifacts Explorer).")
    else:
        constraints_list = art.get("constraints") or []
        # Build a name -> constraint dict map (best-effort)
        name_to_c = {}
        for c in constraints_list:
            if isinstance(c, dict) and c.get("name"):
                name_to_c[str(c.get("name"))] = c

        ledger = art.get("constraint_ledger", {})
        entries = ledger.get("entries", []) if isinstance(ledger, dict) else []
        names = []
        # Prefer ledger order if present (it should reflect evaluation order)
        if entries:
            for e in entries:
                n = str(e.get("name"))
                if n and n not in names:
                    names.append(n)
        else:
            names = sorted(list(name_to_c.keys()))

        if not names:
            st.warning("No constraints found in this artifact.")
        else:
            sel = st.selectbox("Select constraint", names, index=0, key="constraint_inspector_select")

            # Pull both ledger entry (if present) and raw constraint dict (if present)
            entry = None
            if entries:
                for e in entries:
                    if str(e.get("name")) == sel:
                        entry = e
                        break
            c = name_to_c.get(sel, {}) if isinstance(name_to_c.get(sel, {}), dict) else {}

            # Compose a canonical view (prefer ledger fields where available)
            view = {}
            for src in (c, entry or {}):
                if isinstance(src, dict):
                    view.update({k: src.get(k) for k in src.keys()})

            # Core inequality (verbatim fields; no inferred math)
            sense = str(view.get("sense") or "")
            value = view.get("value")
            limit = view.get("limit")
            units = str(view.get("units") or "")
            meaning = str(view.get("meaning") or view.get("note") or "")

            st.subheader("Inequality")
            if sense and value is not None and limit is not None:
                st.code(f"{sel}: value {sense} limit    (value={value}, limit={limit}, units={units})", language="text")
            else:
                st.code(f"{sel}: (insufficient fields to render inequality)", language="text")

            # Pass/fail + margins
            cols = st.columns(4)
            cols[0].metric("passed", str(bool(view.get("passed", False))))
            if view.get("severity") is not None:
                cols[1].metric("severity", str(view.get("severity")))
            if view.get("group") is not None:
                cols[2].metric("group", str(view.get("group")))
            if view.get("dominance_rank") is not None:
                cols[3].metric("dominance_rank", str(view.get("dominance_rank")))

            c1, c2, c3 = st.columns(3)
            if view.get("margin") is not None:
                c1.metric("margin", f"{view.get('margin')}")
            if view.get("margin_frac") is not None:
                c2.metric("margin_frac", f"{view.get('margin_frac')}")
            if view.get("violation_score") is not None:
                c3.metric("violation_score", f"{view.get('violation_score')}")

            st.subheader("Meaning / proxy")
            if meaning.strip():
                st.write(meaning)
            else:
                st.info("No meaning/proxy text is attached to this constraint.")

            # Knobs + dominant inputs
            st.subheader("Knobs / dominant inputs (if present)")
            bb = view.get("best_knobs")
            di = view.get("dominant_inputs")
            kcol1, kcol2 = st.columns(2)
            with kcol1:
                if bb:
                    st.write("**best_knobs**")
                    st.write(bb)
                else:
                    st.caption("best_knobs: (none)")
            with kcol2:
                if di:
                    st.write("**dominant_inputs**")
                    st.write(di)
                else:
                    st.caption("dominant_inputs: (none)")

            # Provenance (constraint-level and artifact-level)
            st.subheader("Provenance (if present)")
            prov = {}
            if isinstance(view.get("provenance"), dict):
                prov["constraint"] = view.get("provenance")
            if isinstance(art.get("provenance"), dict):
                prov["artifact"] = art.get("provenance")
            if prov:
                st.json(prov, expanded=False)
            else:
                st.info("No provenance keys present on this constraint (artifact-level provenance may still exist under artifact.provenance).")

            # Raw views for auditability
            with st.expander("Raw JSON (audit)", expanded=False):
                if isinstance(entry, dict):
                    st.write("**constraint_ledger entry**")
                    st.json(entry, expanded=False)
                if isinstance(c, dict) and c:
                    st.write("**constraints[] item**")
                    st.json(c, expanded=False)


# -----------------------------
# Sensitivity Explorer
# -----------------------------
with tab_sensitivity:
    st.header("Sensitivity Explorer")
    st.caption("Local finite-difference sensitivities around the current point (no model changes).")

    art = st.session_state.get("selected_artifact")
    if not isinstance(art, dict):
        st.info("Load a run artifact first (Run Library or Artifacts Explorer).")
    else:
        inp_d = art.get("inputs", {})
        if not isinstance(inp_d, dict):
            st.error("Artifact inputs missing or invalid.")
        else:
            try:
                base = PointInputs.from_dict(inp_d)
            except Exception:
                # Fallback: try direct constructor with expected keys
                try:
                    base = PointInputs(**{k: inp_d[k] for k in PointInputs.__dataclass_fields__.keys() if k in inp_d})
                except Exception as e:
                    st.error(f"Could not build PointInputs from artifact inputs: {e}")
                    base = None

            if base is not None:
                st.subheader("Base point")
                st.json(base.__dict__)

                # Choose knobs + outputs
                knob_defaults = ["Ip_MA", "fG", "Bt_T", "R0_m", "a_m", "kappa", "Paux_MW", "Ti_keV"]
                available_knobs = [k for k in knob_defaults if k in base.__dict__]
                knobs = st.multiselect("Knobs", available_knobs, default=["Ip_MA", "fG"])
                outputs_default = ["Q_DT_eqv", "H98", "Pfus_MW", "Palpha_MW", "beta_N", "nbar20", "Tbr"]
                outputs = st.multiselect("Outputs", outputs_default, default=["Q_DT_eqv", "H98"])

                h = st.number_input("Step size (absolute)", value=0.05, min_value=1e-6, format="%.6f")

                if st.button("Compute local sensitivities", use_container_width=True):
                    try:
                        from solvers.sensitivity import local_sensitivities
                        from phase1_core import hot_ion_point

                        def evaluator(x: PointInputs):
                            out = hot_ion_point(x)
                            return out if isinstance(out, dict) else {}

                        sens = local_sensitivities(base, params=knobs, outputs=outputs, evaluator=evaluator, h=float(h))
                        # Flatten for table
                        rows = []
                        for o in outputs:
                            for p in knobs:
                                v = sens.get(o, {}).get(p)
                                if v is not None:
                                    rows.append({"output": o, "knob": p, "d(output)/d(knob)": v})
                        if rows:
                            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
                        else:
                            st.info("No sensitivities computed (NaNs or missing outputs).")
                        if "_base" in sens:
                            st.subheader("Base outputs snapshot")
                            st.json(sens["_base"])
                    except Exception as e:
                        st.error(f"Sensitivity computation failed: {e}")

# -----------------------------
# Feasibility Map Viewer
# -----------------------------
with tab_feasmap:
    st.header("Feasibility Map")
    st.caption("Visualize feasibility from study sweeps (heatmap).")

    # Load study index either from session (Run Library) or by path
    p_default = st.session_state.get("selected_study_index_path", "")
    p = st.text_input("Study index.json path", value=p_default)
    idx_data = None
    if p and Path(p).exists():
        try:
            idx_data = json.loads(Path(p).read_text(encoding="utf-8"))
        except Exception as e:
            st.error(f"Could not read study index: {e}")

    if not isinstance(idx_data, dict) or idx_data.get("schema_version") != "study_index.v1":
        st.info("Provide a valid study_out/index.json (schema study_index.v1).")
    else:
        cases = idx_data.get("cases", [])
        study = idx_data.get("study", {})
        sweeps = (study.get("sweeps") if isinstance(study, dict) else None) or []
        # Determine candidate in_ variables for axes
        in_cols = []
        if cases and isinstance(cases, list) and isinstance(cases[0], dict):
            for k in cases[0].keys():
                if k.startswith("in_"):
                    in_cols.append(k)
        # Prefer sweep variables
        sweep_vars = ["in_"+str(s.get("name")) for s in sweeps if isinstance(s, dict) and s.get("name") is not None]
        axis_candidates = [c for c in sweep_vars if c in in_cols] + [c for c in in_cols if c not in sweep_vars]
        if len(axis_candidates) < 2:
            st.warning("Need at least two swept input variables (in_*) to plot a 2D feasibility map.")
        else:
            c1, c2 = st.columns([1,1])
            with c1:
                xcol = st.selectbox("X axis", axis_candidates, index=0)
            with c2:
                ycol = st.selectbox("Y axis", axis_candidates, index=1 if len(axis_candidates)>1 else 0)

            df = pd.DataFrame(cases)
            if "ok" not in df.columns:
                st.error("Study cases table missing 'ok' field.")
            else:
                # Build pivot grid
                xs = sorted(df[xcol].dropna().unique().tolist())
                ys = sorted(df[ycol].dropna().unique().tolist())
                import numpy as np
                grid = np.full((len(ys), len(xs)), np.nan)
                for _, r in df.iterrows():
                    try:
                        xi = xs.index(r[xcol])
                        yi = ys.index(r[ycol])
                        grid[yi, xi] = 1.0 if bool(r["ok"]) else 0.0
                    except Exception:
                        continue

                st.subheader("Feasibility heatmap (1=feasible, 0=infeasible)")
                try:
                    import matplotlib.pyplot as plt  # type: ignore
                    fig, ax = plt.subplots()
                    im = ax.imshow(grid, origin="lower", aspect="auto")
                    ax.set_xticks(range(len(xs)))
                    ax.set_yticks(range(len(ys)))
                    ax.set_xticklabels([str(x) for x in xs], rotation=45, ha="right")
                    ax.set_yticklabels([str(y) for y in ys])
                    ax.set_xlabel(xcol)
                    ax.set_ylabel(ycol)
                    st.pyplot(fig, clear_figure=True)
                except Exception as e:
                    st.error(f"Plot failed: {e}")

                st.subheader("Pick a case to open")
                selx = st.selectbox("X value", xs, index=0)
                sely = st.selectbox("Y value", ys, index=0)
                sub = df[(df[xcol]==selx) & (df[ycol]==sely)]
                if sub.empty:
                    st.info("No case for that cell.")
                else:
                    st.dataframe(sub[["case","ok","iters","message","path"] + [xcol,ycol]], use_container_width=True, hide_index=True)
                    if st.button("Load this case artifact", use_container_width=True):
                        path = str(sub.iloc[0]["path"])
                        try:
                            art = read_run_artifact(Path(path))
                            st.session_state.selected_artifact = art
                            st.session_state.selected_artifact_path = path
                            st.success("Loaded case artifact into session.")
                        except Exception as e:
                            st.error(f"Could not load case artifact: {e}")


# -----------------------------
# UI Upgrade Pack v53 (UI-only): decision/provenance/knobs/regression/study dashboard/maturity/assumptions/export/solver introspection
# -----------------------------
def _get_active_artifact(label: str = "Use loaded artifact in session") -> dict | None:
    "Return the currently active artifact (from session_state or upload)."
    art = st.session_state.get("selected_artifact")
    if isinstance(art, dict) and art:
        st.info("Using artifact loaded into session (Run Library / Feasibility Map).")
        return art
    up = st.file_uploader("Upload shams_run_artifact.json", type=["json"], key=f"active_artifact_upload_{label}")
    return _load_json_from_upload(up)

def _guess_point_inputs_from_artifact(art: dict) -> PointInputs | None:
    "Best-effort extraction of PointInputs from an artifact. Falls back safely."
    if not isinstance(art, dict):
        return None
    cand = {}
    for k in ["inputs", "point", "point_inputs", "design_point", "config", "run_config", "resolved_config"]:
        v = art.get(k)
        if isinstance(v, dict):
            cand.update(v)
    cand.update({k: art.get(k) for k in ["R0_m","a_m","kappa","Bt_T","B0_T","Ip_MA","Ti_keV","fG","Paux_MW","Ti_over_Te","fuel_mode"] if k in art})
    if "B0_T" in cand and "Bt_T" not in cand:
        cand["Bt_T"] = cand["B0_T"]
    if "Ti_Te" in cand and "Ti_over_Te" not in cand:
        cand["Ti_over_Te"] = cand["Ti_Te"]
    if "Ti/Te" in cand and "Ti_over_Te" not in cand:
        cand["Ti_over_Te"] = cand["Ti/Te"]
    try:
        return _make_point_inputs_safe(**cand)
    except Exception:
        return None

def _decision_summary_from_artifact(art: dict) -> dict:
    kpis = art.get("kpis", {}) if isinstance(art.get("kpis"), dict) else {}
    cons = art.get("constraints", []) if isinstance(art.get("constraints"), list) else []
    ledger = art.get("constraint_ledger", {}) if isinstance(art.get("constraint_ledger"), dict) else {}
    feas = bool(art.get("is_feasible")) if "is_feasible" in art else None
    if feas is None:
        feas = all((not bool(c.get("failed"))) for c in cons) if cons else None
    top = ledger.get("top_blockers") if isinstance(ledger.get("top_blockers"), list) else []
    if not top and cons:
        failed = [c for c in cons if c.get("failed")]
        failed = failed[:8]
        top = [{"name": c.get("name"), "group": c.get("group"), "margin": c.get("margin"), "severity": c.get("severity")} for c in failed]
    return {"feasible": feas, "kpis": kpis, "top_blockers": top, "ledger": ledger, "constraints": cons}

def _download_json_button(label: str, data: dict, fname: str, key: str):
    try:
        st.download_button(label, data=json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8"),
                           file_name=fname, mime="application/json", key=key)
    except Exception as e:
        st.warning(f"Download not available: {e}")

with tab_decision:
    st.header("Decision Front Page Builder")
    st.caption("UI-native reconstruction of the decision-grade front-page summary from a run artifact (no physics changes).")

    art = _get_active_artifact("decision")
    if not art:
        st.info("Load an artifact to build the decision summary.")
    else:
        d = _decision_summary_from_artifact(art)
        c1, c2, c3 = st.columns([1,1,1])
        with c1:
            st.metric("Feasibility verdict", "FEASIBLE ✅" if d["feasible"] else ("INFEASIBLE ❌" if d["feasible"] is not None else "UNKNOWN"))
        with c2:
            st.metric("Top KPI: Q", f"{d['kpis'].get('Q_DT_eqv', d['kpis'].get('Q', '—'))}")
        with c3:
            st.metric("Top KPI: Pfus (MW)", f"{d['kpis'].get('P_fus_MW', d['kpis'].get('Pfus_MW', '—'))}")

        st.subheader("Dominant blockers")
        if d["top_blockers"]:
            st.dataframe(_safe_df(d["top_blockers"]), use_container_width=True, hide_index=True)
        else:
            st.write("No blockers found in artifact.")

        with st.expander("Full decision inputs (provenance + schema versions)"):
            prov = art.get("provenance", {}) if isinstance(art.get("provenance"), dict) else {}
            st.json({
                "schema_version": art.get("schema_version"),
                "repo_version": prov.get("repo_version"),
                "git_commit": prov.get("git_commit"),
                "python": prov.get("python"),
                "platform": prov.get("platform"),
                "created_unix": prov.get("created_unix"),
            })

        _download_json_button("Download decision summary JSON", d, "decision_summary.json", "dl_decision_summary")


with tab_nonfeas:
    st.header("Guided Non-Feasibility Mode")
    st.caption("Turn infeasible outcomes into a structured, auditable recovery workflow (UI-only; no physics changes).")

    art = _get_active_artifact("nonfeas")
    if not art:
        st.info("Load an artifact to guide a non-feasibility recovery path.")
    else:
        cons = art.get("constraints", []) if isinstance(art.get("constraints"), list) else []
        kpis = art.get("kpis", {}) if isinstance(art.get("kpis"), dict) else {}

        # Determine hard feasibility
        feasible_hard = None
        if "feasible_hard" in kpis:
            try:
                feasible_hard = bool(kpis.get("feasible_hard"))
            except Exception:
                feasible_hard = None
        if feasible_hard is None and cons:
            try:
                feasible_hard = all(
                    bool(c.get("passed", True))
                    for c in cons
                    if str(c.get("severity", "hard")).lower() == "hard"
                )
            except Exception:
                feasible_hard = None

        if feasible_hard is True:
            st.success("This run is hard-feasible. Guided non-feasibility mode is not needed.")
        else:
            # Get or construct a non-feasibility certificate
            cert = art.get("nonfeasibility_certificate") if isinstance(art.get("nonfeasibility_certificate"), dict) else None
            if not cert:
                hard_failed = [
                    c for c in cons
                    if str(c.get("severity", "hard")).lower() == "hard" and not bool(c.get("passed", True))
                ]

                def _mkey(c):
                    try:
                        return float(c.get("margin", 0.0))
                    except Exception:
                        return 0.0

                hard_failed.sort(key=_mkey)
                cert = {
                    "hard_feasible": False,
                    "dominant_blockers": [{
                        "name": c.get("name", ""),
                        "group": c.get("group", ""),
                        "value": c.get("value"),
                        "limit": c.get("limit"),
                        "sense": c.get("sense"),
                        "margin": c.get("margin"),
                        "meaning": c.get("meaning", ""),
                        "best_knobs": c.get("best_knobs", []),
                        "maturity": c.get("maturity"),
                        "provenance": c.get("provenance"),
                    } for c in hard_failed[:10]],
                    "recommendation": "Move the listed best_knobs (and/or relax assumptions) until all hard constraints pass.",
                }

            st.subheader("Non-Feasibility Certificate")
            st.json(cert)

            t1, t2, t3 = st.tabs(["1) Diagnose", "2) Minimal relaxations", "3) Create a scenario (deck)"])

            with t1:
                st.markdown("### Dominant hard blockers (ranked)")
                blockers = cert.get("dominant_blockers", []) if isinstance(cert.get("dominant_blockers"), list) else []
                if blockers:
                    bdf = _safe_df(blockers)
                    pref = [c for c in ["group", "name", "margin", "value", "limit", "sense", "meaning", "best_knobs", "maturity"] if c in bdf.columns]
                    st.dataframe(bdf[pref] if pref else bdf, use_container_width=True, hide_index=True)
                else:
                    st.warning("No dominant blockers found in certificate.")

                # Solver hints (if present)
                out = art.get("outputs", {}) if isinstance(art.get("outputs"), dict) else {}
                solver = out.get("_solver") if isinstance(out.get("_solver"), dict) else art.get("solver")
                if isinstance(solver, dict) and solver:
                    st.markdown("### Solver hints (from artifact)")
                    show = {k: solver.get(k) for k in ["status", "reason", "clamped", "clamped_on", "residuals", "ui_log"] if k in solver}
                    st.json(show or solver)

                st.markdown("### Action principle")
                st.write(
                    "Fix **hard** blockers first. Soft constraints are advisory unless your decision policy says otherwise. "
                    "Use the knob suggestions as **directional guidance** (not optimization)."
                )

            with t2:
                st.markdown("### Propose a nearest-feasible adjustment (within UI)")
                base = _guess_point_inputs_from_artifact(art)
                if base is None:
                    base = st.session_state.get("last_point_inp")

                if base is None:
                    st.warning("Could not infer PointInputs from artifact. Run Point Designer once or ensure artifact includes inputs.")
                else:
                    st.caption("Choose a dominant blocker, then adjust one or more knobs and re-evaluate.")
                    blockers = cert.get("dominant_blockers", []) if isinstance(cert.get("dominant_blockers"), list) else []
                    if blockers:
                        labels = []
                        for i, b in enumerate(blockers):
                            nm = b.get("name", "") or f"blocker_{i}"
                            mg = b.get("margin")
                            labels.append(f"{i:02d} — {nm} (margin={mg})")
                        bi = st.selectbox("Select blocker", options=list(range(len(blockers))), format_func=lambda i: labels[i], key="nf_blocker_sel")
                        b = blockers[int(bi)]
                        st.markdown("**Suggested knobs (directional):**")
                        st.write(b.get("best_knobs", []) or ["(none provided)"])
                        st.markdown("**Meaning:**")
                        st.write(b.get("meaning", "(no meaning field)"))

                    knob_fields = ["Ip_MA", "fG", "Bt_T", "R0_m", "a_m", "kappa", "Ti_keV", "Paux_MW", "Ti_over_Te"]
                    colA, colB = st.columns([2, 1])
                    with colA:
                        sel_knobs = st.multiselect("Knobs to adjust", options=knob_fields, default=["Ip_MA"], key="nf_knobs")
                    with colB:
                        mode = st.selectbox("Adjustment mode", options=["percent", "absolute"], index=0, key="nf_adj_mode")

                    deltas = {}
                    for k in sel_knobs:
                        v0 = float(getattr(base, k))
                        if mode == "percent":
                            d = st.slider(f"{k} Δ (%)", -50.0, 50.0, 5.0, step=0.5, key=f"nf_d_{k}")
                            deltas[k] = v0 * (1.0 + d / 100.0)
                        else:
                            step = 0.1 if abs(v0) < 10 else 1.0
                            d = st.number_input(f"{k} new value", value=v0, step=step, key=f"nf_abs_{k}")
                            deltas[k] = float(d)

                    fuel_mode = st.selectbox("fuel_mode", options=["DT", "DD"], index=0 if getattr(base, "fuel_mode", "DT") == "DT" else 1, key="nf_fuel_mode")

                    run = st.button("Re-evaluate adjusted point", key="nf_run_eval", use_container_width=True)
                    if run:
                        try:
                            d = base.__dict__.copy()
                            d.update({k: float(v) for k, v in deltas.items()})
                            d["fuel_mode"] = str(fuel_mode)
                            pi = PointInputs(**d)

                            out2 = hot_ion_point(pi, Paux_for_Q_MW=float(getattr(pi, "Paux_MW", 0.0)))
                            cons2 = evaluate_constraints(out2)
                            art2 = build_run_artifact(
                                inputs=dict(pi.__dict__),
                                outputs=dict(out2),
                                constraints=cons2,
                                meta={"mode": "guided_nonfeas"},
                                baseline_inputs=dict(base.__dict__),
                            )
                            st.session_state["nf_last_artifact"] = art2
                            k2 = art2.get("kpis", {}) if isinstance(art2.get("kpis"), dict) else {}
                            st.success(f"Re-evaluated. feasible_hard={k2.get('feasible_hard')}")

                            led = art2.get("constraint_ledger", {}) if isinstance(art2.get("constraint_ledger"), dict) else {}
                            tb = led.get("top_blockers") if isinstance(led.get("top_blockers"), list) else []
                            if tb:
                                st.subheader("New top blockers")
                                st.dataframe(_safe_df(tb), use_container_width=True, hide_index=True)

                            with st.expander("New run artifact (raw)"):
                                st.json(art2)

                            _download_json_button("Download adjusted run artifact", art2, "shams_run_artifact_adjusted.json", "dl_nf_adjusted_artifact")
                        except Exception as e:
                            st.error(f"Re-evaluation failed: {type(e).__name__}: {e}")

            with t3:
                st.markdown("### Create a scenario deck for reproducible follow-up")
                base = _guess_point_inputs_from_artifact(art) or st.session_state.get("last_point_inp")
                last = st.session_state.get("nf_last_artifact")
                if not isinstance(last, dict):
                    st.info("First run an adjustment in 'Minimal relaxations' to generate a proposed follow-up scenario.")
                else:
                    try:
                        import yaml  # type: ignore
                    except Exception:
                        yaml = None  # type: ignore

                    new_inputs = last.get("inputs") if isinstance(last.get("inputs"), dict) else {}
                    base_inputs = dict(base.__dict__) if base is not None else (art.get("inputs") if isinstance(art.get("inputs"), dict) else {})

                    delta = {}
                    for k, v in new_inputs.items():
                        if k in base_inputs and base_inputs.get(k) != v:
                            delta[k] = {"from": base_inputs.get(k), "to": v}

                    st.subheader("Scenario delta (inputs)")
                    st.json(delta if delta else {"note": "No input delta detected."})

                    case_deck = {
                        "schema_version": "case_deck.v1",
                        "name": "guided_nonfeas_followup",
                        "base": {},
                        "point": new_inputs,
                        "notes": {
                            "generated_by": "Guided Non-Feasibility Mode",
                            "source_artifact_schema": art.get("schema_version"),
                        },
                    }

                    deck_txt = yaml.safe_dump(case_deck, sort_keys=False) if yaml is not None else json.dumps(case_deck, indent=2)

                    st.markdown("### Case deck (v1)")
                    st.code(deck_txt, language="yaml" if yaml is not None else "json")

                    st.download_button(
                        "Download case_deck.yaml",
                        data=deck_txt.encode("utf-8"),
                        file_name="case_deck.yaml",
                        mime="text/yaml" if yaml is not None else "application/json",
                        use_container_width=True,
                    )
                    st.download_button(
                        "Download scenario_delta.json",
                        data=json.dumps(delta, indent=2).encode("utf-8"),
                        file_name="scenario_delta.json",
                        mime="application/json",
                        use_container_width=True,
                    )


with tab_cprov:
    st.header("Constraint Provenance Drill-Down")
    st.caption("Click into constraints to see definition fields, fingerprints, and maturity/provenance metadata embedded in the artifact.")

    art = _get_active_artifact("cprov")
    if not art:
        st.info("Load an artifact to inspect constraint provenance.")
    else:
        cons = art.get("constraints", [])
        if not isinstance(cons, list) or not cons:
            st.warning("No 'constraints' list found in artifact.")
        else:
            df = _safe_df(cons)
            pref_cols = [c for c in ["group","name","failed","soft_failed","severity","value","limit","margin","margin_frac","units","fingerprint","provenance_fingerprint","maturity"] if c in df.columns]
            st.dataframe(df[pref_cols] if pref_cols else df, use_container_width=True, hide_index=True)

            names = []
            for i,c in enumerate(cons):
                n = c.get("name") or c.get("id") or f"constraint_{i}"
                names.append(f"{i:03d} — {n}")
            sel = st.selectbox("Select constraint", options=list(range(len(cons))), format_func=lambda i: names[i], key="cprov_sel")
            c = cons[int(sel)]
            st.subheader("Selected constraint (raw)")
            st.json(c)
            if isinstance(c, dict):
                st.markdown("**Fingerprint fields**")
                st.code("\n".join([f"{k}: {c.get(k)}" for k in ["fingerprint","provenance_fingerprint","constraint_fingerprint_sha256"] if k in c] or ["(none found)"]))

with tab_knobs:
    st.header("Knob Trade-Space Explorer")
    st.caption("Explore a 2-knob trade-space by evaluating a small grid around the active point (no optimization; feasibility-first).")

    art = _get_active_artifact("knobs")
    base = _guess_point_inputs_from_artifact(art) if art else None
    if base is None:
        base = st.session_state.get("last_point_inp")

    if base is None:
        st.info("Load an artifact (or run Point Designer) to initialize a base point.")
    else:
        st.markdown("**Base point (editable)**")
        col1, col2, col3 = st.columns(3)
        with col1:
            R0_m = st.number_input("R0 (m)", value=float(base.R0_m), step=0.01, key="knob_R0")
            a_m = st.number_input("a (m)", value=float(base.a_m), step=0.01, key="knob_a")
            kappa = st.number_input("kappa", value=float(base.kappa), step=0.05, key="knob_kappa")
        with col2:
            Bt_T = st.number_input("Bt (T)", value=float(base.Bt_T), step=0.1, key="knob_Bt")
            Ip_MA = st.number_input("Ip (MA)", value=float(base.Ip_MA), step=0.1, key="knob_Ip")
            fG = st.number_input("fG", value=float(base.fG), step=0.01, key="knob_fG")
        with col3:
            Ti_keV = st.number_input("Ti (keV)", value=float(base.Ti_keV), step=0.5, key="knob_Ti")
            Paux_MW = st.number_input("Paux (MW)", value=float(base.Paux_MW), step=1.0, key="knob_Paux")
            Ti_over_Te = st.number_input("Ti/Te", value=float(getattr(base, "Ti_over_Te", 2.0)), step=0.1, key="knob_TiTe")

        fuel_mode = st.selectbox("fuel_mode", options=["DT","DD"], index=0 if getattr(base, "fuel_mode", "DT")=="DT" else 1, key="knob_fuel")

        knobs = ["Ip_MA","fG","Bt_T","R0_m","Paux_MW","Ti_keV"]
        kx = st.selectbox("Knob X", knobs, index=0, key="knob_kx")
        ky = st.selectbox("Knob Y", knobs, index=1, key="knob_ky")

        def _getv(pi: PointInputs, k: str) -> float:
            return float(getattr(pi, k))
        def _setv(pi: PointInputs, k: str, v: float) -> PointInputs:
            d = pi.__dict__.copy()
            d[k]=float(v)
            return PointInputs(**d)

        x0=_getv(base,kx); y0=_getv(base,ky)
        colA,colB=st.columns(2)
        with colA:
            x_span = st.number_input("X span (+/-)", value=0.1*abs(x0) if abs(x0)>0 else 0.1, step=0.01, key="knob_xspan")
        with colB:
            y_span = st.number_input("Y span (+/-)", value=0.1*abs(y0) if abs(y0)>0 else 0.1, step=0.01, key="knob_yspan")
        nx = st.slider("X grid points", 3, 15, 9, key="knob_nx")
        ny = st.slider("Y grid points", 3, 15, 9, key="knob_ny")
        run = st.button("Evaluate grid", key="knob_run", use_container_width=True)

        if run:
            import numpy as np
            xs = np.linspace(x0-x_span, x0+x_span, int(nx))
            ys = np.linspace(y0-y_span, y0+y_span, int(ny))
            rows=[]
            with st.spinner("Evaluating grid..."):
                for xv in xs:
                    for yv in ys:
                        pi = PointInputs(R0_m=float(R0_m), a_m=float(a_m), kappa=float(kappa),
                                         Bt_T=float(Bt_T), Ip_MA=float(Ip_MA), Ti_keV=float(Ti_keV),
                                         fG=float(fG), Paux_MW=float(Paux_MW), Ti_over_Te=float(Ti_over_Te),
                                         fuel_mode=str(fuel_mode))
                        pi = _setv(pi, kx, float(xv))
                        pi = _setv(pi, ky, float(yv))
                        try:
                            out = hot_ion_point(pi)
                            cons = evaluate_constraints(out, point_inputs=pi)
                            ok = all((not bool(c.get("failed"))) for c in cons)
                            top=None
                            if not ok:
                                failed=[c for c in cons if c.get("failed")]
                                if failed:
                                    top=failed[0].get("name")
                            rows.append({kx: float(xv), ky: float(yv), "feasible": bool(ok), "top_blocker": top,
                                         "Q": float(out.get("Q_DT_eqv", out.get("Q", float('nan')))),
                                         "Pfus_MW": float(out.get("P_fus_MW", out.get("Pfus_MW", float('nan'))))})
                        except Exception:
                            rows.append({kx: float(xv), ky: float(yv), "feasible": False, "top_blocker": "eval_error", "Q": float('nan'), "Pfus_MW": float('nan')})
            df=pd.DataFrame(rows)
            st.subheader("Grid results (table)")
            st.dataframe(df, use_container_width=True, hide_index=True)

            try:
                piv = df.pivot(index=ky, columns=kx, values="feasible")
                st.subheader("Feasibility heatmap (True=1 / False=0)")
                st.dataframe(piv.astype(int), use_container_width=True)
            except Exception as e:
                st.warning(f"Could not pivot heatmap: {e}")

with tab_regress:
    st.header("What broke? Regression Viewer")
    st.caption("Compare two artifacts: constraints, ledgers, model sets, and key KPIs. This is UI-only; it doesn't modify artifacts.")

    c1, c2 = st.columns(2)
    with c1:
        upA = st.file_uploader("Artifact A (json)", type=["json"], key="regA")
        artA = _load_json_from_upload(upA)
    with c2:
        upB = st.file_uploader("Artifact B (json)", type=["json"], key="regB")
        artB = _load_json_from_upload(upB)

    if artA and artB:
        def _kpi_df(art):
            k = art.get("kpis", {}) if isinstance(art.get("kpis"), dict) else {}
            return pd.DataFrame([{"kpi": kk, "value": vv} for kk,vv in k.items()]).sort_values("kpi")
        st.subheader("KPI diff")
        dfA=_kpi_df(artA).set_index("kpi")
        dfB=_kpi_df(artB).set_index("kpi")
        join=dfA.join(dfB, lsuffix="_A", rsuffix="_B", how="outer")
        join["delta"]=pd.to_numeric(join["value_B"], errors="coerce")-pd.to_numeric(join["value_A"], errors="coerce")
        st.dataframe(join.reset_index().sort_values("kpi"), use_container_width=True, hide_index=True)

        st.subheader("New / worsened constraint failures")
        consA=artA.get("constraints", []) if isinstance(artA.get("constraints"), list) else []
        consB=artB.get("constraints", []) if isinstance(artB.get("constraints"), list) else []
        def _map(cons):
            m={}
            for c in cons:
                name=c.get("name") or c.get("id")
                if name:
                    m[name]=c
            return m
        mA=_map(consA); mB=_map(consB)
        names=sorted(set(mA.keys())|set(mB.keys()))
        rows=[]
        for n in names:
            a=mA.get(n,{}); b=mB.get(n,{})
            fa=bool(a.get("failed")); fb=bool(b.get("failed"))
            ma=a.get("margin"); mb=b.get("margin")
            rows.append({"name": n, "failed_A": fa, "failed_B": fb, "margin_A": ma, "margin_B": mb,
                         "margin_delta": (mb-ma) if isinstance(ma,(int,float)) and isinstance(mb,(int,float)) else None})
        df=pd.DataFrame(rows)
        df_bad=df[(df["failed_B"]==True) & ((df["failed_A"]==False) | (df["failed_A"].isna()))]
        st.markdown("**New failures in B**")
        st.dataframe(df_bad.sort_values("name"), use_container_width=True, hide_index=True)
        st.markdown("**Largest margin regressions (B-A)**")
        df_reg=df.dropna(subset=["margin_delta"]).sort_values("margin_delta").head(20)
        st.dataframe(df_reg, use_container_width=True, hide_index=True)

        st.subheader("Model set comparison")
        msA=artA.get("model_set"); msB=artB.get("model_set")
        st.json({"model_set_A": msA, "model_set_B": msB})

with tab_study_dash:
    st.header("Study Dashboard")
    st.caption("Manager-grade summary for study outputs (feasible fraction, dominant blockers, robustness).")

    up = st.file_uploader("Upload study index.json (study_index.v1)", type=["json"], key="sd_up")
    idx_data = _load_json_from_upload(up)
    if not idx_data:
        idx_path = st.session_state.get("selected_study_path")
        if idx_path and Path(idx_path).exists():
            try:
                idx_data = json.loads(Path(idx_path).read_text(encoding="utf-8"))
                st.info("Loaded study index from session.")
            except Exception:
                idx_data = None

    if idx_data:
        st.subheader("Study headline")
        st.json({k: idx_data.get(k) for k in ["schema_version","n_cases","elapsed_s","created_unix"] if k in idx_data})
        cases = idx_data.get("cases", [])
        if isinstance(cases, list) and cases:
            df = pd.DataFrame(cases)
            if "ok" in df.columns:
                ok_frac = float(df["ok"].mean())
                st.metric("Feasible fraction", f"{ok_frac:.3f}")
            for col in ["dominant_blocker","top_blocker","blocker"]:
                if col in df.columns:
                    st.subheader("Dominant blocker distribution")
                    hist = df[col].fillna("(none)").value_counts().reset_index()
                    hist.columns=[col,"count"]
                    st.dataframe(hist, use_container_width=True, hide_index=True)
                    break
            st.subheader("Cases table")
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.info("No 'cases' list found in study index. (Older study output?)")

with tab_maturity:
    st.header("Engineering Maturity Heatmap")
    st.caption("Visualize model maturity / validity info embedded in the artifact (model_set + model_registry).")

    art = _get_active_artifact("maturity")
    if not art:
        st.info("Load an artifact to view maturity info.")
    else:
        reg = art.get("model_registry", {})
        ms = art.get("model_set", {})
        rows=[]
        if isinstance(reg, dict):
            entries = reg.get("entries") if isinstance(reg.get("entries"), list) else None
            if entries is None:
                if all(isinstance(v, dict) for v in reg.values()):
                    entries=[{"model_id": k, **v} for k,v in reg.items()]
            if entries:
                selected = set()
                if isinstance(ms, dict):
                    sel = ms.get("selected")
                    if isinstance(sel, dict):
                        selected = set(sel.values()) | set(sel.keys())
                    elif isinstance(sel, list):
                        selected = set(sel)
                for e in entries:
                    mid = e.get("model_id", e.get("id", ""))
                    rows.append({
                        "subsystem": e.get("subsystem", e.get("domain", "")),
                        "model_id": mid,
                        "maturity": e.get("maturity", e.get("maturity_tag", "")),
                        "validity": e.get("validity", e.get("validity_range", "")),
                        "selected": (mid in selected)
                    })
        if rows:
            df=pd.DataFrame(rows)
            st.dataframe(df.sort_values(["subsystem","model_id"]), use_container_width=True, hide_index=True)
            st.markdown("Tip: treat this as a policy gate (e.g., block decisions if maturity < required).")
        else:
            st.info("No model_registry entries found in artifact.")

with tab_assumptions:
    st.header("Assumption Toggle Bar")
    st.caption("Fast scenario exploration by toggling common assumptions and re-evaluating the point (still feasibility-first; no optimization).")

    art = _get_active_artifact("assumptions")
    base = _guess_point_inputs_from_artifact(art) if art else None
    if base is None:
        base = st.session_state.get("last_point_inp")
    if base is None:
        st.info("Load an artifact (or run Point Designer) to use assumption toggles.")
    else:
        col1,col2,col3=st.columns(3)
        with col1:
            fuel = st.selectbox("Fuel mode", ["DT","DD"], index=0 if getattr(base,"fuel_mode","DT")=="DT" else 1, key="ass_fuel")
        with col2:
            ti = st.number_input("Ti (keV)", value=float(base.Ti_keV), step=0.5, key="ass_Ti")
        with col3:
            paux = st.number_input("Paux (MW)", value=float(base.Paux_MW), step=1.0, key="ass_Paux")
        tite = st.number_input("Ti/Te", value=float(getattr(base,"Ti_over_Te", 2.0)), step=0.1, key="ass_TiTe")
        apply = st.button("Apply toggles and evaluate", use_container_width=True, key="ass_run")

        if apply:
            pi = PointInputs(R0_m=float(base.R0_m), a_m=float(base.a_m), kappa=float(base.kappa),
                             Bt_T=float(base.Bt_T), Ip_MA=float(base.Ip_MA), Ti_keV=float(ti),
                             fG=float(base.fG), Paux_MW=float(paux), Ti_over_Te=float(tite),
                             fuel_mode=str(fuel))
            out = hot_ion_point(pi)
            cons = evaluate_constraints(out, point_inputs=pi)
            ok = all((not bool(c.get("failed"))) for c in cons)
            st.metric("Feasible", "YES ✅" if ok else "NO ❌")
            st.subheader("Key outputs")
            st.json({k: out.get(k) for k in ["Q_DT_eqv","P_fus_MW","P_net_MW","betaN","q95","fG"] if k in out})
            st.subheader("Top failed constraints")
            failed=[c for c in cons if c.get("failed")]
            if failed:
                st.dataframe(_safe_df(failed[:10]), use_container_width=True, hide_index=True)
            else:
                st.write("No failed constraints.")

with tab_export:
    st.header("Export / Communication Panel")
    st.caption("One-click export helpers (JSON, CSV, and a one-slide PNG-style summary) with provenance footer.")

    art = _get_active_artifact("export")
    if not art:
        st.info("Load an artifact to export.")
    else:
        _download_json_button("Download run artifact JSON", art, "shams_run_artifact.json", "dl_artifact")
        tables = art.get("tables", {}) if isinstance(art.get("tables"), dict) else {}
        if tables:
            for name, obj in tables.items():
                try:
                    df = _safe_df(obj)
                    st.download_button(f"Download {name}.csv", data=df.to_csv(index=False).encode("utf-8"),
                                       file_name=f"{name}.csv", mime="text/csv", key=f"dl_csv_{name}")
                except Exception:
                    continue
        else:
            st.info("No standardized tables found in artifact ('tables').")

        try:
            import io
            import matplotlib.pyplot as plt
            prov = art.get("provenance", {}) if isinstance(art.get("provenance"), dict) else {}
            d = _decision_summary_from_artifact(art)
            fig = plt.figure(figsize=(10, 5))
            ax = fig.add_subplot(111)
            ax.axis("off")
            title = "SHAMS Decision Summary"
            verdict = "FEASIBLE" if d["feasible"] else "INFEASIBLE"
            ax.text(0.02, 0.92, f"{title} — {verdict}", fontsize=16, weight="bold")
            ax.text(0.02, 0.82, f"Q: {d['kpis'].get('Q_DT_eqv', d['kpis'].get('Q','—'))}    Pfus(MW): {d['kpis'].get('P_fus_MW', d['kpis'].get('Pfus_MW','—'))}", fontsize=12)
            ax.text(0.02, 0.72, "Top blockers:", fontsize=12, weight="bold")
            y=0.66
            for b in (d["top_blockers"] or [])[:6]:
                ax.text(0.04, y, f"- {b.get('group','')}: {b.get('name','')}", fontsize=11)
                y -= 0.06
            footer = f"repo_version={prov.get('repo_version')}  git={prov.get('git_commit')}  python={prov.get('python')}"
            ax.text(0.02, 0.03, footer, fontsize=9)
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
            plt.close(fig)
            st.download_button("Download one-slide summary PNG", data=buf.getvalue(), file_name="shams_one_slide.png",
                               mime="image/png", key="dl_png_slide")
        except Exception as e:
            st.warning(f"PNG summary unavailable: {e}")

with tab_solver:
    st.header("Solver Introspection")
    st.caption("Inspect solver trace/clamp/residual info from artifacts or the last Point Designer run.")

    art = st.session_state.get("selected_artifact")
    if not isinstance(art, dict) or not art:
        st.info("No session artifact loaded. Upload one below to inspect solver annotations.")
        up = st.file_uploader("Upload shams_run_artifact.json", type=["json"], key="solver_up")
        art = _load_json_from_upload(up)

    if art:
        trace = art.get("solver_trace") if isinstance(art.get("solver_trace"), dict) else None
        if trace:
            st.subheader("solver_trace (artifact)")
            st.json(trace)
        else:
            st.subheader("Solver annotations (best-effort)")
            flat = {}
            for k,v in art.items():
                if isinstance(k,str) and (k.startswith("_solver") or k.startswith("_H98") or k.startswith("_Q")):
                    flat[k]=v
            kpis = art.get("kpis", {}) if isinstance(art.get("kpis"), dict) else {}
            for k,v in kpis.items():
                if isinstance(k,str) and (k.startswith("_solver") or k.startswith("_H98") or k.startswith("_Q")):
                    flat[f"kpis.{k}"]=v
            if flat:
                st.json(flat)
            else:
                st.info("No solver trace fields found in artifact.")


# --- Copyright notice
st.markdown('---')
st.caption('© 2026 Afshin Arjhangmehr — SHAMS–FUSION-X')
