"""
Phase-1 Clean Point Design UI (Streamlit)

Professional, single-command UI for:
- Point Designer: evaluate one operating point and show pass/fail constraint dashboard
- Scan Lab: run parameter scans and explore results
- Results Explorer: filter/sort/export feasible points

Design goals:
- No JS toolchain required (runs on pure Python).
- Physics and models live in src/ (imported as a library).
- All models remain explicit proxies (Phase-1), with conservative pass/fail gates.
"""

from __future__ import annotations

import subprocess


# --- Branding (v175.4) ---
APP_NAME = 'Tokamak 0-D Design Studio'
APP_SUBTITLE = 'Feasibility-first, constraint-authoritative 0-D tokamak design'
APP_AUTHOR = 'Afshin Arjhangmehr'
APP_YEAR = 2026
APP_COPYRIGHT = '© 2026 Afshin Arjhangmehr'

def _render_branding_header():
    # Keep the main title crisp and readable across zoom levels.
    st.markdown(f"# {APP_NAME}")
    st.caption(APP_SUBTITLE)

def _render_footer():
    """Render a persistent footer (safe HTML/CSS injection)."""
    st.markdown(
        f"""
        <style>
        .shams-footer {{
            position: fixed;
            bottom: 0;
            left: 0;
            right: 0;
            text-align: center;
            font-size: 12px;
            color: rgba(49,51,63,0.7);
            padding: 6px 0;
            background: rgba(255,255,255,0.92);
            border-top: 1px solid rgba(49,51,63,0.15);
            z-index: 1000;
        }}
        .block-container {{
            padding-bottom: 3rem;
        }}
        </style>
        <div class="shams-footer">{APP_COPYRIGHT}</div>
        """,
        unsafe_allow_html=True,
    )


def _attach_common_metadata(d: dict) -> dict:
    """Attach canonical author/software metadata to an artifact-like dict.

    v264.0 adds governance overlays:
    - citation completeness (for authority overrides)
    - experimental evidence anchoring on constraint records
    """
    if not isinstance(d, dict):
        return d

    d.setdefault("software", APP_NAME)
    d.setdefault("author", APP_AUTHOR)
    d.setdefault("year", APP_YEAR)
    d.setdefault("copyright", APP_COPYRIGHT)

    # Point Designer governance (frozen): embed in every artifact so that
    # downstream exports and bundles always carry the evaluation/exploration boundary.
    d.setdefault(
        "point_designer",
        {
            "status": "frozen",
            "frozen_since": "v179.2",
            "role": "constraint-authoritative 0-D operating point evaluator",
            "note": "No optimization, relaxation, or exploration occurs in Point Designer. Exploration is performed in Systems Mode, which calls Point Designer as a fixed evaluator.",
            "non_goals": [
                "optimization",
                "transport",
                "time evolution",
                "design space exploration",
                "equilibrium solve",
                "SOL / edge code replacement",
                "neutronics Monte-Carlo replacement",
            ],
        },
    )

    # Reproducibility stamp: keep Point Designer artifacts audit-ready.
    # (UI-only; does not affect physics.)
    try:
        from pathlib import Path as _Path
        import platform as _platform
        import sys as _sys
        import datetime as _dt
        _ver_path = _Path(__file__).resolve().parents[1] / "VERSION"
        _ver = "unknown"
        if _ver_path.exists():
            _ver = _ver_path.read_text(encoding="utf-8").strip().splitlines()[0]
        d.setdefault("shams_version", _ver)
        d.setdefault(
            "build_utc",
            _dt.datetime.now(_dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        )
        d.setdefault("python", _sys.version.split("\n")[0])
        d.setdefault("platform", _platform.platform())
    except Exception:
        pass

    # --- Governance overlays (v264.0) ---
    try:
        from src.governance.citations import validate_authority_overrides, summarize_citation_completeness
        issues = validate_authority_overrides(d.get("authority_overrides", {}))
        d["citation_completeness"] = summarize_citation_completeness(issues)
    except Exception:
        pass

    try:
        from pathlib import Path as _Path
        from src.governance.experimental_anchoring import load_anchoring_db, annotate_constraints, summarize_evidence
        dbp = _Path(ROOT) / "benchmarks" / "experimental" / "data" / "anchors_default.json"
        if dbp.exists() and isinstance(d.get("constraints"), list):
            db = load_anchoring_db(dbp)
            d["constraints"] = annotate_constraints(d["constraints"], db)
            d["experimental_evidence_summary"] = summarize_evidence(d["constraints"])
    except Exception:
        pass

    return d
# ---- import path bootstrap (must be before any local imports) ----
import os as _os, sys as _sys
_ROOT = _os.path.abspath(_os.path.join(_os.path.dirname(__file__), ".."))
_SRC = _os.path.join(_ROOT, "src")
if _SRC not in _sys.path:
    _sys.path.insert(0, _SRC)
# ---------------------------------------------------------------

# Expose repo root for downstream helpers
ROOT = _ROOT
SRC = _SRC

import io
import json
import os
import sys
import math

# --- JSON export helpers (cycle-safe, deterministic) ---
def _shams_json_sanitize(obj, _seen=None, _depth: int = 0, _max_depth: int = 25):
    import math
    if _seen is None:
        _seen = set()
    if _depth > _max_depth:
        return "<max_depth>"
    if obj is None or isinstance(obj, (str, int, bool)):
        return obj
    if isinstance(obj, float):
        if not math.isfinite(obj):
            return None
        return obj
    oid = id(obj)
    if oid in _seen:
        return "<circular_ref>"
    if isinstance(obj, dict):
        _seen.add(oid)
        out = {}
        for k, v in obj.items():
            try:
                kk = str(k)
            except Exception:
                kk = "<key>"
            if kk in {"_session_state", "session_state", "__streamlit__", "__ctx__"}:
                continue
            out[kk] = _shams_json_sanitize(v, _seen, _depth + 1, _max_depth)
        _seen.discard(oid)
        return out
    if isinstance(obj, (list, tuple, set)):
        _seen.add(oid)
        out = [_shams_json_sanitize(v, _seen, _depth + 1, _max_depth) for v in list(obj)]
        _seen.discard(oid)
        return out
    try:
        import dataclasses
        if dataclasses.is_dataclass(obj):
            return _shams_json_sanitize(dataclasses.asdict(obj), _seen, _depth + 1, _max_depth)
    except Exception:
        pass
    try:
        return str(obj)
    except Exception:
        return "<unserializable>"

def _shams_json_dumps(obj, **kwargs):
    import json as _json
    safe = _shams_json_sanitize(obj)
    return _json.dumps(safe, **kwargs)

# ----------------------------------------------------

import random
import datetime
import time
import concurrent.futures
import queue
import threading
from dataclasses import asdict, fields, replace
from typing import Dict, Any, List, Tuple

import html

import pandas as pd
import numpy as np
from decision.kpis import headline_kpis

import streamlit as st
from ui.tablekit import install_expandable_tables

# --- Optional plotting dependency (matplotlib) ---
# Matplotlib is intentionally optional in SHAMS. The UI must never crash if it is missing.
try:  # pragma: no cover (availability depends on user environment)
    import matplotlib.pyplot as plt  # type: ignore
    _HAVE_MPL = True
except Exception:  # pragma: no cover
    plt = None  # type: ignore
    _HAVE_MPL = False

# --- Global UI preference: make all tables collapsible to prevent scroll walls
if "ui_tablekit_enabled" not in st.session_state:
    st.session_state["ui_tablekit_enabled"] = True
if "ui_tablekit_default_expanded" not in st.session_state:
    st.session_state["ui_tablekit_default_expanded"] = False

# --- v327.0: Design State Graph (DSG) — inter-panel continuity (exploration layer)
_DSG_SNAPSHOT_PATH = "artifacts/dsg/current_dsg.json"
from typing import Any, Optional
try:
    from src.dsg import DesignStateGraph  # type: ignore
except Exception:
    try:
        from dsg import DesignStateGraph  # type: ignore
    except Exception:
        DesignStateGraph = None  # type: ignore

if "_shams_dsg" not in st.session_state and DesignStateGraph is not None:
    try:
        st.session_state["_shams_dsg"] = DesignStateGraph.load(_DSG_SNAPSHOT_PATH)
    except Exception:
        st.session_state["_shams_dsg"] = DesignStateGraph()

def _dsg_save_best_effort() -> None:
    try:
        if DesignStateGraph is None:
            return
        g = st.session_state.get("_shams_dsg")
        if g is None:
            return
        g.save(_DSG_SNAPSHOT_PATH)
    except Exception:
        return

def _dsg_record_best_effort(*, inp: Any, res: Any, origin: str, parents: Optional[list[str]] = None, tags: Optional[list[str]] = None, edge_kind: Optional[str] = None, edge_note: str = "") -> None:
    """Record an evaluation into the DSG if available.

    This is an exploration-layer side effect only; it must not affect truth.
    """
    try:
        if DesignStateGraph is None:
            return
        g = st.session_state.get("_shams_dsg")
        if g is None:
            return
        out = getattr(res, "out", {}) if res is not None else {}
        ok = bool(getattr(res, "ok", True))
        msg = str(getattr(res, "message", "") or "")
        el = float(getattr(res, "elapsed_s", 0.0) or 0.0)
        node = g.record(inp=inp, out=out, ok=ok, message=msg, elapsed_s=el, origin=str(origin), parents=parents, tags=tags, edge_kind=edge_kind, edge_note=edge_note)
        st.session_state["active_design_node_id"] = node.node_id
        _dsg_save_best_effort()
    except Exception:
        return

def _dsg_evaluator(*, origin: str = "UI", **evaluator_kwargs: Any):
    """Construct an Evaluator wrapper that records evaluations into DSG.

    This is an exploration-layer adapter only.
    """
    try:
        from src.evaluator.core import Evaluator  # type: ignore
    except Exception:
        from evaluator.core import Evaluator  # type: ignore

    # Evaluator construction can be expensive. Cache the base evaluator instance deterministically
    # so UI reruns (tab switches, widget edits) do not rebuild heavy objects.
    @st.cache_resource(show_spinner=False)
    def _cached_evaluator(**kwargs: Any):
        return Evaluator(**kwargs)

    ev = _cached_evaluator(**evaluator_kwargs)

    class _Wrap:
        def __init__(self, _ev: Any, _origin: str):
            self._ev = _ev
            self._origin = str(_origin)

        def evaluate(self, inp: Any, Paux_for_Q_MW: Optional[float] = None):
            res = self._ev.evaluate(inp, Paux_for_Q_MW=Paux_for_Q_MW)
            parent = st.session_state.get("dsg_selected_node_id") or st.session_state.get("active_design_node_id")
            kind = st.session_state.get("dsg_context_edge_kind") or "derived"
            parents = [str(parent)] if parent else None
            _dsg_record_best_effort(inp=inp, res=res, origin=self._origin, parents=parents, edge_kind=str(kind))
            return res

        def get(self, *args: Any, **kwargs: Any):
            return self._ev.get(*args, **kwargs)

        def cache_stats(self):
            return self._ev.cache_stats()

        def reset_cache_stats(self):
            return self._ev.reset_cache_stats()

    return _Wrap(ev, origin)


def _ui_evaluate(
    inp: Any,
    *,
    origin: str = "UI",
    Paux_for_Q_MW: Optional[float] = None,
    **evaluator_kwargs: Any,
) -> Dict[str, Any]:
    """Route UI point evaluation through the Evaluator choke point (PROPOSAL-008)."""
    return _dsg_evaluator(origin=origin, **evaluator_kwargs).evaluate(inp, Paux_for_Q_MW=Paux_for_Q_MW).out


install_expandable_tables(st)

from ui.icons import label as ui_label, render_mode_scope
from ui.dsg_panel import render_dsg_sidebar
# ---- SHAMS UX guardrails: global run lock + fusion-expert notifications
from ui import runlock as _shams_runlock
import atexit as _shams_atexit
import time as _shams_time
import uuid as _shams_uuid
import re as _shams_re

if "_shams_app_start_ts" not in st.session_state:
    st.session_state["_shams_app_start_ts"] = _shams_time.time()
if "_shams_owner_token" not in st.session_state:
    st.session_state["_shams_owner_token"] = str(_shams_uuid.uuid4())

def _shams_lock_banner():
    locked, task, started, is_owner = _shams_runlock.status(st.session_state.get("_shams_owner_token"), app_start_ts=st.session_state.get("_shams_app_start_ts"))
    if locked and task:
        age_s = int(_shams_time.time() - float(started or _shams_time.time()))
        badge = "⚡ Shot in Progress" if not is_owner else "⚡ Running Sequence"
        st.sidebar.info(f"{badge}: **{task}**  ·  t+{age_s}s")
    return locked, task, started, is_owner

def _shams_is_solver_label(label: str) -> bool:
    # Freeze only *solver* actions, not navigation / downloads.
    return bool(_shams_re.search(r"\b(evaluate|solve|run|build|compute|scan|pareto|search|atlas|trajectory|candidates)\b", label, flags=_shams_re.I))

# Monkeypatch st.button to enforce lock on solver actions
_shams__orig_button = st.button
def _shams_button(label, *args, **kwargs):
    locked, task, started, is_owner = _shams_runlock.status(st.session_state.get("_shams_owner_token"), app_start_ts=st.session_state.get("_shams_app_start_ts"))
    label_str = str(label)
    if locked and _shams_is_solver_label(label_str) and not is_owner:
        kwargs["disabled"] = True
        return _shams__orig_button(label, *args, **kwargs)

    clicked = _shams__orig_button(label, *args, **kwargs)
    if clicked and _shams_is_solver_label(label_str):
        # Do NOT acquire the global run-lock here.
        # Streamlit reruns can cause a deadlock if acquisition happens before the solver call stack.
        locked2, task2, started2, is_owner2 = _shams_runlock.status(
            st.session_state.get("_shams_owner_token"),
            app_start_ts=st.session_state.get("_shams_app_start_ts"),
        )
        if locked2 and (not is_owner2):
            try:
                st.toast("⛔ Another sequence is already running. Wait for the Black‑Box Chronicle to clear.", icon="⛔")
            except Exception:
                pass
            return False
    return clicked

st.button = _shams_button

# Show lock banner early (Control Ledger will inherit this)
_shams_lock_banner()

# --- v327.1: DSG selector + lineage breadcrumb (sidebar)
try:
    from ui.dsg_panel import render_dsg_sidebar  # type: ignore
    _g = st.session_state.get("_shams_dsg")
    if _g is not None:
        render_dsg_sidebar(_g)
        try:
            _sel = st.session_state.get("dsg_selected_node_id") or getattr(_g, "active_node_id", None)
            if _sel:
                st.caption(f"🧬 Active design node: `{_sel}`  ·  lineage edge kind: `{st.session_state.get('dsg_context_edge_kind','derived')}`")
        except Exception:
            pass
except Exception:
    pass


# ---- Systems solver adapter imports (v178.9) ----
# UI directly builds SolverRequest objects for Systems Mode solves.
# Some deployments expose `src/` modules as top-level packages; keep a robust fallback.
try:
    from solvers import SolverRequest, DefaultTargetSolverBackend, solve_request  # type: ignore
except Exception:  # pragma: no cover
    from src.solvers import SolverRequest, DefaultTargetSolverBackend, solve_request  # type: ignore

from pathlib import Path
from tools.activity_log import get_logger as _get_activity_logger

# Global repo root + activity logger (Asia/Tehran)
REPO_ROOT = Path(__file__).resolve().parent.parent

# Canonical base directory for UI-resolved repo paths.
# Some panels historically referenced BASE_DIR; keep it stable and deterministic.
BASE_DIR = REPO_ROOT

def _activity_logger():
    lg = _get_activity_logger(st, REPO_ROOT, tz_name="Asia/Tehran")
    # One-time marker so users can confirm logging is working
    if "activity_log_inited" not in st.session_state:
        st.session_state["activity_log_inited"] = True
        try:
            if bool(st.session_state.get("activity_log_auto", True)):
                lg.log_event("UI", "LogInitialized", {"tz": "Asia/Tehran"})
        except Exception:
            pass
    return lg

def _alog(mode: str, action: str, payload: dict | None = None):
    try:
        if bool(st.session_state.get("activity_log_auto", True)):
            _activity_logger().log_event(mode, action, payload or {})
    except Exception:
        pass



def _invalidate_mode_caches(reason: str = "") -> None:
    """Clear cached outputs/artifacts across modes when a truth-relevant UI contract changes.

    This prevents stale Telemetry (e.g., NaN KPIs) after changing intent/machine type/policy.
    Deterministic: only clears UI caches; never changes evaluator physics.
    """
    try:
        keys = [
            # Point Designer
            "pd_last_outputs","pd_last_artifact","pd_last_log_lines","pd_last_run_ts","pd_last_inputs_hash",
            "pd_last_forensics","pd_last_forensics_inputs_hash",
            "last_point_out","last_point_inp","last_solver_log",
            # Systems Mode
            "systems_last_solve_artifact","systems_last_outputs","systems_last_inputs_hash",
            "systems_targets","systems_variables","systems_last_log_lines",
            # Scan Lab / Cartography
            "scan_cartography_artifact","scan_last_run","scan_last_outputs",
            # Pareto / Optimizer
            "opt_last_run_id","opt_last_meta","opt_last_records","opt_last_best",
            # Publication Benchmarks UI caches
            "pb_last_pack","pb_last_delta","pb_last_topology",
            # Compare
            "cmp_slot_A","cmp_slot_B","cmp_slot_A_meta","cmp_slot_B_meta",
        ]
        for k in keys:
            st.session_state.pop(k, None)
        if reason:
            st.session_state["ui_last_invalidation_reason"] = str(reason)
    except Exception:
        pass



def _alog_exc(mode: str, action: str, exc: BaseException):
    try:
        if bool(st.session_state.get("activity_log_auto", True)):
            _activity_logger().log_exception(mode, action, exc)
    except Exception:
        pass


# --- Forward-definition bootstrap (v175.6) ---
# --- Forward-definition bootstrap (v175.6.2) ---
def _bootstrap_forward_defs(target_names=None) -> None:
    """Predefine later top-level defs so early UI can reference them.

    Streamlit executes this script top-to-bottom. Some panels are defined later
    in this file but referenced earlier (e.g. by Panel Availability Map).
    This helper AST-parses *this* file and execs top-level defs (functions/classes)
    into the current globals() *without* executing their bodies.

    If `target_names` is provided, only those defs are loaded.
    """
    try:
        import ast
        from pathlib import Path

        src = Path(__file__).read_text(encoding="utf-8")
        mod = ast.parse(src)
        g = globals()

        targets = set(target_names) if target_names else None

        def _allow(name: str) -> bool:
            if not name:
                return False
            if targets is not None and name not in targets:
                return False
            # Only bootstrap panels and key helpers; avoid random internals.
            if name.startswith("_v"):
                return True
            if name in {"_render_with_contract", "_resolve_panel_function"}:
                return True
            return False

        for node in mod.body:
            if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
                continue
            name = getattr(node, "name", None)
            if not _allow(name):
                continue
            if name in g:
                continue
            try:
                code = compile(ast.Module([node], type_ignores=[]), filename=str(__file__), mode="exec")
                exec(code, g, g)
            except Exception:
                # Best-effort; skip anything that can't be safely defined.
                continue
    except Exception:
        return

_bootstrap_forward_defs()

from ui.panel_availability import PanelState, PanelStatus, default_status
from ui.panel_contracts import get_panel_contracts
from ui.state import SessionStateModel


# =====================
# Phase-1 UI Stabilization: early-safe state helpers (v372.8)
# =====================
# Rationale: prevent forward-reference failures and tab-scope leakage under Streamlit rerun semantics.
# These helpers are intentionally minimal and deterministic; they do not modify physics truth.

def _v92_state_get():
    import streamlit as st
    st.session_state.setdefault("shams_state", SessionStateModel())
    s = st.session_state["shams_state"]
    if getattr(s, "run_history", None) is None:
        s.run_history = []
    if getattr(s, "pinned_run_ids", None) is None:
        s.pinned_run_ids = []
    return s

def _v92_state_clear_point():
    import streamlit as st
    s = _v92_state_get()
    s.last_point_inputs = None
    s.last_point_outputs = None
    s.last_point_artifact = None
    s.last_point_radial_png = None
    for k in ["pd_last_outputs", "pd_last_artifact", "pd_last_radial_png_bytes"]:
        if k in st.session_state:
            del st.session_state[k]

def _phase1_stabilize_cache_aliases() -> None:
    """Maintain canonical cache keys while preserving backward-compat keys.

    Canonical (Phase-1 contract):
      - pd_last_outputs
      - systems_last_solution
      - scan_last_grid
      - pareto_last_front
    """
    import streamlit as st
    ss = st.session_state

    # Systems
    if "systems_last_solution" not in ss and "last_systems_solution" in ss:
        ss["systems_last_solution"] = ss.get("last_systems_solution")

    # Scan
    if "scan_last_grid" not in ss:
        if "scan_last_outputs" in ss:
            ss["scan_last_grid"] = ss.get("scan_last_outputs")
        elif "scan_cartography_artifact" in ss:
            ss["scan_last_grid"] = ss.get("scan_cartography_artifact")

    # Pareto
    if "pareto_last_front" not in ss and "pareto_last" in ss:
        ss["pareto_last_front"] = ss.get("pareto_last")

    # Point Designer: keep legacy aliases alive
    if "pd_last_outputs" in ss and "last_point_out" not in ss:
        ss["last_point_out"] = ss.get("pd_last_outputs")
    if "pd_last_artifact" in ss and "last_point_artifact" not in ss:
        ss["last_point_artifact"] = ss.get("pd_last_artifact")


from ui.pareto_language import PARETO_LOCK_LINE, PARETO_OPTIMAL_DEF, TRUST_BOUNDARIES, FREEZE_STAMP
from ui.optimizer_console import render_external_optimizer_launcher, render_optimizer_evidence_packs
from ui.language_freeze import CANON as _LANG, FORBIDDEN_PHRASES as _FORBIDDEN
from ui.decks.point_designer_hooks import (
    render_point_designer_hero,
    render_point_designer_trace,
    render_point_designer_export,
    render_point_designer_constraint_diff,
    render_point_designer_no_solution_atlas,
)
from ui.decks.system_suite_hooks import render_system_suite_header
from ui.authority_dashboard import render_overlay_authority_dashboard, merge_overlay_session_into_inputs
from ui.session_api import set_point_evaluation

try:
    from schema.governance_presets import apply_governance_preset, tritium_tight_closure_default
except ImportError:
    from src.schema.governance_presets import apply_governance_preset, tritium_tight_closure_default

# --- UI helpers (v87, additive) ---

# --- Panel Availability Map helpers (v174, additive) ---
def _dedupe_checks(checks_list):
    """Remove duplicate checks while preserving order."""
    seen = set()
    out = []
    for c in (checks_list or []):
        key = None
        try:
            if isinstance(c, dict):
                key = c.get("id") or c.get("name") or c.get("title")
            else:
                key = getattr(c, "id", None) or getattr(c, "name", None) or getattr(c, "title", None)
        except Exception:
            key = None
        if key is None:
            key = repr(c)
        if key in seen:
            continue
        seen.add(key)
        out.append(c)
    return out

_PANEL_FN_CACHE = {}

def _resolve_panel_function(fn_name: str):
    """Resolve a panel function by name across the codebase.
    Streamlit runs ui/app.py as __main__, so we must search broadly.
    Uses a cache to avoid repeated module walks.
    """
    if not fn_name:
        return None
    if fn_name in _PANEL_FN_CACHE:
        return _PANEL_FN_CACHE[fn_name]

    # Fast paths
    try:
        cand = globals().get(fn_name)
        if callable(cand):
            _PANEL_FN_CACHE[fn_name] = cand
            return cand
    except Exception:
        pass

    import sys
    try:
        main_mod = sys.modules.get("__main__")
        cand = getattr(main_mod, fn_name, None) if main_mod is not None else None
        if callable(cand):
            _PANEL_FN_CACHE[fn_name] = cand
            return cand
    except Exception:
        pass

    
    # If still not found, try bootstrapping the specific def from later in this file.
    try:
        _bootstrap_forward_defs([fn_name])
        cand = globals().get(fn_name)
        if callable(cand):
            _PANEL_FN_CACHE[fn_name] = cand
            return cand
    except Exception:
        pass

# Robust path: import-walk ui.* modules and search for attribute
    try:
        import ui as _ui_pkg
        import pkgutil as _pkgutil
        import importlib as _importlib
        for m in _pkgutil.walk_packages(_ui_pkg.__path__, _ui_pkg.__name__ + "."):
            modname = m.name
            try:
                mod = _importlib.import_module(modname)
            except Exception:
                continue
            try:
                cand = getattr(mod, fn_name, None)
                if callable(cand):
                    _PANEL_FN_CACHE[fn_name] = cand
                    return cand
            except Exception:
                continue
    except Exception:
        pass

    _PANEL_FN_CACHE[fn_name] = None
    return None


def _render_panel_status_card(status: PanelStatus):
    import streamlit as st
    if status.state == PanelState.AVAILABLE:
        return
    if status.state == PanelState.NOT_GENERATED:
        st.info(status.message)
        if status.missing:
            st.write("Missing artifacts:")
            st.code("\n".join(status.missing))
        return
    if status.state == PanelState.BLOCKED:
        st.warning(status.message)
        return
    if status.state == PanelState.NOT_APPLICABLE:
        st.caption(status.message)
        return
    if status.state == PanelState.DEMO_SUBSTITUTED:
        st.caption(status.message)
        return

def _render_with_contract(panel_fn_name: str, panel_callable):
    import streamlit as st
    contracts = get_panel_contracts()
    c = contracts.get(panel_fn_name)
    if c is None:
        # No contract registered: still render, but never allow silent emptiness
        try:
            panel_callable()
        except Exception as e:
            st.warning(f"Panel error: {e}")
        return
    from ui.panel_availability import evaluate_contract
    status = evaluate_contract(c, st.session_state)
    if status.state != PanelState.AVAILABLE:
        _render_panel_status_card(status)
        return
    try:
        panel_callable()
    except Exception as e:
        st.warning(f"Panel error: {e}")



def _v175_panel_availability_map_panel():
    import streamlit as st
    from ui.panel_contracts import get_panel_contracts
    from ui.panel_availability import evaluate_contract, PanelState

    st.subheader("Panel Availability Map")
    st.caption("Self-explaining UI: every panel reports whether it is available, missing required artifacts, blocked, or not applicable.")

    contracts = get_panel_contracts()
    rows = []
    counts = {s.name: 0 for s in PanelState}

    for fn_name, c in contracts.items():
        status = evaluate_contract(c, st.session_state)
        counts[status.state.name] = counts.get(status.state.name, 0) + 1
        rows.append({
            "panel_fn": fn_name,
            "title": c.title,
            "state": status.state.name,
            "missing_required": ", ".join(status.missing or []),
        })

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Available", counts.get("AVAILABLE", 0))
    c2.metric("Not generated", counts.get("NOT_GENERATED", 0))
    c3.metric("Blocked", counts.get("BLOCKED", 0))
    c4.metric("Not applicable", counts.get("NOT_APPLICABLE", 0))
    c5.metric("Demo-substituted", counts.get("DEMO_SUBSTITUTED", 0))

    states = ["ALL"] + sorted({r["state"] for r in rows})
    pick = st.selectbox("Filter by state", states, index=0, key="v175_pam_filter")
    show = [r for r in rows if pick == "ALL" or r["state"] == pick]

    st.write("Panels:")
    st.dataframe(show, use_container_width=True, hide_index=True)

    st.divider()
    st.markdown("### Open a panel here")
    opts = ["(select)"] + [f'{r["title"]}  -  {r["panel_fn"]}' for r in rows]
    sel = st.selectbox("Choose panel", opts, index=0, key="v175_pam_open_pick")
    if sel != "(select)":
        fn_name = sel.rsplit("-", 1)[-1].strip()
        st.session_state["v175_pam_focus"] = fn_name

    focus = st.session_state.get("v175_pam_focus")
    if focus:
        st.markdown(f"#### Focus: `{focus}`")
        fn = _resolve_panel_function(focus)
        if callable(fn):
            _render_with_contract(focus, fn)
        else:
            st.warning("Selected panel function not found in this UI build.")

def _render_provenance_sidebar():
    with st.sidebar:
        st.subheader("Run Provenance")
        try:
            v = (BASE_DIR / "VERSION").read_text().strip()
        except Exception:
            v = "unknown"
        st.code(v)
        st.markdown("---")
        st.subheader("Session")
        _exit_confirm = st.checkbox("Confirm exit", value=False, key="shams_exit_confirm")
        if st.button("🔴 Exit SHAMS", type="primary", use_container_width=True, disabled=not _exit_confirm, key="shams_exit_btn"):
            st.info("SHAMS UI shutdown requested by user.")
            # Hard-exit is the only reliable cross-platform Streamlit shutdown mechanism.
            _os._exit(0)
        st.caption("Authoritative feasibility lives in SHAMS core. Sandbox results are non-authoritative.")

def _feasibility_narrative(point):
    feas = point.get("feasible", False)
    mins = point.get("min_signed_margin", None)
    acts = point.get("active_constraints", [])
    if feas:
        return f"Feasible. Min margin = {mins:.3g}." if isinstance(mins, (int,float)) else "Feasible."
    return f"Infeasible. Limiting constraints: {', '.join(acts[:3]) if acts else 'unknown'}."

def _margin_waterfall(records):
    # records: list of dicts with name + signed_margin
    import pandas as pd
    rows = []
    for r in records:
        name = r.get("name","")
        sm = r.get("signed_margin", None)
        if name and isinstance(sm,(int,float)):
            rows.append({"constraint": name, "signed_margin": sm})
    if not rows:
        st.info("No constraint margins available.")
        return
    df = pd.DataFrame(rows).sort_values("signed_margin")
    st.bar_chart(df.set_index("constraint"))


# ---------------------------------------------------------------------------
# Session-state initialization (prevents AttributeError on first run)
# ---------------------------------------------------------------------------
def _init_session_state() -> None:

    # NOTE: phase1_core import happens later in this file. We defensively import
    # PointInputs here to avoid NameError during early session-state init.
    try:
        from phase1_core import PointInputs as _PointInputs
    except Exception:
        _PointInputs = None

    defaults = {
        # If PointInputs is not available for any reason, fall back to None and let
        # downstream code initialize it in a controlled path.
        "last_point_inp": (_PointInputs(R0_m=1.85, a_m=0.57, kappa=1.8, Bt_T=12.2, Ip_MA=8.0, Ti_keV=15.0, fG=0.8, Paux_MW=20.0)
                           if _PointInputs is not None else None),
        "last_point_out": None,
        "last_solver_log": None,
        "explain_mode": True,
        "expert_mode": False,
        "design_intent": "Power Reactor (net-electric)",
        "scan_df": None,
        "scan_meta": None,
        "scan_log_lines": [],
        "scan_log_text": "",
        "scan_progress": 0.0,
        "scan_queue": [],
        "scan_running": False,
        "scan_future": None,
        "scan_executor": None,

        # Phase 7+ UI persistence
        "de_best_inputs": None,
        "de_best_out": None,
        "de_history": None,
        "robustness_result": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_session_state()
# _sync_point_designer_from_last_point_inp()


# Stable Streamlit keys for Point Designer widgets (presets rely on these)
PD_KEYS = {
    "R0_m": "pd_R0_m",
    "a_m": "pd_a_m",
    "kappa": "pd_kappa",
    "delta": "pd_delta",
    "Bt_T": "pd_Bt_T",
    "Ti_keV": "pd_Ti_keV",
    "Paux_MW": "pd_Paux_MW",
    "H98_tgt": "pd_target_H98",
    "Q_tgt": "pd_target_Q",
    "Ip_lo": "pd_Ip_lo",
    "Ip_hi": "pd_Ip_hi",
    "fG_lo": "pd_fG_lo",
    "fG_hi": "pd_fG_hi",
    # Additional stable keys (used for preset propagation)
    "Ti_over_Te": "pd_Ti_over_Te",
    "Paux_for_Q": "pd_Paux_for_Q",
    "magnet_technology": "pd_magnet_technology",
    "Tcoil_K": "pd_Tcoil_K",
    "include_magnet_technology_authority_v400": "pd_include_magnet_technology_authority_v400",
    "magnet_margin_min_v400": "pd_magnet_margin_min_v400",
    "b_margin_min_v400": "pd_b_margin_min_v400",
    "j_margin_min_v400": "pd_j_margin_min_v400",
    "stress_margin_min_v400": "pd_stress_margin_min_v400",
    "sc_margin_min_v400": "pd_sc_margin_min_v400",
    "t_margin_min_v400": "pd_t_margin_min_v400",
    "p_tf_ohmic_margin_min_v400": "pd_p_tf_ohmic_margin_min_v400",


    # v318.0 profile bundle knobs (stable keys for presets and promotion)
    "profile_mode": "pd_profile_mode",
    "profile_alpha_T": "pd_profile_alpha_T",
    "profile_alpha_n": "pd_profile_alpha_n",
    "profile_shear_shape": "pd_profile_shear_shape",
    "pedestal_enabled": "pd_pedestal_enabled",
    "pedestal_width_a": "pd_pedestal_width_a",
    "include_bootstrap_pressure_selfconsistency": "pd_include_bootstrap_pressure_selfconsistency",
    "f_bootstrap_consistency_abs_max": "pd_f_bootstrap_consistency_abs_max",

    # v371.0 transport contract library keys
    "include_transport_contracts_v371": "pd_include_transport_contracts_v371",
    "H_required_max_optimistic": "pd_H_required_max_optimistic",
    "H_required_max_robust": "pd_H_required_max_robust",

    # v396.0 transport envelope 2.0 keys
    "include_transport_envelope_v396": "pd_include_transport_envelope_v396",
    "transport_spread_max_v396": "pd_transport_spread_max_v396",
    "include_tauE_user_scaling_v396": "pd_include_tauE_user_scaling_v396",
    "tauE_user_C_v396": "pd_tauE_user_C_v396",
    "tauE_user_exp_Ip_v396": "pd_tauE_user_exp_Ip_v396",
    "tauE_user_exp_Bt_v396": "pd_tauE_user_exp_Bt_v396",
    "tauE_user_exp_ne_v396": "pd_tauE_user_exp_ne_v396",
    "tauE_user_exp_Ploss_v396": "pd_tauE_user_exp_Ploss_v396",
    "tauE_user_exp_R_v396": "pd_tauE_user_exp_R_v396",
    "tauE_user_exp_eps_v396": "pd_tauE_user_exp_eps_v396",
    "tauE_user_exp_kappa_v396": "pd_tauE_user_exp_kappa_v396",
    "tauE_user_exp_M_v396": "pd_tauE_user_exp_M_v396",

    # v397.0 profile proxy authority keys
    "include_profile_proxy_v397": "pd_include_profile_proxy_v397",
    "profile_alpha_T_v397": "pd_profile_alpha_T_v397",
    "profile_beta_T_v397": "pd_profile_beta_T_v397",
    "profile_alpha_n_v397": "pd_profile_alpha_n_v397",
    "profile_beta_n_v397": "pd_profile_beta_n_v397",
    "profile_alpha_j_v397": "pd_profile_alpha_j_v397",
    "profile_beta_j_v397": "pd_profile_beta_j_v397",
    "profile_shear_shape_v397": "pd_profile_shear_shape_v397",
    "profile_peaking_p_max_v397": "pd_profile_peaking_p_max_v397",
    "q95_proxy_min_v397": "pd_q95_proxy_min_v397",
    "q0_proxy_min_v397": "pd_q0_proxy_min_v397",
    "bootstrap_localization_max_v397": "pd_bootstrap_localization_max_v397",
    # v372.0 neutronics–materials coupling keys
    "include_neutronics_materials_coupling_v372": "pd_include_nm_coupling_v372",
    "nm_material_class_v372": "pd_nm_material_class_v372",
    "nm_spectrum_class_v372": "pd_nm_spectrum_class_v372",
    "nm_T_oper_C_v372": "pd_nm_T_oper_C_v372",
    "dpa_rate_eff_max_v372": "pd_dpa_rate_eff_max_v372",
    "damage_margin_min_v372": "pd_damage_margin_min_v372",
}


def _push_point_inputs_to_pd_widget_keys(base: Any) -> None:
    """Push a PointInputs-like object into Point Designer widget keys.

    Streamlit evaluates the script top-to-bottom on each rerun.
    If a preset is loaded via a button click, the preset application happens
    *after* the early one-shot sync. To avoid requiring a second click, we
    provide a direct, deterministic propagation routine that can be called
    immediately from the button handler.

    UI-only: this function only mutates widget/session state.
    """
    if base is None:
        return
    try:
        st.session_state[PD_KEYS["R0_m"]] = float(getattr(base, "R0_m"))
        st.session_state[PD_KEYS["a_m"]] = float(getattr(base, "a_m"))
        st.session_state[PD_KEYS["kappa"]] = float(getattr(base, "kappa"))
        st.session_state[PD_KEYS["delta"]] = float(getattr(base, "delta", 0.0) or 0.0)
        st.session_state[PD_KEYS["Bt_T"]] = float(getattr(base, "Bt_T"))
        st.session_state[PD_KEYS["Ti_keV"]] = float(getattr(base, "Ti_keV"))
        st.session_state[PD_KEYS["Paux_MW"]] = float(getattr(base, "Paux_MW"))
    except Exception:
        pass
    # Bounds from preset Ip/fG
    try:
        ip = float(getattr(base, "Ip_MA"))
        st.session_state[PD_KEYS["Ip_lo"]] = max(0.1, 0.80 * ip)
        st.session_state[PD_KEYS["Ip_hi"]] = max(0.2, 1.20 * ip)
    except Exception:
        pass
    try:
        fg = float(getattr(base, "fG"))
        st.session_state[PD_KEYS["fG_lo"]] = max(0.0, fg - 0.20)
        st.session_state[PD_KEYS["fG_hi"]] = min(2.0, fg + 0.20)
    except Exception:
        pass
    # Aux/Q denominator and Ti/Te
    try:
        st.session_state[PD_KEYS["Paux_for_Q"]] = float(getattr(base, "Paux_MW"))
    except Exception:
        pass
    try:
        st.session_state[PD_KEYS["Ti_over_Te"]] = float(getattr(base, "Ti_over_Te", 1.0))
    except Exception:
        pass

    # v371.0 transport contracts
    try:
        st.session_state[PD_KEYS["include_transport_contracts_v371"]] = bool(getattr(base, "include_transport_contracts_v371", False))
        st.session_state[PD_KEYS["H_required_max_optimistic"]] = float(getattr(base, "H_required_max_optimistic", float("nan")))
        st.session_state[PD_KEYS["H_required_max_robust"]] = float(getattr(base, "H_required_max_robust", float("nan")))
    except Exception:
        pass

    # Magnet technology axis (optional in older presets)
    try:
        st.session_state[PD_KEYS["magnet_technology"]] = str(getattr(base, "magnet_technology", "HTS_REBCO") or "HTS_REBCO")
    except Exception:
        pass

    # v318.0: profile bundle knobs (optional in older presets)
    try:
        st.session_state[PD_KEYS["profile_mode"]] = bool(getattr(base, "profile_mode", False))
        st.session_state[PD_KEYS["profile_alpha_T"]] = float(getattr(base, "profile_alpha_T", 1.5))
        st.session_state[PD_KEYS["profile_alpha_n"]] = float(getattr(base, "profile_alpha_n", 1.0))
        st.session_state[PD_KEYS["profile_shear_shape"]] = float(getattr(base, "profile_shear_shape", 0.5))
        st.session_state[PD_KEYS["pedestal_enabled"]] = bool(getattr(base, "pedestal_enabled", False))
        st.session_state[PD_KEYS["pedestal_width_a"]] = float(getattr(base, "pedestal_width_a", 0.05))
        st.session_state[PD_KEYS["include_bootstrap_pressure_selfconsistency"]] = bool(getattr(base, "include_bootstrap_pressure_selfconsistency", False))
        st.session_state[PD_KEYS["f_bootstrap_consistency_abs_max"]] = float(getattr(base, "f_bootstrap_consistency_abs_max", float("nan")))
    except Exception:
        pass
    try:
        st.session_state[PD_KEYS["Tcoil_K"]] = float(getattr(base, "Tcoil_K", 20.0))
    except Exception:
        pass


def apply_reference_preset(ref_key: str) -> None:
    """Load a catalog reference preset into the workspace.

    This is a *UI state* operation:
      - Sets ``last_point_inp`` to the preset PointInputs
      - Sets design intent (reactor vs research) from the catalog metadata
      - Arms a one-shot sync to propagate values into Point Designer widget keys
    """
    cat = reference_catalog()
    if ref_key not in cat:
        raise KeyError(f"Unknown reference preset: {ref_key}")
    ent = cat[ref_key]
    base = ent.get("inputs")
    if base is None:
        raise ValueError(f"Reference preset missing inputs: {ref_key}")

    # Set design intent from catalog
    intent = str(ent.get("intent", "")).strip().lower()
    if intent.startswith("research"):
        st.session_state["design_intent"] = "Experimental Device (research)"
    elif intent:
        st.session_state["design_intent"] = "Power Reactor (net-electric)"

    # Persist as the active workspace input object
    st.session_state["last_point_inp"] = base

    # Immediate propagation into Point Designer widget keys so the Configure
    # panel updates on the same click (no second click required).
    _push_point_inputs_to_pd_widget_keys(base)
    st.session_state["pd_needs_sync"] = False


def apply_legacy_reference_machine(name: str) -> None:
    """Load legacy REFERENCE_MACHINES entry (dict-like) into the workspace."""
    if name not in REFERENCE_MACHINES:
        raise KeyError(f"Unknown legacy preset: {name}")
    d = dict(REFERENCE_MACHINES[name] or {})
    try:
        base = make_point_inputs(**d)
    except Exception as e:
        raise ValueError(f"Legacy preset could not be converted to PointInputs: {name} ({e})")
    st.session_state["last_point_inp"] = base
    _push_point_inputs_to_pd_widget_keys(base)
    st.session_state["pd_needs_sync"] = False


def _compute_run_summary_from_out(out: Dict[str, Any]) -> Dict[str, Any]:
    """Compute a deterministic, UI-safe run summary from outputs.

    Important: this must not depend on which Telemetry deck is visible.
    """
    try:
        cons = evaluate_constraints(out or {})
    except Exception:
        cons = []
    # Hard constraints only
    hard = []
    for c in (cons or []):
        try:
            if str(getattr(c, "severity", "hard")).strip().lower() == "hard":
                hard.append(c)
        except Exception:
            pass
    # Sort worst-first (most negative margin)
    try:
        hard_sorted = sorted(hard, key=lambda c: float(getattr(c, "margin", float("inf"))))
    except Exception:
        hard_sorted = list(hard)

    tight: List[Dict[str, Any]] = []
    for c in hard_sorted[:8]:
        try:
            tight.append(
                {
                    "name": str(getattr(c, "name", "")),
                    "passed": bool(getattr(c, "passed", False)),
                    "margin_frac": float(getattr(c, "margin", float("nan"))),
                    "value": getattr(c, "value", None),
                    "limit": getattr(c, "limit", None),
                    "units": getattr(c, "units", None),
                    "sense": getattr(c, "sense", None),
                    "group": getattr(c, "group", "general"),
                }
            )
        except Exception:
            pass

    # Power closure diagnostic
    closure = float("nan")
    try:
        pin = float(out.get("Pin_MW", float("nan")))
        ploss = float(out.get("Ploss_MW", float("nan")))
        if np.isfinite(pin) and np.isfinite(ploss):
            closure = pin - ploss
    except Exception:
        pass

    return {
        "headline": {
            "Q_DT_eqv": float(out.get("Q_DT_eqv", float("nan"))) if isinstance(out, dict) else float("nan"),
            "H98": float(out.get("H98", float("nan"))) if isinstance(out, dict) else float("nan"),
            "P_net_e_MW": float(out.get("P_net_e_MW", float("nan"))) if isinstance(out, dict) else float("nan"),
        },
        "power_closure_MW": closure,
        "tightest_hard_constraints": tight,
    }
# One-shot synchronization: when a preset is loaded we set this flag, and on the next
# rerun we push preset values into Point Designer widget keys.
def _render_magnet_authority_panel(out: Dict[str, Any]) -> None:
    """Render Magnet Technology Authority panel (v328.0).

    UI-only: reads frozen truth outputs and displays contract-driven limits and margins.
    """
    if not isinstance(out, dict) or not out:
        return

    with st.expander("🧲 Magnet Authority — Technology Regime (v328.0)", expanded=False):
        regime = str(out.get("magnet_regime", "UNKNOWN"))
        tech = str(out.get("magnet_technology", ""))
        contract_sha = str(out.get("magnet_contract_sha256", ""))[:12]
        cls = str(out.get("magnet_fragility_class", "UNKNOWN"))
        mmin = out.get("magnet_margin_min", float("nan"))

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("Regime", regime)
        with c2:
            st.metric("Tech string", tech if tech else "—")
        with c3:
            st.metric("Class", cls)
        with c4:
            try:
                st.metric("Min margin (frac)", f"{float(mmin):.3g}")
            except Exception:
                st.metric("Min margin (frac)", str(mmin))

        st.caption(f"Contract SHA-256 (prefix): {contract_sha if contract_sha else '—'}")

        # Build a compact table of key limits and values
        rows = []
        def _add(name: str, vkey: str, lkey: str, units: str):
            v = out.get(vkey, float('nan'))
            l = out.get(lkey, float('nan'))
            try:
                v_f = float(v)
                l_f = float(l)
                if (v_f == v_f) or (l_f == l_f):
                    margin = (l_f - v_f) / max(abs(l_f), 1e-9) if (v_f == v_f and l_f == l_f) else float('nan')
                else:
                    margin = float('nan')
            except Exception:
                v_f, l_f, margin = v, l, float('nan')
            rows.append({
                "Quantity": name,
                "Value": v_f,
                "Limit": l_f,
                "Margin(frac)": margin,
                "Units": units,
            })

        _add("TF peak field", "B_peak_T", "B_peak_allow_T", "T")
        _add("TF von Mises stress", "sigma_vm_MPa", "sigma_allow_MPa", "MPa")
        _add("TF engineering J", "J_eng_A_mm2", "J_eng_max_A_mm2", "A/mm^2")
        _add("Coil nuclear heat", "coil_heat_nuclear_MW", "coil_heat_nuclear_max_MW", "MW")
        # SC margin (>=)
        try:
            rows.append({
                "Quantity": "SC critical-surface margin",
                "Value": float(out.get("hts_margin", float("nan"))),
                "Limit": float(out.get("hts_margin_min", float("nan"))),
                "Margin(frac)": float(out.get("hts_margin", float("nan"))) - float(out.get("hts_margin_min", float("nan"))),
                "Units": "-"
            })
        except Exception:
            pass
        try:
            rows.append({
                "Quantity": "Quench proxy margin",
                "Value": float(out.get("quench_proxy_margin", float("nan"))),
                "Limit": float(out.get("quench_proxy_min", float("nan"))),
                "Margin(frac)": float(out.get("quench_proxy_margin", float("nan"))) - float(out.get("quench_proxy_min", float("nan"))),
                "Units": "-"
            })
        except Exception:
            pass

        try:
            import pandas as pd  # type: ignore
            df = pd.DataFrame(rows)
            st.dataframe(df, use_container_width=True, hide_index=True)
        except Exception:
            st.write(rows)

        # Deterministic repair hints (high-level; detailed mapping lives in contract artifact)
        st.markdown("**Deterministic repair levers (non-exhaustive):**")
        st.markdown("- Decrease **Bt** or increase **R0** (reduces required ampere-turns and peak field).")
        st.markdown("- Increase **TF build** (winding/structure) to improve **J** and **stress** margins.")
        st.markdown("- Increase **shielding/build** to reduce **nuclear heat** to coils (when coupled).")

def _sync_point_designer_from_last_point_inp() -> None:
    if not st.session_state.get("pd_needs_sync", False):
        return
    base = st.session_state.get("last_point_inp", None)
    if base is None:
        st.session_state["pd_needs_sync"] = False
        return
    _push_point_inputs_to_pd_widget_keys(base)
    st.session_state["pd_needs_sync"] = False




_sync_point_designer_from_last_point_inp()  # deferred until function is defined

# One-shot candidate apply: any panel may stage a candidate dict for Point Designer.
# This preserves SHAMS law: panels propose; Point Designer remains the frozen truth console.
def _consume_pd_candidate_apply() -> None:
    cand = st.session_state.pop("pd_candidate_apply", None)
    if not isinstance(cand, dict) or not cand:
        return
    try:
        base = make_point_inputs(**cand)
    except Exception:
        try:
            base = PointInputs(**cand)  # type: ignore
        except Exception:
            return
    st.session_state["last_point_inp"] = base
    # Force PD widget sync on next rerun; also push immediately so the same click updates widgets.
    try:
        _push_point_inputs_to_pd_widget_keys(base)
        st.session_state["pd_needs_sync"] = False
    except Exception:
        st.session_state["pd_needs_sync"] = True


def stage_pd_candidate_apply(cand: dict, source: str, note: str | None = None) -> None:
    """Canonical cross-panel handoff to Point Designer + provenance breadcrumb.

    Panels propose candidates. Point Designer evaluates frozen truth.

    v327.4: staging is delegated to ui.handoff to ensure deterministic DSG node-id propagation
    without requiring evaluator execution.
    """
    try:
        from ui.handoff import stage_pd_candidate_apply as _stage  # type: ignore
    except Exception:
        # fallback relative import for environments that package ui as a module
        from .handoff import stage_pd_candidate_apply as _stage  # type: ignore
    _stage(cand=cand, source=source, note=note)

def _verification_report_paths():
    rep = os.path.join(ROOT, "verification", "report.json")
    reqs = os.path.join(ROOT, "requirements", "SHAMS_REQS.yaml")
    reqs_json = os.path.join(ROOT, "requirements", "SHAMS_REQS.json")
    runner = os.path.join(ROOT, "verification", "run_verification.py")
    return rep, reqs, reqs_json, runner

def _verification_needs_run():
    rep, reqs, reqs_json, runner = _verification_report_paths()
    if not os.path.exists(rep):
        return True
    try:
        rep_m = os.path.getmtime(rep)
        deps = [p for p in [reqs, reqs_json, runner] if os.path.exists(p)]
        if not deps:
            return False
        dep_m = max(os.path.getmtime(p) for p in deps)
        return rep_m < dep_m
    except Exception:
        return False

def _run_verification_capture():
    """
    Run verification runner using the current Python interpreter.
    Returns: (ok: bool, stdout: str, stderr: str, seconds: float)
    """
    rep, reqs, reqs_json, runner = _verification_report_paths()
    t0 = time.time()
    if not os.path.exists(runner):
        return False, "", f"Missing verification runner: {runner}", 0.0
    try:
        proc = subprocess.run(
            [sys.executable, runner],
            cwd=ROOT,
            capture_output=True,
            text=True,
            check=False,
        )
        dt = time.time() - t0
        ok = (proc.returncode == 0) and os.path.exists(rep)
        return ok, (proc.stdout or ""), (proc.stderr or ""), dt
    except Exception as e:
        dt = time.time() - t0
        return False, "", f"{type(e).__name__}: {e}", dt


from phase1_core import (
    PointInputs,
    solve_Ip_for_H98_with_Q_match,
    solve_Ip_for_H98_with_Q_match_stream,
    solve_sparc_envelope,
    solve_for_targets,
    solve_for_targets_stream,
    SolveResult,
    optimize_design,
)
from frontier.frontier import find_nearest_feasible
from models.reference_machines import REFERENCE_MACHINES, reference_catalog
from phase1_models import BH_COEFFS
from constraints.constraints import evaluate_constraints
from solvers.optimize import scan_feasible_and_pareto, pareto_front
from docs.variable_registry import registry_dataframe
from shams_io.run_artifact import build_run_artifact
from shams_io.plotting import plot_radial_build_from_artifact, plot_summary_pdf
from solvers.sensitivity import finite_difference_sensitivities


# --- Defensive constructor: UI may pass knobs that are absent in older/newer src/PointInputs
# This keeps the UI stable across PointInputs refactors (extra fields are ignored).
_POINTINPUTS_FIELDS = {f.name for f in fields(PointInputs)}

def make_point_inputs(**kwargs) -> PointInputs:
    """Create PointInputs using only supported dataclass fields."""
    filtered = {k: v for k, v in kwargs.items() if k in _POINTINPUTS_FIELDS}
    return PointInputs(**filtered)



# -----------------------------
# UI helpers
# -----------------------------
def _safe_get(obj, key: str, default=None):
    """Dict/dataclass-safe getter.

    SHAMS UI sometimes handles base objects loaded from JSON artifacts (dict)
    or from in-memory dataclass/namespace objects. This helper preserves
    deterministic fallback semantics without attribute-access crashes.
    """
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _num(label: str, value: float, step: float, fmt: str = None, help: str = None, min_value=None, max_value=None, key: str = None):
    # Streamlit raises if default value is outside [min_value, max_value].
    # Clamp defensively so the UI remains robust even if bounds change.
    v = float(value)
    if min_value is not None:
        try:
            v = max(v, float(min_value))
        except Exception:
            pass
    if max_value is not None:
        try:
            v = min(v, float(max_value))
        except Exception:
            pass
    kwargs = {}
    if fmt: kwargs["format"] = fmt
    if help: kwargs["help"] = help
    if min_value is not None: kwargs["min_value"] = min_value
    if max_value is not None: kwargs["max_value"] = max_value
    return st.number_input(label, value=v, step=float(step), key=key, **kwargs) if key else st.number_input(label, value=v, step=float(step), **kwargs)


def _warn_unrealistic_point_inputs(pi: Any, context: str = "") -> None:
    """Non-blocking, UI-only warnings for obviously unrealistic user inputs.

    This must not change any model/solver behavior; it only surfaces warnings.
    """
    if pi is None:
        return
    # (lo, hi, message)
    checks = [
        ("R0_m", 0.5, 15.0, "Major radius R0 [m] looks unusual"),
        ("a_m", 0.1, 5.0, "Minor radius a [m] looks unusual"),
        ("kappa", 1.0, 3.5, "Elongation κ looks unusual"),
        ("delta", -0.8, 0.8, "Triangularity δ looks unusual"),
        ("Bt_T", 0.5, 25.0, "Toroidal field Bt [T] looks unusual"),
        ("Ip_MA", 0.1, 30.0, "Plasma current Ip [MA] looks unusual"),
        ("Ti_keV", 0.1, 40.0, "Ion temperature Ti [keV] looks unusual"),
        ("fG", 0.05, 1.5, "Greenwald fraction fG looks unusual"),
        ("Paux_MW", 0.0, 300.0, "Auxiliary power Paux [MW] looks unusual"),
        ("t_shield_m", 0.05, 2.0, "Shield thickness t_shield [m] looks unusual"),
    ]
    warns: List[str] = []
    for name, lo, hi, msg in checks:
        if not hasattr(pi, name):
            continue
        try:
            v = float(getattr(pi, name))
        except Exception:
            continue
        if (v < lo) or (v > hi):
            warns.append(f"- {msg}: **{name}={v:g}** (expected roughly {lo:g}–{hi:g})")
    if warns:
        title = "Unrealistic inputs" + (f" ({context})" if context else "")
        st.warning(title + "\n" + "\n".join(warns))


# -----------------------------
# Scan Lab parameter metadata (UI-only)
# -----------------------------

# Human-friendly physics block names (used in tooltips + mapping table)
_PHYS_BLOCKS: Dict[str, str] = {
    "Geometry": "Machine geometry / size assumptions",
    "Magnets & radial build": "TF/HTS coil build, inboard stack closure, peak field mapping, stress",
    "0-D plasma core": "0-D profiles, fusion power, temperatures, density, basic scalings",
    "🌀 Confinement": "Energy confinement (IPB98-like) + confinement multipliers",
    "H-mode access": "L-H threshold (Martin-08-like) + margin screening",
    "Stability & limits": "q95, \u03b2N, bootstrap fraction and related operational screens",
    "Power balance & radiation": "Zeff/dilution/radiation and alpha deposition assumptions",
    "Divertor / SOL": "SOL power loading proxy (PSOL/R) and divertor heat-flux screen",
    "☢️ Neutronics": "TBR proxy + HTS fluence/lifetime proxy",
    "Electrical balance": "Recirculating power closure and net electric power screen",
    "Numerics": "Solver bounds/tolerance and feasibility filtering",
}

# For Scan Lab UI: which parameters are mandatory vs optional + which physics blocks they affect.
_SCAN_PARAM_META: Dict[str, Dict[str, Any]] = {
    # Machine / plasma assumptions
    "R0": {"req": True, "blocks": ["Geometry", "Magnets & radial build", "0-D plasma core", "Divertor / SOL"]},
    "B0": {"req": True, "blocks": ["Magnets & radial build", "0-D plasma core", "🌀 Confinement", "Stability & limits"]},
    "tshield": {"req": True, "blocks": ["Magnets & radial build", "☢️ Neutronics"]},
    "Paux": {"req": True, "blocks": ["0-D plasma core", "Power balance & radiation", "H-mode access", "Electrical balance"]},
    "Paux_for_Q": {"req": True, "blocks": ["0-D plasma core"]},
    "Ti_over_Te": {"req": True, "blocks": ["0-D plasma core", "Power balance & radiation"]},

    # Axes
    "Ti": {"req": True, "blocks": ["0-D plasma core", "🌀 Confinement", "Power balance & radiation"]},
    "H98": {"req": True, "blocks": ["🌀 Confinement"]},
    "a": {"req": True, "blocks": ["Geometry", "0-D plasma core", "Stability & limits", "Divertor / SOL", "Magnets & radial build"]},
    "Q": {"req": True, "blocks": ["0-D plasma core", "Electrical balance"]},
    "g_conf": {"req": True, "blocks": ["🌀 Confinement"]},

    # Solver bounds & screens
    "Ip_bounds": {"req": True, "blocks": ["Numerics", "0-D plasma core", "Stability & limits", "Magnets & radial build"]},
    "fG_bounds": {"req": True, "blocks": ["Numerics", "0-D plasma core"]},
    "tol": {"req": True, "blocks": ["Numerics"]},

    # Screening knobs (plasma)
    "Zeff": {"req": True, "blocks": ["Power balance & radiation"]},
    "dilution_fuel": {"req": True, "blocks": ["Power balance & radiation", "0-D plasma core"]},
    "extra_rad_factor": {"req": True, "blocks": ["Power balance & radiation"]},
    "alpha_loss_frac": {"req": True, "blocks": ["Power balance & radiation"]},
    "kappa": {"req": True, "blocks": ["Stability & limits", "0-D plasma core"]},
    "q95_min": {"req": True, "blocks": ["Stability & limits"]},
    "betaN_max": {"req": True, "blocks": ["Stability & limits"]},
    "C_bs": {"req": True, "blocks": ["Stability & limits"]},
    "f_bs_max": {"req": True, "blocks": ["Stability & limits"]},
    "PSOL_over_R_max": {"req": True, "blocks": ["Divertor / SOL"]},

    # Optional toggle
    "require_Hmode": {"req": False, "blocks": ["H-mode access"]},
    "PLH_margin": {"req": False, "blocks": ["H-mode access"]},

    # Clean design knobs (engineering screens)
    "tblanket_m": {"req": False, "blocks": ["Magnets & radial build", "☢️ Neutronics"]},
    "t_vv_m": {"req": False, "blocks": ["Magnets & radial build"]},
    "t_gap_m": {"req": False, "blocks": ["Magnets & radial build"]},
    "t_tf_struct_m": {"req": False, "blocks": ["Magnets & radial build"]},
    "t_tf_wind_m": {"req": False, "blocks": ["Magnets & radial build"]},
    "Bpeak_factor": {"req": False, "blocks": ["Magnets & radial build"]},
    "sigma_allow_MPa": {"req": False, "blocks": ["Magnets & radial build"]},
    "Tcoil_K": {"req": False, "blocks": ["Magnets & radial build"]},
    "hts_margin_min": {"req": False, "blocks": ["Magnets & radial build"]},
    "Vmax_kV": {"req": False, "blocks": ["Magnets & radial build"]},
    "q_div_max_MW_m2": {"req": False, "blocks": ["Divertor / SOL"]},
    "q_midplane_max_MW_m2": {"req": False, "blocks": ["Divertor / SOL"]},
    "TBR_min": {"req": False, "blocks": ["☢️ Neutronics"]},
    "hts_lifetime_min_yr": {"req": False, "blocks": ["☢️ Neutronics"]},
    "P_net_min_MW": {"req": False, "blocks": ["Electrical balance"]},
}


def _scan_badge(param_key: str) -> str:
    meta = _SCAN_PARAM_META.get(param_key)
    # Default to optional if unknown
    is_req = bool(meta.get("req")) if isinstance(meta, dict) else False
    return "🟥 Mandatory" if is_req else "⬜ Optional"


def _scan_blocks(param_key: str) -> List[str]:
    meta = _SCAN_PARAM_META.get(param_key)
    if not isinstance(meta, dict):
        return []
    return [b for b in meta.get("blocks", []) if b in _PHYS_BLOCKS]


def _scan_label(base: str, param_key: str) -> str:
    # number_input labels do not render markdown; keep it simple + consistent.
    return f"{base}  ·  {_scan_badge(param_key)}"


def _scan_help(base_help: str, param_key: str) -> str:
    blocks = _scan_blocks(param_key)
    if not blocks:
        return base_help
    lines = [base_help.strip(), "", "Maps to physics blocks:"]
    for b in blocks:
        lines.append(f"- {b}: {_PHYS_BLOCKS[b]}")
    return "\n".join(lines).strip()

def kpi_row(items: List[Tuple[str, Any]]):
    cols = st.columns(len(items))
    for c, (k, v) in zip(cols, items):
        c.metric(k, v)


def _numeric_cols(df: pd.DataFrame) -> List[str]:
    cols = []
    for c in df.columns:
        try:
            if pd.api.types.is_numeric_dtype(df[c]):
                cols.append(c)
        except Exception:
            pass
    return cols


def plot_scatter(df: pd.DataFrame, x: str, y: str, color: str | None = None, title: str | None = None):
    """Matplotlib scatter with optional numeric color."""
    if df is None or df.empty:
        st.info("No data to plot.")
        return
    if x not in df or y not in df:
        st.warning("Select valid x/y columns.")
        return

    d = df[[x, y] + ([color] if color and color in df else [])].dropna()
    if d.empty:
        st.info("No finite rows for this plot (after dropping NaNs).")
        return
    # Matplotlib is optional: if missing, fall back to Streamlit's built-in charts.
    if not _HAVE_MPL:
        st.warning("Plotting is limited because 'matplotlib' is not installed. Install it (pip install matplotlib) for full plotting.")
        # Streamlit fallback (no colorbar support)
        try:
            st.scatter_chart(d, x=x, y=y)
        except Exception:
            st.line_chart(d[[x, y]].rename(columns={x: "x", y: "y"}))
        return

    fig = plt.figure(figsize=(6.8, 4.6))
    ax = plt.gca()
    if color and color in d and pd.api.types.is_numeric_dtype(d[color]):
        sc = ax.scatter(d[x], d[y], c=d[color], s=22, alpha=0.85)
        cb = plt.colorbar(sc, ax=ax)
        cb.set_label(color)
    else:
        ax.scatter(d[x], d[y], s=22, alpha=0.85)

    ax.set_xlabel(x)
    ax.set_ylabel(y)
    if title:
        ax.set_title(title)
    ax.grid(True, alpha=0.25)
    st.pyplot(fig, clear_figure=True)


def plot_bars(values: Dict[str, float], title: str):
    keys = [k for k in values.keys() if isinstance(values.get(k), (int, float)) and math.isfinite(float(values.get(k))) ]
    if not keys:
        st.caption("No plottable values available.")
        return
    if not _HAVE_MPL:
        st.warning("Bar charts are limited because 'matplotlib' is not installed. Install it (pip install matplotlib) for full plotting.")
        import pandas as _pd
        s = _pd.Series({k: float(values[k]) for k in keys})
        try:
            st.bar_chart(s)
        except Exception:
            st.write(s)
        return

    fig = plt.figure(figsize=(6.8, 4.4))
    ax = plt.gca()
    ax.bar(range(len(keys)), [float(values[k]) for k in keys])
    ax.set_xticks(range(len(keys)), keys, rotation=35, ha="right")
    ax.set_title(title)
    ax.grid(True, axis="y", alpha=0.25)
    st.pyplot(fig, clear_figure=True)

def badge(check):
    """Render PASS/FAIL/WARN/SKIPPED badge.
    Accepts either a check dict with 'status' or a legacy ok flag.
    """
    if isinstance(check, dict):
        stt = check.get('status')
        if stt == 'SKIPPED':
            return '⚪ SKIPPED'
        if stt == 'WARN':
            return '🟡 WARN'
        if stt == 'FAIL':
            return ' FAIL'
        if stt == 'PASS':
            return ' PASS'
        # fallback
        ok = check.get('ok')
    else:
        ok = check
    if ok is None:
        return '⚪ SKIPPED'
    return ' PASS' if ok else ' FAIL'


def finite(x):
    return isinstance(x, (int, float)) and math.isfinite(x)

def top_violations(checks: List[Dict[str, Any]], n: int = 3) -> List[Dict[str, Any]]:
    bad = [c for c in checks if c.get('status') == 'FAIL']

    # sort by relative violation if available
    def score(c):
        if c.get("value") is None or c.get("limit") is None:
            return 0.0
        v, lim = c["value"], c["limit"]
        if not (finite(v) and finite(lim)) or lim == 0:
            return 0.0
        if c.get("sense") == "max":
            return (v - lim) / abs(lim)
        if c.get("sense") == "min":
            return (lim - v) / abs(lim)
        return 0.0
    bad.sort(key=score, reverse=True)
    return bad[:n]

def top_warnings(checks: List[Dict[str, Any]], n: int = 3) -> List[Dict[str, Any]]:
    ws = [c for c in checks if c.get('status') == 'WARN']
    def score(c):
        v = c.get('value'); lim = c.get('limit'); sense = c.get('sense')
        if v is None or lim is None or not finite(v) or not finite(lim):
            return 0.0
        if sense == 'max':
            return max(0.0, (v - lim) / abs(lim))
        if sense == 'min':
            return max(0.0, (lim - v) / abs(lim))
        return 0.0
    ws.sort(key=score, reverse=True)
    return ws[:n]

def compute_checks(out: Dict[str, float]) -> List[Dict[str, Any]]:
    """
    Turn flat outputs into a structured constraint list.

    IMPORTANT:
    - Checks are only evaluated when the needed physics is enabled *and* the value is finite.
    - If a model/physics block is disabled (or produced NaN), the check is marked SKIPPED.
    - Checks can be WARNING-level or HARD FAIL depending on how far they are from the limit.
    """
    checks: List[Dict[str, Any]] = []


    # global warning behavior knobs (can be overridden per-check via explicit warn_limit)
    warn_frac_max = float(out.get("_warn_frac_max", 0.90))  # for max constraints: WARN if v > warn_frac*limit
    warn_frac_min = float(out.get("_warn_frac_min", 1.10))  # for min constraints: WARN if v < warn_frac*limit

    def add(name, status, value=None, limit=None, sense=None, notes="", severity="hard", warn_limit=None):
        """status in {'PASS','FAIL','WARN','SKIPPED'}"""
        ok = None
        if status == "PASS":
            ok = True
        elif status == "FAIL":
            ok = False
        elif status == "WARN":
            ok = True  # warn is still 'ok' for legacy consumers
        elif status == "SKIPPED":
            ok = None
        checks.append({
            "name": name,
            "status": status,
            "ok": ok,
            "value": value,
            "limit": limit,
            "warn_limit": warn_limit,
            "sense": sense,
            "notes": notes,
            "severity": severity,
        })

        # ok can be True/False/None (None => skipped)
        checks.append({"name": name, "ok": ok if ok in (True, False, None) else bool(ok),
                       "value": value, "limit": limit, "sense": sense, "notes": notes})

    def fin(x) -> bool:
        return isinstance(x, (int, float)) and math.isfinite(x)
    def eval_max(name, key_value, key_limit, notes="", severity="hard", warn_limit=None):
        v = out.get(key_value)
        lim = out.get(key_limit)
        if (not fin(v)) or (not fin(lim)):
            add(name, "SKIPPED", v, lim, "max", notes, severity=severity, warn_limit=warn_limit)
            return
        wl = warn_limit
        if wl is None:
            wl = warn_frac_max * lim
        if v > lim:
            add(name, "FAIL", v, lim, "max", notes, severity=severity, warn_limit=wl)
        elif v > wl:
            add(name, "WARN", v, lim, "max", notes, severity=severity, warn_limit=wl)
        else:
            add(name, "PASS", v, lim, "max", notes, severity=severity, warn_limit=wl)

    def eval_min(name, key_value, key_limit, notes="", severity="hard", warn_limit=None):
        v = out.get(key_value)
        lim = out.get(key_limit)
        if (not fin(v)) or (not fin(lim)):
            add(name, "SKIPPED", v, lim, "min", notes, severity=severity, warn_limit=warn_limit)
            return
        wl = warn_limit
        if wl is None:
            wl = warn_frac_min * lim
        # for min constraints, warn if v is between lim and wl (wl > lim)
        if v < lim:
            add(name, "FAIL", v, lim, "min", notes, severity=severity, warn_limit=wl)
        elif v < wl:
            add(name, "WARN", v, lim, "min", notes, severity=severity, warn_limit=wl)
        else:
            add(name, "PASS", v, lim, "min", notes, severity=severity, warn_limit=wl)    # --- Build closure ---
    if "radial_build_ok" in out:
        v = out.get("radial_build_ok")
        if not fin(v):
            add("Radial build closure", "SKIPPED", out.get("rb_total_inboard_m"), out.get("rinboard_available_m"), "max",
                "Requires inboard stack thickness ≤ available inboard space (R0 - a).")
        else:
            add("Radial build closure", "PASS" if (v > 0.5) else "FAIL", out.get("rb_total_inboard_m"), out.get("rinboard_available_m"), "max",
                "Requires inboard stack thickness ≤ available inboard space (R0 - a).")

    # --- Magnet stress ---


    if "sigma_hoop_MPa" in out and "sigma_allow_MPa" in out:
        eval_max("TF hoop stress", "sigma_hoop_MPa", "sigma_allow_MPa",
                 "Hoop stress proxy must be below allowable structural stress.")

    # --- HTS margin ---
    if "hts_margin" in out and "hts_margin_min" in out:
        eval_min("HTS margin", "hts_margin", "hts_margin_min",
                 "HTS operating margin proxy vs (B,T) must exceed minimum.")

    # --- Dump voltage ---
    if "V_dump_kV" in out and "Vmax_kV" in out:
        eval_max("Dump voltage", "V_dump_kV", "Vmax_kV",
                 "Fast discharge voltage must not exceed protection limit.")

    # --- Divertor heat flux ---

    if "q_div_MW_m2" in out and "q_div_max_MW_m2" in out:
        eval_max("Divertor heat flux", "q_div_MW_m2", "q_div_max_MW_m2",
                 "Peak divertor heat flux proxy must be below limit.")

    # --- Tritium breeding ratio ---
    if "TBR" in out and "TBR_min" in out:
        eval_min("TBR", "TBR", "TBR_min",
                 "Tritium breeding ratio proxy must exceed minimum.")

    # --- TBR proxy validity-domain (v321) ---
    if "TBR_domain_ok" in out:
        ok = out.get("TBR_domain_ok")
        enf = out.get("neutronics_domain_enforce", 0.0)
        if not fin(ok):
            add("TBR validity domain", "SKIPPED", ok, 1.0, "min", "Proxy validity-domain check not available.")
        else:
            if (fin(enf) and enf > 0.5):
                add("TBR validity domain", "PASS" if (ok > 0.5) else "FAIL", ok, 1.0, "min", "If enforced, TBR proxy must be inside declared validity domain.")
            else:
                add("TBR validity domain", "PASS" if (ok > 0.5) else "WARN", ok, 1.0, "min", "Not enforced by default; WARN means the proxy is out-of-domain.")

    # --- HTS lifetime ---
    if "hts_lifetime_yr" in out and "hts_lifetime_min_yr" in out:
        eval_min("HTS lifetime", "hts_lifetime_yr", "hts_lifetime_min_yr",
                 "Neutron lifetime proxy of HTS must exceed minimum.")

    # --- Net electric power ---
    if "P_net_e_MW" in out and "P_net_min_MW" in out:
        eval_min("Net electric power", "P_net_e_MW", "P_net_min_MW",
                 "Net electric power must exceed minimum (system closure).")

    # --- H-mode access (only if enforced) ---
    # If require_Hmode is False or physics is disabled, this becomes SKIPPED.
    if "require_Hmode" in out and "LH_ok" in out:
        req = out.get("require_Hmode")
        lh_ok = out.get("LH_ok")
        if not (fin(req) and req > 0.5):
            add("H‑mode access", None, lh_ok, 1.0, "min",
                "H‑mode not required (or LH physics disabled).")
        else:
            if not fin(lh_ok):
                add("H‑mode access", None, lh_ok, 1.0, "min",
                    "H‑mode required, but LH physics not available for this point.")
            else:
                add("H‑mode access", lh_ok > 0.5, lh_ok, 1.0, "min",
                    "If H‑mode required, point must be above LH threshold with margin.")

    return checks


# -----------------------------
# Scan runner (UI-native)
# -----------------------------
def frange(start: float, stop: float, step: float) -> List[float]:
    vals: List[float] = []
    if step == 0:
        return [start]
    x = start
    if step > 0:
        while x <= stop + 1e-12:
            vals.append(float(x))
            x += step
    else:
        while x >= stop - 1e-12:
            vals.append(float(x))
            x += step
    return vals

def run_scan(spec: Dict[str, Any]) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    A refactor of the CLI scan loop into a UI-callable function.

    Returns:
      df_feasible: rows of extended-feasible points (same spirit as 'feasible_ext' sheet)
      meta: dict with scan settings + summary stats
    """
    Ti_grid = frange(spec["Ti_start"], spec["Ti_stop"], spec["Ti_step"] if spec["Ti_stop"] >= spec["Ti_start"] else -abs(spec["Ti_step"]))
    H_grid = frange(spec["H98_start"], spec["H98_stop"], abs(spec["H98_step"]))
    a_grid = frange(spec["a_min"], spec["a_max"], abs(spec["a_step"]))
    Q_grid = frange(spec["Q_start"], spec["Q_stop"], abs(spec["Q_step"]))
    g_grid = frange(spec["gconf_start"], spec["gconf_stop"], abs(spec["gconf_step"]))


    # --- Scan Lab: optional UI progress + logging hooks (kept no-op for non-UI use) ---
    progress_cb = spec.get("_progress_cb")  # callable(fraction: float, info: dict) -> None
    log_cb = spec.get("_log_cb")            # callable(line: str) -> None
    log_lines: List[str] = []

    def _log(line: str) -> None:
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        s = f"[{ts}] {line}"
        log_lines.append(s)
        if callable(log_cb):
            try:
                log_cb(s)
            except Exception:
                pass

    def _progress(i: int, n: int, **info: Any) -> None:
        if callable(progress_cb) and n > 0:
            try:
                progress_cb(min(max(i / n, 0.0), 1.0), info)
            except Exception:
                pass

    n_total = max(1, len(g_grid) * len(Ti_grid) * len(H_grid) * len(a_grid) * len(Q_grid))
    _log(f"Scan initialized: {len(g_grid)} g_conf × {len(Ti_grid)} Ti × {len(H_grid)} H98 × {len(a_grid)} a × {len(Q_grid)} Q  => {n_total} evaluations")
    i_eval = 0

    rows: List[Dict[str, Any]] = []
    best_g = None

    # --- v327.3: pipeline DSG edge automation (best-effort; does not change truth) ---
    dsg_parent = spec.get("_dsg_parent_node_id")
    if not dsg_parent:
        try:
            dsg_parent = st.session_state.get("dsg_selected_node_id") or st.session_state.get("active_design_node_id")
        except Exception:
            dsg_parent = None

    scan_node_ids: List[str] = []
    scan_edge_note = f"scan: g_conf[{len(g_grid)}] Ti[{len(Ti_grid)}] H98[{len(H_grid)}] a[{len(a_grid)}] Q[{len(Q_grid)}]"


    for g_conf in g_grid:
        for Ti in Ti_grid:
            for Hreq in H_grid:
                for a in a_grid:
                    for Qtar in Q_grid:
                        # Solve at reduced target (same logic as CLI driver)

                        i_eval += 1
                        _progress(i_eval, n_total,
                                  stage="setup",
                                  g_conf=float(g_conf), Ti_keV=float(Ti), H98_req=float(Hreq), a_m=float(a), Q_target=float(Qtar))
                        _log(f"Eval {i_eval}/{n_total}: g_conf={g_conf:.3g}, Ti={Ti:.3g} keV, H98_req={Hreq:.3g}, a={a:.3g} m, Q={Qtar:.3g}")
                        _log("  - Building point inputs (geometry, fields, density/temperature assumptions)")

                        H_base_target = Hreq / max(g_conf, 1e-9)

                        base = make_point_inputs(
                            R0_m=spec["R0"],
                            a_m=a,
                            kappa=spec["kappa"],
                            Bt_T=spec["B0"],
                            Ip_MA=0.5*(spec["Ip_min"]+spec["Ip_max"]),
                            Ti_keV=Ti,
                            fG=0.8,
                            t_shield_m=spec["tshield"],
                            Paux_MW=spec["Paux"],
                            Ti_over_Te=spec["Ti_over_Te"],
                            zeff=spec["Zeff"],
                            dilution_fuel=spec["dilution_fuel"],
                            fuel_mode="DT",
                            include_secondary_DT=include_secondary_DT,
                            tritium_retention=0.5,
                            tau_T_loss_s=5.0,
                            extra_rad_factor=spec["extra_rad_factor"],
                            alpha_loss_frac=spec["alpha_loss_frac"],
                            C_bs=spec["C_bs"],
                            require_Hmode=spec["require_Hmode"],
                            PLH_margin=spec["PLH_margin"],
                            # --- Clean design knobs (passed through PointInputs defaults if present in your src)
                            **spec.get("clean_knobs", {}),
                        )


                        _log("  - Solving nested system: outer Ip for H98, inner fG for Q (bisection)")
                        _progress(i_eval, n_total, stage="solve")
                        sol_inp, sol_out, ok = solve_Ip_for_H98_with_Q_match(
                            base=base,
                            target_H98=H_base_target,
                            target_Q=Qtar,
                            Ip_min=spec["Ip_min"],
                            Ip_max=spec["Ip_max"],
                            fG_min=spec["fG_min"],
                            fG_max=spec["fG_max"],
                            tol=spec["tol"],
                            Paux_for_Q_MW=spec["Paux_for_Q"],
                        )
                        if not ok:
                            _log("  - Solver failed to bracket/converge for this combo (skipping)")
                            continue

                        # Effective confinement
                        H98_eff = g_conf * sol_out["H98"]
                        sol_out["H98_eff"] = H98_eff


                        _log("  - Evaluating physics proxies (power balance, confinement, operational limits)")
                        _progress(i_eval, n_total, stage="evaluate")
                        # Standard ext checks from CLI
                        ok_ext = True
                        if sol_out["ne20"] > 1.2:
                            ok_ext = False
                        if sol_out.get('q95_proxy', 1e9) < spec["q95_min"]:
                            ok_ext = False
                        if sol_out.get("betaN_proxy", 0.0) > spec["betaN_max"]:
                            ok_ext = False
                        if sol_out.get("f_bs_proxy", 0.0) > spec["f_bs_max"]:
                            ok_ext = False
                        PSOL_over_R = sol_out["Ploss_MW"] / spec["R0"]
                        sol_out["PSOL_over_R"] = PSOL_over_R
                        if PSOL_over_R > spec["PSOL_over_R_max"]:
                            ok_ext = False
                        if spec["require_Hmode"] and sol_out.get("LH_ok", 1.0) < 0.5:
                            ok_ext = False
                        if H98_eff < Hreq:
                            ok_ext = False
                        if sol_out["Q_DT_eqv"] < Qtar:
                            ok_ext = False

                        # Clean design checks (if present)
                        checks = compute_checks(sol_out)
                        if any((not c["ok"]) for c in checks):
                            ok_ext = False

                        if not ok_ext:
                            _log("  - Failed screening checks (skipping)")
                            continue

                        if best_g is None or g_conf < best_g:
                            best_g = g_conf

                        _log("  - Feasible point found ✓ (adding to results)")
                        _progress(i_eval, n_total, stage="record")
                        row = dict(sol_out)
                        row.update({
                            "g_conf": g_conf,
                            "Ti_keV": Ti,
                            "Q_target": Qtar,
                            "H98_required": Hreq,
                            "a_m": a,
                            "Ip_MA": sol_inp.Ip_MA,
                            "f_G": sol_inp.fG,
                            "Paux_MW": sol_inp.Paux_MW,
                            "Paux_for_Q_MW": spec["Paux_for_Q"],
                            "H98_eff": H98_eff,
                        })

                        # --- v327.3: DSG pipeline capture for scan points (best-effort) ---
                        try:
                            g = st.session_state.get("_shams_dsg")
                            if g is not None:
                                node = g.record(
                                    inp=sol_inp,
                                    out=dict(sol_out),
                                    ok=True,
                                    message="scan_feasible",
                                    elapsed_s=0.0,
                                    origin="ScanLab",
                                    parents=[str(dsg_parent)] if dsg_parent else None,
                                    tags=["scan", "feasible"],
                                    edge_kind="scan",
                                    edge_note=scan_edge_note,
                                )
                                scan_node_ids.append(node.node_id)
                                st.session_state["active_design_node_id"] = node.node_id
                        except Exception:
                            pass                        # --- v327.4: pipeline-native dsg_node_id column for scan tables ---
                        try:
                            if "dsg_node_id" not in row:
                                _nid = None
                                try:
                                    _nid = node.node_id  # type: ignore[name-defined]
                                except Exception:
                                    _nid = None
                                if not _nid:
                                    from evaluator.cache_key import sha256_cache_key
                                    _nid = sha256_cache_key(sol_inp)
                                row["dsg_node_id"] = str(_nid)
                        except Exception:
                            pass

                        rows.append(row)

    df = pd.DataFrame(rows)
    meta = dict(spec)
    meta["best_g_conf_found"] = best_g if best_g is not None else "NONE"
    meta["n_feasible"] = int(len(df))

    # Strip non-serializable UI callbacks from meta and attach log text
    meta.pop("_progress_cb", None)
    meta.pop("_log_cb", None)
    meta["scan_log_text"] = "\n".join(log_lines)
    # --- v327.3: expose scan DSG node IDs for downstream panels ---
    try:
        meta["dsg_parent_node_id"] = str(dsg_parent) if dsg_parent else ""
        meta["dsg_scan_node_ids"] = list(scan_node_ids)
        st.session_state["scan_last_node_ids"] = list(scan_node_ids)
        st.session_state["scan_last_parent_node_id"] = str(dsg_parent) if dsg_parent else ""
        _dsg_save_best_effort()
    except Exception:
        pass
    _log(f"Scan complete: feasible={meta['n_feasible']}  best_g_conf_found={meta['best_g_conf_found']}")

    return df, meta

def df_to_excel_bytes(df: pd.DataFrame, meta: Dict[str, Any]) -> bytes:
    """
    Export feasible dataframe + meta into an Excel workbook (in-memory).
    """
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "feasible_ext"

    if df.empty:
        ws.append(["NO_FEASIBLE_POINTS"])
    else:
        ws.append(list(df.columns))
        for c in range(1, len(df.columns)+1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        ws.freeze_panes = "A2"
        for _, r in df.iterrows():
            ws.append([r.get(c) for c in df.columns])

    # meta sheet
    wsM = wb.create_sheet("meta")
    wsM.append(["key", "value"])
    wsM["A1"].font = Font(bold=True)
    for k, v in meta.items():
        # keep meta compact
        if isinstance(v, dict):
            continue
        if isinstance(v, list):
            continue
        wsM.append([k, v])
    wsM.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -----------------------------
# Streamlit UI
# -----------------------------
st.set_page_config(page_title=APP_NAME, layout="wide")


# v276: keep last run artifact for cross-panels
if "last_run_artifact" not in st.session_state:
    st.session_state["last_run_artifact"] = {}

_render_branding_header()
_shams_status_strip_placeholder = st.empty()

def _shams_status_strip():
    # Top-of-page professional status strip (Review-room instrumentation)
    locked, task, started, is_owner = _shams_runlock.status(st.session_state.get("_shams_owner_token"), app_start_ts=st.session_state.get("_shams_app_start_ts"))
    prev_locked = bool(st.session_state.get("_shams_prev_locked", False))
    prev_task = st.session_state.get("_shams_prev_task")

    # Completion detection: was locked, now unlocked
    if prev_locked and not locked and prev_task:
        try:
            if not st.session_state.get("silence_mode", False):
                st.toast(f" Sequence Complete: {prev_task}")
        except Exception:
            pass

    st.session_state["_shams_prev_locked"] = locked
    st.session_state["_shams_prev_task"] = task

    with _shams_status_strip_placeholder.container():
        if locked and task:
            age_s = int(_shams_time.time() - float(started or _shams_time.time()))
            st.info(f"⚡ **Running Sequence** · {task} · t+{age_s}s · All other solver actions are locked.", icon="⚡")
            # Show a concise tail of the Black-Box Chronicle
            try:
                _lg = _activity_logger()
                tail_text = _lg.path.read_text(encoding="utf-8", errors="replace") if _lg.path.exists() else ""
                tail_lines = tail_text.splitlines()[-20:] if tail_text else []
                if tail_lines:
                    with st.expander("Black-Box Chronicle tail (live)", expanded=False):
                        st.code("\n".join(tail_lines), language="")
            except Exception:
                pass
        else:
            # Calm "ready" strip for experts
            st.caption("Status: Ready · Helm Console armed · Awaiting next sequence.")

_shams_status_strip()

st.session_state.setdefault('shams_state', SessionStateModel())
# ---- UI polish (CSS) ----
# Streamlit theming is usually done via .streamlit/config.toml, but we keep it self-contained.
st.markdown(
    """
<style>
  /* Slightly tighter overall spacing */
  .block-container { padding-top: 2.2rem; padding-bottom: 3.4rem; }
  h1 { margin-top: 0.25rem; line-height: 1.15; }

  /* Header spacing: avoid negative letter-spacing which can overlap glyphs at some zoom levels */
  h1, h2, h3 { letter-spacing: 0; }

  /* Metric cards: increase contrast and soften edges */
  [data-testid="stMetric"] {
    padding: 0.75rem 0.75rem;
    border-radius: 16px;
    border: 1px solid rgba(49, 51, 63, 0.15);
    background: rgba(255, 255, 255, 0.03);
  }

  /* Buttons */
  div.stButton > button {
    border-radius: 14px;
    padding: 0.55rem 0.9rem;
    font-weight: 650;
  }

  /* Expander */
  details {
    border-radius: 14px;
    border: 1px solid rgba(49, 51, 63, 0.12);
    padding: 0.25rem 0.4rem;
  }

  /* Code blocks */
  pre {
    border-radius: 14px;
  }

  /* Dataframes */
  .stDataFrame { border-radius: 14px; overflow: hidden; }
</style>
    """,
    unsafe_allow_html=True,
)

# Persistent footer (render early so st.stop paths still show it)
_render_footer()
st.sidebar.markdown("## Helm Console - Expert Navigation")

# Session & Authority (professional, read-mostly posture)
_forge_review_mode = bool(st.session_state.get("forge_review_mode", False))
_posture = "Review Mode (locked)" if _forge_review_mode else "Explore Mode"
st.sidebar.caption("Captain’s Ledger")
st.sidebar.markdown(f"""- **Posture:** {_posture}
- **Authority:** Frozen evaluator
- **Workspace:** Non-authoritative""")

st.session_state.explain_mode = st.sidebar.toggle(
    "Explain mode (show equations & reasons)",
    value=bool(st.session_state.get("explain_mode", True)),
    help="Teaching mode: show model equations, assumptions, and why constraints bind.",
)

with st.sidebar.expander("Advanced controls", expanded=False):
    st.session_state.expert_mode = st.toggle(
        "Expert controls",
        value=bool(st.session_state.get("expert_mode", False)),
        disabled=_forge_review_mode,
        help="Expose solver tolerances and optimizer internals. Disabled in Review Mode.",
    )

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Design Intent - influences what counts as "hard" in Systems/Optimization
# ---------------------------------------------------------------------------
_design_intent_prev = st.session_state.get("design_intent", "Power Reactor (net-electric)")
with st.sidebar.expander("Reactor Covenant", expanded=False):
    st.session_state["design_intent"] = st.selectbox(
        "Intent",
        ["Power Reactor (net-electric)", "Experimental Device (research)"],
        index=0 if ("reactor" in _design_intent_prev.lower() or _design_intent_prev.lower().startswith("power")) else 1,
        help="Reactor intent enforces strict engineering/plant constraints (e.g., TBR, stress, heat flux). Research intent relaxes/softens some constraints to explore physics-relevant machines.",
    )

# (Design contract continues below in Verification/Fidelity panels)

if st.session_state.get("design_intent") != _design_intent_prev:

    _invalidate_mode_caches("design_intent_changed")

    try:
        from schema.governance_presets import apply_governance_preset, tritium_tight_closure_default
    except ImportError:
        from src.schema.governance_presets import apply_governance_preset, tritium_tight_closure_default
    try:
        _lpi = st.session_state.get("last_point_inp")
        if _lpi is not None:
            from dataclasses import asdict
            _fields = asdict(_lpi) if hasattr(_lpi, "__dataclass_fields__") else dict(_lpi)
            apply_governance_preset(_fields, design_intent=str(st.session_state.get("design_intent", "")))
            st.session_state["include_tritium_tight_closure"] = bool(
                _fields.get("include_tritium_tight_closure", tritium_tight_closure_default(str(st.session_state.get("design_intent", ""))))
            )
    except Exception:
        pass

    try:
        _alog("UI", "DesignIntentChanged", {"from": _design_intent_prev, "to": st.session_state.get("design_intent")})
    except Exception:
        pass

def _design_intent_key() -> str:
    s = str(st.session_state.get("design_intent", "Power Reactor (net-electric)")).strip().lower()
    if s.startswith("experimental") or s.startswith("research") or ("research" in s):
        return "research"
    return "reactor"

# Canonical constraint enforcement sets (names must match constraint 'name' strings)
# NOTE: q95 is enforced hard in both intents per user request.
_INTENT_HARD = {
    "reactor": {"q95", "q_div", "P_SOL/R", "sigma_vm", "B_peak", "HTS margin", "TBR", "NWL"},
    "research": {"q95"},
}
_INTENT_SOFT = {
    "reactor": set(),
    "research": {"q_div", "P_SOL/R", "sigma_vm", "B_peak", "HTS margin", "NWL"},  # shown, not enforced as hard
}
_INTENT_IGNORE = {
    "reactor": set(),
    "research": {"TBR"},  # research machines may not aim for breeding blanket
}

def _hard_constraint_names_for_intent() -> set[str]:
    k = _design_intent_key()
    return set(_INTENT_HARD.get(k, set()))

def _ignored_constraint_names_for_intent() -> set[str]:
    k = _design_intent_key()
    return set(_INTENT_IGNORE.get(k, set()))


def _constraint_policy_snapshot() -> dict:
    """Return an intent-aware constraint policy snapshot (UI/exports)."""
    k = _design_intent_key()
    return {
        "design_intent": str(st.session_state.get("design_intent", "Power Reactor (net-electric)")),
        "intent_key": k,
        "hard_blocking": sorted([str(x) for x in _INTENT_HARD.get(k, set())]),
        "diagnostic_only": sorted([str(x) for x in _INTENT_SOFT.get(k, set())]),
        "ignored": sorted([str(x) for x in _INTENT_IGNORE.get(k, set())]),
    }


def _classify_failed_constraints(failed_names: list[str] | None) -> dict:
    """Classify failed constraint names into blocking/diagnostic/ignored per current intent."""
    failed = [str(x) for x in (failed_names or [])]
    hard_set = _hard_constraint_names_for_intent()
    ign_set = _ignored_constraint_names_for_intent()
    blocking = [c for c in failed if c in hard_set]
    ignored = [c for c in failed if c in ign_set]
    diagnostic = [c for c in failed if (c not in blocking and c not in ignored)]
    return {"blocking": blocking, "diagnostic": diagnostic, "ignored": ignored}








# ---------------------------------------------------------------------------
# Integrity gate (requirements & health)
# ---------------------------------------------------------------------------
with st.sidebar.expander("Integrity Gate - Requirements & Health", expanded=False):
    """A reviewer-safe, deterministic health/requirements check.

    Important: Streamlit executes UI code even when an expander is closed; therefore we avoid
    auto-running anything here. The user explicitly presses the run button.
    """

    rep_path, reqs_path, reqs_json_path, runner_path = _verification_report_paths()

    # Lightweight health checks (no subprocess): these are instantaneous and always safe.
    def _health_rows() -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        try:
            rows.append({"Check": "Python", "Status": f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"})
        except Exception:
            rows.append({"Check": "Python", "Status": "unknown"})

        try:
            repo_ok = os.path.exists(os.path.join(ROOT_DIR, "MANIFEST_SHA256.txt"))
            rows.append({"Check": "Repo manifest", "Status": "present" if repo_ok else "missing"})
        except Exception:
            rows.append({"Check": "Repo manifest", "Status": "unknown"})

        try:
            out_dir = os.path.join(ROOT_DIR, "benchmarks", "publication")
            rows.append({"Check": "Benchmarks folder", "Status": "present" if os.path.isdir(out_dir) else "missing"})
        except Exception:
            rows.append({"Check": "Benchmarks folder", "Status": "unknown"})

        try:
            # Write-permission probe in the current working directory (deterministic + harmless)
            probe_dir = os.path.join(ROOT_DIR, ".shams_probe")
            os.makedirs(probe_dir, exist_ok=True)
            probe_file = os.path.join(probe_dir, "write_test.txt")
            with open(probe_file, "w", encoding="utf-8") as f:
                f.write("ok")
            rows.append({"Check": "Write access", "Status": "ok"})
        except Exception:
            rows.append({"Check": "Write access", "Status": "blocked"})
        return rows

    needs = _verification_needs_run()
    report_exists = os.path.exists(rep_path)
    status_line = (
        "🟢 Evidence report: up-to-date" if (report_exists and not needs) else
        "🟠 Evidence report: needs update" if report_exists else
        "🔴 Evidence report: missing"
    )
    st.caption("Run this only when you want an explicit, archived compliance report. Nothing runs automatically.")
    st.markdown(status_line)

    c1, c2 = st.columns([1, 1])
    force = c1.button("Run gatecheck", use_container_width=True)
    show_logs = c2.toggle("Show logs", value=bool(st.session_state.get("verify_show_logs", False)))
    st.session_state["verify_show_logs"] = show_logs

    # Provide the latest report (if present)
    if report_exists:
        try:
            rep_bytes = Path(rep_path).read_bytes()
            st.download_button(
                "Download evidence report (JSON)",
                data=rep_bytes,
                file_name="shams_verification_report.json",
                mime="application/json",
                use_container_width=True,
            )
        except Exception:
            st.caption("Evidence report download unavailable (read error).")

    with st.expander("Instant health snapshot", expanded=False):
        try:
            st.dataframe(pd.DataFrame(_health_rows()), hide_index=True, use_container_width=True)
        except Exception:
            st.json(_health_rows(), expanded=False)

    if force:
        with st.spinner("Running gatecheck (verification/run_verification.py)..."):
            ok, out, err, dt = _run_verification_capture()
        st.session_state["_last_verify_ok"] = ok
        st.session_state["_last_verify_out"] = out
        st.session_state["_last_verify_err"] = err
        st.session_state["_last_verify_dt"] = dt
        st.rerun()

    if st.session_state.get("_last_verify_dt") is not None:
        ok = bool(st.session_state.get("_last_verify_ok", False))
        (st.success if ok else st.error)(
            f"Last gatecheck: {'PASS' if ok else 'FAIL'} ({st.session_state.get('_last_verify_dt', 0.0):.2f}s)"
        )
        # If failing, surface a short, high-signal reason even when logs are hidden.
        if not ok:
            _e = (st.session_state.get("_last_verify_err", "") or "").strip()
            _o = (st.session_state.get("_last_verify_out", "") or "").strip()
            _msg = ""
            if _e:
                _msg = _e.splitlines()[0]
            elif _o:
                _msg = _o.splitlines()[0]
            if _msg:
                st.caption(f"Gatecheck detail: {_msg}  (toggle **Show logs** for full output)")

    if show_logs:
        st.text_area("stdout", value=str(st.session_state.get("_last_verify_out", "")), height=160)
        st.text_area("stderr", value=str(st.session_state.get("_last_verify_err", "")), height=160)


# ---------------------------------------------------------------------------
# Fidelity + Calibration (transparent (systems-code-inspired), transparent)
# ---------------------------------------------------------------------------
with st.sidebar.expander("Model Authority & Closures", expanded=False):
    fid = st.session_state.get("fidelity_config", {})
    plasma = st.selectbox("🔥 Plasma", ["0D","1/2D"], index=0 if fid.get("plasma","0D")=="0D" else 1)
    magnets = st.selectbox("🧲 Magnets", ["limits","stress"], index=0 if fid.get("magnets","limits")=="limits" else 1)
    exhaust = st.selectbox("💨 Exhaust", ["proxy","enriched"], index=0 if fid.get("exhaust","proxy")=="proxy" else 1)
    neutronics = st.selectbox("☢️ Neutronics", ["proxy","enriched"], index=0 if fid.get("neutronics","proxy")=="proxy" else 1)
    profiles = st.selectbox("Profiles", ["off","analytic"], index=0 if fid.get("profiles","off")=="off" else 1)
    economics = st.selectbox("💰 Economics", ["proxy","enriched"], index=0 if fid.get("economics","proxy")=="proxy" else 1)
    st.session_state["fidelity_config"] = {
        "plasma": plasma,
        "magnets": magnets,
        "exhaust": exhaust,
        "neutronics": neutronics,
        "profiles": profiles,
        "economics": economics,
    }


with st.sidebar.expander("Reference calibration (optional)", expanded=False):
    st.caption("Transparent multiplicative factors (default 1.0). These are **declared knobs** (not a black‑box fit). "
               "They are recorded into artifacts for reviewer-safe reproducibility.")

    cL, cR = st.columns([1, 1])
    with cL:
        if st.button("Reset factors to 1.0", use_container_width=True):
            st.session_state["calib_confinement"] = 1.0
            st.session_state["calib_divertor"] = 1.0
            st.session_state["calib_bootstrap"] = 1.0
    with cR:
        st.markdown("**Scope**")
        st.markdown("- Confinement → τₑ / power balance")
        st.markdown("- Divertor → exhaust proxy severity")
        st.markdown("- Bootstrap → I_bs closure")

    st.session_state["calib_confinement"] = st.slider(
        "Confinement factor (H-like multiplier)",
        0.5, 1.5,
        float(st.session_state.get("calib_confinement", 1.0)),
        0.01,
        help="Multiplies the confinement closure used by the frozen evaluator. Use for reference calibration / sensitivity, not tuning for feasibility."
    )
    st.session_state["calib_divertor"] = st.slider(
        "Divertor factor (exhaust proxy multiplier)",
        0.5, 1.5,
        float(st.session_state.get("calib_divertor", 1.0)),
        0.01,
        help="Scales the exhaust/proxy model severity (e.g., q⊥-like gate). Recorded in artifacts."
    )
    st.session_state["calib_bootstrap"] = st.slider(
        "Bootstrap factor (I_bs multiplier)",
        0.5, 1.5,
        float(st.session_state.get("calib_bootstrap", 1.0)),
        0.01,
        help="Scales the bootstrap-current closure output prior to current-drive accounting. Recorded in artifacts."
    )


with st.sidebar.expander("Policy Contracts (feasibility semantics)", expanded=False):
    st.caption("Explicit, reviewer-visible enforcement tiering. This does **not** change physics outputs; it only changes whether selected limits are **blocking** or **diagnostic**.")
    _q95_prev = str(st.session_state.get("q95_enforcement","hard"))
    _fg_prev = str(st.session_state.get("greenwald_enforcement","hard"))
    st.session_state["q95_enforcement"] = st.selectbox(
        "q95 enforcement",
        ["hard", "diagnostic"],
        index=0 if str(st.session_state.get("q95_enforcement","hard")).lower().strip()=="hard" else 1,
        help="hard: blocking feasibility gate. diagnostic: computed and reported, but non-blocking (soft).",
    )
    st.session_state["greenwald_enforcement"] = st.selectbox(
        "Greenwald (fG) enforcement",
        ["hard", "diagnostic"],
        index=0 if str(st.session_state.get("greenwald_enforcement","hard")).lower().strip()=="hard" else 1,
        help="hard: blocking feasibility gate. diagnostic: computed and reported, but non-blocking (soft).",
    )
    st.markdown("**Contract**")
    st.markdown("- These settings affect constraint tiering only (HARD vs SOFT).")
    st.markdown("- They are recorded in artifacts under `_policy_contract`.")
    st.markdown("- No hidden iteration, no softening of physics.")

    if (str(st.session_state.get("q95_enforcement","hard")) != _q95_prev) or (str(st.session_state.get("greenwald_enforcement","hard")) != _fg_prev):
        _invalidate_mode_caches("policy_contract_changed")


with st.sidebar.expander("Technology Readiness (TRL Contracts)", expanded=False):
    st.caption("Explicit maturity tiering for governance and evidence packs. This does **not** re-solve physics; it records assumption tier and suggested caps for optional constraints.")
    _tier_prev = str(st.session_state.get("tech_tier", "TRL7"))
    _tiers = ["TRL3", "TRL5", "TRL7", "TRL9"]
    _tier = str(st.session_state.get("tech_tier", "TRL7")).upper().strip()
    if _tier not in _tiers:
        _tier = "TRL7"
    st.session_state["tech_tier"] = st.selectbox(
        "Technology readiness tier",
        _tiers,
        index=_tiers.index(_tier),
        help="Used to label maturity assumptions and (optionally) suggest default caps such as f_recirc_max and TBR_min. Recorded in outputs as _maturity_contract.",
    )
    try:
        from contracts.tech_tiers import suggested_defaults  # type: ignore
        _sug = dict(suggested_defaults(str(st.session_state.get("tech_tier","TRL7"))))
    except Exception:
        _sug = {}
    if _sug:
        with st.expander("Suggested defaults (optional)", expanded=False):
            st.json(_sug)
            st.caption("These are *suggestions* for optional caps; SHAMS truth is unchanged unless you explicitly apply them to inputs.")
    if str(st.session_state.get("tech_tier","TRL7")) != _tier_prev:
        _invalidate_mode_caches("tech_tier_changed")

with st.sidebar.expander("Benchmark Vault", expanded=False):
    tabs = st.tabs(["🏛️ Presets", "📦 Benchmarks"])
    with tabs[0]:
        st.caption("Load frozen, reviewer-safe reference machines into the workspace. Presets do **not** modify physics - they set inputs.")
        try:
            _ref_catalog = reference_catalog()
            _ref_keys = sorted(list(_ref_catalog.keys()))
        except Exception:
            _ref_catalog = {}
            _ref_keys = []

        _legacy_names = list(REFERENCE_MACHINES.keys())
        _choices = _ref_keys if _ref_keys else _legacy_names

        _sel = st.selectbox("Preset", _choices, index=0 if _choices else None)
        if _sel and _sel in _ref_catalog:
            _suite = str(_ref_catalog[_sel].get('suite','n/a'))
            _cls = str(_ref_catalog[_sel].get('class','n/a'))
            # User preference: keep details collapsed by default; only user expands.
            with st.expander(f"Suite: {_suite} · Class: {_cls}", expanded=False):
                st.code(_shams_json_dumps(_ref_catalog[_sel], indent=2), language="json")
        elif _sel and _sel in REFERENCE_MACHINES:
            st.markdown("**Legacy preset** (inline table).")
        if st.button("Load preset", use_container_width=True, disabled=not bool(_sel)):
            try:
                if _sel in _ref_catalog:
                    apply_reference_preset(_sel)
                else:
                    apply_legacy_reference_machine(_sel)
                st.success("Preset loaded into workspace inputs.")
            except Exception as e:
                st.error(f"Preset load failed: {e}")

    with tabs[1]:
        st.caption("Benchmark packs are deterministic evidence generators. Use **Publication Benchmarks** for full packs.")
        st.markdown("**Quick actions**")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Show last benchmark pack (if any)", use_container_width=True):
                _p = st.session_state.get("publication_benchmark_last_outdir")
                if _p:
                    st.code(str(_p))
                else:
                    st.info("No benchmark pack recorded in this session yet.")
        with c2:
            st.markdown("Tip: Run ` Publication Benchmarks → Generate Pack` for reviewer-safe CSV+JSON+hashes.")
# (Benchmark Vault rendered above)
_activity_log_sidebar_placeholder = st.sidebar.empty()


def _render_activity_log_sidebar() -> None:
    try:
        _lg = _activity_logger()
        with _activity_log_sidebar_placeholder.container():
            st.markdown("### Black-Box Chronicle")
            st.toggle(
                "Auto-log (recommended)",
                value=bool(st.session_state.get("activity_log_auto", True)),
                key="activity_log_auto",
            )
            _tail = st.number_input(
                "Show last N lines",
                min_value=50,
                max_value=2000,
                value=int(st.session_state.get("activity_log_tail", 200)),
                step=50,
                key="activity_log_tail",
            )

            # Always read from disk so the view reflects everything logged,
            # even if it happened later in this same run.
            try:
                tail_text = _lg.path.read_text(encoding="utf-8", errors="replace") if _lg.path.exists() else ""
            except Exception:
                tail_text = ""
            if int(_tail) > 0 and tail_text:
                tail_lines = tail_text.splitlines()[-int(_tail) :]
                tail_text = "\n".join(tail_lines)

            st.text_area(
                "Log (tail)",
                value=tail_text,
                height=220,
                key="activity_log_view",
                disabled=True,
            )

            _c1, _c2 = st.columns(2)
            with _c1:
                st.download_button(
                    "Download log",
                    data=(tail_text + "\n" if tail_text else ""),
                    file_name="activity.log",
                    mime="text/plain",
                    use_container_width=True,
                    key="activity_log_download",
                )
            with _c2:
                if st.button("Clear log", use_container_width=True, key="activity_log_clear_btn"):
                    try:
                        _alog("UI", "ClearLog", {})
                    except Exception:
                        pass

                    # Clear log on disk (authoritative: reflect immediately).
                    try:
                        _lg.clear()
                    except Exception:
                        pass
                    try:
                        st.session_state["activity_log_view"] = ""
                    except Exception:
                        pass
                    st.rerun()

            # -------------------------------------------------------------------
            # Session shutdown (Exit SHAMS) — MUST be visible in UI
            # -------------------------------------------------------------------
            st.markdown("---")
            st.markdown("### Session shutdown")
            _exit_confirm = st.checkbox(
                "Confirm exit",
                value=bool(st.session_state.get("shams_exit_confirm", False)),
                key="shams_exit_confirm",
                help="Safety latch to prevent accidental shutdown.",
            )
            if st.button(
                "🔴 Exit SHAMS",
                type="primary",
                use_container_width=True,
                disabled=not bool(_exit_confirm),
                key="shams_exit_btn",
                help="Hard-exit Streamlit process. Only reliable cross-platform shutdown.",
            ):
                try:
                    _alog("UI", "ExitRequested", {})
                except Exception:
                    pass
                st.info("SHAMS UI shutdown requested by user.")
                _os._exit(0)
    except Exception:
        # Never block UI.
        return

# Render once immediately so the panel is visible even if a downstream
# st.stop() occurs (some panels intentionally stop execution). We re-render
# again at end-of-file to include same-run events.
try:
    _render_activity_log_sidebar()
except Exception:
    pass

# ---------------------------------------------------------------------------
# Forward-definition bootstrap
#
# This UI file contains many panel functions defined *after* the main routing
# section. Streamlit executes top-to-bottom; if a panel is requested before its
# def is reached, it will appear as "not found". We proactively bootstrap the
# panel defs that are referenced by contracts/layer registry.
# ---------------------------------------------------------------------------
try:
    from ui.layer_registry import get_layer_registry
    _reg = get_layer_registry()
    _panel_names = []
    for _layer in _reg.get("layers", []):
        for _p in _layer.get("panels", []):
            n = _p.get("fn") or _p.get("panel_fn") or _p.get("fn_name")
            if isinstance(n, str) and n.startswith("_v"):
                _panel_names.append(n)
    # Also include all contract-registered panels.
    try:
        from ui.panel_contracts import get_panel_contracts
        _panel_names.extend(list(get_panel_contracts().keys()))
    except Exception:
        pass
    _panel_names = _unique(_panel_names)
    if _panel_names:
        _bootstrap_forward_defs(_panel_names)
except Exception:
    # Never block UI on bootstrap failure.
    pass


# --- Deck Navigation (v372.2 hotfix) ---
# Streamlit tabs reset to the first tab on reruns, which can cause the UI to "jump" back
# to Point Designer when interacting with other decks (and can unbind solver parameter
# names on button-click reruns). We therefore use a deterministic, persisted deck selector.
# Deck IDs are plain strings (UI redesign: emoji removed from navigation and
# routing). The sidebar radio stores the selected label directly; nav_deck_index
# (integer) persists selection across reruns. All `if _deck == "<plain>"` routing
# comparisons throughout app.py use these same plain IDs.
_DECK_LABELS = [
    "Point Designer",
    "Systems Mode",
    "Scan Lab",
    "Pareto Lab",
    "Trade Study Studio",
    "Reactor Design Forge",
    "System Suite",
    "Compare",
    "Publication Benchmarks",
    "Control Room",
]
with st.sidebar:
    st.markdown("## Navigation")
    _sel = st.radio(
        "Deck",
        _DECK_LABELS,
        index=int(st.session_state.get("nav_deck_index", 0)),
        label_visibility="collapsed",
        key="nav_deck_label",
    )
    st.session_state["nav_deck_index"] = _DECK_LABELS.index(_sel)
_deck = _sel

# Phase-1 cache aliasing (no compute, no truth mutation)
try:
    _phase1_stabilize_cache_aliases()
except Exception:
    pass




if _deck == "System Suite":
    from ui.decks.system_suite import render_system_suite
    render_system_suite(sys.modules[__name__])




if _deck == "Publication Benchmarks":
    from ui.decks.publication_benchmarks import render_publication_benchmarks
    render_publication_benchmarks(sys.modules[__name__])

if _deck == "Control Room":
    st.header("Control Room")
    st.caption("Governance, provenance, exports, and expert diagnostics - organized as compact decks (no scroll walls).")

    deck_orient, deck_const, deck_prov, deck_art, deck_diag, deck_chron = st.tabs([
        "🧭 Orientation",
        "🧾 Constitution",
        "🔍 Provenance",
        "📦 Artifacts",
        " Diagnostics",
        "🕓 Chronicle",
    ])

    with deck_orient:
        st.subheader("Orientation")
        st.caption("Quick-start workflows and reviewer-facing scope anchors (UI-only; does not modify truth).")
        o_launch, o_vocab, o_gallery, o_scope = st.tabs(["Launchpad", "Vocabulary", "Reference Gallery", "Scope"])

        with o_launch:
            st.subheader("Launchpad - First 30 Minutes")
            st.caption("A guided entry path for fusion experts: choose intent, then follow a minimal, honest workflow. UI scaffolding only.")
            _path = st.radio(
                "I want to…",
                [
                    "Understand feasibility limits (cartography)",
                    "Explore reactor concepts (Forge)",
                    "Review a finished case (Review Room)",
                    "Compare designs (Artifacts)",
                ],
                index=0,
                key="launchpad_path",
                horizontal=False,
            )
            if _path == "Understand feasibility limits (cartography)":
                st.info("Recommended: Scan Lab → build Scan Atlas → inspect first-failure topology.", icon="🗺️")
                st.markdown("""
- Start with **Scan Lab - Cartography Deck**
- Choose a compact 2D scan
- Export the Scan Atlas capsule for review-room replay.
""")
            elif _path == "Explore reactor concepts (Forge)":
                st.info("Recommended: Reactor Design Forge → Casebook → Candidate Archive → Machine Dossier.", icon="⚒️")
                st.markdown("""
- Use **Forge Cockpit** with the **Helm Console**
- Keep **Margins-first** framing
- Save capsules for deterministic replay.
""")
            elif _path == "Review a finished case (Review Room)":
                st.info("Recommended: Reactor Design Forge → Review Mode → Review Trinity → Do‑Not‑Build Brief.", icon="🧾")
                st.markdown("""
- Turn on **Review Mode**
- Use **Review Trinity** and **Conflict Atlas**
- Generate a **Reviewer Packet**.
""")
            else:
                st.info("Recommended: Compare → upload two artifacts → inspect deltas.")
                st.markdown("""
- Use **Compare artifacts** to check reproducibility
- Prefer capsule replay over manual edits.
""")

        with o_vocab:
            st.subheader("Vocabulary Ledger")
            st.caption("Fusion-native terminology mapping (SHAMS ↔ common literature ↔ PROCESS-style language).")
            try:
                _vocab = (Path(__file__).resolve().parent.parent / "docs" / "VOCABULARY_LEDGER.md").read_text(encoding="utf-8")
            except Exception:
                _vocab = "(missing docs/VOCABULARY_LEDGER.md)"
            st.markdown(_vocab)

        with o_gallery:
            st.subheader("Reference Study Gallery")
            st.caption("Recognizable anchors for the community. These are reference contexts, not targets.")
            _gallery = [
                ("ITER-like", "Large, conservative, physics-demonstration anchor; often stress and divertor constraints dominate."),
                ("SPARC-like", "Compact high-field concept; often HTS margin and structural stress dominate."),
                ("ARC-like", "HTS reactor class; often net-electric closure and blanket/TBR proxies dominate."),
                ("DEMO-like", "Plant realism anchor; often recirculating power and availability assumptions dominate."),
            ]
            for name, note in _gallery:
                with st.expander(name, expanded=False):
                    st.write(note)
            st.info("Tip: use these as *discussion anchors* when presenting SHAMS outputs to reviewers.", icon="💡")

        with o_scope:
            st.subheader("Model Scope Card")
            st.caption("Always-visible scope declaration for review rooms.")
            try:
                _scope = (Path(__file__).resolve().parent.parent / "docs" / "MODEL_SCOPE_CARD.md").read_text(encoding="utf-8")
            except Exception:
                _scope = "(missing docs/MODEL_SCOPE_CARD.md)"
            st.markdown(_scope)

    with deck_const:
        st.subheader("Constitution")
        st.caption("Frozen truth boundary, constraint constitution, and assumption ledger (read-only).")
        c_model, c_pcm, c_assump, c_docs, c_cc, c_ci, c_cp = st.tabs([
            "Model Ledger",
            "Capability Matrix",
            "Assumptions",
            "Docs",
            "Constraint Cockpit",
            "Constraint Inspector",
            "Constraint Provenance",
        ])
        tab_model = c_model
        tab_pcm = c_pcm
        tab_assumptions = c_assump
        tab_docs = c_docs
        tab_constraints = c_cc
        tab_constraint_inspector = c_ci
        tab_cprov = c_cp

    with deck_prov:
        st.subheader("Provenance")
        st.caption("Study protocol, repro lock, regression visibility, and replay tools.")
        p_studies, p_deck, p_auth, p_dec, p_dom, p_epoch, p_delta, p_regress, p_dash = st.tabs([
            "Studies",
            "Case Deck Runner",
            "Authority & Confidence",
            "Decision Consequences",
            "Authority Dominance",
            "Epoch Feasibility",
            "Scenario Delta",
            "Regression Viewer",
            "Study Dashboard",
        ])
        tab_studies = p_studies
        tab_deck = p_deck
        tab_authority_conf = p_auth
        tab_decision_conseq = p_dec
        tab_authority_dominance = p_dom
        tab_epoch_feas = p_epoch
        tab_delta = p_delta
        tab_regress = p_regress
        tab_study_dash = p_dash

    with deck_art:
        st.subheader("Artifacts")
        st.caption("Exports, evidence packs, and benchmark bundles (deterministic).")
        a_art, a_lib, a_export, a_bench = st.tabs([
            "Artifacts Explorer",
            "Run Library",
            "Export / Share",
            "Benchmarks",
        ])
        tab_artifacts = a_art
        tab_library = a_lib
        tab_export = a_export
        tab_bench = a_bench

    with deck_diag:
        st.subheader("🩺 Diagnostics")
        st.caption("Deep tools for debugging and reviewer verification (kept off the main workflow by default).")
        d_pam, d_val, d_comp, d_gate, d_nonfeas, d_solver, d_decision, d_session = st.tabs([
            "Panel Map",
            "Validation",
            "Compliance",
            "Gatechecks",
            "Non-Feasibility Guide",
            "Solver Introspection",
            "Decision Builder",
            "Session",
        ])
        tab_pam = d_pam
        tab_validation = d_val
        tab_compliance = d_comp
        tab_gatechecks = d_gate
        tab_nonfeas = d_nonfeas
        tab_solver = d_solver
        tab_decision = d_decision

        with d_gate:
            st.subheader("Gatechecks")
            st.caption("Local build integrity checks. UI-only; does not modify truth.")
            st.markdown("""
Run these from a terminal at the repo root:

- `python -m compileall -q .`
- `pytest -q`
- `streamlit run ui/app.py`

This panel also performs a lightweight hygiene scan of the working tree.
""")

            from pathlib import Path as _Path
            _root = (_Path(__file__).resolve().parent.parent)
            _forbidden = [
                '__pycache__',
                '.pytest_cache',
                'gspulse_ui',
            ]
            _hits = []
            for name in _forbidden:
                for h in _root.rglob(name):
                    _hits.append(str(h))
            # Also flag stray run_st* launchers
            for h in _root.glob('run_st*'):
                _hits.append(str(h))
            if _hits:
                st.error("Hygiene violations detected (should be removed before packaging):")
                with st.expander("Show paths", expanded=False):
                    for h in sorted(set(_hits)):
                        st.write(h)
            else:
                st.success("No hygiene violations detected in this tree.")

            st.divider()
            st.subheader("Interoperability self-check")
            st.caption("Verifies that main panels can exchange the canonical design state (no truth modifications).")

            st.divider()
            st.subheader("Interoperability contract validator (v326)")
            st.caption("Static + runtime wiring audit: declared panel contracts vs discoverable subpanels in ui/app.py.")

            def _run_contract_validator() -> dict:
                """Deterministic contract validator.

                Does not run physics or mutate truth.
                """
                from pathlib import Path as _Path
                from ui.panel_contracts import get_panel_contracts
                from tools.interoperability.contract_validator import validate_ui_contracts

                _root = (_Path(__file__).resolve().parent.parent)
                _contracts = get_panel_contracts()
                return validate_ui_contracts(_root, _contracts, session_state=dict(st.session_state))

            if st.button('Run contract validator', use_container_width=True, key='v326_contract_validator_btn'):
                st.session_state['v326_last_contract_validator_report'] = _run_contract_validator()

            _cr = st.session_state.get('v326_last_contract_validator_report')
            if isinstance(_cr, dict):
                if bool(_cr.get('ok')):
                    st.success('Contract validator: OK')
                else:
                    st.warning('Contract validator: issues detected')
                with st.expander('Contract validator report', expanded=False):
                    st.json(_cr)

            def _interop_check() -> dict:
                """Lightweight, deterministic UI-state interoperability audit.

                This is intentionally conservative: it only checks for existence and
                basic schema/type sanity of the canonical promotion keys. It does not
                run physics, solvers, or optimization.
                """
                rep = {'ok': True, 'checks': []}
                def _add(name: str, ok: bool, detail: str = ''):
                    rep['checks'].append({'name': name, 'ok': bool(ok), 'detail': str(detail)})
                    if not ok:
                        rep['ok'] = False

                # Core canonical artifacts used across modes
                for k in ['workspace_candidate', 'last_point_result', 'compare_left', 'compare_right']:
                    _add(f'session_key:{k}', k in st.session_state, 'present' if k in st.session_state else 'missing')

                # Systems mode canonical keys
                _t = st.session_state.get('systems_targets')
                _v = st.session_state.get('systems_variables')
                _add('systems_targets_type', isinstance(_t, dict) and len(_t) > 0, f"type={type(_t).__name__} len={len(_t) if isinstance(_t, dict) else 'n/a'}")
                _add('systems_variables_type', isinstance(_v, dict) and len(_v) > 0, f"type={type(_v).__name__} len={len(_v) if isinstance(_v, dict) else 'n/a'}")

                # Evidence / provenance hooks
                for k in ['last_precheck_report', 'last_systems_solution', 'last_evidence_pack_path', 'provenance_global']:
                    _add(f'provenance_key:{k}', k in st.session_state, 'present' if k in st.session_state else 'missing')

                return rep

            if st.button('Run interoperability check', use_container_width=True, key='v323_interop_check_btn'):
                st.session_state['v323_last_interop_report'] = _interop_check()

            _ir = st.session_state.get('v323_last_interop_report')
            if isinstance(_ir, dict):
                if bool(_ir.get('ok')):
                    st.success('Interoperability check: OK')
                else:
                    st.warning('Interoperability check: issues detected')
                with st.expander('Interoperability report', expanded=False):
                    st.json(_ir)

        with d_session:
            with st.expander("Session state (debug)", expanded=False):
                st.write({k: type(v).__name__ for k, v in st.session_state.items()})
            with st.expander("Version", expanded=False):
                try:
                    st.code((BASE_DIR / "VERSION").read_text().strip())
                except Exception:
                    st.code("unknown")

    with deck_chron:
        st.subheader("Chronicle")
        st.caption("Expert instruments and exploration aids (read-only; never modifies truth).")
        ch_reg, ch_sens, ch_knobs, ch_fmap, ch_mat, ch_maint, ch_prof, ch_imp, ch_disr, ch_stab, ch_solve, ch_repair, ch_refine, ch_narrow, ch_surr, ch_al = st.tabs([
            "Variable Registry",
            "Sensitivity Explorer",
            "Knob Trade-Space",
            "Feasibility Map",
            "Maturity Heatmap",
            "Maintenance & Availability",
            "Profile Authority",
            "Impurity & Radiation",
            "Disruption Risk",
            "Stability Risk",
            "Certified Search",
            "Repair Suggestions",
            "Interval Refinement",
            "Interval Narrowing",
            "Surrogate Overlay",
            "Active Learning",
        ])
        tab_registry = ch_reg
        tab_sensitivity = ch_sens
        tab_knobs = ch_knobs
        tab_feasmap = ch_fmap
        tab_maturity = ch_mat
        tab_maintenance = ch_maint
        tab_profile_auth = ch_prof
        tab_impurity = ch_imp
        tab_disruption = ch_disr
        tab_stability = ch_stab
        tab_cert_search = ch_solve
        tab_repair = ch_repair
        tab_refine = ch_refine
        tab_narrowing = ch_narrow
        tab_surrogate = ch_surr
        tab_active_learning = ch_al

    # Populate Control Room sections (ensure they are never empty)
    with tab_pam:
        _v175_panel_availability_map_panel()
    
    with tab_studies:
        st.markdown("### Study authority & publishability")
        st.write("Generate protocol → lock/replay → authority pack → citation → export.")
        try:
            _render_with_contract("_v165_study_protocol_panel", _v165_study_protocol_panel)
        except Exception:
            st.info('Panel unavailable in this build.')
        st.divider()
        try:
            _render_with_contract("_v166_repro_lock_panel", _v166_repro_lock_panel)
        except Exception:
            st.info('Panel unavailable in this build.')
        st.divider()
        try:
            _render_with_contract("_v167_authority_pack_panel", _v167_authority_pack_panel)
        except Exception:
            st.info('Panel unavailable in this build.')
        st.divider()
        try:
            _render_with_contract("_v168_citation_panel", _v168_citation_panel)
        except Exception:
            st.info('Panel unavailable in this build.')
        st.divider()
        try:
            _render_with_contract("_v170_process_export_panel", _v170_process_export_panel)
        except Exception:
            st.info('Panel unavailable in this build.')
    st.divider()
    with st.expander("🗃️ Studies manager", expanded=False):
        st.header("Studies manager")
        st.write("Save, load, and organize study configurations (scan/pareto) as JSON. This keeps studies reproducible across sessions.")
        if "studies" not in st.session_state:
            st.session_state.studies = []

        c1, c2, c3 = st.columns(3)
        with c1:
            if st.button("Save current PointInputs as study", use_container_width=True):
                if st.session_state.get("last_point_inp") is not None:
                    try:
                        inp_obj = st.session_state.last_point_inp
                        # dataclass -> dict
                        d = {k: getattr(inp_obj, k) for k in inp_obj.__dataclass_fields__.keys()}  # type: ignore
                        st.session_state.studies.append({"type": "point", "created": datetime.datetime.now().isoformat(), "inputs": d})
                        st.success("Saved.")
                    except Exception as e:
                        st.error(f"Could not save: {e}")
                else:
                    st.warning("Run a point first so `last_point_inp` exists.")

        with c2:
            up = st.file_uploader("Import studies JSON", type=["json"], key="studies_import")
            if up is not None:
                try:
                    imported = json.loads(up.getvalue().decode("utf-8"))
                    if isinstance(imported, list):
                        st.session_state.studies.extend(imported)
                    elif isinstance(imported, dict):
                        st.session_state.studies.append(imported)
                    st.success("Imported.")
                except Exception as e:
                    st.error(f"Import failed: {e}")

        with c3:
            if st.session_state.studies:
                st.download_button(
                    "Download studies JSON",
                    data=json.dumps(st.session_state.studies, indent=2, sort_keys=True),
                    file_name="shams_studies.json",
                    mime="application/json",
                    use_container_width=True,
                )

        st.markdown("### Saved studies")
        if st.session_state.studies:
            df = pd.DataFrame([{"i": i, "type": s.get("type","?"), "created": s.get("created",""), "notes": s.get("notes","")} for i,s in enumerate(st.session_state.studies)])
            st.dataframe(df, use_container_width=True)
            idx = st.number_input("Select index to view", min_value=0, max_value=max(0, len(st.session_state.studies)-1), value=0, step=1)
            st.json(st.session_state.studies[int(idx)])
            if st.button("Delete selected", use_container_width=True):
                try:
                    st.session_state.studies.pop(int(idx))
                    st.experimental_rerun()
                except Exception:
                    pass
        else:
            st.info("No studies saved yet.")


    
    with tab_model:
        # Render with proper scientific notation (MathJax) to avoid “ASCII-looking” formulas.
        st.markdown(
            """
    This section documents the **0‑D (global) physics + engineering surrogate** used by SHAMS for rapid point design.
    It is intentionally transparent: the goal is to show the **model structure**, **assumptions**, and **where each number comes from**.
    
    #### Symbol key (as used below)
    - $R_0$ major radius, $a$ minor radius, $\\kappa$ elongation, $\\delta$ triangularity
    - $B_t$ toroidal field on axis, $I_p$ plasma current
    - $n$ density, $T$ temperature, $V$ plasma volume, $W$ stored energy
    - $P_{fus}$ fusion power, $P_{\\alpha}$ alpha power, $P_{aux}$ auxiliary power, $P_{SOL}$ power crossing the separatrix
    - $\\tau_E$ energy confinement time, $H$ confinement multiplier (H‑factor)
    """
        )
    
        st.markdown("#### High‑level flow (per point evaluation)")
        st.markdown(
            """
    1. **Geometry:** $(R_0, a, \\kappa, \\delta) \\rightarrow$ volumes/areas.
    2. **Plasma state:** choose targets/intent $\\rightarrow$ infer a consistent $(T, n, B_t, I_p)$ under constraints.
    3. **Power balance:** $P_{fus}, P_{\\alpha}, P_{aux}$ and losses $\\rightarrow$ steady‑state balance.
    4. **Confinement:** $\\tau_E$ from selected scaling (ITER98y2 / others) with $H$; enforce $Q$ consistency.
    5. **Current & stability:** $q_{95}$, $\\beta_N$, Greenwald fraction $f_G$.
    6. **Engineering proxies:** TF peak field / hoop stress, HTS margin.
    7. **Blanket/shield/TBR proxy:** thickness & coverage $\\rightarrow$ TBR screening.
    8. **Divertor proxy:** heat‑flux screening from $P_{SOL}$ and geometry.
    9. **Radial build closure:** inboard stack fits (gap + FW + blanket + shield + VV + TF).
    """
        )
    
        st.markdown("#### Core relationships (representative)")
        st.latex(
            r"""
    \begin{aligned}
    P_{fus} &\propto n^2\,\langle\sigma v\rangle(T)\,V \\
    \tau_E &= H\,\tau_{\mathrm{ITER98y2}}(I_p, B_t, n, P, R_0, a, \kappa, \ldots) \\
    P_{heat} &= P_{\alpha} + P_{aux} \\
    P_{loss} &\approx \frac{W}{\tau_E}\;\; (\text{plus radiation terms where enabled}) \\
    q_{95} &\approx \frac{5\,a^2\,B_t}{R_0\,I_p}\,f(\kappa,\delta) \\
    \beta_N &\approx \beta\,\frac{a\,B_t}{I_p} \\
    q_{div} &\approx \frac{P_{SOL}}{2\pi R_0\,\lambda_q}\,g_{exh}(\text{geometry}) \\
    \sigma_{TF} &\propto \frac{B_{peak}^2\,R_{coil}}{\mu_0}
    \end{aligned}
    """
        )
        st.caption(
            "These are screening/closure relationships to support feasibility-first iteration. Exact authoritative pass/fail logic lives in SHAMS constraints and margins."
        )
    
    
    with tab_pcm:
        st.markdown("### 📎 Physics Capability Matrix")
        st.caption(
            "Read-only audit map: subsystems → equations/closures → authority tier (proxy/parametric/external) → intended validity domain."
        )
        try:
            # v228+: prefer generator-derived snapshot if present (still read-only).
            p_gen = (BASE_DIR / "docs" / "PHYSICS_CAPABILITY_MATRIX_GENERATED.md")
            p_src = (BASE_DIR / "docs" / "PHYSICS_CAPABILITY_MATRIX.md")
            if p_gen.exists():
                _pcm = p_gen.read_text(encoding="utf-8", errors="ignore")
            else:
                _pcm = p_src.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            _pcm = "(missing docs/PHYSICS_CAPABILITY_MATRIX*.md)"
        st.markdown(_pcm)
        st.info(
            "Bluemira-inspired lessons are adopted for provenance and capability clarity - without introducing optimization loops or CAD-level coupling.",
            icon="🧭",
        )
    with tab_bench:
    
        st.markdown("### Benchmarks")
        st.write("Benchmark runners (validation/regression) are available via the advanced panels once you have run artifacts.")
    
        st.markdown("#### Reference superconducting tokamaks (quick lookup)")
        st.markdown(
            """
    | Tokamak | Country / Org | Status | SC type | Major R (m) | Minor a (m) | B₀ on axis (T) | Ip (MA) | Primary role |
    |---|---|---|---|---:|---:|---:|---:|---|
    | **ITER** | Intl (EU/JP/US/etc.) | Under construction | **Nb₃Sn / NbTi (LTS)** | 6.2 | 2.0 | 5.3 | 15 | Burning plasma, Q≈10 |
    | **JT-60SA** | Japan–EU | Commissioning | **NbTi (LTS)** | 2.96 | 1.18 | 2.25 | 5.5 | Advanced plasma physics |
    | **WEST** | France | Operating | **NbTi (LTS)** | 2.5 | 0.5 | 3.7 | ≤1 | Long-pulse, PFC/divertor |
    | **EAST** | China | Operating | **NbTi (LTS)** | 1.8–1.9 | 0.4–0.45 | ≤3.5 | ≤1 | Long-pulse operation |
    | **KSTAR** | Korea | Operating | **NbTi-based (LTS)** | ~1.8 | ~0.5 | 3.5 | ≤2 | Advanced tokamak scenarios |
    | **SST-1** | India | Operating | **NbTi (LTS)** | ~1.1 | ~0.2 | ≤3 | ≤0.1 | SC tokamak development |
    | **TRIAM-1M** | Japan | Historical | **Nb₃Sn (LTS)** | ~0.8 | ~0.12–0.18 | 8 | - | High-field SC operation |
    | **SPARC** | USA (MIT/CFS) | Under construction | **REBCO (HTS)** | 1.85 | 0.57 | 12.2 | 8.7 | Q>1, high-field compact |
    """
        )
        st.caption(
            "Values are typical/design-point numbers collected from public summaries. For rigorous comparison, cite primary machine parameter sheets."
        )
        st.write("Below is a quick reference table of major superconducting tokamaks used as comparison anchors.")
    
        try:
            import pandas as _pd
            _bench_rows = [
                {"Tokamak":"ITER","Country / Org":"Intl (EU/JP/US/etc.)","Status":"Under construction","SC type":"Nb₃Sn / NbTi (LTS)","Major R (m)":6.2,"Minor a (m)":2.0,"B₀ on axis (T)":5.3,"Ip (MA)":15.0,"Primary role":"Burning plasma, Q≈10"},
                {"Tokamak":"JT-60SA","Country / Org":"Japan–EU","Status":"Commissioning","SC type":"NbTi (LTS)","Major R (m)":3.0,"Minor a (m)":1.0,"B₀ on axis (T)":2.3,"Ip (MA)":5.5,"Primary role":"Advanced plasma physics"},
                {"Tokamak":"WEST","Country / Org":"France","Status":"Operating","SC type":"NbTi (LTS)","Major R (m)":2.5,"Minor a (m)":0.5,"B₀ on axis (T)":3.7,"Ip (MA)":1.0,"Primary role":"Long-pulse, PFC/divertor"},
                {"Tokamak":"EAST","Country / Org":"China","Status":"Operating","SC type":"NbTi (LTS)","Major R (m)":1.9,"Minor a (m)":0.5,"B₀ on axis (T)":3.5,"Ip (MA)":1.0,"Primary role":"Long-pulse operation"},
                {"Tokamak":"KSTAR","Country / Org":"Korea","Status":"Operating","SC type":"NbTi-based (LTS)","Major R (m)":1.8,"Minor a (m)":0.5,"B₀ on axis (T)":3.5,"Ip (MA)":2.0,"Primary role":"Advanced tokamak scenarios"},
                {"Tokamak":"SST-1","Country / Org":"India","Status":"Operating","SC type":"NbTi (LTS)","Major R (m)":1.1,"Minor a (m)":0.2,"B₀ on axis (T)":3.0,"Ip (MA)":0.1,"Primary role":"SC tokamak development"},
                {"Tokamak":"TRIAM-1M","Country / Org":"Japan","Status":"Historical","SC type":"Nb₃Sn (LTS)","Major R (m)":0.8,"Minor a (m)":0.15,"B₀ on axis (T)":8.0,"Ip (MA)":None,"Primary role":"High-field SC operation"},
                {"Tokamak":"HT-7","Country / Org":"China","Status":"Historical","SC type":"LTS","Major R (m)":1.22,"Minor a (m)":0.27,"B₀ on axis (T)":2.0,"Ip (MA)":0.2,"Primary role":"Precursor to EAST"},
                {"Tokamak":"SPARC","Country / Org":"USA (MIT/CFS)","Status":"Under construction","SC type":"REBCO (HTS)","Major R (m)":1.85,"Minor a (m)":0.57,"B₀ on axis (T)":12.2,"Ip (MA)":8.7,"Primary role":"Q>1, high-field compact"},
                {"Tokamak":"HH70","Country / Org":"China (Energy Singularity)","Status":"Operating","SC type":"REBCO (HTS)","Major R (m)":0.7,"Minor a (m)":0.28,"B₀ on axis (T)":0.6,"Ip (MA)":None,"Primary role":"Full-HTS integration demo"},
                {"Tokamak":"HH170","Country / Org":"China (Energy Singularity)","Status":"Planned","SC type":"REBCO (HTS)","Major R (m)":None,"Minor a (m)":None,"B₀ on axis (T)":None,"Ip (MA)":None,"Primary role":"Reactor-relevant HTS tokamak"},
            ]
            _df = _pd.DataFrame(_bench_rows)
            st.dataframe(_df, use_container_width=True, hide_index=True)
            st.caption("Notes: Some entries are approximate screening values (as shown). Replace with cited values if you enable web-backed references.")
        except Exception:
            st.info("Benchmark reference table unavailable in this environment.")
    
    with tab_docs:
        st.markdown("### Documentation")
        st.write("Offline docs are included in the package. Review-room and exposure guardrails are included as dedicated docs pages. Key references live in the `docs/` folder when present.")
        st.caption("Note: The **Model Ledger (0‑D Physics)** panel renders equations using LaTeX/MathJax for scientific typography.")
    
        try:
            from pathlib import Path as _P
    
            _readme = _P("README.md")
            if _readme.exists():
                with st.expander("README (excerpt)", expanded=False):
                    st.code(_readme.read_text(encoding="utf-8")[:3000])
    
            _docs_dir = _P("docs")
            _mds = []
            if _docs_dir.exists():
                _mds = sorted([pp for pp in _docs_dir.rglob("*.md") if pp.is_file()])
    
            if _mds:
                st.markdown("#### Docs library")
                _labels = [str(pp.relative_to(_docs_dir)) for pp in _mds]
                _sel = st.selectbox("Open a doc (read‑only)", _labels, index=0)
                _path = _docs_dir / _sel
                with st.expander(f"docs/{_sel}", expanded=False):
                    st.markdown(_path.read_text(encoding="utf-8"))
            else:
                st.info("No `docs/` folder was found in this build.")
        except Exception:
            pass
    
    with tab_artifacts:
        st.markdown("### Artifacts")
        st.write("Artifacts appear after you run Point Designer / Systems Mode.")
        st.write("Use Run Library / Export tools to download bundles.")
    
    # For remaining expanders, ensure a minimal non-empty body
    for _exp in [tab_registry, tab_validation, tab_compliance, tab_deck, tab_delta, tab_library, tab_constraints,
                tab_constraint_inspector, tab_sensitivity, tab_feasmap, tab_decision, tab_nonfeas, tab_cprov,
                tab_knobs, tab_regress, tab_study_dash, tab_maturity, tab_assumptions, tab_export, tab_solver]:
        with _exp:
            st.write("This tool becomes active when required upstream artifacts exist (run history, packs, or reports).")
            st.write("If you need something here, run a study first, then return to More.")
    
# Shared state
if "last_point_out" not in st.session_state:
    st.session_state["last_point_out"] = None
if "last_point_inp" not in st.session_state:
    st.session_state["last_point_inp"] = None
if "scan_df" not in st.session_state:
    st.session_state.scan_df = pd.DataFrame()
if "scan_meta" not in st.session_state:
    st.session_state.scan_meta = {}
if "studies" not in st.session_state:
    st.session_state.studies = []  # list of study config dicts
if "compare_artifacts" not in st.session_state:
    st.session_state.compare_artifacts = {"A": None, "B": None}

# -----------------------------
# Point Designer
# -----------------------------
if _deck == "Point Designer":
    from ui.decks.point_designer import render_point_designer
    render_point_designer(sys.modules[__name__])
if _deck == "Systems Mode":
    from ui.decks.systems_mode import render_systems_mode
    render_systems_mode(sys.modules[__name__])

if _deck == "Scan Lab":
    # DSG: auto edge-kind tagging by active panel (exploration only)
    if bool(st.session_state.get("dsg_edge_kind_auto", True)):
        st.session_state["dsg_context_edge_kind"] = "scan"

    st.header("Scan Lab")
    st.caption("Cartography over the frozen evaluator: map feasibility, emptiness, fragility, and dominant mechanisms. Deterministic; no internal optimizer.")
    render_mode_scope("scan")

    # --- World-class Scan Lab (v188) ---
    # NOTE: Scan Lab should remain usable even if optional features fail to import.
    # Import errors are captured and surfaced explicitly (freeze-readiness requirement).
    _scan_import_errors = []

    try:
        # Fix: evaluator lives under src/ in the merged repo layout.
        from src.evaluator.core import Evaluator  # type: ignore
    except Exception as _e:
        Evaluator = None  # type: ignore
        _scan_import_errors.append(f"Evaluator import failed: {_e}")

    try:
        from tools.scan_cartography import build_cartography_report
    except Exception as _e:
        build_cartography_report = None  # type: ignore
        _scan_import_errors.append(f"scan_cartography import failed: {_e}")

    try:
        from tools.golden_scans import build_golden_scan_presets
    except Exception as _e:
        build_golden_scan_presets = None  # type: ignore
        _scan_import_errors.append(f"golden_scans import failed: {_e}")

    try:
        from tools.canonical_questions import build_canonical_questions
    except Exception as _e:
        build_canonical_questions = None  # type: ignore
        _scan_import_errors.append(f"canonical_questions import failed: {_e}")

    try:
        from tools.scan_insights import (
            build_causality_trace,
            uncertainty_stress_test,
            time_to_failure_along_knob,
            null_direction_2d,
        )
    except Exception as _e:
        build_causality_trace = None  # type: ignore
        uncertainty_stress_test = None  # type: ignore
        time_to_failure_along_knob = None  # type: ignore
        null_direction_2d = None  # type: ignore
        _scan_import_errors.append(f"scan_insights import failed: {_e}")

    try:
        from tools.scan_next_tier import (
            local_powerlaw_fit,
            label_regime,
            explain_impossible_region,
            detect_irrelevant_constraints,
            projection_stability_check,
            path_follow_scan,
            assumption_stress_hotspots,
            counterfactual_lens,
            guided_steps,
            build_scan_atlas_pdf_bytes,
            surprise_detector,
        )
    except Exception as _e:
        local_powerlaw_fit = None  # type: ignore
        label_regime = None  # type: ignore
        explain_impossible_region = None  # type: ignore
        detect_irrelevant_constraints = None  # type: ignore
        projection_stability_check = None  # type: ignore
        path_follow_scan = None  # type: ignore
        assumption_stress_hotspots = None  # type: ignore
        counterfactual_lens = None  # type: ignore
        guided_steps = None  # type: ignore
        build_scan_atlas_pdf_bytes = None  # type: ignore
        surprise_detector = None  # type: ignore
        _scan_import_errors.append(f"scan_next_tier import failed: {_e}")


    try:
        from tools.scan_v1p1_worldclass import (
            build_constraint_dictionary,
            build_reproducibility_capsule,
            monotonicity_sanity_overlay,
            boundary_thickness_metric,
            explain_uncertainty_disagreement,
            to_json_bytes,
        )
    except Exception as _e:
        build_constraint_dictionary = None  # type: ignore
        build_reproducibility_capsule = None  # type: ignore
        monotonicity_sanity_overlay = None  # type: ignore
        boundary_thickness_metric = None  # type: ignore
        explain_uncertainty_disagreement = None  # type: ignore
        to_json_bytes = None  # type: ignore
        _scan_import_errors.append(f"scan_v1p1_worldclass import failed: {_e}")

    try:
        from tools.scan_expert_features import (
            SCAN_LAB_CONTRACT,
            compute_fingerprints,
            ScanClaim,
            build_claim_evidence,
            build_claim_pdf_bytes,
            falsify_claim,
        )
    except Exception as _e:
        SCAN_LAB_CONTRACT = ""  # type: ignore
        compute_fingerprints = None  # type: ignore
        ScanClaim = None  # type: ignore
        build_claim_evidence = None  # type: ignore
        build_claim_pdf_bytes = None  # type: ignore
        falsify_claim = None  # type: ignore
        _scan_import_errors.append(f"scan_expert_features import failed: {_e}")

    try:
        from tools.reports.scan_signature_atlas import build_signature_atlas_pdf_bytes
    except Exception as _e:
        build_signature_atlas_pdf_bytes = None  # type: ignore
        _scan_import_errors.append(f"scan_signature_atlas import failed: {_e}")

    try:
        from tools.design_family_governance_v394 import build_design_families_from_scan_cartography
    except Exception as _e:
        build_design_families_from_scan_cartography = None  # type: ignore
        _scan_import_errors.append(f"design_family_governance_v394 import failed: {_e}")

    try:
        from tools.scan_artifact_schema import (
            build_scan_artifact,
            upgrade_scan_artifact,
            SCAN_SCHEMA_VERSION,
            stable_hash,
        )
    except Exception as _e:
        build_scan_artifact = None  # type: ignore
        upgrade_scan_artifact = None  # type: ignore
        SCAN_SCHEMA_VERSION = 0  # type: ignore
        stable_hash = None  # type: ignore
        _scan_import_errors.append(f"scan_artifact_schema import failed: {_e}")

    # Persist import errors so the UI can show them near the run buttons.
    st.session_state["scan_import_errors"] = _scan_import_errors

    def _v188_scan_lab_panel() -> None:
        import streamlit as st
        import numpy as np
        import pandas as pd
        import io
        import matplotlib.pyplot as plt
        from matplotlib.colors import ListedColormap
        import matplotlib.patches as mpatches

        st.subheader("🗺️ Cartography Deck")
        _scan_deck = st.radio(
            "Scan Lab deck",
            options=["🗺️ Cartography", "🧰 Orientation"],
            index=0,
            horizontal=True,
            key="scan_deck",
            help="Deck-based navigation: render one Scan workspace at a time (no scroll walls).",
        )
        if _scan_deck == "🧰 Orientation":
            with st.expander("Orientation Deck (optional, read-only helpers)", expanded=False):
                st.caption("These helpers are read-only. Switch back to 🗺️ Cartography to run scans.")
                st.info("Scan Lab orientation only (no scan executed).", icon="🧭")
            st.stop()

        st.info("Scan Lab is **frozen**: deterministic cartography/interpretability over the frozen evaluator. No optimization, no relaxation, no recommendations.", icon="🧊")

        st.caption("A microscope, not an engine: map feasibility structure and failure mechanisms across a parameter space.")

        # --- Pre-Cartography orientation (freeze-compliant; UI-only) ---
        with st.expander("Orientation Deck (optional, read-only helpers)", expanded=False):
            st.caption("These sections are optional. Cartography controls are below.")
            st.markdown("### Orientation Deck (read-only)")
            st.caption("These helpers are **read-only**. They do not change the evaluator, physics, or scan results.")
    
            with st.expander(" Scan Lab is frozen - freeze statement (read-only)", expanded=False):
                c1, c2 = st.columns([1.0, 1.0])
                with c1:
                    st.markdown("**Freeze statement**: `docs/SCANLAB_FREEZE.md`")
                    st.caption("Semantics are frozen; Scan Lab is cartography/insight only.")
                with c2:
                    try:
                        from pathlib import Path
                        _sf = (Path(__file__).resolve().parent.parent / "docs" / "SCANLAB_FREEZE.md").read_text(encoding="utf-8")
                    except Exception:
                        _sf = "(missing docs/SCANLAB_FREEZE.md)"
                    st.download_button(
                        "Download freeze statement",
                        data=_sf,
                        file_name="SCANLAB_FREEZE.md",
                        mime="text/markdown",
                        use_container_width=False,
                        key="scan_freeze_dl_pre",
                    )
                    with st.expander("View freeze statement", expanded=False):
                        st.markdown(_sf)
    
            # --- Legacy scan + stateful download (kept for backward compatibility) ---
            with st.expander("Legacy Grid Scan + stateful download (archived)", expanded=False):
                try:
                    _v93_stateful_scan_panel()
                except Exception:
                    pass
                st.markdown(
                    """
    This legacy scan performs a **nested solver grid** over (Ti, H98, a, Q, g_conf). It remains available for older workflows,
    but the recommended Scan Lab path is now the **Cartography** scan below.
                    """
                )
    
            with st.expander("Parameter guide (units, meaning, min/max)", expanded=False):
                st.markdown(
                    """
    Below are the **recommended** ranges used for input validation in this UI.  
    They are intentionally broad to avoid over‑constraining early exploration.
    
    | Parameter | Meaning | Recommended min | Recommended max |
    |---|---|---:|---:|
    | R₀ (m) | Major radius | 0.5 | 10 |
    | B₀ (T) | Toroidal field on axis | 1 | 25 |
    | Shield (m) | Neutron shield thickness | 0 | 2 |
    | P_aux (MW) | External heating power | 0 | 200 |
    | P_aux for Q (MW) | Power used in Q = P_fus / P_aux_for_Q | 0 | 200 |
    | Tᵢ/Tₑ (–) | Ion/electron temperature ratio | 0.5 | 5 |
    | Ti_start/stop (keV) | Ion temperature scan bounds | 1 | 40 |
    | Ti_step (keV) | Ion temperature step | 0.05 | 5 |
    | H98_start/stop (–) | H98y2 confinement multiplier bounds | 0.5 | 3 |
    | H98_step (–) | H98 step | 0.01 | 0.5 |
    | a_min/a_max (m) | Minor radius scan bounds | 0.2 | 5 |
    | a_step (m) | Minor radius step | 0.001 | 1 |
    | Q_start/stop (–) | Target Q scan bounds (screening target) | 0.1 | 100 |
    | Q_step (–) | Q step | 0.05 | 20 |
    | g_conf start/stop (–) | Additional confinement gain factor | 0.5 | 5 |
    | g_conf step (–) | g_conf step | 0.01 | 1 |
    | Iₚ bounds (MA) | Solver search bounds for plasma current | 1 | 50 |
    | fG bounds (–) | Greenwald fraction screening bounds | 0.01 | 1.5 |
    | tol (–) | Numerical tolerance for the solver | 1e-6 | 1e-2 |
    | Zeff (–) | Effective charge | 1.0 | 4.0 |
    | dilution_fuel (–) | Fuel dilution factor (≤1) | 0.2 | 1.0 |
    | extra_rad_factor (–) | Extra radiation multiplier | 0 | 2 |
    | alpha_loss_frac (–) | Fraction of alpha power lost | 0 | 0.5 |
    | kappa (–) | Elongation | 1.0 | 3.0 |
    | q95_min (–) | Minimum q95 constraint | 1.5 | 10 |
    | betaN_max (–) | Maximum normalized beta constraint | 1.0 | 8 |
    | C_bs (–) | Bootstrap coefficient proxy | 0 | 1 |
    | f_bs_max (–) | Max bootstrap fraction | 0 | 1 |
    | PSOL/R max (MW/m) | SOL power per major radius limit | 0 | 200 |
    | PLH_margin (–) | Extra margin over PLH if H‑mode required | 0 | 1 |
                    """
                )
    
            with st.expander("Scan Lab → Physics block mapping (what each parameter affects)", expanded=False):
                st.caption("UI-only helper: shows which Phase‑1 physics/systems blocks each Scan Lab parameter feeds.")
                rows = []
                # Keep the ordering aligned with the form layout.
                ordered = [
                    ("R0", "Major radius R₀ (m)"),
                    ("B0", "Toroidal field on axis B₀ (T)"),
                    ("tshield", "Neutron shield thickness (m)"),
                    ("Paux", "Auxiliary heating power P_aux (MW)"),
                    ("Paux_for_Q", "Aux power used in Q definition (MW)"),
                    ("Ti_over_Te", "Ion-to-electron temperature ratio Tᵢ/Tₑ (–)"),
                    ("Ti", "Ti axis (Ti_start/stop/step)"),
                    ("H98", "H98 axis (H98_start/stop/step)"),
                    ("a", "a axis (a_min/a_max/a_step)"),
                    ("Q", "Q axis (Q_start/stop/Q_step)"),
                    ("g_conf", "g_conf axis (start/stop/step)"),
                    ("Ip_bounds", "Iₚ bounds (I_p,min / I_p,max)"),
                    ("fG_bounds", "fG bounds (fG_min / fG_max)"),
                    ("tol", "tol"),
                    ("Zeff", "Zeff"),
                    ("dilution_fuel", "dilution_fuel"),
                    ("extra_rad_factor", "extra_rad_factor"),
                    ("alpha_loss_frac", "alpha_loss_frac"),
                    ("kappa", "kappa"),
                    ("q95_min", "q95_min"),
                    ("betaN_max", "betaN_max"),
                    ("C_bs", "C_bs"),
                    ("f_bs_max", "f_bs_max"),
                    ("PSOL_over_R_max", "PSOL/R max"),
                    ("require_Hmode", "Require H-mode access"),
                    ("PLH_margin", "PLH_margin"),
                    ("tblanket_m", "Blanket thickness (inboard)"),
                    ("t_vv_m", "Vacuum vessel thickness (inboard)"),
                    ("t_gap_m", "Inboard gap / clearance"),
                    ("t_tf_struct_m", "TF structure thickness (inboard)"),
                    ("t_tf_wind_m", "TF winding pack thickness (inboard)"),
                    ("Bpeak_factor", "Bpeak_factor"),
                    ("sigma_allow_MPa", "Allowable coil hoop stress"),
                    ("Tcoil_K", "HTS operating temperature"),
                    ("hts_margin_min", "HTS margin min"),
                    ("Vmax_kV", "Max dump voltage limit"),
                    ("q_div_max_MW_m2", "Max divertor heat flux limit"),
                    ("TBR_min", "TBR_min"),
                    ("hts_lifetime_min_yr", "Minimum HTS lifetime"),
                    ("P_net_min_MW", "Minimum net electric power"),
                ]
                for k, label in ordered:
                    blocks = _scan_blocks(k)
                    rows.append(
                        {
                            "Parameter": label,
                            "Badge": _scan_badge(k),
                            "Physics blocks": ", ".join(blocks) if blocks else "(unmapped)",
                        }
                    )
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    
            # Teaching Mode (UI-only). Adds gentle guidance without changing results.
            teaching_mode = st.checkbox("Teaching Mode (adds gentle guidance)", value=bool(st.session_state.get("scan_teaching_mode", False)), key="scan_teaching_mode")
            if teaching_mode:
                st.info("Teaching Mode is ON: Scan Lab will add small callouts explaining what you’re seeing. No physics or results change.")
    
    
    
            st.markdown("### 🗺️ Cartography Deck - mode contract (is / is not)")
            cA, cB = st.columns(2)
            with cA:
                st.markdown("""**What this mode does**
    - Maps the frozen Point Designer truth across a chosen parameter plane
    - Reveals dominant constraints, cliffs, intent splits, and robustness categories
    - Exports replayable scan artifacts (schema v1) and a fixed Scan Lab Atlas""")
            with cB:
                st.markdown("""**What this mode does not do**
    - Optimize, relax constraints, or recommend a best design
    - Apply changes to your base point automatically
    - Redefine physics or hide empty regions""")
    
            # How to think with Scan Lab (philosophy)
            with st.expander("How to think with Scan Lab", expanded=False):
                st.markdown(
                    "**Scan Lab is a microscope, not an engine.**\n"
                    "- It maps the frozen Point Designer truth across a space.\n"
                    "- It does not search for 'best' designs.\n"
                    "- If a region is empty, nature (given assumptions) said *no*.\n\n"
                    "Use it to answer: *What limits me? Where are the cliffs? Which direction helps most?*"
                )
    
            # Contract (always visible, collapsible)
            with st.expander("Scan Lab Contract", expanded=False):
                st.markdown(SCAN_LAB_CONTRACT)
                try:
                    from tools.scan_visual_identity import VISUAL_IDENTITY
                    st.caption(f"Visual semantics frozen: **{VISUAL_IDENTITY.version}**")
                except Exception:
                    st.caption("Visual semantics frozen (Scan Lab v1.0)")
    
            # Restore (artifact -> full UI state)
            with st.expander("Restore Scan Artifact (JSON)", expanded=False):
                st.caption("Upload a previously exported Scan Lab artifact. SHAMS will auto-upgrade it to schema v1 and restore the Scan Lab state.")
                up = st.file_uploader("Upload scan artifact", type=["json"], key="scan_restore_upl")
                if up is not None:
                    try:
                        payload = json.loads(up.getvalue().decode("utf-8"))
                    except Exception as e:
                        payload = None
                        st.error(f"Invalid JSON: {e}")
                    if isinstance(payload, dict) and st.button("Restore this artifact", use_container_width=True, key="scan_restore_btn"):
                        try:
                            art = payload
                            if callable(upgrade_scan_artifact):
                                art = upgrade_scan_artifact(payload)
                            rep = art.get("report") if isinstance(art, dict) else None
                            settings = art.get("settings") if isinstance(art, dict) else None
    
                            if not isinstance(rep, dict):
                                raise ValueError("Artifact missing 'report'")
    
                            # Restore report + artifact
                            st.session_state["scan_cartography_report"] = rep
                            st.session_state["scan_cartography_artifact"] = art
    
                            # Restore scan settings (best effort)
                            if isinstance(settings, dict):
                                st.session_state["scan_cart_x_key"] = settings.get("x_key")
                                st.session_state["scan_cart_y_key"] = settings.get("y_key")
                                st.session_state["scan_cart_x_lo"] = settings.get("x_lo")
                                st.session_state["scan_cart_x_hi"] = settings.get("x_hi")
                                st.session_state["scan_cart_y_lo"] = settings.get("y_lo")
                                st.session_state["scan_cart_y_hi"] = settings.get("y_hi")
                                st.session_state["scan_cart_nx"] = settings.get("nx")
                                st.session_state["scan_cart_ny"] = settings.get("ny")
                                st.session_state["scan_cart_intents"] = settings.get("intents")
                                st.session_state["scan_cart_inc_out"] = settings.get("include_outputs")
    
                            st.success("Restored Scan Lab state from artifact.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Restore failed: {e}")
    
            # Citation-grade provenance / fingerprints
            with st.expander("Provenance (fingerprints)", expanded=False):
                import os
                repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
                fps = {}
                try:
                    if callable(compute_fingerprints):
                        fps = compute_fingerprints(repo_root)
                except Exception:
                    fps = {}
                if fps:
                    st.code(f"fingerprint: {fps.get('fingerprint','n/a')}")
                    st.json(fps, expanded=False)
                else:
                    st.info("Fingerprints unavailable.")
    
            # Freeze readiness (determinism + regression)
            with st.expander("Freeze readiness tools", expanded=False):
                st.caption("These checks do not change physics. They validate determinism and export plumbing.")
                if Evaluator is None or not callable(build_cartography_report) or not callable(stable_hash):
                    st.warning("Freeze readiness checks unavailable (imports missing).")
                else:
                    if st.button("Run quick replay determinism audit", use_container_width=True, key="scan_replay_audit_btn"):
                        try:
                            import numpy as _np
                            ev = _dsg_evaluator(origin="UI", cache_enabled=True)
                            # Small deterministic neighborhood around the current base point
                            base0 = st.session_state.get("last_point_inp")
                            if base0 is None:
                                raise ValueError("No baseline point available.")
                            xk = st.session_state.get("scan_cart_x_key", "Ip_MA")
                            yk = st.session_state.get("scan_cart_y_key", "R0_m")
                            bx = float(getattr(base0, xk, 1.0) or 1.0)
                            by = float(getattr(base0, yk, 1.0) or 1.0)
                            xv = list(_np.linspace(0.95 * bx, 1.05 * bx, 11))
                            yv = list(_np.linspace(0.95 * by, 1.05 * by, 9))
                            intents0 = list(st.session_state.get("scan_cart_intents") or ["Reactor"])
                            rep_a = build_cartography_report(evaluator=ev, base_inputs=base0, x_key=str(xk), y_key=str(yk), x_vals=xv, y_vals=yv, intents=intents0, include_outputs=False)
                            rep_b = build_cartography_report(evaluator=ev, base_inputs=base0, x_key=str(xk), y_key=str(yk), x_vals=xv, y_vals=yv, intents=intents0, include_outputs=False)
                            ha = {
                                "report": stable_hash(rep_a),
                                "dominance": stable_hash(rep_a.get("dominance", {})),
                                "intent_stats": stable_hash(rep_a.get("intent_stats", {})),
                            }
                            hb = {
                                "report": stable_hash(rep_b),
                                "dominance": stable_hash(rep_b.get("dominance", {})),
                                "intent_stats": stable_hash(rep_b.get("intent_stats", {})),
                            }
                            if ha == hb:
                                st.success("Replay determinism audit: PASS")
                            else:
                                st.error("Replay determinism audit: FAIL (hash mismatch)")
                            st.json({"runA": ha, "runB": hb}, expanded=False)
                        except Exception as e:
                            st.error(f"Replay audit failed: {e}")
    
                st.caption("For full freeze gating, run: python scripts/run_scanlab_freeze_qa.py")
    
            # Keyboard quick-jump (expert)
            with st.expander("Expert quick-jump (keyboard)", expanded=False):
                st.caption("Type a letter and press Enter: D=Dominance, F=First-failure, I=Intent split, C=Causality trace")
                cmd = st.text_input("Jump", value="", key="scan_cmd_jump").strip().upper()
                if cmd in ["D","F","I","C"]:
                    st.session_state["scan_view_mode"] = cmd
                    st.success(f"View set to {cmd}.")
    
            base = st.session_state.get("last_point_inp")
            if base is None:
                st.info("Run **Point Designer** first (Scan Lab uses the last evaluated point as the baseline).")
                return
    
            # Canonical questions (teaching / onboarding)
            with st.expander("Canonical questions (teaching)", expanded=False):
                st.caption("A small library of physics questions that Scan Lab can answer. These load scan settings or suggest a view - no designs are applied.")
                qs = []
                try:
                    if callable(build_canonical_questions):
                        qs = build_canonical_questions()
                except Exception:
                    qs = []
                if not qs:
                    st.info("Canonical questions unavailable.")
                else:
                    q_labels = [q.get('question') for q in qs]
                    qpick = st.selectbox("Pick a question", options=q_labels, index=0, key="scan_canon_pick")
                    q = qs[q_labels.index(qpick)]
                    st.write({"hint": q.get("hint"), "suggested_golden_label": q.get("suggested_golden_label")})
                    if q.get("suggested_golden_label"):
                        st.info("Tip: load the suggested golden scan below, then run Cartography.")
    
            # Golden scans (institutional memory)
            with st.expander("Golden scans (teaching + QA)", expanded=False):
                st.caption("One-click canonical landscapes. These load scan settings only; they do not apply designs.")
                presets = []
                try:
                    if callable(build_golden_scan_presets):
                        presets = build_golden_scan_presets(base_inputs=base)
                except Exception:
                    presets = []
                if not presets:
                    st.warning("Golden presets unavailable.")
                else:
                    labels = [p["label"] for p in presets]
                    pick = st.selectbox("Preset", options=labels, index=0, key="scan_golden_pick")
                    gp = presets[labels.index(pick)]
                    st.write({"note": gp.get("note"), "intents": gp.get("intents"), "x": gp.get("x_key"), "y": gp.get("y_key")})
                    if st.button("Load this golden scan", use_container_width=True, key="scan_load_golden"):
                        st.session_state["scan_cart_x_key"] = gp.get("x_key")
                        st.session_state["scan_cart_y_key"] = gp.get("y_key")
                        st.session_state["scan_cart_intents"] = gp.get("intents")
                        st.session_state["scan_cart_x_lo"] = float(gp.get("x_range")[0])
                        st.session_state["scan_cart_x_hi"] = float(gp.get("x_range")[1])
                        st.session_state["scan_cart_y_lo"] = float(gp.get("y_range")[0])
                        st.session_state["scan_cart_y_hi"] = float(gp.get("y_range")[1])
                        st.session_state["scan_cart_nx"] = int(gp.get("n_x"))
                        st.session_state["scan_cart_ny"] = int(gp.get("n_y"))
                        st.session_state["scan_cart_base_override"] = asdict(gp.get("base_inputs")) if gp.get("base_inputs") is not None else None
                        st.success("Loaded golden scan settings.")
    
        st.markdown("---")
        st.markdown("### 🗺️ Cartography")
        st.caption("Produces: constraint-dominance maps, first-failure order, intent-split overlays, robustness labels, and a narrative summary.")

        # Variable pickers
        # Keep list tight and meaningful; these must exist on PointInputs.
        vars2d = [
            ("R0_m", "R0 (m)"),
            ("a_m", "a (m)"),
            ("Bt_T", "B0 (T)"),
            ("Ip_MA", "Ip (MA)"),
            ("fG", "fG (-)"),
            ("Paux_MW", "Paux (MW)"),
            ("kappa", "kappa (-)"),
            ("Ti_keV", "Ti (keV)"),
        ]
        key_to_label = {k: v for k, v in vars2d}
        klist = [k for k, _ in vars2d]

        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            x_key = st.selectbox("x-axis", options=klist, index=klist.index(st.session_state.get("scan_cart_x_key", "Ip_MA")) if st.session_state.get("scan_cart_x_key", "Ip_MA") in klist else 0, format_func=lambda k: f"{k} - {key_to_label.get(k,k)}", key="scan_cart_x")
        with c2:
            y_key = st.selectbox("y-axis", options=klist, index=klist.index(st.session_state.get("scan_cart_y_key", "R0_m")) if st.session_state.get("scan_cart_y_key", "R0_m") in klist else 1, format_func=lambda k: f"{k} - {key_to_label.get(k,k)}", key="scan_cart_y")
        with c3:
            intents = st.multiselect("Intent lenses", options=["Research", "Reactor"], default=st.session_state.get("scan_cart_intents", ["Reactor"]), key="scan_cart_intents")

            # Always-visible intent badge (clarity; no logic change)
            _sel_intents = intents or []
            st.markdown(f"**Intent badge:** {' / '.join([str(x) for x in _sel_intents]) if _sel_intents else '(none)'}")
            st.caption('Projection note: Scan Lab shows a 2D slice. Off-axis constraints may dominate outside this plane.')

        # Ranges + resolution
        def _g(attr: str, default: float) -> float:
            try:
                return float(getattr(base, attr))
            except Exception:
                return float(default)

        bx = _g(x_key, 1.0)
        by = _g(y_key, 1.0)
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            x_lo = st.number_input("x min", value=float(st.session_state.get("scan_cart_x_lo", 0.7 * bx)), step=0.1, key="scan_cart_x_lo")
        with c2:
            x_hi = st.number_input("x max", value=float(st.session_state.get("scan_cart_x_hi", 1.3 * bx)), step=0.1, key="scan_cart_x_hi")
        with c3:
            y_lo = st.number_input("y min", value=float(st.session_state.get("scan_cart_y_lo", 0.7 * by)), step=0.1, key="scan_cart_y_lo")
        with c4:
            y_hi = st.number_input("y max", value=float(st.session_state.get("scan_cart_y_hi", 1.3 * by)), step=0.1, key="scan_cart_y_hi")

        c1, c2, c3 = st.columns([1, 1, 2])
        with c1:
            nx = int(st.slider("Nx", 11, 61, int(st.session_state.get("scan_cart_nx", 31)), 2, key="scan_cart_nx"))
        with c2:
            ny = int(st.slider("Ny", 11, 61, int(st.session_state.get("scan_cart_ny", 25)), 2, key="scan_cart_ny"))
        with c3:
            include_outputs = st.checkbox("Include compact outputs in report", value=False, key="scan_cart_inc_out")

        # Optional base override (loaded from golden scan)
        base_override = st.session_state.get("scan_cart_base_override")
        if isinstance(base_override, dict) and base_override:
            with st.expander("Baseline override (from golden scan)", expanded=False):
                st.json(base_override)

        # Run
        if st.button("Run cartography scan", type="primary", use_container_width=True, key="scan_cart_run"):
            if Evaluator is None or not callable(build_cartography_report):
                st.error("Scan cartography engine unavailable (import error).")
                errs = st.session_state.get("scan_import_errors") or []
                if errs:
                    st.caption("Import details")
                    st.code("\n".join([str(e) for e in errs])
                            [:2000])
                return
            if float(x_hi) <= float(x_lo) or float(y_hi) <= float(y_lo):
                st.error("Invalid bounds: max must be greater than min for both axes.")
                return
            else:
                try:
                    from dataclasses import replace
                    base2 = base
                    if isinstance(base_override, dict) and base_override:
                        try:
                            base2 = replace(base2, **{k: base_override[k] for k in base_override if k in base2.__dict__})
                        except Exception:
                            pass
                    ev = _dsg_evaluator(origin="UI", cache_enabled=True)
                    x_vals = list(np.linspace(float(x_lo), float(x_hi), int(nx)))
                    y_vals = list(np.linspace(float(y_lo), float(y_hi), int(ny)))
                    import time
                    _t0 = time.time()
                    total_pts = max(int(nx) * int(ny), 1)
                    prog = st.progress(0.0, text=f"Stage 1/2 - Evaluating {total_pts} points (frozen evaluator)…")

                    def _cb(done, total):
                        try:
                            dt = max(time.time() - _t0, 1e-6)
                            rate = float(done) / dt
                            eta = (float(total) - float(done)) / max(rate, 1e-6)
                            frac = float(done) / max(float(total), 1.0)
                            prog.progress(min(1.0, max(0.0, frac)), text=f"{done}/{total} pts - {rate:.1f} pt/s - ETA {eta:.0f}s")
                        except Exception:
                            pass

                    rep = build_cartography_report(
                        evaluator=ev,
                        base_inputs=base2,
                        x_key=str(x_key),
                        y_key=str(y_key),
                        x_vals=x_vals,
                        y_vals=y_vals,
                        intents=list(intents or ["Reactor"]),
                        include_outputs=bool(include_outputs),
                        progress_cb=_cb,
                    )
                    try:
                        prog.progress(1.0, text='Stage 2/2 - Computing narrative/topology/artifact…')
                    except Exception:
                        pass
                    rep['run_seconds'] = float(time.time() - _t0)
                    # Attach common metadata and record
                    rep = _attach_common_metadata(rep)
                    st.session_state["scan_cartography_report"] = rep

                    # Freeze-grade Scan Artifact (schema v1)
                    settings = {
                        "x_key": str(x_key),
                        "y_key": str(y_key),
                        "x_lo": float(x_lo),
                        "x_hi": float(x_hi),
                        "y_lo": float(y_lo),
                        "y_hi": float(y_hi),
                        "nx": int(nx),
                        "ny": int(ny),
                        "intents": list(intents or ["Reactor"]),
                        "include_outputs": bool(include_outputs),
                    }
                    artifact = None
                    if callable(build_scan_artifact):
                        try:
                            artifact = build_scan_artifact(
                                report=rep,
                                settings=settings,
                                metadata=dict(rep.get("metadata") or {}),
                                reason_code="run_ok",
                                freeze_statement=f"Scan Lab frozen (schema v{int(SCAN_SCHEMA_VERSION or 1)})",
                            )
                        except Exception:
                            artifact = None
                    if isinstance(artifact, dict):
                        st.session_state["scan_cartography_artifact"] = artifact
                        try:
                            _v98_record_run("scan_cartography", artifact, mode="scan_lab")
                        except Exception:
                            pass
                    else:
                        try:
                            _v98_record_run("scan_cartography", rep, mode="scan_lab")
                        except Exception:
                            pass
                    st.success(f"Scan complete: {rep.get('n_points')} points")
                except Exception as e:
                    st.error(f"Scan failed: {e}")

        rep = st.session_state.get("scan_cartography_report")
        if not isinstance(rep, dict):
            st.info("No cartography scan results yet.")
            return

        # (One-glance truth + intentional emptiness messaging are rendered below.)

        # One-glance truth strip (radical clarity)
        nar_all = rep.get("narrative") or {}
        nar_int = (nar_all.get("intents") or {}) if isinstance(nar_all, dict) else {}
        intents_strip = rep.get("intents") or []
        if not intents_strip:
            intents_strip = ["Reactor"]

        # derive summary for primary intent (first in list)
        it_primary = str(intents_strip[0])
        n0 = nar_int.get(it_primary, {}) if isinstance(nar_int, dict) else {}
        feasible0 = float(n0.get("blocking_feasible_rate", 0.0)) if isinstance(n0, dict) else 0.0
        top0 = (n0.get("dominance_ranked") or []) if isinstance(n0, dict) else []
        dom0 = (top0[0].get("constraint") if top0 else None) or "(none)"
        cliff0 = float(n0.get("cliffiness_proxy", 0.0)) if isinstance(n0, dict) else 0.0

        # robustness verdict (simple, honest): use feasible fraction bands
        if feasible0 >= 0.85:
            rb0 = "Robust"
        elif feasible0 >= 0.55:
            rb0 = "Balanced"
        elif feasible0 >= 0.25:
            rb0 = "Brittle"
        else:
            rb0 = "Knife-edge"

        st.markdown("### One‑glance truth")
        a, b, c, d = st.columns(4)
        a.metric("Dominant constraint", str(dom0))
        b.metric(f"Feasible fraction ({it_primary})", f"{feasible0*100:.0f}%")
        c.metric("Robustness verdict", rb0)
        d.metric("Cliffiness proxy", f"{cliff0:.2f}")

        # The final test: can a user learn something fundamental from one scan?
        with st.expander("One-scan benchmark", expanded=False):
            st.caption("A lightweight self-check for freeze readiness. This is optional and does not affect results.")
            st.checkbox("After one scan, I learned something fundamental about what limits this design space.", key="scan_benchmark_learned")
            st.text_area("If yes: what was it? (optional)", value="", height=90, key="scan_benchmark_note")

        # --- Design Family Governance Engine (v394.0.0) ---
        # Exploration-only: does not affect truth. Deterministic labeling + connected components.
        with st.expander("👪 Design family governance (v394.0.0)", expanded=False):
            st.caption("Extract deterministic design families from the current cartography report (regime-signature labeling + connected components).")
            it_opts = list(intents_strip or ["Reactor"])
            it_sel = st.selectbox("Intent lens", options=it_opts, index=0, key="scan_df_intent_v394")
            min_pts = int(st.slider("Minimum points per family", 4, 80, int(st.session_state.get("scan_df_min_pts_v394", 12)), 1, key="scan_df_min_pts_v394"))
            c1, c2 = st.columns(2)
            with c1:
                run_df = st.button("Build families", type="secondary", use_container_width=True, key="scan_df_build_v394")
            with c2:
                clear_df = st.button("Clear", use_container_width=True, key="scan_df_clear_v394")
            if clear_df:
                st.session_state.pop("scan_design_families_v394", None)
                st.session_state.pop("scan_design_families_v394_cert", None)
                st.success("Cleared design family artifacts.")

            if run_df:
                if not callable(build_design_families_from_scan_cartography):
                    st.error("Design family engine unavailable (import error).")
                else:
                    try:
                        art = build_design_families_from_scan_cartography(rep, intent=str(it_sel), min_points=int(min_pts))
                        # Attach common metadata and record
                        art = _attach_common_metadata(art)
                        st.session_state["scan_design_families_v394"] = art
                        try:
                            from src.certification.design_family_governance_certification_v394 import certify_design_families_v394
                            cert = certify_design_families_v394(artifact=art)
                        except Exception:
                            cert = {"name": "design_family_governance_v394", "verdict": "UNKNOWN"}
                        st.session_state["scan_design_families_v394_cert"] = cert
                        try:
                            _v98_record_run("scan_design_families_v394", {"artifact": art, "cert": cert}, mode="scan_lab")
                        except Exception:
                            pass
                        st.success(f"Built {len(art.get('families') or [])} families.")
                    except Exception as e:
                        st.error(f"Family build failed: {e}")

            art = st.session_state.get("scan_design_families_v394")
            if isinstance(art, dict) and (art.get("families") is not None):
                cert = st.session_state.get("scan_design_families_v394_cert")
                if isinstance(cert, dict):
                    st.markdown("**Certification**")
                    st.json(cert)
                fams = art.get("families") or []
                if isinstance(fams, list) and fams:
                    # Compact table view (avoid scroll walls; default collapsed via expander)
                    rows = []
                    for f in fams:
                        if not isinstance(f, dict):
                            continue
                        rows.append({
                            "family_id": f.get("family_id"),
                            "label": f.get("label"),
                            "n_points": f.get("n_points"),
                            "feasible_frac": f.get("feasible_frac"),
                            "x_range": f"[{f.get('x_min'):.3g}, {f.get('x_max'):.3g}]" if isinstance(f.get('x_min'), (int,float)) and isinstance(f.get('x_max'), (int,float)) else "",
                            "y_range": f"[{f.get('y_min'):.3g}, {f.get('y_max'):.3g}]" if isinstance(f.get('y_min'), (int,float)) and isinstance(f.get('y_max'), (int,float)) else "",
                        })
                    try:
                        import pandas as pd
                        df = pd.DataFrame(rows)
                        st.dataframe(df, use_container_width=True, hide_index=True)
                    except Exception:
                        st.json(rows[:20])
                st.download_button(
                    "Download families JSON",
                    data=json.dumps(art, indent=2, default=str).encode("utf-8"),
                    file_name="shams_scan_design_families_v394.json",
                    mime="application/json",
                    use_container_width=True,
                    key="scan_df_dl_v394",
                )
            else:
                st.info("No design family artifact yet. Run ‘Build families’ after a cartography scan.")

        # Make absence intentional (no-feasible region)
        try:
            all_zero = True
            worst = {}
            for it in intents_strip:
                nn = nar_int.get(str(it), {}) if isinstance(nar_int, dict) else {}
                ff = float(nn.get("blocking_feasible_rate", 0.0)) if isinstance(nn, dict) else 0.0
                if ff > 0.0:
                    all_zero = False
                rk = (nn.get("dominance_ranked") or []) if isinstance(nn, dict) else []
                if rk:
                    worst[str(it)] = rk[0].get("constraint")
            if all_zero:
                st.warning(
                    "No blocking-feasible region exists in this X–Y space (under the selected assumptions)."
                )
                st.markdown(
                    "**Why this is empty (most likely):**\n"
                    + "\n".join([f"- Under **{k}** intent, **{v}** limits essentially everywhere." for k, v in worst.items()])
                )
                st.caption("Try widening bounds, changing axes, or switching intent to test whether this is a structural limit or a policy lens.")
        except Exception:
            pass

        
        
        # --- Post-run Cartography Workbench (v196.3) ---
        # Goal: eliminate scroll-fatigue by treating the map as the center of gravity.
        # This is UI-only; it does not change evaluator truth or scan semantics.

        intents2 = rep.get("intents") or []
        x_vals = rep.get("x_vals") or []
        y_vals = rep.get("y_vals") or []
        pts = rep.get("points") or []
        try:
            grid = {(int(p["i"]), int(p["j"])): p for p in pts if isinstance(p, dict) and "i" in p and "j" in p}
        except Exception:
            grid = {}

        if not intents2:
            intents2 = ["Reactor"]

        # --- Sticky-ish truth bar (best-effort in Streamlit) ---
        try:
            st.markdown(
                """
<style>
/* Make the next container behave like a lightweight "truth bar" */
div[data-testid="stVerticalBlockBorderWrapper"].shams_truthbar { position: sticky; top: 0.5rem; z-index: 50; background: white; }
</style>
                """,
                unsafe_allow_html=True,
            )
        except Exception:
            pass

        # We keep your One‑glance truth metrics above; now we add a compact context line.
        st.caption(f"Post‑run workspace: **{rep.get('x_key')}** vs **{rep.get('y_key')}** · intents: **{' / '.join([str(i) for i in intents2])}** · n={int(rep.get('n_points') or 0)}")

        st.markdown("### 🗺️ Cartography workbench")
        st.caption("Orient → look → probe → explain → compare. (Cartography/interpretability only; no optimization.)")

        # ---------- helpers ----------
        def _cell(intent: str, i: int, j: int) -> dict:
            c = grid.get((int(i), int(j)), {}) if isinstance(grid, dict) else {}
            return ((c.get("intent") or {}).get(str(intent)) or {}) if isinstance(c, dict) else {}

        def _dominance_labels(intent: str):
            labels = set()
            for j in range(len(y_vals)):
                for i in range(len(x_vals)):
                    s = _cell(intent, i, j)
                    if bool(s.get("blocking_feasible")):
                        labels.add("PASS")
                    else:
                        labels.add(str(s.get("dominant_blocking") or "FAIL (unknown)"))
            lab = sorted(list(labels))
            if "PASS" in lab:
                lab = ["PASS"] + [x for x in lab if x != "PASS"]
            return lab

        def _render_dominance_map(intent: str):
            dom = np.empty((len(y_vals), len(x_vals)), dtype=object)
            for j in range(len(y_vals)):
                for i in range(len(x_vals)):
                    s = _cell(intent, i, j)
                    if bool(s.get("blocking_feasible")):
                        dom[j, i] = "PASS"
                    else:
                        dom[j, i] = s.get("dominant_blocking") or "FAIL (unknown)"
            labels = _dominance_labels(intent)
            lut = {lab: k for k, lab in enumerate(labels)}
            z = np.vectorize(lambda s: lut.get(str(s), 0))(dom)

            fig, ax = plt.subplots(figsize=(7.6, 4.4))
            try:
                from tools.scan_visual_identity import build_palette
                palette = build_palette(labels)
            except Exception:
                palette = ['#E0E0E0', '#4C78A8', '#F58518', '#54A24B', '#E45756', '#72B7B2', '#B279A2', '#FF9DA6', '#9D755D', '#BAB0AC', '#2F4B7C', '#7A5195', '#EF5675', '#FFA600']
            if labels and labels[0] == 'PASS':
                palette[0] = '#E0E0E0'
            cmap = ListedColormap(palette[:max(len(labels), 1)])
            ax.imshow(z, origin='lower', aspect='auto', cmap=cmap, vmin=-0.5, vmax=len(labels)-0.5)
            ax.set_xlabel(f"{rep.get('x_key')} - {str(rep.get('x_label') or '')}")
            ax.set_ylabel(f"{rep.get('y_key')} - {str(rep.get('y_label') or '')}")
            ax.set_title(f"Dominant blocking constraint - intent: {intent}")
            return fig, labels

        def _render_feasible_map(intent: str):
            z = np.zeros((len(y_vals), len(x_vals)), dtype=float)
            for j in range(len(y_vals)):
                for i in range(len(x_vals)):
                    s = _cell(intent, i, j)
                    z[j, i] = 1.0 if bool(s.get("blocking_feasible")) else 0.0
            fig, ax = plt.subplots(figsize=(7.6, 4.4))
            ax.imshow(z, origin='lower', aspect='auto')
            ax.set_xlabel(f"{rep.get('x_key')}")
            ax.set_ylabel(f"{rep.get('y_key')}")
            ax.set_title(f"Blocking feasibility (1=feasible, 0=infeasible) - intent: {intent}")
            return fig

        def _render_robustness_proxy(intent: str):
            z = np.full((len(y_vals), len(x_vals)), float("nan"))
            for j in range(len(y_vals)):
                for i in range(len(x_vals)):
                    s = _cell(intent, i, j)
                    try:
                        z[j, i] = float(s.get("local_p_feasible"))
                    except Exception:
                        z[j, i] = float("nan")
            fig, ax = plt.subplots(figsize=(7.6, 4.4))
            ax.imshow(z, origin='lower', aspect='auto')
            ax.set_xlabel(f"{rep.get('x_key')}")
            ax.set_ylabel(f"{rep.get('y_key')}")
            ax.set_title(f"Robustness proxy (local p-feasible) - intent: {intent}")
            return fig

        def _render_operating_contours(intent: str, field: str):
            fc = rep.get("field_cube") if isinstance(rep, dict) else None
            arr = None
            try:
                arr = (fc.get("vars") or {}).get(str(field)) if isinstance(fc, dict) else None
            except Exception:
                arr = None
            if not isinstance(arr, list):
                # fallback: build from per-point outputs (sparse)
                z = np.full((len(y_vals), len(x_vals)), float("nan"))
                for j in range(len(y_vals)):
                    for i in range(len(x_vals)):
                        cell = grid.get((i, j), {}) if isinstance(grid, dict) else {}
                        outs = cell.get("outputs") or {}
                        if isinstance(outs, dict) and field in outs:
                            try:
                                z[j, i] = float(outs.get(field))
                            except Exception:
                                z[j, i] = float("nan")
            else:
                z = np.array(arr, dtype=float)
            fig, ax = plt.subplots(figsize=(7.6, 4.4))
            # Filled contours + feasibility overlay
            try:
                ax.contourf(z, levels=12, origin="lower")
            except Exception:
                ax.imshow(z, origin="lower", aspect="auto")
            ax.set_xlabel(f"{rep.get('x_key')}")
            ax.set_ylabel(f"{rep.get('y_key')}")
            ax.set_title(f"Operating contours: {field} - intent context: {intent}")
            return fig

        # ---------- layout ----------
        nav, mapcol, insp = st.columns([1.05, 2.2, 1.15], gap="large")

        with nav:
            st.markdown("**Navigate**")
            it_active = st.selectbox("Primary intent (view)", options=intents2, index=0, key="scan_wb_intent_active")

            view = st.radio(
                "View",
                options=[
                    "Dominance (blocking)",
                    "Feasibility (blocking)",
                    "Robustness (proxy)",
                    "Operating contours (outputs)",
                ],
                index=0,
                key="scan_wb_view",
            )

            
            out_key = None
            if str(view).startswith("Operating"):
                # Requires include_outputs=True when running the scan
                fc = rep.get("field_cube") if isinstance(rep, dict) else None
                keys = []
                try:
                    keys = sorted(list((fc.get("vars") or {}).keys())) if isinstance(fc, dict) else []
                except Exception:
                    keys = []
                if not keys:
                    st.warning("No output fields found. Re-run the scan with **Include outputs** enabled.")
                else:
                    out_key = st.selectbox("Contour field", options=keys, index=0, key="scan_wb_out_key")
            compare = False
            if len(intents2) >= 2:
                compare = st.checkbox("Compare intents (side-by-side)", value=False, key="scan_wb_compare")

            with st.expander("Legend / meaning", expanded=False):
                st.caption("PASS means blocking-feasible at that cell. Failures are colored by the dominant blocking constraint (worst margin).")
                try:
                    labs = _dominance_labels(it_active)
                    st.write(labs)
                    if labs == ["PASS"]:
                        st.info("All cells are feasible in this slice → dominance map is intentionally neutral/gray.")
                except Exception:
                    pass

            with st.expander("Trust & export", expanded=False):
                st.write({
                    "n_points": rep.get("n_points"),
                    "run_seconds": rep.get("run_seconds"),
                    "report_id": rep.get("id"),
                    "version": rep.get("shams_version"),
                })
                # Downloads (report + artifact) if present
                rep_dl = _shams_json_dumps(rep, indent=2).encode("utf-8")
                st.download_button("Download cartography report (JSON)", data=rep_dl, file_name="shams_cartography_report.json", mime="application/json", use_container_width=True, key="scan_wb_dl_rep")
                art = st.session_state.get("scan_cartography_artifact")
                if isinstance(art, dict):
                    art_dl = _shams_json_dumps(art, indent=2).encode("utf-8")
                    st.download_button("Download scan artifact (JSON, schema v1)", data=art_dl, file_name="shams_scan_artifact_v1.json", mime="application/json", use_container_width=True, key="scan_wb_dl_art")

                    # Optional: boundary segments + field-cube exports (Scan Lab v218 additions)
                    try:
                        bnd = rep.get("boundaries") if isinstance(rep, dict) else None
                        if isinstance(bnd, dict) and bnd:
                            bnd_dl = _shams_json_dumps(bnd, indent=2).encode("utf-8")
                            st.download_button(
                                "Download boundaries (segments JSON)",
                                data=bnd_dl,
                                file_name="shams_scan_boundaries_segments.json",
                                mime="application/json",
                                use_container_width=True,
                                key="scan_wb_dl_boundaries",
                            )
                        fc = rep.get("field_cube") if isinstance(rep, dict) else None
                        if isinstance(fc, dict) and fc:
                            fc_dl = _shams_json_dumps(fc, indent=2).encode("utf-8")
                            st.download_button(
                                "Download field-cube (labelled arrays JSON)",
                                data=fc_dl,
                                file_name="shams_scan_field_cube_v1.json",
                                mime="application/json",
                                use_container_width=True,
                                key="scan_wb_dl_field_cube",
                            )
                    except Exception:
                        pass

        with mapcol:
            if compare and len(intents2) >= 2:
                a, b = st.columns(2)
                for col, it in [(a, intents2[0]), (b, intents2[1])]:
                    with col:
                        if view.startswith("Dominance"):
                            fig, _labs = _render_dominance_map(str(it))
                            st.pyplot(fig, use_container_width=True)
                        elif view.startswith("Feasibility"):
                            fig = _render_feasible_map(str(it))
                            st.pyplot(fig, use_container_width=True)
                        else:
                            if view.startswith("Operating") and out_key:
                                fig = _render_operating_contours(str(it), str(out_key))
                            else:
                                fig = _render_robustness_proxy(str(it))
                            st.pyplot(fig, use_container_width=True)
            else:
                if view.startswith("Dominance"):
                    fig, _labs = _render_dominance_map(str(it_active))
                    st.pyplot(fig, use_container_width=True)
                elif view.startswith("Feasibility"):
                    fig = _render_feasible_map(str(it_active))
                    st.pyplot(fig, use_container_width=True)
                else:
                    if view.startswith("Operating") and out_key:
                        fig = _render_operating_contours(str(it_active), str(out_key))
                    else:
                        fig = _render_robustness_proxy(str(it_active))
                    st.pyplot(fig, use_container_width=True)

            st.caption("Tip: use the Inspector on the right to probe a cell and see the full constraint stack (descriptive only).")

        with insp:
            st.markdown("**Probe / Inspector**")

            if len(x_vals) == 0 or len(y_vals) == 0:
                st.warning("No grid values found in report.")
            else:
                # Probe controls (index-based, reliable across render backends)
                ii = int(st.slider("i (x index)", 0, max(0, len(x_vals) - 1), int(st.session_state.get("scan_wb_i", 0)), 1, key="scan_wb_i"))
                jj = int(st.slider("j (y index)", 0, max(0, len(y_vals) - 1), int(st.session_state.get("scan_wb_j", 0)), 1, key="scan_wb_j"))

                cell0 = grid.get((ii, jj), {}) if isinstance(grid, dict) else {}
                if not cell0:
                    st.warning("Selected cell not found in grid.")
                else:
                    st.write({"x": cell0.get("x"), "y": cell0.get("y"), "i": ii, "j": jj})

                    def _show_intent_block(it: str):
                        s = ((cell0.get("intent") or {}).get(str(it)) or {})
                        st.markdown(f"**Intent:** {it}")
                        st.write({
                            "blocking_feasible": bool(s.get("blocking_feasible")),
                            "dominant_blocking": s.get("dominant_blocking"),
                            "min_blocking_margin": s.get("min_blocking_margin"),
                            "robustness": s.get("robustness"),
                            "local_p_feasible (proxy)": s.get("local_p_feasible"),
                        })
                        fb = s.get("failed_blocking") or []
                        if fb:
                            st.caption("Failed blocking constraints")
                            st.write(list(fb)[:15])
                        mh = (cell0.get("margins_hard") or {}) if isinstance(cell0, dict) else {}
                        if isinstance(mh, dict) and mh:
                            rows = [{"constraint": k, "margin_frac": float(v)} for k, v in mh.items()]
                            rows.sort(key=lambda r: r["margin_frac"])
                            st.caption("Hard-constraint margins (fractional, worst first)")
                            try:
                                import pandas as _pd
                                st.dataframe(_pd.DataFrame(rows), use_container_width=True, height=210)
                            except Exception:
                                st.json(rows[:25], expanded=False)

                    if compare and len(intents2) >= 2:
                        itA, itB = intents2[0], intents2[1]
                        _show_intent_block(str(itA))
                        st.markdown("---")
                        _show_intent_block(str(itB))
                    else:
                        _show_intent_block(str(it_active))

                    # Canonical promotion hook: probe-cell -> Point Designer
                    with st.expander("Promote this probed cell to 🧭 Point Designer", expanded=False):
                        st.caption("Reconstructs a PointInputs candidate from the scan base_inputs + the probed x/y cell and promotes it into Point Designer.")
                        try:
                            base_inputs_dict = rep.get("base_inputs") if isinstance(rep, dict) else None
                            if not isinstance(base_inputs_dict, dict):
                                base_inputs_dict = {}
                            cand = dict(base_inputs_dict)
                            cand[str(rep.get('x_key'))] = float(cell0.get('x'))
                            cand[str(rep.get('y_key'))] = float(cell0.get('y'))
                            st.write({"x_key": str(rep.get('x_key')), "y_key": str(rep.get('y_key')), "x": cand.get(str(rep.get('x_key'))), "y": cand.get(str(rep.get('y_key')))} )
                            if st.button("Promote to Point Designer", use_container_width=True, key="scan_wb_promote_pd"):
                                stage_pd_candidate_apply(cand, source="🗺️ Scan Lab / Workbench Probe", note="Probed scan cell")
                                st.success("Promoted probed cell to Point Designer. Switch to 🧭 Point Designer to review/evaluate.")
                        except Exception as _e:
                            st.info(f"Promotion unavailable: {_e}")

                    fo = cell0.get("failure_order_any") or []
                    if fo:
                        with st.expander("Failure order (hard, worst-first)", expanded=False):
                            st.write(list(fo))

        st.markdown("---")
        with st.expander("Advanced / deep dives (optional)", expanded=False):
            st.caption("Everything below is optional. The workbench above is the primary post-run experience.")
            show_deep = st.checkbox("Show deep dives", value=False, key="scan_show_deep")
        if not bool(show_deep):
            return

# Expert argument tools (claim builder + falsification)
        with st.expander("Expert argument tools (Claim Builder + Falsification)", expanded=False):
            st.caption("Turn scan results into audit-grade, evidence-backed claims. Includes a falsification lens.")

            intents2 = rep.get('intents') or []
            it0 = st.selectbox("Intent", options=intents2 if intents2 else ["Reactor"], index=0, key="scan_claim_intent")
            claim_type = st.selectbox("Claim type", options=["Dominance", "Robustness"], index=0, key="scan_claim_type")

            # Suggested expected value from narrative
            nar = ((rep.get("narrative") or {}).get("intents") or {}).get(it0, {})
            expected_default = ""
            if claim_type == "Dominance":
                expected_default = str((nar.get("dominance_ranked") or [{}])[0].get("constraint") or nar.get("dominant") or "")
            if claim_type == "Robustness":
                expected_default = "Balanced"

            expected = st.text_input("Expected (for falsification)", value=expected_default, key="scan_claim_expected")
            title = st.text_input("Claim title", value=f"Scan claim - {claim_type}", key="scan_claim_title")
            statement = st.text_area(
                "Claim statement",
                value=(
                    f"Under intent {it0}, the scan is dominated by {expected_default or '[constraint]'} across most of the X–Y space."
                    if claim_type == "Dominance" else
                    f"Under intent {it0}, this landscape is {expected_default} in the local neighborhood sense."
                ),
                height=120,
                key="scan_claim_statement",
            )
            notes = st.text_input("Notes (optional)", value="", key="scan_claim_notes")

            # Build evidence and show a stability badge (assumption-sensitivity)
            ev_blob = {}
            try:
                if callable(build_claim_evidence):
                    ev_blob = build_claim_evidence(rep, str(it0))
            except Exception:
                ev_blob = {}

            # Conclusion stability badge (heuristic): if feasible fraction is low or many components -> sensitive
            stability = "Stable"
            try:
                ff = float((ev_blob.get("stats") or {}).get("blocking_feasible_fraction"))
                comps = float((ev_blob.get("cliffs") or {}).get("n_components") or 1)
                if ff < 0.25 or comps >= 3:
                    stability = "Assumption‑sensitive"
                elif ff < 0.5:
                    stability = "Conditionally stable"
            except Exception:
                stability = "Conditionally stable"
            st.info(f"Conclusion stability: **{stability}**")

            # Why nature forces this (synthesis)
            try:
                dom_top = ((ev_blob.get('stats') or {}).get('dominance_top') or [])
                dom_name = dom_top[0][0] if dom_top else (nar.get('dominant') or 'PASS')
                why = (
                    f"Nature forces this landscape to cluster along the **{dom_name}** boundary because it is the dominant blocking limiter across the scan. "
                    f"Where dominance flips, you are crossing a regime boundary: the leverage of {rep.get('x_key')} vs {rep.get('y_key')} changes sign in terms of minimum margin."
                )
                st.caption("Why nature forces this")
                st.write(why)
            except Exception:
                pass

            # Falsification (counterexamples)
            if st.button("Try to falsify this claim", use_container_width=True, key="scan_claim_falsify"):
                try:
                    ct = "dominance" if claim_type == "Dominance" else "robustness"
                    fx = falsify_claim(rep, intent=str(it0), claim_type=ct, expected=str(expected)) if callable(falsify_claim) else {"ok": False, "reason": "falsify_unavailable"}
                    st.session_state["scan_claim_falsify_last"] = fx
                except Exception as e:
                    st.session_state["scan_claim_falsify_last"] = {"ok": False, "reason": str(e)}

            fx = st.session_state.get("scan_claim_falsify_last")
            if isinstance(fx, dict) and fx:
                st.write({"counterexamples": fx.get("n_counterexamples"), "note": fx.get("note")})
                ex = fx.get("examples") or []
                if ex:
                    try:
                        import pandas as pd
                        st.dataframe(pd.DataFrame(ex), use_container_width=True)
                    except Exception:
                        st.json(ex[:10], expanded=False)

            # Export claim as a 1-page PDF slide
            if st.button("Export Claim (1-page PDF)", use_container_width=True, key="scan_claim_export"):
                try:
                    cl = ScanClaim(title=str(title), statement=str(statement), intent=str(it0), claim_type=str(claim_type), notes=str(notes))
                    mp = st.session_state.get("scan_cart_map_pngs") or {}
                    map_png = mp.get(str(it0))
                    import os
                    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
                    fps = compute_fingerprints(repo_root) if callable(compute_fingerprints) else {}
                    pdfb = build_claim_pdf_bytes(claim=cl, evidence=ev_blob, map_png=bytes(map_png) if isinstance(map_png, (bytes, bytearray)) else None, fingerprint=fps) if callable(build_claim_pdf_bytes) else b""
                    st.session_state["scan_claim_pdf_bytes"] = pdfb
                    st.session_state["scan_claim_last"] = {"title": str(title), "statement": str(statement), "intent": str(it0), "type": str(claim_type)}
                except Exception as e:
                    st.error(f"Claim export failed: {e}")

            pdfb = st.session_state.get("scan_claim_pdf_bytes")
            if isinstance(pdfb, (bytes, bytearray)) and len(pdfb) > 100:
                st.download_button("Download claim PDF", data=pdfb, file_name="shams_scan_claim.pdf", mime="application/pdf", use_container_width=True, key="scan_claim_dl")

        # Curated scan library (institutional memory)
        with st.expander("Curated scan library (local)", expanded=False):
            st.caption("Save notable scans locally to build a personal reference library.")
            tag = st.text_input("Tag", value="interesting", key="scan_lib_tag")
            note = st.text_area("Why this scan mattered (one paragraph)", value="", height=90, key="scan_lib_note")
            if st.button("Save scan + note", use_container_width=True, key="scan_lib_save"):
                try:
                    import os, json
                    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
                    lib_path = os.path.join(repo_root, "docs", "scan_library.json")
                    lib = []
                    if os.path.exists(lib_path):
                        lib = json.loads(open(lib_path, "r", encoding="utf-8").read() or "[]")
                    lib.append({
                        "id": rep.get("id"),
                        "tag": str(tag),
                        "note": str(note).strip(),
                        "x": rep.get("x_key"),
                        "y": rep.get("y_key"),
                        "intents": rep.get("intents"),
                        "shams_version": rep.get("shams_version"),
                        "fingerprint": (rep.get("metadata") or {}).get("fingerprints", {}).get("fingerprint") if isinstance(rep.get("metadata"), dict) else None,
                    })
                    with open(lib_path, "w", encoding="utf-8") as f:
                        f.write(_shams_json_dumps(lib, indent=2))
                    st.success(f"Saved to {lib_path}")
                except Exception as e:
                    st.error(f"Save failed: {e}")

        # Next-tier 0-D insight suite (v191)
        with st.expander("Next‑tier insights (0‑D, no optimization)", expanded=False):
            st.caption("These tools turn scans into understanding (laws, regimes, impossibility). They never modify Point Designer truth.")

            insight = st.selectbox(
                "Pick an insight",
                options=[
                    "Local scaling law near a cell",
                    "Regime label at a cell",
                    "Explain impossible / infeasible region",
                    "Constraint irrelevance (never active)",
                    "Assumption stress hotspots (near-threshold)",
                    "Counterfactual lens (drop one constraint)",
                    "Projection stability (vary 3rd variable)",
                    "Path-follow scan (hold a target output)",
                    "Surprise detector (high-entropy neighborhoods)",
                    "Guided insight mode (10‑minute walkthrough)",
                    "Export reference atlas (PDF)",
                    "Export SHAMS Signature Atlas (10 pages)",
                ],
                index=0,
                key="scan_next_tier_pick",
            )

            # point picker shared across views
            try:
                ii = int(st.number_input("Cell i", min_value=0, max_value=max(0, len(x_vals)-1), value=0, step=1, key="scan_nt_i"))
                jj = int(st.number_input("Cell j", min_value=0, max_value=max(0, len(y_vals)-1), value=0, step=1, key="scan_nt_j"))
            except Exception:
                ii, jj = 0, 0

            picked = grid.get((ii, jj), {})

            if insight == "Local scaling law near a cell":
                if not callable(local_powerlaw_fit):
                    st.warning("Local scaling fit unavailable.")
                else:
                    intent_fit = st.selectbox("Intent lens", options=intents2, index=0, key="scan_nt_scal_int")
                    yvar = st.selectbox(
                        "Quantity to fit",
                        options=["min_blocking_margin", "local_p_feasible"] + ["q_div_MW_m2", "sigma_vm_MPa", "B_peak_T", "q95", "TBR", "P_fus_MW"],
                        index=0,
                        key="scan_nt_scal_yvar",
                    )
                    out = local_powerlaw_fit(report=rep, intent=intent_fit, i0=ii, j0=jj, yvar=yvar, radius=2)
                    if out.get("ok"):
                        st.write(out)
                    else:
                        st.warning(out.get("reason"))

            elif insight == "Regime label at a cell":
                if not callable(label_regime):
                    st.warning("Regime labeling unavailable.")
                else:
                    it0 = st.selectbox("Intent lens", options=intents2, index=0, key="scan_nt_reg_int")
                    st.write(label_regime(report=rep, intent=it0, i0=ii, j0=jj))

            elif insight == "Explain impossible / infeasible region":
                if not callable(explain_impossible_region):
                    st.warning("Region explanation unavailable.")
                else:
                    it0 = st.selectbox("Intent lens", options=intents2, index=0, key="scan_nt_imp_int")
                    st.write(explain_impossible_region(report=rep, intent=it0))

            elif insight == "Constraint irrelevance (never active)":
                if not callable(detect_irrelevant_constraints):
                    st.warning("Irrelevance detection unavailable.")
                else:
                    st.write(detect_irrelevant_constraints(report=rep))

            elif insight == "Assumption stress hotspots (near-threshold)":
                if not callable(assumption_stress_hotspots):
                    st.warning("Hotspot detection unavailable.")
                else:
                    it0 = st.selectbox("Intent lens", options=intents2, index=0, key="scan_nt_hot_int")
                    st.write(assumption_stress_hotspots(report=rep, intent=it0, topk=20))

            elif insight == "Counterfactual lens (drop one constraint)":
                if not callable(counterfactual_lens):
                    st.warning("Counterfactual lens unavailable.")
                else:
                    drop = st.text_input("Constraint name to drop (visualization only)", value="TBR", key="scan_nt_drop")
                    it0 = st.selectbox("Intent lens", options=intents2, index=0, key="scan_nt_cf_int")
                    cf = counterfactual_lens(report=rep, intent=it0, drop_constraint=drop)
                    st.write({"removed_constraint": cf.get("removed_constraint"), "ok": cf.get("ok"), "note": cf.get("note")})
                    # basic map: feasible grid under counterfactual
                    try:
                        cf_ok = np.array(cf.get("grid"))
                        fig_cf, ax_cf = plt.subplots(figsize=(7.6, 4.0))
                        ax_cf.imshow(cf_ok.astype(float), origin='lower', aspect='auto')
                        ax_cf.set_title(f"Counterfactual feasible map (drop={drop})")
                        ax_cf.set_xlabel(f"{x_key}")
                        ax_cf.set_ylabel(f"{y_key}")
                        st.pyplot(fig_cf, use_container_width=True)
                    except Exception:
                        pass

            elif insight == "Projection stability (vary 3rd variable)":
                if not callable(projection_stability_check) or Evaluator is None:
                    st.warning("Projection stability check unavailable.")
                else:
                    z_key = st.selectbox("3rd variable (z)", options=klist, index=0, key="scan_nt_z")
                    it0 = st.selectbox("Intent lens", options=intents2, index=0, key="scan_nt_proj_int")
                    rel = float(st.slider("z variation ±%", 1, 20, 5, 1, key="scan_nt_zrel")) / 100.0
                    ev_local = _dsg_evaluator(origin="UI", cache_enabled=True)
                    st.write(projection_stability_check(evaluator=ev_local, base_inputs=base, report=rep, intent=it0, i0=ii, j0=jj, z_key=z_key, rel_step=rel))

            elif insight == "Path-follow scan (hold a target output)":
                if not callable(path_follow_scan) or Evaluator is None:
                    st.warning("Path-follow scan unavailable.")
                else:
                    it0 = st.selectbox("Intent lens", options=intents2, index=0, key="scan_nt_path_int")
                    target_key = st.selectbox("Target output to hold", options=["q95", "B_peak_T", "q_div_MW_m2", "P_fus_MW"], index=0, key="scan_nt_tgt")
                    st.caption("This follows a trajectory by adjusting y to hold the target output approximately constant as x varies.")
                    if st.button("Run path-follow", use_container_width=True, key="scan_nt_run_path"):
                        ev_local = _dsg_evaluator(origin="UI", cache_enabled=True)
                        out = path_follow_scan(evaluator=ev_local, base_inputs=base, x_key=x_key, y_key=y_key, x_vals=list(x_vals), target_output=target_key)
                        st.session_state["scan_path_follow_last"] = out
                    out = st.session_state.get("scan_path_follow_last")
                    if isinstance(out, dict):
                        st.write(out.get("summary"))
                        try:
                            dfp = pd.DataFrame(out.get("path") or [])
                            st.dataframe(dfp, use_container_width=True)
                        except Exception:
                            pass

            elif insight == "Surprise detector (high-entropy neighborhoods)":
                if not callable(surprise_detector):
                    st.warning("Surprise detector unavailable.")
                else:
                    it0 = st.selectbox("Intent lens", options=intents2, index=0, key="scan_nt_sur_int")
                    rad = int(st.slider("Neighborhood radius", 1, 3, 1, 1, key="scan_nt_sur_rad"))
                    st.write(surprise_detector(report=rep, intent=it0, radius=rad))

            elif insight == "Guided insight mode (10‑minute walkthrough)":
                if not callable(guided_steps):
                    st.warning("Guided mode unavailable.")
                else:
                    st.markdown("**Walkthrough steps**")
                    for s in guided_steps():
                        st.write(f"{s.get('step')}. {s.get('title')} - {s.get('hint')}")
                    st.caption("Tip: start with a Golden scan, then follow steps 1→5.")

            elif insight == "Export reference atlas (PDF)":
                if not callable(build_scan_atlas_pdf_bytes):
                    st.warning("Atlas export unavailable.")
                else:
                    st.caption("Build a multi-page PDF atlas from the current scan (one page per intent).")
                    title = st.text_input("Atlas title", value="SHAMS - Scan Lab Atlas", key="scan_nt_atlas_title")
                    if st.button("Build atlas PDF", use_container_width=True, key="scan_nt_build_atlas"):
                        map_pngs = st.session_state.get("scan_cart_map_pngs") or {}
                        pages = []
                        for it0 in intents2:
                            png = map_pngs.get(str(it0))
                            if isinstance(png, (bytes, bytearray)):
                                pages.append({
                                    "report": rep,
                                    "intent": str(it0),
                                    "map_png": bytes(png),
                                    "page_title": f"{title} - {it0}",
                                })
                        pdfb = build_scan_atlas_pdf_bytes(pages=pages, title=str(title))
                        st.session_state["scan_atlas_pdf_bytes"] = pdfb
                    pdfb = st.session_state.get("scan_atlas_pdf_bytes")
                    if isinstance(pdfb, (bytes, bytearray)) and len(pdfb) > 100:
                        st.download_button("Download atlas (PDF)", data=pdfb, file_name="shams_scan_atlas.pdf", mime="application/pdf", use_container_width=True, key="scan_nt_dl_atlas")

            elif insight == "Export SHAMS Signature Atlas (10 pages)":
                if not callable(build_signature_atlas_pdf_bytes):
                    st.warning("Signature atlas export unavailable.")
                else:
                    st.caption("Build the fixed 10-page SHAMS Signature Atlas (contract + provenance + key scan views).")
                    title = st.text_input("Atlas title", value="SHAMS - Scan Lab Signature Atlas", key="scan_sig_atlas_title")
                    if st.button("Build Signature Atlas (10 pages)", use_container_width=True, key="scan_sig_build"):
                        # Map PNGs for intents (from rendered dominance maps)
                        map_pngs = st.session_state.get("scan_cart_map_pngs") or {}
                        # Intent split map PNG (generated when both intents exist)
                        split_png = st.session_state.get("scan_intent_split_png")

                        # Fingerprints (citation-grade)
                        import os
                        repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
                        fps = {}
                        try:
                            if callable(compute_fingerprints):
                                fps = compute_fingerprints(repo_root)
                        except Exception:
                            fps = {}

                        # Optional claim to embed
                        cl = st.session_state.get("scan_claim_last")
                        pdfb = build_signature_atlas_pdf_bytes(
                            report=rep,
                            title=str(title),
                            contract_md=str(SCAN_LAB_CONTRACT),
                            fingerprints=fps,
                            map_png_by_intent={str(k): bytes(v) for k, v in map_pngs.items() if isinstance(v, (bytes, bytearray))},
                            intent_split_png=bytes(split_png) if isinstance(split_png, (bytes, bytearray)) else None,
                            claim=cl if isinstance(cl, dict) else None,
                        )
                        st.session_state["scan_signature_atlas_pdf_bytes"] = pdfb

                        # Publish into docs as SHAMS signature artifact (best-effort)
                        try:
                            out_path = os.path.join(repo_root, "docs", "SHAMS_Scan_Lab_Atlas_Signature.pdf")
                            with open(out_path, "wb") as f:
                                f.write(pdfb)
                            st.success("Published signature atlas into docs/SHAMS_Scan_Lab_Atlas_Signature.pdf")
                        except Exception:
                            pass

                    pdfb = st.session_state.get("scan_signature_atlas_pdf_bytes")
                    if isinstance(pdfb, (bytes, bytearray)) and len(pdfb) > 100:
                        st.download_button("Download Signature Atlas (PDF)", data=pdfb, file_name="shams_scan_signature_atlas.pdf", mime="application/pdf", use_container_width=True, key="scan_sig_dl")

        # PDF one-page summary + PNG map exports per intent
        try:
            from tools.reports.scan_summary import build_scan_summary_pdf_bytes
        except Exception:
            build_scan_summary_pdf_bytes = None  # type: ignore


        # Render maps
        intents2 = rep.get("intents") or []
        x_vals = rep.get("x_vals") or []
        y_vals = rep.get("y_vals") or []
        pts = rep.get("points") or []

        # Build quick lookup
        grid = {(int(p["i"]), int(p["j"])): p for p in pts if isinstance(p, dict) and "i" in p and "j" in p}

        it_primary = str(intents2[0]) if intents2 else 'Reactor'
        # Secondary intent lenses (collapsible) - reduces scroll fatigue.
        for it in intents2:
            if str(it) == str(it_primary):
                continue
            with st.expander(f"Intent lens: {it}", expanded=False):
                nar = ((rep.get("narrative") or {}).get("intents") or {}).get(it, {})
                if nar:
                    st.info(nar.get("plain_language", ""))

            # Topology change alerts (disconnected islands / holes)
                try:
                    topo = (rep.get('topology') or {}).get(it, {})
                    if isinstance(topo, dict) and topo:
                        if int(topo.get('n_components', 1)) > 1:
                            st.warning(f"Topology: {topo.get('n_components')} disconnected feasible islands detected (intent {it}).")
                        if bool(topo.get('has_holes')):
                            st.warning(f"Topology: hole-like infeasible pockets detected (count={topo.get('hole_count')}).")
                except Exception:
                    pass

                # arrays
                dom = np.empty((len(y_vals), len(x_vals)), dtype=object)
                ok = np.zeros((len(y_vals), len(x_vals)), dtype=float)
                mm = np.zeros((len(y_vals), len(x_vals)), dtype=float)
                rb = np.empty((len(y_vals), len(x_vals)), dtype=object)
                for j in range(len(y_vals)):
                    for i in range(len(x_vals)):
                        p = grid.get((i, j), {})
                        s = ((p.get("intent") or {}).get(it) or {})
                        ok[j, i] = 1.0 if bool(s.get("blocking_feasible")) else 0.0
                        # Prevent “all gray” ambiguity when dominance labels are missing.
                        if ok[j, i] > 0.5:
                            dom[j, i] = "PASS"
                        else:
                            dom[j, i] = s.get("dominant_blocking") or "FAIL (unknown)"
                        mm[j, i] = float(s.get("min_blocking_margin") or float("nan"))
                        rb[j, i] = s.get("robustness") or "Unknown"

            # Categorical dominance to integer map
            labels = sorted(list({str(x) for x in dom.flatten().tolist()}))
            lut = {lab: idx for idx, lab in enumerate(labels)}
            z = np.vectorize(lambda s: lut.get(str(s), 0))(dom)

            fig, ax = plt.subplots(figsize=(7.6, 4.4))
            # Stable categorical colormap (PASS is neutral)
            labs = labels
            if 'PASS' in labs:
                labs = ['PASS'] + [x for x in labs if x != 'PASS']
                labels = labs
                lut = {lab: idx for idx, lab in enumerate(labels)}
                z = np.vectorize(lambda s: lut.get(str(s), 0))(dom)
            # Frozen visual semantics: constraint→color mapping (PASS is neutral)
            try:
                from tools.scan_visual_identity import build_palette
                palette = build_palette(labels)
            except Exception:
                palette = ['#E0E0E0', '#4C78A8', '#F58518', '#54A24B', '#E45756', '#72B7B2', '#B279A2', '#FF9DA6', '#9D755D', '#BAB0AC', '#2F4B7C', '#7A5195', '#EF5675', '#FFA600']
            # Guarantee PASS stays neutral even if the mapping changes
            if labels and labels[0] == 'PASS':
                palette[0] = '#E0E0E0'
            cmap = ListedColormap(palette[:max(len(labels), 1)])
            im = ax.imshow(z, origin='lower', aspect='auto', cmap=cmap, vmin=-0.5, vmax=len(labels)-0.5)
            ax.set_xlabel(f"{x_key} - {key_to_label.get(x_key,x_key)}")
            ax.set_ylabel(f"{y_key} - {key_to_label.get(y_key,y_key)}")
            ax.set_title("Constraint‑Dominance Cartography (blocking)")
            # Dominance boundary emphasis
            try:
                b = np.zeros_like(z, dtype=float)
                b[1:, :] |= (z[1:, :] != z[:-1, :])
                b[:, 1:] |= (z[:, 1:] != z[:, :-1])
                ax.contour(b, levels=[0.5], colors='k', linewidths=0.7, origin='lower')
            except Exception:
                pass
            # ticks: keep sparse
            ax.set_xticks([0, len(x_vals)//2, len(x_vals)-1])
            ax.set_xticklabels([f"{x_vals[0]:.3g}", f"{x_vals[len(x_vals)//2]:.3g}", f"{x_vals[-1]:.3g}"])
            ax.set_yticks([0, len(y_vals)//2, len(y_vals)-1])
            ax.set_yticklabels([f"{y_vals[0]:.3g}", f"{y_vals[len(y_vals)//2]:.3g}", f"{y_vals[-1]:.3g}"])
            st.pyplot(fig, use_container_width=True)

            # If everything is PASS, the map is intentionally neutral/gray.
            try:
                if labels == ['PASS']:
                    st.info("This map is neutral/gray because **all sampled points are blocking-feasible** in this slice (dominant constraint = PASS everywhere).")
            except Exception:
                pass

            # Capture a high-DPI PNG for exports (summary PDF / atlas)
            try:
                _buf = io.BytesIO()
                fig.savefig(_buf, format="png", dpi=220, bbox_inches="tight")
                mp = st.session_state.get("scan_cart_map_pngs")
                if not isinstance(mp, dict):
                    mp = {}
                mp[str(it)] = _buf.getvalue()
                st.session_state["scan_cart_map_pngs"] = mp
            except Exception:
                pass

            # Intent-split map: hatched region = Research-feasible but Reactor-infeasible
            if it == 'Reactor' and 'Research' in intents2 and 'Reactor' in intents2:
                try:
                    ok_r = ok
                    ok_s = np.zeros_like(ok)
                    for j in range(len(y_vals)):
                        for i in range(len(x_vals)):
                            p = grid.get((i, j), {})
                            ok_s[j, i] = 1.0 if bool(((p.get('intent') or {}).get('Research') or {}).get('blocking_feasible')) else 0.0
                    only_research = (ok_s > 0.5) & (ok_r < 0.5)
                    if only_research.any():
                        fig_os, ax_os = plt.subplots(figsize=(7.6, 4.4))
                        ax_os.imshow(z, origin='lower', aspect='auto', cmap=cmap, vmin=-0.5, vmax=len(labels)-0.5)
                        ax_os.contourf(only_research.astype(float), levels=[0.5, 1.5], colors='none', hatches=['////'], origin='lower')
                        ax_os.set_title('Intent-split overlay (hatched = Research-only feasible)')
                        st.pyplot(fig_os, use_container_width=True)

                        # Capture PNG for signature atlas
                        try:
                            _buf2 = io.BytesIO()
                            fig_os.savefig(_buf2, format="png", dpi=220, bbox_inches="tight")
                            st.session_state["scan_intent_split_png"] = _buf2.getvalue()
                        except Exception:
                            pass
                except Exception:
                    pass

            with st.expander("Legend (dominant blocking constraint)", expanded=False):
                # Export map for this intent
                try:
                    import io
                    _buf = io.BytesIO()
                    fig.savefig(_buf, format='png', dpi=200, bbox_inches='tight')
                    _buf.seek(0)
                    st.download_button(f'Download map (PNG) - {it}', data=_buf.getvalue(), file_name=f'shams_scan_map_{it.lower()}.png', mime='image/png', use_container_width=True, key=f'scan_map_png_{it}')
                    if callable(build_scan_summary_pdf_bytes):
                        _pdf = build_scan_summary_pdf_bytes(report=rep, intent=str(it), map_png=_buf.getvalue())
                        st.download_button(f'Download 1-page PDF - {it}', data=_pdf, file_name=f'shams_scan_summary_{it.lower()}.pdf', mime='application/pdf', use_container_width=True, key=f'scan_map_pdf_{it}')
                except Exception:
                    pass
                st.write([{"id": int(lut[k]), "constraint": k} for k in labels])

            # Iso-constraint manifolds (margin=0 contours)
            with st.expander("Iso-constraint manifolds (margin=0)", expanded=False):
                st.caption("Shows contour lines where a selected hard-constraint margin crosses zero (approx feasibility boundary).")
                # gather constraint names
                names = set()
                for p in pts:
                    mh = p.get('margins_hard') if isinstance(p, dict) else None
                    if isinstance(mh, dict):
                        for nm in mh.keys():
                            names.add(str(nm))
                names = sorted([n for n in names if n])
                if not names:
                    st.info("Hard-constraint margins not present in this report.")
                else:
                    pick_nm = st.selectbox("Constraint", options=names, index=0, key=f"scan_iso_pick_{it}")
                    M = np.full((len(y_vals), len(x_vals)), np.nan)
                    for j in range(len(y_vals)):
                        for i in range(len(x_vals)):
                            p = grid.get((i, j), {})
                            mh = p.get('margins_hard') if isinstance(p, dict) else None
                            if isinstance(mh, dict) and pick_nm in mh:
                                try:
                                    M[j, i] = float(mh[pick_nm])
                                except Exception:
                                    pass
                    if np.isfinite(M).any():
                        figc, axc = plt.subplots(figsize=(7.6, 4.4))
                        axc.imshow(ok, origin='lower', aspect='auto')
                        try:
                            axc.contour(M, levels=[0.0], colors='k', linewidths=1.0, origin='lower')
                        except Exception:
                            st.info("No margin=0 contour found in the current bounds.")
                        axc.set_title(f"Iso-contour: {pick_nm} margin = 0")
                        st.pyplot(figc, use_container_width=True)
                    else:
                        st.info("No finite margin data for this constraint in the scanned region.")

            # First‑failure topology: show the failure order at a selected point
            with st.expander("First‑Failure Topology (pick a cell)", expanded=False):
                ci, cj = st.columns(2)
                with ci:
                    ii = int(st.slider("i (x index)", 0, max(0, len(x_vals)-1), len(x_vals)//2, 1, key=f"scan_pick_i_{it}"))
                with cj:
                    jj = int(st.slider("j (y index)", 0, max(0, len(y_vals)-1), len(y_vals)//2, 1, key=f"scan_pick_j_{it}"))
                p = grid.get((ii, jj), {})
                st.write({"x": float(p.get("x", float('nan'))), "y": float(p.get("y", float('nan'))), "blocking_feasible": bool(((p.get('intent') or {}).get(it) or {}).get('blocking_feasible'))})
                st.write("Failure order (hard constraints, worst margin first):")
                st.write(p.get("failure_order_any") or [])
                st.write("Intent summary:")
                st.json(((p.get("intent") or {}).get(it) or {}), expanded=False)
                if st.button("Push this point to Point Designer", key=f"scan_push_pd_{it}"):
                    try:
                        # Canonical cross-panel handoff: set pd_candidate_apply.
                        # Point Designer will consume this payload and push it into widget keys.
                        from dataclasses import replace
                        _inp = replace(base, **{x_key: float(p.get("x")), y_key: float(p.get("y"))})
                        try:
                            from dataclasses import asdict
                            stage_pd_candidate_apply(asdict(_inp), source="🗺️ Scan Lab / Topology Picker", note="Picked scan cell")
                        except Exception:
                            stage_pd_candidate_apply(dict(getattr(_inp, "__dict__", {})), source="🗺️ Scan Lab / Topology Picker", note="Picked scan cell")
                        st.session_state["last_point_inp"] = _inp
                        st.success("Promoted this cell to Point Designer. Switch to 🧭 Point Designer to review/evaluate.")
                    except Exception as e:
                        st.error(f"Could not apply point: {e}")

                st.markdown("---")
                st.markdown("##### Local insight tools")
                tabs = st.tabs(["Causality", "Time-to-failure", "Uncertainty", "Null direction"])

                # Build point overrides for this cell
                point_overrides = {x_key: float(p.get("x")), y_key: float(p.get("y"))}

                with tabs[0]:
                    st.caption("Local sensitivity trace for the currently dominant blocking constraint (finite differences).")
                    domc = (((p.get('intent') or {}).get(it) or {}).get('dominant_blocking') or '').strip()
                    if not domc or domc.upper() == 'PASS':
                        st.info("This cell passes blocking constraints; causality trace is most useful on a failing cell.")
                    elif not callable(build_causality_trace) or Evaluator is None:
                        st.info("Causality engine unavailable.")
                    else:
                        knobs = [x_key, y_key, 'R0_m', 'Bt_T', 'Ip_MA', 'fG', 'Paux_MW', 'a_m']
                        knobs = list(dict.fromkeys([k for k in knobs if hasattr(base, k)]))
                        rel_step = st.slider("Sensitivity step (relative)", 0.005, 0.05, 0.01, 0.005, key=f"scan_caus_step_{it}")
                        if st.button("Compute causality trace", key=f"scan_caus_run_{it}", use_container_width=True):
                            try:
                                ev_local = _dsg_evaluator(origin="UI", cache_enabled=True)
                                tr = build_causality_trace(
                                    evaluator=ev_local,
                                    base_inputs=base,
                                    point_overrides=point_overrides,
                                    constraint_name=domc,
                                    knobs=knobs,
                                    rel_step=float(rel_step),
                                )
                                st.json(tr, expanded=False)
                            except Exception as e:
                                st.error(f"Causality trace failed: {e}")

                with tabs[1]:
                    st.caption("How much you can push a knob before the point becomes blocking-infeasible.")
                    if not callable(time_to_failure_along_knob) or Evaluator is None:
                        st.info("Time-to-failure engine unavailable.")
                    else:
                        knob = st.selectbox("Knob", options=[x_key, y_key], index=0, key=f"scan_ttf_knob_{it}")
                        direction = st.radio("Direction", options=["Increase", "Decrease"], horizontal=True, key=f"scan_ttf_dir_{it}")
                        d = 1.0 if direction == "Increase" else -1.0
                        max_rel = st.slider("Max relative push", 0.05, 1.0, 0.5, 0.05, key=f"scan_ttf_max_{it}")
                        if st.button("Compute push-to-fail", key=f"scan_ttf_run_{it}", use_container_width=True):
                            try:
                                ev_local = _dsg_evaluator(origin="UI", cache_enabled=True)
                                rel = time_to_failure_along_knob(
                                    evaluator=ev_local,
                                    base_inputs=base,
                                    point_overrides=point_overrides,
                                    intent=str(it),
                                    knob=str(knob),
                                    direction=float(d),
                                    max_rel=float(max_rel),
                                )
                                if rel is None:
                                    st.info("No failure found within bounds (or point not feasible).")
                                else:
                                    st.success(f"Fails after ≈ {100.0*rel:.1f}% {direction.lower()} in {knob}.")
                            except Exception as e:
                                st.error(f"Time-to-failure failed: {e}")

                with tabs[2]:
                    st.caption("Stress-test this cell under small uncertainty on nuisance inputs. Reports worst-case margin and dominant-constraint probabilities.")
                    if not callable(uncertainty_stress_test) or Evaluator is None:
                        st.info("Uncertainty engine unavailable.")
                    else:
                        nuis_all = [k for k in ['R0_m','a_m','Bt_T','Ip_MA','fG','Paux_MW','kappa','Ti_keV'] if hasattr(base, k)]
                        nuis = st.multiselect("Nuisance keys", options=nuis_all, default=nuis_all[:3], key=f"scan_unc_keys_{it}")
                        rel_unc = st.slider("Relative uncertainty", 0.0, 0.15, 0.03, 0.01, key=f"scan_unc_rel_{it}")
                        n_samples = int(st.slider("Samples", 10, 200, 60, 10, key=f"scan_unc_n_{it}"))
                        seed = int(st.number_input("Seed", value=7, step=1, key=f"scan_unc_seed_{it}"))
                        if st.button("Run uncertainty stress-test", key=f"scan_unc_run_{it}", use_container_width=True):
                            try:
                                ev_local = _dsg_evaluator(origin="UI", cache_enabled=True)
                                u = uncertainty_stress_test(
                                    evaluator=ev_local,
                                    base_inputs=base,
                                    point_overrides=point_overrides,
                                    intent=str(it),
                                    nuisance_keys=list(nuis),
                                    rel_unc=float(rel_unc),
                                    n_samples=int(n_samples),
                                    seed=int(seed),
                                )
                                st.json(u, expanded=False)
                            except Exception as e:
                                st.error(f"Uncertainty test failed: {e}")

                with tabs[3]:
                    st.caption("Null direction = locally flat direction in scan space (perpendicular to margin gradient).")
                    if not callable(null_direction_2d):
                        st.info("Null-direction helper unavailable.")
                    else:
                        try:
                            # local gradient of min margin from neighbors
                            gx = 0.0
                            gy = 0.0
                            if 0 < ii < len(x_vals)-1 and 0 < jj < len(y_vals)-1:
                                gx = (mm[jj, ii+1] - mm[jj, ii-1]) / max((x_vals[ii+1]-x_vals[ii-1]), 1e-12)
                                gy = (mm[jj+1, ii] - mm[jj-1, ii]) / max((y_vals[jj+1]-y_vals[jj-1]), 1e-12)
                            nd = null_direction_2d(gx, gy)
                            st.write({"grad_dir": nd.get('grad_dir'), "flat_dir": nd.get('flat_dir')})
                            st.caption("Interpretation: moving along flat_dir tends to keep min margin nearly constant locally.")
                        except Exception as e:
                            st.error(f"Null direction unavailable: {e}")

            # Optional vector field
            with st.expander("Margin vector field (optional)", expanded=False):
                st.caption("Arrows point toward increasing min blocking margin (local safety direction).")
                if len(x_vals) >= 3 and len(y_vals) >= 3:
                    # finite differences
                    gx = np.zeros_like(mm)
                    gy = np.zeros_like(mm)
                    for j in range(1, len(y_vals)-1):
                        for i in range(1, len(x_vals)-1):
                            gx[j,i] = (mm[j,i+1] - mm[j,i-1]) / max((x_vals[i+1]-x_vals[i-1]), 1e-12)
                            gy[j,i] = (mm[j+1,i] - mm[j-1,i]) / max((y_vals[j+1]-y_vals[j-1]), 1e-12)
                    fig2, ax2 = plt.subplots(figsize=(7.6, 4.4))
                    ax2.imshow(ok, origin="lower", aspect="auto")
                    step = max(1, int(max(len(x_vals), len(y_vals)) / 20))
                    X, Y = np.meshgrid(np.arange(len(x_vals)), np.arange(len(y_vals)))
                    ax2.quiver(X[::step,::step], Y[::step,::step], gx[::step,::step], gy[::step,::step], angles='xy', scale_units='xy', scale=None)
                    ax2.set_title("Vector field over blocking feasibility (background)")
                    ax2.set_xlabel(f"{x_key}")
                    ax2.set_ylabel(f"{y_key}")
                    st.pyplot(fig2, use_container_width=True)
                else:
                    st.info("Increase Nx/Ny to at least 3 for a vector field.")

            # Robustness map (labels)
            with st.expander("Robustness (brutally honest)", expanded=False):
                # summarize counts
                flat = [str(x) for x in rb.flatten().tolist()]
                counts = {k: flat.count(k) for k in sorted(set(flat))}
                st.write({k: f"{v} ({v/len(flat):.0%})" for k, v in counts.items()})

        # Constraint interaction map (coupling view)
        with st.expander("Constraint interaction map (coupling)", expanded=False):
            st.caption("Matrix: how often constraint A appears before B in the local failure order. Descriptive only.")
            inter = rep.get('interaction') or {}
            if isinstance(inter, dict):
                iit = st.selectbox("Intent", options=intents2, index=0, key="scan_inter_intent")
                blob = (inter.get('intents') or {}).get(iit, {}) if isinstance(inter.get('intents'), dict) else {}
                names = blob.get('names') if isinstance(blob, dict) else None
                mat = blob.get('before_counts') if isinstance(blob, dict) else None
                if isinstance(names, list) and isinstance(mat, dict):
                    import pandas as pd
                    dfm = pd.DataFrame(mat)
                    dfm = dfm.reindex(index=names, columns=names)
                    st.dataframe(dfm, use_container_width=True)
                else:
                    st.info("Interaction data unavailable in this report.")
            else:
                st.info("Interaction data unavailable.")

        # Intent split overlay guidance
        if len(intents2) >= 2:
            st.markdown("### Intent-split insight")
            st.caption("Run the same scan with both intents to see how Research-feasible regions differ from Reactor-feasible regions.")

    # Render world-class scan panel
    try:
        _v188_scan_lab_panel()
    except Exception as _e:
        st.error(f"Scan Lab error: {_e}")

    # NOTE: Scan Lab freeze/legacy/parameter-guide/mapping panels are rendered inside Scan Lab
    # (before Cartography) to reduce scroll fatigue and keep the instrument literacy in one place.
if _deck == "Pareto Lab":
    # DSG: auto edge-kind tagging by active panel (exploration only)
    if bool(st.session_state.get("dsg_edge_kind_auto", True)):
        st.session_state["dsg_context_edge_kind"] = "pareto"

    st.header("Pareto Lab")
    st.caption("Trade-off observatory over the feasible set. External optimization is firewalled; truth remains frozen.")
    render_mode_scope("pareto")
    # --- Pareto freeze (read-only semantics) ---
    st.info(PARETO_LOCK_LINE)

    # --- Deck selector (v230.0: external optimizer console) ---
    _pareto_deck_keys = [
        "Internal Pareto Frontier",
        "Robust Pareto Frontier (Phase+UQ)",
        "Regime-Conditioned Pareto Atlas 2.0",
        "Certified Optimization Orchestrator",
        "Feasible Optimizer (External)",
        "Concept Optimization Cockpit",
        "External Optimization Workbench",
        "External Optimization Interpretation",
        "Design Family Narratives",
        "External Optimizer Co-Pilot",
        "External Optimizer Suite",
        "Optimization Evidence Packs",
    ]
    _pareto_deck_labels = {
        "Internal Pareto Frontier": "🧭 Internal Pareto Frontier",
        "Robust Pareto Frontier (Phase+UQ)": "🛡️ Robust Pareto Frontier (Phase+UQ)",
        "Regime-Conditioned Pareto Atlas 2.0": "🧭 Regime-Conditioned Pareto Atlas 2.0",
        "Certified Optimization Orchestrator": "🧾 Certified Optimization Orchestrator",
        "Feasible Optimizer (External)": "🧲 Feasible Optimizer (External)",
        "Concept Optimization Cockpit": "🧬 Concept Optimization Cockpit",
        "External Optimization Workbench": "📈 External Optimization Workbench",
        "External Optimization Interpretation": "🧪 External Optimization Interpretation",
        "Design Family Narratives": "🧬 Design Family Narratives",
        "External Optimizer Co-Pilot": "🧭 External Optimizer Co-Pilot",
        "External Optimizer Suite": "📦 External Optimizer Suite",
        "Optimization Evidence Packs": "🧾 Optimization Evidence Packs",
    }

    # Back-compat for older stored values that included emojis in the raw key.
    _legacy_to_key = {
        "🧬 Concept Optimization Cockpit": "Concept Optimization Cockpit",
        "📈 External Optimization Workbench": "External Optimization Workbench",
        "📦 External Optimizer Suite": "External Optimizer Suite",
    }
    try:
        _legacy = st.session_state.get("pareto_deck_selector_v230")
        if isinstance(_legacy, str) and _legacy in _legacy_to_key:
            st.session_state["pareto_deck_selector_v230"] = _legacy_to_key[_legacy]
    except Exception:
        pass

    _pareto_deck = st.radio(
        "Pareto Lab deck",
        options=_pareto_deck_keys,
        index=0,
        horizontal=True,
        format_func=lambda k: _pareto_deck_labels.get(str(k), str(k)),
        help=(
            "Choose the Pareto Lab deck. External optimizer tooling runs outside the frozen evaluator "
            "and does not modify physics truth."
        ),
        key="pareto_deck_selector_v230",
    )

    if _pareto_deck == "Feasible Optimizer (External)":
        render_external_optimizer_launcher(Path(__file__).resolve().parent.parent)
        st.stop()

    if _pareto_deck == "Certified Optimization Orchestrator":
        from ui.certified_opt_orchestrator import render_certified_optimization_orchestrator

        render_certified_optimization_orchestrator(Path(__file__).resolve().parent.parent)
        st.stop()

    if _pareto_deck == "Concept Optimization Cockpit":
        from ui.concept_opt_cockpit import render_concept_optimization_cockpit

        render_concept_optimization_cockpit(Path(__file__).resolve().parent.parent)
        st.stop()

    if _pareto_deck == "External Optimization Workbench":
        from ui.extopt_workbench import render_extopt_workbench

        render_extopt_workbench(Path(__file__).resolve().parent.parent)
        st.stop()

    if _pareto_deck == "External Optimization Interpretation":
        from ui.extopt_interpretation import render_extopt_interpretation

        render_extopt_interpretation(Path(__file__).resolve().parent.parent)
        st.stop()

    if _pareto_deck == "Design Family Narratives":
        from ui.design_families import render_design_families

        render_design_families(Path(__file__).resolve().parent.parent)
        st.stop()

    if _pareto_deck == "External Optimizer Co-Pilot":
        from ui.extopt_copilot import render_extopt_copilot

        render_extopt_copilot(Path(__file__).resolve().parent.parent)
        st.stop()

    if _pareto_deck == "External Optimizer Suite":
        from ui.extopt_suite import render_extopt_suite

        render_extopt_suite(Path(__file__).resolve().parent.parent)
        st.stop()

    if _pareto_deck == "Optimization Evidence Packs":
        render_optimizer_evidence_packs(Path(__file__).resolve().parent.parent)
        st.stop()

    if _pareto_deck == "Robust Pareto Frontier (Phase+UQ)":
        from ui.robust_pareto_lab import render_robust_pareto_lab

        render_robust_pareto_lab(Path(__file__).resolve().parent.parent)
        st.stop()

    if _pareto_deck == "Regime-Conditioned Pareto Atlas 2.0":
        from ui.regime_conditioned_atlas import render_regime_conditioned_atlas

        render_regime_conditioned_atlas(Path(__file__).resolve().parent.parent)
        st.stop()


    # Summary card (filled after run if results exist)
    with st.container(border=True):
        st.markdown("**Frontier Dashboard**")
        _sum_cols = st.columns(5)
        # placeholders populated later via session state when available
        _ps = st.session_state.get("pareto_last_summary", {})
        _sum_cols[0].metric("Feasible points", _ps.get("n_feasible", "-"))
        _sum_cols[1].metric("Pareto points", _ps.get("n_pareto", "-"))
        _sum_cols[2].metric("Top constraint", _ps.get("top_constraint", "-"))
        _sum_cols[3].metric("Robust mix", _ps.get("robust_mix", "-"))
        _sum_cols[4].metric("Confidence", _ps.get("confidence", "-"))

    # Definition of Pareto optimal (SHAMS-specific)
    with st.expander("ℹ️ What does “Pareto optimal” mean here? (frontier concept)", expanded=False):
        st.markdown(PARETO_OPTIMAL_DEF)

    # Trust boundaries
    with st.expander("Trust boundaries (what you can and cannot conclude)", expanded=False):
        for _t in TRUST_BOUNDARIES:
            st.markdown(f"- {_t}")

        st.caption("Pareto Lab does **not** recommend or select designs. It maps unavoidable trade-offs among **feasible** designs only.")

    # --- Governance (read-only) ---
    with st.expander("Pareto Mode governance (constitution / freeze / contribution rules)", expanded=False):
        st.caption("Read-only governance documents that protect Pareto from drifting into optimization or recommendations.")
        try:
            _c = (Path(__file__).resolve().parent.parent / "docs" / "PARETO_MODE_CONSTITUTION.md").read_text(encoding="utf-8")
            _f = (Path(__file__).resolve().parent.parent / "docs" / "PARETO_V1_FREEZE_DECLARATION.md").read_text(encoding="utf-8")
            _r = (Path(__file__).resolve().parent.parent / "docs" / "PARETO_POST_FREEZE_CONTRIBUTION_RULES.md").read_text(encoding="utf-8")
            _t = (Path(__file__).resolve().parent.parent / "docs" / "PARETO_TEACHING_FREEZE_POLICY.md").read_text(encoding="utf-8")
        except Exception:
            _c=_f=_r=_t="(missing doc file in this build)"
        cols = st.columns(4)
        cols[0].download_button("Download Constitution", data=_c, file_name="PARETO_MODE_CONSTITUTION.md", mime="text/markdown", use_container_width=True)
        cols[1].download_button("Download Freeze", data=_f, file_name="PARETO_V1_FREEZE_DECLARATION.md", mime="text/markdown", use_container_width=True)
        cols[2].download_button("Download Rules", data=_r, file_name="PARETO_POST_FREEZE_CONTRIBUTION_RULES.md", mime="text/markdown", use_container_width=True)
        cols[3].download_button("Download Teaching Policy", data=_t, file_name="PARETO_TEACHING_FREEZE_POLICY.md", mime="text/markdown", use_container_width=True)

    # --- Replay (read-only) ---
    with st.expander("Replay capsule (read-only)", expanded=False):
        st.caption("Load a previously exported Pareto artifact and reproduce the same front without re-sampling. This is audit/review mode.")
        art_file = st.file_uploader("Upload Pareto artifact (.json)", type=["json"], key="pareto_replay_uploader")
        if art_file is not None:
            try:
                art = json.load(art_file)
                st.success("Artifact loaded.")
                st.json({k: art.get(k) for k in ["schema", "version", "intent_mode", "n_samples", "seed", "objectives"] if k in art}, expanded=False)
                _front = pd.DataFrame(art.get("pareto", []) or [])
                _feas = pd.DataFrame(art.get("feasible", []) or [])
                if len(_front):
                    st.markdown("#### Replayed Pareto front")
                    try:
                        import plotly.express as px
                        _objs = list((art.get("objectives") or {}).keys())
                        if len(_objs) >= 2:
                            x, y = _objs[0], _objs[1]
                            fig = px.scatter(_front, x=x, y=y, color="dominant_constraint" if "dominant_constraint" in _front.columns else None, hover_data=_front.columns)
                            st.plotly_chart(fig, use_container_width=True)
                        else:
                            st.dataframe(_front, use_container_width=True)
                    except Exception:
                        st.dataframe(_front, use_container_width=True)
                if len(_feas):
                    st.markdown("#### Replayed feasible set (sampled)")
                    st.dataframe(_feas.head(200), use_container_width=True)
            except Exception as e:
                st.error(f"Could not load artifact: {e}")

    st.markdown(
        "This mode performs a deterministic-feeling **LHS sampling study** inside explicit bounds, filters **intent-aware feasible** points, "
        "and constructs **constraint-annotated Pareto fronts** for explicit objectives."
    )

    # --- Objective Contract (explicit, publishable) ---
    _OBJ_CATALOG = {
        "R0_m": {"units": "m", "desc": "Major radius"},
        "Bt_T": {"units": "T", "desc": "Toroidal field on axis"},
        "Ip_MA": {"units": "MA", "desc": "Plasma current"},
        "fG": {"units": "-", "desc": "Greenwald fraction"},
        "B_peak_T": {"units": "T", "desc": "Peak TF field"},
        "P_e_net_MW": {"units": "MW", "desc": "Net electric power"},
        "Q_DT_eqv": {"units": "-", "desc": "Equivalent DT gain"},
        "q_div_MW_m2": {"units": "MW/m^2", "desc": "Divertor heat-flux proxy"},
        "sigma_vm_MPa": {"units": "MPa", "desc": "Von Mises stress proxy"},
        "hts_margin_cs": {"units": "-", "desc": "HTS margin (critical surface)"},
        "TBR": {"units": "-", "desc": "Tritium breeding ratio"},
    }

    base0 = st.session_state.get("last_point_inp")
    if base0 is None:
        base0 = PointInputs(R0_m=1.85, a_m=0.57, kappa=1.8, Bt_T=12.2, Ip_MA=8.0, Ti_keV=15.0, fG=0.8, Paux_MW=20.0)

    with st.expander("Bounds (sampling hyper-rectangle)", expanded=False):
        st.caption("Bounds are applied to the chosen variables during sampling.")
        bcols = st.columns(4)
        b_R0 = (float(_safe_get(base0, 'R0_m')*0.8), float(_safe_get(base0, 'R0_m')*1.25))
        b_Bt = (float(_safe_get(base0, 'Bt_T')*0.7), float(_safe_get(base0, 'Bt_T')*1.15))
        b_Ip = (float(_safe_get(base0, 'Ip_MA')*0.6), float(_safe_get(base0, 'Ip_MA')*1.6))
        b_fG = (0.3, 1.1)
        R0_lo = _num("R0 min [m]", b_R0[0], 0.01)
        R0_hi = _num("R0 max [m]", b_R0[1], 0.01)
        Bt_lo = _num("Bt min [T]", b_Bt[0], 0.1)
        Bt_hi = _num("Bt max [T]", b_Bt[1], 0.1)
        Ip_lo = _num("Ip min [MA]", b_Ip[0], 0.1)
        Ip_hi = _num("Ip max [MA]", b_Ip[1], 0.1)
        fG_lo = _num("fG min [-]", b_fG[0], 0.05)
        fG_hi = _num("fG max [-]", b_fG[1], 0.05)

        bounds = {
            "R0_m": (float(R0_lo), float(R0_hi)),
            "Bt_T": (float(Bt_lo), float(Bt_hi)),
            "Ip_MA": (float(Ip_lo), float(Ip_hi)),
            "fG": (float(fG_lo), float(fG_hi)),
        }

    with st.expander("Objective Contract (explicit)", expanded=False):
        st.caption("Objectives are explicit and unit-aware. No hidden scoring. The contract is included in exports.")
        # Objective templates (smart presets, not recommendations)
        _OBJ_TEMPLATES = {
            "Custom": None,
            "Reactor - Compact power": {"R0_m":"min","P_e_net_MW":"max","q_div_MW_m2":"min","sigma_vm_MPa":"min","TBR":"max"},
            "Reactor - Max gain": {"Q_DT_eqv":"max","P_e_net_MW":"max","R0_m":"min","q_div_MW_m2":"min"},
            "Research - High current/density": {"Ip_MA":"max","fG":"max","R0_m":"min","Bt_T":"max"},
            "Research - High field": {"Bt_T":"max","B_peak_T":"max","R0_m":"min"},
        }
        tmpl = st.selectbox("Objective template", options=list(_OBJ_TEMPLATES.keys()), index=0, help="Populates objectives with common expert framing. This is not a recommendation.")
        if "pareto_template_last" not in st.session_state:
            st.session_state.pareto_template_last = "Custom"
        # When template changes, update defaults in session_state (deterministic)
        if tmpl != st.session_state.pareto_template_last and _OBJ_TEMPLATES.get(tmpl):
            st.session_state.pareto_sel_objs = list(_OBJ_TEMPLATES[tmpl].keys())
            st.session_state.pareto_obj_senses = dict(_OBJ_TEMPLATES[tmpl])
            st.session_state.pareto_template_last = tmpl
        elif tmpl != st.session_state.pareto_template_last:
            st.session_state.pareto_template_last = tmpl

        intent_mode = st.radio("Design Intent", ["Reactor", "Research", "Both (overlay)"] , index=0, horizontal=True)
        obj_keys = list(_OBJ_CATALOG.keys())
        c1, c2, c3 = st.columns([2, 1, 1])
        with c1:
            sel_objs = st.multiselect("Objectives", options=obj_keys, default=st.session_state.get("pareto_sel_objs", ["R0_m", "B_peak_T", "P_e_net_MW"]), key="pareto_sel_objs")
        with c2:
            st.write("\n")
            robust_margin_thr = float(st.number_input("Robust margin threshold", value=0.10, step=0.05))
        with c3:
            st.write("\n")
            n_samples = int(st.slider("Samples", min_value=50, max_value=4000, value=300, step=50))
        seed = int(st.number_input("Sampling seed", value=1, step=1))

        objectives = {}
        for k in sel_objs:
            meta = _OBJ_CATALOG.get(k, {})
            cols = st.columns([2, 1, 2])
            with cols[0]:
                st.write(f"**{k}**")
            with cols[1]:
                objectives[k] = st.selectbox(f"sense_{k}", ["min", "max"], index=(0 if st.session_state.get("pareto_obj_senses", {}).get(k, ("min" if k in ("R0_m","B_peak_T","q_div_MW_m2","sigma_vm_MPa") else "max"))=="min" else 1), label_visibility="collapsed")
            with cols[2]:
                st.caption(f"{meta.get('desc','')}  [{meta.get('units','-')}]".strip())

        if len(objectives) < 2:
            st.warning("Select at least 2 objectives for a meaningful Pareto front.")

        # Objective sanity validator (warnings only; does not block)
        st.divider()
        with st.expander("Objective sanity checks (warnings only)", expanded=False):
            warns=[]
            if any(k.upper().startswith("TBR") for k in objectives.keys()) and str(intent_mode).startswith("Research"):
                warns.append("TBR is typically **ignored as a blocking constraint** in Research intent. Using TBR as an objective in Research may be uninformative.")
            if any(k in ["P_e_net_MW"] for k in objectives.keys()) and str(intent_mode).startswith("Research"):
                warns.append("Net electric power is usually not a Research driver; ensure this objective is meaningful for your study.")
            if len(set(objectives.keys())) != len(objectives):
                warns.append("Duplicate objective keys detected (this should not happen).")
            if len(warns)==0:
                st.success("No obvious objective-contract red flags.")
            else:
                for w in warns:
                    st.warning(w)

        # Redundancy hint
        st.divider()
        st.caption("Redundancy detection runs after sampling (correlation-based).")

    if st.button("Run Pareto (feasible-only)", type="primary", use_container_width=True):
        import time
        t0=time.time()
        try:
            from solvers.optimize import pareto_optimize
            intents = ["Reactor", "Research"] if str(intent_mode).startswith("Both") else [str(intent_mode)]
            all_runs = []
            all_fronts = []
            all_samples = []
            for it in intents:
                res = pareto_optimize(base0, bounds=bounds, objectives=objectives, n_samples=n_samples, seed=seed, intent_key=it)
                feasible = res.get("feasible", [])
                front = res.get("pareto", [])
                all_samp = res.get("all", [])
                if all_samp:
                    dfA = pd.DataFrame(all_samp)
                    dfA["intent"] = it
                    all_samples.append(dfA)
                if feasible:
                    dfF = pd.DataFrame(feasible)
                    dfF["intent"] = it
                    all_runs.append(dfF)
                if front:
                    dfP = pd.DataFrame(front)
                    dfP["intent"] = it
                    all_fronts.append(dfP)

            dfF_all = pd.concat(all_runs, ignore_index=True) if all_runs else pd.DataFrame()
            dfP_all = pd.concat(all_fronts, ignore_index=True) if all_fronts else pd.DataFrame()
            dfA_all = pd.concat(all_samples, ignore_index=True) if all_samples else pd.DataFrame()
            st.session_state.pareto_last = {
                "objectives": objectives,
                "intent_mode": intent_mode,
                "bounds": bounds,
                "seed": seed,
                "n_samples": n_samples,
                "feasible": dfF_all.to_dict(orient="records") if len(dfF_all) else [],
                "pareto": dfP_all.to_dict(orient="records") if len(dfP_all) else [],
                "robust_margin_thr": robust_margin_thr,
            }

            st.success(f"Done. Feasible points: {len(dfF_all)} / {n_samples*len(intents)}. Pareto points: {len(dfP_all)}. ({time.time()-t0:.1f}s)")
            # Explain why not (if feasibility/front is empty)
            if len(dfF_all) == 0 or len(dfP_all) == 0:
                with st.expander("Explain why not (empty feasibility or empty Pareto front)", expanded=False):
                    if len(dfF_all) == 0:
                        st.warning("No feasible designs were found in the sampled bounds for the selected intent(s). This is not a plotting issue; it means feasibility was not achieved under the frozen evaluator.")
                    elif len(dfP_all) == 0:
                        st.warning("Feasible designs exist, but no non-dominated Pareto set was produced (often due to objective redundancy or insufficient variation).")
                    try:
                        if len(dfA_all) and "first_failure" in dfA_all.columns:
                            vc = dfA_all["first_failure"].fillna("(none)").astype(str).value_counts().head(8)
                            st.markdown("**Top blocking constraints in sampled space (first-failure counts):**")
                            st.dataframe(vc.rename("count").to_frame(), use_container_width=True)
                            st.caption("Tip: If a single constraint dominates all failures, the bounds likely never enter a feasible basin for that intent.")
                        else:
                            st.info("No failure-atlas data available in this run.")
                    except Exception:
                        st.info("Could not summarize failure modes for this run.")

            # --- Sampling honesty panel ---

            with st.expander("Sampling honesty (coverage / density / what was explored)", expanded=False):
                st.caption("Pareto conclusions are only as strong as sampling coverage. This panel reports coverage proxies (no smoothing).")
                if len(dfA_all):
                    st.write({
                        "n_samples_total": int(len(dfA_all)),
                        "n_feasible": int(len(dfF_all)),
                        "feasible_fraction": float(len(dfF_all)/max(len(dfA_all),1)),
                        "intents": intents,
                        "seed": seed,
                    })
                    # Density proxy in objective space (kNN distance)
                    try:
                        obj_keys=list(objectives.keys())
                        if len(obj_keys)>=2 and len(dfF_all)>=10:
                            X=dfF_all[obj_keys].astype(float).to_numpy()
                            k=min(10, len(X)-1)
                            d2=((X[:,None,:]-X[None,:,:])**2).sum(axis=2)
                            np.fill_diagonal(d2, np.inf)
                            knn=np.sort(d2,axis=1)[:,:k]
                            rho=np.sqrt(np.mean(knn,axis=1))
                            st.metric("Median local spacing (objective-space)", float(np.median(rho)))
                            st.metric("95% spacing (thin regions)", float(np.percentile(rho,95)))
                        else:
                            st.info("Density proxy requires ≥10 feasible points and ≥2 objectives.")
                    except Exception as _e:
                        st.info(f"Could not compute density proxy: {_e}")
                else:
                    st.info("No sampling summary available for this run.")

            # --- Failure atlas (infeasible shadow) ---
            with st.expander("Failure atlas (infeasible shadow / what blocks the frontier)", expanded=False):
                st.caption("Faintly shows sampled infeasible points and the first blocking constraint. This does not relax constraints.")
                if len(dfA_all):
                    dfI = dfA_all[~dfA_all.get("is_feasible", False)].copy() if "is_feasible" in dfA_all.columns else pd.DataFrame()
                    if len(dfI) and len(objectives)>=2:
                        obj_keys=list(objectives.keys())
                        x,y=obj_keys[0], obj_keys[1]
                        try:
                            import plotly.express as px
                            figI = px.scatter(dfI, x=x, y=y, color="first_failure" if "first_failure" in dfI.columns else None,
                                              opacity=0.25, hover_data=dfI.columns)
                            st.plotly_chart(figI, use_container_width=True)
                        except Exception:
                            st.dataframe(dfI.head(200), use_container_width=True)
                    else:
                        st.info("No infeasible shadow available (or not enough objectives selected).")
                else:
                    st.info("No infeasible shadow available for this run.")
            if len(dfF_all):
                st.markdown("### Feasible set (intent-aware)")
                st.dataframe(dfF_all, use_container_width=True, height=260)

                # Objective redundancy detection (correlation over feasible set)
                try:
                    corr = dfF_all[list(objectives.keys())].corr(numeric_only=True)
                    bad = []
                    for i, a in enumerate(corr.columns):
                        for j, b in enumerate(corr.columns):
                            if j <= i:
                                continue
                            v = float(corr.loc[a, b])
                            if abs(v) >= 0.92:
                                bad.append((a, b, v))
                    if bad:
                        st.warning("Objective redundancy detected (high correlation on feasible set):")
                        st.dataframe(pd.DataFrame(bad, columns=["objective_a", "objective_b", "corr"]), use_container_width=True)
                except Exception:
                    pass


                # Objective interaction matrix (sign-only couplings on feasible manifold)
                # Non-trade-off region marker (when objectives co-improve)
                try:
                    if "pareto_front" in st.session_state:
                        _pf = st.session_state["pareto_front"]
                        if isinstance(_pf, pd.DataFrame) and len(_pf) >= 10:
                            _okeys = list(objectives.keys())
                            if len(_okeys) >= 2:
                                _c = _pf[_okeys[:2]].corr(numeric_only=True).iloc[0, 1]
                                if _c == _c and float(_c) > 0.80:
                                    st.info("No meaningful trade-off detected in this projection (objectives tend to improve together here).")
                except Exception:
                    pass



                with st.expander("Objective interaction matrix (sign-only, descriptive)", expanded=False):
                    st.caption("Shows how objectives tend to co-vary across the feasible set (not recommendations). '+' means tends to increase together, '-' means trade-off, '~' means weak/none.")
                    try:
                        obj_keys = list(objectives.keys())
                        if len(obj_keys) >= 2 and len(dfF_all) >= 8:
                            C = dfF_all[obj_keys].corr(numeric_only=True)
                            def _sg(v: float) -> str:
                                if not (v == v):
                                    return ""
                                if abs(v) < 0.25:
                                    return "~"
                                return "+" if v > 0 else "-"
                            M = C.copy()
                            for a in obj_keys:
                                for b in obj_keys:
                                    M.loc[a, b] = _sg(float(C.loc[a, b]))
                            st.dataframe(M, use_container_width=True)
                        else:
                            st.info("Not enough feasible points/objectives to form an interaction matrix.")
                    except Exception:
                        st.info("Interaction matrix unavailable in this run.")

                # Epistemic confidence + incompleteness detector (sampling honesty → confidence)
                with st.expander("Confidence & incompleteness (epistemic, not physics)", expanded=False):
                    st.caption("These are sampling-based confidence signals: they estimate where the Pareto picture is solid vs where it may be incomplete due to limited coverage.")
                    try:
                        feas_frac = float(len(dfF_all) / max(int(len(dfA_all)), 1)) if len(dfA_all) else float(len(dfF_all) / max(int(n_samples*len(intents)), 1))
                        st.write({"n_feasible": int(len(dfF_all)), "n_front": int(len(dfP_all)), "feasible_fraction": feas_frac})
                        # Heuristic incompleteness flags
                        flags = []
                        if len(dfF_all) < max(50, 0.01 * n_samples * max(len(intents),1)):
                            flags.append("Feasible sample is sparse; Pareto front may be incomplete.")
                        if feas_frac < 0.002:
                            flags.append("Feasible fraction is very low; consider that the sampled bounds may mostly miss the feasible basin.")
                        if flags:
                            st.warning(" | ".join(flags))
                        else:
                            st.success("No major incompleteness flags triggered by sampling proxies.")
                    except Exception:
                        st.info("Confidence/incompleteness summary unavailable.")

                # Active question suggestions (guides thinking; does not choose designs)
                with st.expander("Possible next questions (guidance, not recommendations)", expanded=False):
                    qs = []
                    try:
                        if 'bad' in locals() and bad:
                            qs.append("Two objectives appear redundant here - would you like to hide one and re-run to sharpen the front?")
                    except Exception:
                        pass
                    try:
                        if len(dfP_all) and "min_signed_margin" in dfP_all.columns:
                            frac_frag = float((pd.to_numeric(dfP_all["min_signed_margin"], errors="coerce") < float(robust_margin_thr)).mean())
                            if frac_frag > 0.5:
                                qs.append("Most Pareto points are fragile under the chosen margin threshold - would you like to view robust-only by default?")
                    except Exception:
                        pass
                    if str(intent_mode).startswith("Both"):
                        qs.append("Would you like to compare intent-split fronts side-by-side (Research vs Reactor) on the same axes?")
                    qs.append("Would you like to export a publication pack (artifact + CSV + narrative + PNG) for this run?")
                    qs.append("Would you like to click a segment and read the regime explanation (what pins this trade-off)?")
                    for q in qs[:6]:
                        st.write("• " + q)

                # Self-audit (continuous proof of integrity)
                with st.expander("Pareto self-audit (read-only integrity checklist)", expanded=False):
                    st.caption("This checklist is informational; it summarizes SHAMS guarantees for this mode.")
                    st.write(
                        {
                            "feasible_only": True,
                            "deterministic": True,
                            "no_recommendations": True,
                            "policy_explicit": True,
                            "intent_explicit": True,
                            "sampling_honesty_reported": True,
                        }
                    )

                # Language calibration (scientific phrasing guardrail)
                with st.expander("Language calibration (how to read statements)", expanded=False):
                    st.caption("All Pareto statements are conditional on bounds, intent, policy, and sampling. They are descriptive (not prescriptive).")
                    st.write("Example: “Stress dominates here” means “Given the selected bounds and policy, σ_vm is most limiting along this segment in the sampled feasible set.”")
            if len(dfP_all):
                st.markdown("### Pareto fronts (constraint-annotated)")
                st.dataframe(dfP_all, use_container_width=True, height=260)

                # Promote a selected Pareto point back into Point Designer (canonical handoff)
                with st.expander("Promote a Pareto point to 🧭 Point Designer", expanded=False):
                    st.caption("Select a point from the Pareto front and promote it into Point Designer inputs (no evaluation performed here).")
                    try:
                        _idxs = list(range(int(len(dfP_all))))
                        _pick = int(st.selectbox("Pareto row index", options=_idxs, index=0, key="pareto_promote_row")) if _idxs else 0
                        _row = dfP_all.iloc[_pick].to_dict() if len(dfP_all) else {}
                        st.write({k: _row.get(k) for k in list(objectives.keys())[:4] + ["dominant_constraint", "min_constraint_margin", "intent"] if k in _row})
                        if st.button("Promote to Point Designer", use_container_width=True, key="pareto_promote_btn"):
                            try:
                                # Reconstruct a full PointInputs dict from the baseline + sampled decision variables.
                                from dataclasses import asdict
                                _base_dict = asdict(base0) if base0 is not None else {}
                            except Exception:
                                _base_dict = dict(getattr(base0, "__dict__", {})) if base0 is not None else {}

                            # Decision variables are the bound keys for this run.
                            for _k in list(bounds.keys()):
                                if _k in _row and _row.get(_k) is not None:
                                    try:
                                        _base_dict[_k] = float(_row.get(_k))
                                    except Exception:
                                        pass

                            stage_pd_candidate_apply(dict(_base_dict), source="📈 Pareto Lab / Internal Pareto", note="Selected Pareto row")
                            st.success("Promoted selected Pareto point to Point Designer. Switch to 🧭 Point Designer to review/evaluate.")
                    except Exception as _e:
                        st.info(f"Promotion UI unavailable for this run: {_e}")

                # Robust envelope (proxy): filter by min_constraint_margin
                dfP_all["robust"] = (pd.to_numeric(dfP_all.get("min_constraint_margin"), errors="coerce") >= robust_margin_thr)
                dfP_robust = dfP_all[dfP_all["robust"]].copy()

                # Freedom-left indicator (2D, selected axes)
                xkey = st.selectbox("x-axis", options=list(objectives.keys()), index=0)
                ykey = st.selectbox("y-axis", options=[k for k in objectives.keys() if k != xkey], index=0)
                ckey = st.selectbox("color", options=["dominant_constraint", "intent", "robust"], index=0)

                def _classify_freedom(df: pd.DataFrame) -> pd.Series:
                    try:
                        d = df.sort_values(xkey)
                        x = pd.to_numeric(d[xkey], errors="coerce").values
                        y = pd.to_numeric(d[ykey], errors="coerce").values
                        dy = np.gradient(y)
                        dx = np.gradient(x)
                        slope = np.abs(dy / (dx + 1e-12))
                        out = []
                        for s in slope:
                            if not np.isfinite(s):
                                out.append("Tight")
                            elif s < 0.15:
                                out.append("Flat")
                            elif s < 0.6:
                                out.append("Tight")
                            else:
                                out.append("Exhausted")
                        return pd.Series(out, index=d.index)
                    except Exception:
                        return pd.Series(["-"]*len(df), index=df.index)

                dfP_all["freedom_left"] = _classify_freedom(dfP_all)
                if len(dfP_robust):
                    dfP_robust["freedom_left"] = _classify_freedom(dfP_robust)

                st.caption("Front segments are annotated with dominant constraint and margin. Robust front is a conservative proxy filter; no uncertainty optimizer is used.")

                # Plot (matplotlib if available, else streamlit)
                try:
                    if _HAVE_MPL and plt is not None:
                        fig = plt.figure()
                        ax = fig.add_subplot(111)
                        # nominal (optionally color by categorical key)
                        if ckey in ("dominant_constraint", "intent") and ckey in dfP_all.columns:
                            cats = list(pd.Series(dfP_all[ckey]).fillna("(none)").astype(str).unique())
                            cmap_vals = pd.Series(dfP_all[ckey]).fillna("(none)").astype(str).map({c:i for i,c in enumerate(cats)})
                            sc = ax.scatter(dfP_all[xkey], dfP_all[ykey], c=cmap_vals, s=20, label="Nominal")
                            cb = fig.colorbar(sc, ax=ax)
                            cb.set_ticks(list(range(len(cats))))
                            cb.set_ticklabels(cats)
                            cb.set_label(ckey)
                        else:
                            ax.scatter(dfP_all[xkey], dfP_all[ykey], s=18, label="Nominal")
                        # robust overlay
                        if len(dfP_robust):
                            ax.scatter(dfP_robust[xkey], dfP_robust[ykey], s=26, marker="x", label="Robust (proxy)")
                        ax.set_xlabel(f"{xkey} [{_OBJ_CATALOG.get(xkey,{}).get('units','-')}]" )
                        ax.set_ylabel(f"{ykey} [{_OBJ_CATALOG.get(ykey,{}).get('units','-')}]" )
                        ax.grid(True, alpha=0.25)
                        ax.legend()
                        st.pyplot(fig, use_container_width=True)
                    else:
                        st.scatter_chart(dfP_all[[xkey, ykey]], x=xkey, y=ykey)
                except Exception:
                    st.scatter_chart(dfP_all[[xkey, ykey]], x=xkey, y=ykey)

                # -----------------------------
                # Pareto v2: world-class interpretability layers (still 0-D, still non-optimizing)
                # -----------------------------
                def _sense_sign(s: str) -> int:
                    return -1 if str(s).lower().strip() == "min" else 1

                def _pareto2(df: pd.DataFrame, xk: str, yk: str, sx: str, sy: str) -> pd.DataFrame:
                    """Return non-dominated set for 2 objectives (stable, no randomness)."""
                    if df is None or len(df) == 0:
                        return df
                    d = df[[c for c in df.columns if c in set(df.columns)]].copy()
                    x = pd.to_numeric(d[xk], errors="coerce").values
                    y = pd.to_numeric(d[yk], errors="coerce").values
                    m = np.isfinite(x) & np.isfinite(y)
                    d = d.loc[m].copy()
                    x = x[m]; y = y[m]
                    # Convert to minimization
                    if str(sx).lower() == "max":
                        x = -x
                    if str(sy).lower() == "max":
                        y = -y
                    order = np.lexsort((y, x))
                    x = x[order]; y = y[order]
                    d = d.iloc[order].copy()
                    best_y = np.inf
                    keep = []
                    for i in range(len(d)):
                        if y[i] < best_y - 1e-12:
                            keep.append(True)
                            best_y = y[i]
                        else:
                            keep.append(False)
                    return d.loc[keep].reset_index(drop=True)

                def _confidence_from_density(df_feas: pd.DataFrame, df_front: pd.DataFrame, xk: str, yk: str) -> pd.Series:
                    """Proxy confidence: kNN distance in objective space (smaller => higher confidence)."""
                    try:
                        if df_feas is None or len(df_feas) < 10 or df_front is None or len(df_front) == 0:
                            return pd.Series([np.nan]*len(df_front), index=df_front.index)
                        F = df_feas[[xk, yk]].copy()
                        P = df_front[[xk, yk]].copy()
                        Fx = pd.to_numeric(F[xk], errors="coerce").values
                        Fy = pd.to_numeric(F[yk], errors="coerce").values
                        Px = pd.to_numeric(P[xk], errors="coerce").values
                        Py = pd.to_numeric(P[yk], errors="coerce").values
                        mF = np.isfinite(Fx) & np.isfinite(Fy)
                        Fx, Fy = Fx[mF], Fy[mF]
                        k = int(max(5, min(25, len(Fx)//30)))
                        out = []
                        for (px, py) in zip(Px, Py):
                            if not (np.isfinite(px) and np.isfinite(py)):
                                out.append(np.nan); continue
                            d2 = (Fx - px)**2 + (Fy - py)**2
                            if len(d2) == 0:
                                out.append(np.nan); continue
                            kk = min(k, len(d2))
                            # partial sort
                            idx = np.argpartition(d2, kk-1)[:kk]
                            md = float(np.mean(np.sqrt(d2[idx]) + 1e-12))
                            out.append(md)
                        # invert and normalize
                        arr = np.asarray(out, dtype=float)
                        if np.all(~np.isfinite(arr)):
                            return pd.Series([np.nan]*len(df_front), index=df_front.index)
                        lo = np.nanmin(arr); hi = np.nanmax(arr)
                        conf = (hi - arr) / (hi - lo + 1e-12)
                        return pd.Series(conf, index=df_front.index)
                    except Exception:
                        return pd.Series([np.nan]*len(df_front), index=df_front.index)

                def _segment_ids(df_front: pd.DataFrame) -> pd.Series:
                    """Assign segment ids where dominant constraint is constant (cliff => new segment)."""
                    if df_front is None or len(df_front) == 0 or "dominant_constraint" not in df_front.columns:
                        return pd.Series([0]*len(df_front), index=df_front.index)
                    dom = df_front["dominant_constraint"].fillna("(none)").astype(str).values
                    seg = []
                    cur = 0
                    prev = dom[0] if len(dom) else "(none)"
                    for d0 in dom:
                        if d0 != prev:
                            cur += 1
                        seg.append(cur)
                        prev = d0
                    return pd.Series(seg, index=df_front.index)

                # Confidence halo + geography tags
                dfP_all = dfP_all.copy()
                dfP_all["confidence"] = _confidence_from_density(dfF_all, dfP_all, xkey, ykey)
                dfP_all["segment_id"] = _segment_ids(dfP_all)
                try:
                    # Geography = cognitive metaphor on frozen data (no new math)
                    geo = []
                    dom = dfP_all["dominant_constraint"].fillna("(none)").astype(str).values if "dominant_constraint" in dfP_all.columns else ["(none)"]*len(dfP_all)
                    for i in range(len(dfP_all)):
                        cliff = (i > 0 and dom[i] != dom[i-1])
                        if cliff:
                            geo.append("Cliff")
                        else:
                            # Ridge proxy: tight margin + stable constraint; Plain: flat + good margin
                            mm = float(dfP_all.iloc[i].get("min_constraint_margin", np.nan))
                            fl = str(dfP_all.iloc[i].get("freedom_left", "-"))
                            if np.isfinite(mm) and mm < max(0.05, robust_margin_thr*0.5):
                                geo.append("Ridge")
                            elif fl == "Flat" and (not np.isfinite(mm) or mm >= robust_margin_thr):
                                geo.append("Plain")
                            else:
                                geo.append("Slope")
                    dfP_all["geography"] = geo
                except Exception:
                    dfP_all["geography"] = ["-"]*len(dfP_all)

                # Quiet opinionated defaults: robust overlay on, fragile points visually de-emphasized
                st.markdown("### Pareto Interpretability Layers")

                # Question-driven exploration (wizard chooses views, not designs)
                qcols = st.columns([2, 2])
                with qcols[0]:
                    question = st.selectbox(
                        "What are you trying to learn?",
                        [
                            "Custom view",
                            "Where does robustness collapse?",
                            "Where is heat exhaust (q_div) limiting?",
                            "Where is stress (σ_vm) limiting?",
                            "Where is TBR policy shaping the trade-off?",
                            "Where do constraints switch (cliffs)?",
                        ],
                        index=0,
                        help="This changes *views and lenses only*. It does not select designs or optimize.",
                    )
                with qcols[1]:
                    focus_metrics = st.multiselect(
                        "Focus metrics (always shown in inspectors)",
                        options=["min_constraint_margin","dominant_constraint","q_div_MW_m2","sigma_vm_MPa","hts_margin_cs","TBR","Q_DT_eqv","P_e_net_MW","B_peak_T"],
                        default=["min_constraint_margin","dominant_constraint","q_div_MW_m2","sigma_vm_MPa"],
                        help="Personalization only: controls what the inspector emphasizes.",
                    )

                # Default lens settings (may be overridden by question-driven preset below)
                _def_geo, _def_conf, _def_policy, _def_teach = True, True, False, False
                if 'question' in locals() and question != "Custom view":
                    if question == "Where does robustness collapse?":
                        _def_geo, _def_conf, _def_policy, _def_teach = True, True, False, True
                    elif question == "Where is heat exhaust (q_div) limiting?":
                        _def_geo, _def_conf, _def_policy, _def_teach = True, True, False, True
                    elif question == "Where is stress (σ_vm) limiting?":
                        _def_geo, _def_conf, _def_policy, _def_teach = True, True, False, True
                    elif question == "Where is TBR policy shaping the trade-off?":
                        _def_geo, _def_conf, _def_policy, _def_teach = True, True, True, True
                    elif question == "Where do constraints switch (cliffs)?":
                        _def_geo, _def_conf, _def_policy, _def_teach = True, True, False, True

                l1, l2, l3, l4 = st.columns([1.2, 1.2, 1.2, 1.4])
                with l1:
                    show_geography = st.checkbox("Geography view", value=_def_geo, help="Terrain metaphor: ridges/cliffs/plains (purely descriptive).")
                with l2:
                    show_conf_halo = st.checkbox("Confidence halo", value=_def_conf, help="Density-based proxy confidence (no smoothing).")
                with l3:
                    show_policy = st.checkbox("Policy change compare", value=_def_policy, help="Compare fronts under explicit policy thresholds (filter-only lens).")
                with l4:
                    teaching_mode = st.checkbox("Teaching mode", value=_def_teach, help="Adds short callouts and guardrails; does not change results.")

                if teaching_mode:
                    st.caption("Tip: A Pareto front is a set of *non-dominated feasible designs*. This tool does not select or recommend designs.")

                # Segment-level explanation
                with st.expander("Explain a front segment", expanded=False):
                    st.caption("Segments are contiguous parts of the front pinned by the same dominant constraint (constraint-switches appear as cliffs).")
                    seg_ids = sorted(dfP_all["segment_id"].unique().tolist()) if len(dfP_all) else [0]
                    sel_seg = int(st.selectbox("Segment", options=seg_ids, index=0))
                    seg_df = dfP_all[dfP_all["segment_id"] == sel_seg].copy()
                    if len(seg_df):
                        domc = str(seg_df["dominant_constraint"].iloc[0]) if "dominant_constraint" in seg_df.columns else "(unknown)"
                        st.write(f"**Dominant constraint:** `{domc}`  |  Points: {len(seg_df)}")
                        # identify strongest driver variable among sampled knobs
                        drivers = [k for k in ("R0_m", "Bt_T", "Ip_MA", "fG") if k in seg_df.columns]
                        driver_msg = ""
                        try:
                            targ = ykey
                            corrs = []
                            for dv in drivers:
                                a = pd.to_numeric(seg_df[dv], errors="coerce")
                                b = pd.to_numeric(seg_df[targ], errors="coerce")
                                if a.notna().sum() > 3 and b.notna().sum() > 3:
                                    corrs.append((dv, float(a.corr(b))))
                            corrs = [c for c in corrs if np.isfinite(c[1])]
                            if corrs:
                                dv, cc = sorted(corrs, key=lambda kv: -abs(kv[1]))[0]
                                driver_msg = f"Within this segment, `{targ}` is most correlated with `{dv}` (corr≈{cc:.2f})."
                        except Exception:
                            pass
                        if driver_msg:
                            st.caption(driver_msg)
                        # causal chain template (descriptive)
                        chain = f"In this segment, pushing `{ykey}` tends to move designs along the front until `{domc}` becomes limiting. "
                        chain += "This reflects coupled 0-D physics and engineering proxies; it is descriptive, not a recommendation."
                        st.write(chain)
                    else:
                        st.info("No points in selected segment.")

                # Objective relevance lens
                with st.expander("Objective relevance lens", expanded=False):
                    st.caption("Shows where objectives genuinely shape the front vs. where they are mostly redundant or flat (descriptive).")
                    try:
                        rel = []
                        for ok in objectives.keys():
                            if ok not in dfP_all.columns:
                                continue
                            vF = float(pd.to_numeric(dfF_all.get(ok, pd.Series([],dtype=float)), errors="coerce").var()) if len(dfF_all) else np.nan
                            vP = float(pd.to_numeric(dfP_all.get(ok, pd.Series([],dtype=float)), errors="coerce").var()) if len(dfP_all) else np.nan
                            ratio = (vP / (vF + 1e-12)) if (np.isfinite(vP) and np.isfinite(vF)) else np.nan
                            rel.append({"objective": ok, "var_front": vP, "var_feasible": vF, "relevance_ratio": ratio})
                        rel_df = pd.DataFrame(rel).sort_values("relevance_ratio", ascending=False)
                        st.dataframe(rel_df, use_container_width=True, hide_index=True)
                        if teaching_mode and len(rel_df):
                            st.caption("High relevance_ratio means the objective varies significantly along the front; low means it may be redundant in this domain.")
                    except Exception:
                        st.info("Relevance lens unavailable.")

                # Policy change compare (filter-only lens; no physics changes)
                dfP_policy = None
                if show_policy:
                    with st.expander("Policy change compare (filter-only lens)", expanded=False):
                        st.caption("This is a policy lens: it filters feasible points using explicit thresholds, then recomputes the non-dominated set. No constraints are relaxed; no evaluator changes are made here.")
                        p1, p2, p3, p4 = st.columns(4)
                        with p1:
                            tbr_min = float(st.number_input("TBR ≥", value=1.10, step=0.01))
                        with p2:
                            sigma_max = float(st.number_input("σ_vm ≤ [MPa]", value=700.0, step=10.0))
                        with p3:
                            qdiv_max = float(st.number_input("q_div ≤ [MW/m²]", value=10.0, step=0.5))
                        with p4:
                            hts_min = float(st.number_input("HTS margin ≥", value=0.10, step=0.01))
                        dpol = dfF_all.copy()
                        if "TBR" in dpol.columns:
                            dpol = dpol[pd.to_numeric(dpol["TBR"], errors="coerce") >= tbr_min]
                        if "sigma_vm_MPa" in dpol.columns:
                            dpol = dpol[pd.to_numeric(dpol["sigma_vm_MPa"], errors="coerce") <= sigma_max]
                        if "q_div_MW_m2" in dpol.columns:
                            dpol = dpol[pd.to_numeric(dpol["q_div_MW_m2"], errors="coerce") <= qdiv_max]
                        if "hts_margin_cs" in dpol.columns:
                            dpol = dpol[pd.to_numeric(dpol["hts_margin_cs"], errors="coerce") >= hts_min]
                        dfP_policy = _pareto2(dpol, xkey, ykey, objectives.get(xkey,"min"), objectives.get(ykey,"min"))
                        st.write(f"Policy-filtered feasible points: {len(dpol)} | policy-front points: {len(dfP_policy) if dfP_policy is not None else 0}")

                # Reference fronts (runtime presets; deterministic)
                with st.expander("Reference fronts (runtime presets)", expanded=False):
                    st.caption("Generate canonical reference fronts for comparison (deterministic presets; no hidden scoring).")
                    ref = st.selectbox("Reference family", ["None", "ITER-like", "SPARC-like", "ARC-like"], index=0)
                    if ref != "None":
                        st.info("Reference fronts are generated on demand using fixed presets and the frozen evaluator. This does not recommend designs.")
                    # This upgrade only provides UI hooks; generation uses the same run button / settings.

                # No free lunch detector
                with st.expander("No free lunch detector", expanded=False):
                    st.caption("Flags regions where both selected objectives appear to improve together - often due to projection, redundancy, or shared drivers.")
                    try:
                        sx = objectives.get(xkey, "min"); sy = objectives.get(ykey, "min")
                        dx = _sense_sign(sx); dy = _sense_sign(sy)
                        d = dfP_all.sort_values(xkey).copy()
                        x = pd.to_numeric(d[xkey], errors="coerce").values
                        y = pd.to_numeric(d[ykey], errors="coerce").values
                        good = []
                        for i in range(1, len(d)):
                            if not (np.isfinite(x[i-1]) and np.isfinite(x[i]) and np.isfinite(y[i-1]) and np.isfinite(y[i])):
                                continue
                            imp_x = dx*(x[i]-x[i-1]) < 0  # improvement
                            imp_y = dy*(y[i]-y[i-1]) < 0
                            if imp_x and imp_y:
                                good.append(int(i))
                        if good:
                            st.warning(f"Detected {len(good)} local steps where both objectives improve together. This may indicate redundancy/shared drivers or a misleading projection.")
                            if teaching_mode:
                                st.caption("Try changing axes, adding a third objective, or inspecting dominance and segment explanations to interpret this correctly.")
                        else:
                            st.success("No obvious 'free lunch' steps detected along the chosen front ordering.")
                    except Exception:
                        st.info("Detector unavailable.")

                # Narrative timeline along the front
                # Knee candidates (descriptive, not "best")
                # Knee candidates (descriptive, not "best")
                with st.expander("Notable compromise regions (knee candidates)", expanded=False):
                    st.caption("Highlights regions of the front where trade-offs tighten rapidly (geometric knee proxy). These are **not** recommendations.")
                    if len(dfP_all) >= 5:
                        try:
                            d = dfP_all[[xkey, ykey, "segment_id", "dominant_constraint", "min_constraint_margin", "confidence", "intent"]].copy()
                            d = d.sort_values(xkey).reset_index(drop=True)
                            x = pd.to_numeric(d[xkey], errors="coerce").values
                            y = pd.to_numeric(d[ykey], errors="coerce").values
                            m = np.isfinite(x) & np.isfinite(y)
                            d = d.loc[m].reset_index(drop=True)
                            x = x[m]; y = y[m]
                            if len(d) >= 5:
                                xn = (x - np.min(x)) / (np.ptp(x) + 1e-12)
                                yn = (y - np.min(y)) / (np.ptp(y) + 1e-12)
                                kappa = np.zeros(len(d))
                                for ii in range(1, len(d)-1):
                                    v1 = np.array([xn[ii]-xn[ii-1], yn[ii]-yn[ii-1]])
                                    v2 = np.array([xn[ii+1]-xn[ii], yn[ii+1]-yn[ii]])
                                    n1 = np.linalg.norm(v1) + 1e-12
                                    n2 = np.linalg.norm(v2) + 1e-12
                                    ang = np.arccos(np.clip(np.dot(v1, v2)/(n1*n2), -1.0, 1.0))
                                    kappa[ii] = float(ang)
                                d["knee_score"] = kappa
                                topk = d.sort_values("knee_score", ascending=False).head(min(8, len(d)))
                                st.dataframe(topk, use_container_width=True, hide_index=True)
                                if teaching_mode:
                                    st.caption("High knee_score means the front bends sharply in the chosen projection; confirm with segment explanations and dominance to avoid projection traps.")
                            else:
                                st.info("Not enough clean points for knee scoring.")
                        except Exception:
                            st.info("Knee scoring unavailable for this run.")
                    else:
                        st.info("Need at least 5 Pareto points to compute knee candidates.")

                with st.expander("Pareto timeline (scrub along the front)", expanded=False):
                    st.caption("Scrub a slider along the front to see objective/constraint transitions as a narrative.")
                    if len(dfP_all):
                        k = int(st.slider("Front index", min_value=0, max_value=max(len(dfP_all)-1, 0), value=min(0, len(dfP_all)-1), step=1))
                        row = dfP_all.iloc[k].to_dict()
                        st.write(f"**Index {k}** | segment={row.get('segment_id')} | geography={row.get('geography')} | dominant={row.get('dominant_constraint')} | margin={row.get('min_constraint_margin')}")
                        st.json({xkey: row.get(xkey), ykey: row.get(ykey), "intent": row.get("intent"), "dominant_constraint": row.get("dominant_constraint"), "confidence": row.get("confidence")}, expanded=False)
                    else:
                        st.info("No front points to scrub.")
                # Point Inspector
                with st.expander("Pareto Point Inspector", expanded=False):
                    idx = int(st.number_input("Row index (in table above)", min_value=0, max_value=max(len(dfP_all)-1,0), value=0, step=1))
                    try:
                        row = dfP_all.iloc[idx].to_dict()
                        st.json(row, expanded=False)
                        # Focus metrics (personalized)
                        try:
                            _fm = focus_metrics if "focus_metrics" in locals() else ["min_constraint_margin","dominant_constraint"]
                            _show = {k: row.get(k) for k in _fm if k in row}
                            if _show:
                                st.markdown("**Focus metrics**")
                                st.dataframe(pd.DataFrame([_show]), use_container_width=True, hide_index=True)
                        except Exception:
                            pass

                        cA, cB, cC = st.columns([1, 1, 1])
                        with cA:
                            if st.button("View in Scan Lab", use_container_width=True, key="pareto_to_scanlab"):
                                # Best-effort cross-link: store a focus payload and instruct user to switch tab.
                                st.session_state.scanlab_focus_from_pareto = {
                                    "inputs": {k: row.get(k) for k in ("R0_m", "Bt_T", "Ip_MA", "fG") if k in row},
                                    "objectives": {k: row.get(k) for k in objectives.keys()},
                                    "intent": str(row.get("intent", intent_mode)),
                                    "dominant_constraint": row.get("dominant_constraint"),
                                    "min_constraint_margin": row.get("min_constraint_margin"),
                                }
                                st.info("Stored a Scan Lab focus hint in session state. Switch to Scan Lab to view the highlighted context.")
                        with cB:
                            st.caption("Read-only: no auto-apply.")
                        st.caption("To inspect physics/constraints deeply, paste these inputs into Point Designer or Systems Mode. Pareto does not auto-apply.")
                        with cC:
                            if st.button("Queue in Systems Mode", use_container_width=True, key="pareto_to_systems"):
                                # Queue a reversible base-apply payload for Systems Mode (no solving here).
                                st.session_state.systems_pending_base_apply = {k: row.get(k) for k in ("R0_m","a_m","kappa","Bt_T","Ip_MA","Ti_keV","fG","Paux_MW","t_shield_m") if k in row}
                                st.session_state.systems_pending_base_apply_source = "Pareto Lab point"
                                st.info("Queued. Switch to Systems Mode and review the pending Apply card (reversible).")

                    except Exception:
                        st.info("Select a valid row index.")

                # Narrative summary (deterministic)
                with st.expander("Trade-off summary (deterministic)", expanded=False):
                    try:
                        dom_counts = dfP_all["dominant_constraint"].value_counts().to_dict() if "dominant_constraint" in dfP_all.columns else {}
                        dom_top = sorted(dom_counts.items(), key=lambda kv: -kv[1])[:3]
                        msg = ""
                        if dom_top:
                            msg += "Dominant-limiting segments (by count along the front): " + ", ".join([f"{k} ({v})" for k,v in dom_top]) + ". "
                        msg += f"Freedom-left classification along chosen axes: Flat={int((dfP_all['freedom_left']=='Flat').sum())}, Tight={int((dfP_all['freedom_left']=='Tight').sum())}, Exhausted={int((dfP_all['freedom_left']=='Exhausted').sum())}. "
                        msg += "These statements are descriptive; no design choice is implied."
                        st.session_state.pareto_narrative_summary = msg
                        st.write(msg)
                    except Exception:
                        st.write("Summary unavailable.")

                # Export artifact (JSON) + CSV
                c1, c2 = st.columns(2)
                with c1:
                    st.download_button(
                        "Download Pareto front (CSV)",
                        data=dfP_all.to_csv(index=False),
                        file_name="shams_pareto_front.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )
                with c2:
                    art = {
                        "schema": "shams.pareto.v1",
                        "version": str(st.session_state.get("app_version","")),
                        "intent_mode": intent_mode,
                        "objectives": {k: {"sense": v, **_OBJ_CATALOG.get(k, {})} for k, v in objectives.items()},
                        "bounds": bounds,
                        "seed": seed,
                        "n_samples": n_samples,
                        "robust_margin_thr": robust_margin_thr,
                        "feasible": dfF_all.to_dict(orient="records") if len(dfF_all) else [],
                        "pareto": dfP_all.to_dict(orient="records") if len(dfP_all) else [],
                    }
                    st.download_button(
                        "Download Pareto artifact (JSON)",
                        data=json.dumps(art, indent=2, sort_keys=True),
                        file_name="shams_pareto_artifact.json",
                        mime="application/json",
                        use_container_width=True,
                    )
                    # Publication-ready export pack (zip): artifact + CSVs + narrative
                    try:
                        import io, zipfile as _zip
                        buf = io.BytesIO()
                        with _zip.ZipFile(buf, "w", compression=_zip.ZIP_DEFLATED) as zf:
                            zf.writestr("pareto/shams_pareto_artifact.json", json.dumps(art, indent=2, sort_keys=True))
                            if len(dfP_all):
                                zf.writestr("pareto/pareto_front.csv", dfP_all.to_csv(index=False))
                            if len(dfF_all):
                                zf.writestr("pareto/feasible_set_sampled.csv", dfF_all.to_csv(index=False))
                            # Narrative summary
                            zf.writestr("pareto/narrative_summary.md", str(st.session_state.get("pareto_narrative_summary","")).strip() or "(no narrative)")
                            # Repro capsule
                            zf.writestr("pareto/README.md", "\n".join([
                                "# SHAMS Pareto Publication Pack",
                                "",
                                "- Includes the JSON artifact (authoritative), CSV exports, and the deterministic narrative summary.",
                                "- Pareto Mode is feasible-only and non-optimizing.",
                                "",
                                f"- intent_mode: {intent_mode}",
                                f"- n_samples: {n_samples}",
                                f"- seed: {seed}",
                                f"- objectives: {list(objectives.keys())}",
                            ]))
                        buf.seek(0)
                        st.download_button(
                            "Download publication pack (.zip)",
                            data=buf.getvalue(),
                            file_name="shams_pareto_publication_pack.zip",
                            mime="application/zip",
                            use_container_width=True,
                        )
                    except Exception:
                        pass
                    # Interactive reproducibility capsule (trust panel)
                    with st.expander("Reproducibility capsule (what is guaranteed / what is not)", expanded=False):
                        st.caption("This makes Pareto audit-ready. It describes guarantees and limitations of the current run (no hidden assumptions).")
                        st.markdown("**Guaranteed**")
                        st.markdown("- Feasibility is evaluated by the frozen Point Designer evaluator (intent-aware).")
                        st.markdown("- Pareto uses explicit objectives only (Objective Contract).")
                        st.markdown("- Sampling is seeded; reruns with same seed/bounds/objectives are reproducible up to floating-point nondeterminism.")
                        st.markdown("**Not guaranteed**")
                        st.markdown("- Global coverage: the front is only as complete as the sampling density.")
                        st.markdown("- Projection honesty: 2D views can hide higher-dimensional dominance changes.")
                        st.markdown("- Counterfactual/policy lenses are filter-only overlays; they do not modify evaluator physics.")
                        st.divider()
                        st.json({
                            "version": str(st.session_state.get("app_version","")),
                            "intent_mode": intent_mode,
                            "seed": seed,
                            "n_samples": n_samples,
                            "bounds": bounds,
                            "objectives": objectives,
                        }, expanded=False)

        except Exception as e:
            st.error(f"Pareto study error: {e}")

    # -----------------------------
    


    # --- Freeze badge (Pareto) ---
    with st.container():
        st.caption(" Pareto Mode v1.0 - Frozen. Descriptive trade-off cartography only.")
        try:
            _pf = (Path(__file__).resolve().parent.parent / "docs" / "PARETO_V1_FREEZE_DECLARATION.md").read_text(encoding="utf-8")
        except Exception:
            _pf = "(missing docs/PARETO_FREEZE.md)"
        st.caption(FREEZE_STAMP)
        st.download_button("Download Pareto Freeze Statement", data=_pf, file_name="PARETO_V1_FREEZE_DECLARATION.md", mime="text/markdown", use_container_width=False)


if _deck == "Trade Study Studio":
    from ui.decks.trade_study_studio import render_trade_study_studio
    render_trade_study_studio(sys.modules[__name__])


if _deck == "Reactor Design Forge":
    st.header("Reactor Design Forge")
    st.caption("Concept assembly + candidate archives + traces. Feeds the frozen evaluator; does not replace it.")
    render_mode_scope("forge")

    # --- Legacy v93 stateful download compatibility (read-only) ---
    try:
        _v93_stateful_sandbox_panel()
    except Exception:
        pass

    st.subheader("🧭 Forge Bridgehead")
    # v208: Review Mode (review-room posture; no knobs/search actions)
    if "forge_review_mode" not in st.session_state:
        st.session_state["forge_review_mode"] = False
    st.session_state["forge_review_mode"] = st.toggle(
        "Review Mode (locks exploration controls)",
        value=bool(st.session_state.get("forge_review_mode")),
        key="forge_review_mode_toggle",
        help="Review Mode is a UI posture: inputs are locked; only review artifacts and comparisons are shown.",
    )

    if bool(st.session_state.get("forge_review_mode")):
        st.info(
            f"{_LANG.get('non_prescriptive_banner')}  \
{_LANG.get('margin_first')}",
            icon="🧑‍⚖️",
        )
    else:
        st.caption("Candidate-generation workspace (external to truth): global pattern → surrogate acceleration → local refinement. All candidates are audited by the frozen evaluator. No relaxation; no auto-apply.")
    st.info(
        "Non-authoritative workspace: this mode produces **candidate archives** + **traces**. "
        "Truth remains in the frozen evaluator. Nothing is applied automatically.",
        icon="🧱",
    )

    # v208: Review Mode locks exploration controls (read-only review posture).
    forge_lock = bool(st.session_state.get("forge_review_mode"))

    # --- Forge deck (no scroll walls) ---
    _forge_deck = st.radio(
        "Forge deck",
        options=["🧩 Intent Compiler", "🚀 Machine Finder", "🔁 Capsules"],
        index=0,
        horizontal=True,
        key="forge_deck",
        help="Deck-based navigation: render one Forge workspace at a time (no scroll walls).",
    )


    # --- Imports (local to keep UI start-up fast) ---
    from src.models.inputs import PointInputs
    from constraints.system import build_constraints_from_outputs
    from tools.process_compat.process_compat import (
        constraints_to_records,
        active_constraints,
        feasibility_flag,
        failure_mode,
    )
    from tools.sandbox.hybrid_engine import (
        Objective, VarSpec, run_hybrid_machine_finder,
        global_de_phase, surrogate_phase, local_refine_phase, surface_surf_phase,
        build_archive, resistance_atlas, variable_correlations, build_feasibility_skeleton,
    )
    from tools.sandbox.optimizer_engines import default_objective_packs
    from tools.sandbox.feasibility_ladder import classify_candidate
    from tools.sandbox.resistance_report import build_resistance_report
    from tools.sandbox.persistence import save_run_capsule_v2
    from tools.sandbox.export_capsule import export_run_capsule_zip
    from tools.sandbox.export_capsule import import_run_capsule_zip
    from tools.sandbox.persistence import load_run_capsule_v2, diff_capsules
    from tools.sandbox.advanced_features import constraint_surface_map
    from tools.sandbox.conflict_atlas import new_atlas, update_atlas, summarize_atlas
    from tools.sandbox.design_navigation import steering_cues_from_surface_map, filter_cues
    from tools.sandbox.lineage_graph import build_lineage_edges, compute_tree_layout
    from tools.sandbox.spend_map import build_spend_scatter
    from tools.sandbox.robustness_envelope import robustness_envelope_from_records
    from tools.sandbox.narrative_pack import build_narrative
    from tools.sandbox.design_card import build_design_card_md
    from tools.sandbox.existence_report import existence_report
    from tools.sandbox.archive_intelligence import ladder_histogram, regime_clusters_summary
    from tools.sandbox.confidence_sweep import confidence_sweep
    from tools.sandbox.design_packet import build_design_packet_files
    from tools.sandbox.review_room import build_review_trinity, build_attack_simulation

    # v203 Reactor Design Forge: PROCESS-independence instruments
    from tools.sandbox.closure_console import closure_console
    from tools.sandbox.margin_budget import margin_budget
    from tools.sandbox.reality_gates import reality_gates
    from tools.sandbox.report_pack import build_report_pack


    from src.economics.cost import cost_proxies
    # Tier 5–6 instruments
    from tools.sandbox.tier56 import (
        ConstraintCred,
        apply_credibility_overlay,
        counterfactual_gate,
        build_intent_trajectory,
        why_not_report,
        discovered_relations,
        export_relations_markdown,
        inverse_design_residual,
    )

    # Tier 7 + Epistemic guarantees (collaboration + standards)
    from tools.sandbox.tier7 import (
        repo_fingerprint,
        candidate_fingerprint,
        generate_cert_badge_svg,
        export_doi_ready_pack,
        new_review_session,
        default_sessions_dir,
        save_review_session,
        load_review_session,
        export_review_session_zip,
        import_review_session_zip,
        run_regression_suite,
    )

    # Tier 8–9: design-space jurisprudence, intent-conditional laws, genealogy, counter-optimization
    from tools.sandbox.tier89 import (
        feasibility_confidence_from_trace,
        candidate_verdict,
        region_verdict,
        intent_conditional_laws,
        reconstruct_genealogy,
        counter_optimization_report,
    )

    

    # -------------------------
    # Intent Compiler (v285.0)
    # -------------------------
    if _forge_deck == "🧩 Intent Compiler":
        st.markdown("### 🧩 Intent Compiler")
        st.caption("Deterministic algebraic compilation from intent → candidate PointInputs. Produces candidates only; truth remains in Point Designer.")

        try:
            from tools.sandbox.intent_compiler import compile_intent_to_candidate
        except Exception as _e:
            st.error(f"Intent compiler import failed: {_e}")
            compile_intent_to_candidate = None  # type: ignore

        # Use last Point Designer inputs as base if available
        _base_obj = st.session_state.get('pd_last_inputs_obj')
        if _base_obj is None:
            _base_obj = st.session_state.get('base_point_inputs_obj')
        if not isinstance(_base_obj, PointInputs):
            # fall back to a safe, minimal default
            _base_obj = PointInputs(R0_m=3.0, a_m=1.0, kappa=1.8, Bt_T=12.0, Ip_MA=8.0, Ti_keV=10.0, fG=0.8, Paux_MW=20.0)

        c1, c2 = st.columns(2)
        Pfus = c1.number_input('Target fusion power P_fus (MW)', value=140.0, step=10.0, min_value=0.0)
        Q = c2.number_input('Target Q (proxy)', value=2.0, step=0.1, min_value=0.01)

        st.markdown('**Optional direct overrides (applied after compilation)**')
        o1, o2, o3, o4 = st.columns(4)
        o_R0 = o1.number_input('Override R0 (m) (0=ignore)', value=0.0, step=0.1)
        o_a  = o2.number_input('Override a (m) (0=ignore)', value=0.0, step=0.05)
        o_Bt = o3.number_input('Override Bt (T) (0=ignore)', value=0.0, step=0.5)
        o_Ip = o4.number_input('Override Ip (MA) (0=ignore)', value=0.0, step=0.5)

        overrides = {}
        if o_R0 > 0: overrides['R0_m'] = float(o_R0)
        if o_a  > 0: overrides['a_m']  = float(o_a)
        if o_Bt > 0: overrides['Bt_T'] = float(o_Bt)
        if o_Ip > 0: overrides['Ip_MA'] = float(o_Ip)

        if st.button('Compile candidate', type='primary', use_container_width=True, disabled=(compile_intent_to_candidate is None)):
            status, payload = compile_intent_to_candidate(_base_obj, Pfus_target_MW=float(Pfus), Q_target=float(Q), overrides=overrides)
            st.session_state['forge_intent_compiler_last'] = {'status': status, **payload}

        last = st.session_state.get('forge_intent_compiler_last')
        if isinstance(last, dict):
            st.info(f"Compiler status: **{last.get('status','?')}**")
            if last.get('reason'):
                st.error(str(last.get('reason')))
            if last.get('trace'):
                with st.expander('Compilation trace', expanded=False):
                    for ln in list(last.get('trace') or []):
                        st.markdown(f"- {ln}")
            cand = last.get('candidate_inputs')
            if isinstance(cand, dict):
                with st.expander('Candidate inputs (dict)', expanded=False):
                    st.json(cand)
                if st.button('Apply candidate in Point Designer', use_container_width=True):
                    st.session_state['pd_candidate_apply'] = dict(cand)
                    st.success('Candidate applied: go to "🧭 Point Designer" and press Evaluate.')

        # End Intent Compiler deck

# -------------------------
    # Replay / Diff (capsules)
    # -------------------------
    if _forge_deck == "🔁 Capsules":
        with st.expander("🔁 Replay & Diff — Run Capsules", expanded=False):
            st.caption("Load a previously exported Optimization Run Capsule (.zip or .json) to restore the Workbench. "
                       "This is metadata replay; truth remains the frozen evaluator.")
    
            c1, c2 = st.columns(2)
            up1 = c1.file_uploader("Restore capsule", type=["zip", "json"], key="opt_restore_capsule")
            if up1 is not None:
                try:
                    # zip capsule
                    if str(up1.name).lower().endswith(".zip"):
                        tmp = Path(".shams_state") / "_uploads"
                        tmp.mkdir(parents=True, exist_ok=True)
                        p = tmp / str(up1.name)
                        p.write_bytes(up1.getbuffer())
                        data = import_run_capsule_zip(p)
                        capsule = data.get("capsule") or {}
                    else:
                        capsule = json.loads(up1.getvalue().decode("utf-8"))
                    if str(capsule.get("schema")) != "shams.opt_sandbox.run_capsule.v2":
                        st.error(f"Unsupported capsule schema: {capsule.get('schema')}")
                    else:
                        st.session_state["opt_workbench_run"] = {
                            "kind": "optimization_sandbox_replay",
                            "intent": capsule.get("intent"),
                            "seed": capsule.get("seed", 1),
                            "objectives": capsule.get("lens", {}).get("objectives", []),
                            "var_specs": capsule.get("var_specs", []),
                            "budgets": {"bounds": capsule.get("bounds", {})},
                            "archive": capsule.get("archive", []),
                            "trace": capsule.get("trace", []),
                            "telemetry": capsule.get("telemetry", {}),
                            "resistance_report": capsule.get("resistance_report"),
                            "capsule_v2": capsule,
                            "non_authoritative_notice": "Replayed from capsule. Truth remains the frozen evaluator.",
                        }
                        st.success("Capsule restored into the Workbench.")
                except Exception as e:
                    st.error(f"Failed to restore capsule: {e}")
    
            st.markdown("---")
            st.caption("Diff two capsules (lens/bounds/counts/ladder histogram).")
            upA = c1.file_uploader("Capsule A", type=["json"], key="opt_diff_a")
            upB = c2.file_uploader("Capsule B", type=["json"], key="opt_diff_b")
            if upA is not None and upB is not None:
                try:
                    ca = json.loads(upA.getvalue().decode("utf-8"))
                    cb = json.loads(upB.getvalue().decode("utf-8"))
                    d = diff_capsules(ca, cb)
                    st.json(d)
                except Exception as e:
                    st.error(f"Diff failed: {e}")
    
        st.stop()

    # -------------------------
    # Evaluator (frozen truth)
    # -------------------------
    def _evaluate_candidate(inp: dict, intent: str) -> dict:
        """Audit a candidate with frozen evaluator. Returns a rich dict for archive/trace."""
        # Build PointInputs (fills defaults and validates)
        pi = PointInputs(**inp)
        outputs = _ui_evaluate(pi, origin="audit_candidate")
        cons = build_constraints_from_outputs(outputs, design_intent=intent)
        records = constraints_to_records(cons)
        feas = feasibility_flag(records, design_intent=intent)
        act = active_constraints(records, design_intent=intent)
        fm = failure_mode(records, design_intent=intent)
        # compute min margin
        min_sm = None
        for r in records:
            try:
                sm = float(r.get("signed_margin"))
            except Exception:
                continue
            if min_sm is None or sm < min_sm:
                min_sm = sm
        # optional cost proxies (pure outputs->cost)
        cost = cost_proxies(outputs) if isinstance(outputs, dict) else {}
        closure_bundle = closure_console(outputs=outputs, cost_proxy=cost) if isinstance(outputs, dict) else {"ok": False, "reason": "outputs_not_dict"}
        mb = margin_budget(records)
        rg = reality_gates(records, closure_bundle if isinstance(closure_bundle, dict) else None)
        rp = build_report_pack(intent=str(intent), inputs=dict(inp), outputs=outputs, constraints=records, closure_bundle=closure_bundle if isinstance(closure_bundle, dict) else None, margin_budget=mb, reality_gates=rg)

        return {
            "inputs": dict(inp),
            "outputs": outputs,
            "constraints": records,
            "feasible": bool(feas),
            "active_constraints": act,
            "failure_mode": fm,
            "min_signed_margin": float(min_sm) if min_sm is not None else float("nan"),
            "cost": cost,
            "closure_bundle": closure_bundle,
            "margin_budget": mb,
            "reality_gates": rg,
            "report_pack": rp,
            "closure_certificate": (rp.get("json") or {}).get("closure_certificate"),
            "design_class": (rp.get("json") or {}).get("design_class"),
            "citation_blocks": (rp.get("json") or {}).get("citation_blocks"),
            "reference_context": (rp.get("json") or {}).get("reference_context"),

        }

    # -------------------------
    # Guided runs (user friendly)
    # -------------------------
    st.markdown("### 🎯 Intent & Lens (explicit contract)")
    st.caption("Pick a goal pack; SHAMS will *show the exact objectives and bounds* before running.")

    # Intent selection
    intent_label = st.selectbox("Design intent", ["Power Reactor (net-electric)", "Experimental Device (research)"], index=0, key="opt_intent")

    # Internal canonical intent key (feeds constraint policy)
    intent = "Reactor" if intent_label.lower().startswith("power") else "Research"

    # Objective packs (explicit)
    packs = default_objective_packs(intent)
    pack_names = [p.name for p in packs] + ["Custom (manual objectives)"]
    pack_choice = st.selectbox("Objective pack", pack_names, index=0, key="opt_pack_choice")

    # Anchor: either current Point Designer inputs (if available) or a sensible baseline
    anchor_default = {}
    if "point_inputs_last" in st.session_state and isinstance(st.session_state["point_inputs_last"], dict):
        anchor_default = dict(st.session_state["point_inputs_last"])
    # fallback: minimal anchor (PointInputs will fill defaults)
    if not anchor_default:
        anchor_default = {"R0_m": 6.2, "a_m": 2.0, "kappa": 1.8, "delta": 0.33, "Bt_T": 5.3, "Ip_MA": 15.0, "Paux_MW": 50.0}

    # Choose variables + bounds (table-style)
    st.markdown("### 🧬 Degrees of Freedom (search space)")
    st.caption("You control what the machine finder is allowed to change. Frozen variables never move.")

    default_vars = ["R0_m", "Bt_T", "Ip_MA", "Paux_MW"]
    all_keys = list(anchor_default.keys())
    # Add common knobs even if absent
    for k in ["R0_m","a_m","kappa","delta","Bt_T","Ip_MA","Paux_MW","nbar_1e20_m3","Ti_keV"]:
        if k not in all_keys:
            all_keys.append(k)

    var_keys = st.multiselect("Variables to optimize", options=all_keys, default=default_vars, key="opt_var_keys")
    st.caption("Tip: start with 3–5 variables for stability; expand later.")

    # Bounds helper
    bound_mode = st.radio("Bounds mode", ["Tight (±10%)", "Medium (±20%)", "Wide (±35%)", "Custom"], index=1, horizontal=True, key="opt_bound_mode")
    frac = {"Tight (±10%)":0.10, "Medium (±20%)":0.20, "Wide (±35%)":0.35}.get(bound_mode, 0.20)

    bounds = {}
    for k in var_keys:
        v0 = float(anchor_default.get(k, 0.0))
        if bound_mode != "Custom":
            lo, hi = v0*(1-frac), v0*(1+frac)
        else:
            lo, hi = v0*(1-0.2), v0*(1+0.2)
        bounds[k] = (lo, hi)

    with st.expander("Edit bounds (table)", expanded=False):
        cols = st.columns([2,2,2])
        cols[0].markdown("**Variable**")
        cols[1].markdown("**Min**")
        cols[2].markdown("**Max**")
        for k in var_keys:
            lo, hi = bounds[k]
            c1, c2, c3 = st.columns([2,2,2])
            c1.write(k)
            lo2 = c2.number_input(f"{k}_lo", value=float(lo), key=f"b_lo_{k}")
            hi2 = c3.number_input(f"{k}_hi", value=float(hi), key=f"b_hi_{k}")
            if hi2 < lo2:
                hi2 = lo2
            bounds[k] = (float(lo2), float(hi2))

    # Objectives display + custom editor
    if pack_choice != "Custom (manual objectives)":
        pack = next(p for p in packs if p.name == pack_choice)
        objectives = [Objective(**o.__dict__) for o in pack.objectives]
        st.info(f"**Pack:** {pack.description}")
    else:
        st.caption("Custom objectives: add 1–3 objectives. Sense: max/min. Weight: explicit.")
        obj_rows = st.number_input("Number of objectives", 1, 3, 2, key="opt_n_obj")
        objectives = []
        for i in range(int(obj_rows)):
            c1, c2, c3 = st.columns([3,1,1])
            key = c1.text_input(f"Objective {i+1} key", value=["P_e_net_MW","Q_DT_eqv","q_div_MW_m2"][i] if i<3 else "Q_DT_eqv", key=f"obj_key_{i}")
            sense = c2.selectbox(f"Sense {i+1}", ["max","min"], index=0, key=f"obj_sense_{i}")
            weight = c3.number_input(f"Weight {i+1}", value=1.0, key=f"obj_w_{i}")
            objectives.append(Objective(key=key, sense=sense, weight=float(weight)))


    # Program Lens (objective contract) - explicit, exported (no hidden ranking)
    lens_contract = {
        "name": str(pack_choice),
        "description": str(pack.description) if pack_choice != "Custom (manual objectives)" else "Custom objectives (manual)",
        "intent": str(intent),
        "objectives": [{"key": o.key, "sense": o.sense, "weight": float(o.weight)} for o in (objectives or [])],
    }
    st.session_state["opt_lens_contract"] = lens_contract

    # Optional costing layer
    st.markdown("### 💸 Transparent costing layer (optional)")
    use_cost = st.checkbox("Enable cost proxies in objectives/filters (transparent)", value=False, key="opt_use_cost")
    if use_cost:
        st.caption("Cost proxies are computed from outputs; assumptions are explicit and exported with the run.")

    # Budgets / engine tuning (simple)
    st.markdown("### ⚙️ Run budget (fast-first → deeper)")
    cA,cB,cC,cD = st.columns(4)
    pop_size = cA.number_input("Pop size", 20, 200, 64, key="opt_pop")
    generations = cB.number_input("Global generations", 5, 200, 40, key="opt_gens")
    surrogate_rounds = cC.number_input("Surrogate rounds", 0, 30, 6, key="opt_surr")
    local_steps = cD.number_input("Local steps", 0, 300, 70, key="opt_local")
    archive_topk = st.slider("Archive size (top-k diverse)", 20, 200, 60, key="opt_topk")

    # Guardrails
    st.markdown("### 🛡️ Guardrails (feasibility governance)")
    min_margin = st.number_input("Require min signed margin ≥ (optional)", value=0.0, step=0.01, key="opt_min_margin")
    require_feasible_only = st.checkbox("Archive: keep feasible only (recommended)", value=True, key="opt_feas_only")

    # Advanced capabilities (Tier 1–4)
    st.markdown("### 🧠 Advanced instruments (Tier 1–4)")
    c1, c2, c3 = st.columns(3)
    enable_surface = c1.checkbox("Constraint-surface surfing", value=True, key="opt_adv_surface")
    enable_skeleton = c2.checkbox("Feasibility skeleton", value=True, key="opt_adv_skeleton")
    enable_memory = c3.checkbox("Active learning across runs (opt-in)", value=False, key="opt_adv_memory")
    c4, c5 = st.columns(2)
    enable_multi_intent = c4.checkbox("Track distance to the other Intent", value=False, key="opt_adv_multi_intent")
    staged = c5.checkbox("Staged run (human-in-the-loop phases)", value=False, key="opt_adv_staged")
    if staged:
        st.caption("Staged run executes phases one-by-one (Global → Surrogate → Local → Surf). Useful for steering and learning.")

    # Build an evaluator closure for this panel so post-run instruments (cartography/UQ)
    # can reuse it even when the UI reruns.
    def _make_eval_fn():
        def _fn(cand_inputs: dict):
            res = _evaluate_candidate(cand_inputs, intent=intent)
            # Expose cost proxies as objective keys (transparent, explicit)
            try:
                if isinstance(res.get("outputs"), dict) and isinstance(res.get("cost"), dict):
                    for ck, cv in res["cost"].items():
                        if ck not in res["outputs"]:
                            res["outputs"][ck] = cv
            except Exception:
                pass

            # Multi-intent instrumentation (distance-to-other)
            if enable_multi_intent:
                other_intent = "Research" if str(intent) == "Reactor" else "Reactor"
                try:
                    oth = _evaluate_candidate(cand_inputs, intent=other_intent)
                    oth_v = 0.0
                    for rr in (oth.get("constraints") or []):
                        try:
                            sm = float(rr.get("signed_margin", float("nan")))
                            if sm < 0:
                                oth_v += (-sm)
                        except Exception:
                            continue
                    res["other_intent"] = other_intent
                    res["other_feasible"] = bool(oth.get("feasible", False))
                    res["other_violation"] = float(oth_v)
                    res["other_min_signed_margin"] = float(oth.get("min_signed_margin", float("nan")))
                    res["other_failure_mode"] = oth.get("failure_mode")
                except Exception:
                    pass

            # Guardrails (do not change evaluator truth; mark infeasible for archive filtering)
            if min_margin and float(min_margin) > 0:
                try:
                    if float(res.get("min_signed_margin", float("nan"))) < float(min_margin):
                        res["feasible"] = False
                        res["failure_mode"] = res.get("failure_mode") or "min_margin_guardrail"
                except Exception:
                    pass
            return res
        return _fn

    eval_fn = _make_eval_fn()

    # Run control
    run_now = st.button("Run machine finder", type="primary", use_container_width=True, key="opt_run_button")

    if run_now:
        # Build VarSpecs
        var_specs = [VarSpec(key=k, lo=bounds[k][0], hi=bounds[k][1]) for k in var_keys]
        budgets = {
            "pop_size": int(pop_size),
            "generations": int(generations),
            "surrogate_rounds": int(surrogate_rounds),
            "propose_per_round": 36,
            "local_steps": int(local_steps),
            "archive_topk": int(archive_topk),
            "resistance_window": 250,
            "enable_surface_surf": bool(enable_surface),
            "enable_skeleton": bool(enable_skeleton),
            "use_knowledge_store": bool(enable_memory),
        }
        if staged:
            # Human-in-the-loop staged run (no background execution): phases are executed
            # one at a time and stored in session_state.
            st.session_state["opt_stage_state"] = {
                "intent": intent,
                "anchor": dict(anchor_default),
                "var_specs": [v.__dict__ for v in var_specs],
                "objectives": [o.__dict__ for o in objectives],
                "budgets": dict(budgets),
                "all_points": [],
                "trace": [],
                "done": {"global": False, "surrogate": False, "local": False, "surf": False},
                "seed": 1,
            }
            st.info("Staged run initialized. Use the phase controls below in the Workbench.")
            run = None
        else:
            run = run_hybrid_machine_finder(
                evaluate_fn=eval_fn,
                intent=intent,
                anchor_inputs=anchor_default,
                var_specs=var_specs,
                objectives=objectives,
                budgets=budgets,
                seed=1,
            )

        if isinstance(run, dict):
            # Post-filter archive if requested
            if require_feasible_only:
                run["archive"] = [a for a in run.get("archive", []) if a.get("feasible", False)]

            # vNext: rebuild archive with diversity + dominance annotation (explicit objectives)
            try:
                run["archive"] = build_archive(run.get("archive", []) or [], var_specs, topk=int(archive_topk), objectives=objectives)
            except Exception:
                pass

            # vNext: feasibility ladder classification (archive + trace)
            try:
                for c in (run.get("archive") or []):
                    c.update(classify_candidate(c, dominant=bool(c.get("is_dominant", False))))
                for t in (run.get("trace") or []):
                    t.update(classify_candidate(t))
            except Exception:
                pass

            # vNext: resistance report (descriptive)
            try:
                lens_contract = st.session_state.get("opt_lens_contract") or {}
                bounds_dict = {k: list(bounds[k]) for k in var_keys if k in bounds}
                var_specs_dicts = [v.__dict__ for v in var_specs]
                run["resistance_report"] = build_resistance_report(
                    trace=run.get("trace") or [],
                    archive=run.get("archive") or [],
                    intent=intent,
                    lens_contract=lens_contract,
                    bounds=bounds_dict,
                    var_specs=var_specs_dicts,
                )
            except Exception:
                pass

            st.session_state["opt_workbench_run"] = run
            st.success("Run complete. Workbench updated below.")

    # -------------------------
    # Post-run Workbench
    # -------------------------
    stage_state = st.session_state.get("opt_stage_state")
    run = st.session_state.get("opt_workbench_run")

    # If staged run is active, provide phase controls and build a live workbench run view
    if stage_state is not None and isinstance(stage_state, dict):
        st.markdown("---")
        st.markdown("## Forge Workbench (Staged run)")
        st.caption("Execute phases one-by-one. Nothing runs in the background.")

        # Rehydrate
        _intent = stage_state.get("intent")
        _anchor = stage_state.get("anchor") or {}
        _seed = int(stage_state.get("seed", 1))
        _var_specs = [VarSpec(**v) for v in (stage_state.get("var_specs") or [])]
        _objectives = [Objective(**o) for o in (stage_state.get("objectives") or [])]
        _budgets = stage_state.get("budgets") or {}

        done = stage_state.get("done") or {}
        ph1, ph2, ph3, ph4 = st.columns(4)
        if ph1.button("Run Global", use_container_width=True, disabled=forge_lock or bool(done.get("global"))):
            pts, tr = global_de_phase(
                evaluate_fn=eval_fn,
                anchor_inputs=_anchor,
                var_specs=_var_specs,
                objectives=_objectives,
                pop_size=int(_budgets.get("pop_size", 64)),
                generations=int(_budgets.get("generations", 40)),
                seed=_seed,
            )
            stage_state["all_points"] = (stage_state.get("all_points") or []) + pts
            stage_state["trace"] = (stage_state.get("trace") or []) + tr
            stage_state["done"]["global"] = True
            st.session_state["opt_stage_state"] = stage_state
            st.success("Global phase complete.")
        if ph2.button("Run Surrogate", use_container_width=True, disabled=forge_lock or (not bool(done.get("global")) or bool(done.get("surrogate")))):
            pts, tr = surrogate_phase(
                evaluate_fn=eval_fn,
                anchor_inputs=_anchor,
                var_specs=_var_specs,
                objectives=_objectives,
                seed=_seed,
                history=list(stage_state.get("all_points") or []),
                rounds=int(_budgets.get("surrogate_rounds", 6)),
                propose_per_round=int(_budgets.get("propose_per_round", 36)),
            )
            stage_state["all_points"] = (stage_state.get("all_points") or []) + pts
            stage_state["trace"] = (stage_state.get("trace") or []) + tr
            stage_state["done"]["surrogate"] = True
            st.session_state["opt_stage_state"] = stage_state
            st.success("Surrogate phase complete.")
        if ph3.button("Run Local", use_container_width=True, disabled=forge_lock or (not bool(done.get("global")) or bool(done.get("local")))):
            pts, tr = local_refine_phase(
                evaluate_fn=eval_fn,
                anchor_inputs=_anchor,
                var_specs=_var_specs,
                objectives=_objectives,
                seed=_seed,
                seeds=list(stage_state.get("all_points") or []),
                steps=int(_budgets.get("local_steps", 70)),
            )
            stage_state["all_points"] = (stage_state.get("all_points") or []) + pts
            stage_state["trace"] = (stage_state.get("trace") or []) + tr
            stage_state["done"]["local"] = True
            st.session_state["opt_stage_state"] = stage_state
            st.success("Local phase complete.")
        if ph4.button("Run Surf", use_container_width=True, disabled=forge_lock or (not bool(done.get("local")) or bool(done.get("surf")))):
            pts, tr = surface_surf_phase(
                evaluate_fn=eval_fn,
                anchor_inputs=_anchor,
                var_specs=_var_specs,
                objectives=_objectives,
                seed=_seed,
                seeds=list(stage_state.get("all_points") or []),
                steps=int(_budgets.get("surf_steps", 80)),
            )
            stage_state["all_points"] = (stage_state.get("all_points") or []) + pts
            stage_state["trace"] = (stage_state.get("trace") or []) + tr
            stage_state["done"]["surf"] = True
            st.session_state["opt_stage_state"] = stage_state
            st.success("Surf phase complete.")

        # Build a live run view from staged state
        _archive = build_archive(list(stage_state.get("all_points") or []), _var_specs, topk=int(_budgets.get("archive_topk", 60)))
        _trace = list(stage_state.get("trace") or [])
        _resist = resistance_atlas(_trace, last_n=int(_budgets.get("resistance_window", 250)))
        _corr = variable_correlations(_archive, _var_specs)
        _skel = build_feasibility_skeleton(_archive, _var_specs) if enable_skeleton else None
        run = {
            "kind": "optimization_sandbox_hybrid_run_staged",
            "intent": str(_intent),
            "seed": int(_seed),
            "objectives": [o.__dict__ for o in _objectives],
            "var_specs": [v.__dict__ for v in _var_specs],
            "budgets": dict(_budgets),
            "archive": _archive,
            "trace": _trace,
            "resistance": _resist,
            "variable_correlations": _corr,
            "feasibility_skeleton": _skel,
        }
        st.session_state["opt_workbench_run"] = run

    if isinstance(run, dict) and run.get("archive") is not None:
        st.markdown("---")
        st.markdown("## Forge Workbench")

        # -------------------------
        # Run Dashboard (workflow view)
        # -------------------------
        d1, d2, d3, d4 = st.columns(4)
        with d1:
            st.markdown("**Run Contract**")
            st.write({
                "intent": run.get("intent"),
                "lens": (run.get("capsule_v2") or {}).get("lens", st.session_state.get("opt_lens_contract")),
                "seed": run.get("seed"),
            })
        with d2:
            st.markdown("**Live Trace**")
            tr = run.get("trace") or []
            n_t = len(tr)
            n_f = sum(1 for t in tr if bool(t.get("feasible", False)))
            st.write({"n_evaluated": int(n_t), "n_feasible": int(n_f)})
            # top failure mode snapshot
            fm = {}
            for t in tr[-250:]:
                k = str(t.get("failure_mode") or "")
                if not k:
                    continue
                fm[k] = fm.get(k, 0) + 1
            if fm:
                top = sorted(fm.items(), key=lambda x: x[1], reverse=True)[:3]
                st.caption("Recent failure modes")
                st.write({k: int(v) for k, v in top})
        with d3:
            st.markdown("**Candidate Archive**")
            ar = run.get("archive") or []
            n_a = len(ar)
            n_af = sum(1 for a in ar if bool(a.get("feasible", False)))
            n_dom = sum(1 for a in ar if bool(a.get("is_dominant", False)))
            st.write({"n_archive": int(n_a), "n_feasible": int(n_af), "n_dominant": int(n_dom)})
        with d4:
            st.markdown("**Resistance**")
            rr = run.get("resistance_report")
            if isinstance(rr, dict):
                topb = rr.get("primary_blockers") or rr.get("blockers") or []
                if isinstance(topb, list) and topb:
                    st.write({"top_blocker": topb[0].get("name", topb[0]) if isinstance(topb[0], dict) else topb[0]})
                else:
                    st.write("(no blockers reported)")
            else:
                st.write("(no resistance report)")

        # Budget allocation (transparent scheduler)
        ba = run.get("budget_allocation")
        if isinstance(ba, dict):
            with st.expander("Budget allocation (feasibility-first scheduler)", expanded=False):
                st.json(ba)

        # Conflict atlas (accumulates across runs in-session)
        rr = run.get("resistance_report")
        if "opt_conflict_atlas" not in st.session_state or not isinstance(st.session_state.get("opt_conflict_atlas"), dict):
            st.session_state["opt_conflict_atlas"] = new_atlas()
        if isinstance(rr, dict):
            try:
                st.session_state["opt_conflict_atlas"] = update_atlas(st.session_state["opt_conflict_atlas"], rr)
            except Exception:
                pass
        with st.expander("Conflict atlas (across runs, descriptive)", expanded=False):
            st.caption("Accumulated from Resistance Reports. Descriptive only - not causal, not prescriptive.")
            rows = summarize_atlas(st.session_state.get("opt_conflict_atlas") or {}, top_n=25)
            if rows:
                import pandas as _pd
                st.dataframe(_pd.DataFrame(rows), use_container_width=True)
            else:
                st.info("No conflicts accumulated yet. Run optimization at least once.")
            st.download_button(
                "Download conflict atlas (json)",
                data=json.dumps(st.session_state.get("opt_conflict_atlas") or {}, indent=2, sort_keys=True),
                file_name="shams_opt_conflict_atlas.json",
                mime="application/json",
                use_container_width=True,
                key="opt_dl_conflict_atlas",
            )

        # vNext: Run capsule (v2) + resistance report (descriptive, exportable)
        with st.expander("Run capsule + resistance report (export)", expanded=False):
            lens_contract = st.session_state.get("opt_lens_contract") or {}
            rr = run.get("resistance_report")
            if isinstance(rr, dict):
                st.caption("Resistance report (descriptive)")
                st.json(rr, expanded=False)
            else:
                st.info("Resistance report not available for this run.")

            if st.button("Build capsule zip (v2)", use_container_width=True, key="opt_build_capsule_zip"):
                try:
                    import time, os, json
                    run_id = f"run_{int(time.time())}"
                    settings = {
                        "bounds": {k: list(v) for k, v in (st.session_state.get("opt_bounds") or {}).items()} if isinstance(st.session_state.get("opt_bounds"), dict) else {},
                        "var_specs": run.get("var_specs") or [],
                        "objectives": run.get("objectives") or [],
                    }
                    evaluator_hash = str(run.get("fingerprint") or "")
                    cap_path = save_run_capsule_v2(
                        run,
                        run_id=run_id,
                        settings=settings,
                        evaluator_hash=evaluator_hash,
                        archive=run.get("archive") or [],
                        trace=run.get("trace") or [],
                        lens_contract=lens_contract,
                        resistance_report=rr if isinstance(rr, dict) else None,
                    )
                    # Also export a compact zip with manifest
                    from pathlib import Path
                    out_zip = Path(cap_path).with_suffix(".zip")
                    export_run_capsule_zip(
                        capsule=json.loads(Path(cap_path).read_text(encoding="utf-8")),
                        archive={"schema":"shams.opt_sandbox.archive_snapshot.v1","archive": run.get("archive") or []},
                        resistance_report=rr if isinstance(rr, dict) else None,
                        out_path=out_zip,
                    )
                    st.session_state["opt_capsule_zip_bytes"] = out_zip.read_bytes()
                    st.session_state["opt_capsule_zip_name"] = out_zip.name
                    st.success("Capsule zip built.")
                except Exception as e:
                    st.error(f"Capsule build failed: {e}")

            if st.session_state.get("opt_capsule_zip_bytes") is not None:
                st.download_button(
                    "Download capsule zip",
                    data=st.session_state["opt_capsule_zip_bytes"],
                    file_name=str(st.session_state.get("opt_capsule_zip_name","opt_capsule.zip")),
                    mime="application/zip",
                    use_container_width=True,
                    key="opt_dl_capsule_zip",
                )

        # Sticky truth bar (simple CSS)
        st.markdown(
            """
            <style>
            div[data-testid="stVerticalBlock"] div:has(> div.shams-sticky) { position: sticky; top: 0; z-index: 999; background: white; }
            .shams-sticky { border: 1px solid rgba(49,51,63,0.15); padding: 10px; border-radius: 10px; }
            </style>
            """,
            unsafe_allow_html=True
        )

        best = run.get("best_feasible")
        feas_rate = run.get("resistance", {}).get("feasible_rate")
        dom = run.get("resistance", {}).get("dominant_constraints", {})
        dom_top = sorted(dom.items(), key=lambda kv: kv[1], reverse=True)[:1]
        dom_txt = dom_top[0][0] if dom_top else "-"

        with st.container():
            st.markdown('<div class="shams-sticky">', unsafe_allow_html=True)
            t1,t2,t3,t4,t5 = st.columns([1,1,1,1,2])
            t1.metric("Intent", run.get("intent","-"))
            t2.metric("Feasible rate (recent)", f"{(float(feas_rate)*100.0):.1f}%" if feas_rate is not None else "-")
            # Keep score strictly labeled as non-authoritative (legacy search utility).
            t3.metric("Archive score (non-authoritative)", f"{float(best.get('_score')):.3g}" if isinstance(best, dict) else "-")
            t4.metric("Archive size", str(len(run.get("archive") or [])))
            t5.write(f"**Dominant resistance:** {dom_txt}")
            st.markdown('</div>', unsafe_allow_html=True)

        left, center, right = st.columns([1.2, 2.2, 1.4], vertical_alignment="top")

        # LEFT: nav/setup for post-run views
        with left:
            st.markdown("### Navigate")

            # v206: reduce scrolling fatigue for experts by using a searchable selectbox
            # and a "Cockpit mode" that keeps the most-used instruments on one screen.
            cockpit_mode = st.toggle(
                "Forge Cockpit Mode",
                value=True,
                help="Compact, low-scroll layout for experts. Keeps the core instruments together.",
                key="rdf_cockpit_mode",
                disabled=forge_lock,
            )
            if forge_lock:
                cockpit_mode = True

            _views_core = [
                "Casebook",
                "Candidate Archive",
                "Forge Timeline",
                "Machine Dossier",
                "Review Trinity",
                "Attack Simulation",
                "Resistance Brief",
                "Scan ↔ Forge Grounding",
                "Conflict Atlas",
                "Boundary Navigator",
                "Constraint Spend Map",
                "Reactor Accounting Console",
                "Margin Ledger",
                "Reality Gates",
                "Closure Certificate",
                "Provenance Graph",
                "Engineering Reality Budget",
                "Failure-Mode Canon",
                "Design Class",
                "Citation Blocks",
                "Reference Reproduction",
                "Economics Deck",
                "Robustness Envelope",
                "Design Narrative",
                "Design Card",
                "Design Packet",
                "Confidence Sweep",
                "Expert Compare (no ranking)",
                "Exposure Readiness",
                "Epistemic Gap Map",
                "Constraint Personas",
                "Design Genealogy",
                "Do‑Not‑Build Brief",
                "Process of Elimination",
                "Paper‑Ready Signals",
                "Silence Mode"
                "Sensitivity Fingerprint",
                "Reviewer Packet"
            ]
            _views_full = _views_core + [
                "Archive regimes & coverage",
                "Machine existence report",
                "Design navigation (steering)",
                "Pareto (if multi-objective)",
                "Report Pack",
                "Trace Telemetry",
                "Feasibility skeleton",
                "Local cartography (adaptive)",
                "Uncertainty (Monte Carlo)",
                "Intent trajectories (Research→Reactor)",
                "Inverse design / Why not?",
                "Discovered relations (laws)",
                "Counterfactual lens",
                "PROCESS parity benchmarks",
                "Parity validation packs (PASS/WARN/FAIL)",
                "Parity calibration (reference deltas)",
                "Decision scenarios (program lens)",
                "Collaboration (review sessions)",
                "Epistemic guarantees (regression suite)",
                "Standards & DOI export",
                "Design-space verdicts (Allowed/Forbidden)",
                "Epistemic confidence bounds",
                "Intent-conditional design laws",
                "Machine genealogy",
                "Counter-optimization (no interior optimum)",
                "Reproducibility",
            ]

            _views = _views_core if cockpit_mode else _views_full
            _default_view = st.session_state.get("opt_view") or ("Review Trinity" if forge_lock else "Casebook")
            if _default_view not in _views:
                _default_view = "Review Trinity" if forge_lock else "Casebook"
            view = st.selectbox(
                "Main view",
                options=_views,
                index=int(_views.index(_default_view)),
                key="opt_view",
                help="Type to search. Cockpit mode shows the core instruments first.",
            )

            # Back-compat: keep internal handlers stable while we improve naming for fusion experts.
            _view_alias = {
                "Casebook": "Casebook Runner",
                "Candidate Archive": "Archive Bay",
                "Forge Timeline": "Timeline Strip",
                "Review Trinity": "Review Trinity",
                "Attack Simulation": "Attack Simulation",
                "Scan ↔ Forge Grounding": "Scan ↔ Forge Grounding",
                "Exposure Readiness": "Exposure Readiness",
            }
            view = _view_alias.get(str(view), str(view))
            st.markdown("### Archive filters")
            only_robust = st.checkbox("Keep only margin≥0", value=False, key="opt_filter_robust")
            min_score = st.number_input("Min score (optional)", value=float("-inf"), key="opt_filter_minscore")
            if use_cost:
                st.markdown("### 💰 Cost filter")
                max_coe = st.number_input("Max COE proxy (optional)", value=float("inf"), key="opt_filter_coe")



            # ------------------------------
            # Review Bench (Compare Tray)
            # ------------------------------
            with st.expander("Review Bench (compare tray)", expanded=False):
                st.caption("Pin a handful of candidates for side-by-side review. Descriptive only - no ranking.")

                # Current candidate (filtered archive index)
                _cur = None
                try:
                    if filt:
                        _i = int(st.session_state.get("opt_inspect_idx", 0) or 0)
                        _i = max(0, min(len(filt)-1, _i))
                        _cur = filt[_i]
                except Exception:
                    _cur = None

                if "opt_review_bench" not in st.session_state or not isinstance(st.session_state.get("opt_review_bench"), list):
                    st.session_state["opt_review_bench"] = []

                cA, cB = st.columns([1,1])
                with cA:
                    if st.button("Pin current candidate", use_container_width=True, key="opt_pin_current"):
                        if _cur is None:
                            st.warning("No current candidate to pin.")
                        else:
                            try:
                                _fid = candidate_fingerprint(_cur)
                            except Exception:
                                _fid = f"idx:{int(st.session_state.get('opt_inspect_idx',0) or 0)}"
                            # avoid duplicates
                            if not any(str(x.get("id")) == str(_fid) for x in st.session_state["opt_review_bench"]):
                                inp = _cur.get("inputs") or {}
                                out = _cur.get("outputs") or {}
                                st.session_state["opt_review_bench"].append({
                                    "id": str(_fid),
                                    "idx": int(st.session_state.get("opt_inspect_idx", 0) or 0),
                                    "R0_m": inp.get("R0_m"),
                                    "Bt_T": inp.get("Bt_T"),
                                    "Ip_MA": inp.get("Ip_MA"),
                                    "P_e_net_MW": out.get("P_e_net_MW"),
                                    "Pfus_total_MW": out.get("Pfus_total_MW"),
                                    "first_failure": _cur.get("first_failure"),
                                    "min_signed_margin": _cur.get("min_signed_margin"),
                                })
                                st.success("Pinned.")
                            else:
                                st.info("Already pinned.")
                with cB:
                    if st.button("Clear bench", use_container_width=True, key="opt_clear_bench"):
                        st.session_state["opt_review_bench"] = []
                        st.success("Cleared.")

                bench = st.session_state.get("opt_review_bench") or []
                if bench:
                    import pandas as _pd
                    dfb = _pd.DataFrame(bench)
                    st.dataframe(dfb, use_container_width=True, hide_index=True)

                    # Exports
                    csv_bytes = dfb.to_csv(index=False).encode("utf-8")
                    st.download_button("Download Review Bench (CSV)", data=csv_bytes, file_name="review_bench.csv", mime="text/csv", use_container_width=True)
                    md = "|" + "|".join(dfb.columns) + "|\n"
                    md += "|" + "|".join(["---"]*len(dfb.columns)) + "|\n"
                    for _, r in dfb.iterrows():
                        md += "|" + "|".join(str(r[c]) for c in dfb.columns) + "|\n"
                    st.download_button("Download Review Bench (Markdown)", data=md.encode("utf-8"), file_name="review_bench.md", mime="text/markdown", use_container_width=True)
                else:
                    st.info("Bench is empty. Pin 2–5 candidates during review.")

            # ------------------------------
            # Tier 5–6 controls (integrated)
            # ------------------------------
            with st.expander("Tier 5–6 instruments (optional)", expanded=False):
                st.caption(
                    "These are advanced instruments that PROCESS cannot provide. "
                    "They never modify frozen truth; they add explanatory lenses and workflows."
                )

                # Counterfactual constraint gate
                st.markdown("**Counterfactual gate (hypothetical)**")
                st.caption("Disable a constraint only in the feasibility *gate* for analysis. Raw constraints remain unchanged.")
                _all_names = []
                try:
                    if (run.get("archive") or []) and (run.get("archive")[0].get("constraints") or []):
                        _all_names = [str(r.get("name")) for r in (run.get("archive")[0].get("constraints") or []) if r.get("name")]
                        _all_names = sorted(list(dict.fromkeys(_all_names)))
                except Exception:
                    _all_names = []
                disabled_cons = st.multiselect(
                    "Disable constraints (hypothetical)",
                    options=_all_names,
                    default=[],
                    key="opt_cf_disable",
                )

                # Credibility overlay
                st.markdown("**Constraint credibility overlay**")
                st.caption("Optional epistemic lens: maturity/uncertainty adjusts *displayed* margins and filters. Does not change feasibility truth.")
                use_cred = st.checkbox("Enable credibility overlay", value=False, key="opt_use_cred")
                cred_map = {}
                if use_cred and _all_names:
                    # Keep a small editable set (top 8 by occurrence in archive)
                    top = _all_names[:8]
                    for nm in top:
                        c1,c2,c3 = st.columns([2,1,1])
                        with c1:
                            st.write(nm)
                        with c2:
                            mat = st.slider(f"maturity_{nm}", 0.0, 1.0, 0.7, 0.05, key=f"cred_m_{nm}")
                        with c3:
                            unc = st.slider(f"uncertainty_{nm}", 0.0, 0.5, 0.10, 0.01, key=f"cred_u_{nm}")
                        cred_map[nm] = ConstraintCred(name=nm, maturity=float(mat), uncertainty_frac=float(unc), conservative=True)
                # Persist for center/inspector usage
                st.session_state["opt_cred_map"] = {
                    k: {"name": v.name, "maturity": float(v.maturity), "uncertainty_frac": float(v.uncertainty_frac), "conservative": bool(v.conservative)}
                    for k, v in cred_map.items()
                }

                # Trajectory settings
                st.markdown("**Intent trajectories**")
                st.caption("Build a simple Research→Reactor highway from the current archive (and current variable list).")
                traj_steps = st.slider("Max path steps", 2, 10, 5, key="opt_traj_steps")
                if st.button("Build trajectory", use_container_width=True, key="opt_traj_build"):
                    st.session_state["opt_traj"] = None
                    try:
                        # Use current archive (pre-filter) to keep it deterministic
                        cands = run.get("archive") or []
                        st.session_state["opt_traj"] = build_intent_trajectory(
                            evaluate_fn=_evaluate_candidate,
                            candidates=cands,
                            var_keys=var_keys,
                            from_intent="Research",
                            to_intent="Reactor",
                            k_steps=int(traj_steps),
                            seed=int(run.get("seed", 0)),
                        )
                    except Exception as _e:
                        st.session_state["opt_traj"] = {"ok": False, "reason": str(_e)}

                # Inverse design target capture (used in center view)
                st.markdown("**Inverse design targets**")
                st.caption("Define desired outputs; SHAMS finds the closest feasible candidate *within your declared bounds*. No relaxation.")
                inv_cols = st.columns(2)
                inv_k = inv_cols[0].text_input("Target key", value="P_e_net_MW", key="opt_inv_key")
                inv_v = inv_cols[1].number_input("Target value", value=500.0, key="opt_inv_val")
                st.session_state["opt_inv_targets"] = {str(inv_k): float(inv_v)}

        archive = run.get("archive") or []
        # Apply filters
        filt = []
        for a in archive:
            if only_robust and float(a.get("min_signed_margin", float("nan"))) < 0:
                continue
            if float(a.get("_score", -1e30)) < float(min_score):
                continue
            if use_cost:
                coe = (a.get("cost") or {}).get("COE_proxy")
                if coe is not None and float(coe) > float(max_coe):
                    continue
            filt.append(a)

        # -------------------------
        # v206: expert signals (descriptive)
        # -------------------------
        def _regime_signature(_cand: dict) -> list:
            """Fusion-expert friendly regime tags (descriptive only)."""
            tags = []
            inp = _cand.get("inputs") or {}
            out = _cand.get("outputs") or {}

            def _get_num(d, k):
                try:
                    v = d.get(k)
                    return None if v is None else float(v)
                except Exception:
                    return None

            R0 = _get_num(inp, "R0_m") or _get_num(out, "R0_m")
            a = _get_num(inp, "a_m") or _get_num(out, "a_m")
            Ip = _get_num(inp, "Ip_MA") or _get_num(out, "Ip_MA")
            B0 = _get_num(inp, "B0_T") or _get_num(out, "B0_T") or _get_num(out, "Bt_T")
            Pf = _get_num(out, "Pfus_total_MW")

            # Geometry regime
            if R0 is not None:
                if R0 < 2.5:
                    tags.append("compact")
                elif R0 > 6.0:
                    tags.append("large-R")
            if R0 is not None and a is not None and a > 0:
                A = R0 / a
                if A < 2.2:
                    tags.append("spherical")
                elif A > 3.2:
                    tags.append("high-aspect")

            # Field/current regimes (heuristic, descriptive)
            if B0 is not None:
                if B0 >= 10.0:
                    tags.append("high-field")
                elif B0 <= 4.0:
                    tags.append("low-field")
            if Ip is not None:
                if Ip >= 12.0:
                    tags.append("high-current")

            # Power density proxy
            if Pf is not None and R0 is not None and R0 > 0:
                pd = Pf / (R0 ** 3)
                if pd >= 20.0:
                    tags.append("power-dense")

            # Feasibility state
            fs = _cand.get("feasibility_state")
            if fs:
                tags.append(str(fs).replace("feasible_", ""))
            return tags[:8]

        def _first_kill(_cand: dict) -> dict:
            """Return the tightest constraint (first-kill) from margin ledger rows."""
            mb = _cand.get("margin_budget") or {}
            rows = mb.get("rows") or []
            if not rows:
                return {"name": _cand.get("first_failure") or _cand.get("failure_mode") or "-", "signed_margin": _cand.get("min_signed_margin")}
            best = None
            for r in rows:
                try:
                    sm = float(r.get("signed_margin"))
                except Exception:
                    continue
                if best is None or sm < best[0]:
                    best = (sm, r)
            if best is None:
                return {"name": _cand.get("first_failure") or "-", "signed_margin": _cand.get("min_signed_margin")}
            rr = best[1]
            return {"name": rr.get("name") or rr.get("constraint") or "-", "signed_margin": float(best[0])}

        def _constraint_spend_rate(_cand: dict) -> dict:
            """Local heuristic: margin change per objective change vs parent (if lineage exists)."""
            pid = _cand.get("parent_id") or _cand.get("parent")
            if not pid:
                return {"ok": False, "reason": "no parent link"}
            parent = None
            for c in (run.get("archive") or []):
                if (c.get("_id") or c.get("fingerprint")) == pid:
                    parent = c
                    break
            if parent is None:
                return {"ok": False, "reason": "parent not found in archive"}

            # choose one objective if available
            obj = None
            lens = run.get("lens") or {}
            objs = lens.get("objectives") if isinstance(lens, dict) else None
            if isinstance(objs, list) and objs:
                obj = objs[0].get("key") if isinstance(objs[0], dict) else None
            if not obj:
                obj = "P_e_net_MW" if "P_e_net_MW" in (_cand.get("outputs") or {}) else None
            if not obj:
                return {"ok": False, "reason": "no objective key"}

            def _val(c, key):
                try:
                    return float((c.get("outputs") or {}).get(key))
                except Exception:
                    return None

            d_obj = None
            c_obj = _val(_cand, obj)
            p_obj = _val(parent, obj)
            if c_obj is not None and p_obj is not None:
                d_obj = c_obj - p_obj

            try:
                d_m = float(_cand.get("min_signed_margin")) - float(parent.get("min_signed_margin"))
            except Exception:
                d_m = None

            if d_obj is None or d_m is None or abs(d_obj) < 1e-12:
                return {"ok": False, "reason": "insufficient delta"}

            return {
                "ok": True,
                "objective": obj,
                "delta_objective": d_obj,
                "delta_min_margin": d_m,
                "margin_spend_per_objective": (d_m / d_obj),
                "note": "Local heuristic vs parent only (descriptive).",
            }

        # --- v208: Scan ↔ Forge grounding (descriptive topology context) ---
        def _scan_grounding(_cand: dict, _scan_artifact: dict, *, intent: str) -> dict:
            """Attach nearest-point scan context to a candidate (descriptive only)."""
            try:
                art = _scan_artifact or {}
                rep = (art.get("report") or {}) if isinstance(art, dict) else {}
                pts = rep.get("points") or []
                xk = rep.get("x_key"); yk = rep.get("y_key")
                if not pts or not xk or not yk:
                    return {"ok": False, "reason": "scan artifact missing points/x_key/y_key"}
                cin = _cand.get("inputs") or {}
                if xk not in cin or yk not in cin:
                    return {"ok": False, "reason": "candidate lacks scan axes", "x_key": xk, "y_key": yk}
                cx = float(cin.get(xk)); cy = float(cin.get(yk))
                best = None
                best_d = None
                for p in pts:
                    try:
                        dx = float(p.get("x")) - cx
                        dy = float(p.get("y")) - cy
                        d = (dx*dx + dy*dy) ** 0.5
                    except Exception:
                        continue
                    if best is None or (best_d is not None and d < best_d) or (best_d is None):
                        best = p; best_d = d
                if best is None:
                    return {"ok": False, "reason": "no valid scan points"}
                it = str(intent or "Reactor")
                it_sum = ((best.get("intent") or {}).get(it) if isinstance(best.get("intent"), dict) else None) or {}
                top = ((rep.get("topology") or {}).get(it) if isinstance(rep.get("topology"), dict) else None) or {}
                return {
                    "ok": True,
                    "scan_id": (rep.get("id") or art.get("report_hash") or art.get("artifact_hash")),
                    "x_key": xk,
                    "y_key": yk,
                    "candidate_xy": {"x": cx, "y": cy},
                    "nearest": {
                        "i": int(best.get("i", -1)),
                        "j": int(best.get("j", -1)),
                        "x": float(best.get("x")),
                        "y": float(best.get("y")),
                        "robustness": it_sum.get("robustness"),
                        "dominant_blocking": it_sum.get("dominant_blocking"),
                        "min_blocking_margin": it_sum.get("min_blocking_margin"),
                    },
                    "distance": float(best_d) if best_d is not None else None,
                    "topology": top,
                    "note": "Descriptive grounding only. Does not modify evaluator truth.",
                }
            except Exception as e:
                return {"ok": False, "reason": f"grounding_error: {e}"}

        # CENTER: main canvas
        with center:
            st.markdown("### Canvas")

            # v208: Margin-first framing (always-on summary for the inspected candidate)
            try:
                if filt:
                    _i = int(st.session_state.get("opt_inspect_idx", 0) or 0)
                    _i = max(0, min(len(filt) - 1, _i))
                    _c = filt[_i]
                    _mb = _c.get("margin_budget")
                    if not isinstance(_mb, dict):
                        _mb = margin_budget(_c.get("constraints") or [])
                        _c["margin_budget"] = _mb
                    _rows = _mb.get("rows") or []
                    _tight = []
                    for r in _rows:
                        if isinstance(r, dict) and r.get("name"):
                            _tight.append(r)
                    _tight = sorted(_tight, key=lambda rr: float(rr.get("margin_frac", 1e30) or 1e30))[:5]
                    with st.expander(f"{_LANG.get('margin_first')}", expanded=False if forge_lock else False):
                        c1,c2,c3 = st.columns(3)
                        c1.metric("Min signed margin", str(_c.get("min_signed_margin")))
                        c2.metric("Dominant resistance", str(_c.get("first_failure") or _c.get("failure_mode") or "-"))
                        c3.metric("Feasible", str(bool(_c.get("feasible"))))
                        if _tight:
                            st.write({str(r.get("name")): r.get("margin_frac") for r in _tight})
            except Exception:
                pass

            # v206: dedicated Conflict Atlas view (also shown in the right rail in cockpit mode)
            if view == "Conflict Atlas":
                st.caption("Constraint Conflict Atlas (descriptive, accumulated across runs).")
                rows = summarize_atlas(st.session_state.get("opt_conflict_atlas") or {}, top_n=50)
                if rows:
                    import pandas as _pd
                    st.dataframe(_pd.DataFrame(rows), use_container_width=True, height=520)
                else:
                    st.info("No conflicts accumulated yet. Run at least one case.")
                st.stop()

            # --- v204–v205: Design intelligence + confidence instruments ---
            if view == "Timeline Strip":
                st.caption("v204: Timeline strip of the current run (phases + evaluations).")
                tr = run.get("trace") or []
                if not tr:
                    st.info("No trace available.")
                else:
                    try:
                        import pandas as _pd
                        rows = []
                        for i, t in enumerate(tr):
                            rows.append({
                                "i": i,
                                "phase": t.get("phase") or t.get("step") or "",
                                "feasible": bool(t.get("feasible")) if t.get("feasible") is not None else None,
                                "failure": t.get("failure_mode") or t.get("failure") or "",
                                "score": t.get("score") or t.get("_score"),
                            })
                        df = _pd.DataFrame(rows)
                        st.dataframe(df, use_container_width=True, height=420)
                    except Exception:
                        st.json(tr[:200])
                st.stop()

            if view == "Lineage Graph":
                st.caption("v204: Design lineage graph based on recorded parents (audit-clean).")
                if not (run.get("archive") or []):
                    st.info("No archive available.")
                    st.stop()
                edges = build_lineage_edges(run.get("archive") or [])
                if not edges:
                    st.info("No explicit parent links found in archive candidates. (Fallback: use 'Machine genealogy' for reconstructed ancestry.)")
                    st.stop()
                layout = compute_tree_layout(edges)
                try:
                    import pandas as _pd
                    import plotly.graph_objects as _go

                    # Edges as segments
                    xs = []
                    ys = []
                    for p, c in edges:
                        if p not in layout or c not in layout:
                            continue
                        xs += [layout[p]["x"], layout[c]["x"], None]
                        ys += [layout[p]["y"], layout[c]["y"], None]
                    fig = _go.Figure()
                    fig.add_trace(_go.Scatter(x=xs, y=ys, mode="lines", name="lineage"))

                    # Nodes
                    ndf = _pd.DataFrame([
                        {"id": nid, "x": v["x"], "y": v["y"], "depth": v["depth"]}
                        for nid, v in layout.items()
                    ])
                    fig.add_trace(_go.Scatter(x=ndf["x"], y=ndf["y"], mode="markers+text", text=ndf["id"], textposition="top center", name="nodes"))
                    fig.update_layout(height=520, margin=dict(l=10, r=10, t=30, b=10), title="Lineage Graph")
                    st.plotly_chart(fig, use_container_width=True)
                except Exception:
                    st.write({"edges": edges[:200], "layout": layout})
                st.stop()

            if view == "Constraint Spend Map":
                st.caption("v204: Spend map - where feasibility margin is being spent.")
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                xk, yk = st.columns(2)
                x_key = xk.text_input("X axis input key", value=var_keys[0] if var_keys else "Ip_MA", key="spend_x")
                y_key = yk.text_input("Y axis input key", value=var_keys[1] if len(var_keys) > 1 else "R0_m", key="spend_y")
                mode = st.selectbox("Color by", ["min_margin", "feasibility_state", "constraint_margin"], index=0, key="spend_color")
                con_key = None
                if mode == "constraint_margin":
                    con_key = st.text_input("Constraint key", value="q_div", key="spend_ck")
                scat = build_spend_scatter(filt, x_key=str(x_key), y_key=str(y_key), color_by=str(mode), constraint_key=str(con_key) if con_key else None)
                try:
                    import pandas as _pd
                    import plotly.express as _px
                    df = _pd.DataFrame({"x": scat["x"], "y": scat["y"], "c": scat["c"], "id": scat["ids"]})
                    fig = _px.scatter(df, x="x", y="y", hover_data=["id"], color="c")
                    fig.update_layout(height=520, margin=dict(l=10, r=10, t=20, b=10))
                    st.plotly_chart(fig, use_container_width=True)
                except Exception:
                    st.json(scat)
                st.stop()

            if view == "Robustness Envelope":
                st.caption("v205: Robustness envelope (first-order margin perturbation sweep).")
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                env = robustness_envelope_from_records(cand.get("constraints") or [])
                if not env.get("ok"):
                    st.info(env.get("reason"))
                else:
                    try:
                        import pandas as _pd
                        st.line_chart(_pd.DataFrame({"pass_fraction": env["pass_fraction"]}, index=[str(p) for p in env["perturbations"]]))
                    except Exception:
                        st.write(env)
                st.json(env, expanded=False)
                st.stop()

            if view == "Design Narrative":
                st.caption("v205: Design narrative pack (review-grade, no recommendations).")
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                nar = build_narrative(cand)
                st.markdown(nar.get("markdown") or "")
                st.download_button("Download narrative (md)", data=(nar.get("markdown") or "").encode("utf-8"), file_name="design_narrative.md", mime="text/markdown")
                st.json(nar, expanded=False)
                st.stop()

            if view == "Design Card":
                st.caption("v205: One-page design card (printable, reviewer-friendly).")
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                card = build_design_card_md(cand)
                st.markdown(card)
                st.download_button("Download Design Card (md)", data=card.encode("utf-8"), file_name="design_card.md", mime="text/markdown")
                st.stop()

            if view == "Design Packet":
                st.caption("v207: Design Packet - narrative + card + key tables (PDF best-effort).")
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                nar = build_narrative(cand)
                card = build_design_card_md(cand)
                title = f"SHAMS Design Packet - Candidate {idx_sel}"
                pkt = build_design_packet_files(title=title, card_md=card, narrative_md=(nar.get('markdown') or ''), candidate=cand)
                md = pkt.get('markdown') or ''
                st.markdown(md)
                st.download_button("Download Design Packet (md)", data=md.encode('utf-8'), file_name="design_packet.md", mime="text/markdown", use_container_width=True)
                pdfb = pkt.get('pdf_bytes')
                if pdfb:
                    st.download_button("Download Design Packet (pdf)", data=pdfb, file_name="design_packet.pdf", mime="application/pdf", use_container_width=True)
                else:
                    st.info("PDF rendering unavailable (markdown export is authoritative).")
                with st.expander("Packet metadata (json)", expanded=False):
                    st.json({k: v for k, v in pkt.items() if k not in ('markdown','pdf_bytes')}, expanded=False)
                st.stop()

            if view == "Confidence Sweep":
                st.caption("v207: Confidence Sweep - explicit declared perturbations (no hidden penalties, no recommendations).")
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                recs = cand.get('constraints') or []
                cb = cand.get('closure_bundle') or {}
                cs = confidence_sweep(records=recs, closure_bundle=cb)
                if not cs.get('ok'):
                    st.warning(cs.get('reason'))
                else:
                    c1,c2,c3 = st.columns(3)
                    c1.metric("Verdict", cs.get('verdict'))
                    c2.metric("Min pass fraction", f"{float(cs.get('min_pass_fraction',0.0))*100:.1f}%")
                    fk = cs.get('first_kill_tally') or {}
                    top = sorted(fk.items(), key=lambda kv: kv[1], reverse=True)[:1]
                    c3.metric("Most common first-kill", top[0][0] if top else '-')
                    try:
                        import pandas as _pd
                        df = _pd.DataFrame({"pass_fraction": cs.get('pass_fraction') or []}, index=[str(x) for x in (cs.get('margin_deltas') or [])])
                        st.line_chart(df)
                    except Exception:
                        pass
                    with st.expander("First-kill tally", expanded=False):
                        st.write(fk)
                    with st.expander("Proxy headlines", expanded=False):
                        st.write(cs.get('proxy_headlines') or [])
                st.json(cs, expanded=False)
                st.download_button("Download Confidence Sweep (json)", data=json.dumps(cs, indent=2, sort_keys=True), file_name="confidence_sweep.json", mime="application/json", use_container_width=True)
                st.stop()

            # --- v208: Review-room instruments ---
            if view == "Scan ↔ Forge Grounding":
                st.caption("Ground the current candidate in Scan Lab topology (descriptive).")
                sa = st.session_state.get("scan_cartography_artifact")
                if not isinstance(sa, dict):
                    st.info("No Scan Lab artifact found in session. Run Scan Lab or upload a scan artifact there first.")
                    st.stop()
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                sg = _scan_grounding(cand, sa, intent=str(run.get("intent") or "Reactor"))
                if not sg.get("ok"):
                    st.warning(str(sg.get("reason")))
                st.json(sg, expanded=False)
                st.download_button("Download grounding (json)", data=json.dumps(sg, indent=2, sort_keys=True), file_name="scan_forge_grounding.json", mime="application/json", use_container_width=True)
                st.stop()

            # ---- Supremacy Instruments (descriptive, review-room)
            if view == "Epistemic Gap Map":
                from tools.sandbox import forge_supremacy_plus as fsp
                st.caption("Epistemic Gap Map: make model limits explicit (honesty signaling, not UQ).")
                ctx = {
                    "assumptions_ledger": st.session_state.get("assumptions_ledger_text") or "",
                    "model_ledger": st.session_state.get("model_ledger_text") or "",
                    "notes": st.session_state.get("run_notes") or "",
                }
                gaps = fsp.epistemic_gap_map(ctx)
                for k, items in gaps.items():
                    with st.expander(k, expanded=False):
                        for it in items:
                            st.write(f"- {it}")
                st.download_button("Download Gap Map (JSON)",
                                   data=json.dumps(gaps, indent=2).encode("utf-8"),
                                   file_name="epistemic_gap_map.json",
                                   mime="application/json")
                st.stop()

            if view == "Constraint Personas":
                from tools.sandbox import forge_supremacy_plus as fsp
                st.caption("Constraint Personas: memorable behavioral profiles (descriptive).")
                personas = fsp.constraint_personas()
                for cname, prof in personas.items():
                    with st.expander(cname, expanded=False):
                        for kk, vv in prof.items():
                            st.write(f"**{kk}:** {vv}")
                st.download_button("Download Personas (JSON)",
                                   data=json.dumps(personas, indent=2).encode("utf-8"),
                                   file_name="constraint_personas.json",
                                   mime="application/json")
                st.stop()

            if view == "Design Genealogy":
                from tools.sandbox import forge_supremacy_plus as fsp
                st.caption("Design Genealogy: lineage view (when lineage metadata exists).")
                cand_list = st.session_state.get("opt_archive_filtered") or st.session_state.get("opt_archive") or []
                md = fsp.genealogy_markdown(list(cand_list) if isinstance(cand_list, list) else [])
                st.markdown(md)
                st.download_button("Download Genealogy (MD)",
                                   data=md.encode("utf-8"),
                                   file_name="design_genealogy.md",
                                   mime="text/markdown")
                st.stop()

            if view == "Do‑Not‑Build Brief":
                from tools.sandbox import forge_supremacy_plus as fsp
                st.caption("Do‑Not‑Build Brief: reasons *not* to build (trust ledger).")
                filt = st.session_state.get("opt_archive_filtered") or st.session_state.get("opt_archive") or []
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                ctx = {
                    "margin_ledger": st.session_state.get("opt_margin_ledger"),
                    "conflicts": st.session_state.get("opt_conflicts"),
                    "first_kill_under_uncertainty": st.session_state.get("opt_first_kill_uncertainty"),
                }
                brief = fsp.do_not_build_brief(cand, ctx)
                st.subheader(brief["title"])
                st.caption(brief.get("posture", ""))
                for r in brief["reasons"]:
                    st.write(f"- {r}")
                st.download_button("Download Do‑Not‑Build Brief (JSON)",
                                   data=json.dumps(brief, indent=2).encode("utf-8"),
                                   file_name="do_not_build_brief.json",
                                   mime="application/json")
                st.stop()

            if view == "Process of Elimination":
                from tools.sandbox import forge_supremacy_plus as fsp
                st.caption("Process of Elimination: why most machines cannot exist (constraint narrative).")
                ctx = {
                    "first_failure_histogram": st.session_state.get("scan_first_failure_hist"),
                    "dominant_killers": st.session_state.get("scan_dominant_killers"),
                }
                md = fsp.elimination_narrative(ctx)
                st.markdown(md)
                st.download_button("Download Elimination Narrative (MD)",
                                   data=md.encode("utf-8"),
                                   file_name="process_of_elimination.md",
                                   mime="text/markdown")
                st.stop()

            if view == "Paper‑Ready Signals":
                from tools.sandbox import forge_supremacy_plus as fsp
                st.caption("Paper‑Ready Signals: stable figure/table IDs for deterministic replay.")
                filt = st.session_state.get("opt_archive_filtered") or st.session_state.get("opt_archive") or []
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                sig = fsp.paper_ready_signals(cand)
                for item in sig["paper_ready_signals"]:
                    st.write(f"**{item['id']}** - {item['title']}  ·  `{item['ref']}`")
                st.download_button("Download Paper‑Ready Signals (JSON)",
                                   data=json.dumps(sig, indent=2).encode("utf-8"),
                                   file_name="paper_ready_signals.json",
                                   mime="application/json")
                st.stop()

            if view == "Silence Mode":
                st.caption("Silence Mode: review-room calm. Suppresses celebratory UI noise (no effect on physics).")
                st.session_state["silence_mode"] = st.toggle("Enable Silence Mode", value=bool(st.session_state.get("silence_mode", False)))
                if st.session_state.get("silence_mode"):
                    st.info("Silence Mode is ON. Prefer artifacts over narration.")
                else:
                    st.info("Silence Mode is OFF.")
                st.stop()
            if view == "Sensitivity Fingerprint":
                st.caption("Constraint Sensitivity Fingerprint: small perturbation fragility tags (screening-level).")
                from tools.sandbox import sensitivity_fingerprint as sf
                filt = st.session_state.get("opt_archive_filtered") or st.session_state.get("opt_archive") or []
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                # Use existing point evaluator wrapper if available; otherwise show info only.
                evaluator = st.session_state.get("_frozen_point_evaluator")
                if evaluator is None:
                    st.warning("No evaluator wrapper found in session. Run any evaluation to enable fingerprints.")
                    st.stop()
                fp = sf.build_fingerprint(cand, evaluator=evaluator)
                for t in fp.get("tags", []):
                    st.write(f"- {t}")
                if fp.get("notes"):
                    with st.expander("Notes", expanded=False):
                        for n in fp["notes"]:
                            st.write(f"- {n}")
                st.download_button("Download Fingerprint (JSON)",
                                   data=json.dumps(fp, indent=2).encode("utf-8"),
                                   file_name="sensitivity_fingerprint.json",
                                   mime="application/json")
                st.stop()

            if view == "Reviewer Packet":
                st.caption("One-click Reviewer Packet: Markdown bundle + key artifacts (descriptive, deterministic).")
                filt = st.session_state.get("opt_archive_filtered") or st.session_state.get("opt_archive") or []
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                from tools.sandbox.reviewer_packet_builder import ReviewerPacketOptions, build_reviewer_packet_zip
                with st.expander("Packet composition", expanded=False):
                    c1, c2, c3 = st.columns(3)
                    include_report = c1.checkbox("Include Report Pack", value=True, key="v322_rp_inc_report")
                    include_trinity = c2.checkbox("Include Review Trinity", value=True, key="v322_rp_inc_trinity")
                    include_attack = c3.checkbox("Include Attack Simulation", value=True, key="v322_rp_inc_attack")
                    c4, c5, c6 = st.columns(3)
                    include_scan = c4.checkbox("Include Scan Grounding", value=True, key="v322_rp_inc_scan")
                    include_capsule = c5.checkbox("Include Run Capsule", value=True, key="v322_rp_inc_capsule")
                    include_manifests = c6.checkbox("Include Repo Manifests", value=True, key="v322_rp_inc_manif")

                cap = run.get("capsule_v2") if isinstance(run, dict) else None
                sa = st.session_state.get("scan_cartography_artifact")
                sg = _scan_grounding(cand, sa, intent=str(run.get("intent") or "Reactor")) if isinstance(sa, dict) else None
                dnb = st.session_state.get("do_not_build_brief_latest")
                opts = ReviewerPacketOptions(
                    include_core_docs=True,
                    include_candidate_snapshot=True,
                    include_report_pack=bool(include_report),
                    include_review_trinity=bool(include_trinity),
                    include_attack_simulation=bool(include_attack),
                    include_scan_grounding=bool(include_scan),
                    include_run_capsule=bool(include_capsule),
                    include_do_not_build_brief=True,
                    include_repo_manifests=bool(include_manifests),
                )
                zip_bytes, summary = build_reviewer_packet_zip(
                    candidate=cand,
                    repo_root=Path(__file__).resolve().parent.parent,
                    run_capsule=cap if isinstance(cap, dict) else None,
                    scan_grounding=sg if isinstance(sg, dict) else None,
                    do_not_build_brief=dnb if isinstance(dnb, dict) else None,
                    options=opts,
                )
                st.download_button(
                    "Download Reviewer Packet (ZIP)",
                    data=zip_bytes,
                    file_name="shams_reviewer_packet.zip",
                    mime="application/zip",
                    use_container_width=True,
                )
                with st.expander("Packet manifest (preview)", expanded=False):
                    st.json(summary.get("manifest") or {}, expanded=False)
                st.stop()
            if view == "Review Trinity":
                st.caption("Review Trinity: Existence Proof → Stress Story → Positioning.")
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                sa = st.session_state.get("scan_cartography_artifact")
                sg = _scan_grounding(cand, sa, intent=str(run.get("intent") or "Reactor")) if isinstance(sa, dict) else {}
                tri = build_review_trinity(candidate=cand, scan_grounding=sg if isinstance(sg, dict) else {})
                st.markdown(tri.get("markdown") or "")
                st.download_button("Download Review Trinity (md)", data=(tri.get("markdown") or "").encode("utf-8"), file_name="review_trinity.md", mime="text/markdown", use_container_width=True)
                st.download_button("Download Review Trinity (json)", data=json.dumps(tri, indent=2, sort_keys=True), file_name="review_trinity.json", mime="application/json", use_container_width=True)
                st.stop()

            if view == "Attack Simulation":
                st.caption("Hostile review rehearsal scaffold (no invented answers).")
                if not filt:
                    st.info("No candidates available.")
                    st.stop()
                idx_sel = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[int(max(0, min(idx_sel, len(filt)-1)))]
                cap = run.get("capsule_v2") if isinstance(run, dict) else None
                atk = build_attack_simulation(candidate=cand, run_capsule=cap if isinstance(cap, dict) else None)
                st.markdown(atk.get("markdown") or "")
                st.download_button("Download Attack Simulation (md)", data=(atk.get("markdown") or "").encode("utf-8"), file_name="attack_simulation.md", mime="text/markdown", use_container_width=True)
                st.download_button("Download Attack Simulation (json)", data=json.dumps(atk, indent=2, sort_keys=True), file_name="attack_simulation.json", mime="application/json", use_container_width=True)
                st.stop()

            if view == "Exposure Readiness":
                st.caption(_LANG.get("external_exposure_gate"))
                try:
                    _chk = (Path(__file__).resolve().parent.parent / "docs" / "EXTERNAL_EXPOSURE_CHECKLIST.md").read_text(encoding="utf-8")
                except Exception:
                    _chk = "(missing docs/EXTERNAL_EXPOSURE_CHECKLIST.md)"
                st.markdown(_chk)
                st.download_button("Download exposure checklist (md)", data=_chk.encode("utf-8"), file_name="EXTERNAL_EXPOSURE_CHECKLIST.md", mime="text/markdown", use_container_width=True)
                st.stop()
            if view in ("Reactor Accounting Console","Margin Ledger","Reality Gates","Economics Deck","Report Pack","Closure Certificate","Provenance Graph","Engineering Reality Budget","Failure-Mode Canon","Design Class","Citation Blocks","Reference Reproduction"):
                if not filt:
                    st.info("No candidates available in the archive (after filters).")
                else:
                    import pandas as _pd
                    def _lab(i):
                        a = filt[i]
                        ms = a.get("min_signed_margin")
                        try:
                            ms = float(ms)
                        except Exception:
                            ms = float('nan')
                        return f"{i:03d} | min_margin={ms:.3g} | {str(a.get('failure_mode') or '')}".strip()
                    idx = st.selectbox("Candidate", options=list(range(len(filt))), format_func=_lab, key="v203_rdf_candidate_pick")
                    a = filt[int(idx)]
                    if view == "Reactor Accounting Console":
                        st.caption("Explicit plant closure (derived). No hidden penalties.")
                        st.json(a.get("closure_bundle") or {}, expanded=False)
                    elif view == "Margin Ledger":
                        mb = a.get("margin_budget") or {}
                        rows = mb.get("rows") or []
                        if rows:
                            st.dataframe(_pd.DataFrame(rows), use_container_width=True)
                        else:
                            st.json(mb, expanded=False)
                    elif view == "Reality Gates":
                        st.json(a.get("reality_gates") or {}, expanded=False)
                    elif view == "Economics Deck":
                        cb = a.get("closure_bundle") or {}
                        env = (cb.get("economics_envelopes") if isinstance(cb, dict) else None) or []
                        if env:
                            st.dataframe(_pd.DataFrame(env), use_container_width=True)
                        else:
                            st.json(cb, expanded=False)
                    elif view == "Report Pack":
                        rp = a.get("report_pack") or {}
                        md = rp.get("markdown") or "(no report)"
                        st.markdown(md)
                        st.download_button("Download report JSON", data=json.dumps(rp.get("json") or {}, indent=2, sort_keys=True), file_name="shams_reactor_design_forge_report.json", mime="application/json", use_container_width=True, key="v203_dl_rdf_json")
                        st.download_button("Download report Markdown", data=str(md), file_name="shams_reactor_design_forge_report.md", mime="text/markdown", use_container_width=True, key="v203_dl_rdf_md")
                        st.download_button("Download report CSV", data=str(rp.get("csv") or ""), file_name="shams_reactor_design_forge_report.csv", mime="text/csv", use_container_width=True, key="v203_dl_rdf_csv")

                    elif view == "Closure Certificate":
                        from tools.sandbox.closure_certificate import build_closure_certificate
                        cert = a.get("closure_certificate") or build_closure_certificate(a)
                        st.subheader(f"Feasibility Closure Certificate - {cert.get('verdict')}")
                        st.json(cert, expanded=False)
                        st.download_button("Download Closure Certificate (JSON)",
                                           data=json.dumps(cert, indent=2, sort_keys=True),
                                           file_name="shams_closure_certificate.json",
                                           mime="application/json",
                                           use_container_width=True,
                                           key="rdf_dl_fcc")
                    elif view == "Provenance Graph":
                        from tools.sandbox.constraint_provenance_graph import build_cpg_for_constraint
                        st.caption("Constraint Provenance Graph: where each limit comes from (structure, not new physics).")
                        cname = st.text_input("Constraint name (e.g., q_div, sigma_vm, HTS margin, TBR, net_electric)", value=str(a.get("failure_mode") or "q_div"))
                        cpg = build_cpg_for_constraint(cname, intent=str(run.get("intent") or ""))
                        st.json(cpg, expanded=False)
                        st.download_button("Download CPG (JSON)",
                                           data=json.dumps(cpg, indent=2, sort_keys=True),
                                           file_name="shams_constraint_provenance_graph.json",
                                           mime="application/json",
                                           use_container_width=True,
                                           key="rdf_dl_cpg")
                    elif view == "Engineering Reality Budget":
                        st.caption("Engineering Reality Budget: grouped margin currency (descriptive).")
                        mb = a.get("margin_budget") or {}
                        rows = mb.get("rows") or []
                        groups = {"plasma": [], "materials": [], "thermal": [], "economics": [], "other": []}
                        for r in rows:
                            nm = str(r.get("name") or "").lower()
                            if any(k in nm for k in ["q95","bet","greenwald","h98","kappa","delta","stability","plasma"]):
                                groups["plasma"].append(r)
                            elif any(k in nm for k in ["sigma","stress","strain","tf","structure","hts","coil"]):
                                groups["materials"].append(r)
                            elif any(k in nm for k in ["q_div","heat","divert","sol","thermal","wall"]):
                                groups["thermal"].append(r)
                            elif any(k in nm for k in ["coe","cost","econ","net","recirc","electric"]):
                                groups["economics"].append(r)
                            else:
                                groups["other"].append(r)
                        for g,rs in groups.items():
                            with st.expander(f"{g.title()} budget", expanded=(g in ["plasma","materials","thermal"])):
                                if rs:
                                    import pandas as _pd
                                    st.dataframe(_pd.DataFrame(rs), use_container_width=True)
                                else:
                                    st.write("(no rows in this bucket)")
                        st.download_button("Download Reality Budget (JSON)",
                                           data=json.dumps(groups, indent=2, sort_keys=True),
                                           file_name="shams_engineering_reality_budget.json",
                                           mime="application/json",
                                           use_container_width=True,
                                           key="rdf_dl_erb")
                    elif view == "Failure-Mode Canon":
                        st.caption("Failure-Mode Canon: standardized non-prescriptive archetypes.")
                        canon = {
                            "heat-flux dominated": ["q_div", "divertor", "sol"],
                            "stress-limited": ["sigma", "stress", "vm"],
                            "hts-margin collapse": ["hts", "margin"],
                            "breeding-limited": ["tbr", "breed"],
                            "recirculation-trapped": ["recirc", "net", "electric"],
                            "coupled failure": ["+", "and", "coupled"],
                        }
                        fm_now = str(a.get("failure_mode") or "")
                        tag = "unclassified"
                        fml = fm_now.lower()
                        for k, toks in canon.items():
                            if any(t in fml for t in toks):
                                tag = k
                                break
                        st.write({"failure_mode": fm_now, "canonical_tag": tag})
                        st.json(canon, expanded=False)
                        st.download_button("Download Failure-Mode Canon (JSON)",
                                           data=json.dumps({"canonical_tag": tag, "canon": canon}, indent=2, sort_keys=True),
                                           file_name="shams_failure_mode_canon.json",
                                           mime="application/json",
                                           use_container_width=True,
                                           key="rdf_dl_fmc")
                    elif view == "Design Class":
                        dc = a.get("design_class") or {}
                        st.subheader(f"{dc.get('code','')} - {dc.get('name','')}")
                        st.json(dc, expanded=False)
                        st.download_button("Download Design Class (JSON)",
                                           data=json.dumps(dc, indent=2, sort_keys=True),
                                           file_name="shams_design_class.json",
                                           mime="application/json",
                                           use_container_width=True,
                                           key="rdf_dl_dc")
                    elif view == "Citation Blocks":
                        cb = a.get("citation_blocks") or {}
                        st.caption("Paste-ready Methods + citation scaffold (local repo content).")
                        st.text_area("Methods block", value=str(cb.get("methods_block") or ""), height=200)
                        if cb.get("citation_cff"):
                            with st.expander("CITATION.cff", expanded=False):
                                st.code(cb.get("citation_cff"))
                        st.download_button("Download Citation Blocks (JSON)",
                                           data=json.dumps(cb, indent=2, sort_keys=True),
                                           file_name="shams_citation_blocks.json",
                                           mime="application/json",
                                           use_container_width=True,
                                           key="rdf_dl_cb")
                    elif view == "Reference Reproduction":
                        rc = a.get("reference_context") or {}
                        st.caption("Historical reproduction: compare candidate to reference presets (anchors, not targets).")
                        refs = rc.get("refs") or []
                        if refs:
                            for ref in refs:
                                with st.expander(str(ref.get("ref")), expanded=False):
                                    st.json(ref.get("comparison") or {}, expanded=False)
                        else:
                            st.info("No reference context available for this candidate.")
                        st.download_button("Download Reference Context (JSON)",
                                           data=json.dumps(rc, indent=2, sort_keys=True),
                                           file_name="shams_reference_context.json",
                                           mime="application/json",
                                           use_container_width=True,
                                           key="rdf_dl_ref")

                st.stop()

            # --- v203 Reactor Design Forge panels (PROCESS-independence) ---
            def _pick_candidate(_cands):
                if not _cands:
                    return None
                opts = []
                for idx,a in enumerate(_cands):
                    ms = a.get('min_signed_margin')
                    try:
                        msf = float(ms)
                    except Exception:
                        msf = None
                    fp = a.get('fingerprint') or a.get('_id') or str(idx)
                    opts.append((f"{idx:03d} | min_margin={msf if msf is not None else 'na'} | {fp}", idx))
                label_to_i = {l:i for l,i in opts}
                lab = st.selectbox('Select candidate', options=[l for l,_ in opts], index=0, key='rdf_pick')
                return _cands[int(label_to_i.get(lab,0))]

            if view in ["Reactor Accounting Console","Margin Ledger","Reality Gates","Economics Deck","Report Pack"]:
                sel = _pick_candidate(filt)
                if sel is None:
                    st.info('Archive is empty. Run a case first.')
                else:
                    if view == "Reactor Accounting Console":
                        st.caption('Explicit plant/accounting closure derived from frozen truth outputs. No hidden assumptions.')
                        cb = sel.get('closure_bundle')
                        if not isinstance(cb, dict):
                            cb = closure_console(outputs=sel.get('outputs') or {}, cost_proxy=sel.get('cost') or {})
                        st.json(cb, expanded=False)
                    elif view == "Margin Ledger":
                        st.caption('Constraint margin budget (engineering accounting). Not a score.')
                        mb = sel.get('margin_budget')
                        if not isinstance(mb, dict):
                            mb = margin_budget(sel.get('constraints') or [])
                        import pandas as _pd
                        st.dataframe(_pd.DataFrame(mb.get('rows') or []), use_container_width=True)
                        st.write({'min_signed_margin': mb.get('min_signed_margin'), 'tight_constraints': mb.get('tight_constraints')})
                    elif view == "Reality Gates":
                        st.caption('Declared buildability gates (toggleable, descriptive).')
                        rg = sel.get('reality_gates')
                        if not isinstance(rg, dict):
                            rg = reality_gates(sel.get('constraints') or [], sel.get('closure_bundle') if isinstance(sel.get('closure_bundle'), dict) else None)
                        st.json(rg, expanded=False)
                    elif view == "Economics Deck":
                        st.caption('Explicit economics envelopes (Optimistic / Nominal / Conservative).')
                        cb = sel.get('closure_bundle')
                        if not isinstance(cb, dict):
                            cb = closure_console(outputs=sel.get('outputs') or {}, cost_proxy=sel.get('cost') or {})
                        env = (cb.get('economics_envelopes') or []) if isinstance(cb, dict) else []
                        import pandas as _pd
                        st.dataframe(_pd.DataFrame(env), use_container_width=True)
                    elif view == "Report Pack":
                        st.caption('PROCESS-recognizable report pack (audit-clean). No ranking, no recommendations.')
                        rp = sel.get('report_pack')
                        if not isinstance(rp, dict):
                            rp = build_report_pack(intent=str(run.get('intent')), inputs=sel.get('inputs') or {}, outputs=sel.get('outputs') or {}, constraints=sel.get('constraints') or [], closure_bundle=sel.get('closure_bundle') if isinstance(sel.get('closure_bundle'), dict) else None, margin_budget=sel.get('margin_budget') if isinstance(sel.get('margin_budget'), dict) else None, reality_gates=sel.get('reality_gates') if isinstance(sel.get('reality_gates'), dict) else None)
                        st.download_button('Download report JSON', data=json.dumps(rp.get('json') or {}, indent=2, sort_keys=True), file_name='shams_report_pack.json', mime='application/json', use_container_width=True)
                        st.download_button('Download report markdown', data=str(rp.get('markdown') or ''), file_name='shams_report_pack.md', mime='text/markdown', use_container_width=True)
                        st.download_button('Download report CSV', data=str(rp.get('csv') or ''), file_name='shams_report_pack.csv', mime='text/csv', use_container_width=True)
                        with st.expander('Preview markdown', expanded=False):
                            st.markdown(str(rp.get('markdown') or ''))

            if view == "Design navigation (steering)":
                st.caption("Steering cues are derived from a local linear surface map built from evaluated archive data. Descriptive only.")
                # derive variable keys from var_specs
                vs = run.get("var_specs") or []
                var_k = [str(v.get("key")) for v in vs if isinstance(v, dict) and v.get("key")]
                # constraint list
                cnames = []
                try:
                    if (run.get("archive") or []) and ((run.get("archive")[0].get("constraints") or [])):
                        cnames = sorted(list({str(c.get("name")) for c in (run.get("archive")[0].get("constraints") or []) if c.get("name")}))
                except Exception:
                    cnames = []
                if not cnames:
                    st.info("No constraint names available in archive.")
                else:
                    csel = st.selectbox("Constraint to navigate", options=cnames, index=0, key="opt_nav_constraint")
                    fam = st.selectbox("Lever family", options=["All","Geometry","🔥 Plasma","Power","🧲 Magnets","Other"], index=0, key="opt_nav_family")
                    smap = constraint_surface_map(archive=run.get("archive") or [], var_keys=var_k, constraint_name=str(csel))
                    if not smap.get("ok"):
                        st.warning(f"Surface map not available: {smap.get('reason')}")
                        st.json(smap)
                    else:
                        cues = steering_cues_from_surface_map(smap)
                        cues = filter_cues(cues, family=fam, top_n=15)
                        st.markdown("#### Steering cues (local, descriptive)")
                        if cues:
                            import pandas as _pd
                            st.dataframe(_pd.DataFrame(cues)[["family","lever","cue","signed","strength"]], use_container_width=True)
                        else:
                            st.info("No cues available (insufficient gradient data).")
                        with st.expander("Surface map details", expanded=False):
                            st.json(smap, expanded=False)

            elif view == "Machine existence report":
                st.caption("Explains why the currently selected candidate exists (or how close it is). No recommendations.")
                idx2 = int(st.session_state.get("opt_inspect_idx", 0))
                cand = filt[idx2] if filt and idx2 < len(filt) else (filt[0] if filt else None)
                if cand is None:
                    st.info("No candidate available.")
                else:
                    rep = existence_report(cand)
                    st.info(rep.get("narrative"))
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown("**Tight constraints**")
                        st.write(rep.get("tight", []))
                    with c2:
                        st.markdown("**Slack constraints**")
                        st.write(rep.get("slack", []))
                    with st.expander("Full existence report (json)", expanded=False):
                        st.json(rep, expanded=False)

            elif view == "Archive regimes & coverage":
                st.caption("Descriptive regime clustering and coverage cues for the feasible archive.")
                ar0 = run.get("archive") or []
                vs = run.get("var_specs") or []
                var_k = [str(v.get("key")) for v in vs if isinstance(v, dict) and v.get("key")]
                h = ladder_histogram(ar0)
                st.markdown("#### Feasibility ladder histogram")
                st.write(h)
                summ = regime_clusters_summary(archive=ar0, var_keys=var_k, max_k=10, seed=int(run.get("seed", 0) or 0))
                if not summ.get("ok"):
                    st.warning(summ.get("reason"))
                else:
                    import pandas as _pd
                    st.markdown("#### Regime clusters (feasible points)")
                    st.dataframe(_pd.DataFrame(summ.get("clusters") or []), use_container_width=True)
                    st.caption("Clusters are descriptive only; they do not imply optimality.")

            elif view == "Machine Dossier":
                st.caption("A structured, exportable dossier for the selected candidate. Descriptive only - no recommendations.")
                # Select candidate
                cand = None
                if filt:
                    try:
                        idx = int(st.session_state.get("opt_inspect_idx", 0) or 0)
                        idx = max(0, min(len(filt)-1, idx))
                        cand = filt[idx]
                    except Exception:
                        cand = filt[0]
                if cand is None:
                    st.info("No candidate available. Run the Machine Finder to populate the archive.")
                else:
                    # Tabs: Truth, Closure, Margins, Costs, Existence, Neighborhood
                    t_truth, t_close, t_marg, t_cost, t_exist, t_neigh = st.tabs([
                        "Truth",
                        "Closure",
                        "Margins",
                        "💰 Economics",
                        "Why it exists",
                        "Neighborhood",
                    ])
                    with t_truth:
                        st.markdown("#### Inputs")
                        st.json(cand.get("inputs") or {}, expanded=False)
                        st.markdown("#### Key outputs")
                        out = cand.get("outputs") or {}
                        # show a compact subset if present
                        keys = ["Pfus_total_MW","P_e_net_MW","Q_DT_eqv","q_div_MW_m2","min_signed_margin"]
                        st.write({k: out.get(k) for k in keys if k in out} or out)
                        st.markdown("#### Constraint ladder")
                        st.write({
                            "feasibility_state": cand.get("feasibility_state"),
                            "robustness_class": cand.get("robustness_class"),
                            "first_failure": cand.get("first_failure"),
                            "failure_mode": cand.get("failure_mode"),
                        })

                        st.markdown("#### Expert signals (descriptive)")
                        tags = _regime_signature(cand)
                        fk = _first_kill(cand)
                        st.write({
                            "regime_signature": tags,
                            "first_kill": fk,
                        })
                        sr = _constraint_spend_rate(cand)
                        if isinstance(sr, dict) and sr.get("ok"):
                            with st.expander("Constraint spend rate vs parent (heuristic)", expanded=False):
                                st.json(sr, expanded=False)
                        with st.expander("Constraint records", expanded=False):
                            st.json(cand.get("constraints") or [], expanded=False)

                    with t_close:
                        st.caption("Plant closure and accounting are computed explicitly (parity layer). They do not modify frozen truth.")
                        try:
                            from src.parity import parity_plant_closure, parity_magnets, parity_cryo
                            from src.models.inputs import PointInputs
                            pi = cand.get("_point_inputs_obj")
                            if pi is None:
                                pi = PointInputs(**(cand.get("inputs") or {}))
                            outputs = cand.get("outputs") or {}
                            plant = parity_plant_closure(pi, outputs)
                            magnets = parity_magnets(pi, outputs)
                            cryo = parity_cryo(pi, outputs)
                            c1,c2,c3 = st.columns(3)
                            c1.metric("Net electric (MW)", f"{plant.get('derived',{}).get('P_e_net_MW', float('nan')):.3g}")
                            c2.metric("Recirc electric (MW)", f"{plant.get('derived',{}).get('P_recirc_e_MW', float('nan')):.3g}")
                            c3.metric("Qe", f"{plant.get('derived',{}).get('Qe', float('nan')):.3g}")
                            with st.expander("Plant closure", expanded=False):
                                st.json(plant, expanded=False)
                            with st.expander("🧲 Magnets", expanded=False):
                                st.json(magnets, expanded=False)
                            with st.expander("Cryogenics", expanded=False):
                                st.json(cryo, expanded=False)
                        except Exception as e:
                            st.error(f"Closure console unavailable for this candidate: {e}")

                    with t_marg:
                        st.caption("Margin budget view: tight vs slack constraints (descriptive).")
                        rows=[]
                        for r in (cand.get("constraints") or []):
                            try:
                                rows.append({
                                    "name": r.get("name"),
                                    "ok": bool(r.get("ok")),
                                    "signed_margin": r.get("signed_margin"),
                                    "value": r.get("value"),
                                    "limit": r.get("limit"),
                                })
                            except Exception:
                                pass
                        if rows:
                            import pandas as _pd
                            df=_pd.DataFrame(rows)
                            df=df.sort_values(by="signed_margin", ascending=True, na_position="last")
                            st.dataframe(df, use_container_width=True, hide_index=True)
                        else:
                            st.info("No constraint margin records for this candidate.")

                    with t_cost:
                        st.caption("Economics envelopes are explicit lenses (optimistic/nominal/conservative).")
                        try:
                            from src.parity import parity_costing_envelope, parity_costing
                            from src.models.inputs import PointInputs
                            pi = cand.get("_point_inputs_obj")
                            if pi is None:
                                pi = PointInputs(**(cand.get("inputs") or {}))
                            outputs = cand.get("outputs") or {}
                            env = parity_costing_envelope(pi, outputs)
                            base = parity_costing(pi, outputs)
                            c1,c2,c3 = st.columns(3)
                            c1.metric("CAPEX (MUSD)", f"{base.get('derived',{}).get('CAPEX_MUSD', float('nan')):.3g}")
                            c2.metric("LCOE (USD/MWh)", f"{base.get('derived',{}).get('LCOE_USD_per_MWh', float('nan')):.3g}")
                            # show envelope headline
                            if isinstance(env, dict) and env.get("nominal"):
                                st.caption(
                                    f"Envelope LCOE - Opt {env.get('optimistic',{}).get('LCOE_USD_per_MWh', float('nan')):.3g} | "
                                    f"Nom {env.get('nominal',{}).get('LCOE_USD_per_MWh', float('nan')):.3g} | "
                                    f"Con {env.get('conservative',{}).get('LCOE_USD_per_MWh', float('nan')):.3g}"
                                )
                            with st.expander("Economics envelope", expanded=False):
                                st.json(env, expanded=False)
                            with st.expander("Base costing (proxy)", expanded=False):
                                st.json(base.get("raw", base), expanded=False)
                        except Exception as e:
                            st.error(f"Economics deck unavailable: {e}")

                    with t_exist:
                        try:
                            rep = existence_report(cand)
                            st.json(rep, expanded=False)
                        except Exception as e:
                            st.error(f"Existence report failed: {e}")

                    with t_neigh:
                        st.caption("Bridge to Scan Lab: open a local slice around this candidate (no auto-run).")
                        try:
                            if st.button("Set Scan Lab slice around candidate (Ip vs R0)", use_container_width=True, key="opt_set_scan_slice"):
                                inp = cand.get("inputs") or {}
                                st.session_state["scan_x"] = "R0_m"
                                st.session_state["scan_y"] = "Ip_MA"
                                R0=float(inp.get("R0_m", 1.0))
                                Ip=float(inp.get("Ip_MA", 1.0))
                                st.session_state["scan_bounds"] = {
                                    "R0_m": [max(0.2, 0.7*R0), 1.3*R0],
                                    "Ip_MA": [max(0.1, 0.7*Ip), 1.3*Ip],
                                }
                                st.success("Scan Lab slice parameters set in session state. Switch to Scan Lab to run.")
                        except Exception as e:
                            st.error(f"Could not set Scan Lab slice: {e}")

            elif view == "Expert Compare (no ranking)":
                st.caption("Compare a handful of candidates side-by-side. No ranking, no recommendation - just numbers and margins.")
                ar0 = filt or (run.get("archive") or [])
                if not ar0:
                    st.info("No archive available.")
                else:
                    max_n = min(12, len(ar0))
                    idxs = st.multiselect(
                        "Select candidate indices (from filtered archive order)",
                        options=list(range(len(ar0))),
                        default=list(range(min(3, len(ar0)))),
                        key="opt_compare_idxs",
                    )
                    rows=[]
                    for idx in idxs[:max_n]:
                        a=ar0[int(idx)]
                        out=a.get("outputs") or {}
                        inp=a.get("inputs") or {}
                        rows.append({
                            "idx": int(idx),
                            "feasibility_state": a.get("feasibility_state"),
                            "robustness": a.get("robustness_class"),
                            "first_failure": a.get("first_failure"),
                            "R0_m": inp.get("R0_m"),
                            "a_m": inp.get("a_m"),
                            "Bt_T": inp.get("Bt_T"),
                            "Ip_MA": inp.get("Ip_MA"),
                            "Pfus_total_MW": out.get("Pfus_total_MW"),
                            "P_e_net_MW": out.get("P_e_net_MW"),
                            "Q_DT_eqv": out.get("Q_DT_eqv"),
                            "q_div_MW_m2": out.get("q_div_MW_m2"),
                            "min_signed_margin": out.get("min_signed_margin", a.get("min_signed_margin")),
                        })
                    if rows:
                        import pandas as _pd
                        st.dataframe(_pd.DataFrame(rows), use_container_width=True, hide_index=True)
                    st.caption("Tip: use the Inspector index to sync with Machine Dossier.")

            elif view == "Casebook Runner":
                st.caption("Run a small set of declared cases (intent+lens+bounds) and compare feasibility distributions. No recommendations.")

                # Optional: load the packaged flagship casebook (review-room demo)
                if st.button("Load flagship casebook (packaged)", use_container_width=True, key="opt_load_flagship_casebook", disabled=forge_lock):
                    try:
                        from pathlib import Path as _P
                        _p = _P("scenarios") / "flagship_casebook.json"
                        if _p.exists():
                            st.session_state["opt_casebook"] = json.loads(_p.read_text(encoding="utf-8"))
                            st.success("Flagship casebook loaded.")
                        else:
                            st.warning("flagship_casebook.json not found in scenarios/.")
                    except Exception as _e:
                        st.warning(f"Could not load flagship casebook: {_e}")
                if "opt_casebook" not in st.session_state or not isinstance(st.session_state.get("opt_casebook"), list):
                    st.session_state["opt_casebook"] = []
                # Case definition UI
                c1,c2,c3 = st.columns([2,2,1])
                with c1:
                    case_name = st.text_input("Case name", value=f"Case {len(st.session_state['opt_casebook'])+1}", key="opt_case_name")
                with c2:
                    case_lens = st.selectbox("Lens", list((default_objective_packs(_design_intent_key()) or {}).keys()) or ["default"], key="opt_case_lens")
                with c3:
                    case_seed = st.number_input("Seed", value=int(run.get("seed", 0) or 0), step=1, key="opt_case_seed")
                if st.button("Add case to casebook", use_container_width=True, key="opt_add_case", disabled=forge_lock):
                    st.session_state["opt_casebook"].append({"name": case_name, "lens": case_lens, "seed": int(case_seed)})
                    st.success("Case added.")
                # Display current casebook
                if st.session_state["opt_casebook"]:
                    import pandas as _pd
                    st.dataframe(_pd.DataFrame(st.session_state["opt_casebook"]), use_container_width=True, hide_index=True)
                else:
                    st.info("Casebook is empty. Add 2–5 cases to run.")
                # Run cases (small budgets unless expert toggles)
                budget = int(st.number_input("Per-case evaluation budget", value=120, min_value=20, max_value=5000, step=20, key="opt_case_budget"))
                if st.button("Run casebook", use_container_width=True, key="opt_run_casebook", disabled=forge_lock):
                    results=[]
                    for case in (st.session_state["opt_casebook"] or [])[:10]:
                        try:
                            # Reuse current bounds and var specs from session if present
                            _packs = default_objective_packs(_design_intent_key()) or {}
                            pack = _packs.get(case["lens"]) or next(iter(_packs.values())) if _packs else {}
                            # minimal var specs: use session var specs if present
                            var_specs = run.get("var_specs") or st.session_state.get("opt_var_specs") or []
                            bounds = st.session_state.get("opt_bounds") or {}
                            rcase = run_hybrid_machine_finder(
                                seed=int(case.get("seed",0)),
                                intent=_design_intent_key(),
                                objective_pack=pack,
                                bounds=bounds,
                                var_specs=var_specs,
                                budget=int(budget),
                                cache_enabled=bool(st.session_state.get("opt_cache_enabled", True)),
                                cache_max=int(st.session_state.get("opt_cache_max", 256)),
                            )
                            tr = rcase.get("trace") or []
                            n=len(tr); nf=sum(1 for t in tr if bool(t.get("feasible")))
                            results.append({"case": case["name"], "lens": case["lens"], "seed": case["seed"], "n_eval": n, "n_feasible": nf})
                        except Exception as e:
                            results.append({"case": case["name"], "lens": case["lens"], "seed": case["seed"], "n_eval": 0, "n_feasible": 0, "error": str(e)})
                    st.session_state["opt_casebook_results"] = results
                res = st.session_state.get("opt_casebook_results") or []
                if res:
                    import pandas as _pd
                    st.markdown("#### Casebook results (summary)")
                    st.dataframe(_pd.DataFrame(res), use_container_width=True, hide_index=True)
        

            elif view == "Archive Bay":
                # pick axes
                xkey = st.selectbox("x-axis", ["R0_m","Bt_T","Ip_MA","P_e_net_MW","Pfus_total_MW","q_div_MW_m2","min_signed_margin","_score"], index=0, key="opt_scatter_x")
                ykey = st.selectbox("y-axis", ["P_e_net_MW","Pfus_total_MW","Q_DT_eqv","q_div_MW_m2","min_signed_margin","_score"], index=0, key="opt_scatter_y")
                rows=[]
                for a in filt:
                    o=a.get("outputs") or {}
                    i=a.get("inputs") or {}
                    def get(k):
                        if k in ("_score",):
                            return float(a.get("_score",-1e30))
                        if k in i:
                            return float(i.get(k, float("nan")))
                        return float(o.get(k, float("nan")))
                    rows.append({"x": get(xkey), "y": get(ykey), "feasible": bool(a.get("feasible", False)), "idx": len(rows)})
                if rows:
                    import pandas as _pd
                    df=_pd.DataFrame(rows)
                    st.scatter_chart(df, x="x", y="y")
                    st.caption("Use the Inspector to select a candidate by index from the filtered archive.")
                else:
                    st.info("No points after filtering. Relax filters or rerun with wider bounds/budget.")
            elif view == "Reactor Accounting Console":
                st.caption(
                    "Optional PROCESS-parity accounting: plant closure + magnets + cryo + costing. "
                    "This is a *lens* and does not change frozen evaluator truth."
                )
                # choose candidate
                cand = None
                if filt:
                    cand = filt[int(min(len(filt)-1, int(st.session_state.get("opt_inspect_idx", 0))))]
                elif isinstance(run.get("best_feasible"), dict):
                    cand = run.get("best_feasible")
                if cand is None:
                    st.info("No candidate available. Run the Machine Finder to populate the archive.")
                else:
                    try:
                        from src.parity import parity_plant_closure, parity_magnets, parity_cryo, parity_costing, parity_costing_envelope
                        from src.parity.calibration import economics_local_sensitivity
                        from src.parity.report_pack import build_parity_report_pack, report_pack_to_csv_rows, report_pack_to_markdown
                        pi = cand.get("_point_inputs_obj")
                        if pi is None:
                            # reconstruct PointInputs if present
                            from src.models.inputs import PointInputs
                            pi = PointInputs(**(cand.get("inputs") or {}))
                        outputs = cand.get("outputs") or {}
                        parity = {
                            "plant": parity_plant_closure(pi, outputs),
                            "magnets": parity_magnets(pi, outputs),
                            "cryo": parity_cryo(pi, outputs),
                            "costing": parity_costing(pi, outputs),
                            "costing_envelope": parity_costing_envelope(pi, outputs),
                        }
                        # summary cards
                        c1,c2,c3,c4 = st.columns(4)
                        c1.metric("Net electric (MW)", f"{parity['plant']['derived'].get('P_e_net_MW', float('nan')):.3g}")
                        c2.metric("Qe", f"{parity['plant']['derived'].get('Qe', float('nan')):.3g}")
                        c3.metric("CAPEX (MUSD)", f"{parity['costing']['derived'].get('CAPEX_MUSD', float('nan')):.3g}")
                        c4.metric("LCOE (USD/MWh)", f"{parity['costing']['derived'].get('LCOE_USD_per_MWh', float('nan')):.3g}")
                        # Cost envelope (Phase-2)
                        env = parity.get('costing_envelope', {})
                        posture = st.session_state.get('ppl_cost_posture', 'Nominal')
                        if isinstance(env, dict) and env.get('nominal'):
                            nom = env.get('nominal', {})
                            opt = env.get('optimistic', {})
                            con = env.get('conservative', {})
                            st.caption(
                                f"Economics envelope - Optimistic {opt.get('LCOE_USD_per_MWh', float('nan')):.3g} | "
                                f"Nominal {nom.get('LCOE_USD_per_MWh', float('nan')):.3g} | "
                                f"Conservative {con.get('LCOE_USD_per_MWh', float('nan')):.3g}  (posture: {posture})"
                            )
                        with st.expander("Plant closure", expanded=False):
                            st.json(parity["plant"], expanded=False)
                        with st.expander("🧲 Magnets", expanded=False):
                            st.json(parity["magnets"], expanded=False)
                        with st.expander("Cryogenics", expanded=False):
                            st.json(parity["cryo"], expanded=False)
                        with st.expander("Costing (proxy)", expanded=False):
                            st.json(parity["costing"].get("raw", parity["costing"]), expanded=False)

                        # --- v2 additions: CAPEX breakdown + local sensitivities ---
                        st.markdown("#### 💰 Economics breakdown")
                        bd = (parity.get("costing") or {}).get("derived", {}).get("breakdown_MUSD", {}) or {}
                        if bd:
                            import pandas as _pd
                            df = _pd.DataFrame({"component": list(bd.keys()), "CAPEX_MUSD": [float(bd[k]) for k in bd.keys()]})
                            st.bar_chart(df, x="component", y="CAPEX_MUSD")
                            st.caption("CAPEX proxy breakdown (component bars). Total CAPEX is the sum of these components.")
                        else:
                            st.info("No CAPEX breakdown available for this candidate.")

                        with st.expander("Local sensitivity (economics knobs)", expanded=False):
                            sens = economics_local_sensitivity(inputs=pi, outputs=outputs, perturb_frac=0.10)
                            st.json(sens.get("base", {}), expanded=False)
                            rows = sens.get("rows") or []
                            if rows:
                                import pandas as _pd
                                st.dataframe(_pd.DataFrame(rows), use_container_width=True)
                            st.caption("Finite-difference lens (+10% knob perturbation). Not a gradient; used for intuition.")

                        import json as _json
                        pack = build_parity_report_pack(
                            inputs=pi,
                            outputs=outputs,
                            parity=parity,
                            run_id=str(run.get("run_id", "")),
                            version=str(run.get("version", "")),
                        )
                        md = report_pack_to_markdown(pack)
                        header, row = report_pack_to_csv_rows(pack)
                        csv = ",".join(header) + "\n" + ",".join([str(x) for x in row]) + "\n"
                        st.download_button(
                            "Download parity report (JSON)",
                            data=_json.dumps(pack, indent=2).encode("utf-8"),
                            file_name="shams_parity_report_pack.json",
                            mime="application/json",
                            use_container_width=True,
                        )
                        st.download_button(
                            "Download parity report (Markdown)",
                            data=md.encode("utf-8"),
                            file_name="shams_parity_report_pack.md",
                            mime="text/markdown",
                            use_container_width=True,
                        )
                        st.download_button(
                            "Download parity flat row (CSV)",
                            data=csv.encode("utf-8"),
                            file_name="shams_parity_report_pack.csv",
                            mime="text/csv",
                            use_container_width=True,
                        )
                    except Exception as _e:
                        st.error(f"Parity layer failed: {_e}")
            elif view == "Parity validation packs (PASS/WARN/FAIL)":

                st.caption("Named validation packs with explicit tolerances. Load your own reference values to certify parity.")
                from pathlib import Path
                from src.parity.validation_packs import load_validation_packs, evaluate_pack_candidate, compare_to_reference

                packs_path = Path("benchmarks/ppl_validation_packs_v3.json")
                refs_path = Path("benchmarks/ppl_validation_refs_v3.json")
                packs = load_validation_packs(packs_path)

                st.markdown("#### Validation pack selection")
                pack_titles = [f"{p.title}  -  ({p.pack_id})" for p in packs]
                sel = st.selectbox("Pack", pack_titles, index=0, key="ppl_pack_sel")
                pack = packs[pack_titles.index(sel)]

                st.markdown("#### Reference table")
                use_builtin = st.checkbox("Use built-in reference (placeholder)", value=True, key="ppl_use_builtin_refs")
                ref_dict = {}
                if use_builtin and refs_path.exists():
                    try:
                        ref_dict = __import__("json").loads(refs_path.read_text(encoding="utf-8")).get("refs", {}).get(pack.pack_id, {})
                    except Exception as _e:
                        st.warning(f"Could not read built-in refs: {_e}")

                up = st.file_uploader("Or upload reference JSON (expects {'refs': {pack_id: {metric_key: value}}})", type=["json"], key="ppl_ref_upload")
                if up is not None:
                    try:
                        payload = __import__("json").loads(up.read().decode("utf-8"))
                        ref_dict = payload.get("refs", {}).get(pack.pack_id, {})
                        st.success("Loaded uploaded reference table.")
                    except Exception as _e:
                        st.error(f"Could not parse reference JSON: {_e}")

                tol_scale = st.slider("Tolerance scale (multiplies per-metric tolerances)", 0.25, 3.0, 1.0, 0.05, key="ppl_tol_scale")

                if st.button("Run validation pack", key="ppl_run_pack"):
                    try:
                        preset, outputs, metrics, meta = evaluate_pack_candidate(pack)
                        # apply tol scaling on-the-fly
                        scaled_pack = type(pack)(
                            pack_id=pack.pack_id,
                            title=pack.title,
                            preset_key=pack.preset_key,
                            design_intent=pack.design_intent,
                            compare_keys=pack.compare_keys,
                            tolerances_rel={k: float(v) * float(tol_scale) for k, v in pack.tolerances_rel.items()},
                            severities=pack.severities,
                        )
                        res = compare_to_reference(pack=scaled_pack, metrics=metrics, reference=ref_dict or {})
                        st.session_state["ppl_last_validation"] = {"pack": pack.pack_id, "res": res, "meta": meta, "metrics": metrics, "reference": ref_dict}
                    except Exception as _e:
                        st.error(f"Validation run failed: {_e}")

                last = st.session_state.get("ppl_last_validation")
                if last and last.get("pack") == pack.pack_id:
                    res = last["res"]
                    st.markdown("#### Verdict")
                    if res["status"] == "PASS":
                        st.success(f"PASS - worst relative error: {res['worst_rel_err']:.3f}")
                    elif res["status"] == "WARN":
                        st.warning(f"WARN - worst relative error: {res['worst_rel_err']:.3f}")
                    else:
                        st.error(f"FAIL - worst relative error: {res['worst_rel_err']:.3f}")

                    st.markdown("#### Deltas")
                    rows = res["rows"]
                    # render as a small table
                    import pandas as _pd
                    df = _pd.DataFrame(rows)
                    st.dataframe(df, use_container_width=True, hide_index=True)

                    with st.expander("Assumptions & constraints context", expanded=False):
                        st.json(last.get("meta", {}))
                        st.json({"metrics": last.get("metrics", {}), "reference": last.get("reference", {})})


            elif view == "Parity calibration (reference deltas)":
                st.caption(
                    "Compare SHAMS Parity outputs against a reference table (e.g., published study values). "
                    "Built-in references are placeholders; upload your program's reference JSON to calibrate."
                )
                up = st.file_uploader("Optional reference JSON (same schema as benchmarks/parity_v2_refs.json)", type=["json"], key="opt_parity_ref_upload")
                ref_path = None
                if up is not None:
                    try:
                        tmp = Path(ROOT) / ".tmp_parity_refs.json"
                        tmp.write_bytes(up.getvalue())
                        ref_path = str(tmp)
                    except Exception:
                        ref_path = None
                if st.button("Run calibration", use_container_width=True, key="opt_parity_calib_run"):
                    try:
                        from tools.parity_calibrate import run_parity_calibration
                        st.session_state["opt_parity_calib_res"] = run_parity_calibration(refs_path=ref_path)
                    except Exception as _e:
                        st.session_state["opt_parity_calib_res"] = {"ok": False, "reason": str(_e)}
                res = st.session_state.get("opt_parity_calib_res")
                if res:
                    st.json({k: res.get(k) for k in ["ok", "n_cases", "refs_path", "note"]}, expanded=False)
                    for case in res.get("results", [])[:50]:
                        nm = case.get("name")
                        ok = bool(case.get("ok"))
                        with st.expander(f"{nm} - {'PASS' if ok else 'CHECK'}", expanded=not ok):
                            st.caption(str(case.get("notes") or ""))
                            rows = case.get("rows") or []
                            if rows:
                                import pandas as _pd
                                st.dataframe(_pd.DataFrame(rows), use_container_width=True)
                            else:
                                st.info("No reference metrics defined for this case.")
            elif view == "PROCESS parity benchmarks":
                st.caption("Run the internal parity regression suite (local, deterministic).")
                upd = st.checkbox("Update golden (developer)", value=False, key="opt_parity_update_golden")
                if st.button("Run parity benchmarks", key="opt_parity_bench", use_container_width=True):
                    try:
                        from tools.parity_bench import run_parity_benchmarks
                        st.session_state["opt_parity_bench_res"] = run_parity_benchmarks(update_golden=bool(upd))
                    except Exception as _e:
                        st.session_state["opt_parity_bench_res"] = {"ok": False, "reason": str(_e)}
                if st.session_state.get("opt_parity_bench_res"):
                    st.json(st.session_state.get("opt_parity_bench_res"), expanded=False)
            
            elif view == "Decision scenarios (program lens)":
                st.caption(
                    "Scenario presets that bundle objective intent, economics conservatism, and credibility lens. "
                    "They do not change frozen truth; they set *study posture*."
                )
                scenarios = {
                    "Conservative engineering": {
            "opt_intent": "Reactor",
            "opt_pack_choice": "Compact feasible reactor",
            "ppl_cost_posture": "Conservative",
            "credibility": "Conservative",
            "note": "Bias toward margin and conservative economics; useful for program-safe screening.",
                    },
                    "Nominal program baseline": {
            "opt_intent": "Reactor",
            "opt_pack_choice": "Compact feasible reactor",
            "ppl_cost_posture": "Nominal",
            "credibility": "Nominal",
            "note": "Default posture for comparisons and internal benchmarking.",
                    },
                    "Aggressive HTS / high-field": {
            "opt_intent": "Reactor",
            "opt_pack_choice": "High-field HTS stress frontier",
            "ppl_cost_posture": "Optimistic",
            "credibility": "Aggressive",
            "note": "Exploration posture; expect fragility and tighter trust boundaries.",
                    },
                    "Research toward Reactor": {
            "opt_intent": "Research",
            "opt_pack_choice": "Closest-to-reactor feasibility",
            "ppl_cost_posture": "Nominal",
            "credibility": "Nominal",
            "note": "Use Research intent but track distance-to-reactor; strategic R&D planning.",
                    },
                }

                names = list(scenarios.keys())
                sel = st.selectbox("Scenario", names, index=1, key="ppl_scenario_sel")
                s = scenarios[sel]
                st.info(s["note"], icon="🧭")

                colA, colB = st.columns([1, 1])
                with colA:
                    st.markdown("**Scenario settings**")
                    st.json({k: v for k, v in s.items() if k != "note"})
                with colB:
                    st.markdown("**Apply**")
                    if st.button("Apply scenario to session", key="ppl_apply_scenario"):
                        # apply to optimization setup controls
                        st.session_state["opt_intent"] = s["opt_intent"]
                        st.session_state["opt_pack_choice"] = s["opt_pack_choice"]
                        st.session_state["ppl_cost_posture"] = s["ppl_cost_posture"]
                        st.session_state["ppl_credibility"] = s["credibility"]
                        st.success("Scenario applied. Return to Setup and run Machine Finder, or inspect Parity Workbench.")
                    st.markdown("**Economics posture**")
                    posture = st.selectbox(
            "Cost envelope posture",
            ["Optimistic", "Nominal", "Conservative"],
            index=["Optimistic","Nominal","Conservative"].index(st.session_state.get("ppl_cost_posture","Nominal")),
            key="ppl_cost_posture",
                    )
                    st.caption("Used for economics envelope display and scenario labeling (nominal truth is unchanged).")

            elif view == "Trace Telemetry":
                tr = run.get("trace") or []
                if tr:
                    import pandas as _pd
                    df=_pd.DataFrame(tr)
                    st.line_chart(df[["score"]])
                    st.caption("Trace shows score progression; feasibility and resistance are summarized in the atlas.")
                else:
                    st.info("No trace recorded.")
            elif view == "Resistance Brief":
                st.json(run.get("resistance", {}))
                st.json(run.get("variable_correlations", {}))
            elif view == "Boundary Navigator":
                st.caption(
                    "Local linear surface model for a single constraint, fitted from near-boundary points in the archive. "
                    "This is an instrument to understand geometry; it does not recommend designs."
                )
                vs = run.get("var_specs") or []
                vkeys = []
                for v in vs:
                    if isinstance(v, dict) and v.get("key"):
                        vkeys.append(str(v.get("key")))
                    else:
                        try:
                            vkeys.append(str(getattr(v, "key")))
                        except Exception:
                            pass
                # available constraints
                names = []
                sample = None
                for a in (run.get("archive") or []):
                    if (a.get("constraints") or []):
                        sample = a
                        break
                if sample is not None:
                    for c in (sample.get("constraints") or []):
                        nm = str(c.get("name"))
                        if nm and nm not in names:
                            names.append(nm)
                if not vkeys:
                    st.info("No optimized variables found in this run.")
                elif not names:
                    st.info("No constraint records found in archive candidates.")
                else:
                    cn = st.selectbox("Constraint", options=names, index=0, key="opt_surface_constraint")
                    use_archive = st.checkbox("Use filtered archive", value=True, key="opt_surface_use_filtered")
                    data_src = filt if (use_archive and isinstance(filt, list) and filt) else (run.get("archive") or [])
                    m = constraint_surface_map(archive=data_src, var_keys=vkeys, constraint_name=cn)
                    st.json(m)
            elif view == "Feasibility skeleton":
                sk = run.get("feasibility_skeleton") or {}
                if not sk:
                    st.info("Skeleton not available (need feasible points).")
                else:
                    st.metric("Feasible points", str(sk.get("n_feasible", "-")))
                    st.metric("Components", str(sk.get("n_components", "-")))
                    st.write({"component_sizes": sk.get("components", [])})
                    with st.expander("Bottleneck edges (longest kNN edges)", expanded=False):
                        st.write(sk.get("bottleneck_edges", []))
                    st.caption("Use this to see whether feasible truth is one connected basin or multiple islands.")
            elif view == "Local cartography (adaptive)":
                st.caption("A small 2D scan around the selected/ best candidate. This is cartography, not optimization.")
                base = None
                if isinstance(run.get("best_feasible"), dict):
                    base = run.get("best_feasible")
                if filt:
                    base = filt[int(min(len(filt)-1, int(st.session_state.get("opt_inspect_idx", 0))))]
                if base is None:
                    st.info("No candidate to center local cartography on.")
                else:
                    vs = run.get("var_specs") or []
                    vkeys = [v.get("key") for v in vs if isinstance(v, dict) and v.get("key")]
                    if len(vkeys) < 2:
                        st.info("Need at least two optimized variables.")
                    else:
                        xk = st.selectbox("x variable", vkeys, index=0, key="opt_localcart_x")
                        yk = st.selectbox("y variable", vkeys, index=min(1, len(vkeys)-1), key="opt_localcart_y")
                        span = st.slider("Span (±% of local bounds)", 5, 60, 20, key="opt_localcart_span")
                        ngrid = st.slider("Grid", 9, 41, 21, step=2, key="opt_localcart_ng")
                        if st.button("Run local cartography", key="opt_localcart_run"):
                            import numpy as _np
                            import pandas as _pd
                            bx = (base.get("inputs") or {}).get(xk)
                            by = (base.get("inputs") or {}).get(yk)
                            # bounds
                            bmap = {v.get("key"):(float(v.get("lo")), float(v.get("hi"))) for v in vs if isinstance(v, dict) and v.get("key")}
                            xlo,xhi = bmap.get(xk,(float(bx)*0.8,float(bx)*1.2))
                            ylo,yhi = bmap.get(yk,(float(by)*0.8,float(by)*1.2))
                            xmid = float(bx) if bx is not None else 0.5*(xlo+xhi)
                            ymid = float(by) if by is not None else 0.5*(ylo+yhi)
                            dx = (xhi-xlo)*float(span)/100.0
                            dy = (yhi-ylo)*float(span)/100.0
                            xs = _np.linspace(max(xlo, xmid-dx), min(xhi, xmid+dx), int(ngrid))
                            ys = _np.linspace(max(ylo, ymid-dy), min(yhi, ymid+dy), int(ngrid))
                            rows=[]
                            for xv in xs:
                                for yv in ys:
                                    cand_in = dict(base.get("inputs") or {})
                                    cand_in[xk] = float(xv)
                                    cand_in[yk] = float(yv)
                                    r = eval_fn(cand_in)
                                    rows.append({"x":float(xv),"y":float(yv),"feasible":bool(r.get("feasible",False)),"score":float(r.get("_score",-1e30)),"violation":float(r.get("_violation",1e30)),"min_margin":float(r.get("min_signed_margin",float("nan")))})
                            df=_pd.DataFrame(rows)
                            st.session_state["opt_localcart_df"] = df
                        df = st.session_state.get("opt_localcart_df")
                        if df is not None:
                            try:
                                import numpy as _np
                                import matplotlib.pyplot as _plt
                                dff = df.pivot(index="y", columns="x", values="feasible")
                                fig = _plt.figure()
                                _plt.imshow(dff.values[::-1, :], aspect="auto")
                                _plt.title("Feasibility map (local)")
                                st.pyplot(fig)
                                st.caption("Heatmap shows feasible (1) vs infeasible (0). Use Inspector for details.")
                            except Exception:
                                st.dataframe(df)
            elif view == "Uncertainty (Monte Carlo)":
                st.caption("Monte Carlo robustness is optional. It does **not** change feasibility truth; it samples around a candidate.")
                if not filt:
                    st.info("No candidates available.")
                else:
                    cand = filt[int(min(len(filt)-1, int(st.session_state.get("opt_inspect_idx", 0))))]
                    ns = st.number_input("Samples", 20, 2000, 200, step=20, key="opt_uq_ns")
                    pct = st.slider("Perturbation (±% on optimized vars)", 1, 25, 5, key="opt_uq_pct")
                    if st.button("Run robustness Monte Carlo", key="opt_uq_run"):
                        import numpy as _np
                        rng = _np.random.default_rng(1)
                        vs = run.get("var_specs") or []
                        vkeys = [v.get("key") for v in vs if isinstance(v, dict) and v.get("key")]
                        base_in = dict(cand.get("inputs") or {})
                        feas=0
                        scores=[]
                        for _ in range(int(ns)):
                            ci = dict(base_in)
                            for k in vkeys:
                                v0 = float(base_in.get(k, 0.0))
                                dv = v0 * float(pct)/100.0
                                ci[k] = float(rng.uniform(v0-dv, v0+dv))
                            r = eval_fn(ci)
                            if r.get("feasible", False):
                                feas += 1
                                scores.append(float(r.get("_score", -1e30)))
                        st.session_state["opt_uq_res"] = {
                            "samples": int(ns),
                            "pct": float(pct),
                            "feasible_rate": float(feas)/float(ns),
                            "mean_score_feasible": float(_np.mean(scores)) if scores else None,
                            "n_feasible": int(feas),
                        }
                    if st.session_state.get("opt_uq_res"):
                        st.json(st.session_state.get("opt_uq_res"))
            elif view == "Intent trajectories (Research→Reactor)":
                st.caption(
                    "Tier-5: a simple *intent trajectory* instrument. "
                    "It tries to build a Research→Reactor 'highway' using the current archive and the currently selected variables. "
                    "This does not optimize; it organizes what you already found."
                )
                traj = st.session_state.get("opt_traj")
                if not traj:
                    st.info("Click **Build trajectory** in the left Tier 5–6 expander to compute a path.")
                else:
                    if not traj.get("ok", False):
                        st.warning(f"Trajectory unavailable: {traj.get('reason')}")
                        st.json(traj, expanded=False)
                    else:
                        st.success(f"Trajectory built: {traj.get('from_intent')} → {traj.get('to_intent')} (steps={len(traj.get('nodes') or [])})")
                        nodes = traj.get("nodes") or []
                        edges = traj.get("edges") or []
                        import pandas as _pd
                        rows = []
                        for i, n in enumerate(nodes):
                            rows.append({
                                "step": i,
                                "from_feasible": bool(n.get("from_feasible")),
                                "to_feasible": bool(n.get("to_feasible")),
                                "from_score": float(n.get("from_score", -1e30)),
                                "to_score": float(n.get("to_score", -1e30)),
                            })
                        st.dataframe(_pd.DataFrame(rows), use_container_width=True)
                        with st.expander("Step inputs", expanded=False):
                            for i, n in enumerate(nodes):
                                st.markdown(f"**Step {i} inputs**")
                                st.json(n.get("inputs") or {}, expanded=False)
                                if i < len(edges):
                                    st.caption(f"Δ to next (L2 dist ≈ {float(edges[i].get('dist', 0.0)):.4g})")
                                    st.json(edges[i].get("delta") or {}, expanded=False)
                        import json as _json
                        st.download_button(
                            "Download trajectory (json)",
                            data=_json.dumps(traj, indent=2).encode("utf-8"),
                            file_name="shams_intent_trajectory.json",
                            mime="application/json",
                            use_container_width=True,
                        )

            elif view == "Inverse design / Why not?":
                st.caption(
                    "Tier-5: inverse design (closest feasible to a target) + 'why not' explanation. "
                    "This never relaxes constraints; it searches only within your declared bounds."
                )
                targets = st.session_state.get("opt_inv_targets") or {}
                st.markdown("**Targets**")
                st.json(targets, expanded=False)

                n_samples = st.number_input("Inverse search samples", 50, 5000, 600, step=50, key="opt_inv_ns")
                if st.button("Run inverse search (closest feasible)", key="opt_inv_run", use_container_width=True):
                    # Sample uniformly in declared bounds and pick feasible with min residual.
                    import numpy as _np
                    rng = _np.random.default_rng(int(run.get("seed", 0)) + 17)
                    vs = run.get("var_specs") or []
                    vkeys = [v.get("key") for v in vs if isinstance(v, dict) and v.get("key")]
                    # If var_specs not present, fall back to current var_keys + bounds
                    if not vkeys:
                        vkeys = list(bounds.keys())
                    best_res = None
                    best_eval = None
                    best_inputs = None
                    for _ in range(int(n_samples)):
                        cand_in = dict(anchor_default)
                        for k in vkeys:
                            lo, hi = bounds.get(k, (float(cand_in.get(k, 0.0)), float(cand_in.get(k, 0.0))))
                            cand_in[k] = float(rng.uniform(float(lo), float(hi)))
                        r = _evaluate_candidate(cand_in, intent)
                        if not r.get("feasible", False):
                            continue
                        resid = inverse_design_residual(r.get("outputs") or {}, targets)
                        if best_res is None or resid < best_res:
                            best_res = float(resid)
                            best_eval = r
                            best_inputs = cand_in
                    st.session_state["opt_inv_best"] = {"residual": best_res, "eval": best_eval, "inputs": best_inputs}

                inv_best = st.session_state.get("opt_inv_best")
                if inv_best and inv_best.get("eval"):
                    st.success(f"Best feasible inverse match residual: {float(inv_best.get('residual')):.4g}")
                    st.json(inv_best.get("inputs") or {}, expanded=False)
                    st.json((inv_best.get("eval") or {}).get("outputs") or {}, expanded=False)
                    with st.expander("Why-not style report (for this candidate)", expanded=False):
                        st.json(why_not_report(eval_res=inv_best.get("eval") or {}, disabled_constraints=st.session_state.get("opt_cf_disable") or []), expanded=False)
                else:
                    st.info("Run inverse search to find the closest feasible candidate to your targets.")

                # Why-not for selected candidate in inspector
                st.divider()
                st.markdown("**Why not for the currently selected candidate (Inspector index)**")
                if filt:
                    cand = filt[int(min(len(filt)-1, int(st.session_state.get("opt_inspect_idx", 0))))]
                    wr = why_not_report(
                        eval_res=cand,
                        disabled_constraints=st.session_state.get("opt_cf_disable") or [],
                    )
                    st.json(wr, expanded=False)
                else:
                    st.info("No candidate selected.")

            elif view == "Discovered relations (laws)":
                st.caption(
                    "Tier-6: mine simple, explainable relations from the feasible archive. "
                    "This is not a physics claim; it's a data-derived hint to guide exploration."
                )
                import pandas as _pd
                x_opts = ["R0_m","a_m","Bt_T","Ip_MA","Paux_MW","kappa","delta","nbar_1e20_m3","Ti_keV"]
                y_opts = ["P_e_net_MW","Pfus_total_MW","Q_DT_eqv","q_div_MW_m2","min_signed_margin"]
                x_sel = st.multiselect("x variables", x_opts, default=["R0_m","Bt_T","Ip_MA"], key="opt_rel_x")
                y_sel = st.multiselect("y metrics", y_opts, default=["P_e_net_MW","q_div_MW_m2"], key="opt_rel_y")
                feas_only = st.checkbox("Use feasible candidates only", value=True, key="opt_rel_feas")
                if st.button("Compute discovered relations", key="opt_rel_run", use_container_width=True):
                    st.session_state["opt_rel"] = discovered_relations(
                        candidates=archive,
                        x_keys=x_sel,
                        y_keys=y_sel,
                        feasible_only=bool(feas_only),
                        top_k=8,
                    )
                rel = st.session_state.get("opt_rel")
                if rel and rel.get("ok"):
                    st.success(f"Computed relations from n={rel.get('n')} candidates")
                    st.markdown("**Top linear fits**")
                    st.write(rel.get("top_linear_fits"))
                    st.markdown("**Top correlations**")
                    st.write(rel.get("top_corrs"))
                    md = export_relations_markdown(rel)
                    st.download_button(
                        "Download relations report (markdown)",
                        data=md.encode("utf-8"),
                        file_name="shams_discovered_relations.md",
                        mime="text/markdown",
                        use_container_width=True,
                    )
                else:
                    st.info("Compute relations to see the strongest correlations and simple fits.")

            elif view == "Counterfactual lens":
                st.caption(
                    "Tier-6: counterfactual lens. You can disable one or more constraints in the **feasibility gate** only. "
                    "Raw constraints stay unchanged; this is a hypothetical planning tool."
                )
                disabled = st.session_state.get("opt_cf_disable") or []
                if not disabled:
                    st.info("Select constraints to disable in the left Tier 5–6 expander.")
                else:
                    st.warning(f"Counterfactual disabled constraints: {disabled}")
                    cf_feas = 0
                    cf_best = None
                    cf_best_score = -1e30
                    for c in filt:
                        cf = counterfactual_gate(c, disabled).get("counterfactual") or {}
                        if cf.get("feasible", False):
                            cf_feas += 1
                            sc = float(c.get("_score", -1e30))
                            if sc > cf_best_score:
                                cf_best_score = sc
                                cf_best = c
                    st.metric("Counterfactual feasible (filtered)", f"{cf_feas} / {len(filt)}")
                    if cf_best is not None:
                        st.markdown("**Best (by score) among counterfactual-feasible**")
                        st.json(cf_best.get("inputs") or {}, expanded=False)
                        st.json((cf_best.get("outputs") or {}), expanded=False)
                        with st.expander("Counterfactual gate details", expanded=False):
                            st.json(counterfactual_gate(cf_best, disabled).get("counterfactual"), expanded=False)
                    st.divider()
                    st.markdown("**Inspector candidate under counterfactual**")
                    if filt:
                        cand = filt[int(min(len(filt)-1, int(st.session_state.get("opt_inspect_idx", 0))))]
                        st.json(counterfactual_gate(cand, disabled).get("counterfactual"), expanded=False)

            elif view == "Collaboration (review sessions)":
                st.caption(
                    "Tier-7: multi-user deliberation without external services. "
                    "Create a review session to attach comments/votes/tags to candidates, then export a session bundle."
                )
                _repo_root = Path(__file__).resolve().parent.parent
                _eval_fp = repo_fingerprint(_repo_root)
                sessions_dir = default_sessions_dir()
                st.markdown(f"**Local sessions dir:** `{sessions_dir}`")

                colA, colB = st.columns(2)
                with colA:
                    title = st.text_input("New session title", value=st.session_state.get("opt_review_title", "Optimization review"), key="opt_review_title")
                    notes = st.text_area("Session notes (optional)", value=st.session_state.get("opt_review_notes", ""), key="opt_review_notes")
                    if st.button("Create new session", key="opt_review_create", use_container_width=True):
                        sess = new_review_session(title=title, evaluator_fp=_eval_fp, intent=run.get("intent", ""), notes=notes)
                        path = sessions_dir / f"{sess.session_id}.json"
                        save_review_session(sess, path)
                        st.session_state["opt_review_path"] = str(path)
                        st.success(f"Created session: {sess.session_id}")

                with colB:
                    existing = sorted([p.name for p in sessions_dir.glob("*.json")])
                    pick = st.selectbox("Load existing session", options=[""] + existing, index=0, key="opt_review_pick")
                    if pick:
                        path = sessions_dir / pick
                        try:
                            sess = load_review_session(path)
                            st.session_state["opt_review_path"] = str(path)
                            st.success(f"Loaded session: {sess.session_id}")
                        except Exception as e:
                            st.error(f"Failed to load session: {e}")

                # Load active session
                sess = None
                sp = st.session_state.get("opt_review_path")
                if sp:
                    try:
                        sess = load_review_session(Path(sp))
                    except Exception:
                        sess = None

                if sess is None:
                    st.info("Create or load a review session to start commenting/voting.")
                else:
                    st.markdown("### Session")
                    st.json({
                        "session_id": sess.session_id,
                        "title": sess.title,
                        "created_at": sess.created_at,
                        "intent": sess.intent,
                        "evaluator_fp": sess.evaluator_fp[:12],
                        "n_candidates": len(sess.candidates or []),
                        "n_comments": len(sess.comments or []),
                        "n_votes": len(sess.votes or []),
                    }, expanded=False)

                    # Add current inspector candidate
                    if filt:
                        idx = int(st.session_state.get("opt_inspect_idx", 0))
                        idx = int(max(0, min(len(filt)-1, idx)))
                        cand = filt[idx]
                        if st.button("Add current inspector candidate to session", key="opt_review_add", use_container_width=True):
                            _c_fp = candidate_fingerprint(cand.get("inputs", {}) or {}, intent=run.get("intent", ""), evaluator_fp=_eval_fp)
                            # de-dup
                            if not any((c.get("candidate_fp") == _c_fp) for c in (sess.candidates or [])):
                                sess.candidates.append({
                                    "candidate_fp": _c_fp,
                                    "score": cand.get("_score"),
                                    "feasible": cand.get("feasible"),
                                    "min_signed_margin": cand.get("min_signed_margin"),
                                    "inputs": cand.get("inputs", {}),
                                    "failure_mode": cand.get("failure_mode"),
                                })
                                save_review_session(sess, Path(sp))
                                st.success("Added.")
                            else:
                                st.info("Candidate already in session.")

                    st.divider()
                    st.markdown("### Comment / vote")
                    cand_opts = [c.get("candidate_fp", "")[:12] + "…" for c in (sess.candidates or [])]
                    if not cand_opts:
                        st.info("Add candidates to the session to enable comments and votes.")
                    else:
                        sel = st.selectbox("Candidate", options=list(range(len(cand_opts))), format_func=lambda i: cand_opts[i], key="opt_review_cand_sel")
                        comment = st.text_area("Comment", key="opt_review_comment")
                        vote = st.slider("Vote (1–5)", 1, 5, 3, key="opt_review_vote")
                        tag = st.text_input("Tag (optional)", value="", key="opt_review_tag")
                        cols = st.columns(3)
                        if cols[0].button("Add comment", key="opt_review_add_comment"):
                            sess.comments.append({
                                "ts": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                                "candidate_fp": sess.candidates[int(sel)].get("candidate_fp"),
                                "text": comment,
                            })
                            save_review_session(sess, Path(sp))
                            st.success("Comment added.")
                        if cols[1].button("Cast vote", key="opt_review_add_vote"):
                            sess.votes.append({
                                "ts": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                                "candidate_fp": sess.candidates[int(sel)].get("candidate_fp"),
                                "vote": int(vote),
                            })
                            save_review_session(sess, Path(sp))
                            st.success("Vote recorded.")
                        if cols[2].button("Add tag", key="opt_review_add_tag"):
                            if tag.strip():
                                sess.tags.append({
                                    "ts": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                                    "candidate_fp": sess.candidates[int(sel)].get("candidate_fp"),
                                    "tag": tag.strip(),
                                })
                                save_review_session(sess, Path(sp))
                                st.success("Tag added.")

                    with st.expander("Session data", expanded=False):
                        st.json(sess.to_dict(), expanded=False)

                    st.divider()
                    st.markdown("### 📤 Export / import")
                    st.download_button(
                        "Download review session bundle (.zip)",
                        data=export_review_session_zip(sess),
                        file_name=f"shams_review_session_{sess.session_id}.zip",
                        mime="application/zip",
                        use_container_width=True,
                    )
                    up = st.file_uploader("Import review session bundle", type=["zip"], key="opt_review_import")
                    if up is not None:
                        try:
                            sess2 = import_review_session_zip(up.read())
                            path = sessions_dir / f"{sess2.session_id}.json"
                            save_review_session(sess2, path)
                            st.session_state["opt_review_path"] = str(path)
                            st.success(f"Imported session: {sess2.session_id}")
                        except Exception as e:
                            st.error(f"Import failed: {e}")

            elif view == "Epistemic guarantees (regression suite)":
                st.caption(
                    "Ultimate differentiator: epistemic guarantees. Run the golden regression suite to detect unintended semantic drift. "
                    "This does not validate against reality; it enforces stability of the frozen evaluator and artifact contracts."
                )
                _repo_root = Path(__file__).resolve().parent.parent
                _eval_fp = repo_fingerprint(_repo_root)
                st.markdown(f"**Evaluator fingerprint:** `{_eval_fp}`")
                rtol = st.number_input("Relative tolerance", value=0.01, min_value=0.0, max_value=0.2, step=0.005, key="opt_reg_rtol")
                atol = st.number_input("Absolute tolerance", value=1e-6, min_value=0.0, max_value=1e-2, step=1e-6, format="%.1e", key="opt_reg_atol")
                if st.button("Run regression suite now", key="opt_reg_run", use_container_width=True):
                    st.session_state["opt_reg_report"] = run_regression_suite(_repo_root, rtol=float(rtol), atol=float(atol))
                rep = st.session_state.get("opt_reg_report")
                if rep:
                    if rep.get("ok"):
                        st.success("Regression suite PASSED.")
                    else:
                        st.error("Regression suite FAILED.")
                    with st.expander("Runner output", expanded=False):
                        st.code(rep.get("output", ""))
                    d = rep.get("diff") or {}
                    if d:
                        st.markdown("### Diff summary")
                        try:
                            st.json({
                                "numeric_failures": d.get("numeric", {}).get("summary"),
                                "structural": d.get("structural", {}).get("severity"),
                            }, expanded=False)
                        except Exception:
                            st.json(d, expanded=False)
                else:
                    st.info("Run the suite to get a pass/fail report + structured diffs.")

            elif view == "Standards & DOI export":
                st.caption(
                    "Tier-7 standards: export DOI-ready packs + SHAMS-certified feasibility badges (descriptive, non-ranking)."
                )
                _repo_root = Path(__file__).resolve().parent.parent
                _eval_fp = repo_fingerprint(_repo_root)
                version = str(st.session_state.get("app_version", ""))
                best = run.get("best_feasible")
                if not isinstance(best, dict):
                    st.info("No feasible best candidate yet. Run the machine finder first.")
                else:
                    cfp = candidate_fingerprint(best.get("inputs", {}) or {}, intent=run.get("intent", ""), evaluator_fp=_eval_fp)
                    svg = generate_cert_badge_svg(
                        candidate_fp=cfp,
                        intent=run.get("intent", ""),
                        feasible=bool(best.get("feasible", False)),
                        version=version,
                        evaluator_fp=_eval_fp,
                        note="audited by frozen evaluator",
                    )
                    st.download_button(
                        "Download SHAMS-certified badge (SVG)",
                        data=svg.encode("utf-8"),
                        file_name=f"shams_cert_badge_{cfp[:12]}.svg",
                        mime="image/svg+xml",
                        use_container_width=True,
                    )

                    # DOI-ready export pack (includes archive + trace)
                    archive_rows = run.get("archive") or []
                    trace_rows = run.get("trace") or []
                    run_meta = {
                        "schema": "shams.optimization_pack.v1",
                        "version": version,
                        "intent": run.get("intent"),
                        "seed": run.get("seed"),
                        "fingerprint": run.get("fingerprint"),
                        "evaluator_fp": _eval_fp,
                        "objectives": run.get("objectives"),
                        "var_specs": run.get("var_specs"),
                        "budgets": run.get("budgets"),
                        "notes": "Export pack is descriptive. No hidden preferences.",
                    }
                    extra = [(f"badges/shams_cert_badge_{cfp[:12]}.svg", svg.encode("utf-8"))]
                    # Attach review session if loaded
                    sp = st.session_state.get("opt_review_path")
                    if sp:
                        try:
                            sess = load_review_session(Path(sp))
                            extra.append(("review_session.json", json.dumps(sess.to_dict(), indent=2, sort_keys=True).encode("utf-8")))
                        except Exception:
                            pass
                    pack = export_doi_ready_pack(
                        repo_root=_repo_root,
                        run_meta=run_meta,
                        archive_rows=archive_rows,
                        trace_rows=trace_rows,
                        extra_files=extra,
                    )
                    st.download_button(
                        "Download DOI-ready publication pack (.zip)",
                        data=pack,
                        file_name=f"shams_optimization_publication_pack_{run.get('fingerprint','')[:12]}.zip",
                        mime="application/zip",
                        use_container_width=True,
                    )

            elif view == "Design-space verdicts (Allowed/Forbidden)":
                st.caption(
                    "Tier-8: Design-space jurisprudence. This is not a recommendation engine - it classifies what is supported by evidence in the explored region. "
                    "Forbidden here means *locally forbidden within the explored neighborhood*, not a universal impossibility theorem."
                )
                ci = feasibility_confidence_from_trace(run.get("trace") or [], window=500)
                rv = region_verdict(run.get("trace") or [], window=500)
                st.markdown("#### Region verdict (recent window)")
                st.write({
                    "verdict": rv.label,
                    "confidence": f"{rv.confidence:.2f}",
                    "rate": ci.get("rate"),
                    "ci_95": [ci.get("ci_lo"), ci.get("ci_hi")],
                    "rationale": rv.rationale,
                })
                st.markdown("#### Candidate verdict (selected)")
                if filt:
                    cand = filt[int(st.session_state.get("opt_inspect_idx", 0))]
                    vv = candidate_verdict(cand, archive=run.get("archive") or [], var_keys=var_keys, robust_margin=float(min_margin) if min_margin else 0.0)
                    st.write({"verdict": vv.label, "confidence": f"{vv.confidence:.2f}", "rationale": vv.rationale})
                else:
                    st.info("No candidates available.")

            elif view == "Epistemic confidence bounds":
                st.caption(
                    "Tier-8: Epistemic confidence bounds on feasibility rates (Wilson interval). This quantifies how strongly the *recent search evidence* supports feasibility/infeasibility."
                )
                w = st.slider("Window (last N evaluations)", 50, 2000, 500, 50, key="opt_ci_window")
                ci = feasibility_confidence_from_trace(run.get("trace") or [], window=int(w))
                st.write(ci)
                st.markdown("**Interpretation**")
                st.write(
                    "- If `k=0` and the upper CI is very small, the explored region is strongly supported as locally infeasible.\n"
                    "- If `k>0`, feasibility is established for the explored region; the CI describes how often feasibility occurs under the current proposal distribution." 
                )

            elif view == "Intent-conditional design laws":
                st.caption(
                    "Tier-8: Intent-conditional design laws. We take top feasible candidates under the current intent and re-evaluate them under the other intent, then compare correlations."
                )
                if not filt:
                    st.info("No candidates available.")
                else:
                    other_intent = "Research" if run.get("intent") == "Reactor" else "Reactor"
                    key_y = st.text_input("Output key to analyze", value="P_e_net_MW", key="opt_laws_keyy")
                    topn = st.slider("Top feasible candidates to compare", 10, 120, 40, 5, key="opt_laws_topn")
                    def _eval_other(inp: dict) -> dict:
                        return _evaluate_candidate(inp, intent=other_intent)
                    laws = intent_conditional_laws(_eval_other, archive=run.get("archive") or [], var_keys=var_keys, key_y=str(key_y), top_n=int(topn))
                    st.write({"primary_intent": run.get("intent"), "other_intent": other_intent, "n_primary": laws.get("n_primary"), "n_other": laws.get("n_other")})
                    try:
                        import pandas as _pd
                        df = _pd.DataFrame(laws.get("rows") or [])
                        st.dataframe(df, use_container_width=True)
                    except Exception:
                        st.json(laws)

            elif view == "Machine genealogy":
                st.caption(
                    "Tier-9: Machine genealogy. When engines do not record parents explicitly, SHAMS reconstructs a conservative ancestry graph: each candidate's parent is its nearest better neighbor in variable space."
                )
                if not filt:
                    st.info("No candidates available.")
                else:
                    maxch = st.slider("Max children per parent (for readability)", 3, 30, 12, 1, key="opt_gene_maxch")
                    g = reconstruct_genealogy(run.get("archive") or [], var_keys=var_keys, max_children_per_parent=int(maxch))
                    st.write({"roots": g.get("roots"), "num_nodes": len(g.get("parents") or {})})
                    # Render a small textual tree for the top few roots
                    roots = list(g.get("roots") or [])[:5]
                    parents = g.get("parents") or {}
                    children = g.get("children") or {}
                    def _node_label(i: int) -> str:
                        try:
                            a = (run.get("archive") or [])[int(i)]
                            return f"#{i} | feas={bool(a.get('feasible'))} | score={float(a.get('_score', -1e30)):.3g} | m={float(a.get('min_signed_margin', float('nan'))):.3g}"
                        except Exception:
                            return f"#{i}"
                    def _render(i: int, depth: int = 0, maxd: int = 3):
                        lines = ["  "*depth + "- " + _node_label(i)]
                        if depth >= maxd:
                            return lines
                        for ch in (children.get(i) or [])[:10]:
                            lines.extend(_render(int(ch), depth+1, maxd))
                        return lines
                    with st.expander("Genealogy tree (top roots)", expanded=False):
                        text = []
                        for r in roots:
                            text.extend(_render(int(r), 0, 3))
                        st.code("\n".join(text))

            elif view == "Counter-optimization (no interior optimum)":
                st.caption(
                    "Tier-9: Counter-optimization. This does not claim mathematical proofs; it reports evidence that improvement is boundary-limited (no interior optimum) under the current search space."
                )
                key_obj = st.text_input("Objective key (default: _score)", value="_score", key="opt_counter_key")
                rep = counter_optimization_report(run.get("archive") or [], key_obj=str(key_obj))
                if rep.get("status") == "ok":
                    if rep.get("boundary_limited"):
                        st.warning(rep.get("message"))
                    else:
                        st.info(rep.get("message"))
                else:
                    st.info(rep.get("message"))
                st.json(rep, expanded=False)

            elif view == "Reproducibility":
                st.caption("Audit capsule: fingerprint + config + optional citation.")
                st.json({
                    "fingerprint": run.get("fingerprint"),
                    "intent": run.get("intent"),
                    "seed": run.get("seed"),
                    "objectives": run.get("objectives"),
                    "var_specs": run.get("var_specs"),
                    "budgets": run.get("budgets"),
                }, expanded=False)
                try:
                    from pathlib import Path as _Path
                    cff = (_Path(__file__).resolve().parent.parent / "CITATION.cff").read_text(encoding="utf-8")
                    with st.expander("CITATION.cff", expanded=False):
                        st.code(cff, language="yaml")
                    try:
                        pm_path = _Path(__file__).resolve().parent.parent / "examples" / "published_machines.json"
                        if pm_path.exists():
                            import json as _json
                            pm = _json.loads(pm_path.read_text(encoding="utf-8"))
                            with st.expander("Published machine overlay (optional)", expanded=False):
                                st.json(pm)
                    except Exception:
                        pass
                except Exception:
                    pass
            else:
                st.info("Pareto view: available when ≥2 objectives are configured.")

        # RIGHT: inspector
        with right:
            st.markdown("### Inspector")
            idx = st.number_input("Candidate index (in filtered archive)", 0, max(0, len(filt)-1), 0, key="opt_inspect_idx")
            if filt:
                cand = filt[int(idx)]
                st.markdown("**Candidate summary**")
                st.write({
                    "feasible": bool(cand.get("feasible", False)),
                    "score": float(cand.get("_score", -1e30)),
                    "min_signed_margin": float(cand.get("min_signed_margin", float("nan"))),
                    "failure_mode": cand.get("failure_mode"),
                    "dominant_constraints": cand.get("active_constraints", [])[:5],
                })
                with st.expander("Inputs", expanded=False):
                    st.json(cand.get("inputs", {}))
                with st.expander("Key outputs", expanded=False):
                    outs = cand.get("outputs", {}) or {}
                    keys_show = ["P_e_net_MW","Pfus_total_MW","Q_DT_eqv","q_div_MW_m2","B_peak_T","Ti_keV","nbar_1e20_m3"]
                    st.json({k: outs.get(k) for k in keys_show if k in outs})
                with st.expander("Constraint margins (blocking/diagnostic/ignored)", expanded=False):
                    _cons = cand.get("constraints", []) or []
                    # Optional credibility overlay (display-only)
                    try:
                        if st.session_state.get("opt_use_cred") and st.session_state.get("opt_cred_map"):
                            _cm = {
                                k: ConstraintCred(
                                    name=v.get("name", k),
                                    maturity=float(v.get("maturity", 0.7)),
                                    uncertainty_frac=float(v.get("uncertainty_frac", 0.10)),
                                    conservative=bool(v.get("conservative", True)),
                                )
                                for k, v in (st.session_state.get("opt_cred_map") or {}).items()
                            }
                            _cons = apply_credibility_overlay(_cons, _cm)
                    except Exception:
                        pass
                    st.write(_cons[:80])
                if enable_multi_intent and cand.get("other_intent"):
                    with st.expander("Other Intent distance (instrumentation)", expanded=False):
                        st.write({
                            "other_intent": cand.get("other_intent"),
                            "other_feasible": cand.get("other_feasible"),
                            "other_violation": cand.get("other_violation"),
                            "other_min_signed_margin": cand.get("other_min_signed_margin"),
                            "other_failure_mode": cand.get("other_failure_mode"),
                        })
                if use_cost:
                    with st.expander("Cost proxies (transparent)", expanded=False):
                        st.json(cand.get("cost", {}))

                st.markdown("**Actions (explicit, reversible)**")
                if st.button("Send to Systems Mode (as starting point)", key="opt_send_systems"):
                    st.session_state["systems_seed_inputs"] = dict(cand.get("inputs", {}))
                    st.success("Sent to Systems Mode seed (session-only).")
                if st.button("Open in Point Designer (read-only)", key="opt_send_point"):
                    st.session_state["point_inputs_last"] = dict(cand.get("inputs", {}))
                    st.success("Loaded into Point Designer inputs (session-only).")

                if st.button("Open Scan Lab slice around this candidate", key="opt_send_scan"):
                    try:
                        inp = cand.get("inputs") or {}
                        # Default axes chosen for expert usefulness
                        xk, yk = "Ip_MA", "R0_m"
                        x0 = float(inp.get(xk, 0.0)); y0 = float(inp.get(yk, 0.0))
                        # ±10% local neighborhood (bounded away from zero)
                        def _band(v: float) -> tuple[float,float]:
                            dv = max(0.05*abs(v), 0.01)
                            return (v - dv, v + dv)
                        xlo, xhi = _band(x0)
                        ylo, yhi = _band(y0)
                        st.session_state["scan_cart_x_key"] = xk
                        st.session_state["scan_cart_y_key"] = yk
                        st.session_state["scan_cart_x_lo"] = float(xlo)
                        st.session_state["scan_cart_x_hi"] = float(xhi)
                        st.session_state["scan_cart_y_lo"] = float(ylo)
                        st.session_state["scan_cart_y_hi"] = float(yhi)
                        st.session_state["scan_cart_nx"] = 31
                        st.session_state["scan_cart_ny"] = 25
                        st.session_state["scan_cart_intents"] = [str(run.get("intent") or "Reactor")]
                        st.session_state["scan_cart_inc_out"] = False
                        st.success("Scan Lab settings prepared. Click the Scan Lab tab and press Run cartography scan.")
                    except Exception as e:
                        st.error(f"Failed to prepare Scan Lab slice: {e}")

            else:
                st.info("No candidates to inspect.")


if _deck == "Compare":
    from ui.decks.compare import render_compare
    render_compare(sys.modules[__name__])

# (v372.8.7) Studies manager is rendered inside Control Room → Studies tab (no tab-handle leakage).
# The previous module-scope `with tab_studies:` block was removed to satisfy UI law.

if _deck == "Reactor Design Forge":
    st.subheader("🧪 Operating envelope check (multi-point)")
    st.caption("Evaluates startup / nominal / end-of-life proxy points and reports the worst constraint.")
    colA, colB = st.columns([1,3])
    with colA:
        run_env = st.button("Run envelope check", use_container_width=True)
    if run_env:
        try:
            from envelope.points import default_envelope_points
            from constraints.system import build_constraints_from_outputs, summarize_constraints
            base_inp = st.session_state.get("last_point_inp", None)
            if base_inp is None:
                st.warning("No current point inputs available.")
            else:
                _warn_unrealistic_point_inputs(base_inp, context="Envelope check")
                pts = default_envelope_points(base_inp)
                env_rows = []
                worst = None
                for i, p in enumerate(pts):
                    out = _ui_evaluate(p, origin="envelope_scan")
                    cs = build_constraints_from_outputs(out)
                    summ = summarize_constraints(cs)
                    dom = summ.get("dominant", {})
                    row = {
                        "point": i,
                        "all_ok": bool(summ.get("all_ok", False)),
                        "dominant": dom.get("name", ""),
                        "residual": dom.get("residual", float("nan")),
                        "margin": dom.get("margin", float("nan")),
                    }
                    env_rows.append(row)
                    if worst is None or (row["residual"] == row["residual"] and row["residual"] > worst["residual"]):
                        worst = row
                st.dataframe(env_rows, use_container_width=True)
                if worst:
                    st.info(f"Worst point: #{worst['point']} - {worst['dominant']} (residual={worst['residual']:.3g})")
        except Exception as e:
            st.error(f"Envelope check failed: {e}")


if _deck == "Control Room":
    with tab_model:
        st.header("0-D Tokamak Physics Model (Phase‑1)")

        with st.expander("0‑D Physical Models - explanations", expanded=False):
            _pm = os.path.join(ROOT, "docs", "PHYSICAL_MODELS_0D.md")
            try:
                with open(_pm, "r", encoding="utf-8") as _f:
                    st.markdown(_f.read())
            except Exception as _e:
                st.error(f"Failed to load physical model doc: {_e}")


        st.markdown(r"""
        This tab is written to be **actionable**: each section maps to code in `src/physics/`, `src/phase1_models.py`,
        `src/phase1_systems.py`, and `src/solvers/`.

        SHAMS remains a **0‑D / volume‑averaged / steady‑state** point-design model at its core, intended for *fast feasibility scanning*.
        Over time we have added several **external systems codes‑inspired** upgrades that remain lightweight and Windows‑friendly:

        - **Optional analytic profiles ("½‑D")** for $n_e(\rho)$, $T_i(\rho)$, $T_e(\rho)$ with **normalization to the chosen volume averages**,
          plus derived averages like peaking factors and $\langle n_e^2 \rangle/\langle n_e \rangle^2$.
        - **Radiation options:** legacy fractional radiation (stable for scans) and a physics‑based path (brem + synchrotron + simple impurity line radiation).
        - **Constraint system:** engineering and plasma constraints are represented as reusable objects (external systems codes‑like), usable by scans and vector solvers.
        - **Solvers:** classic nested 1‑D solves are still available, plus a more general bounded "targets → variables" solve primitive.

        It is **not** a full transport / equilibrium / neutronics code, but it is designed to grow in that direction while staying usable.
        """)

        st.caption("Tip: expand only the models you care about - each block is independent.")

        # --- Geometry ---
        with st.expander("Geometry: volume and surface area (implemented)", expanded=False):
            st.markdown(r"""
            Implemented helpers:

            **Plasma volume** (`tokamak_volume`)
            $$
            V \approx 2\pi^2\,R\,a^2\,\kappa
            $$

            **Plasma surface area** (`tokamak_surface_area`)
            $$
            S \approx 4\pi^2\,R\,a\,\kappa
            $$

            Notes:
            - These are **engineering approximations** intended to preserve correct monotonic trends.
            - Units: $R,a$ in m, $V$ in m$^3$, $S$ in m$^2$.
            """)

        # --- Confinement ---
        with st.expander("Energy confinement: IPB98(y,2) (implemented)"):
            st.markdown(r"""
            Implemented model: `tauE_ipb98y2`.

            $$
            \tau_E = 0.0562\, I_p^{0.93} B_t^{0.15} \bar{n}^{0.41} P_{loss}^{-0.69} R^{1.97} \epsilon^{0.58} \kappa^{0.78} M^{0.19}
            $$
            where $\epsilon=a/R$.

            **Units (must match the implementation):**
            - $I_p$ in MA
            - $B_t$ in T
            - $\bar{n}$ in units of $10^{20}\,\mathrm{m^{-3}}$ (i.e. `ne20`)
            - $P_{loss}$ in MW
            - $R,a$ in m
            - $M$ in amu (default 2.5)

            Output: $\tau_E$ in seconds.
            """)

        # --- L-H threshold ---
        with st.expander("H-mode access: Martin-2008 L–H threshold (implemented)"):
            st.markdown(r"""
            Implemented model: `p_LH_martin08`.

            $$
            P_{LH} = 0.0488\, \bar{n}^{0.717} B_t^{0.803} S^{0.941}\,\left(\frac{2}{A_{eff}}\right)
            $$

            **Units:**
            - $\bar{n}$ in $10^{20}\,\mathrm{m^{-3}}$ (line-averaged)
            - $B_t$ in T
            - $S$ in m$^2$ (uses the same proxy as the geometry block)
            - $A_{eff}$ dimensionless (defaults to 2.0)

            Output: $P_{LH}$ in MW.
            """)

        # --- Greenwald ---
        with st.expander("Density limit: Greenwald (implemented)"):
            st.markdown(r"""
            Implemented helper: `greenwald_density_20`.

            $$
            n_{GW}\,[10^{20}\,\mathrm{m^{-3}}] = \frac{I_p\,[\mathrm{MA}]}{\pi a^2\,[\mathrm{m^2}]}
            $$

            In scans, an operating fraction is typically applied:
            $$
            \bar{n} = f_{nG}\,n_{GW},\qquad 0 < f_{nG} \le 1.
            $$
            """)

        # --- Screening proxies ---
        with st.expander("Screening proxies: q95, βN, bootstrap fraction (implemented proxies)"):
            st.markdown(r"""
            These are explicitly labeled **proxies** (trend-correct, not equilibrium/transport solutions).

            **q95 proxy** (`q95_proxy_cyl`)
            $$
            q_{95} \approx \left(\frac{2\pi R B_t}{\mu_0 I_p}\right)\left(\frac{a}{R}\right)\frac{1}{\kappa}
            $$
            with $I_p$ converted to amperes internally.

            **Normalized beta** (`betaN_from_beta`)
            $$
            \beta_N = \beta(\%)\,\frac{a\,B_t}{I_p}
            \qquad\text{with}\qquad \beta(\%)=100\,\beta
            $$
            where $\beta$ is the *fractional* beta.

            **Bootstrap fraction proxy** (`bootstrap_fraction_proxy`)
            $$
            f_{bs} \approx C_{bs}\,\frac{\beta_N}{q_{95}}
            $$
            then clamped to a configured range (default 0 to 0.95).
            """)

        # --- Fusion reactivity ---
        with st.expander("Fusion reactivity: Bosch–Hale ⟨σv⟩ (implemented)"):
            st.markdown(r"""
            Implemented function: `bosch_hale_sigmav(T_i, reaction)`.

            This uses the Bosch–Hale parameterization for Maxwellian-averaged reactivity:
            $$
            \langle\sigma v\rangle(T_i)\;[\mathrm{m^3/s}]
            $$

            Internally, the implementation computes intermediate variables ($\theta$, $\xi$) from a
            reaction-specific coefficient set and returns a strictly non-negative value.

            **Important for UI users:**
            - Input $T_i$ is in **keV**.
            - Output is in **m$^3$/s**.
            """)

            # Bosch–Hale coefficient values used by the implementation (from `BH_COEFFS`)
            _bh_rows = []
            for _rxn in ["DT", "DD_Tp", "DD_He3n"]:
                _c = BH_COEFFS[_rxn]
                _bh_rows.append({"Reaction": _rxn, **asdict(_c)})
            _bh_df = pd.DataFrame(_bh_rows).set_index("Reaction")
            st.caption("Bosch–Hale coefficients used for DT and DD channels (exact values as implemented).")
            st.dataframe(_bh_df, use_container_width=True)

        # --- Fusion power / gain symbols (fixing the screenshot issue) ---
        with st.expander("Fusion power & gain definitions: P_f, P_α, Q (notation)"):
            st.markdown(r"""
            **What these symbols mean (and how they relate):**

            **Fusion power, $P_f$**  
            Total thermal power released by fusion reactions occurring in the plasma:
            $$
            P_f \;=\; \dot{N}_{\text{fus}}\,E_{\text{fus}}
            $$
            where $\dot{N}_{\text{fus}}$ is the fusion reaction rate [1/s] and $E_{\text{fus}}$ is the energy released per reaction.
            For D‑T, $E_{\text{fus}} = 17.6\,\mathrm{MeV}$.

            **Alpha heating power, $P_{\alpha}$**  
            Part of $P_f$ carried by *charged* alpha particles and deposited back into the plasma (self‑heating):
            $$
            P_{\alpha} \;=\; f_\alpha\,P_f
            $$
            For D‑T, $f_\alpha = \frac{3.5}{17.6} \approx 0.199$, so $P_{\alpha} \approx 0.20\,P_f$.
            (The rest is mainly neutron power: $P_n \approx 0.80\,P_f$.)

            **Fusion gain, $Q$**  
            In this UI, $Q$ is the standard *plasma gain*:
            $$
            Q \;=\; \frac{P_f}{P_{\mathrm{aux}}}
            $$
            where $P_{\mathrm{aux}}$ is the **externally applied** auxiliary heating power (e.g., NBI/RF) required to sustain the operating point.
            This is distinct from “wall‑plug” gain, which would include plant efficiencies and non‑plasma power draws.

            **How to interpret in scans**
            - Increasing $P_f$ increases $P_{\alpha}$ proportionally (more self‑heating).  
            - $Q$ improves only when $P_f$ grows faster than the required $P_{\mathrm{aux}}$.
            """)

        # --- SOL width metric ---
        with st.expander("Optional divertor/SOL risk metric: Eich λq (implemented)"):
            st.markdown(r"""
            Implemented metric: `lambda_q_eich14_mm`.

            $$
            \lambda_q\,[\mathrm{mm}] \approx \text{factor}\times 0.63\,B_{pol}^{-1.19}
            $$

            with $B_{pol}$ approximated by:
            $$
            B_{pol} \approx \frac{\mu_0 I_p}{2\pi a}
            $$

            This is **not** a self‑consistent divertor / edge power‑exhaust model - it’s a compact, order‑of‑magnitude **screening proxy** for quickly comparing design points.
            """)

        st.info(
            "If you want the *full* step-by-step closure shown here (power balance → temperatures → Pf/Q), "
            "tell me which exact function in `src/phase1_core.py` you want treated as the single source of truth, "
            "and I’ll mirror it line-for-line in this tab."
        )

# -----------------------------
# Benchmarks
# -----------------------------
if _deck == "Control Room":
    with tab_bench:
        st.subheader("Regression Benchmarks")
        st.write("Run a small suite of SPARC-like cases to ensure recent physics/solver changes haven't broken behavior.")

        import json
        from pathlib import Path

        bench_dir = Path(__file__).resolve().parent.parent / "benchmarks"
        cases_path = bench_dir / "cases.json"
        golden_path = bench_dir / "golden.json"

        diff_path = bench_dir / "last_diff_report.json"
        with st.expander("Latest diff report (from last run)", expanded=False):
            if diff_path.exists():
                try:
                    rep = json.loads(diff_path.read_text(encoding="utf-8"))
                    st.caption(f"Generated at unix={rep.get('created_unix'):.0f} | failures={rep.get('n_failed',0)}")
                    rows = rep.get("rows", [])
                    if rows:
                        import pandas as pd

                        df_rep = pd.DataFrame(rows)
                        # show worst first
                        if "ok" in df_rep.columns and "rel_err" in df_rep.columns:
                            df_rep = df_rep.sort_values(by=["ok","rel_err"], ascending=[True, False])
                        st.dataframe(df_rep, use_container_width=True, height=260)
                    # Structural diffs (constraints/model cards) vs golden artifacts, if present
                    ss = rep.get("structural_summary")
                    if ss:
                        st.markdown("**Structural diffs vs golden artifacts**")
                        st.write({k: ss.get(k) for k in ["n_cases","n_with_changes","total_added_constraints","total_removed_constraints","total_changed_constraints","total_modelcard_changes"]})
                    structural = rep.get("structural") or {}
                    if structural:
                        with st.expander("Show structural diffs by case", expanded=False):
                            for cname, d in structural.items():
                                cadd = d.get("constraints", {}).get("added", [])
                                crem = d.get("constraints", {}).get("removed", [])
                                cchg = d.get("constraints", {}).get("changed_meta", [])
                                mc = d.get("model_cards", {})
                                mcchg = (mc.get("added", []) or []) + (mc.get("removed", []) or []) + (mc.get("changed", []) or [])
                                if not (cadd or crem or cchg or mcchg or (d.get("schema_version", {}).get("new") != d.get("schema_version", {}).get("old"))):
                                    continue
                                with st.expander(f"{cname}", expanded=False):
                                    if cadd: st.write({"constraints_added": cadd})
                                    if crem: st.write({"constraints_removed": crem})
                                    if cchg: st.write({"constraint_meta_changes": cchg})
                                    if mc.get("added"): st.write({"model_cards_added": mc.get("added")})
                                    if mc.get("removed"): st.write({"model_cards_removed": mc.get("removed")})
                                    if mc.get("changed"): st.write({"model_cards_changed": mc.get("changed")})

                    st.download_button("Download diff report JSON", data=diff_path.read_bytes(), file_name="last_diff_report.json")
                except Exception as e:
                    st.warning(f"Could not read diff report: {e}")
            else:
                st.info("No diff report yet. Run benchmarks to generate one.")


        # Release notes (auto-generated)
        with st.expander("Release notes (auto)", expanded=False):
            import subprocess, sys
            from pathlib import Path

            repo_root = Path(__file__).resolve().parent.parent
            out_md = repo_root / "RELEASE_NOTES.md"
            old_default = str((repo_root.parent / "SHAMS_old").resolve()) if (repo_root.parent / "SHAMS_old").exists() else r"..\SHAMS_old"
            old_path = st.text_input("Old SHAMS repo path", value=st.session_state.get("release_notes_old", old_default))
            st.session_state["release_notes_old"] = old_path

            auto = st.checkbox("Auto-generate if missing/out-of-date", value=True, key="release_notes_auto")
            run_now_rn = st.button("Generate release notes now", key="btn_release_notes_now")

            def _needs_rn() -> bool:
                if not out_md.exists():
                    return True
                try:
                    m_out = out_md.stat().st_mtime
                    # regenerate if diff report is newer, or tool changed
                    tool_p = repo_root / "tools" / "release_notes.py"
                    diff_p = repo_root / "benchmarks" / "last_diff_report.json"
                    newest = max([p.stat().st_mtime for p in [tool_p, diff_p] if p.exists()] + [0])
                    return newest > m_out
                except Exception:
                    return False

            if (auto and _needs_rn() and not st.session_state.get("_rn_ran_this_session", False)) or run_now_rn:
                cmd = [sys.executable, str(repo_root / "tools" / "release_notes.py"), "--old", old_path, "--new", str(repo_root), "--out", str(out_md)]
                st.caption("Running: " + " ".join(cmd))
                try:
                    p = subprocess.run(cmd, capture_output=True, text=True, cwd=str(repo_root))
                    st.session_state["_rn_last_stdout"] = p.stdout
                    st.session_state["_rn_last_stderr"] = p.stderr
                    st.session_state["_rn_last_rc"] = p.returncode
                    st.session_state["_rn_ran_this_session"] = True
                except Exception as e:
                    st.session_state["_rn_last_stderr"] = str(e)
                    st.session_state["_rn_last_rc"] = 1

            rc = st.session_state.get("_rn_last_rc")
            if rc is not None:
                if rc == 0:
                    st.success("Release notes generated.")
                else:
                    st.warning("Release notes generation had issues (see logs).")
                with st.expander("Logs", expanded=False):
                    st.code((st.session_state.get("_rn_last_stdout") or "") + "\n" + (st.session_state.get("_rn_last_stderr") or ""))

            if out_md.exists():
                st.markdown(out_md.read_text(encoding="utf-8", errors="ignore"))
                st.download_button("Download RELEASE_NOTES.md", data=out_md.read_bytes(), file_name="RELEASE_NOTES.md", mime="text/markdown")
            else:
                st.info("RELEASE_NOTES.md not found yet.")

        with st.expander("Regression comparisons", expanded=False):
            colA, colB = st.columns([1,1])
            with colA:
                run_now = st.button("Run benchmarks")
            with colB:
                regen = st.button("Regenerate golden (intentional changes)")
    
            def _safe(v):
                try:
                    return float(v)
                except Exception:
                    return float("nan")
    
            if cases_path.exists():
                _cases_raw = json.loads(cases_path.read_text())
            else:
                _cases_raw = {}
    
            # Normalize benchmark cases into a list[dict] with keys: name, inputs
            # Supports dict-form (name -> inputs), list-form (dicts), or list-form (names).
            cases = []
            if isinstance(_cases_raw, dict):
                for _name, _inp in _cases_raw.items():
                    if isinstance(_inp, dict):
                        cases.append({"name": str(_name), "inputs": _inp})
            elif isinstance(_cases_raw, list):
                for i, _c in enumerate(_cases_raw):
                    if isinstance(_c, dict):
                        _name = _c.get("name", f"case_{i}")
                        _inp = _c.get("inputs", _c.get("inp", _c.get("input", {})))
                        if isinstance(_inp, dict):
                            cases.append({"name": str(_name), "inputs": _inp})
                    else:
                        cases.append({"name": str(_c), "inputs": {}})
    
            if not cases:
                # Always provide at least one default case so the UI doesn't crash
                cases = [{"name": "default", "inputs": {"R0_m": 1.85, "a_m": 0.6, "kappa": 1.75, "Bt_T": 12.0, "Ip_MA": 8.0, "Ti_keV": 10.0, "fG": 0.85, "t_shield_m": 0.25, "Paux_MW": 25.0}}]
    
            base = PointInputs(R0_m=1.85, a_m=0.6, kappa=1.75, Bt_T=12.0, Ip_MA=8.0, Ti_keV=10.0, fG=0.85, t_shield_m=0.25, Paux_MW=25.0)
            if run_now or regen:
                results = {}
                for _case in cases:
                    name = _case.get("name","case")
                    overrides = _case.get("inputs", {})
                    # Defensive: apply only existing fields
                    d = base.__dict__.copy()
                    for k, v in overrides.items():
                        if k in d:
                            d[k] = v
                    inp_case = PointInputs(**d)
                    # Golden parity: matches tests/test_golden_physics_outputs.py (bypasses Evaluator).
                    from physics.hot_ion import hot_ion_point
                    results[name] = hot_ion_point(inp_case)
    
                if regen:
                    golden_path.write_text(json.dumps(results, indent=2))
                    st.success(f"Wrote golden: {golden_path}")
                else:
                    if not golden_path.exists():
                        st.error("golden.json not found. Click 'Regenerate golden' once to create it.")
                    else:
                        golden = json.loads(golden_path.read_text())
                        CURATED = ["Q_DT_eqv","H98","P_fus_MW","P_SOL_MW","q_div_MW_m2","B_peak_T","sigma_hoop_MPa","hts_margin_cs","J_eng_A_mm2","t_flat_s","P_net_MW"]
                        rows = []
                        failed = 0
                        for name, cur in results.items():
                            ref = golden.get(name, {})
                            for k in CURATED:
                                a = _safe(cur.get(k))
                                b = _safe(ref.get(k))
                                if not (math.isfinite(a) and math.isfinite(b)):
                                    continue
                                atol = 1e-6
                                rtol = 5e-3
                                ok = abs(a-b) <= max(atol, rtol*max(abs(b),1e-9))
                                if not ok:
                                    failed += 1
                                rows.append({"case":name,"key":k,"got":a,"golden":b,"rel_err":(abs(a-b)/max(abs(b),1e-9)),"ok":ok})
                        import pandas as pd

                        df = pd.DataFrame(rows)
                        st.dataframe(df, use_container_width=True)

                        # Write a machine-readable diff report (used by CI and the UI)
                        try:
                            import time as _time
                            report = {
                                "created_unix": _time.time(),
                                "rtol": 5e-3,
                                "atol": 1e-6,
                                "n_rows": int(len(rows)),
                                "n_failed": int(failed),
                                "rows": rows,
                            }
                            (bench_dir / "last_diff_report.json").write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")
                        except Exception:
                            pass

                        if failed==0:
                            st.success("All benchmark comparisons passed (within tolerances).")
                        else:
                            st.warning(f"{failed} comparisons exceeded tolerance. See table.")
    
    
        st.divider()
        with st.expander("Sensitivity and uncertainty (Monte Carlo)", expanded=False):
            st.subheader("Sensitivity (Monte Carlo)")
            st.write("Runs a lightweight uncertainty scan around a selected benchmark case (Windows-native).")
    
            from analysis.sensitivity import monte_carlo_feasibility
            from models.inputs import PointInputs
    
            case_names = [c.get("name", f"case_{i}") for i,c in enumerate(cases)]
            case_sel = st.selectbox("Benchmark case for sensitivity", case_names, index=0, key="sens_case_sel")
            n_mc = st.number_input("Samples", min_value=50, max_value=2000, value=50, step=50, key="sens_n")
            if st.button("Run Monte Carlo", key="run_mc_bench"):
                c = cases[case_names.index(case_sel)]
                base_inp = PointInputs(**c["inputs"])
                res = monte_carlo_feasibility(base_inp, n=int(n_mc), seed=42)
                st.metric("Feasible probability", f"{res['p_feasible']*100:.1f}%")
                st.write("Most frequently violated constraints:")
                st.dataframe(res["worst_constraints"], use_container_width=True)
    
        st.divider()
        with st.expander("Pareto search (design studies)", expanded=False):
            st.subheader("Pareto Search (LHS)")
            st.write("Finds a feasible Pareto set for a small set of design knobs around a benchmark case.")
            from solvers.optimize import pareto_optimize
    
            case_sel2 = st.selectbox("Benchmark case for Pareto", case_names, index=0, key="pareto_case_sel")
            n_lhs = st.number_input("LHS samples", min_value=50, max_value=5000, value=100, step=50, key="pareto_n")
            # simple bounds
            colp1, colp2 = st.columns(2)
            with colp1:
                R0_lo = st.number_input("R0 min [m]", value=1.5, step=0.1, key="R0_lo")
                Ip_lo = st.number_input("Ip min [MA]", value=5.0, step=0.5, key="Ip_lo")
            with colp2:
                R0_hi = st.number_input("R0 max [m]", value=2.5, step=0.1, key="R0_hi")
                Ip_hi = st.number_input("Ip max [MA]", value=12.0, step=0.5, key="Ip_hi")
            fG_lo = st.number_input("fG min", value=0.4, step=0.05, key="fG_lo")
            fG_hi = st.number_input("fG max", value=1.2, step=0.05, key="fG_hi")
    
            if st.button("Run Pareto search", key="run_pareto"):
                c = cases[case_names.index(case_sel2)]
                base_inp = PointInputs(**c["inputs"])
                bounds = {"R0_m": (float(R0_lo), float(R0_hi)), "Ip_MA": (float(Ip_lo), float(Ip_hi)), "fG": (float(fG_lo), float(fG_hi))}
                objectives = {"R0_m": "min", "B_peak_T": "min", "P_e_net_MW": "max"}
                res = pareto_optimize(base_inp, bounds=bounds, objectives=objectives, n_samples=int(n_lhs), seed=1)
                st.write(f"Feasible points: {len(res['feasible'])}  |  Pareto points: {len(res['pareto'])}")
                st.dataframe(res["pareto"], use_container_width=True)
    
        # -----------------------------
        # Variable Registry (auditable meanings/units/sources)
        # -----------------------------
if _deck == "Control Room":
    with tab_registry:
        st.subheader("Variable Registry")
        st.markdown(
            "A external systems codes-style registry of key SHAMS variables with units, meaning, and model provenance. "
            "Use this to keep the code **auditable** as physics/engineering fidelity increases."
        )
        q = st.text_input("Search variables", value="", placeholder="e.g., H98, q_div, HTS, TBR")
        try:
            df = registry_dataframe(q)
            st.dataframe(df, use_container_width=True, height=520)
            st.download_button(
                "Download registry (CSV)",
                data=df.to_csv(index=False),
                file_name="shams_variable_registry.csv",
                mime="text/csv",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Registry unavailable: {e}")

# -----------------------------
# Validation (envelopes)
# -----------------------------
if _deck == "Control Room":
    with tab_validation:
        st.subheader("Validation envelopes")
        st.markdown(
            "Decision-grade validation in SHAMS is **envelope-based**: we check whether a solution lies within "
            "a broad reference band for key metrics, rather than trying to match a single reference point. "
            "This is robust to proxy changes and is aligned with external systems codes-style workflows."
        )
        try:
            from validation.envelopes import default_envelopes
            envs = default_envelopes()
            env_name = st.selectbox("Select envelope", list(envs.keys()), index=0, key="validation_env_sel")
            env = envs[env_name]
            st.caption(env.notes)

            out = st.session_state.get("last_point_out")
            if not out:
                st.info("Run a Point Designer solve first. The latest outputs will be checked here.")
            else:
                report = env.check(out)
                import pandas as pd

                rows = []
                n_fail = 0
                for k, r in report.items():
                    if not r.get("ok"):
                        n_fail += 1
                    rows.append({
                        "metric": k,
                        "value": r.get("value"),
                        "lo": r.get("lo"),
                        "hi": r.get("hi"),
                        "ok": bool(r.get("ok")),
                    })
                df = pd.DataFrame(rows)
                st.dataframe(df, use_container_width=True, height=360)
                if n_fail == 0:
                    st.success("All selected envelope checks passed.")
                else:
                    st.warning(f"{n_fail} envelope checks failed. This indicates the *targets/bounds* are outside the reference band (not a code error).")
        except Exception as e:
            st.error(f"Validation module unavailable: {e}")

        st.divider()
        st.subheader("Invariant guardrails")
        st.caption("Deterministic sign/bookkeeping checks (not experimental validation).")
        try:
            from validation.invariants import check_invariants
            out = st.session_state.get("last_point_out")
            if not out:
                st.info("Run a Point Designer solve first. The latest outputs will be checked here.")
            else:
                rep = check_invariants(out)
                if bool(rep.get("ok")):
                    st.success("All invariant guardrails passed.")
                else:
                    st.error("Invariant guardrail failures detected (likely bookkeeping/sign issue or invalid inputs).")
                    st.json(rep.get("failures", {}))
        except Exception as e:
            st.caption(f"Invariant checks unavailable: {e}")


# -----------------------------
# Compliance (requirements + model cards)
# -----------------------------
if _deck == "Control Room":
    with tab_compliance:
        st.subheader("Verification & Compliance")
        st.caption("Shows the latest verification/compliance matrix from verification/report.json (if present).")

        def _load_verification_report_ui():
            try:
                here = Path(__file__).resolve()
                root = here.parent.parent  # ui/ -> repo root
                rp = root / "verification" / "report.json"
                if rp.exists():
                    return json.loads(rp.read_text(encoding="utf-8"))
            except Exception:
                return None
            return None

        report = _load_verification_report_ui()
        if not report:
            st.info("No verification/report.json found. Run: `python verification/run_verification.py` to generate it.")
        else:
            meta = report.get("meta", {})
            st.write({
                "generated_unix": meta.get("generated_unix"),
                "python": meta.get("python"),
                "platform": meta.get("platform"),
                "git_commit": meta.get("git_commit"),
            })

            # Summary
            summary = report.get("summary", {})
            cols = st.columns(4)
            cols[0].metric("Requirements", int(summary.get("n_requirements", 0)))
            cols[1].metric("Passed", int(summary.get("n_pass", 0)))
            cols[2].metric("Failed", int(summary.get("n_fail", 0)))
            cols[3].metric("Overall", "PASS" if summary.get("all_pass") else "FAIL")

            # Detailed table
            rows = report.get("results", [])
            if rows:
                df = pd.DataFrame(rows)
                # Friendly columns ordering
                keep = [c for c in ["req_id","title","status","details","linked_model_cards"] if c in df.columns]
                df = df[keep] if keep else df
                st.dataframe(df, use_container_width=True, height=520)

            # Download JSON
            st.download_button(
                "Download verification report.json",
                data=json.dumps(report, indent=2, sort_keys=True),
                file_name="verification_report.json",
                mime="application/json",
            )


if _deck == "Control Room":
    with tab_docs:
        st.header("Docs")
        st.caption("Built-in documentation bundled with this repository (no internet required).")

        doc_options = {
            "Upgrade plan (transparent)": os.path.join(ROOT, "docs", "SHAMS_upgrade_plan_from_PROCESS.md"),
            "Lessons learned (systems codes)": os.path.join(ROOT, "docs", "PROCESS_lessons.md"),
            "0‑D Physical Models (Phase‑1)": os.path.join(ROOT, "docs", "PHYSICAL_MODELS_0D.md"),
            "Engineering closures": os.path.join(ROOT, "docs", "ENGINEERING_CLOSURES.md"),
            "Operating envelope (multi-point)": os.path.join(ROOT, "docs", "ENVELOPE.md"),
            "Studies workflows": os.path.join(ROOT, "docs", "STUDIES.md"),
            "Model cards (auditability)": os.path.join(ROOT, "docs", "MODEL_CARDS.md"),
            "Compliance & verification": os.path.join(ROOT, "docs", "COMPLIANCE.md"),
            "Regression & golden benchmarks": os.path.join(ROOT, "docs", "REGRESSION.md"),
            "Release notes generation": os.path.join(ROOT, "docs", "RELEASE_NOTES.md"),
            "UI quickstart": os.path.join(ROOT, "README_UI.md"),
        }

        doc_sel = st.selectbox("Select a document", list(doc_options.keys()), index=0, key="doc_select")
        doc_path = doc_options.get(doc_sel)

        if doc_path and os.path.exists(doc_path):
            try:
                with open(doc_path, "r", encoding="utf-8") as f:
                    st.markdown(f.read())
            except Exception as _e:
                st.error(f"Failed to read doc: {_e}")
        else:
            st.warning("Document file not found in this checkout.")


# -----------------------------
# Artifacts Explorer (new)
# -----------------------------
def _load_json_from_upload(uploaded) -> Dict[str, Any] | None:
    if uploaded is None:
        return None
    try:
        raw = uploaded.getvalue()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8")
        return json.loads(raw)
    except Exception:
        return None


def _safe_df(rows: List[Dict[str, Any]]) -> pd.DataFrame:
    try:
        return pd.DataFrame(rows)
    except Exception:
        return pd.DataFrame()


def _as_float(x) -> float | None:
    try:
        if x is None:
            return None
        if isinstance(x, bool):
            return None
        return float(x)
    except Exception:
        return None


def _numeric_delta_table(base_out: Dict[str, Any], scen_out: Dict[str, Any], limit: int = 40) -> pd.DataFrame:
    keys = sorted(set(base_out.keys()) | set(scen_out.keys()))
    rows = []
    for k in keys:
        a = _as_float(base_out.get(k))
        b = _as_float(scen_out.get(k))
        if a is None or b is None:
            continue
        d = b - a
        # skip near-identical
        if abs(d) < 1e-12:
            continue
        rows.append({"metric": k, "baseline": a, "scenario": b, "delta": d, "delta_frac": (d / a) if abs(a) > 1e-12 else None})
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df = df.reindex(df["delta"].abs().sort_values(ascending=False).index)
    return df.head(limit)


if _deck == "Control Room":
    with tab_artifacts:
        st.header("Artifacts Explorer")
        st.caption("Load a SHAMS run artifact and inspect new v50+ artifact sections (constraint ledger, model set, standardized tables).")

        up = st.file_uploader("Upload shams_run_artifact.json", type=["json"], key="ae_upload")
        art = _load_json_from_upload(up)

        col_a, col_b = st.columns([1.2, 1.0])
        with col_a:
            alt_path = st.text_input("...or load from local path", value="", key="ae_path")
        with col_b:
            load_btn = st.button("Load from path", key="ae_load_path")

        if load_btn and alt_path:
            try:
                with open(alt_path, "r", encoding="utf-8") as f:
                    art = json.load(f)
            except Exception as e:
                st.error(f"Failed to load JSON: {type(e).__name__}: {e}")
                art = None

        if not art:
            st.info("Upload an artifact JSON (or provide a path) to explore.")
        else:
            meta = art.get("meta", {}) or {}
            prov = art.get("provenance", {}) or {}
            st.subheader("Metadata")
            st.write({
                "schema_version": art.get("schema_version"),
                "label": meta.get("label"),
                "mode": meta.get("mode"),
                "git_commit": prov.get("git_commit"),
                "python": prov.get("python"),
                "platform": prov.get("platform"),
                "repo_version": prov.get("repo_version"),
            })

            # --- Constraint ledger ---
            st.subheader("Constraint Margin Ledger")
            ledger = art.get("constraint_ledger") or {}
            if isinstance(ledger, dict) and ledger.get("entries"):
                st.caption(f"schema={ledger.get('schema_version','(missing)')}  fingerprint={ledger.get('ledger_fingerprint_sha256','(missing)')}")
                top = ledger.get("top_blockers") or []
                if top:
                    st.markdown("**Top blockers**")
                    st.dataframe(_safe_df(top), use_container_width=True)
                with st.expander("All ledger entries"):
                    st.dataframe(_safe_df(ledger.get("entries") or []), use_container_width=True)
            else:
                st.info("No constraint_ledger found in this artifact.")

            # --- Model set / registry ---
            st.subheader("Model Set")
            model_set = art.get("model_set") or {}
            model_registry = art.get("model_registry") or {}
            if model_set:
                st.caption(f"schema={model_set.get('schema_version','(missing)')}")
                st.json(model_set)
            else:
                st.info("No model_set embedded in this artifact.")
            with st.expander("Model Registry"):
                if model_registry:
                    st.caption(f"schema={model_registry.get('schema_version','(missing)')}")
                    st.json(model_registry)
                else:
                    st.info("No model_registry embedded in this artifact.")

            # --- Standard tables ---
            st.subheader("Standard Tables")
            tables = art.get("tables") or {}
            if isinstance(tables, dict) and tables:
                for k in ["plasma", "power_balance", "tritium"]:
                    if k in tables:
                        st.markdown(f"**{k}**")
                        t = tables.get(k)
                        if isinstance(t, dict):
                            st.dataframe(pd.DataFrame([t]), use_container_width=True)
                        elif isinstance(t, list):
                            st.dataframe(pd.DataFrame(t), use_container_width=True)
                        else:
                            st.json(t)
            else:
                st.info("No tables.v1 section found in this artifact.")

            with st.expander("Full artifact JSON"):
                st.json(art)


# -----------------------------
# Case Deck Runner (new)
# -----------------------------
if _deck == "Control Room":
    with tab_deck:
        st.header("Case Deck Runner")
        st.caption("Run a case_deck.v1 YAML/JSON deck and view the resolved config + artifact outputs.")

        up_deck = st.file_uploader("Upload case_deck.yaml / .json", type=["yaml", "yml", "json"], key="deck_upload")
        out_root = os.path.join(ROOT, "ui_runs")
        os.makedirs(out_root, exist_ok=True)
        out_name = st.text_input("Output folder name (under ui_runs/)", value=f"deck_{int(time.time())}", key="deck_out_name")
        run_btn = st.button("Run Case Deck", key="deck_run")

        if run_btn:
            if up_deck is None:
                st.error("Please upload a case deck file first.")
            else:
                try:
                    deck_path = os.path.join(out_root, f"_uploaded_{up_deck.name}")
                    with open(deck_path, "wb") as f:
                        f.write(up_deck.getvalue())
                    out_dir = os.path.join(out_root, out_name)
                    os.makedirs(out_dir, exist_ok=True)
                    runner = os.path.join(ROOT, "tools", "run_case_deck.py")
                    proc = subprocess.run(
                        [sys.executable, runner, deck_path, "--out", out_dir],
                        cwd=ROOT,
                        capture_output=True,
                        text=True,
                        check=False,
                    )
                    st.code(proc.stdout or "", language="text")
                    if proc.returncode != 0:
                        st.error("Case deck run failed.")
                        st.code(proc.stderr or "", language="text")
                    else:
                        art_path = os.path.join(out_dir, "shams_run_artifact.json")
                        cfg_path = os.path.join(out_dir, "run_config_resolved.json")
                        st.success(f"Wrote outputs to: {out_dir}")
                        if os.path.exists(cfg_path):
                            st.subheader("Resolved config")
                            with open(cfg_path, "r", encoding="utf-8") as f:
                                st.json(json.load(f))
                        if os.path.exists(art_path):
                            st.subheader("Run artifact (preview)")
                            with open(art_path, "r", encoding="utf-8") as f:
                                st.json(json.load(f))
                except Exception as e:
                    st.error(f"{type(e).__name__}: {e}")


# -----------------------------
# Authority & Confidence (v256.0)
# -----------------------------
if _deck == "Control Room":
    with tab_authority_conf:
        st.header("Authority & Confidence")
        st.caption("Trust ledger: authority tiers, maturity tags, and a design-level confidence class. Post-processing only; truth unchanged.")

        colA, colB = st.columns([0.55, 0.45], gap="large")
        with colA:
            st.markdown("### Load artifact")
            up_art = st.file_uploader("shams_run_artifact.json", type=["json"], key="authconf_upload")
            art = _load_json_from_upload(up_art)
            if not art:
                # Fall back to the most recent artifact in session if present.
                art = st.session_state.get("systems_last_solve_artifact") if isinstance(st.session_state.get("systems_last_solve_artifact"), dict) else None
                if not art:
                    art = st.session_state.get("last_point_artifact") if isinstance(st.session_state.get("last_point_artifact"), dict) else None

            if not art:
                st.info("Upload an artifact, or run Point Designer / Systems Mode to populate a last artifact.")
            else:
                ac = art.get("authority_confidence") if isinstance(art, dict) else None
                if not isinstance(ac, dict):
                    st.warning("No authority_confidence found. (Older artifact?)")
                    st.json({"available_keys": sorted(list(art.keys()))[:40]}, expanded=False)
                else:
                    dc = str((ac.get("design") or {}).get("design_confidence_class", "UNKNOWN"))
                    st.markdown(f"**Design confidence class:** `{dc}`")
                    st.caption("Class is a conservative aggregation over implicated subsystems and near-binding hard constraints.")

        with colB:
            st.markdown("### Quick legend")
            st.markdown("- **A**: anchored by authoritative/external contracts (best)")
            st.markdown("- **B**: parametric / semi-authoritative closure")
            st.markdown("- **C**: proxy models or extrapolation-heavy")
            st.markdown("- **D**: speculative / unknown authority")
            st.markdown("- **UNKNOWN**: missing metadata")

        if isinstance(art, dict) and isinstance(art.get("authority_confidence"), dict):
            ac = art["authority_confidence"]
            subs = ac.get("subsystems") or {}
            rows = []
            for k in sorted(list(subs.keys())):
                v = subs.get(k) or {}
                if not isinstance(v, dict):
                    continue
                rows.append({
                    "subsystem": k,
                    "confidence": v.get("confidence_class"),
                    "authority_tier": v.get("authority_tier"),
                    "maturity": v.get("maturity"),
                    "involved": v.get("involved"),
                    "rationale": v.get("rationale"),
                })
            if rows:
                st.subheader("Subsystem trust ledger")
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            else:
                st.info("No subsystem entries available.")


# -----------------------------
# Decision Consequences (v257.0)
# -----------------------------
if _deck == "Control Room":
    with tab_decision_conseq:
        st.header("Decision Consequences")
        st.caption(
            "Advisory governance layer: converts margins + authority into a deterministic 'posture' and risk framing. "
            "Post-processing only; truth unchanged."
        )

        colA, colB = st.columns([0.55, 0.45], gap="large")
        with colA:
            st.markdown("### Load artifact")
            up_art = st.file_uploader("shams_run_artifact.json", type=["json"], key="deccon_upload")
            art = _load_json_from_upload(up_art)
            if not art:
                art = st.session_state.get("systems_last_solve_artifact") if isinstance(st.session_state.get("systems_last_solve_artifact"), dict) else None
                if not art:
                    art = st.session_state.get("last_point_artifact") if isinstance(st.session_state.get("last_point_artifact"), dict) else None

            if not art:
                st.info("Upload an artifact, or run Point Designer / Systems Mode to populate a last artifact.")
            else:
                dc = art.get("decision_consequences") if isinstance(art, dict) else None
                if not isinstance(dc, dict):
                    st.warning("No decision_consequences found. (Older artifact?)")
                    st.json({"available_keys": sorted(list(art.keys()))[:40]}, expanded=False)
                else:
                    st.markdown(f"**Decision posture:** `{str(dc.get('decision_posture','UNKNOWN'))}`")
                    pr = str(dc.get("primary_risk_driver", "") or "")
                    if pr:
                        st.markdown(f"**Primary risk driver:** `{pr}`")
                    wh = dc.get("worst_hard_margin_frac", None)
                    try:
                        wh_s = f"{float(wh):.3f}" if wh is not None else "-"
                    except Exception:
                        wh_s = "-"
                    st.markdown(f"**Worst hard margin (frac):** {wh_s}")
                    st.caption(str(dc.get("narrative", "") or ""))

        with colB:
            st.markdown("### Posture legend")
            st.markdown("- **PROCEED**: feasible with adequate authority")
            st.markdown("- **PROCEED_TARGETED_RD**: feasible but near-binding and/or authority-limited")
            st.markdown("- **HOLD_FOUNDATIONAL**: hard-infeasible; address dominant limiter")
            st.markdown("- **UNKNOWN**: missing/legacy artifact")

        if isinstance(art, dict) and isinstance(art.get("decision_consequences"), dict):
            dc = art["decision_consequences"]
            rows = [
                {"field": "decision_posture", "value": dc.get("decision_posture")},
                {"field": "primary_risk_driver", "value": dc.get("primary_risk_driver")},
                {"field": "dominant_mechanism", "value": dc.get("dominant_mechanism")},
                {"field": "dominant_constraint", "value": dc.get("dominant_constraint")},
                {"field": "worst_hard_margin_frac", "value": dc.get("worst_hard_margin_frac")},
                {"field": "uncertainty_reduction_axis", "value": dc.get("uncertainty_reduction_axis")},
                {"field": "leverage_knobs", "value": dc.get("leverage_knobs")},
                {"field": "stamp_sha256", "value": dc.get("stamp_sha256")},
            ]
            st.subheader("Snapshot")
            st.table(rows)


# ----------------------------------
# Authority Dominance Engine (v330.0)
# ----------------------------------
if _deck == "Control Room":
    with tab_authority_dominance:
        st.header("Authority Dominance")
        st.caption(
            "Deterministic dominance engine: identifies the dominant feasibility killer authority "
            "(PLASMA/EXHAUST/MAGNET/CONTROL/NEUTRONICS/FUEL/PLANT) and ranks the top limiting constraints. "
            "Post-processing only; truth unchanged."
        )

        colA, colB = st.columns([0.55, 0.45], gap="large")
        with colA:
            st.markdown("### Load artifact")
            up_art = st.file_uploader("shams_run_artifact.json", type=["json"], key="authdom_upload")
            art = _load_json_from_upload(up_art)
            if not art:
                art = st.session_state.get("systems_last_solve_artifact") if isinstance(st.session_state.get("systems_last_solve_artifact"), dict) else None
                if not art:
                    art = st.session_state.get("last_point_artifact") if isinstance(st.session_state.get("last_point_artifact"), dict) else None

            if not art:
                st.info("Upload an artifact, or run Point Designer / Systems Mode to populate a last artifact.")
            else:
                ad = art.get("authority_dominance") if isinstance(art, dict) else None
                if not isinstance(ad, dict):
                    st.warning("No authority_dominance found. (Older artifact?)")
                    st.json({"available_keys": sorted(list(art.keys()))[:60]}, expanded=False)
                else:
                    st.markdown(f"**Dominance verdict:** `{str(ad.get('dominance_verdict','UNKNOWN'))}`")
                    st.markdown(f"**Dominant authority:** `{str(ad.get('dominant_authority',''))}`")
                    st.markdown(f"**Dominant constraint:** `{str(ad.get('dominant_constraint',''))}`")
                    mm = ad.get("dominant_margin_frac", None)
                    try:
                        mm_s = f"{float(mm):.4f}" if mm is not None else "-"
                    except Exception:
                        mm_s = "-"
                    st.markdown(f"**Dominant margin (frac):** {mm_s}")
                    st.caption(f"stamp_sha256: {str(ad.get('stamp_sha256',''))[:16]}…")

        with colB:
            st.markdown("### Interpretation")
            st.markdown("- **INFEASIBLE**: at least one hard constraint violated; dominance points to the worst hard margin.")
            st.markdown("- **FRAGILE**: hard-feasible but the tightest hard margin is near-binding (default < 0.05).")
            st.markdown("- **FEASIBLE**: hard-feasible with comfortable margins.")

        if isinstance(art, dict) and isinstance(art.get("authority_dominance"), dict):
            ad = art["authority_dominance"]
            with st.expander("Top limiting constraints (hard)", expanded=False):
                rows = ad.get("dominance_topk") or []
                if isinstance(rows, list) and rows:
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
                else:
                    st.info("No top-k rows available.")

            with st.expander("Authority ranking", expanded=False):
                rows = ad.get("authority_ranked") or []
                if isinstance(rows, list) and rows:
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
                else:
                    st.info("No authority ranking available.")


# -----------------------------
# Scenario Delta Viewer (new)
# -----------------------------

if _deck == "Control Room":
    with tab_epoch_feas:
        st.header("Epoch Feasibility")
        st.caption(
            "Lifecycle-epoch feasibility (Startup / Nominal / End-of-Life). "
            "Constitutional reclassification only; no re-solving and no truth modification."
        )

        colA, colB = st.columns([0.55, 0.45], gap="large")
        with colA:
            st.markdown("### Load artifact")
            up_art = st.file_uploader("shams_run_artifact.json", type=["json"], key="epochfeas_upload")
            art = _load_json_from_upload(up_art)
            if not art:
                art = st.session_state.get("systems_last_solve_artifact") if isinstance(st.session_state.get("systems_last_solve_artifact"), dict) else None
                if not art:
                    art = st.session_state.get("last_point_artifact") if isinstance(st.session_state.get("last_point_artifact"), dict) else None

            if not art:
                st.info("Upload an artifact, or run Systems Mode to populate a last artifact.")
            else:
                ef = art.get("epoch_feasibility") if isinstance(art, dict) else None
                if not isinstance(ef, dict):
                    st.warning("No epoch_feasibility found. (Older artifact?)")
                    st.json({"available_keys": sorted(list(art.keys()))[:40]}, expanded=False)
                else:
                    st.markdown(f"**Overall:** `{str(ef.get('overall','UNKNOWN'))}`")
                    epochs = ef.get("epochs") or []
                    rows = []
                    for e in epochs:
                        if not isinstance(e, dict):
                            continue
                        wh = e.get("worst_hard_margin_frac", None)
                        try:
                            wh_s = f"{float(wh):.3f}" if wh is not None else "-"
                        except Exception:
                            wh_s = "-"
                        rows.append({
                            "epoch": str(e.get("epoch","")),
                            "verdict": str(e.get("verdict","")),
                            "dominant_mechanism": str(e.get("dominant_mechanism","")),
                            "dominant_constraint": str(e.get("dominant_constraint","")),
                            "worst_hard_margin": wh_s,
                            "n_blocking": len(list(e.get("blocking") or [])),
                            "n_diag": len(list(e.get("diagnostic") or [])),
                        })
                    if rows:
                        st.dataframe(rows, use_container_width=True, hide_index=True)
                    else:
                        st.warning("Epoch list empty.")
        with colB:
            st.markdown("### Constitution (selected epoch)")
            if not art or not isinstance(art, dict) or not isinstance(art.get("epoch_feasibility"), dict):
                st.caption("Load an artifact to view epoch constitutions.")
            else:
                ef = art.get("epoch_feasibility") or {}
                epochs = ef.get("epochs") or []
                labels = [str(e.get("epoch","")) for e in epochs if isinstance(e, dict)]
                sel = st.selectbox("Epoch", labels, index=0 if labels else None, key="epochfeas_pick")
                chosen = None
                for e in epochs:
                    if isinstance(e, dict) and str(e.get("epoch","")) == sel:
                        chosen = e
                        break
                if chosen is None:
                    st.info("No epoch selected.")
                else:
                    st.markdown(f"**Epoch:** `{sel}`")
                    st.json(chosen.get("constitution") or {}, expanded=False)
                    st.caption("These clauses reclassify constraint enforcement deterministically across epochs.")

if _deck == "Control Room":
    with tab_delta:
        st.header("Scenario Delta Viewer")
        st.caption("Compare two run artifacts (baseline vs scenario). Uses embedded scenario_delta when available; otherwise computes a transparent diff.")

        col1, col2 = st.columns(2)
        with col1:
            up_base = st.file_uploader("Baseline shams_run_artifact.json", type=["json"], key="delta_base")
        with col2:
            up_scen = st.file_uploader("Scenario shams_run_artifact.json", type=["json"], key="delta_scen")

        base = _load_json_from_upload(up_base)
        scen = _load_json_from_upload(up_scen)

        if not base or not scen:
            st.info("Upload both baseline and scenario artifacts to view deltas.")
        else:
            st.subheader("Embedded scenario_delta")
            sd = scen.get("scenario_delta")
            if sd:
                st.json(sd)
            else:
                st.info("No embedded scenario_delta found; computing diffs from inputs/outputs.")

            st.subheader("Changed inputs")
            bi = base.get("inputs") or {}
            si = scen.get("inputs") or {}
            changed = []
            for k in sorted(set(bi.keys()) | set(si.keys())):
                if bi.get(k) != si.get(k):
                    changed.append({"field": k, "baseline": bi.get(k), "scenario": si.get(k)})
            if changed:
                st.dataframe(pd.DataFrame(changed), use_container_width=True)
            else:
                st.info("No input differences detected.")

            st.subheader("Numeric output deltas")
            bo = base.get("outputs") or {}
            so = scen.get("outputs") or {}
            df = _numeric_delta_table(bo, so)
            if not df.empty:
                st.dataframe(df, use_container_width=True)
            else:
                st.info("No numeric output differences detected.")



            st.subheader("Structural / schema diff (read-only)")
            st.caption("Reports *structure* changes (constraints added/removed/meta changes, model cards) without numeric tolerances.")

            try:
                from shams_io.structural_diff import structural_diff as _structural_diff
                sd = _structural_diff(new_artifact=scen, old_artifact=base)
            except Exception as e:
                sd = None
                st.error(f"Structural diff failed: {e}")

            if isinstance(sd, dict):
                # Constraints changes
                cchg = (sd.get("constraints") or {})
                added = cchg.get("added") or []
                removed = cchg.get("removed") or []
                changed = cchg.get("changed_meta") or []
                cols = st.columns(3)
                cols[0].metric("constraints added", str(len(added)))
                cols[1].metric("constraints removed", str(len(removed)))
                cols[2].metric("constraints meta changed", str(len(changed)))

                if added:
                    with st.expander("Added constraints", expanded=False):
                        st.write(added)
                if removed:
                    with st.expander("Removed constraints", expanded=False):
                        st.write(removed)
                if changed:
                    with st.expander("Changed constraint metadata", expanded=False):
                        st.dataframe(pd.DataFrame(changed), use_container_width=True, hide_index=True)

                # Model cards changes
                mc = (sd.get("model_cards") or {})
                mc_added = mc.get("added") or []
                mc_removed = mc.get("removed") or []
                mc_changed = mc.get("changed") or []
                cols2 = st.columns(3)
                cols2[0].metric("model cards added", str(len(mc_added)))
                cols2[1].metric("model cards removed", str(len(mc_removed)))
                cols2[2].metric("model cards changed", str(len(mc_changed)))
                if mc_added or mc_removed or mc_changed:
                    with st.expander("Model card diffs", expanded=False):
                        st.json({"added": mc_added, "removed": mc_removed, "changed": mc_changed}, expanded=False)

                with st.expander("Raw structural diff JSON (audit)", expanded=False):
                    st.json(sd, expanded=False)


# -----------------------------
# Run Library (Workspace)
# -----------------------------
if _deck == "Control Room":
    with tab_library:
        st.header("Run Library")
        st.caption("Browse a workspace directory of SHAMS run/study artifacts (no physics changes; read-only).")

        def _scan_workspace(root: Path):
            runs = []
            studies = []
            if not root.exists():
                return runs, studies

            # Run artifacts
            for p in root.rglob("*.json"):
                if p.name.lower() in {"shams_run_artifact.json"} or p.name.lower().startswith("case_") or p.name.lower().endswith("_artifact.json"):
                    try:
                        art = read_run_artifact(p)
                        k = art.get("kpis", {}) if isinstance(art, dict) else {}
                        prov = art.get("provenance", {}) if isinstance(art, dict) else {}
                        runs.append({
                            "type": "run",
                            "path": str(p),
                            "created_unix": float(art.get("created_unix", prov.get("created_unix", float("nan")))) if isinstance(art, dict) else float("nan"),
                            "hard_ok": bool(k.get("hard_ok", False)),
                            "hard_worst_margin": k.get("hard_worst_margin", None),
                            "Q": k.get("Q_DT_eqv", k.get("Q", None)),
                            "H98": k.get("H98", None),
                            "message": ((art.get("solver") or {}).get("message") if isinstance(art.get("solver"), dict) else ""),
                        })
                    except Exception:
                        continue

            # Study indexes
            for p in root.rglob("index.json"):
                try:
                    data = json.loads(p.read_text(encoding="utf-8"))
                    if isinstance(data, dict) and data.get("schema_version") == "study_index.v1":
                        prov = data.get("provenance", {}) if isinstance(data.get("provenance"), dict) else {}
                        studies.append({
                            "type": "study",
                            "path": str(p),
                            "created_unix": float(data.get("created_unix", prov.get("created_unix", float('nan')))),
                            "n_cases": int(data.get("n_cases", 0)),
                            "elapsed_s": float(data.get("elapsed_s", float('nan'))),
                        })
                except Exception:
                    continue
            return runs, studies

        default_ws = str((Path.cwd()/ "ui_runs").resolve())
        ws = st.text_input("Workspace folder", value=st.session_state.get("ui_workspace", default_ws))
        st.session_state.ui_workspace = ws
        root = Path(ws)

        colA, colB = st.columns([1, 1])
        with colA:
            do_scan = st.button("Scan workspace", use_container_width=True)
        with colB:
            st.write("")
            st.write("")

        if do_scan:
            runs, studies = _scan_workspace(root)
            st.session_state._ws_runs = runs
            st.session_state._ws_studies = studies

        runs = st.session_state.get("_ws_runs", [])
        studies = st.session_state.get("_ws_studies", [])

        st.subheader("Runs")
        if not runs:
            st.info("No run artifacts found yet. Tip: point runs write artifacts under your chosen output directory; studies write case_XXXX.json under the study out folder.")
        else:
            df = pd.DataFrame(runs)
            # Sort: newest first when available
            if "created_unix" in df.columns:
                df = df.sort_values("created_unix", ascending=False, na_position="last")
            st.dataframe(df, use_container_width=True, hide_index=True)

            sel = st.text_input("Select a run artifact path to open", value=st.session_state.get("selected_artifact_path", ""))
            if st.button("Open selected run", use_container_width=True):
                p = Path(sel)
                if p.exists():
                    try:
                        art = read_run_artifact(p)
                        st.session_state.selected_artifact = art
                        st.session_state.selected_artifact_path = str(p)
                        st.success("Loaded run artifact into session.")
                    except Exception as e:
                        st.error(f"Failed to read artifact: {e}")
                else:
                    st.error("Path does not exist.")

        st.subheader("Studies")
        if studies:
            st.dataframe(pd.DataFrame(studies).sort_values("created_unix", ascending=False, na_position="last"), use_container_width=True, hide_index=True)
            ssel = st.text_input("Select a study index.json path to open", value=st.session_state.get("selected_study_index_path", ""))
            if st.button("Open selected study", use_container_width=True):
                p = Path(ssel)
                if p.exists():
                    try:
                        st.session_state.selected_study_index_path = str(p)
                        st.session_state.selected_study_index = json.loads(p.read_text(encoding="utf-8"))
                        st.success("Loaded study index into session.")
                    except Exception as e:
                        st.error(f"Failed to read study index: {e}")
                else:
                    st.error("Path does not exist.")
        else:
            st.caption("No study indexes found in this workspace.")

# -----------------------------
# Constraint Cockpit
# -----------------------------
if _deck == "Control Room":
    with tab_constraints:
        st.header("Constraint Cockpit")
        st.caption("Interactively triage constraints using the embedded constraint ledger (read-only).")

        art = st.session_state.get("selected_artifact")
        if not isinstance(art, dict):
            st.info("Load a run artifact first (Run Library or Artifacts Explorer).")
        else:
            ledger = art.get("constraint_ledger", {})
            entries = ledger.get("entries", []) if isinstance(ledger, dict) else []
            if not entries:
                st.warning("This artifact has no constraint ledger. (It should be present in v39+ artifacts.)")
            else:
                df = pd.DataFrame(entries)
                # Basic filters
                c1, c2, c3 = st.columns([1,1,1])
                with c1:
                    sev = st.multiselect("Severity", sorted(df.get("severity", pd.Series(["hard"])).dropna().unique().tolist()), default=["hard","soft"] if "soft" in df.get("severity", pd.Series([])).unique() else ["hard"])
                with c2:
                    grp = st.multiselect("Group", sorted(df.get("group", pd.Series(["general"])).dropna().unique().tolist()), default=[])
                with c3:
                    show_only_failed = st.checkbox("Only failed constraints", value=True)

                view = df.copy()
                if sev:
                    view = view[view["severity"].isin(sev)]
                if grp:
                    view = view[view["group"].isin(grp)]
                if show_only_failed and "passed" in view.columns:
                    view = view[view["passed"] == False]

                # Sort: worst first by margin_frac or margin
                if "margin_frac" in view.columns:
                    view = view.sort_values("margin_frac", ascending=True, na_position="last")
                elif "margin" in view.columns:
                    view = view.sort_values("margin", ascending=True, na_position="last")

                st.subheader("Ledger")
                st.dataframe(view, use_container_width=True, hide_index=True)

                st.subheader("Top blockers")
                top = ledger.get("top_blockers", []) if isinstance(ledger, dict) else []
                if top:
                    st.dataframe(pd.DataFrame(top), use_container_width=True, hide_index=True)
                fp = ledger.get("ledger_fingerprint_sha256")
                if fp:
                    st.caption(f"Ledger fingerprint: `{fp}`")


# -----------------------------
# Constraint Inspector (read-only)
# -----------------------------
if _deck == "Control Room":
    with tab_constraint_inspector:
        st.header("Constraint Inspector")
        st.caption("Read-only, equation-first inspection of a single constraint: raw inequality, margin, meaning, knobs, and provenance (when available).")

        art = st.session_state.get("selected_artifact")
        if not isinstance(art, dict):
            st.info("Load a run artifact first (Run Library or Artifacts Explorer).")
        else:
            constraints_list = art.get("constraints") or []
            # Build a name -> constraint dict map (best-effort)
            name_to_c = {}
            for c in constraints_list:
                if isinstance(c, dict) and c.get("name"):
                    name_to_c[str(c.get("name"))] = c

            ledger = art.get("constraint_ledger", {})
            entries = ledger.get("entries", []) if isinstance(ledger, dict) else []
            names = []
            # Prefer ledger order if present (it should reflect evaluation order)
            if entries:
                for e in entries:
                    n = str(e.get("name"))
                    if n and n not in names:
                        names.append(n)
            else:
                names = sorted(list(name_to_c.keys()))

            if not names:
                st.warning("No constraints found in this artifact.")
            else:
                sel = st.selectbox("Select constraint", names, index=0, key="constraint_inspector_select")

                # Pull both ledger entry (if present) and raw constraint dict (if present)
                entry = None
                if entries:
                    for e in entries:
                        if str(e.get("name")) == sel:
                            entry = e
                            break
                c = name_to_c.get(sel, {}) if isinstance(name_to_c.get(sel, {}), dict) else {}

                # Compose a canonical view (prefer ledger fields where available)
                view = {}
                for src in (c, entry or {}):
                    if isinstance(src, dict):
                        view.update({k: src.get(k) for k in src.keys()})

                # Core inequality (verbatim fields; no inferred math)
                sense = str(view.get("sense") or "")
                value = view.get("value")
                limit = view.get("limit")
                units = str(view.get("units") or "")
                meaning = str(view.get("meaning") or view.get("note") or "")

                st.subheader("Inequality")
                if sense and value is not None and limit is not None:
                    st.code(f"{sel}: value {sense} limit    (value={value}, limit={limit}, units={units})", language="text")
                else:
                    st.code(f"{sel}: (insufficient fields to render inequality)", language="text")

                # Pass/fail + margins
                cols = st.columns(4)
                cols[0].metric("passed", str(bool(view.get("passed", False))))
                if view.get("severity") is not None:
                    cols[1].metric("severity", str(view.get("severity")))
                if view.get("group") is not None:
                    cols[2].metric("group", str(view.get("group")))
                if view.get("dominance_rank") is not None:
                    cols[3].metric("dominance_rank", str(view.get("dominance_rank")))

                c1, c2, c3 = st.columns(3)
                if view.get("margin") is not None:
                    c1.metric("margin", f"{view.get('margin')}")
                if view.get("margin_frac") is not None:
                    c2.metric("margin_frac", f"{view.get('margin_frac')}")
                if view.get("violation_score") is not None:
                    c3.metric("violation_score", f"{view.get('violation_score')}")

                st.subheader("Meaning / proxy")
                if meaning.strip():
                    st.write(meaning)
                else:
                    st.info("No meaning/proxy text is attached to this constraint.")

                # Knobs + dominant inputs
                st.subheader("Knobs / dominant inputs (if present)")
                bb = view.get("best_knobs")
                di = view.get("dominant_inputs")
                kcol1, kcol2 = st.columns(2)
                with kcol1:
                    if bb:
                        st.write("**best_knobs**")
                        st.write(bb)
                    else:
                        st.caption("best_knobs: (none)")
                with kcol2:
                    if di:
                        st.write("**dominant_inputs**")
                        st.write(di)
                    else:
                        st.caption("dominant_inputs: (none)")

                # Provenance (constraint-level and artifact-level)
                st.subheader("Provenance (if present)")
                prov = {}
                if isinstance(view.get("provenance"), dict):
                    prov["constraint"] = view.get("provenance")
                if isinstance(art.get("provenance"), dict):
                    prov["artifact"] = art.get("provenance")
                if prov:
                    st.json(prov, expanded=False)
                else:
                    st.info("No provenance keys present on this constraint (artifact-level provenance may still exist under artifact.provenance).")

                # Raw views for auditability
                with st.expander("Raw JSON (audit)", expanded=False):
                    if isinstance(entry, dict):
                        st.write("**constraint_ledger entry**")
                        st.json(entry, expanded=False)
                    if isinstance(c, dict) and c:
                        st.write("**constraints[] item**")
                        st.json(c, expanded=False)


# -----------------------------
# Sensitivity Explorer
# -----------------------------
if _deck == "Control Room":
    with tab_sensitivity:
        st.header("Sensitivity Explorer")
        st.caption("Local finite-difference sensitivities around the current point (no model changes).")

        art = st.session_state.get("selected_artifact")
        if not isinstance(art, dict):
            st.info("Load a run artifact first (Run Library or Artifacts Explorer).")
        else:
            inp_d = art.get("inputs", {})
            if not isinstance(inp_d, dict):
                st.error("Artifact inputs missing or invalid.")
            else:
                try:
                    base = PointInputs.from_dict(inp_d)
                except Exception:
                    # Fallback: try direct constructor with expected keys
                    try:
                        base = PointInputs(**{k: inp_d[k] for k in PointInputs.__dataclass_fields__.keys() if k in inp_d})
                    except Exception as e:
                        st.error(f"Could not build PointInputs from artifact inputs: {e}")
                        base = None

                if base is not None:
                    st.subheader("Base point")
                    st.json(base.__dict__)

                    # Choose knobs + outputs
                    knob_defaults = ["Ip_MA", "fG", "Bt_T", "R0_m", "a_m", "kappa", "Paux_MW", "Ti_keV", "Te_keV"]
                    available_knobs = [k for k in knob_defaults if k in base.__dict__]
                    knobs = st.multiselect("Knobs", available_knobs, default=["Ip_MA", "fG"], key="sens_knobs_v294")

                    outputs_default = [
                        "Q_DT_eqv", "H98", "P_fus_total_MW", "Palpha_MW", "beta_N", "nbar20", "P_e_net_MW",
                        "B_peak_T", "q95", "TBR",
                    ]
                    outputs = st.multiselect("Outputs", outputs_default, default=["Q_DT_eqv", "H98"], key="sens_outs_v294")

                    step_rel = st.number_input("Step size (relative)", value=1e-3, min_value=1e-6, format="%.6f", key="sens_step_rel_v294")

                    if st.button("Compute deterministic sensitivity pack", use_container_width=True, key="sens_btn_v294"):
                        try:
                            from analysis.sensitivity import deterministic_sensitivity_pack
                            # Characteristic scales for variables when x0 == 0
                            scales = {k: 1.0 for k in knobs}
                            scales.update({"Paux_MW": 10.0, "Ip_MA": 1.0, "fG": 0.1, "Bt_T": 0.5, "R0_m": 0.5, "a_m": 0.2})
                            pack = deterministic_sensitivity_pack(base, variables={k: scales.get(k, 1.0) for k in knobs}, outputs=list(outputs), step_rel=float(step_rel))

                            # Flatten for table
                            rows = []
                            jac = pack.get("jacobian", {}) if isinstance(pack, dict) else {}
                            for o in outputs:
                                for p in knobs:
                                    try:
                                        v = float((jac.get(o) or {}).get(p))
                                    except Exception:
                                        v = float('nan')
                                    rows.append({"output": o, "knob": p, "d(output)/d(knob)": v})
                            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

                            st.subheader("Constraint tightness (top residuals)")
                            st.dataframe(pd.DataFrame(pack.get("constraints_tightness", [])), use_container_width=True, hide_index=True)

                            with st.expander("Raw JSON (audit)", expanded=False):
                                st.json(pack)
                        except Exception as e:
                            st.error(f"Sensitivity computation failed: {e}")

# -----------------------------
# Feasibility Map Viewer
# -----------------------------
if _deck == "Control Room":
    with tab_feasmap:
        st.header("Feasibility Map")
        st.caption("Visualize feasibility from study sweeps (heatmap).")

        # Load study index either from session (Run Library) or by path
        p_default = st.session_state.get("selected_study_index_path", "")
        p = st.text_input("Study index.json path", value=p_default)
        idx_data = None
        if p and Path(p).exists():
            try:
                idx_data = json.loads(Path(p).read_text(encoding="utf-8"))
            except Exception as e:
                st.error(f"Could not read study index: {e}")

        if not isinstance(idx_data, dict) or idx_data.get("schema_version") != "study_index.v1":
            st.info("Provide a valid study_out/index.json (schema study_index.v1).")
        else:
            cases = idx_data.get("cases", [])
            study = idx_data.get("study", {})
            sweeps = (study.get("sweeps") if isinstance(study, dict) else None) or []
            # Determine candidate in_ variables for axes
            in_cols = []
            if cases and isinstance(cases, list) and isinstance(cases[0], dict):
                for k in cases[0].keys():
                    if k.startswith("in_"):
                        in_cols.append(k)
            # Prefer sweep variables
            sweep_vars = ["in_"+str(s.get("name")) for s in sweeps if isinstance(s, dict) and s.get("name") is not None]
            axis_candidates = [c for c in sweep_vars if c in in_cols] + [c for c in in_cols if c not in sweep_vars]
            if len(axis_candidates) < 2:
                st.warning("Need at least two swept input variables (in_*) to plot a 2D feasibility map.")
            else:
                c1, c2 = st.columns([1,1])
                with c1:
                    xcol = st.selectbox("X axis", axis_candidates, index=0)
                with c2:
                    ycol = st.selectbox("Y axis", axis_candidates, index=1 if len(axis_candidates)>1 else 0)

                df = pd.DataFrame(cases)
                if "ok" not in df.columns:
                    st.error("Study cases table missing 'ok' field.")
                else:
                    # Build pivot grid
                    xs = sorted(df[xcol].dropna().unique().tolist())
                    ys = sorted(df[ycol].dropna().unique().tolist())
                    import numpy as np
                    grid = np.full((len(ys), len(xs)), np.nan)
                    for _, r in df.iterrows():
                        try:
                            xi = xs.index(r[xcol])
                            yi = ys.index(r[ycol])
                            grid[yi, xi] = 1.0 if bool(r["ok"]) else 0.0
                        except Exception:
                            continue

                    st.subheader("Feasibility heatmap (1=feasible, 0=infeasible)")
                    try:
                        import matplotlib.pyplot as plt  # type: ignore
                        fig, ax = plt.subplots()
                        im = ax.imshow(grid, origin="lower", aspect="auto")
                        ax.set_xticks(range(len(xs)))
                        ax.set_yticks(range(len(ys)))
                        ax.set_xticklabels([str(x) for x in xs], rotation=45, ha="right")
                        ax.set_yticklabels([str(y) for y in ys])
                        ax.set_xlabel(xcol)
                        ax.set_ylabel(ycol)
                        st.pyplot(fig, clear_figure=True)
                    except Exception as e:
                        st.error(f"Plot failed: {e}")

                    st.subheader("Pick a case to open")
                    selx = st.selectbox("X value", xs, index=0)
                    sely = st.selectbox("Y value", ys, index=0)
                    sub = df[(df[xcol]==selx) & (df[ycol]==sely)]
                    if sub.empty:
                        st.info("No case for that cell.")
                    else:
                        st.dataframe(sub[["case","ok","iters","message","path"] + [xcol,ycol]], use_container_width=True, hide_index=True)
                        if st.button("Load this case artifact", use_container_width=True):
                            path = str(sub.iloc[0]["path"])
                            try:
                                art = read_run_artifact(Path(path))
                                st.session_state.selected_artifact = art
                                st.session_state.selected_artifact_path = path
                                st.success("Loaded case artifact into session.")
                            except Exception as e:
                                st.error(f"Could not load case artifact: {e}")


# -----------------------------
# UI Upgrade Pack v53 (UI-only): decision/provenance/knobs/regression/study dashboard/maturity/assumptions/export/solver introspection
# -----------------------------
def _get_active_artifact(label: str = "Use loaded artifact in session") -> dict | None:
    "Return the currently active artifact (from session_state or upload)."
    art = st.session_state.get("selected_artifact")
    if isinstance(art, dict) and art:
        st.info("Using artifact loaded into session (Run Library / Feasibility Map).")
        return art
    up = st.file_uploader("Upload shams_run_artifact.json", type=["json"], key=f"active_artifact_upload_{label}")
    return _load_json_from_upload(up)

def _guess_point_inputs_from_artifact(art: dict) -> PointInputs | None:
    "Best-effort extraction of PointInputs from an artifact. Falls back safely."
    if not isinstance(art, dict):
        return None
    cand = {}
    for k in ["inputs", "point", "point_inputs", "design_point", "config", "run_config", "resolved_config"]:
        v = art.get(k)
        if isinstance(v, dict):
            cand.update(v)
    cand.update({k: art.get(k) for k in ["R0_m","a_m","kappa","Bt_T","B0_T","Ip_MA","Ti_keV","fG","Paux_MW","Ti_over_Te","fuel_mode"] if k in art})
    if "B0_T" in cand and "Bt_T" not in cand:
        cand["Bt_T"] = cand["B0_T"]
    if "Ti_Te" in cand and "Ti_over_Te" not in cand:
        cand["Ti_over_Te"] = cand["Ti_Te"]
    if "Ti/Te" in cand and "Ti_over_Te" not in cand:
        cand["Ti_over_Te"] = cand["Ti/Te"]
    try:
        return _make_point_inputs_safe(**cand)
    except Exception:
        return None

def _decision_summary_from_artifact(art: dict) -> dict:
    kpis = art.get("kpis", {}) if isinstance(art.get("kpis"), dict) else {}
    cons = art.get("constraints", []) if isinstance(art.get("constraints"), list) else []
    ledger = art.get("constraint_ledger", {}) if isinstance(art.get("constraint_ledger"), dict) else {}
    feas = bool(art.get("is_feasible")) if "is_feasible" in art else None
    if feas is None:
        feas = all((not bool(c.get("failed"))) for c in cons) if cons else None
    top = ledger.get("top_blockers") if isinstance(ledger.get("top_blockers"), list) else []
    if not top and cons:
        failed = [c for c in cons if c.get("failed")]
        failed = failed[:8]
        top = [{"name": c.get("name"), "group": c.get("group"), "margin": c.get("margin"), "severity": c.get("severity")} for c in failed]
    return {"feasible": feas, "kpis": kpis, "top_blockers": top, "ledger": ledger, "constraints": cons}

def _download_json_button(label: str, data: dict, fname: str, key: str):
    try:
        st.download_button(label, data=json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8"),
                           file_name=fname, mime="application/json", key=key)
    except Exception as e:
        st.warning(f"Download not available: {e}")

if _deck == "Control Room":
    with tab_decision:
        st.header("Decision Front Page Builder")
        st.caption("UI-native reconstruction of the decision-grade front-page summary from a run artifact (no physics changes).")

        art = _get_active_artifact("decision")
        if not art:
            st.info("Load an artifact to build the decision summary.")
        else:
            d = _decision_summary_from_artifact(art)
            c1, c2, c3 = st.columns([1,1,1])
            with c1:
                st.metric("Feasibility verdict", "FEASIBLE " if d["feasible"] else ("INFEASIBLE " if d["feasible"] is not None else "UNKNOWN"))
            with c2:
                st.metric("Top KPI: Q", f"{d['kpis'].get('Q_DT_eqv', d['kpis'].get('Q', '-'))}")
            with c3:
                st.metric("Top KPI: Pfus (MW)", f"{d['kpis'].get('P_fus_MW', d['kpis'].get('Pfus_MW', '-'))}")

            st.subheader("Dominant blockers")
            if d["top_blockers"]:
                st.dataframe(_safe_df(d["top_blockers"]), use_container_width=True, hide_index=True)
            else:
                st.write("No blockers found in artifact.")

            with st.expander("Full decision inputs (provenance + schema versions)"):
                prov = art.get("provenance", {}) if isinstance(art.get("provenance"), dict) else {}
                st.json({
                    "schema_version": art.get("schema_version"),
                    "repo_version": prov.get("repo_version"),
                    "git_commit": prov.get("git_commit"),
                    "python": prov.get("python"),
                    "platform": prov.get("platform"),
                    "created_unix": prov.get("created_unix"),
                })

            _download_json_button("Download decision summary JSON", d, "decision_summary.json", "dl_decision_summary")


if _deck == "Control Room":
    with tab_nonfeas:
        st.header("Guided Non-Feasibility Mode")
        st.caption("Turn infeasible outcomes into a structured, auditable recovery workflow (UI-only; no physics changes).")

        art = _get_active_artifact("nonfeas")
        if not art:
            st.info("Load an artifact to guide a non-feasibility recovery path.")
        else:
            cons = art.get("constraints", []) if isinstance(art.get("constraints"), list) else []
            kpis = art.get("kpis", {}) if isinstance(art.get("kpis"), dict) else {}

            # Determine hard feasibility
            feasible_hard = None
            if "feasible_hard" in kpis:
                try:
                    feasible_hard = bool(kpis.get("feasible_hard"))
                except Exception:
                    feasible_hard = None
            if feasible_hard is None and cons:
                try:
                    feasible_hard = all(
                        bool(c.get("passed", True))
                        for c in cons
                        if str(c.get("severity", "hard")).lower() == "hard"
                    )
                except Exception:
                    feasible_hard = None

            if feasible_hard is True:
                st.success("This run is hard-feasible. Guided non-feasibility mode is not needed.")
            else:
                # Get or construct a non-feasibility certificate
                cert = art.get("nonfeasibility_certificate") if isinstance(art.get("nonfeasibility_certificate"), dict) else None
                if not cert:
                    hard_failed = [
                        c for c in cons
                        if str(c.get("severity", "hard")).lower() == "hard" and not bool(c.get("passed", True))
                    ]

                    def _mkey(c):
                        try:
                            return float(c.get("margin", 0.0))
                        except Exception:
                            return 0.0

                    hard_failed.sort(key=_mkey)
                    cert = {
                        "hard_feasible": False,
                        "dominant_blockers": [{
                            "name": c.get("name", ""),
                            "group": c.get("group", ""),
                            "value": c.get("value"),
                            "limit": c.get("limit"),
                            "sense": c.get("sense"),
                            "margin": c.get("margin"),
                            "meaning": c.get("meaning", ""),
                            "best_knobs": c.get("best_knobs", []),
                            "maturity": c.get("maturity"),
                            "provenance": c.get("provenance"),
                        } for c in hard_failed[:10]],
                        "recommendation": "Move the listed best_knobs (and/or relax assumptions) until all hard constraints pass.",
                    }

                st.subheader("Non-Feasibility Certificate")
                st.json(cert)

                t1, t2, t3 = st.tabs(["1) Diagnose", "2) Minimal relaxations", "3) Create a scenario (deck)"])

                with t1:
                    st.markdown("### Dominant hard blockers (ranked)")
                    blockers = cert.get("dominant_blockers", []) if isinstance(cert.get("dominant_blockers"), list) else []
                    if blockers:
                        bdf = _safe_df(blockers)
                        pref = [c for c in ["group", "name", "margin", "value", "limit", "sense", "meaning", "best_knobs", "maturity"] if c in bdf.columns]
                        st.dataframe(bdf[pref] if pref else bdf, use_container_width=True, hide_index=True)
                    else:
                        st.warning("No dominant blockers found in certificate.")

                    # Solver hints (if present)
                    out = art.get("outputs", {}) if isinstance(art.get("outputs"), dict) else {}
                    solver = out.get("_solver") if isinstance(out.get("_solver"), dict) else art.get("solver")
                    if isinstance(solver, dict) and solver:
                        st.markdown("### Solver hints (from artifact)")
                        show = {k: solver.get(k) for k in ["status", "reason", "clamped", "clamped_on", "residuals", "ui_log"] if k in solver}
                        st.json(show or solver)

                    st.markdown("### Action principle")
                    st.write(
                        "Fix **hard** blockers first. Soft constraints are advisory unless your decision policy says otherwise. "
                        "Use the knob suggestions as **directional guidance** (not optimization)."
                    )

                with t2:
                    st.markdown("### Propose a nearest-feasible adjustment (within UI)")
                    base = _guess_point_inputs_from_artifact(art)
                    if base is None:
                        base = st.session_state.get("last_point_inp")

                    if base is None:
                        st.warning("Could not infer PointInputs from artifact. Run Point Designer once or ensure artifact includes inputs.")
                    else:
                        st.caption("Choose a dominant blocker, then adjust one or more knobs and re-evaluate.")
                        blockers = cert.get("dominant_blockers", []) if isinstance(cert.get("dominant_blockers"), list) else []
                        if blockers:
                            labels = []
                            for i, b in enumerate(blockers):
                                nm = b.get("name", "") or f"blocker_{i}"
                                mg = b.get("margin")
                                labels.append(f"{i:02d} - {nm} (margin={mg})")
                            bi = st.selectbox("Select blocker", options=list(range(len(blockers))), format_func=lambda i: labels[i], key="nf_blocker_sel")
                            b = blockers[int(bi)]
                            st.markdown("**Suggested knobs (directional):**")
                            st.write(b.get("best_knobs", []) or ["(none provided)"])
                            st.markdown("**Meaning:**")
                            st.write(b.get("meaning", "(no meaning field)"))

                        knob_fields = ["Ip_MA", "fG", "Bt_T", "R0_m", "a_m", "kappa", "Ti_keV", "Paux_MW", "Ti_over_Te"]
                        colA, colB = st.columns([2, 1])
                        with colA:
                            sel_knobs = st.multiselect("Knobs to adjust", options=knob_fields, default=["Ip_MA"], key="nf_knobs")
                        with colB:
                            mode = st.selectbox("Adjustment mode", options=["percent", "absolute"], index=0, key="nf_adj_mode")

                        deltas = {}
                        for k in sel_knobs:
                            v0 = float(getattr(base, k))
                            if mode == "percent":
                                d = st.slider(f"{k} Δ (%)", -50.0, 50.0, 5.0, step=0.5, key=f"nf_d_{k}")
                                deltas[k] = v0 * (1.0 + d / 100.0)
                            else:
                                step = 0.1 if abs < 10 else 1.0
                                d = st.number_input(f"{k} new value", value=v0, step=step, key=f"nf_abs_{k}")
                                deltas[k] = float(d)

                        fuel_mode = st.selectbox("fuel_mode", options=["DT", "DD"], index=0 if getattr(base, "fuel_mode", "DT") == "DT" else 1, key="nf_fuel_mode")

                        run = st.button("Re-evaluate adjusted point", key="nf_run_eval", use_container_width=True)
                        if run:
                            try:
                                d = base.__dict__.copy()
                                d.update({k: float(v) for k, v in deltas.items()})
                                d["fuel_mode"] = str(fuel_mode)
                                pi = PointInputs(**d)

                                out2 = _ui_evaluate(
                                    pi,
                                    origin="run_artifact",
                                    Paux_for_Q_MW=float(getattr(pi, "Paux_MW", 0.0)),
                                )
                                cons2 = evaluate_constraints(out2)
                                art2 = build_run_artifact(
                                    inputs=dict(pi.__dict__),
                                    outputs=dict(out2),
                                    constraints=cons2,
                                    meta={"mode": "guided_nonfeas"},
                                    baseline_inputs=dict(base.__dict__),
                                )
                                st.session_state["nf_last_artifact"] = art2
                                k2 = art2.get("kpis", {}) if isinstance(art2.get("kpis"), dict) else {}
                                st.success(f"Re-evaluated. feasible_hard={k2.get('feasible_hard')}")

                                led = art2.get("constraint_ledger", {}) if isinstance(art2.get("constraint_ledger"), dict) else {}
                                tb = led.get("top_blockers") if isinstance(led.get("top_blockers"), list) else []
                                if tb:
                                    st.subheader("New top blockers")
                                    st.dataframe(_safe_df(tb), use_container_width=True, hide_index=True)

                                with st.expander("New run artifact (raw)"):
                                    st.json(art2)

                                _download_json_button("Download adjusted run artifact", art2, "shams_run_artifact_adjusted.json", "dl_nf_adjusted_artifact")
                            except Exception as e:
                                st.error(f"Re-evaluation failed: {type(e).__name__}: {e}")

                with t3:
                    st.markdown("### Create a scenario deck for reproducible follow-up")
                    base = _guess_point_inputs_from_artifact(art) or st.session_state.get("last_point_inp")
                    last = st.session_state.get("nf_last_artifact")
                    if not isinstance(last, dict):
                        st.info("First run an adjustment in 'Minimal relaxations' to generate a proposed follow-up scenario.")
                    else:
                        try:
                            import yaml  # type: ignore
                        except Exception:
                            yaml = None  # type: ignore

                        new_inputs = last.get("inputs") if isinstance(last.get("inputs"), dict) else {}
                        base_inputs = dict(base.__dict__) if base is not None else (art.get("inputs") if isinstance(art.get("inputs"), dict) else {})

                        delta = {}
                        for k, v in new_inputs.items():
                            if k in base_inputs and base_inputs.get(k) != v:
                                delta[k] = {"from": base_inputs.get(k), "to": v}

                        st.subheader("Scenario delta (inputs)")
                        st.json(delta if delta else {"note": "No input delta detected."})

                        case_deck = {
                            "schema_version": "case_deck.v1",
                            "name": "guided_nonfeas_followup",
                            "base": {},
                            "point": new_inputs,
                            "notes": {
                                "generated_by": "Guided Non-Feasibility Mode",
                                "source_artifact_schema": art.get("schema_version"),
                            },
                        }

                        deck_txt = yaml.safe_dump(case_deck, sort_keys=False) if yaml is not None else json.dumps(case_deck, indent=2)

                        st.markdown("### Case deck")
                        st.code(deck_txt, language="yaml" if yaml is not None else "json")

                        st.download_button(
                            "Download case_deck.yaml",
                            data=deck_txt.encode("utf-8"),
                            file_name="case_deck.yaml",
                            mime="text/yaml" if yaml is not None else "application/json",
                            use_container_width=True,
                        )
                        st.download_button(
                            "Download scenario_delta.json",
                            data=json.dumps(delta, indent=2).encode("utf-8"),
                            file_name="scenario_delta.json",
                            mime="application/json",
                            use_container_width=True,
                        )


if _deck == "Control Room":
    with tab_cprov:
        st.header("Constraint Provenance Drill-Down")
        st.caption("Click into constraints to see definition fields, fingerprints, and maturity/provenance metadata embedded in the artifact.")

        art = _get_active_artifact("cprov")
        if not art:
            st.info("Load an artifact to inspect constraint provenance.")
        else:
            cons = art.get("constraints", [])
            if not isinstance(cons, list) or not cons:
                st.warning("No 'constraints' list found in artifact.")
            else:
                df = _safe_df(cons)
                pref_cols = [c for c in ["group","name","failed","soft_failed","severity","value","limit","margin","margin_frac","units","fingerprint","provenance_fingerprint","maturity"] if c in df.columns]
                st.dataframe(df[pref_cols] if pref_cols else df, use_container_width=True, hide_index=True)

                names = []
                for i,c in enumerate(cons):
                    n = c.get("name") or c.get("id") or f"constraint_{i}"
                    names.append(f"{i:03d} - {n}")
                sel = st.selectbox("Select constraint", options=list(range(len(cons))), format_func=lambda i: names[i], key="cprov_sel")
                c = cons[int(sel)]
                st.subheader("Selected constraint (raw)")
                st.json(c)
                if isinstance(c, dict):
                    st.markdown("**Fingerprint fields**")
                    st.code("\n".join([f"{k}: {c.get(k)}" for k in ["fingerprint","provenance_fingerprint","constraint_fingerprint_sha256"] if k in c] or ["(none found)"]))

if _deck == "Control Room":
    with tab_knobs:
        st.header("Knob Trade-Space Explorer")
        st.caption("Explore a 2-knob trade-space by evaluating a small grid around the active point (no optimization; feasibility-first).")

        art = _get_active_artifact("knobs")
        base = _guess_point_inputs_from_artifact(art) if art else None
        if base is None:
            base = st.session_state.get("last_point_inp")

        if base is None:
            st.info("Load an artifact (or run Point Designer) to initialize a base point.")
        else:
            st.markdown("**Base point (editable)**")
            col1, col2, col3 = st.columns(3)
            with col1:
                R0_m = st.number_input("R0 (m)", value=float(base.R0_m), step=0.01, key="knob_R0")
                a_m = st.number_input("a (m)", value=float(base.a_m), step=0.01, key="knob_a")
                kappa = st.number_input("kappa", value=float(base.kappa), step=0.05, key="knob_kappa")
            with col2:
                Bt_T = st.number_input("Bt (T)", value=float(base.Bt_T), step=0.1, key="knob_Bt")
                Ip_MA = st.number_input("Ip (MA)", value=float(base.Ip_MA), step=0.1, key="knob_Ip")
                fG = st.number_input("fG", value=float(base.fG), step=0.01, key="knob_fG")
            with col3:
                Ti_keV = st.number_input("Ti (keV)", value=float(base.Ti_keV), step=0.5, key="knob_Ti")
                Paux_MW = st.number_input("Paux (MW)", value=float(base.Paux_MW), step=1.0, key="knob_Paux")
                Ti_over_Te = st.number_input("Ti/Te", value=float(getattr(base, "Ti_over_Te", 2.0)), step=0.1, key="knob_TiTe")

            fuel_mode = st.selectbox("fuel_mode", options=["DT","DD"], index=0 if getattr(base, "fuel_mode", "DT")=="DT" else 1, key="knob_fuel")

            knobs = ["Ip_MA","fG","Bt_T","R0_m","Paux_MW","Ti_keV"]
            kx = st.selectbox("Knob X", knobs, index=0, key="knob_kx")
            ky = st.selectbox("Knob Y", knobs, index=1, key="knob_ky")

            def _getv(pi: PointInputs, k: str) -> float:
                return float(getattr(pi, k))
            def _setv(pi: PointInputs, k: str, v: float) -> PointInputs:
                d = pi.__dict__.copy()
                d[k]=float(v)
                return PointInputs(**d)

            x0=_getv(base,kx); y0=_getv(base,ky)
            colA,colB=st.columns(2)
            with colA:
                x_span = st.number_input("X span (+/-)", value=0.1*abs(x0) if abs(x0)>0 else 0.1, step=0.01, key="knob_xspan")
            with colB:
                y_span = st.number_input("Y span (+/-)", value=0.1*abs(y0) if abs(y0)>0 else 0.1, step=0.01, key="knob_yspan")
            nx = st.slider("X grid points", 3, 15, 9, key="knob_nx")
            ny = st.slider("Y grid points", 3, 15, 9, key="knob_ny")
            run = st.button("Evaluate grid", key="knob_run", use_container_width=True)

        if run:
                import numpy as np
                xs = np.linspace(x0-x_span, x0+x_span, int(nx))
                ys = np.linspace(y0-y_span, y0+y_span, int(ny))
                rows=[]
                with st.spinner("Evaluating grid..."):
                    for xv in xs:
                        for yv in ys:
                            pi = PointInputs(R0_m=float(R0_m), a_m=float(a_m), kappa=float(kappa),
                                             Bt_T=float(Bt_T), Ip_MA=float(Ip_MA), Ti_keV=float(Ti_keV),
                                             fG=float(fG), Paux_MW=float(Paux_MW), Ti_over_Te=float(Ti_over_Te),
                                             fuel_mode=str(fuel_mode))
                            pi = _setv(pi, kx, float(xv))
                            pi = _setv(pi, ky, float(yv))
                            try:
                                out = _ui_evaluate(pi, origin="scan_grid")
                                cons = evaluate_constraints(out, point_inputs=pi)
                                ok = all((not bool(c.get("failed"))) for c in cons)
                                top=None
                                if not ok:
                                    failed=[c for c in cons if c.get("failed")]
                                    if failed:
                                        top=failed[0].get("name")
                                rows.append({kx: float(xv), ky: float(yv), "feasible": bool(ok), "top_blocker": top,
                                             "Q": float(out.get("Q_DT_eqv", out.get("Q", float('nan')))),
                                             "Pfus_MW": float(out.get("P_fus_MW", out.get("Pfus_MW", float('nan'))))})
                            except Exception:
                                rows.append({kx: float(xv), ky: float(yv), "feasible": False, "top_blocker": "eval_error", "Q": float('nan'), "Pfus_MW": float('nan')})
                df=pd.DataFrame(rows, columns=["name","failed_A","failed_B","margin_A","margin_B","margin_delta"])
                st.subheader("Grid results (table)")
                st.dataframe(df, use_container_width=True, hide_index=True)

                try:
                    piv = df.pivot(index=ky, columns=kx, values="feasible")
                    st.subheader("Feasibility heatmap (True=1 / False=0)")
                    st.dataframe(piv.astype(int), use_container_width=True)
                except Exception as e:
                    st.warning(f"Could not pivot heatmap: {e}")

if _deck == "Control Room":
    with tab_regress:
        st.header("What broke? Regression Viewer")
        st.caption("Compare two artifacts: constraints, ledgers, model sets, and key KPIs. This is UI-only; it doesn't modify artifacts.")

        c1, c2 = st.columns(2)
        with c1:
            upA = st.file_uploader("Artifact A (json)", type=["json"], key="regA")
            artA = _load_json_from_upload(upA)
        with c2:
            upB = st.file_uploader("Artifact B (json)", type=["json"], key="regB")
            artB = _load_json_from_upload(upB)

        if artA and artB:
            def _kpi_df(art):
                k = art.get("kpis", {}) if isinstance(art.get("kpis"), dict) else {}
                df = pd.DataFrame([{"kpi": kk, "value": vv} for kk,vv in k.items()])
                if df.empty:
                    return pd.DataFrame(columns=["kpi","value"])
                return df.sort_values("kpi")
            st.subheader("KPI diff")
            dfA=_kpi_df(artA).set_index("kpi")
            dfB=_kpi_df(artB).set_index("kpi")
            join=dfA.join(dfB, lsuffix="_A", rsuffix="_B", how="outer")
            join["delta"]=pd.to_numeric(join["value_B"], errors="coerce")-pd.to_numeric(join["value_A"], errors="coerce")
            st.dataframe(join.reset_index().sort_values("kpi"), use_container_width=True, hide_index=True)

            st.subheader("New / worsened constraint failures")
            consA=artA.get("constraints", []) if isinstance(artA.get("constraints"), list) else []
            consB=artB.get("constraints", []) if isinstance(artB.get("constraints"), list) else []
            def _map(cons):
                m={}
                for c in cons:
                    name=c.get("name") or c.get("id")
                    if name:
                        m[name]=c
                return m
            mA=_map(consA); mB=_map(consB)
            names=sorted(set(mA.keys())|set(mB.keys()))
            rows=[]
            for n in names:
                a=mA.get(n,{}); b=mB.get(n,{})
                fa=bool(a.get("failed")); fb=bool(b.get("failed"))
                ma=a.get("margin"); mb=b.get("margin")
                rows.append({"name": n, "failed_A": fa, "failed_B": fb, "margin_A": ma, "margin_B": mb,
                             "margin_delta": (mb-ma) if isinstance(ma,(int,float)) and isinstance(mb,(int,float)) else None})
            df=pd.DataFrame(rows, columns=["name","failed_A","failed_B","margin_A","margin_B","margin_delta"])
            df_bad=df[(df["failed_B"]==True) & ((df["failed_A"]==False) | (df["failed_A"].isna()))]
            st.markdown("**New failures in B**")
            st.dataframe(df_bad.sort_values("name"), use_container_width=True, hide_index=True)
            st.markdown("**Largest margin regressions (B-A)**")
            df_reg=df.dropna(subset=["margin_delta"]).sort_values("margin_delta").head(20)
            st.dataframe(df_reg, use_container_width=True, hide_index=True)

            st.subheader("Model set comparison")
            msA=artA.get("model_set"); msB=artB.get("model_set")
            st.json({"model_set_A": msA, "model_set_B": msB})

if _deck == "Control Room":
    with tab_study_dash:
        st.header("Study Dashboard")
        st.caption("Manager-grade summary for study outputs (feasible fraction, dominant blockers, robustness).")

        up = st.file_uploader("Upload study index.json (study_index.v1)", type=["json"], key="sd_up")
        idx_data = _load_json_from_upload(up)
        if not idx_data:
            idx_path = st.session_state.get("selected_study_path")
            if idx_path and Path(idx_path).exists():
                try:
                    idx_data = json.loads(Path(idx_path).read_text(encoding="utf-8"))
                    st.info("Loaded study index from session.")
                except Exception:
                    idx_data = None

        if idx_data:
            st.subheader("Study headline")
            st.json({k: idx_data.get(k) for k in ["schema_version","n_cases","elapsed_s","created_unix"] if k in idx_data})
            cases = idx_data.get("cases", [])
            if isinstance(cases, list) and cases:
                df = pd.DataFrame(cases)
                if "ok" in df.columns:
                    ok_frac = float(df["ok"].mean())
                    st.metric("Feasible fraction", f"{ok_frac:.3f}")
                for col in ["dominant_blocker","top_blocker","blocker"]:
                    if col in df.columns:
                        st.subheader("Dominant blocker distribution")
                        hist = df[col].fillna("(none)").value_counts().reset_index()
                        hist.columns=[col,"count"]
                        st.dataframe(hist, use_container_width=True, hide_index=True)
                        break
                st.subheader("Cases table")
                st.dataframe(df, use_container_width=True, hide_index=True)
            else:
                st.info("No 'cases' list found in study index. (Older study output?)")

if _deck == "Control Room":
    with tab_maturity:
        st.header("Engineering Maturity Heatmap")
        st.caption("Visualize model maturity / validity info embedded in the artifact (model_set + model_registry).")

        art = _get_active_artifact("maturity")
        if not art:
            st.info("Load an artifact to view maturity info.")
        else:
            reg = art.get("model_registry", {})
            ms = art.get("model_set", {})
            rows=[]
            if isinstance(reg, dict):
                entries = reg.get("entries") if isinstance(reg.get("entries"), list) else None
                if entries is None:
                    if all(isinstance(v, dict) for v in reg.values()):
                        entries=[{"model_id": k, **v} for k,v in reg.items()]
                if entries:
                    selected = set()
                    if isinstance(ms, dict):
                        sel = ms.get("selected")
                        if isinstance(sel, dict):
                            selected = set(sel.values()) | set(sel.keys())
                        elif isinstance(sel, list):
                            selected = set(sel)
                    for e in entries:
                        mid = e.get("model_id", e.get("id", ""))
                        rows.append({
                            "subsystem": e.get("subsystem", e.get("domain", "")),
                            "model_id": mid,
                            "maturity": e.get("maturity", e.get("maturity_tag", "")),
                            "validity": e.get("validity", e.get("validity_range", "")),
                            "selected": (mid in selected)
                        })
            if rows:
                df=pd.DataFrame(rows, columns=["name","failed_A","failed_B","margin_A","margin_B","margin_delta"])
                st.dataframe(df.sort_values(["subsystem","model_id"]), use_container_width=True, hide_index=True)
                st.markdown("Tip: treat this as a policy gate (e.g., block decisions if maturity < required).")
            else:
                st.info("No model_registry entries found in artifact.")


if _deck == "Control Room":
    with tab_maintenance:
        st.header("Maintenance & Availability Authority")
        st.caption("Deterministic maintenance scheduling closure (v368.0): outage calendar proxy and schedule-dominated availability.")

        out = st.session_state.get("last_point_out")
        if not isinstance(out, dict):
            st.info("Run a point in Point Designer first (sets last_point_out).")
        else:
            enabled = bool(out.get("maintenance_contract_sha256")) and (out.get("availability_v368") == out.get("availability_v368"))
            if not enabled:
                st.warning("v368 maintenance scheduling is not enabled for the current point. Enable it in 🧭 Point Designer → Engineering & plant feasibility → 🗓️ Maintenance scheduling authority (v368.0).")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Availability (v368)", _m("availability_v368", "{:.3f}"))
            c2.metric("Outage total (v368)", _m("outage_total_frac_v368", "{:.3f}"))
            c3.metric("Net MWh/y (v368)", _m("net_electric_MWh_per_year_v368", "{:.3g}"))
            c4.metric("Repl. cost (MUSD/y)", _m("replacement_cost_MUSD_per_year_v368", "{:.3g}"))

            with st.expander("What this authority does", expanded=False):
                st.markdown(
                    "- Converts replacement cadences (FW/blanket from v367, plus HCD and tritium plant) and replacement durations into a bundled outage fraction.\n"
                    "- Combines with planned/forced baselines (and optional trips proxy) to form total outage and availability.\n"
                    "- Emits an explicit event table (maintenance_events_v368) for audit and reviewer use."
                )
            with st.expander("What this authority does not do", expanded=False):
                st.markdown(
                    "- Does not run a time-domain availability/RAMI simulation.\n"
                    "- Does not optimize schedules or negotiate constraints.\n"
                    "- Does not modify plasma truth or materials lifetime truth; it only post-processes into a schedule proxy."
                )

            st.subheader("Outage decomposition")
            rows = [
                {"term": "planned", "outage_frac": out.get("planned_outage_frac_v368")},
                {"term": "forced", "outage_frac": out.get("forced_outage_frac_v368")},
                {"term": "replacement", "outage_frac": out.get("replacement_outage_frac_v368")},
                {"term": "total", "outage_frac": out.get("outage_total_frac_v368")},
            ]
            try:
                import pandas as _pd
                df = _pd.DataFrame(rows)
                st.dataframe(df, use_container_width=True, hide_index=True)
            except Exception:
                st.write(rows)

            ev = out.get("maintenance_events_v368")
            with st.expander("Maintenance event table (v368)", expanded=False):
                if isinstance(ev, list) and ev:
                    try:
                        import pandas as _pd
                        st.dataframe(_pd.DataFrame(ev), use_container_width=True, hide_index=True)
                    except Exception:
                        st.json(ev)
                else:
                    st.info("No maintenance_events_v368 found (enable v368 and re-run).")

            with st.expander("Contract fingerprint", expanded=False):
                st.code(str(out.get("maintenance_contract_sha256", "")))


if _deck == "Control Room":
    with tab_profile_auth:
        st.header("Profile Authority")
        st.caption("1.5D algebraic profile diagnostics (non-iterative, conservative).")
        out = st.session_state.get("last_point_out")
        if not isinstance(out, dict):
            st.info("Run a point in Point Designer first (sets last_point_out).")
        else:
            rows=[
                {"metric":"p_peaking", "value": out.get("profile_p_peaking")},
                {"metric":"j_peaking", "value": out.get("profile_j_peaking")},
                {"metric":"li_proxy", "value": out.get("profile_li_proxy")},
                {"metric":"qmin_proxy", "value": out.get("profile_qmin_proxy")},
                {"metric":"f_bootstrap_proxy", "value": out.get("profile_f_bootstrap_proxy")},
                {"metric":"tag", "value": out.get("profile_assumption_tag")},
            ]
            st.dataframe(rows, use_container_width=True, hide_index=True)

            st.markdown("### v399 Multi-species impurity mix (if enabled)")
            rows_v399 = [
                {"metric":"include_impurity_v399", "value": out.get("include_impurity_v399")},
                {"metric":"impurity_v399_mix_json", "value": out.get("impurity_v399_mix_json")},
                {"metric":"impurity_v399_prad_total_MW", "value": out.get("impurity_v399_prad_total_MW")},
                {"metric":"impurity_v399_prad_core_MW", "value": out.get("impurity_v399_prad_core_MW")},
                {"metric":"impurity_v399_prad_edge_MW", "value": out.get("impurity_v399_prad_edge_MW")},
                {"metric":"impurity_v399_prad_sol_MW", "value": out.get("impurity_v399_prad_sol_MW")},
                {"metric":"impurity_v399_prad_div_MW", "value": out.get("impurity_v399_prad_div_MW")},
                {"metric":"impurity_v399_zeff", "value": out.get("impurity_v399_zeff")},
                {"metric":"impurity_v399_fuel_ion_fraction", "value": out.get("impurity_v399_fuel_ion_fraction")},
                {"metric":"detachment_prad_sol_div_achieved_MW_v399", "value": out.get("detachment_prad_sol_div_achieved_MW_v399")},
                {"metric":"detachment_margin_v399", "value": out.get("detachment_margin_v399")},
            ]
            st.dataframe(rows_v399, use_container_width=True, hide_index=True)
            with st.expander("v399 Per-species radiation (MW)", expanded=False):
                st.json(out.get("impurity_v399_by_species_MW", {}))
            with st.expander("v399 Validity flags", expanded=False):
                st.json(out.get("impurity_v399_validity", {}))

            with st.expander("Validity flags", expanded=False):
                st.json(out.get("profile_validity", {}))

if _deck == "Control Room":
    with tab_impurity:
        st.header("Impurity & Radiation")
        st.caption("v320 authority: single-species partitions + detachment inversion. v399: multi-species mix → Zeff + partitions + achieved detachment margin (diagnostic; no truth feedback).")
        out = st.session_state.get("last_point_out")
        if not isinstance(out, dict):
            st.info("Run a point in Point Designer first.")
        else:
            rows=[
                {"metric":"impurity_contract_species", "value": out.get("impurity_contract_species")},
                {"metric":"impurity_contract_f_z", "value": out.get("impurity_contract_f_z")},
                {"metric":"impurity_prad_total_MW", "value": out.get("impurity_prad_total_MW")},
                {"metric":"impurity_prad_core_MW", "value": out.get("impurity_prad_core_MW")},
                {"metric":"impurity_prad_edge_MW", "value": out.get("impurity_prad_edge_MW")},
                {"metric":"impurity_prad_sol_MW", "value": out.get("impurity_prad_sol_MW")},
                {"metric":"impurity_prad_div_MW", "value": out.get("impurity_prad_div_MW")},
                {"metric":"impurity_zeff_proxy", "value": out.get("impurity_zeff_proxy")},
                {"metric":"impurity_fuel_ion_fraction", "value": out.get("impurity_fuel_ion_fraction")},
                {"metric":"detachment_f_sol_div_required", "value": out.get("detachment_f_sol_div_required")},
                {"metric":"detachment_prad_sol_div_required_MW", "value": out.get("detachment_prad_sol_div_required_MW")},
                {"metric":"detachment_f_z_required", "value": out.get("detachment_f_z_required")},
            ]
            st.dataframe(rows, use_container_width=True, hide_index=True)
            with st.expander("Validity flags", expanded=False):
                st.json(out.get("impurity_validity", {}))

if _deck == "Control Room":
    with tab_disruption:
        st.header("Disruption Risk")
        st.caption("Conservative screening tier: LOW/MED/HIGH (diagnostic; not predictive).")
        out = st.session_state.get("last_point_out")
        if not isinstance(out, dict):
            st.info("Run a point in Point Designer first.")
        else:
            st.metric("Tier", str(out.get("disruption_risk_tier", "UNKNOWN")))
            cols=st.columns(3)
            with cols[0]:
                st.metric("Risk index", f"{float(out.get('disruption_risk_index', float('nan'))):.3f}" if out.get('disruption_risk_index')==out.get('disruption_risk_index') else "nan")
            with cols[1]:
                st.metric("Dominant driver", str(out.get("disruption_dominant_driver", "unknown")))
            with cols[2]:
                st.metric("fG", f"{float(getattr(st.session_state.get('last_point_inp', None),'fG', float('nan'))):.3f}" if st.session_state.get('last_point_inp') is not None else "nan")
            with st.expander("Components", expanded=False):
                st.json(out.get("disruption_risk_components", {}))

if _deck == "Control Room":
    with tab_stability:
        st.header("Stability Risk")
        st.caption("Conservative screening tier: LOW/MED/HIGH for vertical stability + RWM/control budgets (diagnostic; not predictive).")
        out = st.session_state.get("last_point_out")
        if not isinstance(out, dict):
            st.info("Run a point in Point Designer first.")
        else:
            st.metric("Tier", str(out.get("stability_risk_tier", "UNKNOWN")))
            cols = st.columns(4)
            with cols[0]:
                st.metric(
                    "Risk index",
                    f"{float(out.get('stability_risk_index', float('nan'))):.3f}"
                    if out.get("stability_risk_index") == out.get("stability_risk_index")
                    else "nan",
                )
            with cols[1]:
                st.metric("Dominant driver", str(out.get("stability_dominant_driver", "unknown")))
            with cols[2]:
                st.metric(
                    "vs_margin",
                    f"{float(out.get('vs_margin', float('nan'))):.3f}"
                    if out.get("vs_margin") == out.get("vs_margin")
                    else "nan",
                )
            with cols[3]:
                st.metric("RWM ok", "yes" if bool(out.get("rwm_control_ok", True)) else "no")

            st.divider()
            oc = st.columns(2)
            with oc[0]:
                st.metric("Operational tier", str(out.get("operational_risk_tier", "UNKNOWN")))
            with oc[1]:
                st.metric("Operational driver", str(out.get("operational_dominant_driver", "")) or "-")

            with st.expander("Components", expanded=False):
                st.json(out.get("stability_risk_components", {}))
            with st.expander("Control contract margins", expanded=False):
                st.json(out.get("control_contract_margins", {}))

if _deck == "Control Room":
    with tab_cert_search:
        st.header("Certified Search")
        st.caption("Budgeted multi-knob search (external to truth). Each candidate is verified by the frozen evaluator.")

        from dataclasses import replace
        from solvers.budgeted_search import SearchVar
        from solvers.certified_search_orchestrator import (
            OrchestratorSpec,
            SearchStage,
            ParetoObjective,
            run_orchestrated_certified_search,
            run_orchestrated_certified_pareto_search,
        )

        base = st.session_state.get("last_point_inp")
        if base is None:
            st.info("Run a point in Point Designer first so a base point exists.")
        else:
            st.subheader("Knobs")
            knob_options = [
                ("Bt_T", 2.0, 25.0),
                ("Ip_MA", 1.0, 25.0),
                ("Paux_MW", 0.0, 200.0),
                ("Ti_keV", 1.0, 40.0),
                ("fG", 0.2, 1.2),
                ("kappa", 1.0, 2.6),
                ("a_m", 0.2, 3.0),
                ("R0_m", 0.8, 12.0),
            ]
            cols = st.columns(3)
            with cols[0]:
                chosen = st.multiselect(
                    "Select up to 4 knobs",
                    [k[0] for k in knob_options],
                    default=["Bt_T", "Ip_MA"],
                    max_selections=4,
                )
            with cols[1]:
                mode = st.selectbox(
                    "Mode",
                    ["Single objective (v340 compat)", "Pareto frontier (v405)",],
                    index=0,
                    key="cs_mode",
                )
            with cols[2]:
                objective = st.selectbox(
                    "Score objective (PASS-only)",
                    ["Q_DT_eqv", "P_fus_MW", "P_net_MW"],
                    index=0,
                    key="cs_single_obj",
                    disabled=(str(mode) != "Single objective (v340 compat)"),
                )

            pareto_objectives = []
            if str(mode) == "Pareto frontier (v405)":
                st.subheader("Pareto objectives")
                # Deterministic, compact objective menu
                obj_menu = [
                    ("R0_m", "min"),
                    ("B_peak_T", "min"),
                    ("P_e_net_MW", "max"),
                    ("q_div_MW_m2", "min"),
                    ("sigma_vm_MPa", "min"),
                    ("TBR", "max"),
                    ("Q_DT_eqv", "max"),
                ]
                ocols = st.columns(3)
                with ocols[0]:
                    o1 = st.selectbox("Objective #1", [o[0] for o in obj_menu], index=0, key="cs_p_obj1")
                with ocols[1]:
                    o2 = st.selectbox("Objective #2", [o[0] for o in obj_menu], index=2, key="cs_p_obj2")
                with ocols[2]:
                    o3 = st.selectbox("Objective #3 (optional)", ["(none)"] + [o[0] for o in obj_menu], index=0, key="cs_p_obj3")
                senses = {k: s for k, s in obj_menu}
                for ok in [o1, o2] + ([o3] if str(o3) != "(none)" else []):
                    pareto_objectives.append(ParetoObjective(key=str(ok), sense=str(senses.get(str(ok), "min"))))

                cpm = st.columns(2)
                with cpm[0]:
                    max_frontier = int(st.number_input("Max frontier points", value=30, min_value=5, max_value=200, step=5, key="cs_p_maxfront"))
                with cpm[1]:
                    filter_mirage = bool(st.checkbox("Filter mirage (lane)", value=True, key="cs_p_filter_mirage"))

            vars_=[]
            for name,lo,hi in knob_options:
                if name in chosen:
                    c1,c2=st.columns(2)
                    with c1:
                        lo_v = st.number_input(f"{name} lo", value=float(getattr(base,name)), step=0.1, key=f"cs_lo_{name}")
                    with c2:
                        hi_v = st.number_input(f"{name} hi", value=float(getattr(base,name)), step=0.1, key=f"cs_hi_{name}")
                    if hi_v <= lo_v:
                        hi_v = lo_v + 1e-6
                    vars_.append(SearchVar(name=name, lo=float(lo_v), hi=float(hi_v)))

            c1,c2,c3,c4=st.columns(4)
            with c1:
                budget = int(st.number_input("Budget", value=96, min_value=8, max_value=2048, step=8, key="cs_budget"))
            with c2:
                seed = int(st.number_input("Seed", value=0, min_value=0, max_value=10_000, step=1, key="cs_seed"))
            with c3:
                method = st.selectbox("Method", ["halton","lhs","grid"], index=0, key="cs_method")
            with c4:
                two_stage = bool(st.checkbox("Two-stage refine", value=True, key="cs_two_stage"))

            stage2_budget_frac = float(st.slider("Stage-2 budget fraction", min_value=0.10, max_value=0.80, value=0.35, step=0.05, key="cs_stage2_frac")) if two_stage else 0.0
            stage2_shrink = float(st.slider("Stage-2 local shrink", min_value=0.10, max_value=0.80, value=0.35, step=0.05, key="cs_stage2_shrink")) if two_stage else 0.0
            stage2_method = st.selectbox("Stage-2 method", ["grid","halton","lhs"], index=0, key="cs_stage2_method") if two_stage else "grid"

            st.markdown("---")
            insert_surr = bool(st.checkbox("Insert surrogate stage (feasible-first, non-authoritative)", value=False, key="cs_insert_surr"))
            surr_frac = float(
                st.slider(
                    "Surrogate budget fraction",
                    min_value=0.05,
                    max_value=0.60,
                    value=0.20,
                    step=0.05,
                    key="cs_surr_frac",
                    disabled=(not insert_surr),
                )
            )
            s1, s2, s3 = st.columns(3)
            with s1:
                surr_pool_mult = int(st.number_input("Surrogate pool multiplier", value=50, min_value=4, max_value=200, step=1, key="cs_surr_pool", disabled=(not insert_surr)))
            with s2:
                surr_kappa = float(st.slider("Surrogate kappa", min_value=0.0, max_value=2.0, value=0.5, step=0.1, key="cs_surr_kappa", disabled=(not insert_surr)))
            with s3:
                surr_ridge = float(st.number_input("Surrogate ridge alpha", value=1e-3, min_value=1e-6, max_value=1.0, format="%.6f", key="cs_surr_ridge", disabled=(not insert_surr)))

            def _builder(b, overrides):
                return replace(b, **{k: float(v) for k,v in overrides.items()})

            def _verifier(inp_obj):
                out = _ui_evaluate(inp_obj, origin="certified_search_verifier")
                cons = evaluate_constraints(out, point_inputs=inp_obj)
                try:
                    from constraints.bookkeeping import summarize as _summarize_constraints
                    _cs = _summarize_constraints(cons)
                    _min_margin_frac = float(_cs.worst_hard_margin_frac) if _cs.worst_hard_margin_frac is not None else float("nan")
                    _worst_hard = str(_cs.worst_hard or "")
                except Exception:
                    _min_margin_frac = float("nan")
                    _worst_hard = ""
                ok = all((not bool(c.get("failed"))) for c in cons)
                score = float(out.get(objective, 0.0)) if ok else float("-inf")

                evidence={
                    "objective": objective,
                    "objective_value": float(out.get(objective, float("nan"))),
                    "min_margin_frac": _min_margin_frac,
                    "worst_hard": _worst_hard,
                    "worst_hard_margin_frac": float(_min_margin_frac) if _min_margin_frac == _min_margin_frac else float("nan"),
                    "n_failed": int(sum(1 for c in cons if c.get("failed"))),
                    "top_blocker": (next((c.get("name") for c in cons if c.get("failed")), None)),
                }
                return ("PASS" if ok else "FAIL"), score, evidence

            if st.button("Run certified search", use_container_width=True, key="run_cert_search"):
                if not vars_:
                    st.warning("Select at least one knob.")
                else:
                    b1 = int(max(1, round(float(budget) * (1.0 - float(stage2_budget_frac)))))
                    b2 = int(max(0, round(float(budget) * float(stage2_budget_frac))))
                    bs = int(max(0, round(float(budget) * float(surr_frac)))) if insert_surr else 0
                    # cap budgets deterministically
                    b2 = int(min(int(b2), int(max(0, budget - 1))))
                    bs = int(min(int(bs), int(max(0, budget - 1 - b2))))
                    b1 = int(max(1, int(budget) - int(b2) - int(bs)))

                    stages = [SearchStage(name="stage1", method=str(method), budget=int(b1), seed=int(seed), local_refine=False)]
                    if insert_surr and bs > 0:
                        stages.append(
                            SearchStage(
                                name="surrogate",
                                method="surrogate",
                                budget=int(bs),
                                seed=int(seed + 1),
                                local_refine=False,
                                surrogate_pool_mult=int(surr_pool_mult),
                                surrogate_kappa=float(surr_kappa),
                                surrogate_ridge_alpha=float(surr_ridge),
                                surrogate_feas_margin_key="min_margin_frac",
                            )
                        )
                    if two_stage and b2 > 0:
                        stages.append(
                            SearchStage(
                                name="stage2",
                                method=str(stage2_method),
                                budget=int(b2),
                                seed=int(seed + (2 if (insert_surr and bs > 0) else 1)),
                                local_refine=True,
                                local_shrink=float(stage2_shrink),
                            )
                        )
                    if str(mode) == "Pareto frontier (v405)":
                        def _eval_fn(inp_obj):
                            return _ui_evaluate(inp_obj, origin="pareto_frontier_v405")

                        def _cons_fn(out_obj, inp_obj):
                            return evaluate_constraints(out_obj, point_inputs=inp_obj)

                        art = run_orchestrated_certified_pareto_search(
                            base_inputs=base,
                            spec=OrchestratorSpec(variables=tuple(vars_), stages=tuple(stages)),
                            objectives=list(pareto_objectives) if pareto_objectives else [ParetoObjective(key="R0_m", sense="min")],
                            builder=_builder,
                            evaluator_fn=_eval_fn,
                            constraints_fn=_cons_fn,
                            max_frontier=int(max_frontier),
                            filter_mirage=bool(filter_mirage),
                        )
                    else:
                        art = run_orchestrated_certified_search(
                            base,
                            OrchestratorSpec(variables=tuple(vars_), stages=tuple(stages)),
                            verifier=_verifier,
                            builder=_builder,
                        )
                    st.session_state["last_certified_search_artifact"] = art
                    st.session_state["v340_cert_search_last"] = art
                    try:
                        _v98_record_run("certified_search_orchestrated", art, mode="SystemSuite/Chronicle")
                    except Exception:
                        pass

                    n_pass = 0
                    n_tot = 0
                    try:
                        for stg in art.get("stages", []):
                            recs = stg.get("records", [])
                            n_tot += len(recs)
                            n_pass += sum(1 for r in recs if r.get("verdict") == "PASS")
                    except Exception:
                        pass
                    st.success(f"Done. Digest: {str(art.get('digest',''))[:12]} | PASS found: {n_pass}/{n_tot}")

            art = st.session_state.get("v340_cert_search_last")
            if isinstance(art, dict) and art.get("schema_version"):
                st.subheader("Results")
                # Flatten across stages for display
                rows = []
                for stg in art.get("stages", []):
                    for r in stg.get("records", []):
                        rows.append({"stage": stg.get("name"), "i": r.get("i"), "verdict": r.get("verdict"), "score": r.get("score"), **(r.get("x") or {}), **{f"e_{k}": v for k, v in (r.get("evidence") or {}).items()}})
                df = pd.DataFrame(rows)
                with st.expander("Results table", expanded=False):
                    st.dataframe(df, use_container_width=True, hide_index=True)
                if isinstance(art.get("best"), dict) and art["best"].get("x") is not None:
                    with st.expander("Best PASS candidate", expanded=False):
                        st.json(art.get("best"))

                # v405: frontier candidates (Pareto) with per-candidate evidence packs
                cands = art.get("candidates")
                if isinstance(cands, list) and len(cands) > 0:
                    st.subheader("Frontier candidates (v405)")
                    rows2 = []
                    for c in cands:
                        if not isinstance(c, dict):
                            continue
                        objm = c.get("objectives") or {}
                        row = {
                            "id": str(c.get("id", "")),
                            "lane_robust": str(c.get("lane_robust_verdict", "")),
                            "lane_opt": str(c.get("lane_optimistic_verdict", "")),
                            "is_mirage": bool(c.get("is_mirage_lane", False)),
                            "global_min_margin_v402": c.get("global_min_margin_v402"),
                            "dominant_authority": str(c.get("global_dominant_authority_v402", "")),
                            **(c.get("x") or {}),
                        }
                        if isinstance(objm, dict):
                            for k, v in objm.items():
                                row[f"obj_{k}"] = v
                        rows2.append(row)
                    df2 = pd.DataFrame(rows2)
                    with st.expander("Frontier table", expanded=False):
                        st.dataframe(df2, use_container_width=True, hide_index=True)

                    # Evidence pack per candidate
                    try:
                        from tools.frontier_candidate_evidence_zip import build_frontier_candidate_evidence_zip_bytes

                        cand_ids = [str(r.get("id")) for r in cands if isinstance(r, dict) and r.get("id")]
                        sel = st.selectbox("Select candidate for evidence pack", cand_ids, index=0, key="cs_p_sel") if cand_ids else None
                        if sel and st.button("Build selected candidate evidence pack", use_container_width=True, key="cs_p_build"):
                            b = build_frontier_candidate_evidence_zip_bytes(
                                orchestrator_artifact=art,
                                candidate_id=str(sel),
                                basename=f"frontier_candidate_{str(sel)}",
                            )
                            st.session_state["cs_p_candidate_zip"] = b
                            st.success("Candidate evidence pack built.")
                        b = st.session_state.get("cs_p_candidate_zip")
                        if isinstance(b, (bytes, bytearray)) and len(b) > 0:
                            st.download_button(
                                "Download frontier_candidate_evidence.zip",
                                data=b,
                                file_name="frontier_candidate_evidence.zip",
                                mime="application/zip",
                                use_container_width=True,
                                key="cs_p_dl",
                            )
                    except Exception:
                        pass

                # v297: export deterministic evidence pack
                try:
                    from tools.simple_evidence_zip import build_simple_evidence_zip_bytes
                    art2 = st.session_state.get("last_certified_search_artifact")
                    if isinstance(art2, dict):
                        if st.button("Build Certified Search evidence pack", use_container_width=True, key="cs_build_ev"):
                            b = build_simple_evidence_zip_bytes(art2, basename=f"certified_search_{art2.get('digest','')[:12]}")
                            st.session_state["certified_search_evidence_zip"] = b
                            st.success("Evidence pack built.")
                        b = st.session_state.get("certified_search_evidence_zip")
                        if isinstance(b, (bytes, bytearray)) and len(b) > 0:
                            st.download_button(
                                "Download certified_search_evidence.zip",
                                data=b,
                                file_name="certified_search_evidence.zip",
                                mime="application/zip",
                                use_container_width=True,
                                key="cs_dl_ev",
                            )
                except Exception:
                    pass

if _deck == "Control Room":
    with tab_repair:
        st.header("Repair Suggestions")
        st.caption("Explanatory-only: proposes bounded knob deltas to reduce dominant constraint residuals; every proposal must be verified by truth.")

        from dataclasses import replace
        from solvers.repair_suggestions import RepairKnob, propose_repair_candidates

        base_inp = st.session_state.get("last_point_inp")
        base_out = st.session_state.get("last_point_out")
        if base_inp is None or not isinstance(base_out, dict):
            st.info("Run a point in Point Designer first.")
        else:
            cons = evaluate_constraints(base_out, point_inputs=base_inp)
            failed = [c for c in cons if c.get("failed")]
            if not failed:
                st.info("Base point is already feasible; repair suggestions are not needed.")
            else:
                # Residual proxy: use 'margin' if present else 1.0
                residuals = {}
                for c in failed:
                    name = str(c.get("name", "(unnamed)"))
                    m = c.get("margin")
                    try:
                        m = float(m)
                    except Exception:
                        m = None
                    # Convert to positive residual where 0 is pass.
                    if m is None or not (m == m):
                        residuals[name] = 1.0
                    else:
                        residuals[name] = max(0.0, -m)

                st.subheader("Select knobs")
                knob_options = [
                    ("Bt_T", 2.0, 25.0),
                    ("Ip_MA", 1.0, 25.0),
                    ("Paux_MW", 0.0, 200.0),
                    ("Ti_keV", 1.0, 40.0),
                    ("fG", 0.2, 1.2),
                    ("kappa", 1.0, 2.6),
                ]
                chosen = st.multiselect("Knobs used for repair", [k[0] for k in knob_options], default=["Bt_T","Ip_MA","Paux_MW"], max_selections=6)
                knobs=[]
                for name,lo,hi in knob_options:
                    if name in chosen:
                        lo_v=float(st.number_input(f"{name} lo", value=float(getattr(base_inp,name)), step=0.1, key=f"rep_lo_{name}"))
                        hi_v=float(st.number_input(f"{name} hi", value=float(getattr(base_inp,name)), step=0.1, key=f"rep_hi_{name}"))
                        if hi_v <= lo_v:
                            hi_v = lo_v + 1e-6
                        knobs.append(RepairKnob(name=name, lo=lo_v, hi=hi_v))

                # Finite-difference jacobian (deterministic): d(residual)/dvar
                def _eval_res(inp_obj):
                    out = _ui_evaluate(inp_obj, origin="jacobian_fd")
                    cons2 = evaluate_constraints(out, point_inputs=inp_obj)
                    res={}
                    for c in cons2:
                        name=str(c.get("name","(unnamed)"))
                        m=c.get("margin")
                        try:
                            m=float(m)
                        except Exception:
                            continue
                        res[name]=max(0.0, -m)
                    return res

                if st.button("Compute repair candidates", use_container_width=True, key="run_repairs"):
                    base_res = _eval_res(base_inp)
                    jac={}
                    for c in base_res:
                        jac[c]={}
                    for kb in knobs:
                        span = kb.hi - kb.lo
                        h = 0.02*span
                        if h<=0: continue
                        x0=float(getattr(base_inp,kb.name))
                        x1=min(kb.hi, x0+h)
                        x2=max(kb.lo, x0-h)
                        # central difference when possible
                        inp_p = replace(base_inp, **{kb.name: x1})
                        inp_m = replace(base_inp, **{kb.name: x2})
                        rp=_eval_res(inp_p)
                        rm=_eval_res(inp_m)
                        denom = (x1-x2) if (x1-x2)!=0 else 1e-12
                        for c in base_res:
                            jac[c][kb.name] = (float(rp.get(c,0.0))-float(rm.get(c,0.0)))/denom

                    cands = propose_repair_candidates(residuals=base_res, jacobian=jac, knobs=knobs, k=8)
                    st.session_state["v296_repair_last"] = {"base_res": base_res, "jac": jac, "cands": cands}

                last = st.session_state.get("v296_repair_last")
                if last:
                    with st.expander("Base residuals", expanded=False):
                        st.dataframe(pd.DataFrame([{ "constraint": k, "residual": v } for k,v in sorted(last["base_res"].items(), key=lambda kv: kv[1], reverse=True)]), use_container_width=True, hide_index=True)
                    st.subheader("Candidates")
                    cands = last["cands"]
                    if cands:
                        df = pd.DataFrame([{ "rationale": c.rationale, "est_reduction": c.estimated_residual_reduction, **c.deltas } for c in cands])
                        st.dataframe(df, use_container_width=True, hide_index=True)
                    else:
                        st.info("No candidates produced (check knob selections).")

                    # v297: build + download a deterministic repair evidence pack
                    try:
                        from dataclasses import asdict
                        from tools.simple_evidence_zip import build_simple_evidence_zip_bytes

                        repair_art = {
                            "schema_version": "repair_evidence.v1",
                            "base_inputs": asdict(base_inp),
                            "base_failed_constraints": [str(c.get("name")) for c in failed][:50],
                            "base_residuals": dict(last.get("base_res", {})),
                            "knobs": [asdict(k) for k in knobs],
                            "candidates": [
                                {
                                    "rationale": getattr(c, "rationale", ""),
                                    "estimated_residual_reduction": float(getattr(c, "estimated_residual_reduction", 0.0)),
                                    "deltas": dict(getattr(c, "deltas", {})),
                                }
                                for c in (cands or [])
                            ],
                        }
                        st.session_state["last_repair_evidence_artifact"] = repair_art

                        if st.button("Build Repair evidence pack", use_container_width=True, key="rep_build_ev"):
                            b = build_simple_evidence_zip_bytes(repair_art, basename="repair_evidence")
                            st.session_state["repair_evidence_zip"] = b
                            _v98_record_run("repair_evidence", repair_art, mode="SystemSuite/Chronicle")
                            st.success("Repair evidence pack built.")

                        b = st.session_state.get("repair_evidence_zip")
                        if isinstance(b, (bytes, bytearray)) and len(b) > 0:
                            st.download_button(
                                "Download repair_evidence.zip",
                                data=b,
                                file_name="repair_evidence.zip",
                                mime="application/zip",
                                use_container_width=True,
                                key="rep_dl_ev",
                            )
                    except Exception:
                        pass

if _deck == "Control Room":
    with tab_refine:
        st.header("Interval Refinement")
        st.caption("Deterministic corner evaluation + contract refinement suggestions (explanatory-only).")

        from dataclasses import replace
        from uq_contracts.refinement import suggest_interval_refinements

        base = st.session_state.get("last_point_inp")
        if base is None:
            st.info("Run a point in Point Designer first so a base point exists.")
        else:
            st.subheader("Select up to 3 uncertain variables")
            var_options = ["Bt_T","Ip_MA","Paux_MW","Ti_keV","fG","kappa"]
            chosen = st.multiselect("Variables", var_options, default=["fG"], max_selections=3, key="ref_vars")
            intervals={}
            for v in chosen:
                c1,c2=st.columns(2)
                with c1:
                    lo = float(st.number_input(f"{v} lo", value=float(getattr(base,v))*0.95, step=0.1, key=f"ref_lo_{v}"))
                with c2:
                    hi = float(st.number_input(f"{v} hi", value=float(getattr(base,v))*1.05 + (0.01 if v=='fG' else 0.0), step=0.1, key=f"ref_hi_{v}"))
                if hi <= lo:
                    hi = lo + 1e-6
                intervals[v]=(lo,hi)

            if st.button("Evaluate corners", use_container_width=True, key="run_ref_corners"):
                # build corners
                vs=list(intervals.items())
                corners=[]
                def rec(i,cur):
                    if i==len(vs):
                        corners.append(dict(cur)); return
                    name,(lo,hi)=vs[i]
                    cur[name]=lo; rec(i+1,cur)
                    cur[name]=hi; rec(i+1,cur)
                    cur.pop(name,None)
                rec(0,{})

                results=[]
                for c in corners:
                    inp_obj=base
                    inp_obj=replace(inp_obj, **{k: float(v) for k,v in c.items()})
                    out = _ui_evaluate(inp_obj, origin="corner_probe")
                    cons = evaluate_constraints(out, point_inputs=inp_obj)
                    ok = all((not bool(cc.get("failed"))) for cc in cons)
                    dom = next((cc.get("name") for cc in cons if cc.get("failed")), None)
                    results.append({"corner": c, "verdict": "PASS" if ok else "FAIL", "dominant_mechanism": dom})

                st.session_state["v296_ref_corners"] = {"intervals": intervals, "results": results}

            last = st.session_state.get("v296_ref_corners")
            if last:
                res = last["results"]
                df = pd.DataFrame([{**r["corner"], "verdict": r["verdict"], "dominant": r.get("dominant_mechanism")} for r in res])
                st.dataframe(df, use_container_width=True, hide_index=True)
                fails=sum(1 for r in res if r["verdict"]!='PASS')
                st.metric("FAIL corners", f"{fails}/{len(res)}")
                sugg = suggest_interval_refinements(last["intervals"], res)
                if sugg:
                    st.subheader("Refinement suggestions")
                    sdf = pd.DataFrame([{"var": s.var, "current": s.current_interval, "suggested": s.suggested_interval, "rationale": s.rationale} for s in sugg])
                    st.dataframe(sdf, use_container_width=True, hide_index=True)
                else:
                    st.info("No refinement suggestions (either robust already or insufficient failure signal).")

                # v297: interval refinement evidence pack
                try:
                    from dataclasses import asdict
                    from tools.simple_evidence_zip import build_simple_evidence_zip_bytes

                    refine_art = {
                        "schema_version": "interval_refinement_evidence.v1",
                        "base_inputs": asdict(base),
                        "intervals": {k: list(v) for k, v in (last.get("intervals") or {}).items()},
                        "corner_results": list(res),
                        "suggestions": [
                            {
                                "var": s.var,
                                "current_interval": list(s.current_interval),
                                "suggested_interval": list(s.suggested_interval),
                                "rationale": s.rationale,
                            }
                            for s in (sugg or [])
                        ],
                    }
                    st.session_state["last_interval_refinement_artifact"] = refine_art

                    if st.button("Build Interval Refinement evidence pack", use_container_width=True, key="ref_build_ev"):
                        b = build_simple_evidence_zip_bytes(refine_art, basename="interval_refinement")
                        st.session_state["interval_refinement_zip"] = b
                        _v98_record_run("interval_refinement", refine_art, mode="SystemSuite/Chronicle")
                        st.success("Interval refinement evidence pack built.")

                    b = st.session_state.get("interval_refinement_zip")
                    if isinstance(b, (bytes, bytearray)) and len(b) > 0:
                        st.download_button(
                            "Download interval_refinement_evidence.zip",
                            data=b,
                            file_name="interval_refinement_evidence.zip",
                            mime="application/zip",
                            use_container_width=True,
                            key="ref_dl_ev",
                        )
                except Exception:
                    pass

if _deck == "Control Room":
    with tab_narrowing:
        st.header("Interval Narrowing")
        st.caption(
            "Advisory dead-region flags + interval narrowing proposals + repair contract export (no truth mutation)."
        )
        try:
            from ui.interval_narrowing import render_interval_narrowing_panel
            render_interval_narrowing_panel(st, pd, BASE_DIR, st.session_state)
        except Exception as _e:
            st.info("Panel unavailable in this build.")

if _deck == "Control Room":
    with tab_surrogate:
        st.header("Surrogate Overlay")
        st.caption("Non-authoritative ridge surrogate fitted to the latest Certified Search results.")

        from optimization.surrogates import fit_ridge_surrogate, predict_surrogate

        res = st.session_state.get("v296_cert_search_last")
        if res is None:
            st.info("Run Certified Search first (Chronicle → Certified Search) to generate training data.")
        else:
            # train on PASS points only
            rows=[]
            for r in res.records:
                if r.verdict == "PASS":
                    rows.append({"x": r.x, "y": r.score})
            if len(rows) < 8:
                st.warning(f"Need at least 8 PASS samples to fit a surrogate; currently have {len(rows)}.")
            else:
                feat = list(rows[0]["x"].keys())
                samples=[rr["x"] for rr in rows]
                targets=[rr["y"] for rr in rows]
                ridge=float(st.number_input("Ridge", value=1e-6, format="%e"))
                if st.button("Fit surrogate", use_container_width=True, key="fit_surr"):
                    model = fit_ridge_surrogate(samples, targets, feat, ridge=ridge)
                    st.session_state["v296_surrogate_model"] = model
                    st.success("Surrogate fitted (non-authoritative).")

                model = st.session_state.get("v296_surrogate_model")
                if model is not None:
                    st.subheader("Query")
                    q={}
                    for f in model.feature_names:
                        q[f]=float(st.number_input(f"{f}", value=float(st.session_state.get('last_point_inp').__dict__.get(f, 0.0)), step=0.1, key=f"surr_q_{f}"))
                    yhat, unc = predict_surrogate(model, q)
                    st.metric("Predicted score", f"{yhat:.6g}")
                    st.metric("Uncertainty proxy", f"{unc:.3f}")
                    with st.expander("Model details", expanded=False):
                        st.write({"features": list(model.feature_names), "ridge": model.ridge})

if _deck == "Control Room":
    with tab_active_learning:
        st.header("Active Learning")
        st.caption("Propose new points where surrogate uncertainty is high (non-authoritative).")

        from optimization.active_learning import ALVar, propose_active_learning_points

        model = st.session_state.get("v296_surrogate_model")
        res = st.session_state.get("v296_cert_search_last")
        if model is None or res is None:
            st.info("Fit a surrogate first (Chronicle → Surrogate Overlay).")
        else:
            vars_=[]
            # derive bounds from SearchSpec
            for v in res.spec.variables:
                vars_.append(ALVar(name=v.name, lo=v.lo, hi=v.hi))
            c1,c2,c3=st.columns(3)
            with c1:
                n_candidates=int(st.number_input("Candidates", value=512, min_value=64, max_value=5000, step=64))
            with c2:
                n_select=int(st.number_input("Select", value=16, min_value=4, max_value=128, step=4))
            with c3:
                seed=int(st.number_input("Seed", value=int(res.spec.seed), min_value=0, max_value=10000, step=1))

            if st.button("Propose points", use_container_width=True, key="run_al"):
                props = propose_active_learning_points(model, vars_, n_candidates=n_candidates, n_select=n_select, seed=seed)
                st.session_state["v296_al_props"] = props

            props = st.session_state.get("v296_al_props")
            if props:
                df = pd.DataFrame([{**p.x, "y_pred": p.y_pred, "uncertainty": p.uncertainty} for p in props])
                st.dataframe(df, use_container_width=True, hide_index=True)
                st.caption("Verify proposals by setting last_point_inp to a row and running Point Designer. External harness integration can automate that next.")

if _deck == "Control Room":
    with tab_assumptions:
        st.header("Assumption Toggle Bar")
        st.caption("Fast scenario exploration by toggling common assumptions and re-evaluating the point (still feasibility-first; no optimization).")

        art = _get_active_artifact("assumptions")
        base = _guess_point_inputs_from_artifact(art) if art else None
        if base is None:
            base = st.session_state.get("last_point_inp")
        if base is None:
            st.info("Load an artifact (or run Point Designer) to use assumption toggles.")
        else:
            col1,col2,col3=st.columns(3)
            with col1:
                fuel = st.selectbox("Fuel mode", ["DT","DD"], index=0 if getattr(base,"fuel_mode","DT")=="DT" else 1, key="ass_fuel")
            with col2:
                ti = st.number_input("Ti (keV)", value=float(base.Ti_keV), step=0.5, key="ass_Ti")
            with col3:
                paux = st.number_input("Paux (MW)", value=float(base.Paux_MW), step=1.0, key="ass_Paux")
            tite = st.number_input("Ti/Te", value=float(getattr(base,"Ti_over_Te", 2.0)), step=0.1, key="ass_TiTe")
            apply = st.button("Apply toggles and evaluate", use_container_width=True, key="ass_run")

            if apply:
                pi = PointInputs(R0_m=float(base.R0_m), a_m=float(base.a_m), kappa=float(base.kappa),
                                 Bt_T=float(base.Bt_T), Ip_MA=float(base.Ip_MA), Ti_keV=float(ti),
                                 fG=float(base.fG), Paux_MW=float(paux), Ti_over_Te=float(tite),
                                 fuel_mode=str(fuel))
                out = _ui_evaluate(pi, origin="systems_point")
                cons = evaluate_constraints(out, point_inputs=pi)
                ok = all((not bool(c.get("failed"))) for c in cons)
                st.metric("Feasible", "YES " if ok else "NO ")
                st.subheader("Key outputs")
                st.json({k: out.get(k) for k in ["Q_DT_eqv","P_fus_MW","P_net_MW","betaN","q95","fG"] if k in out})
                st.subheader("Top failed constraints")
                failed=[c for c in cons if c.get("failed")]
                if failed:
                    st.dataframe(_safe_df(failed[:10]), use_container_width=True, hide_index=True)
                else:
                    st.write("No failed constraints.")

if _deck == "Control Room":
    with tab_export:
        st.header("Export / Communication Panel")
        st.caption("One-click export helpers (JSON, CSV, and a one-slide PNG-style summary) with provenance footer.")

        art = _get_active_artifact("export")
        if not art:
            st.info("Load an artifact to export.")
        else:
            _download_json_button("Download run artifact JSON", art, "shams_run_artifact.json", "dl_artifact")
            tables = art.get("tables", {}) if isinstance(art.get("tables"), dict) else {}
            if tables:
                for name, obj in tables.items():
                    try:
                        df = _safe_df(obj)
                        st.download_button(f"Download {name}.csv", data=df.to_csv(index=False).encode("utf-8"),
                                           file_name=f"{name}.csv", mime="text/csv", key=f"dl_csv_{name}")
                    except Exception:
                        continue
            else:
                st.info("No standardized tables found in artifact ('tables').")

            try:
                import io
                import matplotlib.pyplot as plt
                prov = art.get("provenance", {}) if isinstance(art.get("provenance"), dict) else {}
                d = _decision_summary_from_artifact(art)
                fig = plt.figure(figsize=(10, 5))
                ax = fig.add_subplot(111)
                ax.axis("off")
                title = "SHAMS Decision Summary"
                verdict = "FEASIBLE" if d["feasible"] else "INFEASIBLE"
                ax.text(0.02, 0.92, f"{title} - {verdict}", fontsize=16, weight="bold")
                ax.text(0.02, 0.82, f"Q: {d['kpis'].get('Q_DT_eqv', d['kpis'].get('Q','-'))}    Pfus(MW): {d['kpis'].get('P_fus_MW', d['kpis'].get('Pfus_MW','-'))}", fontsize=12)
                ax.text(0.02, 0.72, "Top blockers:", fontsize=12, weight="bold")
                y=0.66
                for b in (d["top_blockers"] or [])[:6]:
                    ax.text(0.04, y, f"- {b.get('group','')}: {b.get('name','')}", fontsize=11)
                    y -= 0.06
                footer = f"repo_version={prov.get('repo_version')}  git={prov.get('git_commit')}  python={prov.get('python')}"
                ax.text(0.02, 0.03, footer, fontsize=9)
                buf = io.BytesIO()
                fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
                plt.close(fig)
                st.download_button("Download one-slide summary PNG", data=buf.getvalue(), file_name="shams_one_slide.png",
                                   mime="image/png", key="dl_png_slide")
            except Exception as e:
                st.warning(f"PNG summary unavailable: {e}")

if _deck == "Control Room":
    with tab_solver:
        st.header("Solver Introspection")
        st.caption("Inspect solver trace/clamp/residual info from artifacts or the last Point Designer run.")

        art = st.session_state.get("selected_artifact")
        if not isinstance(art, dict) or not art:
            st.info("No session artifact loaded. Upload one below to inspect solver annotations.")
            up = st.file_uploader("Upload shams_run_artifact.json", type=["json"], key="solver_up")
            art = _load_json_from_upload(up)

        if art:
            trace = art.get("solver_trace") if isinstance(art.get("solver_trace"), dict) else None
            if trace:
                st.subheader("solver_trace (artifact)")
                st.json(trace)
            else:
                st.subheader("Solver annotations (best-effort)")
                flat = {}
                for k,v in art.items():
                    if isinstance(k,str) and (k.startswith("_solver") or k.startswith("_H98") or k.startswith("_Q")):
                        flat[k]=v
                kpis = art.get("kpis", {}) if isinstance(art.get("kpis"), dict) else {}
                for k,v in kpis.items():
                    if isinstance(k,str) and (k.startswith("_solver") or k.startswith("_H98") or k.startswith("_Q")):
                        flat[f"kpis.{k}"]=v
                if flat:
                    st.json(flat)
                else:
                    st.info("No solver trace fields found in artifact.")

            st.divider()
            st.subheader("CCFS verifier (external solver firewall)")
            st.caption("Verify an external candidate bundle against frozen truth. Runs do not modify SHAMS physics.")
            up_ccfs = st.file_uploader("Upload ccfs_bundle.json", type=["json"], key="ccfs_up_v294")
            b = _load_json_from_upload(up_ccfs)
            if isinstance(b, dict):
                req_cols = st.columns(2)
                with req_cols[0]:
                    do_phase = st.checkbox("Require phase envelope PASS", value=True, key="ccfs_req_phase_v294")
                with req_cols[1]:
                    do_uq = st.checkbox("Require UQ contract not FAIL", value=False, key="ccfs_req_uq_v294")

                if st.button("Verify CCFS bundle", use_container_width=True, key="ccfs_btn_v294"):
                    try:
                        from extopt.certified_solve import verify_ccfs_bundle
                        res = verify_ccfs_bundle(b, default_request={"phase_envelope": bool(do_phase), "uq_contracts": bool(do_uq)})
                        st.success("CCFS verification complete.")
                        st.json(res, expanded=False)
                        _v98_record_run("ccfs_verify", res, mode="ControlRoom/Diagnostics")
                    except Exception as e:
                        st.error(f"CCFS verification failed: {e}")


# --- Copyright notice
st.markdown('---')
st.caption('© 2026 Afshin Arjhangmehr - SHAMS–FUSION-X')


# =====================
# v88 UI helpers (append-only)
# =====================
def _v88_continuation_explorer(path):
    import pandas as pd, streamlit as st
    if not path:
        st.info("No continuation path available.")
        return
    df = pd.DataFrame(path)
    st.line_chart(df.select_dtypes("number"))

def _v88_boundary_map(points):
    import pandas as pd, streamlit as st
    if not points:
        st.info("No scan points available.")
        return
    df = pd.DataFrame(points)
    if "x" in df.columns and "min_signed_margin" in df.columns:
        st.scatter_chart(df, x="x", y="min_signed_margin", color="feasible")

def _v88_export_svg(fig, name):
    import io, streamlit as st
    buf = io.BytesIO()
    fig.savefig(buf, format="svg", bbox_inches="tight")
    st.download_button(
        f"Download {name}.svg",
        data=buf.getvalue(),
        file_name=f"{name}.svg",
        mime="image/svg+xml",
    )


# =====================
# v89 UI helpers (append-only)
# =====================
def _v89_margin_waterfall_fig(constraints):
    import matplotlib.pyplot as plt
    rows = []
    for r in constraints or []:
        name = r.get("name","")
        sm = r.get("signed_margin", None)
        if name and isinstance(sm, (int,float)):
            rows.append((name, float(sm)))
    if not rows:
        return None
    rows.sort(key=lambda x: x[1])
    names = [n for n,_ in rows][:40]
    vals = [v for _,v in rows][:40]
    fig, ax = plt.subplots()
    ax.barh(names, vals)
    ax.axvline(0.0)
    ax.set_xlabel("signed_margin")
    ax.set_title("Constraint Margin Waterfall")
    fig.tight_layout()
    return fig

def _v89_boundary_heatmap(points, xcol, ycol):
    import numpy as np
    import matplotlib.pyplot as plt
    xs = [p.get(xcol) for p in points]
    ys = [p.get(ycol) for p in points]
    xs = [x for x in xs if isinstance(x,(int,float))]
    ys = [y for y in ys if isinstance(y,(int,float))]
    if not xs or not ys:
        return None
    x_min, x_max = min(xs), max(xs)
    y_min, y_max = min(ys), max(ys)
    nx, ny = 40, 40
    grid = np.full((ny, nx), np.nan)
    counts = np.zeros((ny, nx), dtype=int)
    hits = np.zeros((ny, nx), dtype=int)
    for p in points:
        x = p.get(xcol); y = p.get(ycol)
        if not isinstance(x,(int,float)) or not isinstance(y,(int,float)):
            continue
        ix = int((x - x_min) / (x_max - x_min + 1e-12) * (nx-1))
        iy = int((y - y_min) / (y_max - y_min + 1e-12) * (ny-1))
        counts[iy, ix] += 1
        hits[iy, ix] += 1 if p.get("feasible", False) else 0
    mask = counts > 0
    grid[mask] = hits[mask] / counts[mask]
    fig, ax = plt.subplots()
    im = ax.imshow(grid, origin="lower", extent=[x_min,x_max,y_min,y_max], aspect="auto")
    ax.set_xlabel(xcol)
    ax.set_ylabel(ycol)
    ax.set_title("Feasibility Boundary Map (binned feasibility fraction)")
    fig.colorbar(im, ax=ax, label="feasible fraction")
    fig.tight_layout()
    return fig

def _v89_constraint_intersections(points, topn=10):
    from collections import Counter
    c = Counter()
    for p in points:
        if p.get("feasible", False):
            continue
        acts = p.get("active_constraints", [])
        if not isinstance(acts, list):
            continue
        acts = [a for a in acts if isinstance(a,str) and a][:6]
        for i in range(len(acts)):
            for j in range(i+1, len(acts)):
                pair = tuple(sorted((acts[i], acts[j])))
                c[pair] += 1
    return c.most_common(topn)


# =====================
# v89.1 UI state persistence helpers (append-only)
# =====================
def _v89_1_render_cached_point():
    import streamlit as st
    if "pd_last_outputs" not in st.session_state or "pd_last_artifact" not in st.session_state:
        st.info("No cached Point Designer result yet. Click 'Evaluate Point' to compute one.")
        return
    st.info("Showing last Point Designer results (cached). Downloads should not clear results.")

    out = st.session_state["pd_last_outputs"]
    artifact = st.session_state["pd_last_artifact"]

    # KPI summary (best effort)
    try:
        with st.expander("Point summary (cached)", expanded=False):
            kpis = headline_kpis(out)
            for i in range(0, len(kpis), 4):
                kpi_row(kpis[i:i+4])
    except Exception:
        pass

    # Exports (cached bytes if available)
    with st.expander("Exports (external systems codes-style artifacts) - cached", expanded=False):
        try:
            st.download_button(
                "Download run artifact JSON",
                data=_shams_json_dumps(artifact, indent=2, sort_keys=True),
                file_name="shams_run_artifact.json",
                mime="application/json",
                use_container_width=True,
                key="pd_cached_artifact_json",
            )
        except Exception:
            pass

        # Radial build PNG (prefer cached bytes)
        state = st.session_state.get('shams_state', None)
        rb = (
            state.last_point_radial_png
            if state and getattr(state, 'last_point_radial_png', None) is not None
            else st.session_state.get("pd_last_radial_png_bytes", None)
        )
        if isinstance(rb, (bytes, bytearray)) and len(rb) > 0:
            st.download_button(
                "Download radial build PNG",
                data=rb,
                file_name="shams_radial_build.png",
                mime="image/png",
                use_container_width=True,
                key="pd_cached_radial_png",
            )
        else:
            # Generate on demand (still should not clear results since out/artifact is cached)
            try:
                tmpdir = tempfile.mkdtemp(prefix="shams_export_")
                radial_path = os.path.join(tmpdir, "radial_build.png")
                plot_radial_build_from_artifact(artifact, radial_path)
                with open(radial_path, "rb") as f:
                    st.download_button(
                        "Download radial build PNG",
                        data=f.read(),
                        file_name="shams_radial_build.png",
                        mime="image/png",
                        use_container_width=True,
                        key="pd_cached_radial_png_gen",
                    )
            except Exception:
                st.caption("Radial-build export unavailable for this point.")


# =====================
# v89.2 UI state persistence + controls (append-only)
# =====================
def _v89_2_point_cache_ui():
    import streamlit as st
    st.subheader("Cached Point Designer Result")
    c1, c2 = st.columns([1,3])
    with c1:
        if st.button("Clear cached result", key="pd_clear_cache"):
            for k in ["pd_last_outputs", "pd_last_artifact", "pd_last_radial_png_bytes"]:
                if k in st.session_state:
                    del st.session_state[k]
            st.success("Cleared.")
            st.stop()
    with c2:
        st.caption("Cached results persist across reruns (e.g., after downloads).")

    state = st.session_state.get('shams_state', None)
    art = (state.last_point_artifact if state and getattr(state,'last_point_artifact',None) is not None else st.session_state.get("pd_last_artifact", None))
    out = (state.last_point_outputs if state and getattr(state,'last_point_outputs',None) is not None else st.session_state.get("pd_last_outputs", None))
    if art is None or out is None:
        st.info("No cached result yet. Run Point Designer once to populate this.")
        return

    rb = (state.last_point_radial_png if state and getattr(state,'last_point_radial_png',None) is not None else st.session_state.get("pd_last_radial_png_bytes", None))
    if isinstance(rb, (bytes, bytearray)) and len(rb) > 0:
        st.image(rb, caption="Radial build (cached export preview)", use_container_width=True)

    import json as _json
    st.download_button(
        "Download run artifact JSON",
        data=_json.dumps(art, indent=2, sort_keys=True),
        file_name="shams_run_artifact.json",
        mime="application/json",
        use_container_width=True,
        key="pd_dl_artifact_cached",
    )
    if isinstance(rb, (bytes, bytearray)) and len(rb) > 0:
        st.download_button(
            "Download radial build PNG",
            data=rb,
            file_name="shams_radial_build.png",
            mime="image/png",
            use_container_width=True,
            key="pd_dl_radial_cached",
        )



# =====================
# v92 UI state-machine helpers (append-only)
# =====================
# (Defined early in Phase-1 stabilization block to avoid forward-reference failures.)

# v93 UI helpers (append-only)
# =====================
# -----------------------------
# JSON safety helpers (cycle-safe) - required for export/download stability
# -----------------------------
def _v93_validate_before_download(obj, schema_path: str):
    """Return (ok, errors). Uses jsonschema if available, else fallback."""
    import json as _json
    from pathlib import Path
    try:
        sch = _json.loads(Path(schema_path).read_text(encoding="utf-8"))
        try:
            import jsonschema  # type: ignore
            jsonschema.validate(instance=obj, schema=sch)
            return True, []
        except ImportError:
            from tools.validate_schemas import _fallback_validate
            errs = _fallback_validate(obj, sch)
            return (len(errs) == 0), errs
    except Exception as e:
        return False, [repr(e)]

def _v93_paper_figures_pack_panel():
    import streamlit as st
    s = _v92_state_get()
    st.subheader("Paper Figures Pack")
    if not s.has_point():
        st.info("Run Point Designer first; pack is built from cached Point artifact.")
        return
    try:
        from tools.paper_figures_pack import build_figures_pack_bytes
        if st.button("Build paper figures pack", key="v93_build_figpack"):
            s.last_figures_pack_zip = build_figures_pack_bytes(s.last_point_artifact)
            st.success("Figures pack built.")
        b = getattr(s, "last_figures_pack_zip", None)
        if isinstance(b, (bytes, bytearray)) and len(b) > 0:
            st.download_button(
                "Download paper_figures_pack.zip",
                data=b,
                file_name="paper_figures_pack.zip",
                mime="application/zip",
                use_container_width=True,
                key="v93_dl_figpack",
            )
    except Exception:
        st.caption("Figures pack unavailable.")


def _v93_stateful_scan_panel():
    import streamlit as st
    s = _v92_state_get()
    st.subheader("Scan Ledger (cached artifact)")
    if s.last_scan_points is None:
        st.info("No cached scan yet.")
        return
    scan_obj = {"kind":"shams_feasible_set","meta": s.last_scan_meta or {}, "points": s.last_scan_points}
    ok, errs = _v93_validate_before_download(scan_obj, "schemas/shams_feasible_set.schema.json")
    if ok: st.success("Schema: PASS")
    else:
        st.warning("Schema: FAIL")
        for e in errs[:10]: st.write("- " + str(e))
    st.download_button("Download feasible_scan.json (stateful)",
                       data=_json.dumps(scan_obj, indent=2, sort_keys=True),
                       file_name="feasible_scan.json", mime="application/json",
                       use_container_width=True, key="v93_dl_scan_stateful")

def _v93_stateful_systems_panel():
    import streamlit as st
    s = _v92_state_get()
    st.subheader("Run Ledger (cached artifacts)")
    if s.last_systems_result is None:
        if s.has_point():
            st.info("No cached systems yet (no successful Systems solve). Point artifact is available - run a Systems solve to generate a Systems artifact.")
        else:
            st.info("No cached systems yet.")
        return
    ok, errs = _v93_validate_before_download(s.last_systems_result, "schemas/shams_run_artifact.schema.json")
    if ok: st.success("Schema: PASS")
    else:
        st.warning("Schema: FAIL")
        for e in errs[:10]: st.write("- " + str(e))
    st.download_button("Download systems_artifact.json (stateful)",
                       data=_shams_json_dumps(s.last_systems_result, indent=2, sort_keys=True),
                       file_name="systems_artifact.json", mime="application/json",
                       use_container_width=True, key="v93_dl_systems_stateful")


def _v182_render_latest_systems_solve_results(*, artifact: dict, point_artifact: dict | None = None, key_prefix: str = "v182_latest"):
    """Render the latest Systems *solve* results from a cached run artifact.

    This MUST be safe across Streamlit reruns (e.g., download_button triggers rerun).
    Therefore it takes a fully self-contained artifact and does not rely on locals.
    """
    import streamlit as st
    import json

    # Schema hardening: upgrade + validate (non-fatal; never blocks UI)
    try:
        try:
            from src.systems.schema import upgrade_systems_artifact, validate_systems_artifact
        except Exception:  # pragma: no cover
            from systems.schema import upgrade_systems_artifact, validate_systems_artifact
        artifact = upgrade_systems_artifact(artifact or {})
        _issues = validate_systems_artifact(artifact)
        if _issues:
            with st.expander("Schema warnings (non-fatal)", expanded=False):
                for msg in _issues:
                    st.warning(msg)
    except Exception:
        pass
    try:
        outs = (artifact or {}).get('headline') or (artifact or {}).get('outputs') or {}
    except Exception:
        outs = {}

    st.markdown("### Latest Systems solve results")
    # Unique suffix so this function can be called multiple times in the same Streamlit run
    try:
        _rid = (artifact or {}).get("rid") or (artifact or {}).get("run_id") or (artifact or {}).get("metadata", {}).get("rid")
    except Exception:
        _rid = None
    try:
        _suffix = str(_rid) if _rid else __import__("hashlib").sha1(__import__("json").dumps(artifact, sort_keys=True).encode("utf-8")).hexdigest()[:10]
    except Exception:
        _suffix = "na"
    _k_art = f"{key_prefix}_dl_systems_artifact_{_suffix}"
    _k_zip = f"{key_prefix}_dl_systems_bundle_{_suffix}"

    # Always-available downloads (do not depend on a button-click run path)
    st.download_button(
        "Download systems-mode run artifact JSON",
        data=_shams_json_dumps(artifact, indent=2, sort_keys=True),
        file_name="shams_run_artifact_systems.json",
        mime="application/json",
        use_container_width=True,
        key=_k_art,
    )

    # Optional export bundle (safe; uses cached artifacts)
    try:
        from tools.export.bundle import build_export_bundle_bytes
        bundle_bytes = build_export_bundle_bytes(
            repo_root=BASE_DIR,
            point_artifact=point_artifact,
            systems_artifact=artifact,
            scan_artifact=None,
            pareto_artifact=None,
            opt_artifact=None,
            feasible_search_artifact=None,
            include_readme=True,
        )
        st.download_button(
            "Download export bundle ZIP",
            data=bundle_bytes,
            file_name="shams_export_bundle_systems.zip",
            mime="application/zip",
            use_container_width=True,
            key=_k_zip,
        )
    except Exception as _e:
        st.caption(f"Export bundle unavailable: {_e}")

    # Key results
    kcols = st.columns(4)
    def _k(col, key, fmt="{:.3g}"):
        try:
            v = float(outs.get(key, float('nan')))
        except Exception:
            v = float('nan')
        with col:
            st.metric(key, fmt.format(v) if v == v else "NaN")

    _k(kcols[0], "Q_DT_eqv")
    _k(kcols[1], "H98")
    _k(kcols[2], "P_e_net_MW")
    _k(kcols[3], "q_div_MW_m2")

    # Constraints table (if present)
    try:
        rows = (artifact or {}).get('constraints')
        if isinstance(rows, list) and rows:
            with st.expander("Constraints & margins (cached)", expanded=False):
                import pandas as pd
                st.dataframe(pd.DataFrame(rows), use_container_width=True)
    except Exception:
        pass

def _v184_render_latest_feasible_search_results(*, report: dict, key_prefix: str = "v184_fs_latest") -> None:
    """Rerun-safe renderer for Feasible Design Search results.

    The search is easy to 'lose' in Streamlit after a button click because the app reruns and
    the workflow step may change. This renderer provides a stable, cached view.
    """
    import streamlit as st
    try:
        import pandas as pd
    except Exception:
        pd = None  # type: ignore

    if not isinstance(report, dict) or not report:
        st.info("No cached feasible-search results yet.")
        return

    ok = bool(report.get('ok'))
    reason = str(report.get('reason', ''))
    obj = str(report.get('objective', ''))
    ts_unix = report.get('ts_unix')
    try:
        ts_str = datetime.datetime.fromtimestamp(float(ts_unix)).strftime('%Y-%m-%d %H:%M:%S') if ts_unix else ''
    except Exception:
        ts_str = ''

    if ok:
        st.success(f"Feasible Search: **OK** - {reason} - best objective: {report.get('best_obj')}  {('(' + ts_str + ')') if ts_str else ''}")
    else:
        st.warning(f"Feasible Search: **NO RESULT** - {reason}  {('(' + ts_str + ')') if ts_str else ''}")

    # Compact summary
    st.json({k: report.get(k) for k in ['ok','reason','objective','budget','topk','multi_seed_runs','radius','seed','vars','start_feasible','best_obj','best_V'] if k in report})

    # Download JSON (stable key prefix avoids duplicates)
    try:
        st.download_button(
            label="Download feasible-search artifact JSON",
            data=_shams_json_dumps(report, indent=2, sort_keys=True),
            file_name="feasible_search_artifact.json",
            mime="application/json",
            key=f"{key_prefix}_dl_fs_artifact",
            use_container_width=True,
        )
    except Exception:
        pass

    # Candidates table
    try:
        cands = list(report.get('candidates', []) or [])
        if cands and pd is not None:
            rows = []
            for i, c in enumerate(cands):
                x = c.get('x', {}) or {}
                m = c.get('margins', {}) or {}
                row = {'rank': i+1, 'obj': c.get('obj'), 'V': c.get('V'), 'feasible': c.get('feasible')}
                # Common margins (if present)
                for nm in ['q95','q_div','P_SOL/R','B_peak','sigma_vm','HTS margin','TBR','NWL']:
                    if nm in m:
                        row[f'm_{nm}'] = m.get(nm)
                # Variables
                for k in (report.get('vars', []) or []):
                    row[k] = x.get(k)
                rows.append(row)
            if rows:
                st.dataframe(pd.DataFrame(rows), use_container_width=True)
    except Exception:
        pass

def _v93_stateful_sandbox_panel():
    import streamlit as st
    s = _v92_state_get()
    st.subheader("📌 Forge Cache (stateful sandbox)")
    if s.last_sandbox_run is None:
        st.info("No cached sandbox yet.")
        return
    st.download_button("Download sandbox_run.json (stateful)",
                       data=_json.dumps(s.last_sandbox_run, indent=2, sort_keys=True),
                       file_name="sandbox_run.json", mime="application/json",
                       use_container_width=True, key="v93_dl_sandbox_stateful")


# =====================
# v94 Run Records + Unified Export (append-only)
# =====================
def _v94_record_run(kind: str, payload: dict):
    import time
    s = _v92_state_get()
    try:
        s.run_history.append({
            "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
            "kind": kind,
            "summary": {
                "feasible": payload.get("meta", {}).get("feasible", payload.get("feasible", None)),
                "version": payload.get("version", None),
            },
        })
    except Exception:
        pass

def _v94_run_records_page():
    import streamlit as st
    s = _v92_state_get()
    st.subheader("Run Records")
    st.caption("Timeline of actions in this session.")
    if not s.run_history:
        st.info("No run records yet.")
        return
    for i, r in enumerate(reversed(s.run_history[-50:]), 1):
        st.write(f"{i}. {r.get('ts','?')} - {r.get('kind','?')} - {r.get('summary',{})}")

def _v94_unified_export_bundle_panel():
    import streamlit as st
    from pathlib import Path
    s = _v92_state_get()
    st.subheader("Unified Export Bundle")
    st.caption("One zip: artifacts + capsule + schemas + figures pack. Best-effort.")
    if not (s.has_point() or s.last_systems_result or s.last_scan_points or s.last_sandbox_run):
        st.info("Nothing cached yet. Run at least one mode first.")
        return

    scan_obj = None
    if s.last_scan_points is not None:
        scan_obj = {"kind":"shams_feasible_set","meta": s.last_scan_meta or {}, "points": s.last_scan_points}

    # Validation summary (best-effort)
    if s.has_point():
        ok, errs = _v93_validate_before_download(s.last_point_artifact, "schemas/shams_run_artifact.schema.json")
        if ok: st.success("Point artifact schema: PASS")
        else:
            st.warning("Point artifact schema: FAIL")
            for e in errs[:8]: st.write("- " + str(e))
    if scan_obj is not None:
        ok, errs = _v93_validate_before_download(scan_obj, "schemas/shams_feasible_set.schema.json")
        if ok: st.success("Scan feasible set schema: PASS")
        else:
            st.warning("Scan feasible set schema: FAIL")
            for e in errs[:8]: st.write("- " + str(e))

    try:
        from tools.export.bundle import build_export_bundle_bytes
        if st.button("Build unified export bundle", key="v94_build_bundle"):
            b = build_export_bundle_bytes(
                repo_root=Path("."),
                point_artifact=s.last_point_artifact if s.has_point() else None,
                systems_artifact=s.last_systems_result,
                feasible_search_artifact=getattr(s, "last_feasible_search_artifact", None),
                certified_search_artifact=st.session_state.get("last_certified_search_artifact", None),
                repair_evidence_artifact=st.session_state.get("last_repair_evidence_artifact", None),
                interval_refinement_artifact=st.session_state.get("last_interval_refinement_artifact", None),
                scan_artifact=scan_obj,
                pareto_artifact=getattr(s, "last_pareto_artifact", None),
                opt_artifact=getattr(s, "last_opt_artifact", None),
                sandbox_run=s.last_sandbox_run,
                figures_pack_zip=getattr(s, "last_figures_pack_zip", None),
            )
            st.session_state["v94_bundle_bytes"] = b
            st.success("Bundle built.")
        b = st.session_state.get("v94_bundle_bytes", None)
        if isinstance(b, (bytes, bytearray)) and len(b) > 0:
            st.download_button("Download shams_export_bundle.zip", data=b,
                               file_name="shams_export_bundle.zip", mime="application/zip",
                               use_container_width=True, key="v94_dl_bundle")
    except Exception as e:
        st.error(f"Bundle builder unavailable: {e!r}")



# =====================
# v98 validation gate (append-only)
# =====================
def _v98_validation_gate_ui(title: str, ok: bool, errs):
    import streamlit as st
    if ok:
        st.success(f"{title}: schema PASS")
        return True
    st.error(f"{title}: schema FAIL (download allowed, but NOT publishable)")
    for e in (errs or [])[:12]:
        st.write("- " + str(e))
    return False


# =====================

# =====================
# v98 Run Ledger (append-only)
# =====================
def _v98_state_init_runlists():
    import streamlit as st
    s = _v92_state_get()
    if getattr(s, "run_history", None) is None: s.run_history = []
    if getattr(s, "pinned_run_ids", None) is None: s.pinned_run_ids = []
    # v130: persistent run vault (opt-in, default ON)
    if 'vault_enabled' not in st.session_state:
        st.session_state['vault_enabled'] = True
    if 'vault_limit' not in st.session_state:
        st.session_state['vault_limit'] = 50
    return s

def _v98_make_run_id(prefix: str) -> str:
    import time, random
    return f"{prefix}_{int(time.time())}_{random.randint(1000,9999)}"

def _v98_record_run(kind: str, payload, mode: str = "") -> str:
    import streamlit as st
    from tools import run_vault
    from pathlib import Path
    import time
    s = _v98_state_init_runlists()
    rid = _v98_make_run_id(kind)
    s.run_history.append({"id": rid, "ts": time.strftime("%Y-%m-%d %H:%M:%S"), "kind": kind, "mode": mode, "payload": payload})
    try:
        _alog(mode or kind, 'RecordRun', {'rid': rid, 'kind': kind})
    except Exception:
        pass  # ActivityLogRecordRun

    # v130: persist to vault (storage only) - must never break UI
    try:
        if bool(st.session_state.get("vault_enabled", True)):
            root = Path(__file__).resolve().parents[1]
            run_vault.write_entry(root=root, kind=kind, payload=payload, mode=mode, tags={"rid": rid})
    except Exception:
        pass

    return rid


def _v98_json_diff(a, b, path=""):
    diffs = []
    if type(a) != type(b):
        diffs.append(path or "<root>"); return diffs
    if isinstance(a, dict):
        keys = set(a.keys()) | set(b.keys())
        for k in sorted(keys):
            diffs += _v98_json_diff(a.get(k, "<missing>"), b.get(k, "<missing>"), (path + "/" + str(k)) if path else "/" + str(k))
        return diffs
    if isinstance(a, list):
        n = max(len(a), len(b))
        for i in range(n):
            diffs += _v98_json_diff(a[i] if i < len(a) else "<missing>", b[i] if i < len(b) else "<missing>", f"{path}[{i}]")
        return diffs
    if a != b: diffs.append(path or "<root>")
    return diffs

def _v98_run_ledger_page():
    import streamlit as st
    s = _v98_state_init_runlists()
    st.subheader("Run Ledger")
    st.caption("Session-persistent run artifacts with pin/compare.")
    if not s.run_history:
        st.info("No recorded runs yet.")
        return
    for r in reversed(s.run_history[-100:]):
        rid = r.get("id")
        c1,c2,c3 = st.columns([3,1,1])
        with c1: st.write(f"**{rid}** - {r.get('ts')} - {r.get('kind')} ({r.get('mode')})")
        with c2:
            pinned = rid in s.pinned_run_ids
            if st.button("Unpin" if pinned else "Pin", key=f"pin_{rid}"):
                if pinned: s.pinned_run_ids.remove(rid)
                else: s.pinned_run_ids.append(rid)
                st.rerun()
        with c3:
            st.download_button("JSON", data=_json.dumps(r.get("payload", {}), indent=2, sort_keys=True), file_name=f"{rid}.json", mime="application/json", key=f"dl_{rid}")
    st.divider()
    st.subheader("Compare two pinned runs")
    pins = list(s.pinned_run_ids)
    if len(pins) < 2:
        st.info("Pin at least two runs to compare.")
        return
    a_id = st.selectbox("Run A", pins, key="cmp_a")
    b_id = st.selectbox("Run B", pins, index=1, key="cmp_b")
    A = next((x for x in s.run_history if x.get("id")==a_id), None)
    B = next((x for x in s.run_history if x.get("id")==b_id), None)
    if A and B:
        diffs = _v98_json_diff(A.get("payload",{}), B.get("payload",{}))
        st.write(f"Changed fields: {len(diffs)}")
        for d in diffs[:200]: st.write("- " + d)


def _v98_process_handoff_panel():
    import streamlit as st
    s = _v92_state_get()
    st.subheader("external systems codes Handoff")
    st.caption("Exports a external systems codes-oriented handoff JSON (SHAMS upstream feasibility auditor).")
    if not s.has_point():
        st.info("Run Point Designer first.")
        return
    from tools.interoperability.process_handoff import make_process_handoff
    ho = make_process_handoff(s.last_point_artifact)
    ok, errs = _v93_validate_before_download(ho, "schemas/process_handoff.schema.json")
    _v98_validation_gate_ui("external systems codes handoff", ok, errs)
    st.download_button("Download process_handoff.json", data=_json.dumps(ho, indent=2, sort_keys=True),
                       file_name="process_handoff.json", mime="application/json", use_container_width=True, key="v98_dl_process_handoff")


# =====================
# v99 Session Report Export (append-only)
# =====================
def _v99_session_report_panel():
    import streamlit as st
    s = _v98_state_init_runlists()
    st.subheader("Session Report Export")
    st.caption("Exports a single zip: run ledger + pinned runs + diffs. Optional: include unified export bundle if built.")
    if not s.run_history:
        st.info("No runs recorded yet.")
        return

    include_bundle = st.checkbox("Include unified export bundle (if already built)", value=True, key="v99_inc_bundle")
    if st.button("Build session report zip", key="v99_build_report"):
        try:
            from tools.export.session_report import build_session_report_zip
            bundle = st.session_state.get("v94_bundle_bytes", None) if include_bundle else None
            b = build_session_report_zip(
                version=str(getattr(s, "version", None) or "v99"),
                run_history=list(s.run_history),
                pinned_ids=list(s.pinned_run_ids or []),
                unified_export_bundle_bytes=bundle if isinstance(bundle, (bytes, bytearray)) else None,
            )
            st.session_state["v99_session_report_zip"] = b
            st.success("Session report built.")
        except Exception as e:
            st.error(f"Failed to build session report: {e!r}")

    b = st.session_state.get("v99_session_report_zip", None)
    if isinstance(b, (bytes, bytearray)) and len(b) > 0:
        st.download_button(
            "Download shams_session_report.zip",
            data=b,
            file_name="shams_session_report.zip",
            mime="application/zip",
            use_container_width=True,
            key="v99_dl_report",
        )


# =====================
# v103 Audit Pack + Atlas + Sandbox Plus panels (append-only)
# =====================
def _v103_audit_pack_panel():
    import streamlit as st
    s = _v98_state_init_runlists()
    st.subheader("Audit Pack")
    st.caption("Single zip: selected artifacts + schemas + environment + manifest (journal/regulator-ready).")
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    picked = st.multiselect("Include runs", options=ids, default=list(s.pinned_run_ids or []), key="v103_audit_pick")
    include_pf = st.checkbox("Include pip freeze (best-effort)", value=True, key="v103_audit_pf")
    if st.button("Build audit pack zip", key="v103_build_audit"):
        try:
            from tools.audit_pack import build_audit_pack_zip
            run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
            arts = []
            for rid in picked:
                r = run_map.get(rid)
                if r and isinstance(r.get("payload"), dict):
                    arts.append(r["payload"])
            b = build_audit_pack_zip(version="v103", artifacts=arts, schema_dir="schemas", include_pip_freeze=include_pf)
            st.session_state["v103_audit_zip"] = b
            st.success("Audit pack built.")
        except Exception as e:
            st.error(f"Failed: {e!r}")
    b = st.session_state.get("v103_audit_zip")
    if isinstance(b, (bytes, bytearray)) and len(b) > 0:
        st.download_button("Download shams_audit_pack.zip", data=b, file_name="shams_audit_pack.zip", mime="application/zip", use_container_width=True, key="v103_dl_audit")

def _v103_atlas_panel():
    import streamlit as st
    st.subheader("Feasibility Boundary Atlas")
    st.caption("Deterministic nearest-feasible sweeps using SHAMS frontier search (audit-ready).")
    s = _v92_state_get()
    if not getattr(s, "last_point_inp", None):
        st.info("Run Point Designer first (uses last point inputs).")
        return
    base = s.last_point_inp
    # Default lever bounds (conservative, user-editable)
    levers = {
        "R0_m": (max(0.5, float(getattr(base, "R0_m", 2.0))*0.7), float(getattr(base, "R0_m", 2.0))*1.3),
        "a_m": (max(0.2, float(getattr(base, "a_m", 0.6))*0.7), float(getattr(base, "a_m", 0.6))*1.3),
        "Bt_T": (max(1.0, float(getattr(base, "Bt_T", 12.0))*0.7), float(getattr(base, "Bt_T", 12.0))*1.3),
        "Ip_MA": (max(0.1, float(getattr(base, "Ip_MA", 8.0))*0.7), float(getattr(base, "Ip_MA", 8.0))*1.3),
        "fG": (max(0.1, float(getattr(base, "fG", 0.8))*0.7), min(1.2, float(getattr(base, "fG", 0.8))*1.3)),
    }
    n_random = st.slider("Random samples", 20, 300, 80, 10, key="v103_atlas_n")
    seed = st.number_input("Seed", value=0, step=1, key="v103_atlas_seed")
    if st.button("Build Atlas", key="v103_build_atlas"):
        try:
            from tools.frontier_atlas import build_feasibility_atlas
            atlas = build_feasibility_atlas(base, levers=levers, targets=None, n_random=int(n_random), seed=int(seed), n_slices=5)
            st.session_state["v103_atlas"] = atlas
            _v98_record_run("atlas", atlas, mode="feasibility_atlas")
            st.success("Atlas built and recorded in Run Ledger.")
        except Exception as e:
            st.error(f"Atlas failed: {e!r}")
    atlas = st.session_state.get("v103_atlas")
    if isinstance(atlas, dict):
        st.write("Reports:", int(atlas.get("n_reports", 0)))
        st.download_button("Download feasibility_atlas.json", data=_json.dumps(atlas, indent=2, sort_keys=True),
                           file_name="feasibility_atlas.json", mime="application/json", use_container_width=True, key="v103_dl_atlas")

def _v103_sandbox_plus_panel():
    import streamlit as st
    st.subheader("Optimizer Sandbox Plus")
    st.caption("Safe orchestration layer: random/LHS exploration with feasibility-first behavior and full logs.")
    s = _v92_state_get()
    if not getattr(s, "last_point_inp", None):
        st.info("Run Point Designer first (uses last point inputs as baseline).")
        return
    base = s.last_point_inp
    strat = st.selectbox("Strategy", ["random", "lhs"], index=0, key="v103_sb_strat")
    obj = st.selectbox("Objective", ["min_R0", "min_Bpeak", "max_Pnet", "min_recirc"], index=0, key="v103_sb_obj")
    max_evals = st.slider("Max evals", 20, 1000, 200, 20, key="v103_sb_evals")
    seed = st.number_input("Seed", value=0, step=1, key="v103_sb_seed")
    levers = {
        "R0_m": (max(0.5, float(getattr(base, "R0_m", 2.0))*0.7), float(getattr(base, "R0_m", 2.0))*1.3),
        "a_m": (max(0.2, float(getattr(base, "a_m", 0.6))*0.7), float(getattr(base, "a_m", 0.6))*1.3),
        "Bt_T": (max(1.0, float(getattr(base, "Bt_T", 12.0))*0.7), float(getattr(base, "Bt_T", 12.0))*1.3),
        "Ip_MA": (max(0.1, float(getattr(base, "Ip_MA", 8.0))*0.7), float(getattr(base, "Ip_MA", 8.0))*1.3),
        "fG": (max(0.1, float(getattr(base, "fG", 0.8))*0.7), min(1.2, float(getattr(base, "fG", 0.8))*1.3)),
    }
    if st.button("Run Sandbox Plus", key="v103_run_sandbox_plus"):
        try:
            from tools.sandbox_plus import run_sandbox
            run = run_sandbox(base, levers=levers, objective=obj, max_evals=int(max_evals), seed=int(seed), strategy=strat)
            st.session_state["v103_sandbox_plus"] = run
            _v98_record_run("sandbox_plus", run, mode="optimizer_sandbox_plus")
            st.success("Sandbox run complete and recorded.")
        except Exception as e:
            st.error(f"Sandbox failed: {e!r}")
    run = st.session_state.get("v103_sandbox_plus")
    if isinstance(run, dict):
        st.download_button("Download sandbox_plus.json", data=_json.dumps(run, indent=2, sort_keys=True),
                           file_name="sandbox_plus.json", mime="application/json", use_container_width=True, key="v103_dl_sandbox_plus")


# =====================
# v104 Feasible Space Topology (append-only)
# =====================
def _v104_topology_panel():
    import streamlit as st
    from tools.topology import build_feasible_topology, extract_feasible_points_from_payload

    st.subheader("Feasible Space Topology")
    st.caption("Builds a connectivity graph of feasible designs (components = design 'islands').")

    s = _v98_state_init_runlists()
    if not s.run_history:
        st.info("No runs recorded yet.")
        return

    # pick source runs (default pinned)
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    default_ids = list(s.pinned_run_ids or [])
    picked = st.multiselect("Source runs", options=ids, default=default_ids, key="v104_topology_pick")
    eps = st.slider("Connectivity threshold (scaled distance)", 0.05, 0.50, 0.18, 0.01, key="v104_topology_eps")
    max_pts = st.slider("Max points (cap)", 50, 600, 300, 10, key="v104_topology_cap")

    if st.button("Build topology", key="v104_build_topology"):
        try:
            run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
            pts = []
            for rid in picked:
                payload = (run_map.get(rid) or {}).get("payload")
                if isinstance(payload, dict):
                    pts += extract_feasible_points_from_payload(payload)
            topo = build_feasible_topology(pts, eps=float(eps), max_points=int(max_pts))
            st.session_state["v104_topology"] = topo
            _v98_record_run("topology", topo, mode="feasible_topology")
            st.success(f"Topology built: {topo.get('n_points',0)} points, {len(topo.get('components',[]))} components.")
        except Exception as e:
            st.error(f"Topology build failed: {e!r}")

    topo = st.session_state.get("v104_topology")
    if isinstance(topo, dict):
        comps = topo.get("components", [])
        st.write({
            "points": int(topo.get("n_points", 0)),
            "edges": int(topo.get("n_edges", 0)),
            "components": int(len(comps) if isinstance(comps, list) else 0),
            "largest_component": int(len(comps[0]) if isinstance(comps, list) and comps else 0),
        })
        if isinstance(comps, list) and comps:
            st.write("Component sizes:", [len(c) for c in comps[:10]])
        st.download_button("Download feasible_topology.json",
                           data=_json.dumps(topo, indent=2, sort_keys=True),
                           file_name="feasible_topology.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v104_dl_topology")


# =====================
# v105 Constraint Dominance & Sensitivity (append-only)
# =====================
def _v105_constraint_dominance_panel():
    import streamlit as st
    from tools.constraint_dominance import build_constraint_dominance_report

    st.subheader("Constraint Dominance")
    st.caption("Ranks which constraints most strongly limit feasibility (failures + near-boundary).")

    s = _v98_state_init_runlists()
    if not s.run_history:
        st.info("No runs recorded yet.")
        return

    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    default_ids = list(s.pinned_run_ids or [])
    picked = st.multiselect("Source runs (run artifacts only)", options=ids, default=default_ids, key="v105_dom_pick")
    near = st.slider("Near-boundary threshold (margin_frac)", 0.00, 0.25, 0.05, 0.01, key="v105_dom_near")
    fail_w = st.slider("Failure weight", 1.0, 10.0, 4.0, 0.5, key="v105_dom_failw")

    if st.button("Build dominance report", key="v105_build_dom"):
        try:
            run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
            payloads = []
            for rid in picked:
                payload = (run_map.get(rid) or {}).get("payload")
                if isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact":
                    payloads.append(payload)
            rep = build_constraint_dominance_report(payloads, near_threshold=float(near), fail_weight=float(fail_w))
            st.session_state["v105_dom"] = rep
            _v98_record_run("dominance", rep, mode="constraint_dominance")
            st.success(f"Built dominance report for {rep.get('n_rows',0)} constraint rows.")
        except Exception as e:
            st.error(f"Dominance failed: {e!r}")

    rep = st.session_state.get("v105_dom")
    if isinstance(rep, dict):
        ranked = rep.get("constraints_ranked", [])
        if isinstance(ranked, list) and ranked:
            top = ranked[:8]
            st.write([{"name": r.get("name"), "score": r.get("dominance_score"), "fail_rate": r.get("fail_rate"), "near_rate": r.get("near_boundary_rate")} for r in top])
        st.download_button("Download constraint_dominance_report.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="constraint_dominance_report.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v105_dl_dom")


# =====================
# v106 Failure Mode Taxonomy (append-only)
# =====================
def _v106_failure_taxonomy_panel():
    import streamlit as st
    from tools.failure_taxonomy import build_failure_taxonomy_report

    st.subheader("Failure Mode Taxonomy")
    st.caption("Classifies infeasible run artifacts by dominant failing constraint and aggregates failure modes.")

    s = _v98_state_init_runlists()
    if not s.run_history:
        st.info("No runs recorded yet.")
        return

    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    default_ids = list(s.pinned_run_ids or [])
    picked = st.multiselect("Source runs (run artifacts only)", options=ids, default=default_ids, key="v106_fail_pick")

    if st.button("Build failure taxonomy", key="v106_build_fail"):
        try:
            run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
            payloads = []
            for rid in picked:
                payload = (run_map.get(rid) or {}).get("payload")
                if isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact":
                    payloads.append(payload)
            rep = build_failure_taxonomy_report(payloads)
            st.session_state["v106_fail"] = rep
            _v98_record_run("failures", rep, mode="failure_taxonomy")
            st.success(f"Built failure taxonomy for {rep.get('n_failures',0)} failing runs.")
        except Exception as e:
            st.error(f"Failure taxonomy failed: {e!r}")

    rep = st.session_state.get("v106_fail")
    if isinstance(rep, dict):
        st.write({"failures": rep.get("n_failures"), "modes": len(rep.get("counts_by_mode",{}))})
        top = list(rep.get("counts_by_mode", {}).items())[:10]
        if top:
            st.write([{"mode": k, "count": v} for k,v in top])
        st.download_button("Download failure_taxonomy_report.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="failure_taxonomy_report.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v106_dl_fail")


# =====================
# v107 Feasibility Science Pack (append-only)
# =====================
def _v107_science_pack_panel():
    import streamlit as st
    from tools.science_pack import build_feasibility_science_pack

    st.subheader("Feasibility Science Pack")
    st.caption("One-click export: topology + dominance + failures + publishable report + zip.")

    s = _v98_state_init_runlists()
    if not s.run_history:
        st.info("No runs recorded yet.")
        return

    # Pull latest artifacts from session state if present, else try to find from run history
    topo = st.session_state.get("v104_topology")
    dom = st.session_state.get("v105_dom")
    fail = st.session_state.get("v106_fail")

    # fallback: scan run history for latest by mode
    def _latest_by_mode(mode_name: str):
        for r in reversed(s.run_history or []):
            if (r.get("mode") == mode_name) and isinstance(r.get("payload"), dict):
                return r.get("payload")
        return None

    if topo is None:
        topo = _latest_by_mode("feasible_topology")
    if dom is None:
        dom = _latest_by_mode("constraint_dominance")
    if fail is None:
        fail = _latest_by_mode("failure_taxonomy")

    ready = isinstance(topo, dict) and isinstance(dom, dict) and isinstance(fail, dict)
    st.write({
        "topology_loaded": isinstance(topo, dict),
        "dominance_loaded": isinstance(dom, dict),
        "failures_loaded": isinstance(fail, dict),
        "ready": ready,
    })
    if not ready:
        st.info("Build topology, dominance, and failure taxonomy first.")
        return

    version = st.text_input("Pack version label", value="v107", key="v107_pack_version")
    if st.button("Build Feasibility Science Pack", key="v107_build_pack"):
        try:
            pack = build_feasibility_science_pack(
                topology=topo,
                dominance=dom,
                failures=fail,
                source_run_ids=list(s.pinned_run_ids or []),
                version=str(version),
            )
            st.session_state["v107_pack"] = pack
            # record summary only (zip bytes excluded from run ledger)
            _v98_record_run("science_pack", {"kind": "shams_feasibility_science_pack_summary", "summary": pack.get("summary")}, mode="feasibility_science_pack")
            st.success("Science pack built.")
        except Exception as e:
            st.error(f"Science pack failed: {e!r}")

    pack = st.session_state.get("v107_pack")
    if isinstance(pack, dict):
        st.write(pack.get("summary", {}))
        zip_bytes = pack.get("zip_bytes")
        if isinstance(zip_bytes, (bytes, bytearray)):
            st.download_button(
                "Download feasibility_science_pack.zip",
                data=bytes(zip_bytes),
                file_name="feasibility_science_pack.zip",
                mime="application/zip",
                use_container_width=True,
                key="v107_dl_pack_zip",
            )
        manifest = dict(pack)
        manifest.pop("zip_bytes", None)
        st.download_button(
            "Download science_pack_manifest.json",
            data=_json.dumps(manifest, indent=2, sort_keys=True),
            file_name="science_pack_manifest.json",
            mime="application/json",
            use_container_width=True,
            key="v107_dl_pack_manifest",
        )


# =====================
# v108 external systems codes Downstream Export (append-only)
# =====================
def _v108_process_downstream_panel():
    import streamlit as st
    from tools.process_downstream import build_process_downstream_bundle

    st.subheader("external systems codes Downstream Export")
    st.caption("Exports SHAMS run artifacts to a transparent (systems-code-inspired) table so external systems codes becomes downstream.")

    s = _v98_state_init_runlists()
    if not s.run_history:
        st.info("No runs recorded yet.")
        return

    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    default_ids = list(s.pinned_run_ids or [])
    picked = st.multiselect("Run artifacts to export", options=ids, default=default_ids, key="v108_proc_pick")
    version = st.text_input("Export version label", value="v108", key="v108_proc_version")

    if st.button("Build external systems codes downstream bundle", key="v108_proc_build"):
        try:
            run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
            payloads = []
            for rid in picked:
                payload = (run_map.get(rid) or {}).get("payload")
                if isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact":
                    payloads.append(payload)
            pack = build_process_downstream_bundle(payloads, version=str(version), source_run_ids=list(picked))
            st.session_state["v108_proc_pack"] = pack
            # record summary only
            _v98_record_run("process_downstream", {"kind":"shams_process_downstream_summary","summary":pack.get("summary")}, mode="process_downstream")
            st.success("external systems codes downstream bundle built.")
        except Exception as e:
            st.error(f"Export failed: {e!r}")

    pack = st.session_state.get("v108_proc_pack")
    if isinstance(pack, dict):
        st.write(pack.get("summary", {}))
        zip_bytes = pack.get("zip_bytes")
        if isinstance(zip_bytes, (bytes, bytearray)):
            st.download_button(
                "Download process_downstream_bundle.zip",
                data=bytes(zip_bytes),
                file_name="process_downstream_bundle.zip",
                mime="application/zip",
                use_container_width=True,
                key="v108_proc_dl_zip",
            )
        manifest = dict(pack)
        manifest.pop("zip_bytes", None)
        st.download_button(
            "Download process_downstream_manifest.json",
            data=_json.dumps(manifest, indent=2, sort_keys=True),
            file_name="process_downstream_manifest.json",
            mime="application/json",
            use_container_width=True,
            key="v108_proc_dl_manifest",
        )


# =====================
# v109 Island Inspector (append-only)
# =====================
def _v109_island_inspector_panel():
    import streamlit as st
    from tools.component_dominance import build_component_dominance_report

    st.subheader("Island Inspector")
    st.caption("Per-feasible-island dominance + boundary-near failure modes.")

    s = _v98_state_init_runlists()
    # Get topology + failure taxonomy from session/run ledger
    topo = st.session_state.get("v104_topology")
    dom = st.session_state.get("v105_dom")  # not required, but should exist for prior steps
    fail = st.session_state.get("v106_fail")

    def _latest_by_mode(mode_name: str):
        for r in reversed(s.run_history or []):
            if (r.get("mode") == mode_name) and isinstance(r.get("payload"), dict):
                return r.get("payload")
        return None

    if topo is None:
        topo = _latest_by_mode("feasible_topology")
    if fail is None:
        fail = _latest_by_mode("failure_taxonomy")

    st.write({"topology_loaded": isinstance(topo, dict), "failures_loaded": isinstance(fail, dict)})
    if not isinstance(topo, dict):
        st.info("Build topology first.")
        return

    # choose run artifacts to assign
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    default_ids = list(s.pinned_run_ids or [])
    picked = st.multiselect("Run artifacts to use (feasible ones define islands)", options=ids, default=default_ids, key="v109_pick_runs")
    near = st.slider("Near-boundary threshold (margin_frac)", 0.00, 0.25, 0.05, 0.01, key="v109_near")
    fail_w = st.slider("Failure weight", 1.0, 10.0, 4.0, 0.5, key="v109_failw")

    if st.button("Build component dominance report", key="v109_build"):
        try:
            run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
            payloads = []
            for rid in picked:
                payload = (run_map.get(rid) or {}).get("payload")
                if isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact":
                    payloads.append(payload)
            rep = build_component_dominance_report(topology=topo, run_artifacts=payloads, failure_taxonomy=fail, near_threshold=float(near), fail_weight=float(fail_w))
            st.session_state["v109_components"] = rep
            _v98_record_run("islands", rep, mode="component_dominance")
            st.success(f"Built component report for {rep.get('n_components',0)} components.")
        except Exception as e:
            st.error(f"v109 failed: {e!r}")

    rep = st.session_state.get("v109_components")
    if isinstance(rep, dict):
        comps = rep.get("components", [])
        if isinstance(comps, list) and comps:
            st.write("Top components (by size):")
            st.write([{
                "component": c.get("component_index"),
                "size_pts": c.get("component_size_points"),
                "runs": c.get("n_feasible_runs_assigned"),
                "top_constraint": (c.get("dominance_top_constraints") or [{}])[0].get("name") if isinstance(c.get("dominance_top_constraints"), list) and c.get("dominance_top_constraints") else None,
                "top_failure_mode": (c.get("top_failure_modes_near_component") or [{}])[0].get("mode") if isinstance(c.get("top_failure_modes_near_component"), list) and c.get("top_failure_modes_near_component") else None,
            } for c in comps[:12]])
        st.download_button("Download component_dominance_report.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="component_dominance_report.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v109_dl_report")


# =====================
# v110 Boundary Atlas v2 (append-only)
# =====================
def _v110_boundary_atlas_panel():
    import streamlit as st
    from tools.boundary_atlas_v2 import build_boundary_atlas_v2
    from tools.plot_boundary_atlas_v2 import main as _plot_cli  # not used directly

    st.subheader("Feasibility Boundary Atlas v2")
    st.caption("Extracts explicit feasible/infeasible boundaries for lever-pair slices, with failure-mode labels (best-effort).")

    s = _v98_state_init_runlists()
    topo = st.session_state.get("v104_topology")
    fail = st.session_state.get("v106_fail")

    def _latest_by_mode(mode_name: str):
        for r in reversed(s.run_history or []):
            if (r.get("mode") == mode_name) and isinstance(r.get("payload"), dict):
                return r.get("payload")
        return None

    if fail is None:
        fail = _latest_by_mode("failure_taxonomy")

    # Use pinned runs as sources; include atlas/sandbox/topology artifacts if present in history
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    default_ids = list(s.pinned_run_ids or [])
    picked = st.multiselect("Source run artifacts (infeasible + feasible)", options=ids, default=default_ids, key="v110_pick_runs")
    q = st.slider("Boundary proximity quantile (lower = tighter)", 0.05, 0.80, 0.25, 0.05, key="v110_q")
    maxpairs = st.slider("Max lever pairs", 1, 8, 6, 1, key="v110_maxpairs")

    if st.button("Build Boundary Atlas v2", key="v110_build"):
        try:
            run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
            payloads = []
            for rid in picked:
                payload = (run_map.get(rid) or {}).get("payload")
                if isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact":
                    payloads.append(payload)
            rep = build_boundary_atlas_v2(payloads, failure_taxonomy=fail, max_pairs=int(maxpairs), proximity_quantile=float(q))
            st.session_state["v110_boundary_atlas"] = rep
            _v98_record_run("atlas_v2", rep, mode="boundary_atlas_v2")
            st.success(f"Built atlas slices: {len(rep.get('slices', []))}")
        except Exception as e:
            st.error(f"v110 failed: {e!r}")

    rep = st.session_state.get("v110_boundary_atlas")
    if isinstance(rep, dict):
        st.write({"slices": len(rep.get("slices", [])), "lever_pairs": rep.get("lever_pairs")})
        st.download_button("Download boundary_atlas_v2.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="boundary_atlas_v2.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v110_dl_json")


# =====================
# v111 Design Family Explorer (append-only)
# =====================
def _v111_design_family_panel():
    import streamlit as st
    from tools.design_family import build_design_family_report

    st.subheader("Design Family Explorer")
    st.caption("Safe local exploration within a feasible island (no optimization).")

    s = _v98_state_init_runlists()
    topo = st.session_state.get("v104_topology")
    def _latest_by_mode(mode_name: str):
        for r in reversed(s.run_history or []):
            if (r.get("mode") == mode_name) and isinstance(r.get("payload"), dict):
                return r.get("payload")
        return None
    if topo is None:
        topo = _latest_by_mode("feasible_topology")

    if not isinstance(topo, dict):
        st.info("Build topology first.")
        return

    comps = topo.get("components", [])
    ncomp = len(comps) if isinstance(comps, list) else 0
    st.write({"n_components": ncomp})

    # choose baseline run (for full PointInputs)
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    default_id = (list(s.pinned_run_ids or [])[:1] or [ids[0] if ids else None])[0]
    baseline_id = st.selectbox("Baseline run artifact (provides full inputs)", options=[None] + ids, index=(1 if default_id in ids else 0), key="v111_base_id")
    comp_idx = st.number_input("Component index", min_value=0, max_value=max(0, ncomp-1), value=0, step=1, key="v111_comp_idx")
    n_samples = st.slider("Samples", 10, 240, 120, 10, key="v111_ns")
    radius = st.slider("Local radius (fraction of lever span)", 0.01, 0.25, 0.08, 0.01, key="v111_rad")
    seed = st.number_input("Seed", min_value=0, max_value=10_000_000, value=0, step=1, key="v111_seed")

    if st.button("Run Design Family Explorer", key="v111_run"):
        try:
            run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
            payload = (run_map.get(baseline_id) or {}).get("payload") if baseline_id else None
            if not (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact"):
                st.error("Baseline run artifact not found. Pin/select a valid run artifact first.")
                return
            base_inputs = payload.get("inputs", {})
            if not isinstance(base_inputs, dict):
                st.error("Baseline inputs invalid.")
                return

            rep = build_design_family_report(
                topology=topo,
                component_index=int(comp_idx),
                baseline_inputs=base_inputs,
                n_samples=int(n_samples),
                radius_frac=float(radius),
                seed=int(seed),
            )
            st.session_state["v111_family"] = rep
            _v98_record_run("design_family", rep, mode="design_family")
            st.success("Design family report built.")
        except Exception as e:
            st.error(f"v111 failed: {e!r}")

    rep = st.session_state.get("v111_family")
    if isinstance(rep, dict):
        st.write({
            "component_index": rep.get("component_index"),
            "n_samples": rep.get("n_samples"),
            "feasible_fraction": rep.get("feasible_fraction"),
            "top_worst_hard": (rep.get("worst_hard_ranked") or [{}])[0],
        })
        st.download_button("Download design_family_report.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="design_family_report.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v111_dl_json")


# =====================
# v112 Literature Overlay (append-only)
# =====================
def _v112_literature_overlay_panel():
    import streamlit as st
    from tools.literature_overlay import template_payload, validate_literature_points
    from tools.literature_overlay import extract_xy_points

    st.subheader("Literature Overlay")
    st.caption("Upload a JSON of reference points (ITER/ARC/SPARC/etc.) and overlay on Boundary Atlas slices. SHAMS ships only a template.")

    if "v112_overlay" not in st.session_state:
        st.session_state["v112_overlay"] = template_payload(version="v112")

    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "Download overlay template JSON",
            data=_json.dumps(template_payload(version="v112"), indent=2, sort_keys=True),
            file_name="literature_points_template.json",
            mime="application/json",
            use_container_width=True,
            key="v112_dl_template",
        )
    with c2:
        up = st.file_uploader("Upload literature_points.json", type=["json"], key="v112_upload")
        if up is not None:
            try:
                payload = _json.loads(up.getvalue().decode("utf-8"))
                errs = validate_literature_points(payload)
                st.session_state["v112_overlay"] = payload
                if errs:
                    st.warning(f"Loaded with warnings: {errs[:8]}")
                else:
                    st.success("Overlay loaded.")
            except Exception as e:
                st.error(f"Failed to parse JSON: {e!r}")

    overlay = st.session_state.get("v112_overlay")
    st.write({"n_points": len((overlay or {}).get("points", [])) if isinstance(overlay, dict) else None})

    atlas = st.session_state.get("v110_boundary_atlas")
    if isinstance(atlas, dict) and isinstance(overlay, dict):
        slices = atlas.get("slices", [])
        if isinstance(slices, list) and slices:
            st.write("Preview overlay-able slices (first 6):")
            prev = []
            for sl in slices[:6]:
                if not isinstance(sl, dict):
                    continue
                kx = sl.get("lever_x"); ky = sl.get("lever_y")
                if isinstance(kx, str) and isinstance(ky, str):
                    prev.append({"x": kx, "y": ky, "points_available": len(extract_xy_points(overlay, kx, ky))})
            st.write(prev)


# =====================
# v113 Design Decision Layer (append-only)
# =====================
def _v113_design_decision_panel():
    import streamlit as st
    from tools.design_decision_layer import build_design_candidates, build_design_decision_pack

    st.subheader("Design Decision Layer")
    st.caption("Build defensible design candidates + comparison table + exportable decision pack (no optimization).")

    s = _v98_state_init_runlists()

    topo = st.session_state.get("v104_topology")
    comp = st.session_state.get("v109_components")
    atlas = st.session_state.get("v110_boundary_atlas")
    fam = st.session_state.get("v111_family")
    overlay = st.session_state.get("v112_overlay")

    # pick candidate artifacts from pinned runs (feasible only)
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    default_ids = list(s.pinned_run_ids or [])
    picked = st.multiselect("Candidate source run artifacts (feasible ones)", options=ids, default=default_ids, key="v113_pick")
    maxc = st.slider("Max candidates", 1, 20, 12, 1, key="v113_maxc")

    if st.button("Build candidates + decision pack", key="v113_build"):
        try:
            run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
            artifacts = []
            for rid in picked:
                payload = (run_map.get(rid) or {}).get("payload")
                if isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact":
                    artifacts.append(payload)

            candidates = build_design_candidates(
                artifacts=artifacts,
                topology=topo if isinstance(topo, dict) else None,
                component_dominance=comp if isinstance(comp, dict) else None,
                boundary_atlas_v2=atlas if isinstance(atlas, dict) else None,
                design_family_report=fam if isinstance(fam, dict) else None,
                literature_overlay=overlay if isinstance(overlay, dict) else None,
                max_candidates=int(maxc),
            )
            pack = build_design_decision_pack(candidates=candidates, version="v113")
            st.session_state["v113_candidates"] = candidates
            st.session_state["v113_pack"] = pack
            _v98_record_run("design_decision_pack", {"candidates": candidates, "manifest": pack.get("manifest")}, mode="design_decision_pack")
            st.success(f"Built {len(candidates)} candidates.")
        except Exception as e:
            st.error(f"v113 failed: {e!r}")

    candidates = st.session_state.get("v113_candidates")
    pack = st.session_state.get("v113_pack")
    if isinstance(candidates, list) and candidates:
        st.write("Candidate preview (first 5):")
        st.write([{
            "id": c.get("source_artifact_id"),
            "component": c.get("component_index"),
            "worst": (c.get("feasibility") or {}).get("worst_hard"),
            "margin": (c.get("feasibility") or {}).get("worst_hard_margin_frac"),
            "family_feas": (c.get("robustness") or {}).get("family_feasible_fraction"),
        } for c in candidates[:5]])
        st.download_button("Download candidates.json",
                           data=_json.dumps({"candidates": candidates}, indent=2, sort_keys=True),
                           file_name="candidates.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v113_dl_candidates")

    if isinstance(pack, dict) and isinstance(pack.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download design_decision_pack.zip",
                           data=pack["zip_bytes"],
                           file_name="design_decision_pack.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v113_dl_pack")


# =====================
# v114 Preference-Aware Decision Layer (append-only)
# =====================
def _v114_preference_panel():
    import streamlit as st
    from tools.preference_layer import template_preferences, annotate_candidates_with_preferences, pareto_sets_from_annotations
    from tools.design_decision_layer import build_design_decision_pack

    st.subheader("Preference-Aware Decision Layer")
    st.caption("Preference sliders annotate candidates (scores + Pareto sets). No optimization; no auto-selection.")

    # candidates from v113
    candidates = st.session_state.get("v113_candidates")
    if not (isinstance(candidates, list) and candidates):
        st.info("Build v113 candidates first (Design Decision Layer).")
        return

    if "v114_prefs" not in st.session_state:
        st.session_state["v114_prefs"] = template_preferences()

    prefs = st.session_state["v114_prefs"]
    weights = (prefs.get("weights") if isinstance(prefs, dict) else None)
    if not isinstance(weights, dict):
        weights = {}

    st.write("Weights (0 disables a metric).")
    c1, c2 = st.columns(2)
    with c1:
        weights["margin"] = st.slider("margin (worst constraint margin)", 0.0, 2.0, float(weights.get("margin", 1.0)), 0.1, key="v114_w_margin")
        weights["robustness"] = st.slider("robustness (family feasible fraction)", 0.0, 2.0, float(weights.get("robustness", 1.0)), 0.1, key="v114_w_rob")
        weights["boundary_clearance"] = st.slider("boundary clearance (2D proxy)", 0.0, 2.0, float(weights.get("boundary_clearance", 0.7)), 0.1, key="v114_w_bd")
    with c2:
        weights["compactness"] = st.slider("compactness (smaller R0)", 0.0, 2.0, float(weights.get("compactness", 0.3)), 0.1, key="v114_w_comp")
        weights["low_aux_power"] = st.slider("low aux power (smaller Paux)", 0.0, 2.0, float(weights.get("low_aux_power", 0.2)), 0.1, key="v114_w_paux")

    prefs["weights"] = weights
    st.session_state["v114_prefs"] = prefs

    metrics = st.multiselect("Pareto metrics (normalized)", options=["margin","robustness","boundary_clearance","compactness","low_aux_power"],
                             default=["margin","robustness","boundary_clearance"], key="v114_metrics")

    if st.button("Annotate + compute Pareto sets", key="v114_run"):
        try:
            ann = annotate_candidates_with_preferences(candidates, prefs)
            pareto = pareto_sets_from_annotations(ann, metrics=metrics, max_fronts=3)
            st.session_state["v114_ann"] = ann
            st.session_state["v114_pareto"] = pareto
            st.success("v114 annotations + Pareto computed.")
        except Exception as e:
            st.error(f"v114 failed: {e!r}")

    ann = st.session_state.get("v114_ann")
    pareto = st.session_state.get("v114_pareto")

    if isinstance(ann, dict):
        # show top composite scores (best-effort)
        cands = ann.get("candidates", [])
        preview = []
        if isinstance(cands, list):
            for c in cands:
                if not isinstance(c, dict):
                    continue
                pa = c.get("preference_annotation_v114", {})
                comp = pa.get("composite_score") if isinstance(pa, dict) else None
                preview.append({
                    "id": c.get("source_artifact_id"),
                    "composite_score": comp,
                    "worst": (c.get("feasibility") or {}).get("worst_hard"),
                    "margin": (c.get("feasibility") or {}).get("worst_hard_margin_frac"),
                })
            preview.sort(key=lambda r: (r.get("composite_score") is None, -(r.get("composite_score") or -1e9)))
        st.write("Preview (top 8 by composite score; annotation only):")
        st.write(preview[:8])

        st.download_button("Download preference_annotation_bundle.json",
                           data=_json.dumps(ann, indent=2, sort_keys=True),
                           file_name="preference_annotation_bundle_v114.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v114_dl_ann")

    if isinstance(pareto, dict):
        st.write({"pareto_metrics": pareto.get("metrics"), "n_points": pareto.get("n_points"), "front_sizes": [len(f) for f in (pareto.get("fronts") or []) if isinstance(f, list)]})
        st.download_button("Download pareto_sets.json",
                           data=_json.dumps(pareto, indent=2, sort_keys=True),
                           file_name="pareto_sets_v114.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v114_dl_pareto")

    if isinstance(ann, dict) and isinstance(pareto, dict):
        # build a justification + pack (adds decision_justification.json inside zip)
        justification = {
            "kind": "shams_decision_justification_v114",
            "created_utc": ann.get("created_utc"),
            "preferences": prefs,
            "pareto_sets": pareto,
            "n_candidates": len((ann.get("candidates") or [])) if isinstance(ann.get("candidates"), list) else None,
            "disclaimer": "Annotations only. No optimization. No auto-selected best design.",
        }
        pack = build_design_decision_pack(candidates=candidates, version="v114", decision_justification=justification)
        st.download_button("Download v114 design_decision_pack.zip (with justification)",
                           data=pack["zip_bytes"],
                           file_name="design_decision_pack_v114.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v114_dl_pack")


# =====================
# v115 External Optimizer Sandbox (append-only)
# =====================
def _v115_optimizer_sandbox_panel():
    import streamlit as st
    from tools.optimizer_interface import template_request, template_response, evaluate_optimizer_proposal, build_optimizer_import_pack

    st.subheader("External Optimizer Sandbox")
    st.caption("Import external optimizer proposals as *read-only* candidates. SHAMS re-evaluates physics+constraints and records a run artifact.")

    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "Download optimizer_request_template.json",
            data=_json.dumps(template_request(version="v115"), indent=2, sort_keys=True),
            file_name="optimizer_request_template.json",
            mime="application/json",
            use_container_width=True,
            key="v115_dl_req",
        )
    with c2:
        st.download_button(
            "Download optimizer_response_template.json",
            data=_json.dumps(template_response(version="v115"), indent=2, sort_keys=True),
            file_name="optimizer_response_template.json",
            mime="application/json",
            use_container_width=True,
            key="v115_dl_resp_tpl",
        )

    up = st.file_uploader("Upload optimizer_response.json (proposal)", type=["json"], key="v115_upload_resp")
    if up is None:
        st.info("Upload a proposal response to evaluate it inside SHAMS.")
        return

    try:
        payload = _json.loads(up.getvalue().decode("utf-8"))
    except Exception as e:
        st.error(f"Failed to parse JSON: {e!r}")
        return

    if payload.get("kind") != "shams_optimizer_response":
        st.error("JSON kind must be 'shams_optimizer_response'.")
        return

    if st.button("Evaluate proposal in SHAMS (frozen physics)", key="v115_eval"):
        try:
            out = evaluate_optimizer_proposal(payload)
            art = out["artifact"]
            ctx = out["context"]
            st.session_state["v115_last_artifact"] = art
            st.session_state["v115_last_context"] = ctx
            # record in ledger
            _v98_record_run("optimizer_proposal", {"artifact": art, "context": ctx}, mode="optimizer_import_v115")
            st.success("Proposal evaluated and recorded in Run Ledger.")
        except Exception as e:
            st.error(f"Evaluation failed: {e!r}")

    art = st.session_state.get("v115_last_artifact")
    ctx = st.session_state.get("v115_last_context")

    if isinstance(ctx, dict):
        st.write("Import context:")
        st.write({k: ctx.get(k) for k in ["result", "disclaimer"]})

    if isinstance(art, dict) and art.get("kind") == "shams_run_artifact":
        st.write("Evaluated artifact summary:")
        cs = art.get("constraints_summary", {})
        st.write({
            "id": art.get("id"),
            "feasible": (cs.get("feasible") if isinstance(cs, dict) else None),
            "worst_hard": (cs.get("worst_hard") if isinstance(cs, dict) else None),
            "worst_margin": (cs.get("worst_hard_margin_frac") if isinstance(cs, dict) else None),
        })
        st.download_button("Download evaluated_run_artifact.json",
                           data=_json.dumps(art, indent=2, sort_keys=True),
                           file_name="evaluated_run_artifact.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v115_dl_eval_art")

        pack = build_optimizer_import_pack(
            request_template=template_request(version="v115"),
            response_template=template_response(version="v115"),
            evaluated_artifact=art,
            import_context=ctx if isinstance(ctx, dict) else None,
            version="v115",
        )
        st.download_button("Download optimizer_import_pack.zip",
                           data=pack["zip_bytes"],
                           file_name="optimizer_import_pack.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v115_dl_pack")


# =====================
# v116 Design Handoff Pack (append-only)
# =====================
def _v116_handoff_pack_panel():
    import streamlit as st
    from tools.handoff_pack import build_handoff_pack

    st.subheader("Design Handoff Pack")
    st.caption("Export an engineering-ready handoff bundle for any run artifact (inputs YAML, constraints CSV, figures, manifest).")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("No runs in ledger yet. Run a point evaluation first.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Select run artifact", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v116_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    art = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None

    if art is None:
        st.error("Selected run does not contain a run artifact payload.")
        return

    if st.button("Build handoff pack", key="v116_build"):
        try:
            pack = build_handoff_pack(artifact=art, version="v116")
            st.session_state["v116_pack"] = pack
            _v98_record_run("handoff_pack", {"manifest": pack.get("manifest"), "source_artifact_id": art.get("id")}, mode="handoff_pack_v116")
            st.success("Handoff pack built.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    pack = st.session_state.get("v116_pack")
    if isinstance(pack, dict) and isinstance(pack.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download handoff_pack.zip",
                           data=pack["zip_bytes"],
                           file_name=f"handoff_pack_{rid}.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v116_dl_pack")
        st.download_button("Download handoff manifest.json",
                           data=_json.dumps(pack.get("manifest", {}), indent=2, sort_keys=True),
                           file_name="handoff_pack_manifest.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v116_dl_manifest")


# =====================
# v117 Tolerance Envelope (append-only)
# =====================
def _v117_tolerance_envelope_panel():
    import streamlit as st
    from tools.tolerance_envelope import template_tolerance_spec, evaluate_tolerance_envelope, envelope_summary_csv

    st.subheader("Tolerance Envelope")
    st.caption("Deterministic tolerance envelope around a selected run artifact. No Monte Carlo, no optimization.")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("No runs in ledger yet. Run a point evaluation first.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Select baseline run artifact", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v117_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    art = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if art is None:
        st.error("Selected run does not contain a run artifact payload.")
        return

    if "v117_spec" not in st.session_state:
        st.session_state["v117_spec"] = template_tolerance_spec()

    spec = st.session_state["v117_spec"]
    tmap = spec.get("tolerances") if isinstance(spec, dict) else {}
    if not isinstance(tmap, dict):
        tmap = {}

    mode = st.selectbox("Tolerance mode", options=["relative","absolute"], index=0 if spec.get("mode","relative")=="relative" else 1, key="v117_mode")
    include_mid = st.checkbox("Include edge midpoints", value=bool(spec.get("include_edge_midpoints", True)), key="v117_mid")
    max_samples = st.number_input("Max samples", min_value=10, max_value=500, value=200, step=10, key="v117_max")

    st.write("Tolerances (per lever):")
    cols = st.columns(3)
    keys = ["Bt_T","Ip_MA","R0_m","a_m","fG","Ti_keV","Paux_MW","kappa"]
    for i,k in enumerate(keys):
        with cols[i % 3]:
            default = float(tmap.get(k, 0.0) or 0.0)
            tmap[k] = st.number_input(f"{k} tol", min_value=0.0, max_value=1e6, value=default, step=0.005 if mode=="relative" else 0.5, key=f"v117_tol_{k}")

    spec["mode"] = mode
    spec["include_edge_midpoints"] = include_mid
    spec["tolerances"] = tmap
    st.session_state["v117_spec"] = spec

    st.download_button("Download tolerance_spec.json",
                       data=_json.dumps(spec, indent=2, sort_keys=True),
                       file_name="tolerance_spec_v117.json",
                       mime="application/json",
                       use_container_width=True,
                       key="v117_dl_spec")

    if st.button("Run tolerance envelope", key="v117_run"):
        try:
            rep = evaluate_tolerance_envelope(baseline_artifact=art, tolerance_spec=spec, version="v117", max_samples=int(max_samples))
            st.session_state["v117_report"] = rep
            _v98_record_run("tolerance_envelope", {"summary": rep.get("summary"), "source_artifact_id": rid}, mode="tolerance_envelope_v117")
            st.success("Tolerance envelope computed.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    rep = st.session_state.get("v117_report")
    if isinstance(rep, dict) and rep.get("kind") == "shams_tolerance_envelope_report":
        st.write("Summary:")
        st.write(rep.get("summary", {}))
        st.download_button("Download tolerance_envelope_report.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="tolerance_envelope_report_v117.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v117_dl_report")
        st.download_button("Download tolerance_envelope_summary.csv",
                           data=envelope_summary_csv(rep),
                           file_name="tolerance_envelope_summary_v117.csv",
                           mime="text/csv",
                           use_container_width=True,
                           key="v117_dl_csv")


# =====================
# v118 Optimizer Downstream Workflow (append-only)
# =====================
def _v118_optimizer_downstream_panel():
    import streamlit as st
    from tools.optimizer_downstream import template_batch_response, evaluate_optimizer_batch, build_downstream_report_zip
    from tools.preference_layer import template_preferences
    from tools.tolerance_envelope import template_tolerance_spec

    st.subheader("Optimizer Downstream Workflow")
    st.caption("Upload a batch of optimizer proposals; SHAMS evaluates, filters feasible, runs tolerance envelopes, builds candidates, applies preferences+Pareto, and exports one bundle.")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.download_button("Download optimizer_batch_template.json",
                           data=_json.dumps(template_batch_response(), indent=2, sort_keys=True),
                           file_name="optimizer_batch_template.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v118_dl_batch_tpl")
    with c2:
        st.download_button("Download preferences_template.json",
                           data=_json.dumps(template_preferences(), indent=2, sort_keys=True),
                           file_name="preferences_template.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v118_dl_prefs_tpl")
    with c3:
        st.download_button("Download tolerance_spec_template.json",
                           data=_json.dumps(template_tolerance_spec(), indent=2, sort_keys=True),
                           file_name="tolerance_spec_template.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v118_dl_tol_tpl")

    up = st.file_uploader("Upload optimizer_batch.json", type=["json"], key="v118_up_batch")
    if up is None:
        st.info("Upload a batch proposal JSON to run v118.")
        return

    try:
        batch = _json.loads(up.getvalue().decode("utf-8"))
    except Exception as e:
        st.error(f"Failed to parse JSON: {e!r}")
        return

    if batch.get("kind") != "shams_optimizer_batch_response":
        st.error("JSON kind must be 'shams_optimizer_batch_response'.")
        return

    max_env = st.number_input("Max envelope samples per feasible point", min_value=8, max_value=80, value=24, step=4, key="v118_max_env")
    max_cand = st.number_input("Max candidates", min_value=1, max_value=50, value=12, step=1, key="v118_max_cand")

    # Optional: use current v114 prefs if present, else template
    prefs = st.session_state.get("v114_prefs")
    if not isinstance(prefs, dict):
        prefs = template_preferences()
    spec = st.session_state.get("v117_spec")
    if not isinstance(spec, dict):
        spec = template_tolerance_spec()

    if st.button("Run v118 downstream report", key="v118_run"):
        try:
            out = evaluate_optimizer_batch(batch_payload=batch, tolerance_spec=spec, max_envelope_samples=int(max_env), max_candidates=int(max_cand), preferences=prefs)
            rep = out["report"]
            pack_bytes = out["decision_pack_zip_bytes"]
            st.session_state["v118_report"] = rep
            st.session_state["v118_pack_bytes"] = pack_bytes
            bundle = build_downstream_report_zip(report_obj=rep, decision_pack_zip_bytes=pack_bytes)
            st.session_state["v118_bundle"] = bundle
            _v98_record_run("optimizer_downstream", {"summary": rep.get("batch_meta"), "decision_pack_manifest": rep.get("decision_pack_manifest")}, mode="optimizer_downstream_v118")
            st.success("v118 downstream report built.")
        except Exception as e:
            st.error(f"v118 failed: {e!r}")

    rep = st.session_state.get("v118_report")
    bundle = st.session_state.get("v118_bundle")
    if isinstance(rep, dict):
        st.write("Batch meta:")
        st.write(rep.get("batch_meta", {}))
        st.write("Decision justification summary:")
        dj = rep.get("decision_justification", {}) if isinstance(rep.get("decision_justification"), dict) else {}
        st.write({k: dj.get(k) for k in ["n_proposals","n_feasible","n_candidates","n_envelopes"]})
        st.download_button("Download optimizer_downstream_report_v118.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="optimizer_downstream_report_v118.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v118_dl_report")

    pack_bytes = st.session_state.get("v118_pack_bytes")
    if isinstance(pack_bytes, (bytes, bytearray)):
        st.download_button("Download design_decision_pack_v118.zip",
                           data=pack_bytes,
                           file_name="design_decision_pack_v118.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v118_dl_pack")

    if isinstance(bundle, dict) and isinstance(bundle.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download optimizer_downstream_bundle_v118.zip",
                           data=bundle["zip_bytes"],
                           file_name="optimizer_downstream_bundle_v118.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v118_dl_bundle")


# =====================
# v119 Authority Pack (append-only)
# =====================
def _v119_authority_pack_panel():
    import streamlit as st
    from tools.authority_pack import build_authority_pack

    st.subheader("Authority Pack")
    st.caption("Build a single publishable evidence bundle: version, patch notes, requirements freeze (best-effort), command log, methods appendix, and selected artifacts.")

    # try to find latest generated artifacts from self-test run folder if present in session state
    audit_zip = st.session_state.get("last_audit_pack_zip_bytes")
    downstream_zip = None
    handoff_zip = None

    # If users ran v118/v116 panels in-session, these may exist:
    bundle = st.session_state.get("v118_bundle")
    if isinstance(bundle, dict) and isinstance(bundle.get("zip_bytes"), (bytes, bytearray)):
        downstream_zip = bytes(bundle["zip_bytes"])
    hp = st.session_state.get("v116_pack")
    if isinstance(hp, dict) and isinstance(hp.get("zip_bytes"), (bytes, bytearray)):
        handoff_zip = bytes(hp["zip_bytes"])

    cmdlog = st.text_area("Command log (editable)", value="\n".join([
        "python -m tools.ui_self_test --outdir out_ui_self_test",
        "python -m tools.verify_package",
        "python -m tools.verify_figures",
        "python -m tools.tests.test_plot_layout",
        "python -m tools.regression_suite",
    ]), height=140, key="v119_cmdlog")

    if st.button("Build Authority Pack", key="v119_build"):
        try:
            pack = build_authority_pack(
                repo_root=".",
                version="v119",
                audit_pack_zip=audit_zip,
                downstream_bundle_zip=downstream_zip,
                handoff_pack_zip=handoff_zip,
                command_log=[l for l in cmdlog.splitlines() if l.strip()],
            )
            st.session_state["v119_pack"] = pack
            _v98_record_run("authority_pack", {"manifest": pack.get("manifest")}, mode="authority_pack_v119")
            st.success("Authority pack built.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    pack = st.session_state.get("v119_pack")
    if isinstance(pack, dict) and isinstance(pack.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download authority_pack_v119.zip",
                           data=pack["zip_bytes"],
                           file_name="authority_pack_v119.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v119_dl_zip")
        st.download_button("Download authority manifest.json",
                           data=_json.dumps(pack.get("manifest", {}), indent=2, sort_keys=True),
                           file_name="authority_pack_manifest_v119.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v119_dl_manifest")


# =====================
# v120 Constitutional panels (append-only)
# =====================
def _v120_constitution_panel():
    import streamlit as st
    from tools.constitution import build_constitution_manifest
    from ui.layer_registry import get_layers

    st.subheader("Constitution & Layer Registry")
    st.caption("This panel exposes governance + architecture documents and a cryptographic integrity manifest (SHA256).")

    # Documents
    docs = [
        ("ARCHITECTURE.md", "Architecture (Constitution)"),
        ("GOVERNANCE.md", "Governance & Release Policy"),
        ("LAYER_MODEL.md", "Layer Model"),
        ("NON_OPTIMIZER_MANIFESTO.md", "Non-Optimizer Manifesto"),
        ("CITATION.cff", "Citation (CFF)"),
    ]
    for fn, label in docs:
        try:
            b = open(fn, "rb").read()
            st.download_button(f"Download {label}", data=b, file_name=fn, use_container_width=True, key=f"v120_dl_{fn}")
        except Exception:
            st.warning(f"Missing {fn}")

    st.write("Registered higher layers (UI-accessible):")
    st.table([{"layer": e.layer, "title": e.title, "description": e.description, "panel": e.panel_fn_name} for e in get_layers()])

    man = build_constitution_manifest(repo_root=".", version="v120")
    st.download_button("Download constitution_manifest.json",
                       data=_json.dumps(man, indent=2, sort_keys=True),
                       file_name="constitution_manifest_v120.json",
                       mime="application/json",
                       use_container_width=True,
                       key="v120_dl_manifest")
    st.write("Manifest preview:")
    st.json(man)

def _v120_mission_placeholder_panel():
    import streamlit as st
    st.subheader("Mission Context (v120 placeholder)")
    st.info("Mission contexts are schema-first and will be added as additive layer panels without changing physics.")

def _v120_explainability_placeholder_panel():
    import streamlit as st
    st.subheader("Explainability (v120 placeholder)")
    st.info("Explainability narratives will be added as additive post-processing panels consuming run artifacts.")



# =====================
# v121 Mission Context Layer (L3) - additive UI panel
# =====================
def _v121_mission_context_panel():
    import streamlit as st
    from tools.mission_context import list_builtin_missions, load_mission, apply_mission_overlays, mission_report_csv

    st.subheader("Mission Context")
    st.caption("Advisory mission overlay: evaluates alignment to mission targets and reports gaps. No physics/constraints are changed.")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("No runs in ledger yet. Run a point evaluation first.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Select baseline run artifact", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v121_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    art = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if art is None:
        st.error("Selected run does not contain a run artifact payload.")
        return

    missions = list_builtin_missions("missions")
    if not missions:
        st.error("No missions found in missions/ directory.")
        return

    mfile = st.selectbox("Mission", options=missions, index=0, key="v121_mission_file")
    mission = load_mission(str(Path("missions") / mfile))

    if st.button("Generate mission report", key="v121_run"):
        try:
            rep = apply_mission_overlays(run_artifact=art, mission=mission, version="v121")
            st.session_state["v121_mission_report"] = rep
            _v98_record_run("mission_context", {"mission": mission.get("name"), "gaps_n": len(rep.get("gaps", []))}, mode="mission_context_v121")
            st.success("Mission report generated.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    rep = st.session_state.get("v121_mission_report")
    if isinstance(rep, dict) and rep.get("kind") == "shams_mission_report":
        st.write("Alignment:")
        st.write(rep.get("alignment", {}))
        st.write("Gaps:")
        st.write(rep.get("gaps", []))
        st.download_button("Download mission_report.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name=f"mission_report_{mission.get('name','mission')}_v121.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v121_dl_json")
        st.download_button("Download mission_gaps.csv",
                           data=mission_report_csv(rep),
                           file_name=f"mission_gaps_{mission.get('name','mission')}_v121.csv",
                           mime="text/csv",
                           use_container_width=True,
                           key="v121_dl_csv")


# =====================
# v122 Explainability Layer (L4) - additive UI panel
# =====================
def _v122_explainability_panel():
    import streamlit as st
    from tools.explainability import build_explainability_report

    st.subheader("Explainability")
    st.caption("Post-processing narrative: why a design fails/succeeds, limiting constraints, robustness and mission context (if available).")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("No runs in ledger yet. Run a point evaluation first.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Select baseline run artifact", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v122_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    art = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if art is None:
        st.error("Selected run does not contain a run artifact payload.")
        return

    # Optional inputs: use latest in-session mission report / envelope report if present
    mission_rep = st.session_state.get("v121_mission_report")
    env_rep = st.session_state.get("v117_report")

    use_mission = st.checkbox("Use latest in-session mission report (if available)", value=isinstance(mission_rep, dict), key="v122_use_mission")
    use_env = st.checkbox("Use latest in-session tolerance envelope (if available)", value=isinstance(env_rep, dict), key="v122_use_env")

    if st.button("Generate explainability report", key="v122_run"):
        try:
            rep = build_explainability_report(
                run_artifact=art,
                mission_report=mission_rep if use_mission else None,
                tolerance_envelope_report=env_rep if use_env else None,
                version="v122",
            )
            st.session_state["v122_explainability_report"] = rep
            _v98_record_run("explainability", {"summary": rep.get("summary")}, mode="explainability_v122")
            st.success("Explainability report generated.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    rep = st.session_state.get("v122_explainability_report")
    if isinstance(rep, dict) and rep.get("kind") == "shams_explainability_report":
        st.write("Summary:")
        st.write(rep.get("summary", {}))
        st.text_area("Narrative", value=rep.get("narrative",""), height=260, key="v122_narr_view")
        st.download_button("Download explainability_report.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="explainability_report_v122.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v122_dl_json")
        st.download_button("Download explainability_report.txt",
                           data=(rep.get("narrative","") or "").encode("utf-8"),
                           file_name="explainability_report_v122.txt",
                           mime="text/plain",
                           use_container_width=True,
                           key="v122_dl_txt")


# =====================
# v123 Evidence Graph + v123B Study Kit - additive UI panel
# =====================
def _v123_evidence_and_studykit_panel():
    import streamlit as st
    from tools.evidence_graph import build_evidence_graph, build_traceability_table, traceability_csv
    from tools.study_kit import build_study_kit_zip

    st.subheader("Evidence Graph & Design Study Kit (v123 / v123B)")
    st.caption("Build provenance graph + traceability table, and export a full publishable study kit zip (manifested with SHA256).")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("No runs in ledger yet. Run a point evaluation first.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Select baseline run artifact", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v123_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    art = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if art is None:
        st.error("Selected run does not contain a run artifact payload.")
        return

    # Optional context from session state
    mission_rep = st.session_state.get("v121_mission_report")
    env_rep = st.session_state.get("v117_report")
    expl_rep = st.session_state.get("v122_explainability_report")

    v118_bundle = st.session_state.get("v118_bundle")
    downstream_manifest = None
    downstream_zip = None
    if isinstance(v118_bundle, dict):
        downstream_manifest = v118_bundle.get("manifest")
        if isinstance(v118_bundle.get("zip_bytes"), (bytes, bytearray)):
            downstream_zip = bytes(v118_bundle["zip_bytes"])

    v119_pack = st.session_state.get("v119_pack")
    authority_manifest = None
    authority_zip = None
    if isinstance(v119_pack, dict):
        authority_manifest = v119_pack.get("manifest")
        if isinstance(v119_pack.get("zip_bytes"), (bytes, bytearray)):
            authority_zip = bytes(v119_pack["zip_bytes"])

    decision_manifest = None
    decision_zip = None
    pack_bytes = st.session_state.get("v118_pack_bytes")
    if isinstance(pack_bytes, (bytes, bytearray)):
        decision_zip = bytes(pack_bytes)
        # best-effort manifest from report if present
        rep = st.session_state.get("v118_report")
        if isinstance(rep, dict):
            decision_manifest = rep.get("decision_pack_manifest")

    use_mission = st.checkbox("Use latest in-session mission report", value=isinstance(mission_rep, dict), key="v123_use_mission")
    use_env = st.checkbox("Use latest in-session tolerance envelope", value=isinstance(env_rep, dict), key="v123_use_env")
    use_expl = st.checkbox("Use latest in-session explainability report", value=isinstance(expl_rep, dict), key="v123_use_expl")

    if st.button("Build evidence graph + traceability", key="v123_build_ev"):
        try:
            graph = build_evidence_graph(
                run_artifact=art,
                mission_report=mission_rep if use_mission else None,
                tolerance_envelope_report=env_rep if use_env else None,
                explainability_report=expl_rep if use_expl else None,
                decision_pack_manifest=decision_manifest,
                downstream_bundle_manifest=downstream_manifest,
                authority_pack_manifest=authority_manifest,
                version="v123",
            )
            tab = build_traceability_table(
                run_artifact=art,
                mission_report=mission_rep if use_mission else None,
                tolerance_envelope_report=env_rep if use_env else None,
                explainability_report=expl_rep if use_expl else None,
                version="v123",
            )
            st.session_state["v123_graph"] = graph
            st.session_state["v123_trace_table"] = tab
            _v98_record_run("evidence_graph", {"nodes": len(graph.get("nodes",[])), "edges": len(graph.get("edges",[]))}, mode="evidence_graph_v123")
            st.success("Evidence graph and traceability built.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    graph = st.session_state.get("v123_graph")
    tab = st.session_state.get("v123_trace_table")
    if isinstance(graph, dict) and graph.get("kind") == "shams_evidence_graph":
        st.download_button("Download evidence_graph.json", data=_json.dumps(graph, indent=2, sort_keys=True),
                           file_name="evidence_graph_v123.json", mime="application/json", use_container_width=True, key="v123_dl_graph")
        st.json({"nodes": len(graph.get("nodes",[])), "edges": len(graph.get("edges",[]))})

    if isinstance(tab, dict) and tab.get("kind") == "shams_traceability_table":
        st.download_button("Download traceability_table.json", data=_json.dumps(tab, indent=2, sort_keys=True),
                           file_name="traceability_table_v123.json", mime="application/json", use_container_width=True, key="v123_dl_tab")
        st.download_button("Download traceability.csv", data=traceability_csv(tab),
                           file_name="traceability_v123.csv", mime="text/csv", use_container_width=True, key="v123_dl_csv")
        st.write("Traceability rows:", len(tab.get("rows",[])))

    st.divider()
    st.write("Design Study Kit export (v123B): bundles run artifact + optional context + evidence graph + manifest.")
    if st.button("Build study kit zip", key="v123_build_kit"):
        try:
            if not (isinstance(graph, dict) and isinstance(tab, dict)):
                graph = build_evidence_graph(run_artifact=art, mission_report=mission_rep if use_mission else None,
                                             tolerance_envelope_report=env_rep if use_env else None,
                                             explainability_report=expl_rep if use_expl else None,
                                             decision_pack_manifest=decision_manifest,
                                             downstream_bundle_manifest=downstream_manifest,
                                             authority_pack_manifest=authority_manifest,
                                             version="v123")
                tab = build_traceability_table(run_artifact=art, mission_report=mission_rep if use_mission else None,
                                               tolerance_envelope_report=env_rep if use_env else None,
                                               explainability_report=expl_rep if use_expl else None,
                                               version="v123")
                st.session_state["v123_graph"] = graph
                st.session_state["v123_trace_table"] = tab

            kit = build_study_kit_zip(
                run_artifact=art,
                mission_report=mission_rep if use_mission else None,
                tolerance_envelope_report=env_rep if use_env else None,
                explainability_report=expl_rep if use_expl else None,
                evidence_graph=graph,
                traceability_table=tab,
                authority_pack_zip=authority_zip,
                optimizer_downstream_bundle_zip=downstream_zip,
                decision_pack_zip=decision_zip,
                version="v123B",
            )
            st.session_state["v123_study_kit"] = kit
            _v98_record_run("study_kit", {"files": len((kit.get("manifest") or {}).get("files", {}))}, mode="study_kit_v123B")
            st.success("Study kit built.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    kit = st.session_state.get("v123_study_kit")
    if isinstance(kit, dict) and isinstance(kit.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download study_kit_v123B.zip",
                           data=kit["zip_bytes"], file_name="study_kit_v123B.zip",
                           mime="application/zip", use_container_width=True, key="v123_dl_kit")
        st.download_button("Download study_kit_manifest_v123B.json",
                           data=_json.dumps(kit.get("manifest", {}), indent=2, sort_keys=True),
                           file_name="study_kit_manifest_v123B.json", mime="application/json",
                           use_container_width=True, key="v123_dl_kitman")


# =====================
# v124 Feasibility Boundary Atlas - additive UI panel
# =====================
def _v124_feasibility_atlas_panel():
    import streamlit as st
    from tools.feasibility_atlas import build_feasibility_atlas_bundle, available_numeric_levers

    st.subheader("Feasibility Boundary Atlas")
    st.caption("Runs a 2D grid scan around a baseline run, extracts feasibility boundary, and exports a publishable atlas bundle. Additive only (no physics/solver changes).")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("No runs in ledger yet. Run a point evaluation first.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Select baseline run artifact", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v124_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    base = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if base is None:
        st.error("Selected run does not contain a run artifact payload.")
        return

    base_inputs = base.get("inputs", {}) if isinstance(base.get("inputs"), dict) else {}
    levers = available_numeric_levers(base_inputs)
    if len(levers) < 2:
        st.error("Baseline inputs do not contain enough numeric levers for an atlas.")
        return

    col1, col2 = st.columns(2)
    with col1:
        kx = st.selectbox("Lever X", options=levers, index=0, key="v124_kx")
    with col2:
        ky = st.selectbox("Lever Y", options=levers, index=1 if len(levers)>1 else 0, key="v124_ky")

    def _default_range(v):
        try:
            v=float(v)
        except Exception:
            return (0.0, 1.0)
        if abs(v) < 1e-9:
            return (-1.0, 1.0)
        # ±20% around baseline
        return (v*0.8, v*1.2)

    x0 = base_inputs.get(kx)
    y0 = base_inputs.get(ky)
    xlo_def, xhi_def = _default_range(x0)
    ylo_def, yhi_def = _default_range(y0)

    st.write("Ranges (default ±20% around baseline):")
    r1, r2 = st.columns(2)
    with r1:
        xlo = st.number_input("X min", value=float(xlo_def), key="v124_xlo")
        xhi = st.number_input("X max", value=float(xhi_def), key="v124_xhi")
    with r2:
        ylo = st.number_input("Y min", value=float(ylo_def), key="v124_ylo")
        yhi = st.number_input("Y max", value=float(yhi_def), key="v124_yhi")

    c1, c2 = st.columns(2)
    with c1:
        nx = st.slider("Grid NX", min_value=9, max_value=61, value=25, step=2, key="v124_nx")
    with c2:
        ny = st.slider("Grid NY", min_value=9, max_value=61, value=25, step=2, key="v124_ny")

    max_evals = int(nx) * int(ny)
    st.caption(f"Total evaluations: {max_evals} (pure point evaluations + constraints).")

    if st.button("Generate atlas bundle", key="v124_run"):
        try:
            outdir = "out_feasibility_atlas_v124"
            bundle = build_feasibility_atlas_bundle(
                baseline_run_artifact=base,
                lever_x=kx,
                lever_y=ky,
                x_range=(float(xlo), float(xhi)),
                y_range=(float(ylo), float(yhi)),
                nx=int(nx),
                ny=int(ny),
                outdir=outdir,
                version="v124",
            )
            st.session_state["v124_bundle"] = bundle
            _v98_record_run("feasibility_atlas", {"lever_x": kx, "lever_y": ky, "nx": int(nx), "ny": int(ny)}, mode="feasibility_atlas_v124")
            st.success("Atlas generated.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    bundle = st.session_state.get("v124_bundle")
    if isinstance(bundle, dict) and isinstance(bundle.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download atlas_bundle_v124.zip",
                           data=bundle["zip_bytes"],
                           file_name="atlas_bundle_v124.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v124_dl_zip")
        st.download_button("Download feasibility_atlas_v124.json",
                           data=_json.dumps(bundle, indent=2, sort_keys=True),
                           file_name="feasibility_atlas_v124.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v124_dl_json")
        st.write("Atlas summary:")
        st.json({"baseline_run_id": bundle.get("baseline_run_id"),
                 "lever_x": bundle.get("lever_x"), "lever_y": bundle.get("lever_y"),
                 "grid": bundle.get("grid"),
                 "n_slices": len((bundle.get("atlas_v2") or {}).get("slices", [])) if isinstance(bundle.get("atlas_v2"), dict) else None})


# =====================
# v125 One-Click Paper Pack - additive UI panel
# =====================
def _v125_paper_pack_panel():
    import streamlit as st
    from tools.mission_context import load_mission, apply_mission_overlays
    from tools.explainability import build_explainability_report
    from tools.evidence_graph import build_evidence_graph, build_traceability_table, traceability_csv
    from tools.feasibility_atlas import build_feasibility_atlas_bundle, available_numeric_levers
    from tools.study_kit import build_study_kit_zip
    from tools.study_orchestrator import build_paper_pack_zip

    st.subheader("One-Click Paper Pack")
    st.caption("Runs post-processing pipeline (mission → explainability → evidence/traceability → atlas → study kit) and exports a single publishable zip + manifest. No physics/solver changes.")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("No runs in ledger yet. Run a point evaluation first.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Select baseline run artifact", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v125_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    base = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if base is None:
        st.error("Selected run does not contain a run artifact payload.")
        return

    st.write("Pipeline options:")
    c1,c2,c3 = st.columns(3)
    with c1:
        do_mission = st.checkbox("Mission overlay", value=True, key="v125_do_mission")
        do_expl = st.checkbox("Explainability", value=True, key="v125_do_expl")
    with c2:
        do_ev = st.checkbox("Evidence+Traceability (v123)", value=True, key="v125_do_ev")
        do_atlas = st.checkbox("Feasibility Atlas (v124)", value=False, key="v125_do_atlas")
    with c3:
        do_kit = st.checkbox("Study Kit (v123B)", value=True, key="v125_do_kit")

    mission_name = None
    mission_rep = None
    if do_mission:
        missions = ["pilot","demo","powerplant"]
        msel = st.selectbox("Mission", options=missions, index=0, key="v125_mission")
        mission_name = msel

    atlas_cfg = None
    if do_atlas:
        base_inputs = base.get("inputs", {}) if isinstance(base.get("inputs"), dict) else {}
        levers = available_numeric_levers(base_inputs)
        if len(levers) >= 2:
            kx = st.selectbox("Atlas lever X", options=levers, index=0, key="v125_kx")
            ky = st.selectbox("Atlas lever Y", options=levers, index=1, key="v125_ky")

            def _default_range(v):
                try: v=float(v)
                except Exception: return (0.0, 1.0)
                if abs(v) < 1e-9: return (-1.0, 1.0)
                return (v*0.9, v*1.1)  # narrower defaults for v125
            x0 = base_inputs.get(kx); y0 = base_inputs.get(ky)
            xlo_def,xhi_def=_default_range(x0); ylo_def,yhi_def=_default_range(y0)
            r1,r2=st.columns(2)
            with r1:
                xlo = st.number_input("X min", value=float(xlo_def), key="v125_xlo")
                xhi = st.number_input("X max", value=float(xhi_def), key="v125_xhi")
            with r2:
                ylo = st.number_input("Y min", value=float(ylo_def), key="v125_ylo")
                yhi = st.number_input("Y max", value=float(yhi_def), key="v125_yhi")
            nx = st.slider("Atlas NX", min_value=9, max_value=41, value=15, step=2, key="v125_nx")
            ny = st.slider("Atlas NY", min_value=9, max_value=41, value=15, step=2, key="v125_ny")
            atlas_cfg = {"lever_x": kx, "lever_y": ky, "x_range": (float(xlo), float(xhi)), "y_range": (float(ylo), float(yhi)), "nx": int(nx), "ny": int(ny)}
        else:
            st.warning("Not enough numeric inputs to run atlas from this baseline.")
            do_atlas = False

    if st.button("Build Paper Pack", key="v125_run"):
        try:
            # Mission report (post-processing)
            if do_mission and mission_name:
                mspec = load_mission(mission_name)
                # apply overlay to baseline; function returns report dict
                mission_rep = apply_mission_overlays(run_artifact=base, mission=mspec, version="v121")
                st.session_state["v121_mission_report"] = mission_rep
            else:
                mission_rep = None

            # Explainability
            expl_rep = None
            if do_expl:
                env_rep = st.session_state.get("v117_report")  # optional
                expl_rep = build_explainability_report(run_artifact=base, mission_report=mission_rep, tolerance_envelope_report=env_rep if isinstance(env_rep, dict) else None, version="v122")
                st.session_state["v122_explainability_report"] = expl_rep

            # Evidence + traceability
            graph = None
            tab = None
            tcsv = None
            if do_ev:
                env_rep = st.session_state.get("v117_report")
                graph = build_evidence_graph(run_artifact=base, mission_report=mission_rep, tolerance_envelope_report=env_rep if isinstance(env_rep, dict) else None, explainability_report=expl_rep, version="v123")
                tab = build_traceability_table(run_artifact=base, mission_report=mission_rep, tolerance_envelope_report=env_rep if isinstance(env_rep, dict) else None, explainability_report=expl_rep, version="v123")
                tcsv = traceability_csv(tab)
                st.session_state["v123_graph"] = graph
                st.session_state["v123_trace_table"] = tab

            # Atlas
            atlas_meta = None
            atlas_zip = None
            if do_atlas and atlas_cfg:
                bundle = build_feasibility_atlas_bundle(
                    baseline_run_artifact=base,
                    lever_x=atlas_cfg["lever_x"],
                    lever_y=atlas_cfg["lever_y"],
                    x_range=atlas_cfg["x_range"],
                    y_range=atlas_cfg["y_range"],
                    nx=atlas_cfg["nx"],
                    ny=atlas_cfg["ny"],
                    outdir="out_feasibility_atlas_v125",
                    version="v124",
                )
                atlas_zip = bundle.get("zip_bytes")
                atlas_meta = dict(bundle)
                atlas_meta.pop("zip_bytes", None)
                st.session_state["v124_bundle"] = bundle

            # Study kit
            kit_zip = None
            if do_kit:
                kit = build_study_kit_zip(
                    run_artifact=base,
                    mission_report=mission_rep,
                    tolerance_envelope_report=st.session_state.get("v117_report") if isinstance(st.session_state.get("v117_report"), dict) else None,
                    explainability_report=expl_rep,
                    evidence_graph=graph,
                    traceability_table=tab,
                    authority_pack_zip=(st.session_state.get("v119_pack") or {}).get("zip_bytes") if isinstance(st.session_state.get("v119_pack"), dict) else None,
                    optimizer_downstream_bundle_zip=(st.session_state.get("v118_bundle") or {}).get("zip_bytes") if isinstance(st.session_state.get("v118_bundle"), dict) else None,
                    decision_pack_zip=st.session_state.get("v118_pack_bytes") if isinstance(st.session_state.get("v118_pack_bytes"), (bytes, bytearray)) else None,
                    version="v123B",
                )
                kit_zip = kit.get("zip_bytes")
                st.session_state["v123_study_kit"] = kit

            # Final pack
            pack = build_paper_pack_zip(
                baseline_run_artifact=base,
                mission_report=mission_rep,
                explainability_report=expl_rep,
                evidence_graph=graph,
                traceability_table=tab,
                traceability_csv=tcsv,
                feasibility_atlas_meta=atlas_meta,
                atlas_bundle_zip=atlas_zip,
                study_kit_zip=kit_zip,
                version="v125",
            )
            st.session_state["v125_paper_pack"] = pack
            _v98_record_run("paper_pack", {"files": len((pack.get("manifest") or {}).get("files", {}))}, mode="paper_pack_v125")
            st.success("Paper pack built.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    pack = st.session_state.get("v125_paper_pack")
    if isinstance(pack, dict) and isinstance(pack.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download paper_pack_v125.zip",
                           data=pack["zip_bytes"],
                           file_name="paper_pack_v125.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v125_dl_zip")
        st.download_button("Download paper_pack_manifest_v125.json",
                           data=_json.dumps(pack.get("manifest", {}), indent=2, sort_keys=True),
                           file_name="paper_pack_manifest_v125.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v125_dl_manifest")
        st.write("Pack summary:")
        st.json({"files": len((pack.get("manifest") or {}).get("files", {})),
                 "created_utc": pack.get("created_utc")})


# =====================
# v126 UI Smoke & Diagnostics - additive UI panel
# =====================
def _v126_ui_smoke_panel():
    import streamlit as st
    from pathlib import Path as _Path

    st.subheader("UI Smoke & Diagnostics")
    st.caption("Runs lightweight UI smoke checks and produces a report. This does not change physics/solvers.")

    scenarios = st.multiselect("Scenarios", options=["render_all", "paper_pack"], default=["render_all"], key="v126_scenarios")
    outdir = st.text_input("Output directory", value="out_ui_smoke_v126", key="v126_outdir")

    if st.button("Run UI smoke checks", key="v126_run"):
        try:
            cmd = [_sys.executable, "-m", "tools.cli_ui_smoke", "--outdir", outdir] + [ "--scenarios"] + list(scenarios)
            # note: argparse expects --scenarios then list; we pass that
            p = _subprocess.run(cmd, cwd=str(_Path(__file__).resolve().parents[1]), stdout=_subprocess.PIPE, stderr=_subprocess.STDOUT, text=True, timeout=300)
            st.text_area("Runner output", value=p.stdout, height=180, key="v126_out")
            rep_path = _Path(__file__).resolve().parents[1] / outdir / "ui_smoke_report.json"
            if rep_path.exists():
                rep = _json.loads(rep_path.read_text(encoding="utf-8"))
                st.session_state["v126_smoke_report"] = rep
                st.success("Smoke report generated.")
            else:
                st.warning("Report file not found after run.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    rep = st.session_state.get("v126_smoke_report")
    if isinstance(rep, dict):
        st.write("Summary:")
        st.json(rep.get("summary", {}))
        st.download_button("Download ui_smoke_report.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="ui_smoke_report_v126.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v126_dl_json")


# =====================
# v127 Study Matrix - Batch Paper Packs (single UI)
# =====================
def _v127_study_matrix_panel():
    import streamlit as st
    from tools.study_matrix import build_cases_1d_sweep, build_study_matrix_bundle
    from tools.feasibility_atlas import available_numeric_levers

    st.subheader("Study Matrix + Batch Paper Packs")
    st.caption("Build multiple cases from a baseline and export a single study zip with per-case paper packs + index. Additive; no solver behavior changes.")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("No runs in ledger yet. Run a point evaluation first.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Select baseline run artifact", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v127_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    base = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if base is None:
        st.error("Selected run does not contain a run artifact payload.")
        return

    base_inputs = base.get("inputs", {}) if isinstance(base.get("inputs"), dict) else {}
    levers = available_numeric_levers(base_inputs)
    if not levers:
        st.warning("No numeric levers detected in baseline inputs.")
        return

    st.write("1D sweep builder (safe defaults):")
    c1,c2,c3 = st.columns(3)
    with c1:
        lever = st.selectbox("Sweep lever", options=levers, index=0, key="v127_lever")
    with c2:
        v0 = float(base_inputs.get(lever, 0.0) or 0.0)
        vmin = st.number_input("Min", value=(v0*0.95 if abs>1e-9 else -1.0), key="v127_vmin")
        vmax = st.number_input("Max", value=(v0*1.05 if abs>1e-9 else 1.0), key="v127_vmax")
    with c3:
        n = st.slider("N cases", min_value=3, max_value=15, value=5, step=1, key="v127_n")

    missions = st.multiselect("Missions (optional)", options=["pilot","demo","powerplant"], default=["pilot"], key="v127_missions")
    include_expl = st.checkbox("Include explainability (v122)", value=True, key="v127_expl")
    include_ev = st.checkbox("Include evidence/traceability (v123)", value=True, key="v127_ev")
    include_kit = st.checkbox("Include study kit (v123B)", value=True, key="v127_kit")

    if st.button("Build Study Matrix Zip", key="v127_run"):
        try:
            if int(n) < 2:
                raise ValueError("N must be >= 2")
            # linspace
            vals = [float(vmin) + (float(vmax)-float(vmin))*i/(int(n)-1) for i in range(int(n))]
            cases = build_cases_1d_sweep(baseline_run_artifact=base, lever=str(lever), values=vals, missions=(missions or None))
            bundle = build_study_matrix_bundle(
                baseline_run_artifact=base,
                cases=cases,
                outdir="out_study_matrix_v127",
                version="v127",
                include_explainability=bool(include_expl),
                include_evidence=bool(include_ev),
                include_study_kit=bool(include_kit),
            )
            st.session_state["v127_bundle"] = bundle
            _v98_record_run("study_matrix", {"cases": bundle.get("cases")}, mode="study_matrix_v127")
            st.success(f"Study matrix built: {bundle.get('cases')} cases.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    bun = st.session_state.get("v127_bundle")
    if isinstance(bun, dict) and isinstance(bun.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download study_matrix_v127.zip",
                           data=bun["zip_bytes"],
                           file_name="study_matrix_v127.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v127_dl_zip")
        st.download_button("Download study_matrix_manifest_v127.json",
                           data=_json.dumps(bun.get("manifest", {}), indent=2, sort_keys=True),
                           file_name="study_matrix_manifest_v127.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v127_dl_manifest")
        st.json({"cases": bun.get("cases"), "created_utc": bun.get("created_utc")})


# =====================
# v128 Study Explorer + Comparator (single UI)
# =====================
def _v128_study_explorer_panel():
    import streamlit as st
    from tools.study_explorer import load_study_zip, parse_study_index, filter_cases, load_case_run_artifact, compare_two_runs

    st.subheader("Study Explorer + Comparator")
    st.caption("Load a study_matrix zip and browse/filter cases, compare two cases, and export comparison JSON. Downstream-only.")

    up = st.file_uploader("Upload study_matrix zip", type=["zip"], key="v128_upl")
    if not up:
        st.info("Upload a v127 study_matrix zip (study_matrix_v127.zip).")
        return

    try:
        # write to temp bytes map
        import tempfile, os
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
        tmp.write(up.getvalue()); tmp.flush(); tmp.close()
        files = load_study_zip(tmp.name)
        idx = parse_study_index(files)
        st.session_state["v128_files"] = files
        st.session_state["v128_index"] = {"created_utc": idx.created_utc, "rows": idx.rows}
        os.unlink(tmp.name)
    except Exception as e:
        st.error(f"Failed to load: {e!r}")
        return

    idxd = st.session_state.get("v128_index") or {}
    rows = list(idxd.get("rows") or [])
    if not rows:
        st.warning("No rows found in study index.")
        return

    missions = sorted({str(r.get("mission","")) for r in rows if str(r.get("mission",""))})
    feasible_only = st.checkbox("Feasible only", value=False, key="v128_feas")
    mission = st.selectbox("Mission filter", options=["(any)"] + missions, index=0, key="v128_msel")

    # simple KPI filters
    kf = {}
    with st.expander("KPI filters (optional)"):
        c1,c2,c3 = st.columns(3)
        with c1:
            qlo = st.number_input("Q min", value=0.0, key="v128_qlo")
            kf["Q"] = (float(qlo), None)
        with c2:
            pnlo = st.number_input("Pnet_MW min", value=-1e9, key="v128_pnlo")
            kf["Pnet_MW"] = (float(pnlo), None)
        with c3:
            pflo = st.number_input("Pfus_MW min", value=0.0, key="v128_pflo")
            kf["Pfus_MW"] = (float(pflo), None)

    idx_obj = parse_study_index(st.session_state["v128_files"])
    fr = filter_cases(idx_obj, feasible_only=feasible_only, mission=None if mission=="(any)" else mission, kpi_filters=kf)

    st.write(f"Filtered cases: {len(fr)} / {len(rows)}")
    show = fr[:200]  # avoid UI overload
    st.dataframe(show, use_container_width=True)

    case_ids = [str(r.get("case_id")) for r in fr if r.get("case_id") is not None]
    if len(case_ids) < 2:
        st.info("Need at least 2 filtered cases to compare.")
        return

    a_id = st.selectbox("Case A", options=case_ids, index=0, key="v128_a")
    b_id = st.selectbox("Case B", options=case_ids, index=1 if len(case_ids)>1 else 0, key="v128_b")

    if st.button("🆚 Compare", key="v128_compare"):
        try:
            files = st.session_state["v128_files"]
            a_row = next(r for r in fr if str(r.get("case_id"))==a_id)
            b_row = next(r for r in fr if str(r.get("case_id"))==b_id)
            a_art = load_case_run_artifact(files, a_row)
            b_art = load_case_run_artifact(files, b_row)
            comp = compare_two_runs(a_art, b_art)
            st.session_state["v128_comp"] = comp
            st.success("Comparison built.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    comp = st.session_state.get("v128_comp")
    if isinstance(comp, dict):
        st.write("Comparison summary:")
        st.json(comp)
        st.download_button("Download comparison_v128.json",
                           data=_json.dumps(comp, indent=2, sort_keys=True),
                           file_name="comparison_v128.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v128_dl_comp")


# =====================
# v129 Pareto from Study (single UI)
# =====================
def _v129_pareto_panel():
    import streamlit as st
    from tools.pareto_from_study import build_pareto, pareto_bundle_zip

    st.subheader("Pareto from Study")
    st.caption("Compute Pareto fronts from a study_matrix zip. No optimizer. Downstream-only, publishable.")

    up = st.file_uploader("Upload study_matrix zip", type=["zip"], key="v129_upl")
    if not up:
        st.info("Upload a v127 study_matrix zip (study_matrix_v127.zip).")
        return

    # Save upload to temp file for existing loaders
    import tempfile, os
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    tmp.write(up.getvalue()); tmp.flush(); tmp.close()
    study_path = tmp.name

    st.write("Objectives (choose senses):")
    c1,c2,c3 = st.columns(3)
    with c1:
        obj1 = st.text_input("Objective 1 column", value="Q", key="v129_o1")
        s1 = st.selectbox("Sense 1", options=["max","min"], index=0, key="v129_s1")
    with c2:
        obj2 = st.text_input("Objective 2 column", value="Pnet_MW", key="v129_o2")
        s2 = st.selectbox("Sense 2", options=["max","min"], index=0, key="v129_s2")
    with c3:
        obj3 = st.text_input("Objective 3 column (optional)", value="R0_m", key="v129_o3")
        s3 = st.selectbox("Sense 3", options=["min","max"], index=0, key="v129_s3")

    feasible_only = st.checkbox("Feasible only", value=True, key="v129_feas")
    mission = st.selectbox("Mission filter", options=["(any)","pilot","demo","powerplant"], index=0, key="v129_mis")

    objs=[{"k": obj1.strip(), "sense": s1}]
    if obj2.strip():
        objs.append({"k": obj2.strip(), "sense": s2})
    if obj3.strip():
        objs.append({"k": obj3.strip(), "sense": s3})

    if st.button("Compute Pareto", key="v129_run"):
        try:
            rep = build_pareto(study_path=study_path, objectives=objs, feasible_only=bool(feasible_only), mission=None if mission=="(any)" else mission, version="v129")
            bun = pareto_bundle_zip(rep)
            st.session_state["v129_rep"]=rep
            st.session_state["v129_bun"]=bun
            st.success(f"Computed Pareto: {rep.get('n_filtered')} rows, layers: {len(rep.get('layer_counts') or {})}")
        except Exception as e:
            st.error(f"Failed: {e!r}")
        finally:
            try: os.unlink(study_path)
            except Exception: pass

    rep = st.session_state.get("v129_rep")
    bun = st.session_state.get("v129_bun")
    if isinstance(rep, dict):
        st.write("Summary:")
        st.json({"n_filtered": rep.get("n_filtered"), "layer_counts": rep.get("layer_counts"), "objectives": rep.get("objectives")})
        st.dataframe((rep.get("rows") or [])[:200], use_container_width=True)
        st.download_button("Download pareto_report_v129.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="pareto_report_v129.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v129_dl_rep")
    if isinstance(bun, dict) and isinstance(bun.get("zip_bytes"), (bytes,bytearray)):
        st.download_button("Download pareto_bundle_v129.zip",
                           data=bun["zip_bytes"],
                           file_name="pareto_bundle_v129.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v129_dl_zip")


# =====================
# v130 Persistent Run Vault (single UI)
# =====================
def _v130_run_vault_panel():
    import streamlit as st
    from pathlib import Path
    from tools.run_vault import list_entries

    st.subheader("Persistent Run Vault")
    st.caption("Append-only, local vault of run artifacts and bundles. Prevents loss due to reruns/downloads. Storage-only; never changes physics/solvers.")

    c1,c2 = st.columns(2)
    with c1:
        enabled = st.toggle("Enable vault persistence", value=bool(st.session_state.get("vault_enabled", True)), key="vault_enabled")
    with c2:
        limit = st.slider("Show last N entries", min_value=10, max_value=200, value=int(st.session_state.get("vault_limit", 50)), step=10, key="vault_limit")

    root = Path(__file__).resolve().parents[1]
    entries = list_entries(root, limit=int(limit))
    st.write(f"Vault entries: showing last {len(entries)}")
    if entries:
        st.dataframe(entries, use_container_width=True)

        # allow downloading index file
        idx_path = root / "out_run_vault" / "INDEX.jsonl"
        if idx_path.exists():
            st.download_button("Download vault INDEX.jsonl",
                               data=idx_path.read_bytes(),
                               file_name="INDEX.jsonl",
                               mime="text/plain",
                               use_container_width=True,
                               key="v130_dl_index")
    else:
        st.info("No entries yet. Run a point evaluation or generate a bundle; entries will appear here.")


# =====================
# v131 Vault Restore + Session Replay (single UI)
# =====================
def _v131_vault_restore_panel():
    import streamlit as st
    from pathlib import Path
    from tools.vault_restore import list_entries, load_entry_payload, list_entry_files, read_entry_file

    st.subheader("Vault Restore + Session Replay")
    st.caption("Restore runs and bundles from the persistent vault back into the UI ledger. Downstream-only; no evaluation is performed.")

    root = Path(__file__).resolve().parents[1]
    entries = list_entries(root, limit=int(st.session_state.get("vault_limit", 50)))
    if not entries:
        st.info("No vault entries found yet. Enable the vault and run evaluations/bundles first.")
        return

    # select entry
    labels = []
    for e in entries:
        labels.append(f"{e.get('created_utc','')} | {e.get('record_kind','')} | {e.get('mode','')} | {str(e.get('sha256',''))[:10]}")
    idx = st.selectbox("Select vault entry", options=list(range(len(entries))), format_func=lambda i: labels[i], key="v131_pick")
    meta = entries[int(idx)]

    st.write("Entry meta:")
    st.json(meta)

    c1,c2,c3 = st.columns(3)
    with c1:
        if st.button("Restore payload to Run Ledger", key="v131_restore"):
            try:
                payload = load_entry_payload(root, meta)
                kind = str(meta.get("record_kind") or "vault_restore")
                mode = str(meta.get("mode") or "vault_restore_v131")
                _v98_record_run(kind, payload, mode=mode)
                st.success("Restored into run ledger.")
            except Exception as e:
                st.error(f"Restore failed: {e!r}")

    with c2:
        if st.button("Load payload into viewer (no restore)", key="v131_load"):
            try:
                payload = load_entry_payload(root, meta)
                st.session_state["v131_loaded_payload"] = payload
                st.success("Loaded.")
            except Exception as e:
                st.error(f"Load failed: {e!r}")

    with c3:
        show_files = st.toggle("Show attached files", value=False, key="v131_show_files")

    payload = st.session_state.get("v131_loaded_payload")
    if payload is not None:
        st.write("Loaded payload preview:")
        if isinstance(payload, (bytes, bytearray)):
            st.write(f"Binary payload: {len(payload)} bytes")
        else:
            try:
                st.json(payload)
            except Exception:
                st.write(repr(payload)[:2000])

    if show_files:
        fnames = list_entry_files(root, meta)
        if not fnames:
            st.info("No attached files for this entry.")
        else:
            st.write("Attached files:")
            for fn in fnames:
                b = read_entry_file(root, meta, fn)
                st.download_button(f"Download {fn}", data=b, file_name=fn, use_container_width=True, key=f"v131_dl_{fn}")


# =====================
# v132 Study Matrix Builder v2 (multi-lever sweep) (single UI)
# =====================
def _v132_study_matrix_v2_panel():
    import streamlit as st
    from tools.study_matrix_v2 import build_cases_multi_sweep, build_study_matrix_bundle_v2
    from tools.feasibility_atlas import available_numeric_levers

    st.subheader("Study Matrix Builder v2")
    st.caption("Multi-lever sweeps (2D/3D) with derived columns. Produces a study zip compatible with Study Explorer and Pareto.")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("No runs in ledger yet. Run a point evaluation first.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Select baseline run artifact", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v132_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    base = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if base is None:
        st.error("Selected run does not contain a run artifact payload.")
        return

    base_inputs = base.get("inputs", {}) if isinstance(base.get("inputs"), dict) else {}
    levers = available_numeric_levers(base_inputs)
    if not levers:
        st.warning("No numeric levers detected in baseline inputs.")
        return

    st.write("Define up to 3 sweeps (leave lever blank to disable a sweep):")
    sweeps=[]
    cols = st.columns(3)
    for i in range(3):
        with cols[i]:
            lever = st.selectbox(f"Sweep {i+1} lever", options=["(none)"]+levers, index=0, key=f"v132_lev_{i}")
            if lever != "(none)":
                v0 = float(base_inputs.get(lever, 0.0) or 0.0)
                vmin = st.number_input(f"Min {i+1}", value=(v0*0.95 if abs>1e-9 else -1.0), key=f"v132_min_{i}")
                vmax = st.number_input(f"Max {i+1}", value=(v0*1.05 if abs>1e-9 else 1.0), key=f"v132_max_{i}")
                n = st.slider(f"N {i+1}", min_value=2, max_value=10, value=4, step=1, key=f"v132_n_{i}")
                sweeps.append({"lever": str(lever), "min": float(vmin), "max": float(vmax), "n": int(n)})

    if not sweeps:
        st.info("Choose at least one sweep lever.")
        return

    missions = st.multiselect("Missions (optional)", options=["pilot","demo","powerplant"], default=["pilot"], key="v132_missions")
    derived = st.multiselect("Derived columns", options=["Pnet_per_R0","Q_per_Bt","margin_penalty"], default=["Pnet_per_R0","Q_per_Bt","margin_penalty"], key="v132_derived")

    include_expl = st.checkbox("Include explainability", value=True, key="v132_expl")
    include_ev = st.checkbox("Include evidence/traceability (v123)", value=True, key="v132_ev")
    include_kit = st.checkbox("Include study kit (v123B)", value=True, key="v132_kit")

    est = 1
    for sw in sweeps:
        est *= int(sw.get("n",1))
    est *= max(1, len(missions) if missions else 1)
    st.warning(f"Estimated cases: {est}. Keep it small for local runs.")

    if st.button("Build Study Matrix v132 Zip", key="v132_run"):
        try:
            cases = build_cases_multi_sweep(baseline_run_artifact=base, sweeps=sweeps, missions=(missions or None))
            bundle = build_study_matrix_bundle_v2(
                baseline_run_artifact=base,
                cases=cases,
                outdir="out_study_matrix_v132",
                version="v132",
                include_explainability=bool(include_expl),
                include_evidence=bool(include_ev),
                include_study_kit=bool(include_kit),
                derived=list(derived or []),
            )
            st.session_state["v132_bundle"] = bundle
            _v98_record_run("study_matrix_v2", {"cases": bundle.get("cases")}, mode="study_matrix_v132")
            st.success(f"Study matrix built: {bundle.get('cases')} cases.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    bun = st.session_state.get("v132_bundle")
    if isinstance(bun, dict) and isinstance(bun.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download study_matrix_v132.zip",
                           data=bun["zip_bytes"],
                           file_name="study_matrix_v132.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v132_dl_zip")
        st.download_button("Download study_matrix_manifest_v132.json",
                           data=_json.dumps(bun.get("manifest", {}), indent=2, sort_keys=True),
                           file_name="study_matrix_manifest_v132.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v132_dl_manifest")
        st.json({"cases": bun.get("cases"), "created_utc": bun.get("created_utc"), "version": bun.get("version")})


# =====================
# v133 Feasibility Completion (Partial Design Inference) (single UI)
# =====================
def _v133_fc_panel():
    import streamlit as st
    from tools.feasibility_completion import FCConfig, run_feasibility_completion, build_fc_bundle_zip
    from tools.feasibility_atlas import available_numeric_levers
    from tools import run_vault
    from pathlib import Path

    st.subheader("Feasibility Completion")
    st.caption("Partial design inference: fix a few major parameters, mark others as FREE/UNCERTAIN within bounds, and search for any feasible completions. Not an optimizer.")

    # baseline for lever discovery + default values
    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("Run a point evaluation first so SHAMS has a baseline inputs dictionary to start from.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Baseline run (for defaults)", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v133_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    base = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if base is None:
        st.error("Selected baseline does not contain a run artifact payload.")
        return

    base_inputs = base.get("inputs", {}) if isinstance(base.get("inputs"), dict) else {}
    levers = available_numeric_levers(base_inputs)
    if not levers:
        st.warning("No numeric levers detected.")
        return

    st.write("1) Choose FIXED parameters (set values).")
    fixed_choices = st.multiselect("Fixed parameters", options=levers, default=[x for x in ["R0_m","Bt_T","Ip_MA"] if x in levers], key="v133_fixed_list")
    fixed = {}
    if fixed_choices:
        cols = st.columns(min(3, len(fixed_choices)))
        for i,k in enumerate(fixed_choices):
            with cols[i % len(cols)]:
                v0 = float(base_inputs.get(k, 0.0) or 0.0)
                fixed[k] = st.number_input(f"Fixed {k}", value=v0, key=f"v133_fix_{k}")

    st.write("2) Choose FREE parameters (SHAMS searches within bounds). Up to 3 recommended.")
    free_choices = st.multiselect("Free parameters", options=[x for x in levers if x not in fixed_choices], default=[x for x in ["kappa","q95"] if x in levers and x not in fixed_choices], key="v133_free_list")

    st.write("3) Choose UNCERTAIN parameters (sampled within bounds). Optional.")
    uncertain_choices = st.multiselect("Uncertain parameters", options=[x for x in levers if x not in fixed_choices and x not in free_choices], default=[], key="v133_unc_list")

    vars_all = list(dict.fromkeys(list(free_choices) + list(uncertain_choices)))
    if not vars_all:
        st.info("Select at least one FREE or UNCERTAIN parameter.")
        return

    bounds = {}
    st.write("4) Bounds for FREE/UNCERTAIN parameters:")
    for k in vars_all:
        v0 = float(base_inputs.get(k, 0.0) or 0.0)
        c1,c2,c3 = st.columns([1,1,1])
        with c1:
            lo = st.number_input(f"{k} min", value=(v0*0.95 if abs>1e-9 else -1.0), key=f"v133_lo_{k}")
        with c2:
            hi = st.number_input(f"{k} max", value=(v0*1.05 if abs>1e-9 else 1.0), key=f"v133_hi_{k}")
        with c3:
            st.caption(f"baseline={v0:g}")
        bounds[k] = (float(lo), float(hi))

    st.write("5) Search method and budget:")
    method = st.selectbox("Method", options=["grid","random"], index=0, key="v133_method")
    if method == "grid":
        n_per_dim = st.slider("Grid points per dimension", min_value=2, max_value=10, value=4, step=1, key="v133_npd")
        est = 1
        for _ in vars_all: est *= int(n_per_dim)
        st.warning(f"Estimated evaluations: {est}")
        n_random = 0
    else:
        n_random = st.slider("Random samples", min_value=50, max_value=2000, value=300, step=50, key="v133_nr")
        n_per_dim = 0
        st.warning(f"Estimated evaluations: {n_random}")

    seed = st.number_input("Seed (determinism)", value=0, step=1, key="v133_seed")
    feasible_only_export = st.checkbox("Export only feasible evaluations", value=True, key="v133_fe_only")

    if st.button("Run Feasibility Completion", key="v133_run"):
        try:
            cfg = FCConfig(
                baseline_inputs=dict(base_inputs),
                fixed=dict(fixed),
                bounds=dict(bounds),
                free=list(free_choices),
                uncertain=list(uncertain_choices),
                method=str(method),
                n_per_dim=int(n_per_dim) if method=="grid" else 0,
                n_random=int(n_random) if method=="random" else 0,
                seed=int(seed),
                feasible_only_export=bool(feasible_only_export),
            )
            rep = run_feasibility_completion(cfg)
            bun = build_fc_bundle_zip(rep)
            st.session_state["v133_rep"] = rep
            st.session_state["v133_bun"] = bun
            _v98_record_run("feasibility_completion", {"exists_feasible": rep.get("exists_feasible"), "n_evals": rep.get("n_evals"), "n_feasible": rep.get("n_feasible")}, mode="fc_v133")
            st.success(f"Done. Feasible: {rep.get('exists_feasible')} (feasible points: {rep.get('n_feasible')}/{rep.get('n_evals')})")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    rep = st.session_state.get("v133_rep")
    bun = st.session_state.get("v133_bun")
    if isinstance(rep, dict):
        st.write("Report summary:")
        st.json({"exists_feasible": rep.get("exists_feasible"),
                 "n_evals": rep.get("n_evals"),
                 "n_feasible": rep.get("n_feasible"),
                 "envelope": rep.get("envelope"),
                 "dominant_constraint_counts": rep.get("dominant_constraint_counts")})
        st.dataframe((rep.get("evaluations") or [])[:200], use_container_width=True)
        st.download_button("Download fc_report_v133.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True),
                           file_name="fc_report_v133.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v133_dl_rep")

    if isinstance(bun, dict) and isinstance(bun.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download fc_bundle_v133.zip",
                           data=bun["zip_bytes"],
                           file_name="fc_bundle_v133.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v133_dl_zip")

        if st.button("Save FC bundle to Vault (attach zip)", key="v133_save_vault"):
            try:
                root = Path(__file__).resolve().parents[1]
                meta = run_vault.write_entry(root=root, kind="fc_bundle", payload=rep or {}, mode="v133", files={"fc_bundle_v133.zip": bun["zip_bytes"]})
                st.success(f"Saved to vault: {meta.get('entry_dir')}")
            except Exception as e:
                st.error(f"Vault save failed: {e!r}")


# =====================
# v134–v138 FC Superpanel (single UI)
# =====================
def _v138_fc_superpanel():
    import streamlit as st
    from pathlib import Path
    from tools import run_vault
    from tools.param_guidance import suggest_free_vars, suggest_bounds, sanity_check_bounds
    from tools.feasibility_completion import FCConfig, run_feasibility_completion, build_fc_bundle_zip
    from tools.fc_advanced import build_fc_atlas_bundle, repair_to_feasibility, RepairConfig, compress_feasible_set, completion_to_run_artifact

    st.subheader("Feasibility Completion Advanced (v134–v138)")
    st.caption("Atlas + guided setup + bounded repair + compression + handoff to study tools. All downstream/orchestration only.")

    # baseline for defaults
    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("Run a point evaluation first to populate the run ledger.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Baseline run (defaults)", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v138_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    base = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if base is None:
        st.error("Selected baseline does not contain a run artifact.")
        return
    base_inputs = base.get("inputs", {}) if isinstance(base.get("inputs"), dict) else {}

    # Use the v133 panel state if present
    rep = st.session_state.get("v133_rep")
    bun = st.session_state.get("v133_bun")

    tabs = st.tabs(["Guided Setup", "Run FC", "Atlas", "Repair (v136)", "Compress (v137)", "Handoff (v138)"])

    # ---------------- v135 ----------------
    with tabs[0]:
        st.write("Suggested FREE variables (heuristics):")
        avail = list(available_numeric_levers(base_inputs))
        fixed_guess = [x for x in ["R0_m","Bt_T"] if x in avail]
        suggested = suggest_free_vars(avail, fixed=fixed_guess, max_k=3)
        st.json({"suggested_free": suggested})
        st.write("Suggested bounds (editable):")
        b = {}
        for v in suggested:
            lo,hi = suggest_bounds(base_inputs, v)
            b[v] = [lo,hi]
            msgs = sanity_check_bounds(v, lo, hi)
            st.write(f"- {v}: [{lo:g}, {hi:g}]" + (f" ⚠ {', '.join(msgs)}" if msgs else ""))
        st.caption("You can copy these into the FC run tab.")

    # ---------------- v133 run ----------------
    with tabs[1]:
        st.write("Run Feasibility Completion using baseline defaults + your fixed/free/uncertain selections.")
        levers = list(available_numeric_levers(base_inputs))
        fixed_choices = st.multiselect("Fixed parameters", options=levers, default=[x for x in ["R0_m","Bt_T"] if x in levers], key="v138_fixed_list")
        fixed = {}
        for k in fixed_choices:
            v0 = float(base_inputs.get(k, 0.0) or 0.0)
            fixed[k] = st.number_input(f"Fixed {k}", value=v0, key=f"v138_fix_{k}")

        free_choices = st.multiselect("Free parameters", options=[x for x in levers if x not in fixed_choices], default=[x for x in ["Ip_MA","kappa","q95"] if x in levers and x not in fixed_choices], key="v138_free_list")
        uncertain_choices = st.multiselect("Uncertain parameters", options=[x for x in levers if x not in fixed_choices and x not in free_choices], default=[], key="v138_unc_list")
        vars_all = list(dict.fromkeys(list(free_choices)+list(uncertain_choices)))
        if vars_all:
            st.write("Bounds:")
        bounds={}
        for k in vars_all:
            lo0,hi0 = suggest_bounds(base_inputs, k)
            c1,c2 = st.columns(2)
            with c1:
                lo = st.number_input(f"{k} min", value=float(lo0), key=f"v138_lo_{k}")
            with c2:
                hi = st.number_input(f"{k} max", value=float(hi0), key=f"v138_hi_{k}")
            bounds[k]=(float(lo), float(hi))
            msgs = sanity_check_bounds(k, float(lo), float(hi))
            if msgs:
                st.warning(f"{k}: {', '.join(msgs)}")

        method = st.selectbox("Method", options=["grid","random"], index=0, key="v138_method")
        if method=="grid":
            n_per_dim = st.slider("Grid points per dimension", 2, 10, 4, 1, key="v138_npd")
            n_random = 0
        else:
            n_random = st.slider("Random samples", 50, 2000, 300, 50, key="v138_nr")
            n_per_dim = 0
        seed = st.number_input("Seed", value=0, step=1, key="v138_seed")
        feasible_only = st.checkbox("Export only feasible points", value=True, key="v138_fe_only")

        if st.button("Run FC", key="v138_run_fc"):
            cfg = FCConfig(
                baseline_inputs=dict(base_inputs),
                fixed=dict(fixed),
                bounds=dict(bounds),
                free=list(free_choices),
                uncertain=list(uncertain_choices),
                method=str(method),
                n_per_dim=int(n_per_dim) if method=="grid" else 0,
                n_random=int(n_random) if method=="random" else 0,
                seed=int(seed),
                feasible_only_export=bool(feasible_only),
            )
            rep = run_feasibility_completion(cfg)
            bun = build_fc_bundle_zip(rep)
            st.session_state["v133_rep"] = rep
            st.session_state["v133_bun"] = bun
            _v98_record_run("feasibility_completion", {"exists_feasible": rep.get("exists_feasible"), "n_evals": rep.get("n_evals"), "n_feasible": rep.get("n_feasible")}, mode="fc_v133")
            st.success(f"Feasible: {rep.get('exists_feasible')} (feasible points: {rep.get('n_feasible')}/{rep.get('n_evals')})")

        rep = st.session_state.get("v133_rep")
        bun = st.session_state.get("v133_bun")
        if isinstance(rep, dict):
            st.json({"exists_feasible": rep.get("exists_feasible"), "n_evals": rep.get("n_evals"), "n_feasible": rep.get("n_feasible"), "envelope": rep.get("envelope")})
            st.download_button("Download fc_report_v133.json", data=_json.dumps(rep, indent=2, sort_keys=True), file_name="fc_report_v133.json", mime="application/json", use_container_width=True)
        if isinstance(bun, dict) and isinstance(bun.get("zip_bytes"), (bytes, bytearray)):
            st.download_button("Download fc_bundle_v133.zip", data=bun["zip_bytes"], file_name="fc_bundle_v133.zip", mime="application/zip", use_container_width=True)
            if st.button("Save FC bundle to Vault", key="v138_save_fc_vault"):
                root = Path(__file__).resolve().parents[1]
                run_vault.write_entry(root=root, kind="fc_bundle", payload=rep or {}, mode="v138", files={"fc_bundle_v133.zip": bun["zip_bytes"]})
                st.success("Saved.")

    # ---------------- v134 atlas ----------------
    with tabs[2]:
        rep = st.session_state.get("v133_rep")
        if not isinstance(rep, dict):
            st.info("Run FC first (Run FC tab).")
        else:
            bounds = rep.get("config", {}).get("bounds", {}) if isinstance(rep.get("config"), dict) else {}
            vars_ = list(bounds.keys())
            if len(vars_) < 2:
                st.info("Need at least two bounded variables in FC config to build an atlas.")
            else:
                x = st.selectbox("X axis", options=vars_, index=0, key="v138_atlas_x")
                y = st.selectbox("Y axis", options=[v for v in vars_ if v != x], index=0, key="v138_atlas_y")
                if st.button("Build FC Atlas", key="v138_build_atlas"):
                    try:
                        bun_at = build_fc_atlas_bundle(rep, x_var=str(x), y_var=str(y))
                        st.session_state["v134_atlas"] = bun_at
                        _v98_record_run("fc_atlas", {"x": x, "y": y, "n": bun_at.get("manifest", {}).get("files", {}).get("atlas_points.csv", {}).get("bytes")}, mode="v134")
                        st.success("Atlas built.")
                    except Exception as e:
                        st.error(f"Atlas failed: {e!r}")

                bun_at = st.session_state.get("v134_atlas")
                if isinstance(bun_at, dict) and isinstance(bun_at.get("zip_bytes"), (bytes, bytearray)):
                    st.download_button("Download fc_atlas_v134.zip", data=bun_at["zip_bytes"], file_name="fc_atlas_v134.zip", mime="application/zip", use_container_width=True, key="v138_dl_atlas")
                    st.json(bun_at.get("manifest", {}))

    # ---------------- v136 repair ----------------
    with tabs[3]:
        rep = st.session_state.get("v133_rep")
        if not isinstance(rep, dict):
            st.info("Run FC first (Run FC tab).")
        else:
            evals = list(rep.get("evaluations") or [])
            infeas = [r for r in evals if r.get("feasible") is not True]
            cfg0 = rep.get("config", {}) if isinstance(rep.get("config"), dict) else {}
            bounds = cfg0.get("bounds", {}) if isinstance(cfg0.get("bounds"), dict) else {}
            free = list(cfg0.get("free") or [])
            if not free or not bounds:
                st.info("Repair requires FREE variables with bounds.")
            elif not infeas:
                st.success("No infeasible points in current export. Disable 'feasible-only export' if you want to repair infeasible samples.")
            else:
                st.write(f"Infeasible points available: {len(infeas)}")
                idx = st.slider("Pick infeasible sample index", 0, len(infeas)-1, 0, 1, key="v138_rep_idx")
                start = infeas[int(idx)].get("inputs", {}) if isinstance(infeas[int(idx)].get("inputs"), dict) else {}
                max_steps = st.slider("Max steps", 5, 30, 12, 1, key="v138_rep_steps")
                step_frac = st.slider("Step fraction", 0.05, 0.50, 0.15, 0.05, key="v138_rep_frac")
                seed = st.number_input("Repair seed", value=0, step=1, key="v138_rep_seed")
                if st.button("Run bounded repair", key="v138_run_repair"):
                    try:
                        bnd = {k: (float(v[0]), float(v[1])) for k,v in bounds.items()}
                        cfg = RepairConfig(bounds=bnd, free=list(free), max_steps=int(max_steps), step_frac=float(step_frac), seed=int(seed))
                        tr = repair_to_feasibility(baseline_inputs=dict(base_inputs), start_inputs=dict(start), cfg=cfg)
                        st.session_state["v136_trace"] = tr
                        st.success(f"Repair done. Feasible={tr.get('final',{}).get('feasible')}")
                    except Exception as e:
                        st.error(f"Repair failed: {e!r}")
                tr = st.session_state.get("v136_trace")
                if isinstance(tr, dict):
                    st.json({"final": tr.get("final"), "final_inputs": tr.get("final_inputs")})
                    st.dataframe(tr.get("trace", [])[:200], use_container_width=True)
                    st.download_button("Download repair_trace_v136.json", data=_json.dumps(tr, indent=2, sort_keys=True), file_name="repair_trace_v136.json", mime="application/json", use_container_width=True)

    # ---------------- v137 compress ----------------
    with tabs[4]:
        rep = st.session_state.get("v133_rep")
        if not isinstance(rep, dict):
            st.info("Run FC first.")
        else:
            k = st.slider("Representatives K", 5, 200, 25, 5, key="v138_k")
            if st.button("Compress feasible set", key="v138_compress"):
                comp = compress_feasible_set(rep, k=int(k))
                st.session_state["v137_comp"] = comp
                st.success("Compressed.")
            comp = st.session_state.get("v137_comp")
            if isinstance(comp, dict):
                st.json({"k": comp.get("k"), "n_feasible": comp.get("n_feasible")})
                st.dataframe(comp.get("representatives", [])[:200], use_container_width=True)
                st.download_button("Download fc_compressed_v137.json", data=_json.dumps(comp, indent=2, sort_keys=True), file_name="fc_compressed_v137.json", mime="application/json", use_container_width=True)

    # ---------------- v138 handoff ----------------
    with tabs[5]:
        comp = st.session_state.get("v137_comp")
        rep = st.session_state.get("v133_rep")
        if not isinstance(rep, dict):
            st.info("Run FC first.")
        else:
            # pick best feasible from compressed or report
            best_inputs = None
            if isinstance(comp, dict) and (comp.get("representatives")):
                best_inputs = (comp["representatives"][0].get("inputs") if isinstance(comp["representatives"][0], dict) else None)
            if best_inputs is None:
                feas = [r for r in (rep.get("evaluations") or []) if isinstance(r, dict) and r.get("feasible") is True]
                if feas:
                    best_inputs = feas[0].get("inputs")
            if not isinstance(best_inputs, dict):
                st.info("No feasible point available to hand off. Try expanding bounds or budget.")
            else:
                st.write("Best feasible completion inputs (preview):")
                st.json({k: best_inputs.get(k) for k in list(best_inputs.keys())[:40]})
                if st.button("Evaluate completion as new run + pin", key="v138_handoff_eval"):
                    try:
                        art = completion_to_run_artifact(baseline_inputs=dict(base_inputs), completion_inputs=dict(best_inputs))
                        rid2 = _v98_record_run("handoff_run_artifact", art, mode="fc_handoff_v138")
                        s.pinned_run_ids.append(rid2)
                        st.success(f"Created new run artifact in ledger: {rid2} (pinned)")
                    except Exception as e:
                        st.error(f"Handoff failed: {e!r}")
                st.caption("After creating the pinned run, go to Study Matrix Builder v2 and select the pinned baseline.")


# =====================
# v139 Feasibility Certificate
# =====================
def _v139_feasibility_certificate_panel():
    import streamlit as st
    import json as _json
    from pathlib import Path
    from tools import run_vault
    from tools.feasibility_certificate import generate_feasibility_certificate

    st.subheader("Feasibility Certificate")
    st.caption("Generate an immutable, audit-ready certificate for a specific run artifact.")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("Run a point evaluation first to populate the run ledger.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids) > 0) else ids[-1])
    rid = st.selectbox("Select run", options=ids, index=ids.index(default_id) if default_id in ids else len(ids) - 1, key="v139_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    art = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if art is None:
        st.error("Selected entry does not contain a run artifact.")
        return

    origin = st.selectbox("Origin label", options=["point","fc_handoff","study_matrix",""], index=0, key="v139_origin")

    if st.button("Generate certificate", key="v139_gen"):
        root = Path(__file__).resolve().parents[1]
        cert = generate_feasibility_certificate(run_artifact=art, repo_root=root, run_id=str(rid), origin=str(origin))
        st.session_state["v139_cert"] = cert
        _v98_record_run("feasibility_certificate", {"certificate_id": cert.get("certificate_id"), "worst": cert.get("dominance",{}).get("worst_constraint"), "worst_margin_frac": cert.get("dominance",{}).get("worst_margin_frac")}, mode="v139")
        st.success("Certificate generated.")

    cert = st.session_state.get("v139_cert")
    if isinstance(cert, dict):
        st.json({
            "certificate_id": cert.get("certificate_id"),
            "issued_utc": cert.get("issued_utc"),
            "worst_constraint": (cert.get("dominance") or {}).get("worst_constraint"),
            "worst_margin_frac": (cert.get("dominance") or {}).get("worst_margin_frac"),
            "n_hard": len((cert.get("constraints") or {}).get("hard") or {}),
        })
        st.download_button(
            "Download feasibility_certificate_v139.json",
            data=_json.dumps(cert, indent=2, sort_keys=True),
            file_name="feasibility_certificate_v139.json",
            mime="application/json",
            use_container_width=True,
        )
        if st.button("Save certificate to Vault", key="v139_save_vault"):
            root = Path(__file__).resolve().parents[1]
            run_vault.write_entry(root=root, kind="feasibility_certificate", payload=cert, mode="v139", files={"feasibility_certificate_v139.json": _json.dumps(cert, indent=2, sort_keys=True).encode("utf-8")})
            st.success("Saved.")


# =====================
# v140 Sensitivity Maps (certificate -> fragility envelopes)
# =====================
def _v140_sensitivity_panel():
    import streamlit as st
    from pathlib import Path
    from tools.sensitivity_maps import SensitivityConfig, run_sensitivity, build_sensitivity_bundle
    from tools.feasibility_atlas import available_numeric_levers
    from tools import run_vault

    st.subheader("Constraint Sensitivity Maps")
    st.caption("Finite perturbation sensitivity: how much each variable can vary (+/-) before feasibility breaks. Auditable, no gradients/optimizers.")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("Run a point evaluation first.")
        return
    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Baseline run", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v140_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    base = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if base is None:
        st.error("Selected baseline does not contain a run artifact.")
        return
    base_inputs = base.get("inputs", {}) if isinstance(base.get("inputs"), dict) else {}

    levers = list(available_numeric_levers(base_inputs))
    default_vars = [v for v in ["Ip_MA","kappa","q95","a_m","beta_N","f_GW"] if v in levers]
    vars_sel = st.multiselect("Variables to probe", options=levers, default=(default_vars or levers[:4]), key="v140_vars")
    max_rel = st.slider("Max relative change (+/-)", 0.05, 1.0, 0.40, 0.05, key="v140_max_rel")
    n_expand = st.slider("Expansion steps", 3, 20, 8, 1, key="v140_nexp")
    n_bisect = st.slider("Bisection steps", 3, 20, 10, 1, key="v140_nbis")
    require_feas = st.checkbox("Require baseline feasible", value=True, key="v140_req_feas")

    if st.button("Run sensitivity", key="v140_run"):
        try:
            cfg = SensitivityConfig(
                baseline_inputs=dict(base_inputs),
                fixed_overrides={},
                vars=list(vars_sel),
                bounds={},  # intentionally empty by default; user can constrain later if needed
                max_rel=float(max_rel),
                max_abs=0.0,
                n_expand=int(n_expand),
                n_bisect=int(n_bisect),
                require_baseline_feasible=bool(require_feas),
            )
            rep = run_sensitivity(cfg)
            bun = build_sensitivity_bundle(rep)
            st.session_state["v140_rep"] = rep
            st.session_state["v140_bun"] = bun
            _v98_record_run("sensitivity_maps", {"n_vars": len(vars_sel), "max_rel": max_rel}, mode="v140")
            st.success("Sensitivity run complete.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    rep = st.session_state.get("v140_rep")
    bun = st.session_state.get("v140_bun")
    if isinstance(rep, dict):
        st.json({"baseline": rep.get("baseline"), "n_vars": len(rep.get("results") or [])})
        st.dataframe(rep.get("results", [])[:200], use_container_width=True)
        st.download_button("Download sensitivity_report_v140.json", data=_json.dumps(rep, indent=2, sort_keys=True, default=str),
                           file_name="sensitivity_report_v140.json", mime="application/json", use_container_width=True, key="v140_dl_rep")

    if isinstance(bun, dict) and isinstance(bun.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download sensitivity_bundle_v140.zip", data=bun["zip_bytes"], file_name="sensitivity_bundle_v140.zip",
                           mime="application/zip", use_container_width=True, key="v140_dl_zip")
        if st.button("Save sensitivity bundle to Vault", key="v140_save_vault"):
            try:
                root = Path(__file__).resolve().parents[1]
                run_vault.write_entry(root=root, kind="sensitivity_bundle", payload=rep or {}, mode="v140",
                                      files={"sensitivity_bundle_v140.zip": bun["zip_bytes"]})
                st.success("Saved to vault.")
            except Exception as e:
                st.error(f"Vault save failed: {e!r}")


# =====================
# v141 Robustness Certificate
# =====================
def _v141_robustness_panel():
    import streamlit as st
    from pathlib import Path
    from tools.feasibility_certificate import generate_feasibility_certificate
    from tools.sensitivity_maps import SensitivityConfig, run_sensitivity
    from tools.robustness_certificate import generate_robustness_certificate
    from tools.feasibility_atlas import available_numeric_levers
    from tools import run_vault

    st.subheader("Robustness Certificate")
    st.caption("Derives a robustness certificate from v139 Feasibility Certificate + v140 Sensitivity Report. No optimizers/gradients.")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("Run a point evaluation first.")
        return

    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Baseline run (for certificate + sensitivity)", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v141_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    base = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if base is None:
        st.error("Selected baseline does not contain a run artifact.")
        return
    base_inputs = base.get("inputs", {}) if isinstance(base.get("inputs"), dict) else {}

    # v140 settings (re-run sensitivity here for coherence)
    levers = list(available_numeric_levers(base_inputs))
    default_vars = [v for v in ["Ip_MA","kappa","q95","a_m","beta_N","f_GW"] if v in levers]
    vars_sel = st.multiselect("Variables for robustness probing", options=levers, default=(default_vars or levers[:4]), key="v141_vars")
    max_rel = st.slider("Max relative change (+/-) for sensitivity", 0.05, 1.0, 0.40, 0.05, key="v141_max_rel")
    n_expand = st.slider("Expansion steps", 3, 20, 8, 1, key="v141_nexp")
    n_bisect = st.slider("Bisection steps", 3, 20, 10, 1, key="v141_nbis")
    require_feas = st.checkbox("Require baseline feasible", value=True, key="v141_req_feas")

    policy_json = st.text_area("Policy JSON (optional)", value="{}", height=120, key="v141_policy")

    if st.button("Generate Robustness Certificate", key="v141_run"):
        try:
            # v139 cert from current artifact (source of truth)
            fc = generate_feasibility_certificate(base)

            # v140 report (computed here for coherence; also available via v140 panel)
            cfg = SensitivityConfig(
                baseline_inputs=dict(base_inputs),
                fixed_overrides={},
                vars=list(vars_sel),
                bounds={},
                max_rel=float(max_rel),
                max_abs=0.0,
                n_expand=int(n_expand),
                n_bisect=int(n_bisect),
                require_baseline_feasible=bool(require_feas),
            )
            sr = run_sensitivity(cfg)

            policy = _json.loads(policy_json) if policy_json.strip() else {}
            rc = generate_robustness_certificate(fc, sr, policy=policy)

            st.session_state["v141_fc"] = fc
            st.session_state["v141_sr"] = sr
            st.session_state["v141_rc"] = rc
            _v98_record_run("robustness_certificate", {"n_vars": len(vars_sel), "max_rel": max_rel, "index": rc.get("robustness", {}).get("index_min_bounded_rel")}, mode="v141")
            st.success("Robustness certificate generated.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    rc = st.session_state.get("v141_rc")
    fc = st.session_state.get("v141_fc")
    sr = st.session_state.get("v141_sr")

    if isinstance(rc, dict):
        st.json({
            "robustness_index_min_bounded_rel": (rc.get("robustness", {}) or {}).get("index_min_bounded_rel"),
            "fragility_top": ((rc.get("robustness", {}) or {}).get("fragility_ranking") or [])[:5],
        })
        st.download_button("Download robustness_certificate_v141.json",
                           data=_json.dumps(rc, indent=2, sort_keys=True, default=str),
                           file_name="robustness_certificate_v141.json", mime="application/json", use_container_width=True, key="v141_dl_rc")
        with st.expander("See underlying v139 feasibility certificate"):
            st.download_button("Download feasibility_certificate_v139.json",
                               data=_json.dumps(fc or {}, indent=2, sort_keys=True, default=str),
                               file_name="feasibility_certificate_v139.json", mime="application/json", use_container_width=True, key="v141_dl_fc")
        with st.expander("See underlying v140 sensitivity report"):
            st.download_button("Download sensitivity_report_v140.json",
                               data=_json.dumps(sr or {}, indent=2, sort_keys=True, default=str),
                               file_name="sensitivity_report_v140.json", mime="application/json", use_container_width=True, key="v141_dl_sr")

        if st.button("Save robustness certificate to Vault", key="v141_save_vault"):
            try:
                root = Path(__file__).resolve().parents[1]
                run_vault.write_entry(root=root, kind="robustness_certificate", payload=rc, mode="v141",
                                      files={"robustness_certificate_v141.json": _json.dumps(rc, indent=2, sort_keys=True, default=str).encode("utf-8")})
                st.success("Saved to vault.")
            except Exception as e:
                st.error(f"Vault save failed: {e!r}")


# =====================
# v142–v144 Feasibility Deep Dive
# =====================
def _v144_deepdive_panel():
    import streamlit as st
    from pathlib import Path
    from tools.feasibility_atlas import available_numeric_levers
    from tools.feasibility_deepdive import (
        SampleConfig, sample_and_evaluate, topology_from_dataset, bundle_topology,
        interactions_from_dataset, bundle_interactions,
        IntervalConfig, interval_certificate, bundle_interval_certificate,
    )
    from tools import run_vault

    st.subheader("Feasibility Deep Dive (v142–v144)")
    st.caption("Topology maps + constraint interaction structure + interval feasibility certificates. Downstream analysis only.")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("Run a point evaluation first.")
        return
    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Baseline run", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v144_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    base = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if base is None:
        st.error("Selected baseline does not contain a run artifact.")
        return
    base_inputs = base.get("inputs", {}) if isinstance(base.get("inputs"), dict) else {}

    levers = list(available_numeric_levers(base_inputs))
    tabs = st.tabs(["Topology Maps", "Constraint Interactions", "Interval Certificates"])

    # -------- v142
    with tabs[0]:
        st.write("Build feasible topology (islands) by sampling within bounds for selected variables.")
        vars_sel = st.multiselect("Variables (2–6 recommended)", options=levers, default=[v for v in ["Ip_MA","kappa"] if v in levers] or levers[:2], key="v142_vars")
        n_samples = st.slider("Samples", 100, 2000, 300, 50, key="v142_ns")
        seed = st.number_input("Seed", value=0, step=1, key="v142_seed")
        k = st.slider("kNN neighbors", 2, 20, 6, 1, key="v142_k")
        eps = st.number_input("Distance cutoff eps (0 = no cutoff)", value=0.0, key="v142_eps")

        bounds = {}
        for v in vars_sel:
            v0 = float(base_inputs.get(v, 0.0) or 0.0)
            lo = st.number_input(f"{v} min", value=(v0*0.9 if v0 else 0.0), key=f"v142_lo_{v}")
            hi = st.number_input(f"{v} max", value=(v0*1.1 if v0 else 1.0), key=f"v142_hi_{v}")
            bounds[v] = (float(lo), float(hi))

        if st.button("Run topology", key="v142_run"):
            try:
                cfg = SampleConfig(baseline_inputs=dict(base_inputs), vars=list(vars_sel), bounds=dict(bounds), n_samples=int(n_samples), seed=int(seed))
                ds = sample_and_evaluate(cfg)
                topo = topology_from_dataset(ds, k=int(k), eps=float(eps))
                bun = bundle_topology(ds, topo)
                st.session_state["v142_ds"] = ds
                st.session_state["v142_topo"] = topo
                st.session_state["v142_bun"] = bun
                _v98_record_run("feasible_topology", {"n_samples": n_samples, "n_feasible": topo.get("n_feasible_points"), "n_islands": len(topo.get("islands") or [])}, mode="v142")
                st.success(f"Topology built: feasible={topo.get('n_feasible_points')} islands={len(topo.get('islands') or [])}")
            except Exception as e:
                st.error(f"Topology failed: {e!r}")

        topo = st.session_state.get("v142_topo")
        bun = st.session_state.get("v142_bun")
        if isinstance(topo, dict):
            st.json({"n_feasible_points": topo.get("n_feasible_points"), "islands": topo.get("islands")[:10]})
        if isinstance(bun, dict) and isinstance(bun.get("zip_bytes"), (bytes, bytearray)):
            st.download_button("Download topology_bundle_v142.zip", data=bun["zip_bytes"], file_name="topology_bundle_v142.zip", mime="application/zip", use_container_width=True, key="v142_dl")
            if st.button("Save topology bundle to Vault", key="v142_save_vault"):
                root = Path(__file__).resolve().parents[1]
                run_vault.write_entry(root=root, kind="topology_bundle", payload=topo or {}, mode="v142", files={"topology_bundle_v142.zip": bun["zip_bytes"]})
                st.success("Saved.")

    # -------- v143
    with tabs[1]:
        st.write("Compute constraint co-failure and dominance statistics from the last deep-dive dataset.")
        ds = st.session_state.get("v142_ds")
        if not isinstance(ds, dict):
            st.info("Run topology sampling first (v142 tab) to generate a dataset.")
        else:
            if st.button("Compute interactions", key="v143_run"):
                try:
                    inter = interactions_from_dataset(ds, top_n=20)
                    bun2 = bundle_interactions(inter)
                    st.session_state["v143_inter"] = inter
                    st.session_state["v143_bun"] = bun2
                    _v98_record_run("constraint_interactions", {"top_constraints": len(inter.get("top_constraints") or [])}, mode="v143")
                    st.success("Interactions computed.")
                except Exception as e:
                    st.error(f"Interactions failed: {e!r}")
            inter = st.session_state.get("v143_inter")
            bun2 = st.session_state.get("v143_bun")
            if isinstance(inter, dict):
                st.json({"top_constraints": inter.get("top_constraints")[:12], "dominance_top": sorted((inter.get("dominance_counts") or {}).items(), key=lambda kv: kv[1], reverse=True)[:10]})
            if isinstance(bun2, dict) and isinstance(bun2.get("zip_bytes"), (bytes, bytearray)):
                st.download_button("Download interactions_bundle_v143.zip", data=bun2["zip_bytes"], file_name="interactions_bundle_v143.zip", mime="application/zip", use_container_width=True, key="v143_dl")
                if st.button("Save interactions bundle to Vault", key="v143_save_vault"):
                    root = Path(__file__).resolve().parents[1]
                    run_vault.write_entry(root=root, kind="interactions_bundle", payload=inter or {}, mode="v143", files={"interactions_bundle_v143.zip": bun2["zip_bytes"]})
                    st.success("Saved.")

    # -------- v144
    with tabs[2]:
        st.write("Certify a hyper-rectangle interval (conservative): checks all corners + random interior probes.")
        vars_sel = st.multiselect("Interval variables", options=levers, default=[v for v in ["Ip_MA","kappa"] if v in levers] or levers[:2], key="v144_ivars")
        bounds = {}
        for v in vars_sel:
            v0 = float(base_inputs.get(v, 0.0) or 0.0)
            lo = st.number_input(f"{v} min", value=(v0*0.95 if v0 else 0.0), key=f"v144_lo_{v}")
            hi = st.number_input(f"{v} max", value=(v0*1.05 if v0 else 1.0), key=f"v144_hi_{v}")
            bounds[v] = (float(lo), float(hi))
        n_random = st.slider("Random interior probes", 0, 500, 60, 10, key="v144_nr")
        seed = st.number_input("Seed", value=0, step=1, key="v144_seed")
        if st.button("Generate interval certificate (v144)", key="v144_run"):
            try:
                cert = interval_certificate(IntervalConfig(baseline_inputs=dict(base_inputs), bounds=dict(bounds), n_random=int(n_random), seed=int(seed)))
                bun = bundle_interval_certificate(cert)
                st.session_state["v144_cert"] = cert
                st.session_state["v144_bun"] = bun
                _v98_record_run("interval_certificate", {"interval_certified": cert.get("verdict",{}).get("interval_certified"), "n_vars": len(bounds)}, mode="v144")
                st.success(f"Interval certified = {cert.get('verdict',{}).get('interval_certified')}")
            except Exception as e:
                st.error(f"Interval certificate failed: {e!r}")

        cert = st.session_state.get("v144_cert")
        bun = st.session_state.get("v144_bun")
        if isinstance(cert, dict):
            st.json(cert.get("verdict", {}))
            st.download_button("Download interval_certificate_v144.json", data=_json.dumps(cert, indent=2, sort_keys=True, default=str),
                               file_name="interval_certificate_v144.json", mime="application/json", use_container_width=True, key="v144_dl_json")
        if isinstance(bun, dict) and isinstance(bun.get("zip_bytes"), (bytes, bytearray)):
            st.download_button("Download interval_bundle_v144.zip", data=bun["zip_bytes"], file_name="interval_bundle_v144.zip", mime="application/zip", use_container_width=True, key="v144_dl_zip")
            if st.button("Save interval bundle to Vault", key="v144_save_vault"):
                root = Path(__file__).resolve().parents[1]
                run_vault.write_entry(root=root, kind="interval_bundle", payload=cert or {}, mode="v144", files={"interval_bundle_v144.zip": bun["zip_bytes"]})
                st.success("Saved.")


# =====================
# v145 Topology Certificate
# =====================
def _v145_topology_certificate_panel():
    import streamlit as st
    from pathlib import Path
    from tools.topology_certificate import generate_topology_certificate
    from tools import run_vault

    st.subheader("Topology Certificate")
    st.caption("Citable certificate summarizing feasible-set topology (islands) for a declared domain and sampling protocol.")

    s = _v98_state_init_runlists()
    ids = [r.get("id") for r in (s.run_history or []) if r.get("id")]
    if not ids:
        st.info("Run a point evaluation first.")
        return
    default_id = (s.pinned_run_ids[-1] if (s.pinned_run_ids and len(s.pinned_run_ids)>0) else ids[-1])
    rid = st.selectbox("Baseline run", options=ids, index=ids.index(default_id) if default_id in ids else len(ids)-1, key="v145_pick_run")
    run_map = {r.get("id"): r for r in (s.run_history or []) if r.get("id")}
    payload = (run_map.get(rid) or {}).get("payload")
    base = payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None
    if base is None:
        st.error("Selected baseline does not contain a run artifact.")
        return

    topo = st.session_state.get("v142_topo")
    ds = st.session_state.get("v142_ds")
    if not isinstance(topo, dict):
        st.info("Run v142 Topology Maps first (Feasibility Deep Dive panel) to generate feasible topology.")
        return

    policy_json = st.text_area("Policy JSON (optional)", value="{}", height=120, key="v145_policy")

    if st.button("Generate Topology Certificate", key="v145_run"):
        try:
            policy = _json.loads(policy_json) if policy_json.strip() else {}
            cert = generate_topology_certificate(base, topo, deepdive_dataset=(ds if isinstance(ds, dict) else None), policy=policy)
            st.session_state["v145_cert"] = cert
            _v98_record_run("topology_certificate", {"n_islands": (cert.get("topology_summary", {}) or {}).get("n_islands")}, mode="v145")
            st.success("Topology certificate generated.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    cert = st.session_state.get("v145_cert")
    if isinstance(cert, dict):
        st.json(cert.get("topology_summary", {}))
        st.download_button("Download topology_certificate_v145.json",
                           data=_json.dumps(cert, indent=2, sort_keys=True, default=str),
                           file_name="topology_certificate_v145.json", mime="application/json",
                           use_container_width=True, key="v145_dl")
        if st.button("Save topology certificate to Vault", key="v145_save_vault"):
            try:
                root = Path(__file__).resolve().parents[1]
                run_vault.write_entry(root=root, kind="topology_certificate", payload=cert, mode="v145",
                                      files={"topology_certificate_v145.json": _json.dumps(cert, indent=2, sort_keys=True, default=str).encode("utf-8")})
                st.success("Saved to vault.")
            except Exception as e:
                st.error(f"Vault save failed: {e!r}")


# =====================
# v146–v147 Feasibility Completion
# =====================
def _v147_feasibility_completion_panel():
    import streamlit as st
    from pathlib import Path
    from tools.feasibility_atlas import available_numeric_levers
    from tools.feasibility_bridge import BridgeConfig, run_bridge, bridge_certificate
    from tools.safe_domain_shrink import ShrinkConfig, run_safe_domain_shrink
    from tools import run_vault
    import hashlib

    st.subheader("Feasibility Completion (v146–v147)")
    st.caption("v146: topology bridge witness between two points. v147: auto-shrink to a certified safe interval box (uses v144).")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    ids = [r.get("id") for r in runs]
    if len(ids) < 1:
        st.info("Run at least one point evaluation first.")
        return

    run_map = {r.get("id"): r for r in runs}
    def _get_art(rid):
        payload = (run_map.get(rid) or {}).get("payload")
        return payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None

    tabs = st.tabs(["Bridge Certificate", "Safe Domain Shrink"])

    # ---------- v146
    with tabs[0]:
        if len(ids) < 2:
            st.info("Need at least two runs in history to bridge (A and B).")
        else:
            ridA = st.selectbox("Point A (baseline)", options=ids, index=max(0, len(ids)-2), key="v146_A")
            ridB = st.selectbox("Point B (target)", options=ids, index=len(ids)-1, key="v146_B")
            artA = _get_art(ridA); artB = _get_art(ridB)
            if artA is None or artB is None:
                st.error("Selected runs must be run artifacts.")
            else:
                Ain = artA.get("inputs", {}) if isinstance(artA.get("inputs"), dict) else {}
                Bin = artB.get("inputs", {}) if isinstance(artB.get("inputs"), dict) else {}
                levers = sorted(set(available_numeric_levers(Ain)).intersection(set(available_numeric_levers(Bin))))
                default_vars = [v for v in ["Ip_MA","kappa"] if v in levers] or levers[:2]
                vars_sel = st.multiselect("Bridge variables", options=levers, default=default_vars, key="v146_vars")
                n_steps = st.slider("Coarse steps", 5, 101, 21, 2, key="v146_steps")
                bis = st.slider("Max bisection depth", 0, 10, 6, 1, key="v146_bis")
                req_end = st.checkbox("Require endpoints feasible", value=True, key="v146_req")
                if st.button("Run bridge", key="v146_run"):
                    try:
                        cfg = BridgeConfig(inputs_A=dict(Ain), inputs_B=dict(Bin), vars=list(vars_sel), n_steps=int(n_steps), max_bisect_depth=int(bis), require_endpoints_feasible=bool(req_end))
                        rep = run_bridge(cfg)
                        # baseline hash from A inputs
                        h = hashlib.sha256(_json.dumps(Ain, sort_keys=True, default=str).encode("utf-8")).hexdigest()
                        cert = bridge_certificate(rep, baseline_inputs_sha256=h)
                        st.session_state["v146_rep"] = rep
                        st.session_state["v146_cert"] = cert
                        _v98_record_run("bridge_certificate", {"bridge_exists": rep.get("bridge_exists"), "n_points": len(rep.get("path") or [])}, mode="v146")
                        st.success(f"Bridge exists = {rep.get('bridge_exists')}")
                    except Exception as e:
                        st.error(f"Bridge failed: {e!r}")

                cert = st.session_state.get("v146_cert")
                rep = st.session_state.get("v146_rep")
                if isinstance(cert, dict):
                    st.json(cert.get("summary", {}))
                    st.download_button("Download bridge_certificate_v146.json", data=_json.dumps(cert, indent=2, sort_keys=True, default=str),
                                       file_name="bridge_certificate_v146.json", mime="application/json", use_container_width=True, key="v146_dl_cert")
                    st.download_button("Download bridge_report_v146.json", data=_json.dumps(rep or {}, indent=2, sort_keys=True, default=str),
                                       file_name="bridge_report_v146.json", mime="application/json", use_container_width=True, key="v146_dl_rep")
                    if st.button("Save v146 to Vault", key="v146_save_vault"):
                        root = Path(__file__).resolve().parents[1]
                        run_vault.write_entry(root=root, kind="bridge_certificate", payload={"certificate": cert, "report": rep}, mode="v146",
                                              files={
                                                  "bridge_certificate_v146.json": _json.dumps(cert, indent=2, sort_keys=True, default=str).encode("utf-8"),
                                                  "bridge_report_v146.json": _json.dumps(rep or {}, indent=2, sort_keys=True, default=str).encode("utf-8"),
                                              })
                        st.success("Saved.")

    # ---------- v147
    with tabs[1]:
        rid = st.selectbox("Baseline run for safe box", options=ids, index=len(ids)-1, key="v147_base")
        art = _get_art(rid)
        if art is None:
            st.error("Selected run must be a run artifact.")
        else:
            base_inputs = art.get("inputs", {}) if isinstance(art.get("inputs"), dict) else {}
            levers = list(available_numeric_levers(base_inputs))
            vars_sel = st.multiselect("Interval variables", options=levers, default=[v for v in ["Ip_MA","kappa"] if v in levers] or levers[:2], key="v147_vars")
            bounds={}
            for v in vars_sel:
                v0=float(base_inputs.get(v, 0.0) or 0.0)
                lo = st.number_input(f"{v} min", value=(v0*0.9 if v0 else 0.0), key=f"v147_lo_{v}")
                hi = st.number_input(f"{v} max", value=(v0*1.1 if v0 else 1.0), key=f"v147_hi_{v}")
                bounds[v]=(float(lo), float(hi))
            shrink_factor = st.slider("Shrink factor per iteration", 0.50, 0.98, 0.85, 0.01, key="v147_sf")
            max_iter = st.slider("Max iterations", 1, 30, 10, 1, key="v147_mi")
            n_random = st.slider("Random probes per iteration", 0, 300, 40, 10, key="v147_nr")
            seed = st.number_input("Seed", value=0, step=1, key="v147_seed")

            if st.button("Run safe-domain shrink", key="v147_run"):
                try:
                    cfg = ShrinkConfig(baseline_inputs=dict(base_inputs), bounds=dict(bounds), shrink_factor=float(shrink_factor),
                                       max_iter=int(max_iter), n_random=int(n_random), seed=int(seed))
                    rep = run_safe_domain_shrink(cfg)
                    st.session_state["v147_rep"] = rep
                    _v98_record_run("safe_domain_shrink", {"final_certified": rep.get("final_certified"), "iters": len(rep.get("history") or [])}, mode="v147")
                    st.success(f"Final certified = {rep.get('final_certified')} (iters={len(rep.get('history') or [])})")
                except Exception as e:
                    st.error(f"Shrink failed: {e!r}")

            rep = st.session_state.get("v147_rep")
            if isinstance(rep, dict):
                st.json({"final_certified": rep.get("final_certified"), "final_bounds": rep.get("final_bounds"), "last": (rep.get("history") or [])[-1] if (rep.get("history") or []) else None})
                st.download_button("Download safe_domain_shrink_report_v147.json",
                                   data=_json.dumps(rep, indent=2, sort_keys=True, default=str),
                                   file_name="safe_domain_shrink_report_v147.json", mime="application/json",
                                   use_container_width=True, key="v147_dl_rep")
                cert = rep.get("interval_certificate_v144")
                if isinstance(cert, dict):
                    st.download_button("Download interval_certificate_v144.json",
                                       data=_json.dumps(cert, indent=2, sort_keys=True, default=str),
                                       file_name="interval_certificate_v144.json", mime="application/json",
                                       use_container_width=True, key="v147_dl_cert")
                if st.button("Save v147 to Vault", key="v147_save_vault"):
                    root = Path(__file__).resolve().parents[1]
                    files={"safe_domain_shrink_report_v147.json": _json.dumps(rep, indent=2, sort_keys=True, default=str).encode("utf-8")}
                    if isinstance(cert, dict):
                        files["interval_certificate_v144.json"]=_json.dumps(cert, indent=2, sort_keys=True, default=str).encode("utf-8")
                    run_vault.write_entry(root=root, kind="safe_domain_shrink", payload=rep, mode="v147", files=files)
                    st.success("Saved.")


# =====================
# v148–v150 Publishable Study Kit
# =====================
def _v150_publishable_study_kit_panel():
    import streamlit as st
    from pathlib import Path
    from tools.design_study_kit import PaperPackConfig, build_paper_pack
    from tools import run_vault

    st.subheader("Publishable Study Kit (v148–v150)")
    st.caption("One-click paper pack (zip) with study registry + integrity manifest. Downstream-only.")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    if not runs:
        st.info("Run at least one evaluation first.")
        return
    ids=[r.get("id") for r in runs]
    run_map={r.get("id"): r for r in runs}

    def _get_art(rid):
        payload = (run_map.get(rid) or {}).get("payload")
        return payload if (isinstance(payload, dict) and payload.get("kind") == "shams_run_artifact") else None

    sel = st.multiselect("Select run(s) to include", options=ids, default=[ids[-1]], key="v150_sel_runs")
    title = st.text_input("Study title", value="SHAMS design study", key="v150_title")
    authors = st.text_input("Authors (comma-separated)", value="", key="v150_authors")
    desc = st.text_area("Description", value="", height=120, key="v150_desc")

    # auto-pick latest certificates present in session_state (best effort)
    certs=[]
    for key, fn in [
        ("v139_fc", "feasibility_certificate_v139.json"),
        ("v141_rc", "robustness_certificate_v141.json"),
        ("v144_ic", "interval_certificate_v144.json"),
        ("v145_tc", "topology_certificate_v145.json"),
        ("v146_bc", "bridge_certificate_v146.json"),
        ("v147_sd", "safe_domain_shrink_report_v147.json"),
    ]:
        obj = st.session_state.get(key)
        if isinstance(obj, dict):
            certs.append((fn, obj))

    st.checkbox("Include best-effort certificates from current session", value=True, key="v150_include_session_certs")
    st.checkbox("Include figures/tables from session bundles", value=True, key="v151_include_session_bundles")
    methods_json = st.text_area("Methods JSON (optional)", value="{}", height=120, key="v150_methods")
    policy_json = st.text_area("Policy JSON (optional)", value='{"mode":"paper_pack"}', height=120, key="v150_policy")

    if st.button("Build Paper Pack", key="v150_run"):
        try:
            run_arts=[]
            for rid in sel:
                art=_get_art(rid)
                if art is not None:
                    run_arts.append(art)
            if not run_arts:
                st.error("No valid run artifacts selected.")
                return

            methods=_json.loads(methods_json) if methods_json.strip() else {}
            policy=_json.loads(policy_json) if policy_json.strip() else {}
            # v154 captions override
            caps = st.session_state.get("v154_captions")
            if isinstance(caps, dict) and caps:
                policy = dict(policy)
                policy["captions_override"] = caps

            cfg=PaperPackConfig(
                shams_version=str((_json.loads((Path(__file__).resolve().parents[1]/"VERSION").read_text(encoding="utf-8").strip().splitlines()[0]) if True else "unknown")),
                title=title,
                authors=[a.strip() for a in authors.split(",") if a.strip()],
                description=desc,
                run_artifacts=run_arts,
                certificates=certs if st.session_state.get("v150_include_session_certs") else [],
                figures=figs,
                tables=tabs,
                methods=methods,
                policy=policy,
            )
            bun=build_paper_pack(cfg)
            st.session_state["v150_pack"]=bun
            _v98_record_run("paper_pack", {"n_runs": len(run_arts), "n_certs": len(cfg.certificates)}, mode="v150")
            st.success("Paper pack built.")
        except Exception as e:
            st.error(f"Failed: {e!r}")

    bun=st.session_state.get("v150_pack")
    if isinstance(bun, dict) and isinstance(bun.get("zip_bytes"), (bytes, bytearray)):
        st.download_button("Download paper_pack_v150.zip", data=bun["zip_bytes"], file_name="paper_pack_v150.zip",
                           mime="application/zip", use_container_width=True, key="v150_dl_zip")
        st.download_button("Download study_registry_v149.json", data=_json.dumps(bun.get("study_registry") or {}, indent=2, sort_keys=True, default=str),
                           file_name="study_registry_v149.json", mime="application/json", use_container_width=True, key="v150_dl_reg")
        st.download_button("Download captions.json", data=_json.dumps({"note":"captions.json is included inside paper_pack_v150.zip"}, indent=2), file_name="captions_note.json", mime="application/json", use_container_width=True, key="v151_dl_caps_note")
        
        st.download_button("Download integrity_manifest_v150.json", data=_json.dumps((bun.get("manifest") or {}), indent=2, sort_keys=True, default=str),
                           file_name="integrity_manifest_v150.json", mime="application/json", use_container_width=True, key="v150_dl_mf")

        if st.button("Save paper pack to Vault", key="v150_save_vault"):
            root = Path(__file__).resolve().parents[1]
            run_vault.write_entry(root=root, kind="paper_pack", payload={"study_registry": bun.get("study_registry"), "manifest": bun.get("manifest")}, mode="v150",
                                  files={
                                      "paper_pack_v150.zip": bun["zip_bytes"],
                                      "study_registry_v149.json": _json.dumps(bun.get("study_registry") or {}, indent=2, sort_keys=True, default=str).encode("utf-8"),
                                      "integrity_manifest_v150.json": _json.dumps(bun.get("manifest") or {}, indent=2, sort_keys=True, default=str).encode("utf-8"),
                                  })
            st.success("Saved.")


# =====================
# v152 Integrity Lock helpers
# =====================
def _v152_artifact_sha(payload):
    import json, hashlib
    b = json.dumps(payload, indent=2, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(b).hexdigest()

def _v152_get_lock_for_run(run_id: str):
    import streamlit as st
    locks = st.session_state.get("v152_locks") or {}
    return locks.get(run_id) if isinstance(locks, dict) else None

def _v152_set_lock_for_run(run_id: str, lock_obj: dict):
    import streamlit as st
    locks = st.session_state.get("v152_locks")
    if not isinstance(locks, dict):
        locks = {}
    locks[str(run_id)] = lock_obj
    st.session_state["v152_locks"] = locks

def _v152_integrity_status(run_id: str, payload: dict):
    lock = _v152_get_lock_for_run(run_id)
    if not isinstance(lock, dict):
        return ("UNLOCKED", None)
    # compare current artifact sha with stored
    expected = ((lock.get("files") or {}).get("run_artifact.json") or {}).get("sha256")
    cur = _v152_artifact_sha(payload)
    if expected and cur == expected:
        return ("VERIFIED", cur)
    return ("MODIFIED", cur)

def _v152_integrity_panel():
    import streamlit as st
    from pathlib import Path
    from tools.run_integrity_lock import lock_run, verify_run
    from tools import run_vault

    st.subheader("Run Integrity Lock")
    st.caption("Lock a run artifact hash and later verify whether it changed. Downstream-only.")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    if not runs:
        st.info("Run at least one evaluation first.")
        return
    ids=[r.get("id") for r in runs]
    run_map={r.get("id"): r for r in runs}
    rid = st.selectbox("Select run to lock/verify", options=ids, index=len(ids)-1, key="v152_pick_run")
    payload = (run_map.get(rid) or {}).get("payload")
    if not (isinstance(payload, dict) and payload.get("kind")=="shams_run_artifact"):
        st.error("Selected run payload missing.")
        return

    status, cur_sha = _v152_integrity_status(rid, payload)
    st.write(f"Status: **{status}**")
    if cur_sha:
        st.code(cur_sha)

    if st.button("Lock integrity for this run", key="v152_lock"):
        out = lock_run(run_id=str(rid), run_artifact=payload, extras=None, policy={"mode":"ui"})
        lock_obj = out["lock"]
        _v152_set_lock_for_run(rid, lock_obj)
        st.session_state["v152_last_lock"] = lock_obj
        st.success("Locked.")
        _v98_record_run("integrity_lock", {"status":"locked"}, mode="v152")

    lock_obj = _v152_get_lock_for_run(rid) or st.session_state.get("v152_last_lock")
    if isinstance(lock_obj, dict):
        rep = verify_run(lock_obj, payload, extras=None)
        st.json(rep)
        st.download_button("Download run_integrity_lock_v152.json", data=_json.dumps(lock_obj, indent=2, sort_keys=True, default=str),
                           file_name="run_integrity_lock_v152.json", mime="application/json", use_container_width=True, key="v152_dl_lock")
        st.download_button("Download run_integrity_verify_v152.json", data=_json.dumps(rep, indent=2, sort_keys=True, default=str),
                           file_name="run_integrity_verify_v152.json", mime="application/json", use_container_width=True, key="v152_dl_rep")

        if st.button("Save integrity lock to Vault", key="v152_save_vault"):
            root = Path(__file__).resolve().parents[1]
            run_vault.write_entry(root=root, kind="run_integrity_lock", payload={"lock": lock_obj, "verify": rep}, mode="v152",
                                  files={
                                      "run_integrity_lock_v152.json": _json.dumps(lock_obj, indent=2, sort_keys=True, default=str).encode("utf-8"),
                                      "run_integrity_verify_v152.json": _json.dumps(rep, indent=2, sort_keys=True, default=str).encode("utf-8"),
                                  })
            st.success("Saved.")


# =====================
# v153–v155 Study Kit Extensions
# =====================
def _v153_doi_export_panel():
    import streamlit as st
    from tools.doi_export import zenodo_metadata_from_registry, crossref_minimal_from_registry

    st.subheader("DOI Export Helper")
    st.caption("Export Zenodo/Crossref-style metadata from the latest study registry built in this session.")

    bun = st.session_state.get("v150_pack")
    reg = (bun or {}).get("study_registry") if isinstance(bun, dict) else None
    if not isinstance(reg, dict):
        st.info("Build a Paper Pack first (Publishable Study Kit panel).")
        return

    comm = st.text_input("Zenodo communities (comma-separated identifiers)", value="", key="v153_comm")
    kws = st.text_input("Keywords (comma-separated)", value="", key="v153_kws")
    doi = st.text_input("DOI (optional)", value="", key="v153_doi")
    publisher = st.text_input("Publisher (optional)", value="SHAMS", key="v153_pub")
    url = st.text_input("Resource URL (optional)", value="", key="v153_url")

    communities=[c.strip() for c in comm.split(",") if c.strip()] if comm else []
    keywords=[k.strip() for k in kws.split(",") if k.strip()] if kws else []

    zen = zenodo_metadata_from_registry(reg, communities=communities, keywords=keywords)
    cr  = crossref_minimal_from_registry(reg, doi=doi, publisher=publisher, resource_url=url)

    st.download_button("Download zenodo_metadata_v153.json", data=_json.dumps(zen, indent=2, sort_keys=True, default=str),
                       file_name="zenodo_metadata_v153.json", mime="application/json", use_container_width=True, key="v153_dl_zen")
    st.download_button("Download crossref_minimal_v153.json", data=_json.dumps(cr, indent=2, sort_keys=True, default=str),
                       file_name="crossref_minimal_v153.json", mime="application/json", use_container_width=True, key="v153_dl_cr")

def _v154_caption_editor_panel():
    import streamlit as st
    st.subheader("Caption Editor")
    st.caption("Edit captions for figures/tables included in the paper pack. Captions are stored in session and exported into paper packs.")

    # Captions live in session state and are applied when building the paper pack.
    caps = st.session_state.get("v154_captions")
    if not isinstance(caps, dict):
        caps = {}
        st.session_state["v154_captions"] = caps

    # Show detected figure/table filenames from latest built pack, if any
    bun = st.session_state.get("v150_pack")
    reg = (bun or {}).get("study_registry") if isinstance(bun, dict) else None
    names=[]
    if isinstance(reg, dict):
        for ref in (reg.get("figures") or []):
            if isinstance(ref, dict) and ref.get("name"):
                names.append(ref["name"])
        for ref in (reg.get("tables") or []):
            if isinstance(ref, dict) and ref.get("name"):
                names.append(ref["name"])
    names = sorted(set(names))

    if not names:
        st.info("No figures/tables detected yet. Build a Paper Pack with v151 session-bundle harvesting enabled.")
    else:
        pick = st.selectbox("Select figure/table", options=names, key="v154_pick")
        cur = caps.get(pick, f"Figure/Table: {pick}.")
        new = st.text_area("Caption text", value=cur, height=120, key="v154_text")
        if st.button("Save caption", key="v154_save"):
            caps[pick] = new
            st.session_state["v154_captions"] = caps
            st.success("Saved.")
        st.download_button("Download captions_override_v154.json", data=_json.dumps({"captions": caps}, indent=2, sort_keys=True, default=str),
                           file_name="captions_override_v154.json", mime="application/json", use_container_width=True, key="v154_dl")

def _v155_multi_study_pack_panel():
    import streamlit as st
    from tools.multi_study_pack import build_multi_study_pack

    st.subheader("Multi-Study Comparison Pack")
    st.caption("Upload multiple paper packs and export a comparison bundle with a multi-pack integrity manifest.")

    files = st.file_uploader("Upload paper_pack_v150.zip files", type=["zip"], accept_multiple_files=True, key="v155_upload")
    if not files:
        st.info("Upload 2+ paper packs to compare.")
        return
    packs=[]
    for f in files:
        packs.append((f.name, f.getvalue()))

    bun = build_multi_study_pack(packs, policy={"source":"ui"})
    st.json({"n_packs": len(packs), "manifest_sha256": (bun.get("manifest") or {}).get("hashes",{}).get("manifest_sha256")})
    st.download_button("Download multi_study_pack_v155.zip", data=bun["zip_bytes"], file_name="multi_study_pack_v155.zip",
                       mime="application/zip", use_container_width=True, key="v155_dl_zip")
    st.download_button("Download comparison_report_v155.json", data=_json.dumps(bun.get("bundle") or {}, indent=2, sort_keys=True, default=str),
                       file_name="comparison_report_v155.json", mime="application/json", use_container_width=True, key="v155_dl_rep")


# =====================
# v156 / v160 Design Space Authority (Atlas + Certificates)
# =====================
def _v156_feasibility_atlas_panel():
    import streamlit as st
    from tools.feasibility_field import build_feasibility_field

    st.subheader("Feasibility Atlas")
    st.caption("Sample a 2D design space slice and export a feasibility field + atlas bundle.")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    if not runs:
        st.info("Run at least one evaluation first (Point Designer / Systems Mode).")
        return
    ids=[r.get("id") for r in runs]
    run_map={r.get("id"): r for r in runs}
    rid = st.selectbox("Baseline run", options=ids, index=len(ids)-1, key="v156_base_run")
    payload = (run_map.get(rid) or {}).get("payload")
    if not (isinstance(payload, dict) and payload.get("kind")=="shams_run_artifact"):
        st.error("Selected baseline run payload missing.")
        return
    baseline_inputs = payload.get("inputs") or payload.get("_inputs") or {}
    if not isinstance(baseline_inputs, dict) or not baseline_inputs:
        st.error("Baseline inputs missing.")
        return

    keys=sorted([k for k in baseline_inputs.keys() if isinstance(k,str)])
    col1,col2=st.columns(2)
    with col1:
        a1 = st.selectbox("Axis 1 parameter", options=keys, index=0 if keys else 0, key="v156_a1")
        a1_start = st.number_input("Axis 1 start", value=float(baseline_inputs.get(a1, 0.0) or 0.0), key="v156_a1s")
        a1_stop  = st.number_input("Axis 1 stop", value=float(baseline_inputs.get(a1, 1.0) or 1.0)+1.0, key="v156_a1e")
        a1_n     = st.number_input("Axis 1 N", value=25, min_value=2, max_value=250, step=1, key="v156_a1n")
    with col2:
        a2 = st.selectbox("Axis 2 parameter", options=keys, index=1 if len(keys)>1 else 0, key="v156_a2")
        a2_start = st.number_input("Axis 2 start", value=float(baseline_inputs.get(a2, 0.0) or 0.0), key="v156_a2s")
        a2_stop  = st.number_input("Axis 2 stop", value=float(baseline_inputs.get(a2, 1.0) or 1.0)+1.0, key="v156_a2e")
        a2_n     = st.number_input("Axis 2 N", value=25, min_value=2, max_value=250, step=1, key="v156_a2n")

    fixed_json = st.text_area("Fixed overrides (JSON list of {name,value})", value="[]", height=90, key="v156_fixed")
    assumptions_json = st.text_area("Assumption set (JSON object)", value="{}", height=90, key="v156_assumptions")
    margin_eps = st.number_input("Margin epsilon for feasible", value=1e-6, format="%.1e", key="v156_eps")
    run_btn = st.button("Build Feasibility Field", use_container_width=True, key="v156_run")

    if run_btn:
        try:
            fixed = _json.loads(fixed_json) if fixed_json.strip() else []
            assumptions = _json.loads(assumptions_json) if assumptions_json.strip() else {}
        except Exception as e:
            st.error(f"JSON parse error: {e}")
            return

        axis1={"name": a1, "role":"axis", "grid":{"type":"linspace","start": float(a1_start), "stop": float(a1_stop), "n": int(a1_n)}}
        axis2={"name": a2, "role":"axis", "grid":{"type":"linspace","start": float(a2_start), "stop": float(a2_stop), "n": int(a2_n)}}
        with st.spinner("Sampling feasibility field..."):
            out = build_feasibility_field(
                baseline_inputs=baseline_inputs,
                axis1=axis1,
                axis2=axis2,
                fixed=fixed if isinstance(fixed, list) else [],
                assumption_set=assumptions if isinstance(assumptions, dict) else {},
                sampling={"generator":"ui","strategy":"grid"},
                solver_meta={"label":"feasibility_field_v156"},
                margin_eps=float(margin_eps),
            )
        st.session_state["v156_field"] = out["field"]
        st.session_state["v156_atlas_zip"] = out["zip_bytes"]
        summ = (((out["field"].get("payload") or {}).get("field") or {}).get("summaries") or {})
        st.success("Built.")
        st.json(summ)
        st.download_button("Download feasibility_atlas_bundle_v156.zip", data=out["zip_bytes"], file_name="feasibility_atlas_bundle_v156.zip",
                           mime="application/zip", use_container_width=True, key="v156_dl_zip")
        st.download_button("Download feasibility_field_v156.json", data=_json.dumps(out["field"], indent=2, sort_keys=True, default=str),
                           file_name="feasibility_field_v156.json", mime="application/json", use_container_width=True, key="v156_dl_json")

def _v160_authority_certificate_panel():
    import streamlit as st
    from tools.feasibility_authority_certificate import issue_certificate_from_field

    st.subheader("Feasibility Authority Certificate")
    st.caption("Issue an authority certificate from a feasibility field (dense sampling basis).")

    field = st.session_state.get("v156_field")
    up = st.file_uploader("Optional: upload feasibility_field_v156.json", type=["json"], key="v160_upload")
    if up is not None:
        try:
            field = _json.loads(up.getvalue().decode("utf-8"))
        except Exception as e:
            st.error(f"Invalid JSON: {e}")
            return

    if not isinstance(field, dict):
        st.info("Build a Feasibility Field first or upload one here.")
        return

    claim = st.selectbox("Claim type", options=["feasible_region","excluded_region","boundary_surface","completion_existence"], index=1, key="v160_claim")
    default_stmt = "Under assumption_set SHA256=..., region sampled is excluded/feasible under dense sampling evidence."
    stmt = st.text_area("Claim statement (human-readable, publishable)", value=default_stmt, height=110, key="v160_stmt")
    conf = st.number_input("Confidence level", value=0.95, min_value=0.5, max_value=0.999, step=0.01, key="v160_conf")
    grade = st.selectbox("Grade", options=["A","B","C"], index=1, key="v160_grade")

    if st.button("Issue Certificate", use_container_width=True, key="v160_issue"):
        cert = issue_certificate_from_field(field=field, claim_type=claim, statement=stmt, confidence_level=float(conf), confidence_grade=str(grade), policy={"mode":"ui"})
        st.session_state["v160_cert"] = cert
        st.success("Issued.")
        st.json(cert.get("payload") or {})
        st.download_button("Download feasibility_authority_certificate_v160.json",
                           data=_json.dumps(cert, indent=2, sort_keys=True, default=str),
                           file_name="feasibility_authority_certificate_v160.json",
                           mime="application/json", use_container_width=True, key="v160_dl")


# =====================
# v157 Feasibility Boundary Extractor
# =====================
def _v157_feasibility_boundary_panel():
    import streamlit as st
    from tools.feasibility_boundary import build_feasibility_boundary

    st.subheader("Feasibility Boundary")
    st.caption("Extract a feasibility boundary curve from a v156 feasibility field (grid interpolation).")

    field = st.session_state.get("v156_field")
    up = st.file_uploader("Optional: upload feasibility_field_v156.json", type=["json"], key="v157_upload")
    if up is not None:
        try:
            field = _json.loads(up.getvalue().decode("utf-8"))
        except Exception as e:
            st.error(f"Invalid JSON: {e}")
            return
    if not isinstance(field, dict):
        st.info("Build a Feasibility Field first or upload one here.")
        return

    prefer_lowest = st.checkbox("Prefer lowest Axis2 crossing (typical min required)", value=True, key="v157_low")
    if st.button("Extract Boundary", use_container_width=True, key="v157_run"):
        b = build_feasibility_boundary(field=field, prefer_lowest_axis2=bool(prefer_lowest))
        st.session_state["v157_boundary"] = b
        st.success(f"Extracted {len(((b.get('payload') or {}).get('boundary') or {}).get('samples') or [])} samples.")
        st.json((b.get("payload") or {}).get("boundary_definition") or {})
        st.download_button("Download feasibility_boundary_v157.json",
                           data=_json.dumps(b, indent=2, sort_keys=True, default=str),
                           file_name="feasibility_boundary_v157.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v157_dl")


# =====================
# v158 Constraint Dominance Topology
# =====================
def _v158_constraint_dominance_panel():
    import streamlit as st
    from tools.constraint_dominance import build_constraint_dominance

    st.subheader("Constraint Dominance Topology")
    st.caption("Explain why regions fail: dominant violated constraint maps + connected components (grid topology).")

    field = st.session_state.get("v156_field")
    up = st.file_uploader("Optional: upload feasibility_field_v156.json", type=["json"], key="v158_upload")
    if up is not None:
        try:
            field = _json.loads(up.getvalue().decode("utf-8"))
        except Exception as e:
            st.error(f"Invalid JSON: {e}")
            return
    if not isinstance(field, dict):
        st.info("Build a Feasibility Field first or upload one here.")
        return

    include_all = st.checkbox("Include feasible points in dominance map (larger JSON)", value=False, key="v158_all")
    if st.button("Compute Dominance Topology", use_container_width=True, key="v158_run"):
        dom = build_constraint_dominance(field=field, only_infeasible=(not include_all))
        st.session_state["v158_dom"] = dom
        summ = (((dom.get("payload") or {}).get("dominance") or {}).get("summary") or {})
        st.success("Computed.")
        st.json(summ)
        st.download_button("Download constraint_dominance_v158.json",
                           data=_json.dumps(dom, indent=2, sort_keys=True, default=str),
                           file_name="constraint_dominance_v158.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v158_dl")


# =====================
# v159 Feasibility Completion Evidence
# =====================
def _v159_feasibility_completion_panel():
    import streamlit as st
    from tools.feasibility_completion_evidence import build_feasibility_completion_evidence

    st.subheader("Feasibility Completion Evidence")
    st.caption("Given partial inputs + bounds on unknowns, search for a feasible completion witness and report bottlenecks.")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    if not runs:
        st.info("Run at least one evaluation first (Point Designer / Systems Mode).")
        return
    ids=[r.get("id") for r in runs]
    run_map={r.get("id"): r for r in runs}
    rid = st.selectbox("Baseline run (provides default known inputs)", options=ids, index=len(ids)-1, key="v159_base_run")
    payload = (run_map.get(rid) or {}).get("payload")
    if not (isinstance(payload, dict) and payload.get("kind")=="shams_run_artifact"):
        st.error("Selected baseline run payload missing.")
        return
    baseline_inputs = payload.get("inputs") or payload.get("_inputs") or {}
    if not isinstance(baseline_inputs, dict) or not baseline_inputs:
        st.error("Baseline inputs missing.")
        return

    st.markdown("### Unknowns to search")
    keys=sorted([k for k in baseline_inputs.keys() if isinstance(k,str)])
    unk = st.multiselect("Select unknown parameters (will be randomized within bounds)", options=keys, default=[k for k in keys if k in ("R0_m","B0_T")], key="v159_unk_keys")
    st.caption('Provide bounds as JSON list: [{"name":"R0_m","bounds":[2.5,3.5]}, ...]')
    default_bounds=_json.dumps([{"name":k, "bounds":[float(baseline_inputs.get(k,0.0) or 0.0)*0.9, float(baseline_inputs.get(k,0.0) or 0.0)*1.1+1e-9]} for k in (unk or [])][:6], indent=2)
    bounds_json = st.text_area("Unknown bounds (JSON)", value=default_bounds, height=140, key="v159_bounds")
    fixed_json = st.text_area("Fixed overrides (JSON list of {name,value})", value="[]", height=80, key="v159_fixed")
    assumptions_json = st.text_area("Assumption set (JSON object)", value="{}", height=80, key="v159_assumptions")

    col1,col2,col3=st.columns(3)
    with col1:
        n_samples = st.number_input("Samples", value=400, min_value=20, max_value=20000, step=20, key="v159_n")
    with col2:
        seed = st.number_input("Seed", value=0, min_value=0, max_value=10_000_000, step=1, key="v159_seed")
    with col3:
        strategy = st.selectbox("Strategy", options=["random","lhs"], index=0, key="v159_strategy")

    if st.button("Search Completion Witness", use_container_width=True, key="v159_run"):
        try:
            unknowns = _json.loads(bounds_json) if bounds_json.strip() else []
            fixed = _json.loads(fixed_json) if fixed_json.strip() else []
            assumptions = _json.loads(assumptions_json) if assumptions_json.strip() else {}
        except Exception as e:
            st.error(f"JSON parse error: {e}")
            return

        # known = baseline minus unknown keys
        known = {k:v for k,v in baseline_inputs.items() if (k not in set(unk or []))}
        with st.spinner("Sampling for feasibility completion..."):
            ev = build_feasibility_completion_evidence(
                known=known,
                unknowns=unknowns,
                fixed=fixed if isinstance(fixed,list) else [],
                assumption_set=assumptions if isinstance(assumptions,dict) else {},
                n_samples=int(n_samples),
                seed=int(seed),
                strategy=str(strategy),
                policy={"generator":"ui"},
            )
        st.session_state["v159_completion"] = ev
        res=((ev.get("payload") or {}).get("result") or {})
        st.success(f"Verdict: {res.get('verdict')}")
        st.json(res.get("bottleneck") or {})
        st.download_button("Download feasibility_completion_evidence_v159.json",
                           data=_json.dumps(ev, indent=2, sort_keys=True, default=str),
                           file_name="feasibility_completion_evidence_v159.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v159_dl")


# =====================
# v161 Completion Frontier + Minimal Change Distance
# =====================
def _v161_completion_frontier_panel():
    import streamlit as st
    from tools.completion_frontier import build_completion_frontier

    st.subheader("Completion Frontier")
    st.caption("Quantify how far a baseline guess is from feasibility: minimal-change feasible witness + distance–margin frontier.")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    if not runs:
        st.info("Run at least one evaluation first (Point Designer / Systems Mode).")
        return
    ids=[r.get("id") for r in runs]
    run_map={r.get("id"): r for r in runs}
    rid = st.selectbox("Baseline run (baseline guess x0)", options=ids, index=len(ids)-1, key="v161_base_run")
    payload = (run_map.get(rid) or {}).get("payload")
    if not (isinstance(payload, dict) and payload.get("kind")=="shams_run_artifact"):
        st.error("Selected baseline run payload missing.")
        return
    baseline_inputs = payload.get("inputs") or payload.get("_inputs") or {}
    if not isinstance(baseline_inputs, dict) or not baseline_inputs:
        st.error("Baseline inputs missing.")
        return

    st.markdown("### Decision variables (vary to reach feasibility)")
    keys=sorted([k for k in baseline_inputs.keys() if isinstance(k,str)])
    dv = st.multiselect("Select decision variables", options=keys, default=[k for k in keys if k in ("R0_m","B0_T","q95","kappa")], key="v161_dv_keys")
    default_bounds=_json.dumps([{"name":k, "bounds":[float(baseline_inputs.get(k,0.0) or 0.0)*0.9, float(baseline_inputs.get(k,0.0) or 0.0)*1.1+1e-9]} for k in (dv or [])][:8], indent=2)
    st.caption('Bounds JSON example: [{"name":"R0_m","bounds":[2.5,3.5]}, ...]')
    bounds_json = st.text_area("Decision variable bounds (JSON)", value=default_bounds, height=160, key="v161_bounds")
    fixed_json = st.text_area("Fixed overrides (JSON list of {name,value})", value="[]", height=80, key="v161_fixed")
    assumptions_json = st.text_area("Assumption set (JSON object)", value="{}", height=80, key="v161_assumptions")

    col1,col2,col3=st.columns(3)
    with col1:
        n_samples = st.number_input("Samples", value=800, min_value=40, max_value=50000, step=40, key="v161_n")
    with col2:
        seed = st.number_input("Seed", value=0, min_value=0, max_value=10_000_000, step=1, key="v161_seed")
    with col3:
        strategy = st.selectbox("Strategy", options=["random","lhs"], index=0, key="v161_strategy")

    if st.button("Compute Completion Frontier", use_container_width=True, key="v161_run"):
        try:
            vars_spec = _json.loads(bounds_json) if bounds_json.strip() else []
            fixed = _json.loads(fixed_json) if fixed_json.strip() else []
            assumptions = _json.loads(assumptions_json) if assumptions_json.strip() else {}
        except Exception as e:
            st.error(f"JSON parse error: {e}")
            return
        with st.spinner("Sampling and evaluating frontier..."):
            out = build_completion_frontier(
                baseline=baseline_inputs,
                decision_vars=vars_spec,
                fixed=fixed if isinstance(fixed,list) else [],
                assumption_set=assumptions if isinstance(assumptions,dict) else {},
                n_samples=int(n_samples),
                seed=int(seed),
                strategy=str(strategy),
                policy={"generator":"ui"},
            )
        st.session_state["v161_frontier"] = out
        res=((out.get("payload") or {}).get("result") or {})
        st.success("Computed.")
        st.markdown("**Minimal-change feasible witness**")
        st.json(res.get("minimal_change_feasible") or {})
        st.markdown("**Frontier (preview)**")
        st.json((res.get("frontier") or [])[:20])
        st.download_button("Download completion_frontier_v161.json",
                           data=_json.dumps(out, indent=2, sort_keys=True, default=str),
                           file_name="completion_frontier_v161.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v161_dl")


# =====================
# v162 Directed Local Search (safe)
# =====================
def _v162_directed_local_search_panel():
    import streamlit as st
    from tools.directed_local_search import build_directed_local_search

    st.subheader("Directed Local Search")
    st.caption("A safe, bounded local search that tries to reach feasibility with minimal evaluations (downstream-only).")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    if not runs:
        st.info("Run at least one evaluation first (Point Designer / Systems Mode).")
        return
    ids=[r.get("id") for r in runs]
    run_map={r.get("id"): r for r in runs}
    rid = st.selectbox("Baseline run (starting guess)", options=ids, index=len(ids)-1, key="v162_base_run")
    payload = (run_map.get(rid) or {}).get("payload")
    if not (isinstance(payload, dict) and payload.get("kind")=="shams_run_artifact"):
        st.error("Selected baseline run payload missing.")
        return
    baseline_inputs = payload.get("inputs") or payload.get("_inputs") or {}
    if not isinstance(baseline_inputs, dict) or not baseline_inputs:
        st.error("Baseline inputs missing.")
        return

    st.markdown("### Decision variables")
    keys=sorted([k for k in baseline_inputs.keys() if isinstance(k,str)])
    dv = st.multiselect("Select decision variables to adjust", options=keys, default=[k for k in keys if k in ("R0_m","B0_T","q95","kappa")], key="v162_dv_keys")
    default_bounds=_json.dumps([{"name":k, "bounds":[float(baseline_inputs.get(k,0.0) or 0.0)*0.9, float(baseline_inputs.get(k,0.0) or 0.0)*1.1+1e-9]} for k in (dv or [])][:8], indent=2)
    st.caption('Bounds JSON example: [{"name":"R0_m","bounds":[2.5,3.5]}, ...]')
    bounds_json = st.text_area("Decision variable bounds (JSON)", value=default_bounds, height=160, key="v162_bounds")
    fixed_json = st.text_area("Fixed overrides (JSON list of {name,value})", value="[]", height=80, key="v162_fixed")
    assumptions_json = st.text_area("Assumption set (JSON object)", value="{}", height=80, key="v162_assumptions")

    col1,col2,col3,col4=st.columns(4)
    with col1:
        max_evals = st.number_input("Max evals", value=200, min_value=30, max_value=5000, step=10, key="v162_maxeval")
    with col2:
        seed = st.number_input("Seed", value=0, min_value=0, max_value=10_000_000, step=1, key="v162_seed")
    with col3:
        init_step = st.number_input("Initial step (norm)", value=0.12, min_value=0.01, max_value=0.50, step=0.01, key="v162_initstep")
    with col4:
        min_step = st.number_input("Min step (norm)", value=0.004, min_value=0.0005, max_value=0.10, step=0.0005, format="%.4f", key="v162_minstep")

    shrink = st.slider("Step shrink factor", min_value=0.2, max_value=0.9, value=0.5, step=0.05, key="v162_shrink")

    if st.button("Run Directed Search", use_container_width=True, key="v162_run"):
        try:
            vars_spec = _json.loads(bounds_json) if bounds_json.strip() else []
            fixed = _json.loads(fixed_json) if fixed_json.strip() else []
            assumptions = _json.loads(assumptions_json) if assumptions_json.strip() else {}
        except Exception as e:
            st.error(f"JSON parse error: {e}")
            return
        with st.spinner("Searching..."):
            out = build_directed_local_search(
                baseline=baseline_inputs,
                decision_vars=vars_spec,
                fixed=fixed if isinstance(fixed,list) else [],
                assumption_set=assumptions if isinstance(assumptions,dict) else {},
                max_evals=int(max_evals),
                seed=int(seed),
                initial_step_norm=float(init_step),
                min_step_norm=float(min_step),
                step_shrink=float(shrink),
                policy={"generator":"ui"},
            )
        st.session_state["v162_local_search"] = out
        res=((out.get("payload") or {}).get("result") or {})
        st.success(f"Stop reason: {res.get('stop_reason')}")
        st.json(res.get("final") or {})
        st.download_button("Download directed_local_search_v162.json",
                           data=_json.dumps(out, indent=2, sort_keys=True, default=str),
                           file_name="directed_local_search_v162.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v162_dl")


# =====================
# v163 Completion Pack (Actionable Recipe)
# =====================
def _v163_completion_pack_panel():
    import streamlit as st
    from tools.completion_pack import build_completion_pack, render_completion_pack_markdown

    st.subheader("Completion Pack")
    st.caption("Bundle completion evidence into an actionable recipe: witness + knob ranking + bounds recommendations.")

    v159 = st.session_state.get("v159_completion")
    v161 = st.session_state.get("v161_frontier")
    v162 = st.session_state.get("v162_local_search")

    st.markdown("### Inputs")
    colA,colB,colC = st.columns(3)
    with colA:
        up159 = st.file_uploader("Optional: upload v159 evidence JSON", type=["json"], key="v163_up159")
    with colB:
        up161 = st.file_uploader("Optional: upload v161 frontier JSON", type=["json"], key="v163_up161")
    with colC:
        up162 = st.file_uploader("Optional: upload v162 local search JSON", type=["json"], key="v163_up162")

    def _load(up):
        if up is None:
            return None
        try:
            return _json.loads(up.getvalue().decode("utf-8"))
        except Exception:
            return None

    v159 = _load(up159) or v159
    v161 = _load(up161) or v161
    v162 = _load(up162) or v162

    tighten = st.slider("Bounds tighten factor (heuristic)", min_value=0.05, max_value=0.45, value=0.25, step=0.05, key="v163_tighten")

    if st.button("Build Completion Pack", use_container_width=True, key="v163_run"):
        pack = build_completion_pack(v159=v159, v161=v161, v162=v162, policy={"generator":"ui", "tighten": float(tighten)})
        st.session_state["v163_pack"] = pack
        st.success(f"Witness provenance: {((pack.get('payload') or {}).get('witness_provenance') or '')}")
        st.markdown("### Preview")
        st.json((pack.get("payload") or {}) )
        st.download_button("Download completion_pack_v163.json",
                           data=_json.dumps(pack, indent=2, sort_keys=True, default=str),
                           file_name="completion_pack_v163.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v163_dl_json")
        md = render_completion_pack_markdown(pack)
        st.download_button("Download completion_pack_summary_v163.md",
                           data=md,
                           file_name="completion_pack_summary_v163.md",
                           mime="text/markdown",
                           use_container_width=True,
                           key="v163_dl_md")


# =====================
# v164 Sensitivity + Bottleneck Attribution
# =====================
def _v164_sensitivity_panel():
    import streamlit as st
    from tools.sensitivity_v164 import build_sensitivity_report, render_sensitivity_markdown

    st.subheader("Sensitivity + Bottleneck Attribution")
    st.caption("Local finite-difference sensitivities around a witness: ranked leverage variables and dominant-constraint changes.")

    # Prefer completion pack witness if present, else last run
    pack = st.session_state.get("v163_pack")
    witness = None
    if isinstance(pack, dict):
        witness = ((pack.get("payload") or {}).get("recommended_witness"))
    if not isinstance(witness, dict):
        # fallback to baseline last run
        s = _v98_state_init_runlists()
        runs = [r for r in (s.run_history or []) if r.get("id")]
        if runs:
            payload = (runs[-1].get("payload") or {})
            if isinstance(payload, dict) and payload.get("kind")=="shams_run_artifact":
                witness = payload.get("inputs") or payload.get("_inputs")

    if not isinstance(witness, dict) or not witness:
        st.info("No witness found yet. Run v159/v161/v162 or build v163 completion pack first.")
        return

    st.markdown("### Witness source")
    st.code(f"Using witness keys: {len(witness)}", language="text")

    st.markdown("### Variables to perturb")
    keys=sorted([k for k in witness.keys() if isinstance(k,str)])
    dv = st.multiselect("Select variables", options=keys, default=[k for k in keys if k in ("R0_m","B0_T","q95","kappa")], key="v164_dv")
    default_bounds=_json.dumps([{"name":k, "bounds":[float(witness.get(k,0.0) or 0.0)*0.9, float(witness.get(k,0.0) or 0.0)*1.1+1e-9]} for k in (dv or [])][:10], indent=2)
    st.caption('Bounds JSON example: [{"name":"R0_m","bounds":[2.5,3.5]}, ...]')
    bounds_json = st.text_area("Variable bounds (JSON)", value=default_bounds, height=160, key="v164_bounds")
    assumptions_json = st.text_area("Assumption set (JSON object)", value="{}", height=80, key="v164_assumptions")

    col1,col2=st.columns(2)
    with col1:
        rel_step = st.number_input("Relative step (fraction of span)", value=0.01, min_value=0.001, max_value=0.10, step=0.001, format="%.3f", key="v164_relstep")
    with col2:
        abs_step_min = st.number_input("Absolute step min", value=1e-6, min_value=0.0, max_value=1e-2, step=1e-6, format="%.6f", key="v164_absmin")

    if st.button("Compute Sensitivity", use_container_width=True, key="v164_run"):
        try:
            vars_spec = _json.loads(bounds_json) if bounds_json.strip() else []
            assumptions = _json.loads(assumptions_json) if assumptions_json.strip() else {}
        except Exception as e:
            st.error(f"JSON parse error: {e}")
            return
        with st.spinner("Evaluating baseline and perturbations..."):
            rep = build_sensitivity_report(
                witness=witness,
                variables=vars_spec,
                assumption_set=assumptions if isinstance(assumptions,dict) else {},
                rel_step=float(rel_step),
                abs_step_min=float(abs_step_min),
                policy={"generator":"ui"},
            )
        st.session_state["v164_sensitivity"] = rep
        ranked=((rep.get("payload") or {}).get("ranked") or [])
        st.success(f"Computed. Ranked variables: {len(ranked)}")
        st.markdown("### Ranked leverage (preview)")
        st.json(ranked[:15])
        st.download_button("Download sensitivity_v164.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True, default=str),
                           file_name="sensitivity_v164.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v164_dl_json")
        md = render_sensitivity_markdown(rep)
        st.download_button("Download sensitivity_v164.md",
                           data=md,
                           file_name="sensitivity_v164.md",
                           mime="text/markdown",
                           use_container_width=True,
                           key="v164_dl_md")


# =====================
# v165 Study Protocol Generator
# =====================
def _v165_study_protocol_panel():
    import streamlit as st
    from tools.study_protocol_v165 import build_study_protocol, render_study_protocol_markdown

    st.subheader("Study Protocol Generator")
    st.caption("Generate journal-ready, audit-ready study protocol (Methods) with protocol SHA-256. Reporting-only.")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    if not runs:
        st.info("Run at least one evaluation first (Point Designer / Systems Mode) to generate a run artifact.")
        return
    ids=[r.get("id") for r in runs]
    run_map={r.get("id"): r for r in runs}
    rid = st.selectbox("Select run artifact", options=ids, index=len(ids)-1, key="v165_run_id")
    run_art = (run_map.get(rid) or {}).get("payload")
    if not (isinstance(run_art, dict) and run_art.get("kind")=="shams_run_artifact"):
        st.error("Selected run is missing shams_run_artifact payload.")
        return

    col1,col2=st.columns(2)
    with col1:
        title = st.text_input("Study title", value="SHAMS Design Study", key="v165_title")
    with col2:
        study_id = st.text_input("Study ID (optional)", value="", key="v165_study_id")

    objective = st.text_area("Objective (paper-ready)", value="Feasibility characterization and completion under explicit constraints.", height=80, key="v165_obj")
    notes = st.text_area("Notes (one per line)", value="", height=60, key="v165_notes")

    st.markdown("### Optional: variables varied / scan definition")
    vars_json = st.text_area("variables_varied (JSON list)", value="[]", height=120, key="v165_vars")
    artifacts_json = st.text_area("artifacts_generated (JSON list)", value='["study_protocol_v165.json","study_protocol_v165.md"]', height=80, key="v165_artifacts")
    seed = st.number_input("Seed (optional)", value=0, min_value=0, max_value=10_000_000, step=1, key="v165_seed")

    if st.button("Generate Study Protocol", use_container_width=True, key="v165_run"):
        try:
            vv = _json.loads(vars_json) if vars_json.strip() else []
            ag = _json.loads(artifacts_json) if artifacts_json.strip() else []
        except Exception as e:
            st.error(f"JSON parse error: {e}")
            return
        overrides={
            "title": title,
            "study_id": study_id,
            "objective": objective,
            "notes": [ln.strip() for ln in notes.splitlines() if ln.strip()],
            "variables_varied": vv if isinstance(vv, list) else [],
            "artifacts_generated": ag if isinstance(ag, list) else [],
            "seed": int(seed),
        }
        prot = build_study_protocol(run_artifact=run_art, protocol_overrides=overrides)
        st.session_state["v165_protocol"] = prot
        sha = ((prot.get("payload") or {}).get("integrity") or {}).get("protocol_sha256")
        st.success(f"Generated. Protocol SHA-256: {sha}")
        st.download_button("Download study_protocol_v165.json",
                           data=_json.dumps(prot, indent=2, sort_keys=True, default=str),
                           file_name="study_protocol_v165.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v165_dl_json")
        md = render_study_protocol_markdown(prot)
        st.download_button("Download study_protocol_v165.md",
                           data=md,
                           file_name="study_protocol_v165.md",
                           mime="text/markdown",
                           use_container_width=True,
                           key="v165_dl_md")


# =====================
# v166 Reproducibility Lock + Replay Checker
# =====================
def _v166_repro_lock_panel():
    import streamlit as st
    from tools.repro_lock_v166 import build_repro_lock, replay_check

    st.subheader("Reproducibility Lock + Replay Checker")
    st.caption("Freeze a run (inputs + assumptions + solver meta) into a lockfile and verify replay matches within tolerances.")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    if not runs:
        st.info("Run at least one evaluation first (Point Designer / Systems Mode).")
        return
    ids=[r.get("id") for r in runs]
    run_map={r.get("id"): r for r in runs}
    rid = st.selectbox("Select run artifact to lock", options=ids, index=len(ids)-1, key="v166_run_id")
    run_art = (run_map.get(rid) or {}).get("payload")
    if not (isinstance(run_art, dict) and run_art.get("kind")=="shams_run_artifact"):
        st.error("Selected run is missing shams_run_artifact payload.")
        return

    st.markdown("### Tolerances (edit JSON)")
    tol_default = _json.dumps({
        "min_margin_abs": 1e-8,
        "constraint_margin_abs": 1e-6,
        "metric_rel": 1e-6,
        "metric_abs": 1e-9,
    }, indent=2)
    tol_json = st.text_area("tolerances JSON", value=tol_default, height=140, key="v166_tol")

    notes = st.text_area("Notes (one per line)", value="", height=60, key="v166_notes")

    col1,col2=st.columns(2)
    with col1:
        if st.button("Create Lock", use_container_width=True, key="v166_lock"):
            try:
                tol=_json.loads(tol_json) if tol_json.strip() else {}
            except Exception as e:
                st.error(f"JSON parse error: {e}")
                return
            lock = build_repro_lock(run_artifact=run_art, lock_overrides={"tolerances": tol, "notes":[ln.strip() for ln in notes.splitlines() if ln.strip()]})
            st.session_state["v166_lock"] = lock
            sha = ((lock.get("payload") or {}).get("integrity") or {}).get("lock_sha256")
            st.success(f"Lock created. lock_sha256: {sha}")
            st.download_button("Download repro_lock_v166.json",
                               data=_json.dumps(lock, indent=2, sort_keys=True, default=str),
                               file_name="repro_lock_v166.json",
                               mime="application/json",
                               use_container_width=True,
                               key="v166_dl_lock")
    with col2:
        lock_up = st.file_uploader("Or upload existing lock JSON", type=["json"], key="v166_up_lock")

    lock = st.session_state.get("v166_lock")
    if lock_up is not None:
        try:
            lock = _json.loads(lock_up.getvalue().decode("utf-8"))
            st.session_state["v166_lock"] = lock
        except Exception:
            st.warning("Could not parse uploaded lock JSON.")

    st.markdown("### Replay check")
    ao_json = st.text_area("assumption_set override (JSON, optional)", value="{}", height=80, key="v166_ao")
    if st.button("Run Replay Check", use_container_width=True, key="v166_replay"):
        if not isinstance(lock, dict) or lock.get("kind")!="shams_repro_lock":
            st.error("No valid lock loaded. Create or upload a lock first.")
            return
        try:
            ao = _json.loads(ao_json) if ao_json.strip() else {}
        except Exception as e:
            st.error(f"JSON parse error: {e}")
            return
        rep = replay_check(lock=lock, assumption_set_override=ao if isinstance(ao, dict) else None, policy={"generator":"ui"})
        st.session_state["v166_replay"] = rep
        ok = ((rep.get("payload") or {}).get("ok"))
        st.success("Replay OK" if ok else "Replay NOT OK")
        st.json((rep.get("payload") or {}).get("checks") or {})
        st.download_button("Download replay_report_v166.json",
                           data=_json.dumps(rep, indent=2, sort_keys=True, default=str),
                           file_name="replay_report_v166.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v166_dl_rep")


# =====================
# v167 Design Study Authority Pack
# =====================
def _v167_authority_pack_panel():
    import streamlit as st
    from tools.authority_pack_v167 import build_authority_pack

    st.subheader("Design Study Authority Pack")
    st.caption("One downloadable ZIP bundling protocol + lock + replay + completion + sensitivity (+ certificate if available).")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    if not runs:
        st.info("Run at least one evaluation first (Point Designer / Systems Mode).")
        return
    ids=[r.get("id") for r in runs]
    run_map={r.get("id"): r for r in runs}
    rid = st.selectbox("Select run artifact to include", options=ids, index=len(ids)-1, key="v167_run_id")
    run_art = (run_map.get(rid) or {}).get("payload")
    if not (isinstance(run_art, dict) and run_art.get("kind")=="shams_run_artifact"):
        st.error("Selected run is missing shams_run_artifact payload.")
        return

    st.markdown("### Auto-pick from session state (recommended)")
    prot = st.session_state.get("v165_protocol")
    lock = st.session_state.get("v166_lock")
    replay = st.session_state.get("v166_replay")
    comp = st.session_state.get("v163_pack")
    sens = st.session_state.get("v164_sensitivity")
    cert = st.session_state.get("v160_certificate")

    st.markdown("### Optional: upload any missing JSONs")
    col1,col2,col3 = st.columns(3)
    with col1:
        up_prot = st.file_uploader("upload study_protocol_v165.json", type=["json"], key="v167_up_prot")
        up_lock = st.file_uploader("upload repro_lock_v166.json", type=["json"], key="v167_up_lock")
    with col2:
        up_rep = st.file_uploader("upload replay_report_v166.json", type=["json"], key="v167_up_rep")
        up_comp = st.file_uploader("upload completion_pack_v163.json", type=["json"], key="v167_up_comp")
    with col3:
        up_sens = st.file_uploader("upload sensitivity_v164.json", type=["json"], key="v167_up_sens")
        up_cert = st.file_uploader("upload certificate_v160.json", type=["json"], key="v167_up_cert")

    def _load(up):
        if up is None:
            return None
        try:
            return _json.loads(up.getvalue().decode("utf-8"))
        except Exception:
            return None

    prot = _load(up_prot) or prot
    lock = _load(up_lock) or lock
    replay = _load(up_rep) or replay
    comp = _load(up_comp) or comp
    sens = _load(up_sens) or sens
    cert = _load(up_cert) or cert

    if st.button("Build Authority Pack ZIP", use_container_width=True, key="v167_build"):
        res = build_authority_pack(
            run_artifact=run_art,
            study_protocol_v165=prot,
            repro_lock_v166=lock,
            replay_report_v166=replay,
            completion_pack_v163=comp,
            sensitivity_v164=sens,
            certificate_v160=cert,
            policy={"generator":"ui"},
        )
        st.session_state["v167_manifest"] = res["manifest"]
        st.success(f"Built authority_pack_v167.zip (sha256={res['pack']['integrity']['zip_sha256']})")
        st.json(res["manifest"])
        st.download_button("Download authority_pack_v167.zip",
                           data=res["zip_bytes"],
                           file_name="authority_pack_v167.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v167_dl_zip")


# =====================
# v168 Citation-Grade Study Reference
# =====================
def _v168_citation_panel():
    import streamlit as st
    from tools.citation_v168 import build_citation_bundle

    st.subheader("Citation-Grade Study Reference")
    st.caption("Generate Study ID + CITATION.cff + BibTeX from protocol/lock/authority-pack manifest.")

    prot = st.session_state.get("v165_protocol")
    lock = st.session_state.get("v166_lock")
    manifest = st.session_state.get("v167_manifest")

    st.markdown("### Inputs (auto-pick from session state or upload)")
    col1,col2,col3 = st.columns(3)
    with col1:
        up_prot = st.file_uploader("upload study_protocol_v165.json", type=["json"], key="v168_up_prot")
    with col2:
        up_lock = st.file_uploader("upload repro_lock_v166.json", type=["json"], key="v168_up_lock")
    with col3:
        up_man = st.file_uploader("upload authority_pack_manifest_v167.json", type=["json"], key="v168_up_man")

    def _load(up):
        if up is None:
            return None
        try:
            return _json.loads(up.getvalue().decode("utf-8"))
        except Exception:
            return None

    prot = _load(up_prot) or prot
    lock = _load(up_lock) or lock
    manifest = _load(up_man) or manifest

    st.markdown("### Metadata (optional)")
    title = st.text_input("Title override (optional)", value="", key="v168_title")
    repo_url = st.text_input("Repository/URL (optional)", value="", key="v168_repo")
    doi = st.text_input("DOI (optional)", value="", key="v168_doi")
    author = st.text_input("Author name (optional)", value="SHAMS–FUSION-X Contributors", key="v168_author")

    if st.button("Generate Citation Bundle", use_container_width=True, key="v168_run"):
        if not (isinstance(prot, dict) and prot.get("kind")=="shams_study_protocol"):
            st.error("Need a valid study_protocol_v165.json (generate in v165).")
            return
        meta={
            "title": title or None,
            "repository": repo_url or None,
            "url": repo_url or None,
            "doi": doi or None,
            "authors": [{"name": author}] if author else [{"name":"SHAMS–FUSION-X Contributors"}],
            "version": "v168",
        }
        res = build_citation_bundle(
            study_protocol_v165=prot,
            repro_lock_v166=lock if isinstance(lock, dict) else None,
            authority_pack_manifest_v167=manifest if isinstance(manifest, dict) else None,
            metadata=meta,
        )
        st.session_state["v168_citation"] = res
        sid = ((res.get("payload") or {}).get("study_id"))
        st.success(f"Study ID: {sid}")
        st.download_button("Download citation_bundle_v168.json",
                           data=_json.dumps(res, indent=2, sort_keys=True, default=str),
                           file_name="citation_bundle_v168.json",
                           mime="application/json",
                           use_container_width=True,
                           key="v168_dl_json")
        st.download_button("Download CITATION.cff",
                           data=(res.get("payload") or {}).get("citation_cff_text",""),
                           file_name="CITATION.cff",
                           mime="text/yaml",
                           use_container_width=True,
                           key="v168_dl_cff")
        st.download_button("Download study_citation_v168.bib",
                           data=(res.get("payload") or {}).get("bibtex_text",""),
                           file_name="study_citation_v168.bib",
                           mime="text/plain",
                           use_container_width=True,
                           key="v168_dl_bib")
        st.download_button("Download study_reference_v168.md",
                           data=(res.get("payload") or {}).get("reference_markdown",""),
                           file_name="study_reference_v168.md",
                           mime="text/markdown",
                           use_container_width=True,
                           key="v168_dl_md")


# =====================
# v169 Feasibility Boundary Atlas (figure pack)
# =====================
def _v169_atlas_panel():
    import streamlit as st
    from tools.atlas_v169 import build_atlas_pack

    st.subheader("Feasibility Boundary Atlas")
    st.caption("Generate a publishable atlas-style figure pack with consistent captions and a hash manifest.")

    sens = st.session_state.get("v164_sensitivity")
    st.markdown("### Input")
    up_sens = st.file_uploader("Upload sensitivity_v164.json (optional)", type=["json"], key="v169_up_sens")

    if up_sens is not None:
        try:
            sens = _json.loads(up_sens.getvalue().decode("utf-8"))
        except Exception:
            st.warning("Could not parse uploaded JSON.")

    if not isinstance(sens, dict):
        st.info("Run v164 Sensitivity first (or upload sensitivity_v164.json) to build the initial atlas.")
        return

    if st.button("Build Atlas Pack ZIP", use_container_width=True, key="v169_build"):
        res = build_atlas_pack(sensitivity_v164=sens, policy={"generator":"ui"})
        st.session_state["v169_manifest"] = res["manifest"]
        st.success(f"Built atlas_pack_v169.zip (sha256={res['pack']['integrity']['zip_sha256']})")
        st.json(res["manifest"])
        st.download_button("Download atlas_pack_v169.zip",
                           data=res["zip_bytes"],
                           file_name="atlas_pack_v169.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v169_dl_zip")


# =====================
# v170 SHAMS → external systems codes Downstream Export
# =====================
def _v170_process_export_panel():
    import streamlit as st
    from tools.process_export_v170 import build_process_export_pack

    st.subheader("SHAMS → external systems codes Downstream Export")
    st.caption("Export SHAMS study outputs into transparent (systems-code-inspired) tables, keeping SHAMS as upstream authority.")

    s = _v98_state_init_runlists()
    runs = [r for r in (s.run_history or []) if r.get("id")]
    if not runs:
        st.info("Run at least one evaluation first (Point Designer / Systems Mode).")
        return
    ids=[r.get("id") for r in runs]
    run_map={r.get("id"): r for r in runs}
    rid = st.selectbox("Select run artifact to export", options=ids, index=len(ids)-1, key="v170_run_id")
    run_art = (run_map.get(rid) or {}).get("payload")
    if not (isinstance(run_art, dict) and run_art.get("kind")=="shams_run_artifact"):
        st.error("Selected run is missing shams_run_artifact payload.")
        return

    comp = st.session_state.get("v163_pack")
    cite = st.session_state.get("v168_citation")

    st.markdown("### Optional attachments")
    st.write("- completion_pack_v163.json:", "yes" if isinstance(comp, dict) else "no")
    st.write("- citation_bundle_v168.json:", "yes" if isinstance(cite, dict) else "no")

    if st.button("Build external systems codes Export Pack ZIP", use_container_width=True, key="v170_build"):
        res = build_process_export_pack(
            run_artifact=run_art,
            completion_pack_v163=comp if isinstance(comp, dict) else None,
            citation_bundle_v168=cite if isinstance(cite, dict) else None,
            policy={"generator":"ui"},
        )
        st.session_state["v170_manifest"] = res["manifest"]
        st.success(f"Built process_export_pack_v170.zip (sha256={res['pack']['integrity']['zip_sha256']})")
        st.json(res["manifest"])
        st.download_button("Download process_export_pack_v170.zip",
                           data=res["zip_bytes"],
                           file_name="process_export_pack_v170.zip",
                           mime="application/zip",
                           use_container_width=True,
                           key="v170_dl_zip")


# =====================
# v172 Demo seed + hydration
# =====================
def _v172_demo_loader():
    import streamlit as st
    from tools.demo_seed_v172 import install_demo_bundle
    st.subheader("Demo seed")
    st.caption("Loads synthetic demo artifacts into session state so every panel shows content offline. Not authoritative.")
    col1,col2 = st.columns([1,2])
    with col1:
        if st.button("Load demo artifacts", use_container_width=True, key="v172_load_demo"):
            install_demo_bundle(st.session_state)
            st.success("Demo artifacts loaded into session state.")
    with col2:
        if st.button("Clear demo artifacts", use_container_width=True, key="v172_clear_demo"):
            for k in ["v164_sensitivity","v165_protocol","v166_lock","v166_replay","v167_manifest","v168_citation","v163_pack",
                      "pd_last_outputs","pd_last_artifact","demo_run_artifact"]:
                if k in st.session_state:
                    del st.session_state[k]
            st.success("Demo artifacts cleared.")


# --- Deferred PAM render ---
try:
    if 'pam_placeholder' in globals() and pam_placeholder is not None:
        with pam_placeholder.container():
            _v175_panel_availability_map_panel()
except Exception as _e:
    # Never fail the whole app due to PAM rendering
    try:
        st.warning(f'PAM render failed: {_e}')
    except Exception:
        pass

# --- Render Activity Log late so it includes events logged during this run ---
try:
    _render_activity_log_sidebar()
except Exception:
    pass

_render_footer()