
"""
phase1_hf_compact_conf_improvement_scan_v2.py

Driver: scan confinement gain g_conf and find minimum improvement that yields >=1
"clean" extended-feasible operating point.

This script is drop-in compatible with the previous Phase-1 scan conventions:
- CLI args are preserved and extended (new constraints are additive).
- Output workbook uses sheets:
    - feasible_ext
    - failures (optional)
    - meta
with autosized columns + frozen header rows.

Physics/solvers:
- Delegates point physics and solvers to phase1_core.py.
- Added "clean point design" screening is computed inside hot_ion_point() and
  enforced here as explicit constraints.

NOTE:
- All added models are *screening proxies* (see phase1_systems.py).
- The point of the scan is to search design space quickly and transparently.
"""

from __future__ import annotations

import argparse
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font

from phase1_core import PointInputs, solve_Ip_for_H98_with_Q_match


def frange(start: float, stop: float, step: float) -> List[float]:
    """
    Inclusive-ish float range using step sign.
    Matches the intent of numpy.arange in the original scripts.
    """
    vals: List[float] = []
    if step == 0:
        return [start]
    x = start
    if step > 0:
        while x <= stop + 1e-12:
            vals.append(float(x))
            x += step
    else:
        while x >= stop - 1e-12:
            vals.append(float(x))
            x += step
    return vals


def autosize_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def main() -> None:
    p = argparse.ArgumentParser(
        description="Scan confinement gain g_conf and find minimum improvement that yields >=1 clean-feasible point."
    )

    # ---------------------------------------------------------------------
    # Primary machine knobs (Phase-1)
    # ---------------------------------------------------------------------
    p.add_argument("--R0", type=float, default=1.81)
    p.add_argument("--B0", type=float, default=10.0)
    p.add_argument("--tshield", type=float, default=0.8)

    p.add_argument("--Paux", type=float, default=48.0)
    p.add_argument("--Paux_for_Q", type=float, default=48.0)

    p.add_argument("--Ti_over_Te", type=float, default=2.0)

    # ---------------------------------------------------------------------
    # Scan axes (Phase-1)
    # ---------------------------------------------------------------------
    p.add_argument("--Ti_start", type=float, default=16.0)
    p.add_argument("--Ti_stop", type=float, default=8.0)
    p.add_argument("--Ti_step", type=float, default=0.25)

    p.add_argument("--H98_start", type=float, default=1.0)
    p.add_argument("--H98_stop", type=float, default=1.7)
    p.add_argument("--H98_step", type=float, default=0.05)

    p.add_argument("--a_min", type=float, default=0.45)
    p.add_argument("--a_max", type=float, default=0.72)
    p.add_argument("--a_step", type=float, default=0.01)

    p.add_argument("--Q_start", type=float, default=0.8)
    p.add_argument("--Q_stop", type=float, default=2.0)
    p.add_argument("--Q_step", type=float, default=0.1)

    # Solver bounds
    p.add_argument("--Ip_min", type=float, default=10.0)
    p.add_argument("--Ip_max", type=float, default=120.0)
    p.add_argument("--fG_min", type=float, default=0.01)
    p.add_argument("--fG_max", type=float, default=1.20)
    p.add_argument("--tol", type=float, default=1e-3)

    # Confinement-improvement scan
    p.add_argument("--gconf_start", type=float, default=1.0)
    p.add_argument("--gconf_stop", type=float, default=2.5)
    p.add_argument("--gconf_step", type=float, default=0.05)

    # Output
    p.add_argument("--out_xlsx", type=str, default="phase1_hf_compact_conf_scan.xlsx")
    p.add_argument("--write_failures", action="store_true", default=False)

    # ---------------------------------------------------------------------
    # Plasma/physics knobs (Phase-1 extended)
    # ---------------------------------------------------------------------
    p.add_argument("--Zeff", type=float, default=1.8)
    p.add_argument("--dilution_fuel", type=float, default=0.85)
    p.add_argument("--alpha_loss_frac", type=float, default=0.05)

    # Radiated fractions (new, replaces ad-hoc "extra_rad_factor")
    p.add_argument("--f_rad_core", type=float, default=0.20)
    p.add_argument("--f_rad_div", type=float, default=0.30)

    # Shape/proxy constraints
    p.add_argument("--kappa", type=float, default=1.8)
    p.add_argument("--q95_min", type=float, default=3.0)
    p.add_argument("--betaN_max", type=float, default=4.0)

    p.add_argument("--C_bs", type=float, default=0.15)
    p.add_argument("--f_bs_max", type=float, default=0.60)

    # H-mode access constraint
    p.add_argument("--require_Hmode", action="store_true", default=False)
    p.add_argument("--PLH_margin", type=float, default=0.0)

    # SOL power density proxy constraint
    p.add_argument("--PSOL_over_R_max", type=float, default=80.0)

    # ---------------------------------------------------------------------
    # (1) Radial build + magnet stress + Bpeak mapping
    # ---------------------------------------------------------------------
    p.add_argument("--t_fw", type=float, default=0.02)
    p.add_argument("--t_blanket", type=float, default=0.50)
    p.add_argument("--t_vv", type=float, default=0.05)
    p.add_argument("--t_gap", type=float, default=0.03)
    p.add_argument("--t_tf_wind", type=float, default=0.20)
    p.add_argument("--t_tf_struct", type=float, default=0.15)

    p.add_argument("--Bpeak_factor", type=float, default=1.05)
    p.add_argument("--sigma_allow_MPa", type=float, default=800.0)

    # ---------------------------------------------------------------------
    # (2) HTS critical margin (B,T) + quench/dump constraint proxy
    # ---------------------------------------------------------------------
    p.add_argument("--Tcoil_K", type=float, default=20.0)
    p.add_argument("--hts_margin_min", type=float, default=1.2)
    p.add_argument("--N_tf_turns", type=int, default=1)
    p.add_argument("--tau_dump_s", type=float, default=10.0)
    p.add_argument("--Vmax_kV", type=float, default=20.0)
    p.add_argument("--tf_energy_volume_factor", type=float, default=1.0)

    # ---------------------------------------------------------------------
    # (3) Divertor heat flux constraint (λq + L∥ + detachment/radiation fraction)
    # ---------------------------------------------------------------------
    p.add_argument("--lambda_q_factor", type=float, default=1.0)
    p.add_argument("--flux_expansion", type=float, default=5.0)
    p.add_argument("--q_div_max", type=float, default=10.0)
    p.add_argument("--f_Lpar", type=float, default=1.0)

    # ---------------------------------------------------------------------
    # (4) Neutronics lifetime/TBR feasibility (shield thickness & coverage)
    # ---------------------------------------------------------------------
    p.add_argument("--blanket_coverage", type=float, default=0.80)
    p.add_argument("--TBR_min", type=float, default=1.05)
    p.add_argument("--TBR_lambda_m", type=float, default=0.30)
    p.add_argument("--TBR_multiplier", type=float, default=1.10)

    p.add_argument("--hts_fluence_limit_n_m2", type=float, default=3.0e22)
    p.add_argument("--atten_len_m", type=float, default=0.25)
    p.add_argument("--f_geom_to_tf", type=float, default=0.05)
    p.add_argument("--lifetime_min_yr", type=float, default=3.0)

    # ---------------------------------------------------------------------
    # (5) Recirculating power closure -> net power
    # ---------------------------------------------------------------------
    p.add_argument("--eta_elec", type=float, default=0.40)
    p.add_argument("--eta_aux_wallplug", type=float, default=0.40)
    p.add_argument("--eta_cd_wallplug", type=float, default=0.33)
    p.add_argument("--eta_cd_MA_per_MW", type=float, default=0.05)
    p.add_argument("--steady_state", action="store_true", default=True)
    p.add_argument("--no_steady_state", action="store_true", default=False)

    p.add_argument("--R_joint_ohm", type=float, default=1e-9)
    p.add_argument("--N_joints", type=int, default=0)
    p.add_argument("--static_cold_W", type=float, default=2.0e4)
    p.add_argument("--W_elec_per_Wcold", type=float, default=300.0)

    p.add_argument("--pump_frac_of_gross", type=float, default=0.03)
    p.add_argument("--P_net_min_MW", type=float, default=-1e9)

    args = p.parse_args()

    steady_state = args.steady_state and (not args.no_steady_state)

    Ti_grid = frange(args.Ti_start, args.Ti_stop, -abs(args.Ti_step) if args.Ti_stop < args.Ti_start else abs(args.Ti_step))
    H_grid = frange(args.H98_start, args.H98_stop, abs(args.H98_step))
    a_grid = frange(args.a_min, args.a_max, abs(args.a_step))
    Q_grid = frange(args.Q_start, args.Q_stop, abs(args.Q_step))
    g_grid = frange(args.gconf_start, args.gconf_stop, abs(args.gconf_step))

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("feasible_ext")

    # Columns: keep original ordering first, then append "clean design" fields.
    cols = [
        # scan indices
        "g_conf","Ti_keV","Q_target","H98_required","a_m",

        # solved point
        "R0_m","B0_T","kappa","Ip_MA","f_G","t_shield_m","ne20",

        # performance
        "Pfus_DD_MW","Pfus_DT_eqv_MW","Pfus_DT_adj_MW","Pfus_total_MW",
        "Q_DT_eqv","H98","H98_eff",

        # power
        "Paux_MW","Paux_for_Q_MW","Palpha_MW","Pin_MW",
        "Prad_core_MW","P_SOL_MW","Ploss_MW",
        "P_LH_MW","LH_ok",

        # screening proxies
        "q95_proxy","betaN_proxy","f_bs_proxy",
        "PSOL_over_R","S_n_W_m2","P_n_captured_MW","eps_n",

        # (1) build/magnet
        "radial_build_ok","rb_total_inboard_m","R_coil_inner_m",
        "B_peak_T","sigma_hoop_MPa","sigma_allow_MPa",

        # (2) HTS + dump
        "Tcoil_K","hts_margin","hts_margin_min",
        "E_tf_MJ","I_tf_A","V_dump_kV","Vmax_kV","tau_dump_s","N_tf_turns",

        # (3) divertor
        "lambda_q_mm","Lpar_m","flux_expansion","f_rad_div",
        "q_div_MW_m2","q_div_max_MW_m2",

        # (4) neutronics
        "TBR","TBR_min","blanket_coverage",
        "hts_fluence_per_fpy_n_m2","hts_fluence_limit_n_m2","hts_lifetime_yr","lifetime_min_yr",

        # (5) systems power
        "P_gross_e_MW","P_recirc_MW","P_net_e_MW","P_net_min_MW",
        "I_CD_MA","P_CD_launch_MW","P_CD_wall_MW","P_aux_wall_MW","P_cryo_e_MW","P_pumps_MW",
    ]

    ws.append(cols)
    for c in range(1, len(cols)+1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    ws.freeze_panes = "A2"

    fail_rows: List[Dict] = []
    best_g = None

    for g_conf in g_grid:
        base_ok = 0
        ext_ok = 0

        for Ti in Ti_grid:
            for Hreq in H_grid:
                for a in a_grid:
                    for Qtar in Q_grid:
                        # Required H98 can be met via confinement multiplier g_conf:
                        # We solve for the base H98 that, when multiplied, reaches Hreq.
                        H_base_target = Hreq / max(g_conf, 1e-9)

                        base = PointInputs(
                            # Phase-1 knobs
                            R0_m=args.R0, a_m=a, kappa=args.kappa, Bt_T=args.B0,
                            Ip_MA=0.5*(args.Ip_min+args.Ip_max),
                            Ti_keV=Ti, fG=0.8,
                            t_shield_m=args.tshield,
                            Paux_MW=args.Paux,
                            Ti_over_Te=args.Ti_over_Te,

                            # plasma knobs
                            zeff=args.Zeff,
                            dilution_fuel=args.dilution_fuel,
                            alpha_loss_frac=args.alpha_loss_frac,
                            f_rad_core=args.f_rad_core,
                            f_rad_div=args.f_rad_div,

                            # screening knobs
                            C_bs=args.C_bs,
                            require_Hmode=args.require_Hmode,
                            PLH_margin=args.PLH_margin,

                            # (1) build/magnet
                            t_fw_m=args.t_fw,
                            t_blanket_m=args.t_blanket,
                            t_vv_m=args.t_vv,
                            t_gap_m=args.t_gap,
                            t_tf_wind_m=args.t_tf_wind,
                            t_tf_struct_m=args.t_tf_struct,
                            Bpeak_factor=args.Bpeak_factor,
                            sigma_allow_MPa=args.sigma_allow_MPa,

                            # (2) HTS/dump
                            Tcoil_K=args.Tcoil_K,
                            hts_margin_min=args.hts_margin_min,
                            N_tf_turns=args.N_tf_turns,
                            tau_dump_s=args.tau_dump_s,
                            Vmax_kV=args.Vmax_kV,
                            tf_energy_volume_factor=args.tf_energy_volume_factor,

                            # (3) divertor
                            lambda_q_factor=args.lambda_q_factor,
                            flux_expansion=args.flux_expansion,
                            q_div_max_MW_m2=args.q_div_max,
                            f_Lpar=args.f_Lpar,

                            # (4) neutronics
                            blanket_coverage=args.blanket_coverage,
                            TBR_min=args.TBR_min,
                            TBR_lambda_m=args.TBR_lambda_m,
                            TBR_multiplier=args.TBR_multiplier,
                            hts_fluence_limit_n_m2=args.hts_fluence_limit_n_m2,
                            atten_len_m=args.atten_len_m,
                            f_geom_to_tf=args.f_geom_to_tf,
                            lifetime_min_yr=args.lifetime_min_yr,

                            # (5) net power closure
                            eta_elec=args.eta_elec,
                            eta_aux_wallplug=args.eta_aux_wallplug,
                            eta_cd_wallplug=args.eta_cd_wallplug,
                            eta_cd_MA_per_MW=args.eta_cd_MA_per_MW,
                            steady_state=steady_state,
                            R_joint_ohm=args.R_joint_ohm,
                            N_joints=args.N_joints,
                            static_cold_W=args.static_cold_W,
                            W_elec_per_Wcold=args.W_elec_per_Wcold,
                            pump_frac_of_gross=args.pump_frac_of_gross,
                            P_net_min_MW=args.P_net_min_MW,
                        )

                        sol_inp, sol_out, ok = solve_Ip_for_H98_with_Q_match(
                            base=base,
                            target_H98=H_base_target,
                            target_Q=Qtar,
                            Ip_min=args.Ip_min,
                            Ip_max=args.Ip_max,
                            fG_min=args.fG_min,
                            fG_max=args.fG_max,
                            tol=args.tol,
                            Paux_for_Q_MW=args.Paux_for_Q,
                        )
                        if not ok:
                            if args.write_failures:
                                fail_rows.append({
                                    "g_conf": g_conf, "Ti_keV": Ti, "Q_target": Qtar, "H98_required": Hreq, "a_m": a,
                                    "stage": "solve", "msg": "no_bracket_or_nan"
                                })
                            continue
                        base_ok += 1

                        # Effective confinement with g_conf
                        H98_eff = g_conf * sol_out["H98"]

                        # Extended checks
                        stage = None
                        msg = None
                        ok_ext = True

                        # Original extended constraints
                        if sol_out["ne20"] > 1.2:
                            ok_ext = False; stage="n20"; msg=f"ne20={sol_out['ne20']:.3f}"

                        if ok_ext and sol_out.get("q95_proxy", 1e9) < args.q95_min:
                            ok_ext = False; stage="q95"; msg=f"q95={sol_out['q95_proxy']:.3f}"

                        if ok_ext and sol_out.get("betaN_proxy", 0.0) > args.betaN_max:
                            ok_ext = False; stage="betaN"; msg=f"betaN={sol_out['betaN_proxy']:.3f}"

                        if ok_ext and sol_out.get("f_bs_proxy", 0.0) > args.f_bs_max:
                            ok_ext = False; stage="f_bs"; msg=f"f_bs={sol_out['f_bs_proxy']:.3f}"

                        PSOL_over_R = sol_out["P_SOL_MW"] / args.R0
                        if ok_ext and PSOL_over_R > args.PSOL_over_R_max:
                            ok_ext = False; stage="PSOL/R"; msg=f"{PSOL_over_R:.3f}"

                        if ok_ext and args.require_Hmode and sol_out.get("LH_ok", 1.0) < 0.5:
                            ok_ext = False; stage="LH"; msg="Hmode_access_failed"

                        if ok_ext and H98_eff < Hreq:
                            ok_ext = False; stage="H98"; msg=f"H98_eff={H98_eff:.3f} < {Hreq:.3f}"

                        if ok_ext and sol_out["Q_DT_eqv"] < Qtar:
                            ok_ext = False; stage="Q"; msg=f"Q={sol_out['Q_DT_eqv']:.3f}"

                        # -------------------------------------------------
                        # Added "clean point design" constraints
                        # -------------------------------------------------

                        if ok_ext and sol_out.get("radial_build_ok", 0.0) < 0.5:
                            ok_ext = False; stage="build"; msg="radial_build_failed"

                        if ok_ext and sol_out.get("sigma_hoop_MPa", 0.0) > sol_out.get("sigma_allow_MPa", 1e9):
                            ok_ext = False; stage="stress"; msg=f"sigma={sol_out['sigma_hoop_MPa']:.1f}MPa"

                        if ok_ext and sol_out.get("hts_margin", 0.0) < sol_out.get("hts_margin_min", 0.0):
                            ok_ext = False; stage="HTS"; msg=f"margin={sol_out['hts_margin']:.3f}"

                        if ok_ext and sol_out.get("V_dump_kV", 0.0) > sol_out.get("Vmax_kV", 1e9):
                            ok_ext = False; stage="dumpV"; msg=f"Vdump={sol_out['V_dump_kV']:.2f}kV"

                        if ok_ext and sol_out.get("q_div_MW_m2", 0.0) > sol_out.get("q_div_max_MW_m2", 1e9):
                            ok_ext = False; stage="divertor"; msg=f"qdiv={sol_out['q_div_MW_m2']:.2f}"

                        if ok_ext and sol_out.get("TBR", 0.0) < sol_out.get("TBR_min", 0.0):
                            ok_ext = False; stage="TBR"; msg=f"TBR={sol_out['TBR']:.3f}"

                        if ok_ext and sol_out.get("hts_lifetime_yr", 0.0) < sol_out.get("lifetime_min_yr", 0.0):
                            ok_ext = False; stage="life"; msg=f"yr={sol_out['hts_lifetime_yr']:.2f}"

                        # Optional net power constraint
                        Pnet_min = sol_out.get("P_net_min_MW", -1e9)
                        if ok_ext and Pnet_min > -1e8 and sol_out.get("P_net_e_MW", -1e9) < Pnet_min:
                            ok_ext = False; stage="Pnet"; msg=f"Pnet={sol_out['P_net_e_MW']:.2f}"

                        if not ok_ext:
                            if args.write_failures:
                                fail_rows.append({
                                    "g_conf": g_conf, "Ti_keV": Ti, "Q_target": Qtar, "H98_required": Hreq, "a_m": a,
                                    "stage": stage, "msg": msg
                                })
                            continue

                        # Record feasible row
                        ext_ok += 1
                        if best_g is None or g_conf < best_g:
                            best_g = g_conf

                        row = [sol_out.get(k) if k in sol_out else None for k in cols]
                        # Fix scan index columns explicitly
                        row[0] = g_conf
                        row[1] = Ti
                        row[2] = Qtar
                        row[3] = Hreq
                        row[4] = a
                        row[7] = args.kappa
                        row[17] = H98_eff
                        row[19] = args.Paux_for_Q
                        row[31] = PSOL_over_R
                        ws.append(row)

        print(f"g_conf={g_conf:.2f}  base_ok={base_ok:4d}  ext_ok={ext_ok:4d}")

    autosize_columns(ws)

    if args.write_failures:
        wsF = wb.create_sheet("failures")
        fail_cols = ["g_conf", "Ti_keV", "Q_target", "H98_required", "a_m", "stage", "msg"]
        wsF.append(fail_cols)
        for c in range(1, len(fail_cols) + 1):
            wsF.cell(row=1, column=c).font = Font(bold=True)
        for r in fail_rows:
            wsF.append([r.get(k) for k in fail_cols])
        wsF.freeze_panes = "A2"
        autosize_columns(wsF)

    wsM = wb.create_sheet("meta")
    wsM.append(["key", "value"])
    wsM["A1"].font = Font(bold=True)
    meta = {k: getattr(args, k) for k in vars(args).keys()}
    meta["steady_state_effective"] = steady_state
    meta["best_g_conf_found"] = best_g if best_g is not None else "NONE"
    for k, v in meta.items():
        wsM.append([k, v])
    wsM.freeze_panes = "A2"
    autosize_columns(wsM)

    wb.save(args.out_xlsx)
    print(f"\nWrote: {args.out_xlsx}")


if __name__ == "__main__":
    main()
